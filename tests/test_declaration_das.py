# -*- coding: utf-8 -*-
"""Tests — Déclaration Annuelle des Salaires (DAS) + gating abonnement Cabinet."""
import io
from datetime import date, datetime

import pytest
import openpyxl

from app import app as flask_app
from models import (db, Plan, Tenant, Utilisateur, CategorieEmploi,
                    Salarie, PeriodePaie, BulletinPaie,
                    Prestataire, FacturePrestataire, ContratPrestation)


@pytest.fixture
def app():
    flask_app.config.update({"TESTING": True, "WTF_CSRF_ENABLED": False,
                             "SERVER_NAME": "localhost"})
    with flask_app.app_context():
        db.drop_all()
        db.create_all()
        _seed()
        yield flask_app
        db.session.remove()
        db.drop_all()


@pytest.fixture
def client(app):
    return app.test_client()


def _seed():
    cabinet = Plan(code="CABINET", nom="Cabinet", prix_mensuel=100000,
                   max_salaries=None, max_utilisateurs=10, actif=True)
    starter = Plan(code="STARTER", nom="Starter", prix_mensuel=15000,
                   max_salaries=10, max_utilisateurs=1, actif=True)
    db.session.add_all([cabinet, starter]); db.session.flush()

    # Tenant Cabinet (accès DAS)
    t_cab = Tenant(slug="cabinet-co", denomination="CABINET CO", sigle="CAB",
                   numero_cnss="CNSS-CAB", convention="COMMERCE",
                   plan_id=cabinet.id, statut="ACTIF", token_api="tk_cab")
    # Tenant Starter (pas d'accès DAS)
    t_star = Tenant(slug="petit-co", denomination="PETIT CO", sigle="PET",
                    plan_id=starter.id, statut="ACTIF", token_api="tk_star")
    db.session.add_all([t_cab, t_star]); db.session.flush()

    cat = CategorieEmploi(tenant_id=t_cab.id, code="E3", libelle="Exéc 3",
                          salaire_minimum=105300)
    db.session.add(cat); db.session.flush()

    admin_cab = Utilisateur(nom="A", prenom="Cab", email="admin@cab.ga",
                            role="TENANT_ADMIN", tenant_id=t_cab.id,
                            actif=True, email_verifie=True)
    admin_cab.set_password("MotDePasse1")
    admin_star = Utilisateur(nom="A", prenom="Star", email="admin@star.ga",
                             role="TENANT_ADMIN", tenant_id=t_star.id,
                             actif=True, email_verifie=True)
    admin_star.set_password("MotDePasse1")
    db.session.add_all([admin_cab, admin_star]); db.session.flush()

    s = Salarie(tenant_id=t_cab.id, matricule="M001", nom="NDONG", prenom="Paul",
                nationalite="GABONAISE", sexe="M", date_naissance=date(1990, 5, 1),
                date_embauche=date(2020, 3, 1), situation_matrimoniale="Marié",
                nb_enfants=2, numero_cnss="CN-001", emploi="Vendeur",
                categorie_id=cat.id, statut="ACTIF")
    db.session.add(s); db.session.flush()

    for mois in (1, 2):
        p = PeriodePaie(tenant_id=t_cab.id, annee=2025, mois=mois,
                        libelle_mois=PeriodePaie.MOIS_NOMS[mois], statut="OUVERT")
        db.session.add(p); db.session.flush()
        db.session.add(BulletinPaie(
            tenant_id=t_cab.id, salarie_id=s.id, periode_id=p.id,
            salaire_base=300000, salaire_brut=400000,
            indem_logement=50000, indem_transport=30000, allocations_conge=20000,
            tcs=8000, irpp=25000, cfp=1500, fnh=4500,
            net_a_payer=330000, statut="VALIDÉ", date_validation=datetime.utcnow()))

    # ── Prestataires (honoraires) ────────────────────────────────────────────
    p_local = Prestataire(tenant_id=t_cab.id, code="PR001", type_personne="PHYSIQUE",
                          categorie="FREELANCE", raison_sociale="MBADINGA Consulting",
                          nif="NIF-LOC", activite="Conseil", resident=True,
                          assujetti_tva=True, ville="Libreville", pays="Gabon",
                          telephone="01-02-03-04", statut="ACTIF")
    p_etr = Prestataire(tenant_id=t_cab.id, code="PR002", type_personne="MORALE",
                        categorie="SOUS_TRAITANT", raison_sociale="GLOBAL TECH LTD",
                        nif="NIF-ETR", activite="Ingénierie", resident=False,
                        assujetti_tva=False, ville="Paris", pays="France", statut="ACTIF")
    db.session.add_all([p_local, p_etr]); db.session.flush()

    ct = ContratPrestation(tenant_id=t_cab.id, prestataire_id=p_local.id,
                           objet="Mission de conseil RH", montant=1000000,
                           date_debut=date(2025, 1, 10), statut="EN_COURS")
    db.session.add(ct)

    # Facture locale : HT 1 000 000, TVA 18%, RAS 9,5%
    f1 = FacturePrestataire(tenant_id=t_cab.id, prestataire_id=p_local.id,
                            numero="F-2025-001", date_facture=date(2025, 3, 15),
                            montant_ht=1000000, taux_tva=18, taux_retenue=9.5,
                            statut="PAYEE")
    f1.calculer()
    # Facture étrangère : HT 2 000 000, TVA 0, RAS 20%
    f2 = FacturePrestataire(tenant_id=t_cab.id, prestataire_id=p_etr.id,
                            numero="F-2025-002", date_facture=date(2025, 6, 20),
                            montant_ht=2000000, taux_tva=0, taux_retenue=20,
                            statut="EN_ATTENTE")
    f2.calculer()
    db.session.add_all([f1, f2])
    db.session.commit()


def _login(client, email):
    return client.post("/login", data={"email": email, "password": "MotDePasse1"},
                       follow_redirects=False)


# ── Agrégation ────────────────────────────────────────────────────────────────
def test_agregation_das(app):
    from declaration_das import agreger_das
    import models as M
    t = Tenant.query.filter_by(slug="cabinet-co").first()
    lignes, totaux = agreger_das(t, 2025, models=M)
    assert totaux["nb_salaries"] == 1
    l = lignes[0]
    assert l["total_1a5"] == 800000
    assert l["brut_presence"] == 660000          # 800k − 100k AN − 40k congé
    assert l["av_logement"] == 100000
    assert l["brut_conge"] == 40000
    assert l["ni_transport"] == 60000
    assert l["tcts"] == 16000 and l["irpp"] == 50000
    assert l["nationalite"] == 1 and l["sexe"] == 1 and l["situation"] == 1
    # Cohérence comptable
    av = l["av_logement"] + l["av_eau"] + l["av_dom"] + l["av_nour"]
    assert l["brut_presence"] + av + l["brut_conge"] == l["total_1a5"]


def test_das_vide_leve_exception(app):
    from declaration_das import agreger_das, DASVide
    import models as M
    t = Tenant.query.filter_by(slug="cabinet-co").first()
    with pytest.raises(DASVide):
        agreger_das(t, 2099, models=M)   # exercice sans bulletin


def test_generation_excel(app):
    from declaration_das import generer_das_excel
    import models as M
    t = Tenant.query.filter_by(slug="cabinet-co").first()
    contenu = generer_das_excel(t, 2025, models=M)
    wb = openpyxl.load_workbook(io.BytesIO(contenu))
    for feuille in ("Paramètres", "ID19 - Détail salaires", "ID20 - Récapitulatif"):
        assert feuille in wb.sheetnames


# ── Gating abonnement ─────────────────────────────────────────────────────────
def test_das_accessible_cabinet(client):
    _login(client, "admin@cab.ga")
    r = client.get("/declaration-das?annee=2025")
    assert r.status_code == 200
    assert b"NDONG" in r.data


def test_das_refuse_starter(client):
    _login(client, "admin@star.ga")
    r = client.get("/declaration-das?annee=2025", follow_redirects=False)
    # plan_required redirige vers la page d'abonnement
    assert r.status_code in (301, 302)
    assert "/paiement" in r.headers.get("Location", "")


def test_export_excel_refuse_starter(client):
    _login(client, "admin@star.ga")
    r = client.get("/declaration-das/excel?annee=2025", follow_redirects=False)
    assert r.status_code in (301, 302)
    assert "/paiement" in r.headers.get("Location", "")


def test_export_excel_cabinet_telecharge(client):
    _login(client, "admin@cab.ga")
    r = client.get("/declaration-das/excel?annee=2025")
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers.get("Content-Type", "")
    assert "DAS_" in r.headers.get("Content-Disposition", "")


# ── Honoraires (prestataires) ─────────────────────────────────────────────────
def test_agregation_honoraires(app):
    from declaration_das import agreger_honoraires
    import models as M
    t = Tenant.query.filter_by(slug="cabinet-co").first()
    lignes, tot = agreger_honoraires(t, 2025, models=M)
    assert tot["nb_prestataires"] == 2
    assert tot["montant_ht"] == 3000000             # 1M + 2M
    assert tot["tva"] == 180000                     # 18% de 1M ; étranger 0%
    # RAS : local 9,5% de 1M = 95 000 ; étranger 20% de 2M = 400 000
    assert tot["retenue_local"] == 95000
    assert tot["retenue_etranger"] == 400000
    assert tot["retenue_total"] == 495000
    par_nif = {l["nif"]: l for l in lignes}
    assert par_nif["NIF-LOC"]["resident"] is True
    assert par_nif["NIF-ETR"]["resident"] is False
    assert par_nif["NIF-ETR"]["retenue_etranger"] == 400000


def test_excel_contient_feuille_honoraires(app):
    from declaration_das import generer_das_excel
    import models as M
    t = Tenant.query.filter_by(slug="cabinet-co").first()
    contenu = generer_das_excel(t, 2025, models=M)
    wb = openpyxl.load_workbook(io.BytesIO(contenu))
    assert "ID23-24 - Honoraires" in wb.sheetnames


def test_ecran_affiche_honoraires(client):
    _login(client, "admin@cab.ga")
    r = client.get("/declaration-das?annee=2025")
    assert r.status_code == 200
    assert b"MBADINGA" in r.data and b"GLOBAL TECH" in r.data
    assert "Honoraires".encode() in r.data
