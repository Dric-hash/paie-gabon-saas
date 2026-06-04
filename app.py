import os
"""
app.py — SaaS Paie Gabon — Multi-tenant
"""
from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_file, abort, session
from sqlalchemy.orm import joinedload
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from datetime import datetime, date, timedelta
from functools import wraps
import io, os, secrets as sec, threading
import hmac

from models import (db, Plan, Tenant, Utilisateur, CategorieEmploi, Salarie,
                    Contrat, PeriodePaie, BulletinPaie, RubriquePaie, Conge,
                    Acompte, Journalier, Pointage, FeuillePaieJournalier,
                    Site, AffectationSite, Paiement, OAuthClient)
from calculs_paie import calculer_bulletin, calculer_masse_salariale
from flask_mail import Mail, Message
from flask_wtf.csrf import CSRFProtect


# ══════════════════════════════════════════════════════════════════════════════
# CACHE — KPIs Dashboard
# Utilise Redis si REDIS_URL est définie, sinon désactivé (no-op).
# Le cache en mémoire (dict Python) était incompatible avec --workers >1 :
# chaque worker Gunicorn a sa propre copie, les invalidations ne se propagent pas.
# ══════════════════════════════════════════════════════════════════════════════
_redis_client = None
_REDIS_URL = os.environ.get("REDIS_URL", "")
if _REDIS_URL:
    try:
        import redis as _redis_lib
        _redis_client = _redis_lib.from_url(_REDIS_URL, decode_responses=True,
                                             socket_connect_timeout=2,
                                             socket_timeout=2)
        _redis_client.ping()
        print(f"[CACHE] Redis connecté.")
    except Exception as _e:
        print(f"[CACHE] Connexion Redis échouée ({_e}). Cache désactivé.")
        _redis_client = None
else:
    print("[CACHE] REDIS_URL non définie. Cache dashboard désactivé (données fraîches à chaque requête).")

import json as _json

def _cache_get(key: str):
    if not _redis_client:
        return None
    try:
        raw = _redis_client.get(key)
        return _json.loads(raw) if raw else None
    except Exception:
        return None

def _cache_set(key: str, value, ttl_seconds: int = 300):
    if not _redis_client:
        return
    try:
        _redis_client.setex(key, ttl_seconds, _json.dumps(value, default=str))
    except Exception:
        pass

def _cache_delete(key_prefix: str):
    if not _redis_client:
        return
    try:
        cursor = 0
        while True:
            cursor, keys = _redis_client.scan(cursor, match=f"{key_prefix}*", count=100)
            if keys:
                _redis_client.delete(*keys)
            if cursor == 0:
                break
    except Exception:
        pass

# TTL par type de données (secondes)
TTL_KPIS_DASH   = 300   # KPIs emploi + bulletins : 5 min
TTL_EVOLUTION   = 600   # Courbes 6 mois : 10 min (ça ne change pas souvent)
TTL_CATS_STATS  = 600   # Répartition par catégorie : 10 min
TTL_TOP_SAL     = 300   # Top 5 salariés : 5 min
TTL_ALERTES     = 120   # Alertes : 2 min (sensibles)

app = Flask(__name__)
_secret = os.environ.get("SECRET_KEY", "")
if not _secret:
    import sys
    if os.environ.get("FLASK_ENV") == "production" or os.environ.get("RAILWAY_ENVIRONMENT"):
        print("ERREUR CRITIQUE : la variable d'environnement SECRET_KEY n'est pas définie. "
              "Démarrage interrompu.", file=sys.stderr)
        sys.exit(1)
    else:
        # Dev/test uniquement — génère une clé aléatoire à chaque redémarrage
        # (les sessions ne survivent pas aux redémarrages, c'est voulu en dev)
        _secret = sec.token_hex(32)
        print("[DEV] SECRET_KEY non définie — clé temporaire générée. "
              "Définissez SECRET_KEY en production.", file=sys.stderr)
app.config["SECRET_KEY"] = _secret
_db_url = os.environ.get("DATABASE_URL", "sqlite:///saas_paie.db")
if _db_url.startswith("postgres://"):
    _db_url = _db_url.replace("postgres://", "postgresql://", 1)
app.config["SQLALCHEMY_DATABASE_URI"]        = _db_url
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
app.config["SQLALCHEMY_ENGINE_OPTIONS"] = {
    "pool_size":     5,
    "max_overflow":  10,
    "pool_timeout":  30,
    "pool_recycle":  1800,
    "pool_pre_ping": True,
}
db.init_app(app)

# Configuration email Gmail
app.config["MAIL_SERVER"]   = "smtp.gmail.com"
app.config["MAIL_PORT"]     = 587
app.config["MAIL_USE_TLS"]  = True
app.config["MAIL_USERNAME"] = os.environ.get("MAIL_USERNAME", "")
app.config["MAIL_PASSWORD"] = os.environ.get("MAIL_PASSWORD", "")
app.config["MAIL_DEFAULT_SENDER"] = os.environ.get("MAIL_USERNAME", "noreply@paiegalon.ga")
app.config["MAIL_SUPPRESS_SEND"] = not bool(os.environ.get("MAIL_USERNAME", ""))

# ── Sécurité sessions ─────────────────────────────────────────────────────────
app.config["PERMANENT_SESSION_LIFETIME"] = timedelta(minutes=int(os.environ.get("SESSION_TIMEOUT_MINUTES", "60")))
app.config["SESSION_COOKIE_HTTPONLY"]    = True   # Protège contre XSS
app.config["SESSION_COOKIE_SAMESITE"]   = "Lax"  # Protège contre CSRF
# En prod Railway, HTTPS est garanti
app.config["SESSION_COOKIE_SECURE"]     = os.environ.get("RAILWAY_ENVIRONMENT") == "production"

# ── Middleware inactivité session ──────────────────────────────────────────────
@app.before_request
def gerer_session_inactivite():
    """Déconnecter l'utilisateur après X minutes d'inactivité."""
    # Ne pas vérifier sur les routes publiques
    excluded = ["/login", "/inscription", "/confirmer-email", "/mot-de-passe-oublie",
                "/reinitialiser-mdp", "/politique-confidentialite", "/static"]
    if any(request.path.startswith(p) for p in excluded):
        return
    if current_user.is_authenticated:
        now      = datetime.utcnow()
        derniere = session.get("derniere_activite")
        timeout  = int(os.environ.get("SESSION_TIMEOUT_MINUTES", "60"))
        if derniere:
            try:
                elapsed = (now - datetime.fromisoformat(derniere)).total_seconds() / 60
                if elapsed > timeout:
                    logout_user()
                    session.clear()
                    flash(f"Session expirée après {timeout} min d'inactivité. Reconnectez-vous.", "error")
                    return redirect(url_for("login"))
            except Exception:
                pass
        session["derniere_activite"] = now.isoformat()
        session.permanent = True


mail = Mail(app)
csrf = CSRFProtect(app)
# Accepter le token CSRF aussi via le header HTTP (pour les appels fetch/AJAX)
app.config["WTF_CSRF_CHECK_DEFAULT"] = True
app.config["WTF_CSRF_TIME_LIMIT"]    = 3600  # 1h

# ── Logging structuré ─────────────────────────────────────────────────────────
import logging, sys as _sys
_log_level = os.environ.get("LOG_LEVEL", "INFO").upper()
logging.basicConfig(
    stream=_sys.stdout,
    level=getattr(logging, _log_level, logging.INFO),
    format="%(asctime)s [%(levelname)s] %(name)s — %(message)s",
    datefmt="%Y-%m-%dT%H:%M:%S",
)
logger = logging.getLogger("paiegalon")
# Réduire le bruit de SQLAlchemy en production
if os.environ.get("SQLALCHEMY_ECHO") != "1":
    logging.getLogger("sqlalchemy.engine").setLevel(logging.WARNING)

login_manager = LoginManager(app)
login_manager.login_view = "login"
login_manager.login_message = "Veuillez vous connecter."

@login_manager.user_loader
def load_user(uid): return Utilisateur.query.get(int(uid))

def super_admin_required(f):
    @wraps(f)
    def d(*a,**k):
        if not current_user.is_authenticated or not current_user.is_super_admin: abort(403)
        return f(*a,**k)
    return d

def _parse_date(val):
    if not val: return None
    if isinstance(val, date): return val
    if isinstance(val, datetime): return val.date()
    try: return datetime.strptime(str(val).strip()[:10], "%Y-%m-%d").date()
    except: return None

def calculer_parts_irpp(situation_matrimoniale: str, nb_enfants: int) -> float:
    situation = (situation_matrimoniale or "").upper().strip()
    nb_enf = int(nb_enfants or 0)
    if "CELIBATAIRE" in situation and "AEAC" in situation: parts = 1.5
    elif "CELIBATAIRE" in situation: parts = 1.0
    elif "DIVORCE" in situation and "AEAC" in situation: parts = 1.5
    elif "DIVORCE" in situation: parts = 1.0
    elif "MARIE" in situation or "MARIÉ" in situation: parts = 2.0
    elif "VEUF" in situation and "2 ANS" in situation: parts = 1.5
    elif "VEUF" in situation: parts = 2.0
    else: parts = 1.0
    parts += nb_enf * 0.5
    return round(parts, 1)

def tenant_required(f):
    @wraps(f)
    def d(*a,**k):
        if not current_user.is_authenticated: return redirect(url_for("login"))
        if not current_user.is_super_admin and (not current_user.tenant_id or not current_user.tenant or current_user.tenant.statut not in ("ACTIF","ESSAI","PAIEMENT_EN_ATTENTE")):
            flash("Compte suspendu ou non associé à une entreprise.","error")
            return redirect(url_for("login"))
        return f(*a,**k)
    return d

def can_edit(f):
    @wraps(f)
    def d(*a,**k):
        if not current_user.can_edit: abort(403)
        return f(*a,**k)
    return d

def get_tenant():
    if current_user.is_super_admin: return None
    return current_user.tenant

# ── Envoi email asynchrone (ne bloque pas le serveur) ─────────────────────────
def send_email_async(msg):
    """Envoie un email dans un thread séparé pour ne pas bloquer Gunicorn."""
    def run(app_context, message):
        with app_context:
            try:
                mail.send(message)
            except Exception as e:
                print(f"[EMAIL ERROR] {e}")
    t = threading.Thread(target=run, args=(app.app_context(), msg))
    t.daemon = True
    t.start()

# ── Auth ──────────────────────────────────────────────────────────────────────
@app.route("/")
def index():
    if current_user.is_authenticated:
        return redirect(url_for("admin_dashboard") if current_user.is_super_admin else url_for("dashboard"))
    return redirect(url_for("login"))

@app.route("/login", methods=["GET","POST"])
def login():
    if current_user.is_authenticated: return redirect(url_for("index"))
    if request.method == "POST":
        email = request.form.get("email","").strip().lower()
        pw    = request.form.get("password","")
        user  = Utilisateur.query.filter_by(email=email, actif=True).first()

        # ── Protection anti-brute-force ───────────────────────────────────────
        now = datetime.utcnow()
        MAX_ECHECS   = int(os.environ.get("LOGIN_MAX_ECHECS",   "5"))
        BLOCAGE_MIN  = int(os.environ.get("LOGIN_BLOCAGE_MIN",  "15"))

        if user:
            # Vérifier si le compte est bloqué
            if user.compte_bloque_jusqu and now < user.compte_bloque_jusqu:
                reste = int((user.compte_bloque_jusqu - now).total_seconds() / 60) + 1
                flash(f"Compte temporairement bloqué. Réessayez dans {reste} min.", "error")
                return render_template("auth/login.html")

            if user.check_password(pw):
                # Succès : réinitialiser les compteurs
                user.nb_echecs_connexion = 0
                user.compte_bloque_jusqu = None
                user.derniere_connexion  = now
                db.session.commit()
                login_user(user)
                return redirect(url_for("index"))
            else:
                # Échec : incrémenter et éventuellement bloquer
                user.nb_echecs_connexion = (user.nb_echecs_connexion or 0) + 1
                if user.nb_echecs_connexion >= MAX_ECHECS:
                    user.compte_bloque_jusqu = now + timedelta(minutes=BLOCAGE_MIN)
                    user.nb_echecs_connexion = 0
                    db.session.commit()
                    flash(f"Trop de tentatives. Compte bloqué {BLOCAGE_MIN} minutes.", "error")
                    return render_template("auth/login.html")
                db.session.commit()

        flash("Email ou mot de passe incorrect.","error")
    return render_template("auth/login.html")

@app.route("/inscription", methods=["GET","POST"])
def inscription():
    plans = Plan.query.filter_by(actif=True).all()
    if request.method == "POST":
        email = request.form.get("email","").strip().lower()
        if Utilisateur.query.filter_by(email=email).first():
            flash("Email déjà utilisé.","error")
            return render_template("auth/inscription.html", plans=plans)
        plan = Plan.query.get(request.form.get("plan_id","")) or Plan.query.filter_by(code="STARTER").first()
        denom = request.form.get("denomination","").strip()
        slug_base = denom.lower().replace(" ","_")[:30]
        slug = slug_base; i=1
        while Tenant.query.filter_by(slug=slug).first(): slug=f"{slug_base}_{i}"; i+=1
        t = Tenant(slug=slug, denomination=denom.upper(),
                   sigle=request.form.get("sigle","").strip().upper(),
                   activite=request.form.get("activite","").strip(),
                   nif=request.form.get("nif","").strip(),
                   telephone=request.form.get("telephone","").strip(),
                   ville=request.form.get("ville","Libreville"),
                   pays="Gabon", plan_id=plan.id if plan else None,
                   statut="ESSAI", date_expiration=datetime.utcnow()+timedelta(days=30))
        t.token_api = sec.token_hex(32)
        db.session.add(t); db.session.flush()
        for code,lib in [("C1","Ouvriers"),("C2","Techniciens"),("C3","Conducteurs de Travaux"),("C4","Cadres")]:
            db.session.add(CategorieEmploi(tenant_id=t.id,code=code,libelle=lib))
        admin = Utilisateur(nom=request.form.get("nom","").strip().upper(),
                            prenom=request.form.get("prenom","").strip(),
                            email=email, role="TENANT_ADMIN", tenant_id=t.id, actif=True)
        admin.set_password(request.form.get("password",""))
        # Générer token de confirmation email
        token_conf = sec.token_urlsafe(32)
        admin.token_confirmation        = token_conf
        admin.token_confirmation_expiry = datetime.utcnow() + timedelta(hours=48)
        admin.email_verifie             = False
        db.session.add(admin); db.session.commit()

        # Envoyer email de confirmation
        lien = url_for("confirmer_email", token=token_conf, _external=True)
        if os.environ.get("MAIL_USERNAME"):
            msg = Message(
                subject="✅ Confirmez votre inscription — PaieGabon",
                recipients=[email],
                html=f"""
                <div style="font-family:Arial,sans-serif;max-width:520px;margin:0 auto">
                  <div style="background:#1a2332;padding:1.5rem;border-radius:.75rem .75rem 0 0;text-align:center">
                    <h1 style="color:white;margin:0;font-size:1.25rem">PaieGabon</h1>
                    <p style="color:rgba(255,255,255,.7);margin:.25rem 0 0;font-size:.875rem">Ameriack I.T. Solutions</p>
                  </div>
                  <div style="background:white;padding:2rem;border:1px solid #e5e7eb;border-top:none;border-radius:0 0 .75rem .75rem">
                    <h2 style="color:#111827;margin:0 0 1rem">Bienvenue, {admin.prenom} !</h2>
                    <p style="color:#6b7280;line-height:1.6">
                      Votre compte <strong>{denom.upper()}</strong> a été créé avec succès.<br/>
                      Cliquez sur le bouton ci-dessous pour confirmer votre adresse email.
                    </p>
                    <div style="text-align:center;margin:1.5rem 0">
                      <a href="{lien}" style="background:#1a2332;color:white;padding:.875rem 2rem;border-radius:.75rem;font-weight:700;text-decoration:none;font-size:1rem">
                        ✅ Confirmer mon email
                      </a>
                    </div>
                    <p style="color:#9ca3af;font-size:.75rem;text-align:center">
                      Ce lien expire dans 48h.<br/>
                      Si vous n'avez pas créé ce compte, ignorez cet email.
                    </p>
                  </div>
                </div>""",
                sender=app.config["MAIL_DEFAULT_SENDER"]
            )
            send_email_async(msg)
            flash(f"Compte créé ! Un email de confirmation a été envoyé à {email}. Vérifiez votre boîte mail.", "success")
        else:
            admin.email_verifie = True  # En dev sans mail, activer directement
            db.session.commit()
            flash("Bienvenue ! Essai gratuit de 30 jours activé.", "success")
        login_user(admin)
        return redirect(url_for("dashboard"))
    return render_template("auth/inscription.html", plans=plans)

@app.route("/logout")
@login_required
def logout():
    logout_user()
    session.clear()
    return redirect(url_for("login"))

# ── Confirmation email après inscription ──────────────────────────────────────
@app.route("/confirmer-email/<token>")
def confirmer_email(token):
    u = Utilisateur.query.filter_by(token_confirmation=token).first()
    if not u:
        flash("Lien de confirmation invalide ou déjà utilisé.", "error")
        return redirect(url_for("login"))
    if u.token_confirmation_expiry and datetime.utcnow() > u.token_confirmation_expiry:
        flash("Ce lien a expiré (48h). Reconnectez-vous pour en demander un nouveau.", "error")
        return redirect(url_for("login"))
    u.email_verifie             = True
    u.token_confirmation        = None
    u.token_confirmation_expiry = None
    db.session.commit()
    flash("✅ Email confirmé ! Votre compte est pleinement activé.", "success")
    if current_user.is_authenticated:
        return redirect(url_for("dashboard"))
    return redirect(url_for("login"))

@app.route("/renvoyer-confirmation")
@login_required
def renvoyer_confirmation():
    """Renvoyer l'email de confirmation si non vérifié."""
    u = current_user
    if u.email_verifie:
        flash("Votre email est déjà confirmé.", "info")
        return redirect(url_for("dashboard"))
    if not os.environ.get("MAIL_USERNAME"):
        u.email_verifie = True
        db.session.commit()
        flash("Mode développement : email validé automatiquement.", "success")
        return redirect(url_for("dashboard"))
    token = sec.token_urlsafe(32)
    u.token_confirmation        = token
    u.token_confirmation_expiry = datetime.utcnow() + timedelta(hours=48)
    db.session.commit()
    lien = url_for("confirmer_email", token=token, _external=True)
    msg  = Message(
        subject="✅ Confirmez votre email — PaieGabon",
        recipients=[u.email],
        html=f'<p>Cliquez ici pour confirmer : <a href="{lien}">{lien}</a></p>',
        sender=app.config["MAIL_DEFAULT_SENDER"]
    )
    send_email_async(msg)
    flash(f"Email de confirmation renvoyé à {u.email}.", "success")
    return redirect(url_for("dashboard"))

# ── Modifier email de connexion ───────────────────────────────────────────────
@app.route("/profil/changer-email", methods=["GET","POST"])
@login_required
def changer_email():
    if current_user.is_super_admin:
        return redirect(url_for("admin_dashboard"))
    t = get_tenant()
    if request.method == "POST":
        nouvel_email = request.form.get("nouvel_email","").strip().lower()
        mot_de_passe = request.form.get("mot_de_passe","")
        if not nouvel_email or "@" not in nouvel_email:
            flash("Adresse email invalide.", "error")
            return render_template("auth/changer_email.html", tenant=t)
        if not current_user.check_password(mot_de_passe):
            flash("Mot de passe incorrect.", "error")
            return render_template("auth/changer_email.html", tenant=t)
        if Utilisateur.query.filter_by(email=nouvel_email).first():
            flash("Cette adresse email est déjà utilisée par un autre compte.", "error")
            return render_template("auth/changer_email.html", tenant=t)
        token = sec.token_urlsafe(32)
        current_user.nouvel_email_en_attente = nouvel_email
        current_user.token_changement_email  = token
        current_user.token_changement_expiry = datetime.utcnow() + timedelta(hours=24)
        db.session.commit()
        lien = url_for("confirmer_changement_email", token=token, _external=True)
        if os.environ.get("MAIL_USERNAME"):
            msg = Message(
                subject="📧 Confirmation changement d'email — PaieGabon",
                recipients=[nouvel_email],
                html=f"""
                <div style="font-family:Arial,sans-serif;max-width:520px;margin:0 auto">
                  <div style="background:#1a2332;padding:1.5rem;text-align:center;border-radius:.75rem .75rem 0 0">
                    <h1 style="color:white;margin:0">PaieGabon</h1>
                  </div>
                  <div style="background:white;padding:2rem;border:1px solid #e5e7eb;border-top:none;border-radius:0 0 .75rem .75rem">
                    <h2 style="color:#111827">Confirmation de changement d'email</h2>
                    <p style="color:#6b7280">Vous avez demandé à changer votre email de connexion vers :<br/>
                    <strong>{nouvel_email}</strong></p>
                    <p style="color:#6b7280">Cliquez sur le lien ci-dessous pour confirmer ce changement :</p>
                    <div style="text-align:center;margin:1.5rem 0">
                      <a href="{lien}" style="background:#1a2332;color:white;padding:.875rem 2rem;border-radius:.75rem;font-weight:700;text-decoration:none">
                        ✅ Confirmer le nouvel email
                      </a>
                    </div>
                    <p style="color:#9ca3af;font-size:.75rem;text-align:center">
                      Lien valable 24h. Si ce n'est pas vous, ignorez cet email.
                    </p>
                  </div>
                </div>""",
                sender=app.config["MAIL_DEFAULT_SENDER"]
            )
            send_email_async(msg)
            flash(f"Un lien de confirmation a été envoyé à {nouvel_email}. Cliquez dessus pour valider le changement.", "success")
        else:
            current_user.email = nouvel_email
            current_user.nouvel_email_en_attente = None
            current_user.token_changement_email  = None
            db.session.commit()
            flash(f"Email mis à jour : {nouvel_email}", "success")
        return redirect(url_for("parametres"))
    return render_template("auth/changer_email.html", tenant=t)

@app.route("/profil/confirmer-email/<token>")
@login_required
def confirmer_changement_email(token):
    u = Utilisateur.query.filter_by(token_changement_email=token).first()
    if not u:
        flash("Lien invalide ou déjà utilisé.", "error")
        return redirect(url_for("parametres"))
    if u.token_changement_expiry and datetime.utcnow() > u.token_changement_expiry:
        flash("Ce lien a expiré (24h). Refaites la demande.", "error")
        return redirect(url_for("changer_email"))
    ancien_email = u.email
    u.email                     = u.nouvel_email_en_attente
    u.nouvel_email_en_attente   = None
    u.token_changement_email    = None
    u.token_changement_expiry   = None
    db.session.commit()
    flash(f"✅ Email mis à jour ({ancien_email} → {u.email}). Reconnectez-vous.", "success")
    logout_user()
    session.clear()
    return redirect(url_for("login"))

# ── Super-Admin ───────────────────────────────────────────────────────────────
@app.route("/admin")
@super_admin_required
def admin_dashboard():
    from datetime import timedelta
    now     = datetime.utcnow()
    MOIS_FR = ["","Jan","Fév","Mar","Avr","Mai","Jun","Jul","Aoû","Sep","Oct","Nov","Déc"]

    # ── Tous les tenants (base de tout) ─────────────────────────────────────
    tenants      = Tenant.query.options(joinedload(Tenant.plan)).order_by(Tenant.date_inscription.desc()).all()
    nb_tenants   = len(tenants)
    nb_actifs    = sum(1 for t in tenants if t.statut == "ACTIF")
    nb_essai     = sum(1 for t in tenants if t.statut == "ESSAI")
    nb_suspendus = sum(1 for t in tenants if t.statut == "SUSPENDU")

    # ── MRR / ARR ────────────────────────────────────────────────────────────
    mrr = sum((float(t.plan.prix_mensuel) if t.plan else 0) for t in tenants if t.statut == "ACTIF")
    arr = round(mrr * 12)

    # ── Croissance MoM (mois en cours vs mois précédent) ────────────────────
    debut_mois_c = datetime(now.year, now.month, 1)
    debut_mois_p = (debut_mois_c.replace(day=1) - timedelta(days=1)).replace(day=1)
    inscrits_ce_mois  = sum(1 for t in tenants if t.date_inscription and t.date_inscription >= debut_mois_c)
    inscrits_mois_prec= sum(1 for t in tenants if t.date_inscription and debut_mois_p <= t.date_inscription < debut_mois_c)
    croissance_pct    = round(((inscrits_ce_mois - inscrits_mois_prec) / max(inscrits_mois_prec, 1)) * 100)

    # ── Churn rate (suspendus / total actifs + suspendus) ────────────────────
    churn_rate = round(nb_suspendus / max(nb_actifs + nb_suspendus, 1) * 100, 1)

    # ── ARPU (revenu moyen par client actif) ─────────────────────────────────
    arpu = round(mrr / max(nb_actifs, 1))

    # ── Taux conversion ──────────────────────────────────────────────────────
    taux_conversion = round((nb_actifs / max(nb_tenants, 1)) * 100)

    # ── Essais urgents (< 7 jours) ───────────────────────────────────────────
    essais_urgents = [
        t for t in tenants
        if t.statut == "ESSAI" and t.date_expiration and t.date_expiration <= now + timedelta(days=7)
    ]

    # ── Activité totale ──────────────────────────────────────────────────────
    total_sal = db.session.query(db.func.count(Salarie.id)).scalar() or 0
    total_bul = db.session.query(db.func.count(BulletinPaie.id)).scalar() or 0
    total_ptg = db.session.query(db.func.count(Pointage.id)).scalar() or 0

    # ── Score d'activité par tenant (bulletins + pointages ce mois) ─────────
    sal_par_tenant     = {}
    bul_par_tenant     = {}
    score_par_tenant   = {}
    for t in tenants:
        nb_s = Salarie.query.filter_by(tenant_id=t.id, statut="ACTIF").count()
        nb_b = BulletinPaie.query.filter_by(tenant_id=t.id).count()
        nb_b_mois = BulletinPaie.query.join(PeriodePaie).filter(
            PeriodePaie.tenant_id == t.id,
            PeriodePaie.annee == now.year, PeriodePaie.mois == now.month
        ).count()
        nb_p_mois = Pointage.query.filter_by(tenant_id=t.id).filter(
            Pointage.date_pointage >= debut_mois_c.date()
        ).count()
        sal_par_tenant[t.id]   = nb_s
        bul_par_tenant[t.id]   = nb_b
        # Score 0-100 : bulletins ce mois (50%) + pointages ce mois (30%) + salariés (20%)
        score = min(100, int(
            min(nb_b_mois / max(nb_s, 1) * 50, 50) +
            min(nb_p_mois / max(nb_s * 20, 1) * 30, 30) +
            min(nb_s / 10 * 20, 20)
        ))
        score_par_tenant[t.id] = score

    # ── Revenus 6 mois glissants + nb inscriptions ───────────────────────────
    revenus_6mois     = []
    inscrits_6mois    = []
    bul_6mois         = []
    mois_labels       = []
    for i in range(5, -1, -1):
        d     = now - timedelta(days=i * 30)
        debut = datetime(d.year, d.month, 1)
        fin   = (debut + timedelta(days=32)).replace(day=1)
        mois_labels.append(f"{MOIS_FR[d.month]} {str(d.year)[2:]}")
        rev = sum(
            (float(t.plan.prix_mensuel) if t.plan else 0)
            for t in tenants
            if t.statut == "ACTIF" and t.date_inscription and t.date_inscription <= fin
        )
        revenus_6mois.append(int(rev))
        inscrits_6mois.append(sum(
            1 for t in tenants if t.date_inscription and debut <= t.date_inscription < fin
        ))
        nb_bul = BulletinPaie.query.join(PeriodePaie).filter(
            PeriodePaie.annee == d.year, PeriodePaie.mois == d.month
        ).count()
        bul_6mois.append(nb_bul)

    # ── Répartition par plan ─────────────────────────────────────────────────
    plans_tous = Plan.query.filter_by(actif=True).all()
    repartition_plans = [
        {"nom": p.nom, "nb": sum(1 for t in tenants if t.plan_id == p.id and t.statut == "ACTIF"),
         "prix": float(p.prix_mensuel)}
        for p in plans_tous
        if sum(1 for t in tenants if t.plan_id == p.id and t.statut == "ACTIF") > 0
    ]

    # ── Top tenants les plus actifs ───────────────────────────────────────────
    top_tenants = sorted(
        [t for t in tenants if t.statut == "ACTIF"],
        key=lambda t: score_par_tenant.get(t.id, 0),
        reverse=True
    )[:5]

    # ── Nouvelles inscriptions ce mois ───────────────────────────────────────
    nouveaux_ce_mois = [t for t in tenants if t.date_inscription and t.date_inscription >= debut_mois_c]

    return render_template("admin/dashboard.html",
        tenants=tenants, nb_tenants=nb_tenants,
        nb_actifs=nb_actifs, nb_essai=nb_essai, nb_suspendus=nb_suspendus,
        mrr=mrr, arr=arr, arpu=arpu, churn_rate=churn_rate,
        croissance_pct=croissance_pct, inscrits_ce_mois=inscrits_ce_mois,
        taux_conversion=taux_conversion,
        total_sal=total_sal, total_bul=total_bul, total_ptg=total_ptg,
        essais_urgents=essais_urgents,
        revenus_6mois=revenus_6mois, inscrits_6mois=inscrits_6mois,
        mois_labels=mois_labels, repartition_plans=repartition_plans,
        bul_6mois=bul_6mois, bul_labels=mois_labels,
        sal_par_tenant=sal_par_tenant, bul_par_tenant=bul_par_tenant,
        score_par_tenant=score_par_tenant, top_tenants=top_tenants,
        nouveaux_ce_mois=nouveaux_ce_mois,
        bulletins_par_tenant=bul_par_tenant,
        now=now)

@app.route("/admin/tenants")
@super_admin_required
def admin_tenants():
    q=request.args.get("q",""); statut=request.args.get("statut","")
    query=Tenant.query
    if q: query=query.filter(Tenant.denomination.ilike(f"%{q}%"))
    if statut: query=query.filter_by(statut=statut)
    return render_template("admin/tenants.html", tenants=query.order_by(Tenant.date_inscription.desc()).all(),
        plans=Plan.query.all(), q=q, statut=statut)

@app.route("/admin/tenants/<int:id>")
@super_admin_required
def admin_tenant_detail(id):
    t = Tenant.query.get_or_404(id)
    # Dernières périodes pour l'historique activité
    periodes_recentes = PeriodePaie.query.filter_by(tenant_id=id)        .order_by(PeriodePaie.annee.desc(), PeriodePaie.mois.desc()).limit(6).all()
    nb_journaliers = Journalier.query.filter_by(tenant_id=id).count()
    return render_template("admin/tenant_detail.html", tenant=t,
        nb_salaries=Salarie.query.filter_by(tenant_id=id).count(),
        nb_bulletins=BulletinPaie.query.filter_by(tenant_id=id).count(),
        users=Utilisateur.query.filter_by(tenant_id=id).all(),
        plans=Plan.query.all(),
        periodes_recentes=periodes_recentes,
        nb_journaliers=nb_journaliers,
        now=datetime.utcnow())

@app.route("/admin/tenants/<int:id>/statut", methods=["POST"])
@super_admin_required
def admin_tenant_statut(id):
    t=Tenant.query.get_or_404(id)
    t.statut=request.form.get("statut",t.statut)
    if request.form.get("plan_id"): t.plan_id=int(request.form["plan_id"])
    # Date d'expiration essai
    date_exp = request.form.get("date_expiration","").strip()
    if date_exp:
        try:
            from datetime import datetime as _dt
            t.date_expiration = _dt.strptime(date_exp, "%Y-%m-%d")
        except: pass
    elif request.form.get("clear_expiration"):
        t.date_expiration = None
    db.session.commit(); flash(f"{t.denomination} mis à jour.","success")
    return redirect(url_for("admin_tenant_detail",id=id))

@app.route("/admin/tenants/<int:id>/notes", methods=["POST"])
@super_admin_required
def admin_tenant_notes(id):
    t = Tenant.query.get_or_404(id)
    t.notes = request.form.get("notes", "").strip()
    db.session.commit()
    flash("Notes mises à jour.", "success")
    return redirect(url_for("admin_tenant_detail", id=id))


@app.route("/admin/tenants/<int:id>/regenerer-token", methods=["POST"])
@super_admin_required
def admin_regenerer_token(id):
    """Génère ou régénère le token API d'un tenant."""
    t = Tenant.query.get_or_404(id)
    ancien = t.token_api
    t.token_api = sec.token_hex(32)
    db.session.commit()
    action = "généré" if not ancien else "régénéré"
    flash(f"Token API {action} pour {t.denomination}. "
          f"Transmettez-le de façon sécurisée au développeur RH du client.", "success")
    logger.info(f"[SuperAdmin] Token API {action} — tenant={t.id} par {current_user.email}")
    return redirect(url_for("admin_tenant_detail", id=id))

@app.route("/admin/tenants/<int:id>/impersonate")
@super_admin_required
def admin_impersonate(id):
    u=Utilisateur.query.filter_by(tenant_id=id,role="TENANT_ADMIN").first()
    if not u: flash("Aucun admin trouvé.","error"); return redirect(url_for("admin_tenant_detail",id=id))
    logout_user(); login_user(u)
    flash(f"Connecté en tant que {u.nom_complet} ({u.tenant.denomination})","warning")
    return redirect(url_for("dashboard"))

@app.route("/admin/plans", methods=["GET","POST"])
@super_admin_required
def admin_plans():
    if request.method=="POST":
        p=Plan(code=request.form["code"].strip().upper(),
               nom=request.form["nom"].strip(),
               prix_mensuel=float(request.form.get("prix_mensuel",0)),
               max_salaries=int(request.form["max_salaries"]) if request.form.get("max_salaries") else None,
               max_utilisateurs=int(request.form["max_utilisateurs"]) if request.form.get("max_utilisateurs") else None,
               description=request.form.get("description",""))
        db.session.add(p); db.session.commit(); flash("Plan créé avec succès.","success")
    plans = Plan.query.order_by(Plan.prix_mensuel.asc()).all()
    # Nb clients actifs + revenus par plan
    stats_plans = {}
    for p in plans:
        nb_actifs  = Tenant.query.filter_by(plan_id=p.id, statut="ACTIF").count()
        nb_essai   = Tenant.query.filter_by(plan_id=p.id, statut="ESSAI").count()
        nb_total   = Tenant.query.filter_by(plan_id=p.id).count()
        revenus    = nb_actifs * float(p.prix_mensuel)
        stats_plans[p.id] = {"nb_actifs": nb_actifs, "nb_essai": nb_essai,
                              "nb_total": nb_total, "revenus": revenus}
    return render_template("admin/plans.html", plans=plans, stats_plans=stats_plans)

@app.route("/admin/plans/<int:id>/modifier", methods=["POST"])
@super_admin_required
def admin_plan_modifier(id):
    p = Plan.query.get_or_404(id)
    p.nom          = request.form.get("nom", p.nom).strip()
    p.prix_mensuel = float(request.form.get("prix_mensuel", p.prix_mensuel))
    p.max_salaries     = int(request.form["max_salaries"]) if request.form.get("max_salaries") else None
    p.max_utilisateurs = int(request.form["max_utilisateurs"]) if request.form.get("max_utilisateurs") else None
    p.description  = request.form.get("description", p.description or "")
    db.session.commit()
    flash(f"Plan « {p.nom} » modifié avec succès.", "success")
    return redirect(url_for("admin_plans"))

@app.route("/admin/plans/<int:id>/toggle", methods=["POST"])
@super_admin_required
def admin_plan_toggle(id):
    p = Plan.query.get_or_404(id)
    p.actif = not p.actif
    db.session.commit()
    etat = "activé" if p.actif else "désactivé"
    flash(f"Plan « {p.nom} » {etat}.", "success")
    return redirect(url_for("admin_plans"))

@app.route("/admin/plans/<int:id>/supprimer", methods=["POST"])
@super_admin_required
def admin_plan_supprimer(id):
    p = Plan.query.get_or_404(id)
    nb_clients = Tenant.query.filter_by(plan_id=id).count()
    if nb_clients > 0:
        flash(f"Impossible de supprimer : {nb_clients} entreprise(s) utilisent ce plan.", "error")
        return redirect(url_for("admin_plans"))
    nom = p.nom
    db.session.delete(p); db.session.commit()
    flash(f"Plan « {nom} » supprimé.", "success")
    return redirect(url_for("admin_plans"))

@app.route("/admin/stats")
@super_admin_required
def admin_stats():
    tenants = Tenant.query.order_by(Tenant.denomination).all()
    stats = []
    total_revenus  = 0
    total_salaries = 0
    total_bulletins= 0
    total_journaliers = 0
    nb_actif = nb_essai = nb_suspendu = nb_attente = 0

    for t in tenants:
        nb_sal = Salarie.query.filter_by(tenant_id=t.id, statut="ACTIF").count()
        nb_bul = BulletinPaie.query.filter_by(tenant_id=t.id).count()
        nb_per = PeriodePaie.query.filter_by(tenant_id=t.id).count()
        nb_jou = Journalier.query.filter_by(tenant_id=t.id).count()
        rev    = float(t.plan.prix_mensuel) if t.plan and t.statut == "ACTIF" else 0
        total_revenus   += rev
        total_salaries  += nb_sal
        total_bulletins += nb_bul
        total_journaliers += nb_jou
        if   t.statut == "ACTIF":                nb_actif   += 1
        elif t.statut == "ESSAI":                nb_essai   += 1
        elif t.statut == "SUSPENDU":             nb_suspendu+= 1
        elif t.statut == "PAIEMENT_EN_ATTENTE":  nb_attente += 1
        stats.append({
            "tenant": t, "nb_salaries": nb_sal, "nb_bulletins": nb_bul,
            "nb_periodes": nb_per, "nb_journaliers": nb_jou, "revenus": rev
        })

    # Top 8 pour graphique — trié en Python
    top8 = sorted(stats, key=lambda x: x["nb_bulletins"], reverse=True)[:8]
    top8_labels = [s["tenant"].denomination[:18] for s in top8]
    top8_data   = [s["nb_bulletins"] for s in top8]

    taux_conv = round(nb_actif / len(tenants) * 100) if tenants else 0

    return render_template("admin/stats.html",
        stats=stats,
        total_revenus=total_revenus,
        total_salaries=total_salaries,
        total_bulletins=total_bulletins,
        total_journaliers=total_journaliers,
        nb_actifs=nb_actif,
        nb_essai=nb_essai,
        nb_suspendu=nb_suspendu,
        nb_attente=nb_attente,
        taux_conv=taux_conv,
        top8_labels=top8_labels,
        top8_data=top8_data,
        now=datetime.utcnow())

@app.route("/admin/import", methods=["GET","POST"])
@login_required
@super_admin_required
def admin_import_excel():
    from werkzeug.utils import secure_filename
    tenants = Tenant.query.order_by(Tenant.denomination).all()
    resultats = None
    if request.method == "POST":
        tenant_id = request.form.get("tenant_id", type=int)
        tenant = Tenant.query.get_or_404(tenant_id)
        f = request.files.get("excel_file")
        if not f or not f.filename.endswith(".xlsx"):
            flash("Fichier invalide. Utilisez un fichier .xlsx", "error")
            return render_template("admin/import_excel.html", tenants=tenants)
        imp_societe  = "import_societe"  in request.form
        imp_salaries = "import_salaries" in request.form
        imp_bulletins= "import_bulletins"in request.form
        ecraser      = "ecraser"         in request.form
        try:
            from openpyxl import load_workbook
            from datetime import datetime as dt2, date as d2
            wb = load_workbook(io.BytesIO(f.read()), read_only=True, data_only=True)
            nb_sal=nb_bul=nb_per=nb_err=0
            def nv(v,d=0):
                try: return float(v) if v is not None else d
                except: return d
            def dv(v):
                if v is None: return None
                if isinstance(v,dt2): return v.date()
                if isinstance(v,d2): return v
                try: return dt2.strptime(str(v)[:10],"%Y-%m-%d").date()
                except: return None
            if imp_societe and "INFOS SOCIETE" in wb.sheetnames:
                ws=wb["INFOS SOCIETE"]; infos={}
                for row in ws.iter_rows(values_only=True):
                    if row[1] and row[2]: infos[str(row[1]).strip().upper()]=row[2]
                tenant.denomination=str(infos.get("DENOMINATION SOCIALE",tenant.denomination)).strip().upper()
                tenant.sigle=str(infos.get("SIGLE",tenant.sigle or "")).strip()
                tenant.activite=str(infos.get("ACTIVITE",tenant.activite or "")).strip()
                tenant.nif=str(infos.get("NIF",tenant.nif or "")).strip()
                tenant.adresse=str(infos.get("ADRESSE",tenant.adresse or "")).strip()
                tenant.ville=str(infos.get("VILLE",tenant.ville or "Libreville")).strip()
            cats={}
            for code,lib in [("C1","Ouvriers"),("C2","Techniciens"),("C3","Conducteurs"),("C4","Cadres")]:
                cat=CategorieEmploi.query.filter_by(tenant_id=tenant.id,code=code).first()
                if not cat:
                    cat=CategorieEmploi(tenant_id=tenant.id,code=code,libelle=lib)
                    db.session.add(cat); db.session.flush()
                cats[code]=cat
            salaries_map={}
            if imp_salaries and "INFOS SALARIES" in wb.sheetnames:
                ws=wb["INFOS SALARIES"]; header=None
                for row in ws.iter_rows(values_only=True):
                    if not any(v is not None for v in row): continue
                    if header is None: header=row; continue
                    if row[0] is None: continue
                    matricule=str(row[0]).strip().upper()
                    if not matricule: continue
                    sal=Salarie.query.filter_by(tenant_id=tenant.id,matricule=matricule).first()
                    if sal and not ecraser: salaries_map[matricule]=sal; continue
                    if not sal: sal=Salarie(tenant_id=tenant.id,matricule=matricule); db.session.add(sal)
                    sal.nom=str(row[1]).strip().upper() if row[1] else "—"
                    sal.prenom=str(row[2]).strip() if row[2] else "—"
                    sal.telephone=str(row[3]).strip() if row[3] else None
                    sal.nationalite=str(row[5]).strip().upper() if row[5] else "GABONAISE"
                    sal.sexe=str(row[6]).strip().upper() if row[6] else "M"
                    sal.date_naissance=dv(row[7])
                    sal.date_embauche=dv(row[9]) or d2(2024,8,1)
                    sal.date_cessation=dv(row[10])
                    sal.situation_matrimoniale=str(row[11]).strip().upper() if row[11] else None
                    sal.nb_enfants=int(row[12]) if row[12] and str(row[12]).replace('.','').isdigit() else 0
                    sal.nombre_parts=float(row[13]) if row[13] else 1.0
                    sal.numero_cnss=str(row[14]).strip() if row[14] else None
                    sal.numero_cnamgs=str(row[15]).strip() if row[15] else None
                    sal.emploi=str(row[16]).strip().upper() if row[16] else None
                    sal.nb_enfants_moins_16ans=int(row[18]) if row[18] and str(row[18]).replace('.','').isdigit() else 0
                    sal.assujetti_cnamgs=str(row[19]).strip().upper()=="OUI" if row[19] else True
                    sal.statut="INACTIF" if dv(row[10]) else "ACTIF"
                    cat_code=str(row[17]).strip().upper() if row[17] else "C1"
                    sal.categorie_id=cats.get(cat_code,cats.get("C1")).id
                    db.session.flush(); salaries_map[matricule]=sal; nb_sal+=1
            if not imp_salaries:
                for s in Salarie.query.filter_by(tenant_id=tenant.id).all():
                    salaries_map[s.matricule]=s
            MOIS={"JANVIER":1,"FÉVRIER":2,"FEVRIER":2,"MARS":3,"AVRIL":4,"MAI":5,"JUIN":6,
                  "JUILLET":7,"AOÛT":8,"AOUT":8,"SEPTEMBRE":9,"OCTOBRE":10,"NOVEMBRE":11,"DÉCEMBRE":12,"DECEMBRE":12}
            periodes_cache={}
            if imp_bulletins and "DONNEES DU BULLETIN" in wb.sheetnames:
                ws=wb["DONNEES DU BULLETIN"]
                for i,row in enumerate(ws.iter_rows(values_only=True)):
                    if i==0: continue
                    if not any(v is not None for v in row): continue
                    if row[4] is None: continue
                    matricule=str(row[4]).strip().upper().replace(' - ','-').replace('- ','-').replace(' -','-')
                    if matricule not in salaries_map: nb_err+=1; continue
                    annee=int(row[1]) if row[1] else None
                    mois_str=str(row[2]).strip().upper() if row[2] else ""
                    mois=MOIS.get(mois_str)
                    if not annee or not mois: nb_err+=1; continue
                    pk=f"{annee}-{mois}"
                    if pk not in periodes_cache:
                        p=PeriodePaie.query.filter_by(tenant_id=tenant.id,annee=annee,mois=mois).first()
                        if not p:
                            mois_nom=[k for k,v in MOIS.items() if v==mois and len(k)>4][0]
                            p=PeriodePaie(tenant_id=tenant.id,annee=annee,mois=mois,
                                libelle_mois=mois_nom,trimestre=f"T{(mois-1)//3+1}",statut="CLÔTURÉ")
                            db.session.add(p); db.session.flush(); nb_per+=1
                        periodes_cache[pk]=p
                    sal=salaries_map[matricule]
                    bul=BulletinPaie.query.filter_by(tenant_id=tenant.id,salarie_id=sal.id,periode_id=periodes_cache[pk].id).first()
                    if bul and not ecraser: continue
                    if not bul: bul=BulletinPaie(tenant_id=tenant.id,salarie_id=sal.id,periode_id=periodes_cache[pk].id); db.session.add(bul)
                    bul.nb_jours_travailles=int(nv(row[42]))
                    bul.salaire_base=nv(row[5]); bul.heures_sup_10=nv(row[7]); bul.heures_sup_30=nv(row[9])
                    bul.heures_sup_40=nv(row[11]); bul.heures_sup_70=nv(row[13]); bul.absences=nv(row[15])
                    bul.sursalaire=nv(row[17]); bul.prime_caisse=nv(row[19]); bul.carburant=nv(row[21])
                    bul.prime_anciennete=nv(row[23]); bul.indem_logement=nv(row[25])
                    bul.indem_domesticite=nv(row[26]); bul.indem_eau_electricite=nv(row[27])
                    bul.indem_nourriture=nv(row[28]); bul.prime_rendement=nv(row[29])
                    bul.prime_assiduité=nv(row[31]); bul.prime_qualite=nv(row[33])
                    bul.prime_performance=nv(row[35]); bul.prime_transport=nv(row[37])
                    bul.prime_responsabilite=nv(row[39]); bul.allocations_conge=nv(row[41])
                    bul.salaire_brut=nv(row[53]); bul.base_cnss=nv(row[54])
                    bul.cnss_salarie=nv(row[55]); bul.cnss_patronale=nv(row[56])
                    bul.base_cnamgs=nv(row[59]); bul.cnamgs_salarie=nv(row[60]); bul.cnamgs_patronale=nv(row[61])
                    bul.fnh=nv(row[62]); bul.cfp=nv(row[63]); bul.base_tcs=nv(row[72]); bul.tcs=nv(row[73])
                    bul.net_avant_irpp=nv(row[74]); bul.base_irpp=nv(row[75]); bul.irpp=nv(row[76])
                    bul.salaire_net=nv(row[77]); bul.prime_panier=nv(row[78]); bul.indem_transport=nv(row[79])
                    bul.indem_representation=nv(row[80]); bul.prime_salisure=nv(row[81])
                    bul.acompte=nv(row[82]); bul.net_a_payer=nv(row[83]) if len(row)>83 else 0
                    bul.statut="VALIDÉ"; bul.date_validation=datetime.utcnow(); nb_bul+=1
            db.session.commit()
            resultats={"nb_salaries":nb_sal,"nb_bulletins":nb_bul,"nb_periodes":nb_per,"erreurs":nb_err}
            flash(f"Import réussi ! {nb_sal} salariés, {nb_bul} bulletins, {nb_per} périodes.","success")
        except Exception as e:
            db.session.rollback()
            flash(f"Erreur : {str(e)}","error")
    return render_template("admin/import_excel.html", tenants=tenants, resultats=resultats)

@app.route("/admin/update-taux", methods=["POST"])
@super_admin_required
def admin_update_taux():
    taux = {"CNSS":(0.05,0.18),"CNAMGS":(0.02,0.041),"FNH":(0.0,0.03),"TCS":(0.05,0.0),"CFP":(0.0,0.005)}
    nb = 0
    for code,(sal,pat) in taux.items():
        for r in RubriquePaie.query.filter_by(code=code).all():
            r.taux_salarie=sal; r.taux_patronal=pat; nb+=1
    db.session.commit()
    flash(f"Taux mis a jour ({nb} rubriques).", "success")
    return redirect(url_for("admin_rubriques"))

@app.route("/admin/rubriques", methods=["GET","POST"])
@super_admin_required
def admin_rubriques():
    if request.method=="POST":
        r=RubriquePaie(code=request.form["code"].strip().upper(),
            libelle=request.form["libelle"].strip(), type=request.form.get("type","COTISATION"),
            taux_salarie=float(request.form["taux_salarie"]) if request.form.get("taux_salarie") else None,
            taux_patronal=float(request.form["taux_patronal"]) if request.form.get("taux_patronal") else None,
            plafond_mensuel=float(request.form["plafond_mensuel"]) if request.form.get("plafond_mensuel") else None)
        db.session.add(r); db.session.commit(); flash("Rubrique créée.","success")
    return render_template("admin/rubriques.html", rubriques=RubriquePaie.query.all())

@app.route("/admin/tenants/<int:id>/supprimer", methods=["POST"])
@super_admin_required
def admin_tenant_supprimer(id):
    t = Tenant.query.get_or_404(id)
    nom = t.denomination
    try:
        # ── 1. Pointages journaliers (FK bloquante) ───────────────────────
        for j in Journalier.query.filter_by(tenant_id=id).all():
            Pointage.query.filter_by(journalier_id=j.id).delete()
            FeuillePaieJournalier.query.filter_by(journalier_id=j.id).delete()
            AffectationSite.query.filter_by(journalier_id=j.id).delete()
        Journalier.query.filter_by(tenant_id=id).delete()

        # ── 2. Données salariés ────────────────────────────────────────────
        for s in Salarie.query.filter_by(tenant_id=id).all():
            BulletinPaie.query.filter_by(salarie_id=s.id).delete()
            Contrat.query.filter_by(salarie_id=s.id).delete()
            Pointage.query.filter_by(salarie_id=s.id).delete()
            Acompte.query.filter_by(salarie_id=s.id).delete()
            Conge.query.filter_by(salarie_id=s.id).delete()
            AffectationSite.query.filter_by(salarie_id=s.id).delete()
        Salarie.query.filter_by(tenant_id=id).delete()

        # ── 3. Reste du tenant ─────────────────────────────────────────────
        AffectationSite.query.filter_by(tenant_id=id).delete()
        from models import Site
        Site.query.filter_by(tenant_id=id).delete()
        PeriodePaie.query.filter_by(tenant_id=id).delete()
        CategorieEmploi.query.filter_by(tenant_id=id).delete()
        Acompte.query.filter_by(tenant_id=id).delete()
        Conge.query.filter_by(tenant_id=id).delete()
        Utilisateur.query.filter_by(tenant_id=id).delete()

        db.session.delete(t)
        db.session.commit()
        flash(f"Entreprise « {nom} » supprimée définitivement.", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"Erreur: {str(e)}", "error")
    return redirect(url_for("admin_tenants"))

# ── Dashboard ─────────────────────────────────────────────────────────────────
@app.route("/dashboard")
@login_required
def dashboard():
    if current_user.is_super_admin: return redirect(url_for("admin_dashboard"))
    t=get_tenant()
    if not t: flash("Aucune entreprise associée.","error"); return redirect(url_for("login"))
    now=datetime.now()

    # ── KPIs emploi (cache TTL=5min) ─────────────────────────────────────────
    _ck_kpis = f"{t.id}:kpis_emploi"
    kpis_cached = _cache_get(_ck_kpis)
    if kpis_cached:
        nb_actifs, nb_inactifs, nb_total, nb_journaliers, nb_new_mois = kpis_cached
    else:
        from sqlalchemy import func
        _sal_q = db.session.query(
            func.sum(db.cast(Salarie.statut == "ACTIF",   db.Integer)).label("actifs"),
            func.sum(db.cast(Salarie.statut == "INACTIF", db.Integer)).label("inactifs"),
            func.count().label("total"),
        ).filter(Salarie.tenant_id == t.id).one()
        nb_actifs   = int(_sal_q.actifs   or 0)
        nb_inactifs = int(_sal_q.inactifs or 0)
        nb_total    = int(_sal_q.total    or 0)
        nb_journaliers = Journalier.query.filter_by(tenant_id=t.id, statut="ACTIF").count()
        debut_mois  = datetime(now.year, now.month, 1).date()
        nb_new_mois = Salarie.query.filter(Salarie.tenant_id==t.id,
                                           Salarie.date_embauche>=debut_mois).count()
        _cache_set(_ck_kpis, (nb_actifs, nb_inactifs, nb_total, nb_journaliers, nb_new_mois),
                   TTL_KPIS_DASH)
    nb_total_employes = nb_actifs + nb_journaliers
    debut_mois  = datetime(now.year, now.month, 1).date()
    periode = PeriodePaie.query.filter_by(tenant_id=t.id, annee=now.year, mois=now.month).first()
    masse={}; nb_v=nb_b=nb_p=0
    if periode:
        buls = BulletinPaie.query.filter_by(tenant_id=t.id, periode_id=periode.id).all()
        masse = calculer_masse_salariale(buls)
        nb_v = sum(1 for b in buls if b.statut=="VALIDÉ")
        nb_p = sum(1 for b in buls if b.statut=="PAYÉ")
        nb_b = sum(1 for b in buls if b.statut=="BROUILLON")
    _ck_evo = f"{t.id}:evolution_{now.year}_{now.month}"
    evolution = _cache_get(_ck_evo)
    if evolution is None:
        evolution = []
        mois_noms = ["","Jan","Fév","Mar","Avr","Mai","Jun","Jul","Aoû","Sep","Oct","Nov","Déc"]
        for i in range(5, -1, -1):
            m = now.month - i; y = now.year
            while m <= 0: m += 12; y -= 1
            p = PeriodePaie.query.filter_by(tenant_id=t.id, annee=y, mois=m).first()
            if p:
                buls_p = BulletinPaie.query.filter_by(tenant_id=t.id, periode_id=p.id).all()
                total_net    = sum(float(b.net_a_payer    or 0) for b in buls_p)
                total_brut   = sum(float(b.salaire_brut   or 0) for b in buls_p)
                total_charges= sum(float(b.cnss_patronale or 0)+float(b.cnamgs_patronale or 0)
                                   +float(b.fnh or 0)+float(b.cfp or 0) for b in buls_p)
            else:
                total_net=total_brut=total_charges=0; buls_p=[]
            evolution.append({"mois":mois_noms[m],"annee":y,"brut":round(total_brut),
                              "net":round(total_net),"charges":round(total_charges),
                              "nb_bulletins":len(buls_p)})
        _cache_set(_ck_evo, evolution, TTL_EVOLUTION)
    top_salaries = []
    if periode:
        top_salaries = BulletinPaie.query.filter_by(tenant_id=t.id, periode_id=periode.id).order_by(BulletinPaie.net_a_payer.desc()).limit(5).all()
    from sqlalchemy import func
    _ck_cats = f"{t.id}:cats_stats"
    cats_stats = _cache_get(_ck_cats)
    if cats_stats is None:
        cats_stats = db.session.query(
            CategorieEmploi.code, CategorieEmploi.libelle,
            func.count(Salarie.id).label("nb")
        ).join(Salarie, Salarie.categorie_id==CategorieEmploi.id)\
         .filter(Salarie.tenant_id==t.id, Salarie.statut=="ACTIF")\
         .group_by(CategorieEmploi.code, CategorieEmploi.libelle).all()
        _cache_set(_ck_cats, cats_stats, TTL_CATS_STATS)
    derniers = BulletinPaie.query.filter_by(tenant_id=t.id).order_by(BulletinPaie.date_creation.desc()).limit(6).all()
    # ══════════════════════════════════════════════════════════════════════════
    # ALERTES INTELLIGENTES
    # ══════════════════════════════════════════════════════════════════════════
    import calendar
    alertes = []
    MOIS_NOMS_LONG = ["","Janvier","Février","Mars","Avril","Mai","Juin",
                      "Juillet","Août","Septembre","Octobre","Novembre","Décembre"]
    mois_courant = MOIS_NOMS_LONG[now.month]
    debut_mois_d = date(now.year, now.month, 1)
    fin_mois_d   = date(now.year, now.month, calendar.monthrange(now.year, now.month)[1])

    # ── 1. Quota employés dépassé ou proche ──────────────────────────────────
    q_info = t.quota_employes_info
    if q_info.get("max"):
        pct = round(q_info["total"] / q_info["max"] * 100)
        if q_info["plein"]:
            alertes.append({"type":"danger","icone":"🚫","titre":"Limite d'employés atteinte",
                "msg":f"Vous avez atteint la limite de {q_info['max']} employés de votre plan {t.plan.nom}. "
                      f"Impossible d'ajouter de nouveaux travailleurs.",
                "lien":"/parametres","lien_texte":"Changer de plan"})
        elif pct >= 80:
            alertes.append({"type":"warning","icone":"⚠️","titre":"Quota employés bientôt atteint",
                "msg":f"{q_info['total']}/{q_info['max']} employés utilisés ({pct}%). "
                      f"Il reste {q_info['max'] - q_info['total']} place(s).",
                "lien":"/parametres","lien_texte":"Voir le plan"})

    # ── 2. Période non créée pour le mois courant ─────────────────────────────
    if not periode:
        alertes.append({"type":"warning","icone":"📅","titre":f"Période {mois_courant} {now.year} manquante",
            "msg":f"Aucune période de paie n'est ouverte pour {mois_courant} {now.year}. "
                  f"Les bulletins ne peuvent pas être saisis.",
            "lien":"/periodes","lien_texte":"Créer la période"})

    # ── 3. Période précédente non clôturée ────────────────────────────────────
    mois_prec = now.month - 1 or 12
    annee_prec = now.year if now.month > 1 else now.year - 1
    periode_prec = PeriodePaie.query.filter_by(
        tenant_id=t.id, annee=annee_prec, mois=mois_prec).first()
    if periode_prec and periode_prec.statut not in ("CLÔTURÉE", "CLOTUREE", "PAYÉE"):
        buls_prec = BulletinPaie.query.filter_by(
            tenant_id=t.id, periode_id=periode_prec.id).all()
        nb_non_clos = sum(1 for b in buls_prec if b.statut in ("BROUILLON", "VALIDÉ"))
        if nb_non_clos > 0:
            alertes.append({"type":"warning","icone":"🔓","titre":f"Période {MOIS_NOMS_LONG[mois_prec]} non clôturée",
                "msg":f"{nb_non_clos} bulletin(s) de {MOIS_NOMS_LONG[mois_prec]} {annee_prec} "
                      f"ne sont pas encore payés.",
                "lien":f"/bulletins?periode_id={periode_prec.id}","lien_texte":"Voir les bulletins"})

    # ── 4. Bulletins en brouillon ce mois ─────────────────────────────────────
    if nb_b > 0:
        alertes.append({"type":"warning","icone":"📝","titre":f"{nb_b} brouillon(s) à valider",
            "msg":f"{nb_b} bulletin(s) de {mois_courant} sont en brouillon et doivent être validés.",
            "lien":f"/bulletins?periode_id={periode.id}&statut=BROUILLON" if periode else "/bulletins",
            "lien_texte":"Valider maintenant"})

    # ── 5. Salariés actifs sans bulletin ce mois ──────────────────────────────
    if periode:
        ids_avec_bulletin = {b.salarie_id for b in
            BulletinPaie.query.filter_by(tenant_id=t.id, periode_id=periode.id).all()}
        salaries_sans_bulletin = Salarie.query.filter_by(
            tenant_id=t.id, statut="ACTIF"
        ).filter(~Salarie.id.in_(ids_avec_bulletin)).all() if ids_avec_bulletin else             Salarie.query.filter_by(tenant_id=t.id, statut="ACTIF").all()
        nb_sans = len(salaries_sans_bulletin)
        if nb_sans > 0:
            exemples = ", ".join(s.nom_complet for s in salaries_sans_bulletin[:3])
            if nb_sans > 3: exemples += f" et {nb_sans-3} autre(s)"
            alertes.append({"type":"info","icone":"👤","titre":f"{nb_sans} salarié(s) sans bulletin ce mois",
                "msg":f"{exemples}.",
                "lien":f"/bulletins/saisie","lien_texte":"Saisir un bulletin"})

    # ── 6. Journaliers avec feuilles en attente de paiement ───────────────────
    feuilles_att = FeuillePaieJournalier.query.filter_by(
        tenant_id=t.id, statut="EN_ATTENTE").all()
    nb_feuilles_att = len(feuilles_att)
    if nb_feuilles_att > 0:
        montant_att = sum(float(f.montant_brut or 0) for f in feuilles_att)
        alertes.append({"type":"warning","icone":"🦺","titre":f"{nb_feuilles_att} feuille(s) journaliers non payée(s)",
            "msg":f"Total en attente : {int(montant_att):,} FCFA.",
            "lien":"/journaliers/paie","lien_texte":"Payer maintenant"})

    # ── 7. Journaliers actifs sans pointage cette semaine ─────────────────────
    lundi = (now.date() - timedelta(days=now.weekday()))
    samedi = lundi + timedelta(days=5)
    nb_jour_actifs = Journalier.query.filter_by(tenant_id=t.id, statut="ACTIF").count()
    if nb_jour_actifs > 0:
        ids_pointes_sem = {p.journalier_id for p in
            Pointage.query.filter_by(tenant_id=t.id)
            .filter(Pointage.journalier_id.isnot(None),
                    Pointage.date_pointage >= lundi,
                    Pointage.date_pointage <= samedi).all()}
        nb_non_pointes_jour = nb_jour_actifs - len(ids_pointes_sem)
        if nb_non_pointes_jour > 0 and now.weekday() >= 1:   # après lundi
            alertes.append({"type":"info","icone":"📋","titre":f"{nb_non_pointes_jour} journalier(s) non pointé(s) cette semaine",
                "msg":f"Semaine du {lundi.strftime('%d/%m')} : {nb_non_pointes_jour} journalier(s) "
                      f"sans pointage enregistré.",
                "lien":f"/pointage/recap-semaine","lien_texte":"Voir le récap"})

    # ── 8. Acomptes en attente de déduction ───────────────────────────────────
    acomptes_att = Acompte.query.filter_by(tenant_id=t.id, statut="EN_ATTENTE").count()
    if acomptes_att > 0:
        alertes.append({"type":"info","icone":"💸","titre":f"{acomptes_att} acompte(s) en attente",
            "msg":f"{acomptes_att} acompte(s) n'ont pas encore été déduits des bulletins.",
            "lien":"/acomptes","lien_texte":"Voir les acomptes"})

    # ── 9. Statut abonnement/essai — TOUJOURS affiché ───────────────────────
    exp_date = t.date_expiration
    jours_restants = (exp_date.date() - now.date()).days if exp_date else None

    if t.statut == "ESSAI":
        if jours_restants is None or jours_restants <= 0:
            alertes.append({"type":"danger","icone":"🚫",
                "titre":"Période d'essai expirée",
                "msg":"Votre essai gratuit a expiré. Souscrivez maintenant pour continuer.",
                "lien":"/parametres","lien_texte":"Souscrire maintenant"})
        elif jours_restants <= 7:
            alertes.append({"type":"danger","icone":"⏰",
                "titre":f"Essai : {jours_restants} jour(s) restant(s)",
                "msg":f"Expire le {exp_date.strftime('%d/%m/%Y')}. Souscrivez pour ne pas perdre vos données.",
                "lien":"/parametres","lien_texte":"Souscrire maintenant"})
        elif jours_restants <= 14:
            alertes.append({"type":"warning","icone":"⏳",
                "titre":f"Essai : {jours_restants} jour(s) restant(s)",
                "msg":f"Votre période d'essai se termine le {exp_date.strftime('%d/%m/%Y')}.",
                "lien":"/parametres","lien_texte":"Voir les plans"})
        else:
            alertes.append({"type":"info","icone":"🧪",
                "titre":f"Période d'essai — {jours_restants} jour(s) restant(s)",
                "msg":f"Essai gratuit jusqu'au {exp_date.strftime('%d/%m/%Y')}. Plan actuel : {t.plan.nom}.",
                "lien":"/parametres","lien_texte":"Voir les plans"})

    elif t.statut == "ACTIF":
        if jours_restants is None or jours_restants <= 0:
            alertes.append({"type":"danger","icone":"🔒",
                "titre":"Abonnement expiré",
                "msg":"Votre abonnement a expiré. Renouvelez pour continuer.",
                "lien":"/parametres","lien_texte":"Renouveler"})
        elif jours_restants <= 7:
            alertes.append({"type":"danger","icone":"⏰",
                "titre":f"Abonnement : {jours_restants} jour(s) restant(s)",
                "msg":f"Expire le {exp_date.strftime('%d/%m/%Y')}. Renouvelez dès maintenant.",
                "lien":"/parametres","lien_texte":"Renouveler"})
        elif jours_restants <= 30:
            alertes.append({"type":"warning","icone":"📅",
                "titre":f"Abonnement : {jours_restants} jour(s) restant(s)",
                "msg":f"Expire le {exp_date.strftime('%d/%m/%Y')}.",
                "lien":"/parametres","lien_texte":"Renouveler"})
        else:
            alertes.append({"type":"info","icone":"✅",
                "titre":f"Abonnement actif — {jours_restants} jour(s) restant(s)",
                "msg":f"Plan {t.plan.nom} valide jusqu'au {exp_date.strftime('%d/%m/%Y')}.",
                "lien":"/parametres","lien_texte":"Gérer l'abonnement"})

    elif t.statut == "PAIEMENT_EN_ATTENTE":
        alertes.append({"type":"warning","icone":"💳",
            "titre":"Paiement en attente",
            "msg":"Votre paiement est en cours de traitement. L'accès complet sera rétabli dès confirmation.",
            "lien":"/parametres","lien_texte":"Contacter le support"})

    # Trier par priorité : danger > warning > info
    _prio = {"danger": 0, "warning": 1, "info": 2}
    alertes.sort(key=lambda a: _prio.get(a["type"], 3))

    # Compteurs pour l'en-tête
    nb_alertes_critiques = sum(1 for a in alertes if a["type"] == "danger")
    nb_alertes_warning   = sum(1 for a in alertes if a["type"] == "warning")

    return render_template("tenant/dashboard.html", tenant=t,
        nb_actifs=nb_actifs, nb_inactifs=nb_inactifs, nb_total=nb_total,
        nb_journaliers=nb_journaliers, nb_total_employes=nb_total_employes,
        nb_new_mois=nb_new_mois, periode=periode, masse=masse,
        nb_valides=nb_v, nb_payes=nb_p, nb_brouillon=nb_b,
        evolution=evolution, top_salaries=top_salaries,
        cats_stats=cats_stats, derniers=derniers, alertes=alertes,
        nb_alertes_critiques=nb_alertes_critiques,
        nb_alertes_warning=nb_alertes_warning, now=now)

# ── Salariés ──────────────────────────────────────────────────────────────────
@app.route("/salaries")
@login_required
def salaries():
    if current_user.is_super_admin: return redirect(url_for("admin_dashboard"))
    t = get_tenant()
    if not t: return redirect(url_for("login"))
    q      = request.args.get("q", "")
    statut = request.args.get("statut", "")
    page   = request.args.get("page", 1, type=int)
    query  = Salarie.query.filter_by(tenant_id=t.id)
    if q:      query = query.filter(db.or_(Salarie.nom.ilike(f"%{q}%"), Salarie.prenom.ilike(f"%{q}%"), Salarie.matricule.ilike(f"%{q}%")))
    if statut: query = query.filter_by(statut=statut)
    query = query.options(joinedload(Salarie.categorie))
    pagination = query.order_by(Salarie.nom).paginate(page=page, per_page=25, error_out=False)
    _args  = {k: v for k, v in request.args.items() if k != 'page'}
    _base  = request.path + '?' + '&'.join(f'{k}={v}' for k, v in _args.items())
    _sep   = '&' if _args else '?'
    return render_template("tenant/salaries.html",
        salaries=pagination.items, pagination=pagination,
        categories=CategorieEmploi.query.filter_by(tenant_id=t.id).all(),
        q=q, statut=statut, tenant=t,
        pagination_base=_base + _sep)




@app.route("/manifest.json")
def pwa_manifest():
    """Sert le manifest PWA."""
    import json
    manifest_path = os.path.join(os.path.dirname(__file__), "static", "manifest.json")
    if os.path.exists(manifest_path):
        with open(manifest_path) as f:
            data = json.load(f)
        from flask import Response
        return Response(json.dumps(data), mimetype="application/manifest+json")
    return jsonify({}), 404


@app.route("/sw.js")
def pwa_sw():
    """Sert le Service Worker depuis la racine (obligatoire pour scope /)."""
    sw_path = os.path.join(os.path.dirname(__file__), "static", "sw.js")
    if os.path.exists(sw_path):
        from flask import send_file, Response
        with open(sw_path) as f: content = f.read()
        return Response(content, mimetype="application/javascript",
            headers={"Service-Worker-Allowed": "/"})
    return "// sw not found", 404


@app.route("/offline")
def pwa_offline():
    """Page affichée quand l'utilisateur est hors ligne."""
    t = get_tenant() if current_user.is_authenticated else None
    return render_template("tenant/offline.html", tenant=t)


@app.route("/simulateur")
@login_required
def simulateur_paie():
    """Simulateur de paie interactif."""
    if current_user.is_super_admin: return redirect(url_for("admin_dashboard"))
    t = get_tenant()
    if not t: return redirect(url_for("login"))
    salaries_list = Salarie.query.filter_by(tenant_id=t.id, statut="ACTIF")        .options(joinedload(Salarie.categorie)).order_by(Salarie.nom).all()
    return render_template("tenant/simulateur.html", tenant=t, salaries=salaries_list)


@app.route("/api/simuler-paie", methods=["POST"])
@login_required
def api_simuler_paie():
    """API JSON : simuler un bulletin de paie complet sans le sauvegarder."""
    t = get_tenant()
    if not t: return jsonify({"erreur": "non connecté"})
    from calculs_paie import calculer_bulletin, calculer_heures_sup_btp
    try:
        d = request.get_json(force=True) or {}
        # Accepter aussi le form-data
        if not d:
            d = {k: request.form.get(k) for k in request.form}

        def flt(key, default=0):
            try: return float(str(d.get(key) or default).replace(",",".") or default)
            except: return float(default)

        # Calcul des heures sup si mode BTP demandé
        mode_btp = d.get("mode_btp") == "1"
        sal_base = flt("salaire_base")

        if mode_btp and sal_base > 0:
            from calculs_paie import calculer_heures_sup_btp
            btp = calculer_heures_sup_btp(sal_base,
                h10=flt("h10") or None, h30=flt("h30") or None,
                h40=flt("h40"), h70=flt("h70"))
            d["heures_sup_10"] = btp["montant_10"]
            d["heures_sup_30"] = btp["montant_30"]
            d["heures_sup_40"] = btp["montant_40"]
            d["heures_sup_70"] = btp["montant_70"]

        # Nombre de parts IRPP
        sal_id = d.get("salarie_id")
        nb_parts = 1.0
        if sal_id:
            s = Salarie.query.filter_by(id=int(sal_id), tenant_id=t.id).first()
            if s: nb_parts = float(s.nombre_parts or 1)
        nb_parts = flt("nb_parts") or nb_parts

        result = calculer_bulletin(d, nb_parts=nb_parts)

        # Enrichir avec détails BTP
        from calculs_paie import calculer_taux_horaire, H_NORMALES_MENSUEL
        if sal_base > 0:
            th = calculer_taux_horaire(sal_base)
            result["taux_horaire"]     = round(th, 2)
            result["h_normales_mensuel"] = H_NORMALES_MENSUEL

        # Ajouter libellés pour l'affichage
        result["gains_detail"] = [
            {"label": "Salaire de base",          "montant": result["salaire_base"]},
            {"label": "H.sup +10%",               "montant": result["heures_sup_10"]},
            {"label": "H.sup +30%",               "montant": result["heures_sup_30"]},
            {"label": "H.sup +40% (nuit/dim.)",   "montant": result["heures_sup_40"]},
            {"label": "H.sup +70% (fériés)",      "montant": result["heures_sup_70"]},
            {"label": "Sursalaire",               "montant": result["sursalaire"]},
            {"label": "Prime de transport",       "montant": result["prime_transport"]},
            {"label": "Prime de responsabilité",  "montant": result["prime_responsabilite"]},
            {"label": "Indemnité logement",       "montant": result["indem_logement"]},
            {"label": "Prime d'ancienneté",      "montant": result["prime_anciennete"]},
            {"label": "Autres primes",            "montant": sum([
                result["prime_caisse"], result["carburant"],
                result["prime_rendement"], result["prime_qualite"],
                result["prime_performance"], result["prime_assiduité"],
            ])},
        ]
        result["gains_detail"] = [g for g in result["gains_detail"] if g["montant"] > 0]

        result["retenues_detail"] = [
            {"label": f"CNSS salarié (5% / base {int(result['base_cnss']):,} FCFA)", "montant": result["cnss_salarie"]},
            {"label": f"CNAMGS salarié (2% / base {int(result['base_cnamgs']):,} FCFA)", "montant": result["cnamgs_salarie"]},
            {"label": f"TCS (5% / base {int(result['base_tcs']):,} FCFA)",             "montant": result["tcs"]},
            {"label": f"IRPP ({nb_parts} part(s))",                                    "montant": result["irpp"]},
            {"label": "Acompte",                                                        "montant": result["acompte"]},
            {"label": "Retenue absences",                                               "montant": result["absences"]},
        ]
        result["retenues_detail"] = [r for r in result["retenues_detail"] if r["montant"] > 0]

        result["charges_pat_detail"] = [
            {"label": "CNSS patronal (18%)",     "montant": result["cnss_patronale"]},
            {"label": "CNAMGS patronal (4.1%)",  "montant": result["cnamgs_patronale"]},
            {"label": "FNH (3%)",                "montant": result["fnh"]},
            {"label": "CFP (0.5%)",              "montant": result["cfp"]},
        ]

        return jsonify(result)
    except Exception as e:
        import traceback
        return jsonify({"erreur": str(e), "trace": traceback.format_exc()[-300:]})

@app.route("/salaries/import", methods=["GET","POST"])
@login_required
def salaries_import():
    """Import en masse de salariés depuis un fichier Excel."""
    if current_user.is_super_admin: return redirect(url_for("admin_dashboard"))
    t = get_tenant()
    if not t: return redirect(url_for("login"))

    if request.method == "GET":
        categories = CategorieEmploi.query.filter_by(tenant_id=t.id).all()
        return render_template("tenant/salaries_import.html",
            tenant=t, categories=categories)

    # ── POST : traitement du fichier ─────────────────────────────────────────
    fichier = request.files.get("fichier")
    if not fichier or not fichier.filename.endswith((".xlsx", ".xls")):
        flash("❌ Fichier invalide. Utilisez le modèle Excel fourni (.xlsx).", "error")
        return redirect(url_for("salaries_import"))

    mode = request.form.get("mode", "ignorer")  # ignorer | ecraser

    import openpyxl
    from datetime import date as date_type

    def parse_date(val):
        if not val: return None
        if isinstance(val, (date_type, datetime)): return val if isinstance(val, date_type) else val.date()
        s = str(val).strip()
        for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d/%m/%y"):
            try: return datetime.strptime(s, fmt).date()
            except: pass
        return None

    def clean(val): return str(val).strip() if val not in (None, "") else None

    try:
        wb = openpyxl.load_workbook(fichier, data_only=True)
        ws = wb.active
    except Exception as e:
        flash(f"❌ Erreur lecture fichier : {e}", "error")
        return redirect(url_for("salaries_import"))

    # Trouver la ligne d'en-tête (ligne avec "MATRICULE")
    header_row = None
    for row_idx in range(1, 10):
        row_vals = [str(ws.cell(row_idx, c).value or "").upper().strip() for c in range(1, 20)]
        if "MATRICULE" in row_vals:
            header_row = row_idx
            break

    if not header_row:
        flash("❌ Entête non trouvée. Utilisez le modèle Excel fourni.", "error")
        return redirect(url_for("salaries_import"))

    # Mapper les colonnes
    headers = {}
    for col in range(1, ws.max_column + 1):
        val = str(ws.cell(header_row, col).value or "").upper().strip()
        if val: headers[val] = col

    required = ["MATRICULE", "NOM", "PRENOM", "EMPLOI", "DATE_EMBAUCHE"]
    for req in required:
        if req not in headers:
            flash(f"❌ Colonne obligatoire manquante : {req}", "error")
            return redirect(url_for("salaries_import"))

    # Charger les catégories du tenant
    cats = {c.code.upper(): c for c in CategorieEmploi.query.filter_by(tenant_id=t.id).all()}

    # ── Vérifier quota ────────────────────────────────────────────────────────
    nb_existants = Salarie.query.filter_by(tenant_id=t.id, statut="ACTIF").count()
    quota = t.quota_employes_info

    # Traiter les lignes de données
    nb_crees = nb_maj = nb_erreurs = nb_ignores = 0
    erreurs   = []
    avertissements = []

    for row_idx in range(header_row + 1, ws.max_row + 1):
        def get(col_name):
            c = headers.get(col_name)
            return ws.cell(row_idx, c).value if c else None

        matricule = clean(get("MATRICULE"))
        if not matricule: continue  # ligne vide

        nom    = clean(get("NOM"))
        prenom = clean(get("PRENOM"))
        emploi = clean(get("EMPLOI"))
        date_e = parse_date(get("DATE_EMBAUCHE"))

        # Validation champs obligatoires
        if not all([nom, prenom, emploi, date_e]):
            erreurs.append(f"Ligne {row_idx} ({matricule}) : champs obligatoires manquants")
            nb_erreurs += 1
            continue

        # Vérifier si matricule existe
        existant = Salarie.query.filter_by(tenant_id=t.id, matricule=matricule).first()

        if existant and mode == "ignorer":
            nb_ignores += 1
            continue

        # Récupérer catégorie
        cat_code = clean(get("CATEGORIE"))
        cat = cats.get(cat_code.upper()) if cat_code else None

        # Salaire de base → créer/maj contrat
        salaire_raw = get("SALAIRE_BASE")
        salaire_base = None
        if salaire_raw:
            try: salaire_base = float(str(salaire_raw).replace(" ","").replace(",","."))
            except: pass

        # Quota check pour nouvelles créations
        if not existant and quota.get("max"):
            if nb_existants + nb_crees >= quota["max"]:
                erreurs.append(f"Quota atteint ({quota['max']} employés). Import arrêté à la ligne {row_idx}.")
                break

        data = dict(
            tenant_id              = t.id,
            matricule              = matricule.upper(),
            nom                    = nom.upper(),
            prenom                 = prenom,
            emploi                 = emploi.upper(),
            date_embauche          = date_e,
            categorie_id           = cat.id if cat else None,
            telephone              = clean(get("TELEPHONE")),
            sexe                   = clean(get("SEXE")),
            date_naissance         = parse_date(get("DATE_NAISSANCE")),
            situation_matrimoniale = clean(get("SITUATION_MAT")),
            nationalite            = clean(get("NATIONALITE")) or "GABONAISE",
            numero_cnss            = clean(get("NUMERO_CNSS")),
            numero_cnamgs          = clean(get("NUMERO_CNAMGS")),
            email                  = clean(get("EMAIL")),
            adresse                = clean(get("ADRESSE")),
            statut                 = "ACTIF",
        )
        try: data["nb_enfants"]   = int(float(str(get("NB_ENFANTS") or 0)))
        except: data["nb_enfants"] = 0
        try: data["nombre_parts"] = float(str(get("NOMBRE_PARTS") or 1).replace(",","."))
        except: data["nombre_parts"] = 1

        try:
            if existant:
                for k, v in data.items():
                    if v is not None: setattr(existant, k, v)
                s_obj = existant
                nb_maj += 1
            else:
                s_obj = Salarie(**data)
                db.session.add(s_obj)
                db.session.flush()
                nb_crees += 1

            # Créer/maj contrat si salaire fourni
            if salaire_base and salaire_base > 0:
                contrat = next((c for c in (s_obj.contrats if s_obj.id else [])), None)
                if not contrat:
                    contrat = Contrat(tenant_id=t.id, salarie_id=s_obj.id,
                                      date_debut=date_e, actif=True)
                    db.session.add(contrat)
                contrat.salaire_base = salaire_base
                contrat.type_contrat = "CDI"
                contrat.actif        = True

        except Exception as e:
            db.session.rollback()
            erreurs.append(f"Ligne {row_idx} ({matricule}) : {str(e)[:80]}")
            nb_erreurs += 1
            continue

    db.session.commit()
    _cache_delete(f"{t.id}:")  # Invalider cache

    # Message résumé
    msg_parts = []
    if nb_crees:   msg_parts.append(f"✅ {nb_crees} salarié(s) créé(s)")
    if nb_maj:     msg_parts.append(f"🔄 {nb_maj} mis à jour")
    if nb_ignores: msg_parts.append(f"⏭️ {nb_ignores} ignoré(s) (déjà existants)")
    if nb_erreurs: msg_parts.append(f"❌ {nb_erreurs} erreur(s)")
    flash(" · ".join(msg_parts) or "Aucune donnée importée.", "success" if not nb_erreurs else "error")

    for err in erreurs[:5]:
        flash(f"⚠️ {err}", "error")

    return redirect(url_for("salaries"))


@app.route("/salaries/import/modele")
@login_required
def salaries_import_modele():
    """Télécharger le modèle Excel vierge."""
    import os
    modele_path = os.path.join(os.path.dirname(__file__), "modele_import_salaries.xlsx")
    if os.path.exists(modele_path):
        from flask import send_file
        return send_file(modele_path, as_attachment=True,
                         download_name="modele_import_salaries.xlsx",
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    flash("Modèle non disponible.", "error")
    return redirect(url_for("salaries_import"))

@app.route("/salaries/nouveau", methods=["GET","POST"])
@login_required
def salarie_nouveau():
    if current_user.is_super_admin: return redirect(url_for("admin_dashboard"))
    t=get_tenant()
    if not t: return redirect(url_for("login"))
    if not current_user.can_edit: abort(403)
    cats=CategorieEmploi.query.filter_by(tenant_id=t.id).all()
    # Vérifier quota dès le GET (bloquer accès au formulaire)
    q = t.quota_employes_info
    if q["max"] and q["plein"]:
        flash(
            f"Limite atteinte — Plan « {t.plan.nom} » : {q['max']} employé(s) maximum "
            f"({q['salaries']} salarié(s) + {q['journaliers']} journalier(s)). "
            f"Passez au plan supérieur.", "error"
        )
        return redirect(url_for("salaries"))
    if request.method=="POST":
        if not t.peut_ajouter_employe:
            flash(f"Limite atteinte ({t.plan.max_salaries} employés). Passez au plan supérieur.","error")
            return redirect(url_for("salaries"))
        s=Salarie(tenant_id=t.id,
            matricule=request.form["matricule"].strip().upper(),
            categorie_id=request.form.get("categorie_id") or None,
            nom=request.form["nom"].strip().upper(), prenom=request.form["prenom"].strip(),
            telephone=request.form.get("telephone"), email=request.form.get("email","").strip() or None,
            nationalite=request.form.get("nationalite","GABONAISE"),
            sexe=request.form.get("sexe"),
            date_naissance=_pd(request.form.get("date_naissance")),
            date_embauche=_pd(request.form["date_embauche"]),
            situation_matrimoniale=request.form.get("situation_matrimoniale"),
            nb_enfants=int(request.form.get("nb_enfants") or 0),
            nb_enfants_moins_16ans=int(request.form.get("nb_enfants_moins_16ans") or 0),
            nombre_parts=float(request.form.get("nombre_parts") or 1),
            numero_cnss=request.form.get("numero_cnss"), numero_cnamgs=request.form.get("numero_cnamgs"),
            emploi=request.form.get("emploi"), assujetti_cnamgs=request.form.get("assujetti_cnamgs")=="OUI", statut="ACTIF")
        db.session.add(s)
        sb=float(request.form.get("salaire_base") or 0)
        if sb: db.session.add(Contrat(tenant_id=t.id,salarie=s,type_contrat=request.form.get("type_contrat","CDI"),date_debut=s.date_embauche,salaire_base=sb,poste=s.emploi,actif=True))
        db.session.commit(); flash(f"Salarié {s.nom_complet} créé.","success")
        return redirect(url_for("salarie_detail",id=s.id))
    sites = Site.query.filter_by(tenant_id=t.id, actif=True).order_by(Site.nom).all()
    return render_template("tenant/salarie_form.html", salarie=None, categories=cats,
        action="nouveau", tenant=t, sites=sites, aff_actuelle=None)

@app.route("/salaries/<int:id>")
@login_required
def salarie_detail(id):
    if current_user.is_super_admin: return redirect(url_for("admin_dashboard"))
    t = get_tenant()
    if not t: return redirect(url_for("login"))
    s = Salarie.query.filter_by(id=id, tenant_id=t.id).first_or_404()
    bulletins = BulletinPaie.query.filter_by(salarie_id=id, tenant_id=t.id)        .order_by(BulletinPaie.date_creation.desc()).all()
    contrat = Contrat.query.filter_by(salarie_id=id, tenant_id=t.id, actif=True).first()
    conge   = Conge.query.filter_by(salarie_id=id, tenant_id=t.id,
                                     annee=datetime.now().year).first()
    total_brut = sum(float(b.salaire_brut or 0) for b in bulletins)
    total_net  = sum(float(b.net_a_payer  or 0) for b in bulletins)
    total_cnss = sum(float(b.cnss_salarie or 0) for b in bulletins)
    total_irpp = sum(float(b.irpp         or 0) for b in bulletins)
    nb_mois    = len(bulletins)
    anciennete_jours = (datetime.now().date() - s.date_embauche).days if s.date_embauche else 0
    anciennete_ans   = anciennete_jours // 365
    anciennete_mois  = (anciennete_jours % 365) // 30

    # ── Historique des pointages (30 derniers jours) ──────────────────────────
    nb_jours = request.args.get("nb_jours", type=int, default=30)
    nb_jours = min(max(nb_jours, 7), 90)          # borne 7-90 jours
    date_fin   = datetime.now().date()
    date_debut = date_fin - timedelta(days=nb_jours - 1)

    pts_hist = Pointage.query.filter_by(tenant_id=t.id, salarie_id=id)        .filter(Pointage.date_pointage >= date_debut,
                Pointage.date_pointage <= date_fin)        .order_by(Pointage.date_pointage.desc()).all()

    # Stats synthèse
    nb_presences  = sum(1 for p in pts_hist if p.present)
    nb_absences   = sum(1 for p in pts_hist if p.absent)
    nb_non_pointes = nb_jours - len(pts_hist)
    h_normales_tot = round(sum(float(p.heures_normales or 0) for p in pts_hist if p.present), 1)
    h_sup_tot      = round(sum(
        float(p.heures_sup_10 or 0) + float(p.heures_sup_30 or 0) +
        float(p.heures_sup_40 or 0) + float(p.heures_sup_70 or 0)
        for p in pts_hist if p.present), 1)
    taux_presence = round(nb_presences / (nb_presences + nb_absences) * 100
                          ) if (nb_presences + nb_absences) > 0 else 0

    return render_template("tenant/salarie_detail.html",
        salarie=s, tenant=t, bulletins=bulletins, contrat=contrat, conge=conge,
        total_brut=total_brut, total_net=total_net, total_cnss=total_cnss,
        total_irpp=total_irpp, nb_mois=nb_mois,
        anciennete_ans=anciennete_ans, anciennete_mois=anciennete_mois,
        # Historique pointage
        pts_hist=pts_hist, nb_jours=nb_jours,
        nb_presences=nb_presences, nb_absences=nb_absences,
        nb_non_pointes=nb_non_pointes,
        h_normales_tot=h_normales_tot, h_sup_tot=h_sup_tot,
        taux_presence=taux_presence,
        date_debut_hist=date_debut, date_fin_hist=date_fin)

@app.route("/salaries/<int:id>/modifier", methods=["GET","POST"])
@login_required
def salarie_modifier(id):
    if current_user.is_super_admin: return redirect(url_for("admin_dashboard"))
    t=get_tenant()
    if not t: return redirect(url_for("login"))
    if not current_user.can_edit: abort(403)
    s = Salarie.query.filter_by(id=id, tenant_id=t.id).first_or_404()
    cats = CategorieEmploi.query.filter_by(tenant_id=t.id).all()
    if request.method=="POST":
        for f,v in [("nom",request.form["nom"].strip().upper()),("prenom",request.form["prenom"].strip()),
            ("telephone",request.form.get("telephone")),
            ("email",request.form.get("email","").strip() or None),
            ("nationalite",request.form.get("nationalite")),
            ("sexe",request.form.get("sexe")),("date_naissance",_pd(request.form.get("date_naissance"))),
            ("situation_matrimoniale",request.form.get("situation_matrimoniale")),
            ("nb_enfants",int(request.form.get("nb_enfants") or 0)),
            ("nombre_parts",calculer_parts_irpp(request.form.get("situation_matrimoniale",""),int(request.form.get("nb_enfants",0) or 0))),
            ("numero_cnss",request.form.get("numero_cnss")),("numero_cnamgs",request.form.get("numero_cnamgs")),
            ("emploi",request.form.get("emploi")),("categorie_id",request.form.get("categorie_id") or None),
            ("statut",request.form.get("statut","ACTIF")),("date_modification",datetime.utcnow())]:
            setattr(s,f,v)
        db.session.commit()
        # ── Affectation site ──────────────────────────────────────────────
        site_id = request.form.get("site_id", type=int)
        if site_id:
            aff_prev = AffectationSite.query.filter_by(
                salarie_id=s.id, tenant_id=t.id, actif=True).first()
            if aff_prev and aff_prev.site_id != site_id:
                aff_prev.actif    = False
                aff_prev.date_fin = date.today()
                aff_prev.motif    = "Réaffecté via formulaire salarié"
            if not aff_prev or aff_prev.site_id != site_id:
                db.session.add(AffectationSite(
                    tenant_id=t.id, site_id=site_id, salarie_id=s.id,
                    date_debut=date.today(), actif=True,
                    cree_par=current_user.email))
            db.session.commit()
        elif request.form.get("retirer_site"):
            aff = AffectationSite.query.filter_by(
                salarie_id=s.id, tenant_id=t.id, actif=True).first()
            if aff:
                aff.actif    = False
                aff.date_fin = date.today()
                aff.motif    = "Retiré via formulaire salarié"
                db.session.commit()
        flash("Fiche mise à jour.", "success")
        return redirect(url_for("salarie_detail", id=s.id))
    # Récupérer site actuel + liste des sites
    aff_actuelle = AffectationSite.query.filter_by(
        salarie_id=id, tenant_id=t.id, actif=True).first()
    sites = Site.query.filter_by(tenant_id=t.id, actif=True).order_by(Site.nom).all()
    return render_template("tenant/salarie_form.html", salarie=s, categories=cats,
        action="modifier", tenant=t, sites=sites, aff_actuelle=aff_actuelle)


@app.route("/salaries/<int:id>/supprimer", methods=["POST"])
@login_required
def salarie_supprimer(id):
    if current_user.is_super_admin: return redirect(url_for("admin_dashboard"))
    t = get_tenant()
    if not t: return redirect(url_for("login"))
    if not current_user.is_tenant_admin:
        flash("Seul l'administrateur peut supprimer un salarié.", "error")
        return redirect(url_for("salaries"))
    s = Salarie.query.filter_by(id=id, tenant_id=t.id).first_or_404()
    nom = s.nom_complet
    bulletins_actifs = BulletinPaie.query.filter_by(salarie_id=id).filter(
        BulletinPaie.statut.in_(["VALIDÉ","PAYÉ"])).count()
    if bulletins_actifs > 0:
        flash(f"Impossible de supprimer {nom} : {bulletins_actifs} bulletin(s) validé(s). Passez-le en INACTIF.", "error")
        return redirect(url_for("salarie_detail", id=id))
    try:
        BulletinPaie.query.filter_by(salarie_id=id).delete()
        Contrat.query.filter_by(salarie_id=id).delete()
        Pointage.query.filter_by(salarie_id=id).delete()
        Acompte.query.filter_by(salarie_id=id).delete()
        Conge.query.filter_by(salarie_id=id).delete()
        db.session.delete(s); db.session.commit()
        flash(f"Salarié {nom} supprimé.", "success")
    except Exception as e:
        db.session.rollback(); flash(f"Erreur: {str(e)}", "error")
        return redirect(url_for("salarie_detail", id=id))
    return redirect(url_for("salaries"))

# ── Bulletins ─────────────────────────────────────────────────────────────────
@app.route("/bulletins")
@login_required
def bulletins():
    if current_user.is_super_admin: return redirect(url_for("admin_dashboard"))
    t=get_tenant()
    if not t: return redirect(url_for("login"))
    pid          = request.args.get("periode_id", type=int)
    sf           = request.args.get("statut", "")
    site_filtre_id = request.args.get("site_id", type=int)
    periodes     = PeriodePaie.query.filter_by(tenant_id=t.id)                    .order_by(PeriodePaie.annee.desc(), PeriodePaie.mois.desc()).all()
    sites_list   = Site.query.filter_by(tenant_id=t.id, actif=True).order_by(Site.nom).all()
    site_filtre  = Site.query.get(site_filtre_id) if site_filtre_id else None
    ps = None; buls = []; masse = {}; pagination = None

    if pid:
        ps = PeriodePaie.query.filter_by(id=pid, tenant_id=t.id).first_or_404()
        q = BulletinPaie.query.options(
            joinedload(BulletinPaie.salarie),
            joinedload(BulletinPaie.periode),
        ).filter_by(tenant_id=t.id, periode_id=pid)
        if sf:
            q = q.filter_by(statut=sf)

        # ── Filtre par site ──────────────────────────────────────────────────
        if site_filtre_id:
            # Récupérer les IDs des salariés affectés à ce site
            ids_sal = [a.salarie_id for a in AffectationSite.query.filter_by(
                tenant_id=t.id, site_id=site_filtre_id, actif=True
            ).filter(AffectationSite.salarie_id.isnot(None)).all()]
            q = q.filter(BulletinPaie.salarie_id.in_(ids_sal))

        page_bul   = request.args.get("page", 1, type=int)
        # Une seule query base avec le join, puis on pagine
        q_joined   = q.join(Salarie).order_by(Salarie.nom)
        buls_tous  = q_joined.all()
        masse      = calculer_masse_salariale(buls_tous)
        pagination = q_joined.paginate(page=page_bul, per_page=25, error_out=False)
        buls       = pagination.items

    # Affectation site de chaque salarié pour affichage dans le tableau
    aff_sal = {a.salarie_id: a.site for a in AffectationSite.query.filter_by(
        tenant_id=t.id, actif=True
    ).filter(AffectationSite.salarie_id.isnot(None)).all()}

    _args = {k: v for k, v in request.args.items() if k != 'page'}
    _base = request.path + '?' + '&'.join(f'{k}={v}' for k, v in _args.items())
    _sep  = '&' if _args else '?'
    return render_template("tenant/bulletins.html",
        periodes=periodes, periode_sel=ps,
        bulletins=buls, masse=masse, statut_filtre=sf,
        sites=sites_list, site_filtre=site_filtre, aff_sal=aff_sal,
        pagination=pagination if pid else None,
        pagination_base=_base + _sep,
        tenant=t)

@app.route("/bulletins/valider-lot", methods=["POST"])
@login_required
def bulletins_valider_lot():
    """Valider une sélection de bulletins ou tous les brouillons d'une période."""
    if current_user.is_super_admin: return redirect(url_for("admin_dashboard"))
    t = get_tenant()
    if not t: return redirect(url_for("login"))
    if not current_user.can_edit: abort(403)

    pid      = request.form.get("periode_id", type=int)
    site_id  = request.form.get("site_id",    type=int)
    action   = request.form.get("action_lot", "valider")
    ids_str  = request.form.get("bulletin_ids", "")
    ids_sel  = [int(x) for x in ids_str.split(",") if x.strip().isdigit()]

    if not pid:
        flash("Période manquante.", "error")
        return redirect(url_for("bulletins"))

    # ── CORRECTION BUG : si aucun ID sélectionné + action valider → refuser ──
    # Avant : ids_sel vide → validait TOUS les bulletins de la période
    # Maintenant : ids_sel vide → uniquement pour les actions "tout valider" explicites
    if not ids_sel and action == "valider":
        # Vérifier que c'est bien une demande "tout valider" (bouton dédié)
        tout_valider = request.form.get("tout_valider", "0")
        if tout_valider != "1":
            flash("Aucun bulletin sélectionné. Cochez des bulletins ou utilisez 'Tout valider'.", "warning")
            return redirect(f"/bulletins?periode_id={pid}" + (f"&site_id={site_id}" if site_id else ""))

    # Charger les bulletins ciblés
    if ids_sel:
        buls = BulletinPaie.query.filter(
            BulletinPaie.id.in_(ids_sel),
            BulletinPaie.tenant_id == t.id
        ).all()
    else:
        q = BulletinPaie.query.filter_by(tenant_id=t.id, periode_id=pid, statut="BROUILLON")
        if site_id:
            ids_sal = [a.salarie_id for a in AffectationSite.query.filter_by(
                tenant_id=t.id, site_id=site_id, actif=True
            ).filter(AffectationSite.salarie_id.isnot(None)).all()]
            q = q.filter(BulletinPaie.salarie_id.in_(ids_sal))
        buls = q.all()

    nb = 0
    if action == "valider":
        for b in buls:
            if b.statut != "BROUILLON": continue
            b.statut          = "VALIDÉ"
            b.date_validation = datetime.utcnow()
            for a in Acompte.query.filter_by(
                tenant_id=t.id, salarie_id=b.salarie_id,
                mois=b.periode.mois, annee=b.periode.annee, statut="EN_ATTENTE").all():
                a.statut = "DEDUIT"
            nb += 1
        msg = f"✅ {nb} bulletin(s) validé(s)."

    elif action == "annuler_validation":
        # ── NOUVEAU : annuler la validation → repasser en BROUILLON ──────────
        for b in buls:
            if b.statut not in ("VALIDÉ", "VALIDE"): continue
            b.statut          = "BROUILLON"
            b.date_validation = None
            # Remettre les acomptes déduits en attente
            for a in Acompte.query.filter_by(
                tenant_id=t.id, salarie_id=b.salarie_id,
                mois=b.periode.mois, annee=b.periode.annee, statut="DEDUIT").all():
                a.statut = "EN_ATTENTE"
            nb += 1
        msg = f"↩️ {nb} bulletin(s) remis en brouillon."

    elif action == "payer":
        for b in buls:
            if b.statut not in ("VALIDÉ", "VALIDE", "BROUILLON"): continue
            b.statut = "PAYÉ"
            nb += 1
        msg = f"💰 {nb} bulletin(s) marqué(s) comme payé(s)."

    elif action == "supprimer_brouillons":
        for b in buls:
            if b.statut != "BROUILLON": continue
            db.session.delete(b); nb += 1
        msg = f"🗑️ {nb} brouillon(s) supprimé(s)."

    else:
        flash("Action inconnue.", "error")
        return redirect(f"/bulletins?periode_id={pid}")

    db.session.commit()
    _cache_delete(f"{t.id}:")
    flash(msg, "success")
    redir = f"/bulletins?periode_id={pid}"
    if site_id: redir += f"&site_id={site_id}"
    return redirect(redir)

@app.route("/bulletins/saisie", methods=["GET","POST"])
@login_required
def bulletin_saisie():
    if current_user.is_super_admin: return redirect(url_for("admin_dashboard"))
    t=get_tenant()
    if not t: return redirect(url_for("login"))
    if not current_user.can_edit: abort(403)
    sals=Salarie.query.filter_by(tenant_id=t.id,statut="ACTIF").order_by(Salarie.nom).all()
    pers=PeriodePaie.query.filter_by(tenant_id=t.id,statut="OUVERT").order_by(PeriodePaie.annee.desc(),PeriodePaie.mois.desc()).all()
    if request.method=="POST":
        sid=int(request.form["salarie_id"]); pid=int(request.form["periode_id"])
        s=Salarie.query.filter_by(id=sid,tenant_id=t.id).first_or_404()
        periode = PeriodePaie.query.filter_by(id=pid, tenant_id=t.id).first_or_404()
        acomptes_en_attente = Acompte.query.filter_by(
            tenant_id=t.id, salarie_id=sid,
            mois=periode.mois, annee=periode.annee, statut="EN_ATTENTE").all()
        total_acomptes = sum(float(a.montant) for a in acomptes_en_attente)
        donnees={}
        for k,v in request.form.items():
            # Exclure les champs non-numériques et les champs base_/taux_ (traités séparément)
            if k in ("salarie_id","periode_id","csrf_token","action","nb_jours_travailles"):
                continue
            if k.startswith("base_") or k.startswith("taux_"):
                continue
            try:
                donnees[k] = float(v) if v else 0
            except (ValueError, TypeError):
                donnees[k] = 0
        if total_acomptes > 0:
            donnees["acompte"] = max(donnees.get("acompte", 0), total_acomptes)
        res=calculer_bulletin(donnees,nb_parts=float(s.nombre_parts or 1))
        ex=BulletinPaie.query.filter_by(tenant_id=t.id,salarie_id=sid,periode_id=pid).first()
        b=ex or BulletinPaie(tenant_id=t.id,salarie_id=sid,periode_id=pid)
        if not ex: db.session.add(b)
        for k,v in res.items():
            if not k.startswith("_") and hasattr(b,k): setattr(b,k,v)
        b.nb_jours_travailles=int(request.form.get("nb_jours_travailles") or 0)
        # ✅ Sauvegarder base et taux saisis manuellement pour chaque rubrique
        RUBRIQUES_BT = ["salaire_base","heures_sup_10","heures_sup_30","heures_sup_40","heures_sup_70",
            "absences","sursalaire","prime_caisse","carburant","prime_anciennete",
            "indem_logement","indem_domesticite","indem_eau_electricite","indem_nourriture",
            "prime_transport","prime_responsabilite","prime_rendement","prime_assiduité",
            "prime_qualite","prime_performance","allocations_conge",
            "indem_compensatrice_conge","indem_services_rendus",
            "indem_compensatrice_preavis","indem_licenciement"]
        for r in RUBRIQUES_BT:
            base_val = request.form.get(f"base_{r}", "")
            taux_val = request.form.get(f"taux_{r}", "")
            if hasattr(b, f"base_{r}"):
                try: setattr(b, f"base_{r}", float(base_val) if base_val else None)
                except: pass
            if hasattr(b, f"taux_{r}"):
                setattr(b, f"taux_{r}", taux_val.strip()[:20] if taux_val else "")
        action=request.form.get("action","brouillon")
        if action=="valider":
            b.statut="VALIDÉ"; b.date_validation=datetime.utcnow()
            for a in acomptes_en_attente: a.statut = "DEDUIT"
        else:
            b.statut="BROUILLON"
        db.session.commit()
        if total_acomptes > 0:
            flash(f"Bulletin sauvegardé. Acompte de {int(total_acomptes):,} FCFA déduit automatiquement.".replace(",", " "), "success")
        else:
            flash(f"Bulletin {'validé' if b.statut=='VALIDÉ' else 'sauvegardé'}.","success")
        return redirect(url_for("bulletin_detail",id=b.id))
    sid=request.args.get("salarie_id",type=int)
    ss=Salarie.query.filter_by(id=sid,tenant_id=t.id).first() if sid else None
    c=Contrat.query.filter_by(salarie_id=sid,tenant_id=t.id,actif=True).first() if sid else None
    acomptes_attente = Acompte.query.filter_by(tenant_id=t.id, salarie_id=sid, statut="EN_ATTENTE").all() if sid else []
    total_acomptes = sum(float(a.montant) for a in acomptes_attente)
    return render_template("tenant/bulletin_saisie.html", salaries=sals, periodes=pers, salarie_sel=ss, contrat=c, tenant=t,
        acomptes_attente=acomptes_attente, total_acomptes=total_acomptes)

@app.route("/bulletins/<int:id>")
@login_required
def bulletin_detail(id):
    if current_user.is_super_admin: return redirect(url_for("admin_dashboard"))
    t=get_tenant()
    if not t: return redirect(url_for("login"))
    return render_template("tenant/bulletin_detail.html",
        bulletin=BulletinPaie.query.filter_by(id=id,tenant_id=t.id).first_or_404(), tenant=t)

@app.route("/bulletins/<int:id>/valider", methods=["POST"])
@login_required
def bulletin_valider(id):
    if current_user.is_super_admin: return redirect(url_for("admin_dashboard"))
    t = get_tenant()
    if not t: return redirect(url_for("login"))
    b = BulletinPaie.query.filter_by(id=id, tenant_id=t.id).first_or_404()
    if b.statut == "VALIDÉ":
        flash("Ce bulletin est déjà validé.", "info")
        return redirect(url_for("bulletin_detail", id=id))
    acomptes = Acompte.query.filter_by(
        tenant_id=t.id, salarie_id=b.salarie_id,
        mois=b.periode.mois, annee=b.periode.annee, statut="EN_ATTENTE").all()
    for a in acomptes: a.statut = "DEDUIT"
    b.statut = "VALIDÉ"; b.date_validation = datetime.utcnow()
    db.session.commit()
    flash("Bulletin validé avec succès.", "success")
    return redirect(url_for("bulletin_detail", id=id))

@app.route("/bulletins/<int:id>/payer", methods=["POST"])
@login_required
def bulletin_paye(id):
    if current_user.is_super_admin: return redirect(url_for("admin_dashboard"))
    t = get_tenant()
    if not t: return redirect(url_for("login"))
    b = BulletinPaie.query.filter_by(id=id, tenant_id=t.id).first_or_404()
    b.statut = "PAYÉ"; db.session.commit()
    flash("Bulletin marqué comme payé.", "success")
    return redirect(url_for("bulletin_detail", id=id))

@app.route("/bulletins/<int:id>/supprimer", methods=["POST"])
@login_required
def bulletin_supprimer(id):
    if current_user.is_super_admin:
        b = BulletinPaie.query.get_or_404(id)
        salarie_id = b.salarie_id
        db.session.delete(b); db.session.commit()
        flash("Bulletin supprimé (super admin).", "success")
        return redirect(url_for("salarie_detail", id=salarie_id))
    t = get_tenant()
    if not t: return redirect(url_for("login"))
    b = BulletinPaie.query.filter_by(id=id, tenant_id=t.id).first_or_404()
    if b.statut == "VALIDÉ":
        flash("Impossible de supprimer un bulletin validé.", "error")
        return redirect(url_for("bulletin_detail", id=id))
    db.session.delete(b); db.session.commit()
    flash("Bulletin supprimé.", "success")
    return redirect(url_for("bulletins"))

@app.route("/bulletins/<int:id>/pdf")
@login_required
def bulletin_pdf(id):
    """Génère et retourne le bulletin en PDF téléchargeable."""
    if current_user.is_super_admin:
        b = BulletinPaie.query.get_or_404(id)
        t = b.salarie.tenant
    else:
        t = get_tenant()
        if not t: return redirect(url_for("login"))
        b = BulletinPaie.query.filter_by(id=id, tenant_id=t.id).first_or_404()
    try:
        from pdf_bulletin import generer_bulletin_pdf
        pdf_bytes = generer_bulletin_pdf(b, t)
        nom_fichier = (
            f"bulletin_{b.salarie.nom}_{b.salarie.prenom}_{b.periode.annee}_{b.periode.mois:02d}.pdf"
            .replace(" ", "_")
        )
        from flask import Response
        return Response(
            pdf_bytes,
            mimetype="application/pdf",
            headers={
                "Content-Disposition": f'attachment; filename="{nom_fichier}"',
                "Content-Length": str(len(pdf_bytes)),
            }
        )
    except Exception as e:
        flash(f"Erreur génération PDF : {e}", "error")
        return redirect(url_for("bulletin_detail", id=id))


@app.route("/bulletins/<int:id>/imprimer")
@login_required
def bulletin_imprimer(id):
    if current_user.is_super_admin:
        b = BulletinPaie.query.get_or_404(id)
        t = b.salarie.tenant
    else:
        t = get_tenant()
        if not t: return redirect(url_for("login"))
        b = BulletinPaie.query.filter_by(id=id, tenant_id=t.id).first_or_404()
    # Récupérer le modèle avec fallback sécurisé
    try:
        modele = t.modele_bulletin or "classique"
    except Exception:
        modele = "classique"
    if modele not in ("classique", "moderne", "minimaliste"):
        modele = "classique"
    template_map = {
        "classique":   "tenant/bulletin_print.html",
        "moderne":     "tenant/bulletin_print_moderne.html",
        "minimaliste": "tenant/bulletin_print_minimaliste.html",
    }
    import os
    template = template_map[modele]
    # Vérifier que le fichier template existe sur le serveur
    tpl_path = os.path.join(os.path.dirname(__file__), "templates", template)
    if not os.path.exists(tpl_path):
        template = "tenant/bulletin_print.html"
    return render_template(template, bulletin=b, tenant=t)

# ✅ ENVOI EMAIL ASYNCHRONE — ne bloque plus le serveur
@app.route("/bulletins/<int:id>/envoyer-email", methods=["POST"])
@login_required
def bulletin_envoyer_email(id):
    if current_user.is_super_admin: return redirect(url_for("admin_dashboard"))
    t = get_tenant()
    if not t: return redirect(url_for("login"))
    b = BulletinPaie.query.filter_by(id=id, tenant_id=t.id).first_or_404()
    s = b.salarie
    dest_email = request.form.get("email_dest", "").strip()
    if not dest_email and s.email:
        dest_email = s.email
    if not dest_email:
        flash(f"{s.nom_complet} n'a pas d'adresse email. Renseignez-en une dans le formulaire.", "error")
        return redirect(url_for("bulletin_detail", id=id))
    if not os.environ.get("MAIL_USERNAME"):
        flash("Email non configuré sur le serveur (MAIL_USERNAME manquant).", "error")
        return redirect(url_for("bulletin_detail", id=id))
    try:
        corps = (f"Bonjour {s.prenom},\n\n"
                 f"Veuillez trouver votre bulletin de paie pour : {b.periode.libelle_complet}\n\n"
                 f"Salaire brut : {int(b.salaire_brut or 0)} FCFA\n"
                 f"Net a payer  : {int(b.net_a_payer or 0)} FCFA\n\n"
                 f"Cordialement,\n{t.denomination}")
        msg = Message(
            subject=f"Bulletin de paie {b.periode.libelle_complet} — {t.denomination}",
            recipients=[dest_email],
            body=corps,
            sender=app.config["MAIL_DEFAULT_SENDER"]
        )
        # ✅ Envoi dans un thread séparé → le serveur répond immédiatement
        send_email_async(msg)
        flash(f"Email en cours d'envoi à {dest_email}.", "success")
    except Exception as e:
        flash(f"Erreur préparation email: {str(e)}", "error")
    return redirect(url_for("bulletin_detail", id=id))

@app.route("/bulletins/envoyer-tous", methods=["POST"])
@login_required
def bulletins_envoyer_tous():
    if current_user.is_super_admin: return redirect(url_for("admin_dashboard"))
    t = get_tenant()
    if not t: return redirect(url_for("login"))
    periode_id = request.form.get("periode_id", type=int)
    if not periode_id: flash("Période manquante.", "error"); return redirect(url_for("bulletins"))
    buls = BulletinPaie.query.filter_by(tenant_id=t.id, periode_id=periode_id).all()
    nb_ok=0; nb_sans_email=0
    for b in buls:
        if not b.salarie.email: nb_sans_email+=1; continue
        try:
            corps = (f"Bonjour {b.salarie.prenom},\n\n"
                     f"Bulletin {b.periode.libelle_complet}\n"
                     f"Net a payer : {int(b.net_a_payer or 0)} FCFA\n\n"
                     f"Cordialement, {t.denomination}")
            msg = Message(subject=f"Bulletin {b.periode.libelle_complet}",
                recipients=[b.salarie.email], body=corps,
                sender=app.config["MAIL_DEFAULT_SENDER"])
            send_email_async(msg)
            nb_ok+=1
        except Exception as e:
            print(f"Erreur email {b.salarie.email}: {e}")
    flash(f"{nb_ok} email(s) en cours d'envoi. {nb_sans_email} salarié(s) sans email.", "success")
    return redirect(url_for("bulletins"))

# ── Périodes ──────────────────────────────────────────────────────────────────
@app.route("/periodes")
@login_required
def periodes():
    if current_user.is_super_admin: return redirect(url_for("admin_dashboard"))
    t=get_tenant()
    if not t: return redirect(url_for("login"))
    return render_template("tenant/periodes.html", tenant=t,
        periodes=PeriodePaie.query.filter_by(tenant_id=t.id).order_by(PeriodePaie.annee.desc(),PeriodePaie.mois.desc()).all(),
        now=datetime.now())

@app.route("/periodes/nouvelle", methods=["POST"])
@tenant_required
@can_edit
def periode_nouvelle():
    t=get_tenant(); annee=int(request.form["annee"]); mois=int(request.form["mois"])
    noms=PeriodePaie.MOIS_NOMS
    if PeriodePaie.query.filter_by(tenant_id=t.id,annee=annee,mois=mois).first(): flash("Période existante.","warning")
    else:
        db.session.add(PeriodePaie(tenant_id=t.id,annee=annee,mois=mois,libelle_mois=noms[mois],
            trimestre=f"T{(mois-1)//3+1}",statut="OUVERT",date_ouverture=datetime.utcnow()))
        db.session.commit(); flash(f"Période {noms[mois]} {annee} créée.","success")
    return redirect(url_for("periodes"))

@app.route("/periodes/<int:id>/cloturer", methods=["POST"])
@tenant_required
@can_edit
def periode_cloturer(id):
    t=get_tenant(); p=PeriodePaie.query.filter_by(id=id,tenant_id=t.id).first_or_404()
    p.statut="CLÔTURÉ"; p.date_cloture=datetime.utcnow(); db.session.commit()
    flash("Période clôturée.","success"); return redirect(url_for("periodes"))

# ── Paiement abonnement ───────────────────────────────────────────────────────
@app.route("/paiement")
@login_required
def paiement():
    if current_user.is_super_admin: return redirect(url_for("admin_dashboard"))
    t = get_tenant()
    if not t: return redirect(url_for("login"))
    plans = Plan.query.filter_by(actif=True).order_by(Plan.prix_mensuel).all()
    historique = Paiement.query.filter_by(tenant_id=t.id)\
        .order_by(Paiement.date_creation.desc()).limit(10).all()
    return render_template("tenant/paiement.html", tenant=t, plans=plans,
                           historique=historique)


# ── Airtel Money — Initiation ──────────────────────────────────────────────────
@app.route("/paiement/airtel/initier", methods=["POST"])
@login_required
def paiement_airtel_initier():
    """
    Lance une demande de paiement STK Push Airtel Money.
    Le client reçoit une notification USSD sur son téléphone.
    """
    if current_user.is_super_admin: return redirect(url_for("admin_dashboard"))
    t = get_tenant()
    if not t: return redirect(url_for("login"))

    telephone  = request.form.get("telephone", "").strip()
    duree      = int(request.form.get("duree", 1) or 1)
    plan_id    = request.form.get("plan_id", type=int) or (t.plan_id)

    plan = Plan.query.get(plan_id) if plan_id else t.plan
    if not plan:
        flash("Plan introuvable.", "error")
        return redirect(url_for("paiement"))

    if not telephone:
        flash("Veuillez saisir votre numéro Airtel Money.", "error")
        return redirect(url_for("paiement"))

    montant = float(plan.prix_mensuel) * duree

    # Générer une référence unique
    import uuid
    reference = f"AM-{t.id}-{uuid.uuid4().hex[:10].upper()}"

    # Enregistrer la tentative en base
    p = Paiement(
        tenant_id=t.id,
        moyen="AIRTEL_MONEY",
        montant=montant,
        duree_mois=duree,
        plan_id=plan.id,
        reference_interne=reference,
        telephone=telephone,
        statut="EN_ATTENTE",
    )
    db.session.add(p)
    db.session.commit()

    # Appeler l'API Airtel
    try:
        from airtel_money import initier_paiement, AirtelConfigError
        resultat = initier_paiement(
            reference=reference,
            telephone=telephone,
            montant=montant,
            description=f"Abonnement PaieGabon {plan.nom} — {duree} mois",
        )
        p.reference_externe = resultat.get("transaction_id")
        import json
        p.reponse_raw = json.dumps(resultat.get("raw", {}))

        if resultat["success"]:
            db.session.commit()
            logger.info(f"[Paiement] Airtel initié — ref={reference} tenant={t.id}")
            flash(
                f"Demande de paiement envoyée sur le {telephone}. "
                "Validez sur votre téléphone dans les 2 minutes.",
                "success"
            )
            return redirect(url_for("paiement_airtel_attente", reference=reference))
        else:
            p.statut = "ECHEC"
            p.notes  = resultat["message"]
            db.session.commit()
            flash(f"Échec : {resultat['message']}", "error")
            return redirect(url_for("paiement"))

    except Exception as e:
        p.statut = "ECHEC"
        p.notes  = str(e)
        db.session.commit()
        logger.error(f"[Paiement] Erreur Airtel : {e}")
        flash(f"Erreur de connexion Airtel Money. Réessayez ou contactez le support.", "error")
        return redirect(url_for("paiement"))


# ── Airtel Money — Page d'attente ──────────────────────────────────────────────
@app.route("/paiement/airtel/attente/<reference>")
@login_required
def paiement_airtel_attente(reference):
    """
    Page d'attente affichée après l'initiation.
    Fait un polling automatique toutes les 5 secondes via AJAX.
    """
    if current_user.is_super_admin: return redirect(url_for("admin_dashboard"))
    t = get_tenant()
    if not t: return redirect(url_for("login"))
    p = Paiement.query.filter_by(
        reference_interne=reference, tenant_id=t.id
    ).first_or_404()
    return render_template("tenant/paiement_attente.html", paiement=p, tenant=t)


# ── Airtel Money — Vérification statut (AJAX polling) ─────────────────────────
@app.route("/paiement/airtel/statut/<reference>")
@login_required
def paiement_airtel_statut(reference):
    """
    Endpoint JSON pour le polling côté client.
    Retourne le statut actuel de la transaction.
    """
    t = get_tenant()
    if not t: return jsonify({"statut": "ERREUR", "message": "Non connecté"}), 401

    p = Paiement.query.filter_by(
        reference_interne=reference, tenant_id=t.id
    ).first_or_404()

    # Si déjà confirmé en base, retourner directement
    if p.statut == "SUCCES":
        return jsonify({"statut": "SUCCES", "message": "Paiement confirmé !"})
    if p.statut == "ECHEC":
        return jsonify({"statut": "ECHEC", "message": p.notes or "Paiement refusé."})
    if p.statut == "EXPIRE":
        return jsonify({"statut": "EXPIRE", "message": "Délai dépassé. Recommencez."})

    # Interroger l'API Airtel si on a un transaction_id
    if p.reference_externe:
        try:
            from airtel_money import verifier_statut
            r = verifier_statut(p.reference_externe)
            if r["statut"] == "SUCCESS":
                _activer_abonnement(p)
                return jsonify({"statut": "SUCCES", "message": "Paiement confirmé !"})
            elif r["statut"] in ("FAILED", "EXPIRED"):
                p.statut = "ECHEC" if r["statut"] == "FAILED" else "EXPIRE"
                p.notes  = r.get("message", "")
                db.session.commit()
                return jsonify({"statut": p.statut, "message": p.notes})
        except Exception as e:
            logger.warning(f"[Paiement] Polling Airtel erreur : {e}")

    return jsonify({"statut": "EN_ATTENTE", "message": "En attente de confirmation…"})


# ── Airtel Money — Webhook (callback automatique d'Airtel) ────────────────────
@app.route("/webhook/airtel", methods=["POST"])
@csrf.exempt  # Les webhooks externes ne peuvent pas envoyer de token CSRF
def webhook_airtel():
    """
    Reçoit les notifications automatiques d'Airtel après paiement du client.
    Airtel appelle cette URL avec le résultat de la transaction.
    """
    from airtel_money import valider_signature_webhook
    import json

    payload_bytes = request.get_data()
    signature     = request.headers.get("X-Airtel-Signature", "")

    # Vérifier la signature si configurée
    if not valider_signature_webhook(payload_bytes, signature):
        logger.warning("[Webhook Airtel] Signature invalide — requête ignorée.")
        return jsonify({"status": "SIGNATURE_INVALIDE"}), 401

    try:
        data = request.get_json(force=True) or {}
        logger.info(f"[Webhook Airtel] Reçu : {data}")

        # Extraire les infos de la transaction
        txn   = data.get("transaction", {}) or data.get("data", {}).get("transaction", {})
        ref   = txn.get("id") or txn.get("reference") or data.get("reference", "")
        statut_api = (txn.get("status") or data.get("status", {}).get("code", "")).upper()

        if not ref:
            logger.warning("[Webhook Airtel] Référence manquante dans le payload.")
            return jsonify({"status": "REF_MANQUANTE"}), 400

        # Retrouver le paiement
        p = Paiement.query.filter(
            (Paiement.reference_interne == ref) |
            (Paiement.reference_externe == ref)
        ).first()

        if not p:
            logger.warning(f"[Webhook Airtel] Paiement introuvable pour ref={ref}")
            return jsonify({"status": "INTROUVABLE"}), 404

        if p.statut == "SUCCES":
            # Idempotence — déjà traité
            return jsonify({"status": "DEJA_TRAITE"}), 200

        p.reponse_raw = json.dumps(data)

        if statut_api in ("TS", "SUCCESS", "200"):
            _activer_abonnement(p)
            logger.info(f"[Webhook Airtel] Succès — ref={ref} tenant={p.tenant_id}")
        else:
            p.statut = "ECHEC"
            p.notes  = f"Code Airtel : {statut_api}"
            db.session.commit()
            logger.info(f"[Webhook Airtel] Échec — ref={ref} code={statut_api}")

        return jsonify({"status": "OK"}), 200

    except Exception as e:
        logger.error(f"[Webhook Airtel] Erreur traitement : {e}")
        db.session.rollback()
        return jsonify({"status": "ERREUR_INTERNE"}), 500


# ── Helper : activer l'abonnement après paiement confirmé ─────────────────────
def _activer_abonnement(paiement: "Paiement"):
    """
    Appelé après confirmation d'un paiement (webhook ou polling).
    Met à jour le tenant : statut ACTIF, date_expiration prolongée.
    Envoie un email de confirmation.
    """
    from datetime import timezone
    p = paiement
    p.statut           = "SUCCES"
    p.date_confirmation = datetime.utcnow()

    t = p.tenant
    now = datetime.utcnow()

    # Prolonger depuis aujourd'hui ou depuis la date d'expiration si future
    base = t.date_expiration if (t.date_expiration and t.date_expiration > now) else now
    t.date_expiration = base + timedelta(days=30 * p.duree_mois)
    t.statut = "ACTIF"

    if p.plan_id:
        t.plan_id = p.plan_id

    db.session.commit()
    _cache_delete(f"{t.id}:")  # invalider le cache dashboard

    logger.info(
        f"[Abonnement] Tenant {t.id} activé jusqu'au "
        f"{t.date_expiration.strftime('%d/%m/%Y')} — "
        f"{p.duree_mois} mois via {p.moyen}"
    )

    # Email de confirmation
    try:
        msg = Message(
            subject=f"Abonnement PaieGabon activé — {t.denomination}",
            recipients=[u.email for u in t.utilisateurs if u.role == "TENANT_ADMIN" and u.email],
            body=(
                f"Bonjour,\n\n"
                f"Votre paiement de {float(p.montant):,.0f} FCFA a été confirmé.\n"
                f"Abonnement actif jusqu'au : {t.date_expiration.strftime('%d/%m/%Y')}\n"
                f"Référence : {p.reference_interne}\n\n"
                f"Merci de votre confiance.\n"
                f"L'équipe PaieGabon"
            ),
        )
        send_email_async(msg)
    except Exception as e:
        logger.warning(f"[Abonnement] Email de confirmation non envoyé : {e}")


@app.route("/paiement/confirmer", methods=["POST"])
@login_required
def paiement_confirmer():
    """Route de compatibilité — paiement manuel (admin valide manuellement)."""
    if current_user.is_super_admin: return redirect(url_for("admin_dashboard"))
    t = get_tenant()
    if not t: return redirect(url_for("login"))
    mode      = request.form.get("mode", "MANUEL")
    reference = request.form.get("reference", "").strip()
    duree     = int(request.form.get("duree", 1) or 1)
    if not reference:
        flash("Veuillez indiquer une référence de transaction.", "error")
        return redirect(url_for("paiement"))
    import uuid
    ref_interne = f"MAN-{t.id}-{uuid.uuid4().hex[:8].upper()}"
    p = Paiement(
        tenant_id=t.id, moyen="MANUEL", montant=float(t.plan.prix_mensuel) * duree if t.plan else 0,
        duree_mois=duree, plan_id=t.plan_id, reference_interne=ref_interne,
        reference_externe=reference, statut="EN_ATTENTE",
        notes=f"Paiement manuel déclaré par {current_user.email}",
    )
    db.session.add(p)
    t.statut = "PAIEMENT_EN_ATTENTE"
    db.session.commit()
    flash(f"Paiement {mode} (réf. {reference}) enregistré. Activation sous 48h après vérification.", "success")
    return redirect(url_for("parametres"))


# ══════════════════════════════════════════════════════════════════════════════
# CINETPAY — Paiement multi-opérateurs (Airtel, Moov, Visa, Mastercard)
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/paiement/cinetpay/initier", methods=["POST"])
@login_required
def paiement_cinetpay_initier():
    """
    Initie un paiement CinetPay.
    Crée une session et redirige le client vers la page de paiement CinetPay
    où il choisit son moyen : Airtel Money, Moov Money ou carte bancaire.
    """
    if current_user.is_super_admin: return redirect(url_for("admin_dashboard"))
    t = get_tenant()
    if not t: return redirect(url_for("login"))

    duree   = int(request.form.get("duree", 1) or 1)
    plan_id = request.form.get("plan_id", type=int) or t.plan_id
    plan    = Plan.query.get(plan_id) if plan_id else t.plan

    if not plan:
        flash("Plan introuvable.", "error")
        return redirect(url_for("paiement"))

    montant = float(plan.prix_mensuel) * duree

    import uuid
    reference = f"CP-{t.id}-{uuid.uuid4().hex[:10].upper()}"

    # Récupérer l'admin du tenant pour pré-remplir les infos client
    admin = Utilisateur.query.filter_by(tenant_id=t.id, role="TENANT_ADMIN").first()
    nom_client   = admin.nom_complet if admin else t.denomination
    email_client = admin.email if admin else ""

    # Enregistrer la tentative
    p = Paiement(
        tenant_id=t.id,
        moyen="CINETPAY",
        montant=montant,
        duree_mois=duree,
        plan_id=plan.id,
        reference_interne=reference,
        statut="EN_ATTENTE",
    )
    db.session.add(p)
    db.session.commit()

    try:
        from cinetpay import initier_paiement, CinetPayConfigError
        resultat = initier_paiement(
            reference=reference,
            montant=montant,
            description=f"PaieGabon {plan.nom} — {duree} mois — {t.denomination}",
            nom_client=nom_client,
            email_client=email_client,
        )

        import json
        p.reponse_raw = json.dumps(resultat.get("raw", {}))

        if resultat["success"]:
            p.reference_externe = resultat.get("payment_token", "")
            db.session.commit()
            logger.info(f"[CinetPay] Session créée — ref={reference} tenant={t.id}")
            # Rediriger directement vers la page CinetPay
            return redirect(resultat["payment_url"])
        else:
            p.statut = "ECHEC"
            p.notes  = resultat["message"]
            db.session.commit()
            flash(f"Erreur CinetPay : {resultat['message']}", "error")
            return redirect(url_for("paiement"))

    except Exception as e:
        p.statut = "ECHEC"
        p.notes  = str(e)
        db.session.commit()
        logger.error(f"[CinetPay] Erreur initiation : {e}")
        flash("Erreur de connexion CinetPay. Réessayez ou contactez le support.", "error")
        return redirect(url_for("paiement"))


@app.route("/paiement/cinetpay/retour")
@login_required
def paiement_cinetpay_retour():
    """
    Page de retour après la page de paiement CinetPay.
    CinetPay redirige ici après que le client ait terminé (succès ou annulation).
    On affiche un message d'attente pendant que le webhook confirme.
    """
    if current_user.is_super_admin: return redirect(url_for("admin_dashboard"))
    t = get_tenant()
    if not t: return redirect(url_for("login"))

    transaction_id = request.args.get("transaction_id", "")
    # Chercher le paiement par référence interne ou externe
    p = None
    if transaction_id:
        p = Paiement.query.filter(
            (Paiement.reference_interne == transaction_id) |
            (Paiement.reference_externe == transaction_id),
            Paiement.tenant_id == t.id
        ).first()

    # Vérification immédiate du statut
    if p and p.statut == "EN_ATTENTE" and (p.reference_interne or p.reference_externe):
        try:
            from cinetpay import verifier_statut
            ref = p.reference_interne
            r   = verifier_statut(ref)
            if r["statut"] == "ACCEPTED":
                _activer_abonnement(p)
                flash("Paiement confirmé ! Votre abonnement est actif.", "success")
                return redirect(url_for("dashboard"))
            elif r["statut"] in ("REFUSED", "CANCELLED"):
                p.statut = "ECHEC"
                p.notes  = r.get("message", "Paiement refusé ou annulé.")
                db.session.commit()
        except Exception as e:
            logger.warning(f"[CinetPay] Vérification retour échouée : {e}")

    if p and p.statut == "SUCCES":
        flash("Paiement confirmé ! Votre abonnement est actif.", "success")
        return redirect(url_for("dashboard"))

    # Afficher la page d'attente (le webhook va confirmer dans quelques secondes)
    return render_template("tenant/paiement_cinetpay_retour.html",
                           paiement=p, tenant=t,
                           transaction_id=transaction_id)


@app.route("/paiement/cinetpay/statut/<reference>")
@login_required
def paiement_cinetpay_statut(reference):
    """Endpoint JSON pour le polling côté client sur la page de retour."""
    t = get_tenant()
    if not t: return jsonify({"statut": "ERREUR"}), 401

    p = Paiement.query.filter(
        (Paiement.reference_interne == reference),
        Paiement.tenant_id == t.id
    ).first_or_404()

    if p.statut == "SUCCES":
        return jsonify({"statut": "SUCCES", "message": "Paiement confirmé !"})
    if p.statut == "ECHEC":
        return jsonify({"statut": "ECHEC", "message": p.notes or "Paiement refusé."})

    # Vérification active
    try:
        from cinetpay import verifier_statut
        r = verifier_statut(reference)
        if r["statut"] == "ACCEPTED":
            _activer_abonnement(p)
            return jsonify({"statut": "SUCCES", "message": "Paiement confirmé !"})
        elif r["statut"] in ("REFUSED", "CANCELLED"):
            p.statut = "ECHEC"
            p.notes  = r.get("message", "")
            db.session.commit()
            return jsonify({"statut": "ECHEC", "message": p.notes})
    except Exception as e:
        logger.warning(f"[CinetPay] Polling statut erreur : {e}")

    return jsonify({"statut": "EN_ATTENTE", "message": "Vérification en cours…"})


@app.route("/webhook/cinetpay", methods=["POST"])
@csrf.exempt
def webhook_cinetpay():
    """
    Reçoit les notifications automatiques de CinetPay.
    Appelé par CinetPay dès que le paiement est confirmé ou refusé.
    """
    import json
    try:
        data = request.get_json(force=True) or request.form.to_dict()
        logger.info(f"[Webhook CinetPay] Reçu : {data}")

        from cinetpay import valider_webhook
        if not valider_webhook(data):
            return jsonify({"status": "SITE_ID_INVALIDE"}), 401

        # Extraire la référence de transaction
        ref = (data.get("cpm_trans_id") or data.get("transaction_id")
               or data.get("metadata", ""))
        statut_api = (data.get("cpm_result") or data.get("status") or "").upper()

        if not ref:
            logger.warning("[Webhook CinetPay] Référence manquante.")
            return jsonify({"status": "REF_MANQUANTE"}), 400

        p = Paiement.query.filter_by(reference_interne=ref).first()
        if not p:
            # Essayer avec reference_externe
            token = data.get("cpm_payment_config") or data.get("payment_token", "")
            p = Paiement.query.filter_by(reference_externe=token).first() if token else None

        if not p:
            logger.warning(f"[Webhook CinetPay] Paiement introuvable ref={ref}")
            return jsonify({"status": "INTROUVABLE"}), 404

        if p.statut == "SUCCES":
            return jsonify({"status": "DEJA_TRAITE"}), 200

        p.reponse_raw = json.dumps(data)

        # "00" = succès chez CinetPay
        if statut_api in ("00", "ACCEPTED", "SUCCESS"):
            _activer_abonnement(p)
            logger.info(f"[Webhook CinetPay] Succès — ref={ref} tenant={p.tenant_id}")
        else:
            p.statut = "ECHEC"
            p.notes  = f"Code CinetPay : {statut_api}"
            db.session.commit()
            logger.info(f"[Webhook CinetPay] Échec — ref={ref} code={statut_api}")

        return jsonify({"status": "OK"}), 200

    except Exception as e:
        logger.error(f"[Webhook CinetPay] Erreur : {e}")
        db.session.rollback()
@app.route("/parametres")
@tenant_required
def parametres():
    t=get_tenant()
    # Passer tous les plans actifs pour l'onglet abonnement
    plans_dispo = Plan.query.filter_by(actif=True).order_by(Plan.prix_mensuel.asc()).all()
    return render_template("tenant/parametres.html", tenant=t,
        rubriques=RubriquePaie.query.filter_by(actif=True).all(),
        categories=CategorieEmploi.query.filter_by(tenant_id=t.id).all(),
        users=Utilisateur.query.filter_by(tenant_id=t.id).all(),
        plans_dispo=plans_dispo)

@app.route("/parametres/logo", methods=["POST"])
@login_required
def parametres_logo():
    if current_user.is_super_admin: return redirect(url_for("admin_dashboard"))
    t = get_tenant()
    if not t: return redirect(url_for("login"))
    logo_file = request.files.get("logo")
    if logo_file and logo_file.filename:
        import base64
        file_data = logo_file.read()
        if len(file_data) > 1_000_000:
            flash("Fichier trop volumineux. Maximum 1 Mo.", "error")
            return redirect(url_for("parametres"))

        # ── Validation de l'extension ─────────────────────────────────────────
        ext = logo_file.filename.rsplit(".", 1)[-1].lower() if "." in logo_file.filename else ""
        EXTENSIONS_AUTORISEES = {"png", "jpg", "jpeg", "gif", "webp"}
        if ext not in EXTENSIONS_AUTORISEES:
            flash("Format non autorisé. Utilisez PNG, JPG, JPEG, GIF ou WEBP (pas SVG).", "error")
            return redirect(url_for("parametres"))

        # ── Validation du MIME réel (magic bytes) — pas seulement l'extension ─
        MAGIC = {
            b"\x89PNG":   "image/png",
            b"\xff\xd8\xff": "image/jpeg",
            b"GIF8":      "image/gif",
            b"RIFF":      None,  # WebP — vérification complémentaire ci-dessous
        }
        detected_mime = None
        for magic, mime in MAGIC.items():
            if file_data[:len(magic)] == magic:
                if magic == b"RIFF" and file_data[8:12] == b"WEBP":
                    detected_mime = "image/webp"
                else:
                    detected_mime = mime
                break
        if not detected_mime:
            flash("Le contenu du fichier ne correspond pas à une image valide.", "error")
            return redirect(url_for("parametres"))

        b64 = base64.b64encode(file_data).decode("utf-8")
        logo_data = f"data:{detected_mime};base64,{b64}"
        try:
            db.session.execute(db.text("UPDATE tenants SET logo_url = :logo WHERE id = :id"),{"logo": logo_data, "id": t.id})
            db.session.commit(); db.session.expire(t)
            flash("Logo mis à jour avec succès.", "success")
        except Exception as e:
            db.session.rollback(); flash(f"Erreur: {str(e)}", "error")
    else:
        flash("Aucun fichier sélectionné.", "error")
    return redirect(url_for("parametres"))

@app.route("/parametres/logo/supprimer", methods=["POST"])
@login_required
def parametres_logo_supprimer():
    t = get_tenant()
    if not t: return redirect(url_for("login"))
    t.logo_url = None; db.session.commit()
    flash("Logo supprime.", "success")
    return redirect(url_for("parametres"))

@app.route("/parametres/modele-bulletin", methods=["POST"])
@login_required
def parametres_modele_bulletin():
    """Changer le modèle d'impression des bulletins."""
    if current_user.is_super_admin: return redirect(url_for("admin_dashboard"))
    t = get_tenant()
    if not t: return redirect(url_for("login"))
    modele = request.form.get("modele_bulletin", "classique")
    if modele not in ("classique", "moderne", "minimaliste"):
        modele = "classique"
    t.modele_bulletin = modele
    db.session.commit()
    flash(f"Modèle d'impression « {modele.capitalize()} » appliqué.", "success")
    return redirect(url_for("parametres"))

@app.route("/parametres/societe", methods=["POST"])
@tenant_required
@can_edit
def parametres_societe():
    t=get_tenant()
    for f in ["denomination","sigle","activite","secteur","nif","numero_cnss","numero_cnamgs","adresse","boite_postale","telephone","ville","region"]:
        try: setattr(t,f,request.form.get(f,"").strip() or None)
        except: pass
    db.session.commit(); flash("Informations mises à jour.","success")
    return redirect(url_for("parametres"))


@app.route("/parametres/demande-changement-plan", methods=["POST"])
@login_required
def demande_changement_plan():
    if current_user.is_super_admin: return redirect(url_for("admin_dashboard"))
    t = get_tenant()
    if not t: return redirect(url_for("login"))
    if not current_user.is_tenant_admin: abort(403)
    plan_souhaite_id = request.form.get("plan_id", type=int)
    motif = request.form.get("motif", "").strip()
    plan_souhaite = Plan.query.get(plan_souhaite_id) if plan_souhaite_id else None
    if not plan_souhaite:
        flash("Plan invalide.", "error")
        return redirect(url_for("parametres"))
    # Enregistrer la demande dans les notes + changer statut
    note_demande = (
        f"[DEMANDE CHANGEMENT PLAN — {datetime.now().strftime('%d/%m/%Y %H:%M')}] "
        f"Plan souhaité : {plan_souhaite.nom} ({int(plan_souhaite.prix_mensuel):,} FCFA/mois). "
        f"Motif : {motif or 'Non précisé'}. "
        f"Demandé par : {current_user.nom_complet} ({current_user.email})."
    )
    # Ajouter à la suite des notes existantes
    t.notes = (t.notes or "") + ("\n" if t.notes else "") + note_demande
    t.statut = "PAIEMENT_EN_ATTENTE"
    db.session.commit()
    flash(f"Demande de passage au plan « {plan_souhaite.nom} » enregistrée. L'équipe PaieGabon vous contactera sous 24h pour finaliser.", "success")
    return redirect(url_for("parametres"))

@app.route("/parametres/annuler-abonnement", methods=["POST"])
@login_required
def annuler_abonnement():
    if current_user.is_super_admin: return redirect(url_for("admin_dashboard"))
    t = get_tenant()
    if not t: return redirect(url_for("login"))
    if not current_user.is_tenant_admin: abort(403)
    motif = request.form.get("motif", "").strip()
    t.statut = "ANNULATION_DEMANDEE"
    t.notes = f"Annulation demandée le {datetime.now().strftime('%d/%m/%Y')}. Motif: {motif}"
    db.session.commit()
    flash("Demande d annulation enregistrée. L equipe PaieGabon vous contactera sous 48h.", "success")
    return redirect(url_for("parametres"))

# ── Utilisateurs ──────────────────────────────────────────────────────────────
@app.route("/utilisateurs")
@login_required
def utilisateurs():
    if current_user.is_super_admin: return redirect(url_for("admin_dashboard"))
    t = get_tenant()
    if not t: return redirect(url_for("login"))
    liste = Utilisateur.query.filter_by(tenant_id=t.id).order_by(Utilisateur.nom).all()
    return render_template("tenant/utilisateurs.html", tenant=t, utilisateurs=liste, users=liste)

@app.route("/utilisateurs/nouveau", methods=["GET","POST"])
@login_required
def utilisateur_nouveau():
    if current_user.is_super_admin: return redirect(url_for("admin_dashboard"))
    t = get_tenant()
    if not t: return redirect(url_for("login"))
    if not current_user.is_tenant_admin:
        flash("Réservé à l administrateur.", "error")
        return redirect(url_for("utilisateurs"))
    # ── Vérifier la limite dès le GET (bloquer l'accès au formulaire) ───────
    if t.plan and t.plan.max_utilisateurs:
        nb_actuel = Utilisateur.query.filter_by(tenant_id=t.id, actif=True).count()
        if nb_actuel >= t.plan.max_utilisateurs:
            flash(
                f"Limite atteinte — Plan « {t.plan.nom} » : "
                f"{t.plan.max_utilisateurs} utilisateur(s) maximum "
                f"(vous en avez {nb_actuel}). "
                f"Passez au plan supérieur pour en ajouter d'autres.",
                "error"
            )
            return redirect(url_for("utilisateurs"))

    if request.method == "GET":
        nb_utilisateurs = Utilisateur.query.filter_by(tenant_id=t.id, actif=True).count()
        return render_template("tenant/utilisateur_form.html", tenant=t,
            nb_utilisateurs=nb_utilisateurs)
    email = request.form.get("email", "").strip().lower()
    nom = request.form.get("nom", "").strip().upper()
    prenom = request.form.get("prenom", "").strip()
    password = request.form.get("password", "")
    role = request.form.get("role", "GESTIONNAIRE")
    if not email or not nom or not password:
        flash("Veuillez remplir tous les champs.", "error")
        return render_template("tenant/utilisateur_form.html", tenant=t)
    if Utilisateur.query.filter_by(email=email).first():
        flash("Email déjà utilisé.", "error")
        return render_template("tenant/utilisateur_form.html", tenant=t)
    u = Utilisateur(nom=nom, prenom=prenom, email=email, role=role, tenant_id=t.id, actif=True)
    u.set_password(password)
    db.session.add(u); db.session.commit()
    flash(f"Utilisateur {u.nom_complet} créé.", "success")
    return redirect(url_for("utilisateurs"))

@app.route("/utilisateurs/<int:id>/toggle", methods=["POST"])
@login_required
def utilisateur_toggle(id):
    """Activer / désactiver un utilisateur."""
    if current_user.is_super_admin: return redirect(url_for("admin_dashboard"))
    t = get_tenant()
    if not t: return redirect(url_for("login"))
    if not current_user.is_tenant_admin:
        flash("Réservé à l'administrateur.", "error")
        return redirect(url_for("utilisateurs"))
    u = Utilisateur.query.filter_by(id=id, tenant_id=t.id).first_or_404()
    if u.id == current_user.id:
        flash("Vous ne pouvez pas vous désactiver vous-même.", "error")
        return redirect(url_for("utilisateurs"))
    u.actif = not u.actif
    db.session.commit()
    etat = "activé" if u.actif else "désactivé"
    flash(f"Utilisateur {u.nom_complet} {etat}.", "success")
    return redirect(url_for("utilisateurs"))

# ── Journaliers ───────────────────────────────────────────────────────────────
@app.route("/journaliers")
@login_required
def journaliers():
    if current_user.is_super_admin: return redirect(url_for("admin_dashboard"))
    t = get_tenant()
    if not t: return redirect(url_for("login"))
    q    = request.args.get("q", "")
    page = request.args.get("page", 1, type=int)
    query = Journalier.query.filter_by(tenant_id=t.id)
    if q: query = query.filter(db.or_(Journalier.nom.ilike(f"%{q}%"), Journalier.prenom.ilike(f"%{q}%"), Journalier.profession.ilike(f"%{q}%")))
    pagination = query.order_by(Journalier.nom).paginate(page=page, per_page=25, error_out=False)
    _args = {k: v for k, v in request.args.items() if k != 'page'}
    _base = request.path + '?' + '&'.join(f'{k}={v}' for k, v in _args.items())
    _sep  = '&' if _args else '?'
    return render_template("tenant/journaliers.html",
        tenant=t, journaliers=pagination.items, pagination=pagination, q=q,
        pagination_base=_base + _sep)

@app.route("/journaliers/nouveau", methods=["GET","POST"])
@login_required
def journalier_nouveau():
    if current_user.is_super_admin: return redirect(url_for("admin_dashboard"))
    t = get_tenant()
    if not t: return redirect(url_for("login"))
    # Vérifier quota dès le GET
    q = t.quota_employes_info
    if q["max"] and q["plein"]:
        flash(
            f"Limite atteinte — Plan « {t.plan.nom} » : {q['max']} employé(s) maximum "
            f"({q['salaries']} salarié(s) + {q['journaliers']} journalier(s)). "
            f"Passez au plan supérieur.", "error"
        )
        return redirect(url_for("journaliers"))
    if request.method == "POST":
        if not t.peut_ajouter_employe:
            flash(f"Limite atteinte ({t.plan.max_salaries} employés). Passez au plan supérieur.","error")
            return redirect(url_for("journaliers"))
        j = Journalier(tenant_id=t.id,
            nom=request.form["nom"].strip().upper(),
            prenom=request.form["prenom"].strip(),
            telephone=request.form.get("telephone","").strip(),
            profession=request.form.get("profession","").strip().upper(),
            taux_horaire=float(request.form.get("taux_horaire",0) or 0),
            date_embauche=_parse_date(request.form.get("date_embauche")),
            date_debut=   _parse_date(request.form.get("date_debut")),
            date_fin=     _parse_date(request.form.get("date_fin")),
            nationalite=  request.form.get("nationalite","").strip() or None,
            statut="ACTIF")
        db.session.add(j); db.session.commit()
        flash(f"Journalier {j.nom_complet} créé.", "success")
        return redirect(url_for("journaliers"))
    return render_template("tenant/journalier_form.html", tenant=t, journalier=None)

@app.route("/journaliers/<int:id>")
@login_required
def journalier_detail(id):
    """Fiche détail d'un journalier avec historique de pointage."""
    if current_user.is_super_admin: return redirect(url_for("admin_dashboard"))
    t = get_tenant()
    if not t: return redirect(url_for("login"))
    j = Journalier.query.filter_by(id=id, tenant_id=t.id).first_or_404()

    # Feuilles de paie
    feuilles = FeuillePaieJournalier.query.filter_by(
        journalier_id=id, tenant_id=t.id
    ).order_by(FeuillePaieJournalier.date_fin.desc()).all()
    total_percu = sum(float(f.montant_brut or 0) for f in feuilles if f.statut == "PAYÉ")

    # Affectation site courante
    aff = AffectationSite.query.filter_by(
        journalier_id=id, tenant_id=t.id, actif=True).first()

    # ── Historique des pointages ──────────────────────────────────────────────
    nb_jours = request.args.get("nb_jours", type=int, default=30)
    nb_jours = min(max(nb_jours, 7), 90)
    date_fin   = datetime.now().date()
    date_debut = date_fin - timedelta(days=nb_jours - 1)

    pts_hist = Pointage.query.filter_by(tenant_id=t.id, journalier_id=id)        .filter(Pointage.date_pointage >= date_debut,
                Pointage.date_pointage <= date_fin)        .order_by(Pointage.date_pointage.desc()).all()

    nb_presences   = sum(1 for p in pts_hist if p.present)
    nb_absences    = sum(1 for p in pts_hist if p.absent)
    nb_non_pointes = nb_jours - len(pts_hist)
    h_normales_tot = round(sum(float(p.heures_normales or 0) for p in pts_hist if p.present), 1)
    h_sup_tot      = round(sum(float(p.heures_sup or 0) for p in pts_hist if p.present), 1)
    taux_presence  = round(nb_presences / (nb_presences + nb_absences) * 100
                           ) if (nb_presences + nb_absences) > 0 else 0

    return render_template("tenant/journalier_detail.html",
        journalier=j, tenant=t, feuilles=feuilles,
        total_percu=total_percu, aff=aff,
        pts_hist=pts_hist, nb_jours=nb_jours,
        nb_presences=nb_presences, nb_absences=nb_absences,
        nb_non_pointes=nb_non_pointes,
        h_normales_tot=h_normales_tot, h_sup_tot=h_sup_tot,
        taux_presence=taux_presence,
        date_debut_hist=date_debut, date_fin_hist=date_fin)

@app.route("/journaliers/<int:id>/modifier", methods=["GET","POST"])
@login_required
def journalier_modifier(id):
    t = get_tenant()
    if not t: return redirect(url_for("login"))
    j = Journalier.query.filter_by(id=id, tenant_id=t.id).first_or_404()
    if request.method == "POST":
        j.nom=request.form["nom"].strip().upper(); j.prenom=request.form["prenom"].strip()
        j.telephone=request.form.get("telephone","").strip()
        j.profession=request.form.get("profession","").strip().upper()
        j.taux_horaire=float(request.form.get("taux_horaire",0) or 0)
        j.date_embauche=_parse_date(request.form.get("date_embauche"))
        j.date_debut=   _parse_date(request.form.get("date_debut"))
        j.date_fin=     _parse_date(request.form.get("date_fin"))
        j.nationalite=  request.form.get("nationalite","").strip() or None
        j.statut=request.form.get("statut","ACTIF")
        # ── Affectation site ──────────────────────────────────────────────
        site_id = request.form.get("site_id", type=int)
        if site_id:
            aff_prev = AffectationSite.query.filter_by(
                journalier_id=j.id, tenant_id=t.id, actif=True).first()
            if aff_prev and aff_prev.site_id != site_id:
                aff_prev.actif    = False
                aff_prev.date_fin = date.today()
                aff_prev.motif    = "Réaffecté via formulaire journalier"
            if not aff_prev or aff_prev.site_id != site_id:
                db.session.add(AffectationSite(
                    tenant_id=t.id, site_id=site_id, journalier_id=j.id,
                    date_debut=date.today(), actif=True,
                    cree_par=current_user.email))
        elif request.form.get("retirer_site"):
            aff = AffectationSite.query.filter_by(
                journalier_id=j.id, tenant_id=t.id, actif=True).first()
            if aff:
                aff.actif    = False
                aff.date_fin = date.today()
                aff.motif    = "Retiré via formulaire journalier"
        db.session.commit()
        flash("Journalier mis à jour.", "success")
        return redirect(url_for("journaliers"))
    aff_actuelle = AffectationSite.query.filter_by(
        journalier_id=id, tenant_id=t.id, actif=True).first()
    sites = Site.query.filter_by(tenant_id=t.id, actif=True).order_by(Site.nom).all()
    return render_template("tenant/journalier_form.html", tenant=t, journalier=j,
        sites=sites, aff_actuelle=aff_actuelle)

# ── Pointage ──────────────────────────────────────────────────────────────────
@app.route("/pointage")
@login_required
def pointage():
    if current_user.is_super_admin: return redirect(url_for("admin_dashboard"))
    t = get_tenant()
    if not t: return redirect(url_for("login"))
    now = datetime.now()
    date_str = request.args.get("date", now.strftime("%Y-%m-%d"))
    try: date_sel = datetime.strptime(date_str, "%Y-%m-%d").date()
    except: date_sel = now.date()
    # ── Filtre par site ───────────────────────────────────────────────────────
    site_filtre_id = request.args.get("site_id", type=int)
    sites_list = Site.query.filter_by(tenant_id=t.id, actif=True).order_by(Site.nom).all()
    site_filtre = Site.query.get(site_filtre_id) if site_filtre_id else None

    if site_filtre_id:
        # Salariés affectés à ce site
        ids_sal = [a.salarie_id for a in AffectationSite.query.filter_by(
            tenant_id=t.id, site_id=site_filtre_id, actif=True
        ).filter(AffectationSite.salarie_id.isnot(None)).all()]
        ids_jour = [a.journalier_id for a in AffectationSite.query.filter_by(
            tenant_id=t.id, site_id=site_filtre_id, actif=True
        ).filter(AffectationSite.journalier_id.isnot(None)).all()]
        salaries_list   = Salarie.query.filter(
            Salarie.tenant_id==t.id, Salarie.statut=="ACTIF",
            Salarie.id.in_(ids_sal)
        ).order_by(Salarie.nom).all()
        journaliers_list = Journalier.query.filter(
            Journalier.tenant_id==t.id, Journalier.statut=="ACTIF",
            Journalier.id.in_(ids_jour)
        ).order_by(Journalier.nom).all()
    else:
        salaries_list    = Salarie.query.filter_by(tenant_id=t.id, statut="ACTIF").order_by(Salarie.nom).all()
        journaliers_list = Journalier.query.filter_by(tenant_id=t.id, statut="ACTIF").order_by(Journalier.nom).all()

    pts_salaries    = {p.salarie_id:    p for p in Pointage.query.filter_by(tenant_id=t.id, date_pointage=date_sel).filter(Pointage.salarie_id.isnot(None)).all()}
    pts_journaliers = {p.journalier_id: p for p in Pointage.query.filter_by(tenant_id=t.id, date_pointage=date_sel).filter(Pointage.journalier_id.isnot(None)).all()}
    nb_presents_sal  = sum(1 for p in pts_salaries.values()    if p.present)
    nb_presents_jour = sum(1 for p in pts_journaliers.values() if p.present)
    nb_absents       = sum(1 for p in list(pts_salaries.values())+list(pts_journaliers.values()) if p.absent)
    lundi   = date_sel - timedelta(days=date_sel.weekday())
    semaine = [lundi + timedelta(days=i) for i in range(6)]

    # Affectation site de chaque travailleur pour affichage dans le pointage
    aff_sal  = {a.salarie_id:    a.site for a in AffectationSite.query.filter_by(tenant_id=t.id, actif=True).filter(AffectationSite.salarie_id.isnot(None)).all()}
    aff_jour = {a.journalier_id: a.site for a in AffectationSite.query.filter_by(tenant_id=t.id, actif=True).filter(AffectationSite.journalier_id.isnot(None)).all()}

    return render_template("tenant/pointage.html",
        tenant=t, date_sel=date_sel, semaine=semaine,
        date_hier=(date_sel  - timedelta(days=1)).strftime("%Y-%m-%d"),
        date_demain=(date_sel + timedelta(days=1)).strftime("%Y-%m-%d"),
        salaries=salaries_list, journaliers=journaliers_list,
        pts_salaries=pts_salaries, pts_journaliers=pts_journaliers,
        nb_presents_sal=nb_presents_sal, nb_presents_jour=nb_presents_jour,
        nb_absents=nb_absents, now=now,
        sites=sites_list, site_filtre=site_filtre,
        aff_sal=aff_sal, aff_jour=aff_jour)

@app.route("/pointage/individuel", methods=["GET","POST"])
@login_required
def pointage_individuel():
    """Pointage d'un seul salarié ou journalier."""
    if current_user.is_super_admin: return redirect(url_for("admin_dashboard"))
    t = get_tenant()
    if not t: return redirect(url_for("login"))

    date_str  = request.args.get("date", datetime.now().strftime("%Y-%m-%d"))
    type_w    = request.args.get("type", "sal")   # "sal" ou "jour"
    worker_id = request.args.get("id", type=int)

    try:   date_sel = datetime.strptime(date_str, "%Y-%m-%d").date()
    except: date_sel = datetime.now().date()

    if request.method == "POST":
        # Sauvegarder le pointage individuel
        date_p = datetime.strptime(
            request.form.get("date_pointage", date_str), "%Y-%m-%d").date()
        wtype  = request.form.get("worker_type", "sal")
        wid    = int(request.form.get("worker_id", 0))
        present = request.form.get("present") == "1"
        absent  = not present
        def _hm(val):
            """Valider et retourner un horaire HH:MM ou None."""
            v = (val or "").strip()
            if not v: return None
            import re
            return v if re.match(r"^\d{1,2}:\d{2}$", v) else None

        def _diff_hm(debut, fin):
            """Calculer la différence en heures entre deux horaires HH:MM."""
            try:
                h1,m1 = map(int, debut.split(":")); h2,m2 = map(int, fin.split(":"))
                diff = (h2*60+m2 - h1*60-m1) / 60
                return max(0.0, round(diff, 2))
            except: return 0.0

        # Horaires saisis
        em = _hm(request.form.get("entree_matin"))
        sm = _hm(request.form.get("sortie_matin"))
        ea = _hm(request.form.get("entree_apmidi"))
        sa = _hm(request.form.get("sortie_apmidi"))
        es = _hm(request.form.get("entree_sup"))
        ss = _hm(request.form.get("sortie_sup"))

        # Calcul auto des heures normales depuis les horaires
        h_normales_auto = 0.0
        if em and sm: h_normales_auto += _diff_hm(em, sm)
        if ea and sa: h_normales_auto += _diff_hm(ea, sa)
        h_normales_man = float(request.form.get("heures_normales", 0) or 0)
        # Priorité aux horaires si saisis, sinon saisie manuelle
        heures_normales_final = round(h_normales_auto, 2) if h_normales_auto > 0 else h_normales_man or 8

        # Heures sup depuis horaires
        h_sup_horaire = _diff_hm(es, ss) if (es and ss) else 0.0

        type_jour = request.form.get("type_jour", "NORMAL")

        # Reclasser les heures selon le type de jour
        h_sup_10_man = float(request.form.get("heures_sup_10",0) or 0)
        h_sup_30_man = float(request.form.get("heures_sup_30",0) or 0)
        h_sup_40_man = float(request.form.get("heures_sup_40",0) or 0)
        h_sup_70_man = float(request.form.get("heures_sup_70",0) or 0)

        if type_jour == "DIMANCHE":
            # Tout va en +40% (dimanche)
            h_sup_40_final = round(h_sup_horaire + heures_normales_final, 2)
            h_sup_10_final = 0; h_sup_30_final = 0; h_sup_70_final = 0
            heures_normales_final = 0
        elif type_jour == "FERIE":
            # Tout va en +70% (jour férié)
            h_sup_70_final = round(h_sup_horaire + heures_normales_final, 2)
            h_sup_10_final = 0; h_sup_30_final = 0; h_sup_40_final = 0
            heures_normales_final = 0
        elif type_jour in ("CHOME_PAYE", "CHOME_RECUPERABLE"):
            # Présent mais jour chômé → heures normales conservées, pas de sup
            h_sup_10_final = 0; h_sup_30_final = 0; h_sup_40_final = 0; h_sup_70_final = 0
        else:
            # NORMAL
            h_sup_10_final = round(h_sup_horaire, 2) if h_sup_horaire > 0 else h_sup_10_man
            h_sup_30_final = h_sup_30_man
            h_sup_40_final = h_sup_40_man
            h_sup_70_final = h_sup_70_man

        kwargs = dict(
            heures_normales = heures_normales_final,
            motif_absence   = request.form.get("motif_absence","") if absent else None,
            entree_matin    = em, sortie_matin  = sm,
            entree_apmidi   = ea, sortie_apmidi = sa,
            entree_sup      = es, sortie_sup    = ss,
            type_jour       = type_jour,
        )
        if wtype == "sal":
            kwargs.update(dict(
                heures_sup_10 = h_sup_10_final,
                heures_sup_30 = h_sup_30_final,
                heures_sup_40 = h_sup_40_final,
                heures_sup_70 = h_sup_70_final,
            ))
            pt = Pointage.query.filter_by(
                tenant_id=t.id, date_pointage=date_p, salarie_id=wid).first()
            if not pt:
                pt = Pointage(tenant_id=t.id, date_pointage=date_p, salarie_id=wid)
                db.session.add(pt)
        else:
            # Journalier : gérer heures_sup selon type_jour
            if type_jour in ("DIMANCHE", "FERIE"):
                kwargs["heures_sup"]      = round(h_sup_horaire + heures_normales_final, 2)
                kwargs["heures_normales"] = 0
            else:
                kwargs["heures_sup"] = float(request.form.get("heures_sup",0) or 0)
            pt = Pointage.query.filter_by(
                tenant_id=t.id, date_pointage=date_p, journalier_id=wid).first()
            if not pt:
                pt = Pointage(tenant_id=t.id, date_pointage=date_p, journalier_id=wid)
                db.session.add(pt)
        pt.present = present
        pt.absent  = absent
        for k, v in kwargs.items():
            setattr(pt, k, v)
        db.session.commit()
        worker_name = (Salarie.query.get(wid) or Journalier.query.get(wid)).nom_complet
        flash(f"✅ Pointage de {worker_name} enregistré.", "success")
        # Rester sur la même page pour pointer la personne suivante
        redir = request.form.get("next_url") or f"/pointage/individuel?date={date_p}&type={wtype}&id={wid}"
        return redirect(redir)

    # GET : charger le travailleur sélectionné
    worker = pt_existant = None
    historique_30j = []
    stats_30j = {"presences": 0, "absences": 0, "h_normales": 0.0,
                 "h_sup": 0.0, "taux": 0}

    if worker_id:
        if type_w == "sal":
            worker = Salarie.query.filter_by(id=worker_id, tenant_id=t.id).first()
            pt_existant = Pointage.query.filter_by(
                tenant_id=t.id, date_pointage=date_sel, salarie_id=worker_id).first()
            if worker:
                date_debut_30 = date_sel - timedelta(days=29)
                historique_30j = Pointage.query.filter_by(
                    tenant_id=t.id, salarie_id=worker_id
                ).filter(
                    Pointage.date_pointage >= date_debut_30,
                    Pointage.date_pointage <= date_sel
                ).order_by(Pointage.date_pointage.desc()).all()
        else:
            worker = Journalier.query.filter_by(id=worker_id, tenant_id=t.id).first()
            pt_existant = Pointage.query.filter_by(
                tenant_id=t.id, date_pointage=date_sel, journalier_id=worker_id).first()
            if worker:
                date_debut_30 = date_sel - timedelta(days=29)
                historique_30j = Pointage.query.filter_by(
                    tenant_id=t.id, journalier_id=worker_id
                ).filter(
                    Pointage.date_pointage >= date_debut_30,
                    Pointage.date_pointage <= date_sel
                ).order_by(Pointage.date_pointage.desc()).all()

        # Stats sur les 30 jours
        if historique_30j:
            nb_p = sum(1 for p in historique_30j if p.present)
            nb_a = sum(1 for p in historique_30j if p.absent)
            hn   = round(sum(float(p.heures_normales or 0) for p in historique_30j if p.present), 1)
            if type_w == "sal":
                hs = round(sum(
                    float(p.heures_sup_10 or 0) + float(p.heures_sup_30 or 0) +
                    float(p.heures_sup_40 or 0) + float(p.heures_sup_70 or 0)
                    for p in historique_30j if p.present), 1)
            else:
                hs = round(sum(float(p.heures_sup or 0) for p in historique_30j if p.present), 1)
            total_ptg = nb_p + nb_a
            stats_30j = {
                "presences": nb_p, "absences": nb_a,
                "h_normales": hn, "h_sup": hs,
                "taux": round(nb_p / total_ptg * 100) if total_ptg > 0 else 0
            }

    # Listes pour la recherche
    salaries_list    = Salarie.query.filter_by(
        tenant_id=t.id, statut="ACTIF").order_by(Salarie.nom).all()
    journaliers_list = Journalier.query.filter_by(
        tenant_id=t.id, statut="ACTIF").order_by(Journalier.nom).all()

    return render_template("tenant/pointage_individuel.html",
        tenant=t, date_sel=date_sel,
        date_hier=(date_sel - timedelta(days=1)).strftime("%Y-%m-%d"),
        date_demain=(date_sel + timedelta(days=1)).strftime("%Y-%m-%d"),
        type_w=type_w, worker=worker, pt_existant=pt_existant,
        historique_30j=historique_30j, stats_30j=stats_30j,
        salaries=salaries_list, journaliers=journaliers_list,
        now=datetime.now())

@app.route("/pointage/supprimer/<int:ptg_id>", methods=["POST"])
@login_required
def pointage_supprimer(ptg_id):
    """Supprimer un pointage individuel."""
    if current_user.is_super_admin: return redirect(url_for("admin_dashboard"))
    t = get_tenant()
    if not t: return redirect(url_for("login"))
    pt = Pointage.query.filter_by(id=ptg_id, tenant_id=t.id).first_or_404()
    # Mémoriser le contexte pour rediriger au bon endroit
    next_url = request.form.get("next_url", "/pointage")
    date_str = str(pt.date_pointage)
    type_w   = "sal" if pt.salarie_id else "jour"
    wid      = pt.salarie_id or pt.journalier_id
    db.session.delete(pt)
    db.session.commit()
    flash("🗑️ Pointage supprimé.", "success")
    return redirect(next_url or f"/pointage/individuel?date={date_str}&type={type_w}&id={wid}")

@app.route("/pointage/sauvegarder", methods=["POST"])
@login_required
def pointage_sauvegarder():
    t = get_tenant()
    if not t: return redirect(url_for("login"))
    date_str = request.form.get("date_pointage")
    try: date_p = datetime.strptime(date_str, "%Y-%m-%d").date()
    except: flash("Date invalide.", "error"); return redirect(url_for("pointage"))
    nb = 0
    sel_sal  = {v for k,v in request.form.items() if k.startswith("sel_sal_")}
    sel_jour = {v for k,v in request.form.items() if k.startswith("sel_jour_")}
    for key, val in request.form.items():
        if key.startswith("sal_present_"):
            sid_str = key.replace("sal_present_","")
            if sid_str not in sel_sal: continue
            sid = int(sid_str); present = val == "1"; absent = not present
            pt = Pointage.query.filter_by(tenant_id=t.id, date_pointage=date_p, salarie_id=sid).first()
            if not pt: pt = Pointage(tenant_id=t.id, date_pointage=date_p, salarie_id=sid); db.session.add(pt)
            pt.present=present; pt.absent=absent
            pt.heures_normales = float(request.form.get(f"sal_heures_{sid}", 8) or 8)
            pt.heures_sup_10   = float(request.form.get(f"sal_sup10_{sid}", 0) or 0)
            pt.heures_sup_30   = float(request.form.get(f"sal_sup30_{sid}", 0) or 0)
            pt.heures_sup_40   = float(request.form.get(f"sal_sup40_{sid}", 0) or 0)
            pt.heures_sup_70   = float(request.form.get(f"sal_sup70_{sid}", 0) or 0)
            pt.motif_absence   = request.form.get(f"sal_motif_{sid}", "") if absent else None
            nb += 1
        if key.startswith("jour_present_"):
            jid_str = key.replace("jour_present_","")
            if jid_str not in sel_jour: continue
            jid = int(jid_str); present = val == "1"; absent = not present
            pt = Pointage.query.filter_by(tenant_id=t.id, date_pointage=date_p, journalier_id=jid).first()
            if not pt: pt = Pointage(tenant_id=t.id, date_pointage=date_p, journalier_id=jid); db.session.add(pt)
            pt.present=present; pt.absent=absent
            pt.heures_normales = float(request.form.get(f"jour_heures_{jid}", 8) or 8)
            pt.heures_sup      = float(request.form.get(f"jour_sup_{jid}", 0) or 0)
            pt.motif_absence   = request.form.get(f"jour_motif_{jid}", "") if absent else None
            nb += 1
    db.session.commit()
    flash(f"Pointage du {date_p.strftime('%d/%m/%Y')} sauvegardé ({nb} lignes).", "success")
    return redirect(url_for("pointage", date=date_str))

# ── Paie journaliers ──────────────────────────────────────────────────────────
@app.route("/journaliers/paie")
@login_required
def journaliers_paie():
    if current_user.is_super_admin: return redirect(url_for("admin_dashboard"))
    t = get_tenant()
    if not t: return redirect(url_for("login"))

    # ── Filtre par site ──────────────────────────────────────────────────────
    site_filtre_id = request.args.get("site_id", type=int)
    sites_list     = Site.query.filter_by(tenant_id=t.id, actif=True).order_by(Site.nom).all()
    site_filtre    = Site.query.get(site_filtre_id) if site_filtre_id else None
    statut_filtre  = request.args.get("statut", "")

    if site_filtre_id:
        ids_jour = [a.journalier_id for a in AffectationSite.query.filter_by(
            tenant_id=t.id, site_id=site_filtre_id, actif=True
        ).filter(AffectationSite.journalier_id.isnot(None)).all()]
        journaliers_list = Journalier.query.filter(
            Journalier.tenant_id==t.id, Journalier.statut=="ACTIF",
            Journalier.id.in_(ids_jour)
        ).order_by(Journalier.nom).all()
    else:
        journaliers_list = Journalier.query.filter_by(tenant_id=t.id, statut="ACTIF").order_by(Journalier.nom).all()

    # ── Feuilles filtrées ────────────────────────────────────────────────────
    q_feuilles = FeuillePaieJournalier.query.filter_by(tenant_id=t.id)
    if site_filtre_id:
        ids_jour_all = [a.journalier_id for a in AffectationSite.query.filter_by(
            tenant_id=t.id, site_id=site_filtre_id
        ).filter(AffectationSite.journalier_id.isnot(None)).all()]
        q_feuilles = q_feuilles.filter(FeuillePaieJournalier.journalier_id.in_(ids_jour_all))
    if statut_filtre:
        q_feuilles = q_feuilles.filter_by(statut=statut_filtre)
    page_f      = request.args.get("page", 1, type=int)
    # KPIs sur toutes les feuilles (sans pagination)
    feuilles_tous    = q_feuilles.order_by(FeuillePaieJournalier.date_fin.desc()).all()
    total_en_attente = sum(float(f.montant_brut or 0) for f in feuilles_tous if f.statut == "EN_ATTENTE")
    total_paye       = sum(float(f.montant_brut or 0) for f in feuilles_tous if f.statut == "PAYÉ")
    nb_en_attente    = sum(1 for f in feuilles_tous if f.statut == "EN_ATTENTE")
    q_feuilles = q_feuilles.options(joinedload(FeuillePaieJournalier.journalier))
    pagination_f     = q_feuilles.order_by(FeuillePaieJournalier.date_fin.desc()).paginate(page=page_f, per_page=25, error_out=False)
    feuilles         = pagination_f.items

    # Affectation site de chaque journalier (pour affichage dans la liste)
    aff_jour = {a.journalier_id: a.site for a in AffectationSite.query.filter_by(
        tenant_id=t.id, actif=True
    ).filter(AffectationSite.journalier_id.isnot(None)).all()}

    _args = {k: v for k, v in request.args.items() if k != 'page'}
    _base = request.path + '?' + '&'.join(f'{k}={v}' for k, v in _args.items())
    _sep  = '&' if _args else '?'
    return render_template("tenant/journaliers_paie.html",
        tenant=t, feuilles=feuilles, journaliers=journaliers_list,
        sites=sites_list, site_filtre=site_filtre, statut_filtre=statut_filtre,
        total_en_attente=total_en_attente, total_paye=total_paye,
        nb_en_attente=nb_en_attente, aff_jour=aff_jour,
        pagination=pagination_f, pagination_base=_base + _sep,
        now=datetime.now())

@app.route("/journaliers/paie/generer", methods=["POST"])
@login_required
def journaliers_paie_generer():
    t = get_tenant()
    if not t: return redirect(url_for("login"))
    date_debut = _parse_date(request.form.get("date_debut"))
    date_fin   = _parse_date(request.form.get("date_fin"))
    if not date_debut or not date_fin: flash("Dates invalides.", "error"); return redirect(url_for("journaliers_paie"))
    site_id      = request.form.get("site_id", type=int)
    taux_custom  = {}  # taux personnalisés par journalier
    heures_custom = {} # heures manuelles par journalier
    ids_coches   = request.form.getlist("journalier_ids")

    for key, val in request.form.items():
        if key.startswith("taux_") and val:
            try: taux_custom[int(key[5:])] = float(val)
            except: pass
        if key.startswith("heures_") and val:
            try: heures_custom[int(key[7:])] = float(val)
            except: pass

    if ids_coches:
        journaliers_a_payer = Journalier.query.filter(
            Journalier.tenant_id==t.id,
            Journalier.id.in_([int(i) for i in ids_coches])
        ).all()
    elif site_id:
        ids_site = [a.journalier_id for a in AffectationSite.query.filter_by(
            tenant_id=t.id, site_id=site_id, actif=True
        ).filter(AffectationSite.journalier_id.isnot(None)).all()]
        journaliers_a_payer = Journalier.query.filter(
            Journalier.tenant_id==t.id, Journalier.statut=="ACTIF",
            Journalier.id.in_(ids_site)
        ).all()
    else:
        journaliers_a_payer = Journalier.query.filter_by(tenant_id=t.id, statut="ACTIF").all()

    nb = 0
    for j in journaliers_a_payer:
        if str(j.id) not in ids_coches and ids_coches:
            continue
        taux = taux_custom.get(j.id, float(j.taux_horaire or 0))
        if j.id in heures_custom:
            total_h  = heures_custom[j.id]
            nb_jours = 1  # heures manuelles = considéré comme 1 entrée
        else:
            pts = Pointage.query.filter_by(tenant_id=t.id, journalier_id=j.id)                  .filter(Pointage.date_pointage>=date_debut, Pointage.date_pointage<=date_fin,
                          Pointage.present==True).all()
            total_h  = sum(float(p.heures_normales or 0)+float(p.heures_sup or 0) for p in pts)
            nb_jours = len(pts)
        if total_h <= 0 and nb_jours == 0: continue
        if FeuillePaieJournalier.query.filter_by(
            tenant_id=t.id, journalier_id=j.id,
            date_debut=date_debut, date_fin=date_fin).first(): continue
        db.session.add(FeuillePaieJournalier(
            tenant_id=t.id, journalier_id=j.id,
            date_debut=date_debut, date_fin=date_fin,
            nb_jours=nb_jours, total_heures=total_h,
            taux_horaire=taux, montant_brut=round(total_h*taux, 2),
            statut="EN_ATTENTE"))
        nb += 1
    db.session.commit()
    flash(f"{nb} feuille(s) générée(s).", "success")
    redirect_url = url_for("journaliers_paie")
    if site_id:
        redirect_url += f"?site_id={site_id}"
    return redirect(redirect_url)

@app.route("/journaliers/paie/<int:id>/payer", methods=["POST"])
@login_required
def journalier_payer(id):
    t = get_tenant()
    if not t: return redirect(url_for("login"))
    f = FeuillePaieJournalier.query.filter_by(id=id, tenant_id=t.id).first_or_404()
    f.statut="PAYÉ"; f.date_paiement=datetime.now().date(); db.session.commit()
    flash(f"Paiement de {f.journalier.nom_complet} enregistré.", "success")
    return redirect(url_for("journaliers_paie"))

@app.route("/journaliers/paie/<int:id>/modifier", methods=["POST"])
@login_required
def journalier_feuille_modifier(id):
    t = get_tenant()
    if not t: return redirect(url_for("login"))
    f = FeuillePaieJournalier.query.filter_by(id=id, tenant_id=t.id).first_or_404()
    f.montant_brut = float(request.form.get("montant_brut", f.montant_brut) or f.montant_brut)
    f.observation  = request.form.get("observation", "").strip()
    db.session.commit(); flash("Feuille modifiée.", "success")
    return redirect(url_for("journaliers_paie"))

@app.route("/journaliers/paie/<int:id>/supprimer", methods=["POST"])
@login_required
def journalier_feuille_supprimer(id):
    t = get_tenant()
    if not t: return redirect(url_for("login"))
    f = FeuillePaieJournalier.query.filter_by(id=id, tenant_id=t.id).first_or_404()
    db.session.delete(f); db.session.commit()
    flash("Feuille supprimée.", "success")
    return redirect(url_for("journaliers_paie"))

@app.route("/journaliers/paie/export")
@login_required
def journaliers_paie_export():
    """Export Excel des feuilles de paie journalier — filtré par site et/ou période."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    import io, calendar

    t = get_tenant()
    if not t: return redirect(url_for("login"))

    # ── Paramètres de filtre ──────────────────────────────────────────────────
    site_id    = request.args.get("site_id",    type=int)
    statut_f   = request.args.get("statut",     "")
    date_debut = _parse_date(request.args.get("date_debut", ""))
    date_fin   = _parse_date(request.args.get("date_fin",   ""))
    site       = Site.query.get(site_id) if site_id else None

    # ── Requête ───────────────────────────────────────────────────────────────
    q = FeuillePaieJournalier.query.filter_by(tenant_id=t.id)
    if site_id:
        ids_j = [a.journalier_id for a in AffectationSite.query.filter_by(
            tenant_id=t.id, site_id=site_id
        ).filter(AffectationSite.journalier_id.isnot(None)).all()]
        q = q.filter(FeuillePaieJournalier.journalier_id.in_(ids_j))
    if date_debut:
        q = q.filter(FeuillePaieJournalier.date_debut >= date_debut)
    if date_fin:
        q = q.filter(FeuillePaieJournalier.date_fin   <= date_fin)
    if statut_f:
        q = q.filter_by(statut=statut_f)
    feuilles = q.order_by(
        FeuillePaieJournalier.date_fin.desc(),
        FeuillePaieJournalier.journalier_id
    ).all()

    # ── Affectation site de chaque journalier ─────────────────────────────────
    aff_map = {}
    for a in AffectationSite.query.filter_by(tenant_id=t.id).filter(
            AffectationSite.journalier_id.isnot(None)).all():
        if a.journalier_id not in aff_map:
            aff_map[a.journalier_id] = a.site.nom if a.site else "—"

    # ── Styles communs ────────────────────────────────────────────────────────
    HDR_FONT   = Font(bold=True, color="FFFFFF", size=9)
    HDR_FILL   = PatternFill("solid", fgColor="1a2332")
    HDR_ALIGN  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    BODY_FONT  = Font(size=9)
    EVEN_FILL  = PatternFill("solid", fgColor="F7F8FA")
    TOTAL_FONT = Font(bold=True, size=10, color="FFFFFF")
    TOTAL_FILL = PatternFill("solid", fgColor="1a2332")
    MONEY_FMT  = '#,##0'
    thin       = Side(style="thin", color="D1D5DB")
    BORDER     = Border(left=thin, right=thin, top=thin, bottom=thin)
    CENTER     = Alignment(horizontal="center")
    RIGHT      = Alignment(horizontal="right")

    wb = Workbook()

    # ══════════════════════════════════════════════════════════════════════════
    # ONGLET 1 — Détail complet
    # ══════════════════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Détail"
    ws.freeze_panes = "A4"

    # Titre
    titre = (f"PAIE JOURNALIERS — {t.denomination}"
             + (f" — {site.nom}" if site else "")
             + (f" — {date_debut.strftime('%d/%m/%Y')} au {date_fin.strftime('%d/%m/%Y')}" if date_debut and date_fin else "")
             + f" — Édité le {datetime.now().strftime('%d/%m/%Y à %H:%M')}")
    ws.merge_cells("A1:K1")
    ws["A1"] = titre
    ws["A1"].font = Font(bold=True, size=12, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="1a2332")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    ws.append([])  # ligne vide

    # En-têtes
    hdrs = ["Journalier","Profession","Site","Période du","au",
            "Nb jours","Total heures","Taux/h (FCFA)","Montant brut (FCFA)","Statut","Date paiement"]
    ws.append(hdrs)
    for c_idx, h in enumerate(hdrs, 1):
        cell = ws.cell(row=3, column=c_idx, value=h)
        cell.font  = HDR_FONT
        cell.fill  = HDR_FILL
        cell.alignment = HDR_ALIGN
        cell.border = BORDER
    ws.row_dimensions[3].height = 18

    # Données
    total_montant  = 0
    total_jours    = 0
    total_heures   = 0
    for row_idx, f in enumerate(feuilles, 4):
        site_nom  = aff_map.get(f.journalier_id, "—")
        montant   = float(f.montant_brut  or 0)
        heures    = float(f.total_heures  or 0)
        jours     = int(f.nb_jours or 0)
        total_montant += montant
        total_jours   += jours
        total_heures  += heures
        row_data = [
            f.journalier.nom_complet,
            f.journalier.profession or "—",
            site_nom,
            f.date_debut.strftime("%d/%m/%Y") if f.date_debut else "",
            f.date_fin.strftime("%d/%m/%Y")   if f.date_fin   else "",
            jours,
            round(heures, 2),
            float(f.taux_horaire or 0),
            montant,
            f.statut,
            f.date_paiement.strftime("%d/%m/%Y") if f.date_paiement else "",
        ]
        ws.append(row_data)
        is_even = (row_idx % 2 == 0)
        for c_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=c_idx)
            cell.font   = BODY_FONT
            cell.border = BORDER
            if is_even: cell.fill = EVEN_FILL
            # Formats numériques
            if c_idx in (6, 7):   cell.alignment = CENTER
            if c_idx in (8, 9):
                cell.number_format = MONEY_FMT
                cell.alignment     = RIGHT
            # Statut coloré
            if c_idx == 10:
                cell.alignment = CENTER
                if val == "PAYÉ":
                    cell.font = Font(bold=True, color="065F46", size=9)
                else:
                    cell.font = Font(bold=True, color="92400E", size=9)

    # Ligne totaux
    ws.append([])
    tr = ws.max_row + 1
    totals = ["", "", "", "", "TOTAL", total_jours, round(total_heures,2),
              "", total_montant, "", ""]
    ws.append(totals)
    for c_idx, val in enumerate(totals, 1):
        cell = ws.cell(row=tr, column=c_idx)
        cell.font   = TOTAL_FONT
        cell.fill   = TOTAL_FILL
        cell.border = BORDER
        if c_idx in (8, 9):
            cell.number_format = MONEY_FMT
            cell.alignment     = RIGHT
        if c_idx == 5:
            cell.alignment = RIGHT

    # Largeurs colonnes
    col_widths = [28, 18, 20, 13, 13, 10, 13, 16, 20, 14, 15]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = w

    # ══════════════════════════════════════════════════════════════════════════
    # ONGLET 2 — Récap par site
    # ══════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Récap par site")
    ws2.freeze_panes = "A3"

    ws2.merge_cells("A1:F1")
    ws2["A1"] = f"RÉCAPITULATIF PAR SITE — {t.denomination}"
    ws2["A1"].font = Font(bold=True, size=12, color="FFFFFF")
    ws2["A1"].fill = PatternFill("solid", fgColor="374151")
    ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 20

    hdrs2 = ["Site","Nb journaliers","Nb jours total","Total heures","Montant brut (FCFA)","Statut"]
    ws2.append(hdrs2)
    for c_idx, h in enumerate(hdrs2, 1):
        cell = ws2.cell(row=2, column=c_idx, value=h)
        cell.font  = HDR_FONT
        cell.fill  = PatternFill("solid", fgColor="374151")
        cell.alignment = HDR_ALIGN
        cell.border = BORDER

    # Grouper par site
    from collections import defaultdict
    by_site = defaultdict(lambda: {"journaliers": set(), "jours": 0, "heures": 0.0,
                                    "montant": 0.0, "nb_payes": 0, "nb_total": 0})
    for f in feuilles:
        s_nom = aff_map.get(f.journalier_id, "Sans site")
        by_site[s_nom]["journaliers"].add(f.journalier_id)
        by_site[s_nom]["jours"]    += int(f.nb_jours or 0)
        by_site[s_nom]["heures"]   += float(f.total_heures or 0)
        by_site[s_nom]["montant"]  += float(f.montant_brut or 0)
        by_site[s_nom]["nb_total"] += 1
        if f.statut == "PAYÉ": by_site[s_nom]["nb_payes"] += 1

    grand_total = 0
    for row_idx, (s_nom, data) in enumerate(sorted(by_site.items()), 3):
        pct_paye = int(data["nb_payes"] / data["nb_total"] * 100) if data["nb_total"] else 0
        statut_txt = f"{data['nb_payes']}/{data['nb_total']} payé(s) ({pct_paye}%)"
        row_data = [
            s_nom,
            len(data["journaliers"]),
            data["jours"],
            round(data["heures"], 2),
            round(data["montant"], 2),
            statut_txt,
        ]
        ws2.append(row_data)
        grand_total += data["montant"]
        is_even = (row_idx % 2 == 0)
        for c_idx, val in enumerate(row_data, 1):
            cell = ws2.cell(row=row_idx, column=c_idx)
            cell.font   = BODY_FONT
            cell.border = BORDER
            if is_even: cell.fill = EVEN_FILL
            if c_idx == 5:
                cell.number_format = MONEY_FMT
                cell.alignment     = RIGHT
            if c_idx in (2, 3, 4):
                cell.alignment = CENTER

    # Total récap
    ws2.append([])
    tr2 = ws2.max_row + 1
    ws2.append(["TOTAL GÉNÉRAL", "", "", "", grand_total, ""])
    for c_idx in range(1, 7):
        cell = ws2.cell(row=tr2, column=c_idx)
        cell.font   = TOTAL_FONT
        cell.fill   = PatternFill("solid", fgColor="374151")
        cell.border = BORDER
        if c_idx == 5:
            cell.number_format = MONEY_FMT
            cell.alignment     = RIGHT

    for i, w in enumerate([28, 16, 14, 14, 22, 22], 1):
        ws2.column_dimensions[ws2.cell(1, i).column_letter].width = w

    # ══════════════════════════════════════════════════════════════════════════
    # ONGLET 3 — Récap par journalier
    # ══════════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Récap par journalier")
    ws3.freeze_panes = "A3"

    ws3.merge_cells("A1:G1")
    ws3["A1"] = f"RÉCAPITULATIF PAR JOURNALIER — {t.denomination}"
    ws3["A1"].font = Font(bold=True, size=12, color="FFFFFF")
    ws3["A1"].fill = PatternFill("solid", fgColor="065F46")
    ws3["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 20

    hdrs3 = ["Journalier","Profession","Site","Nb périodes","Nb jours","Total heures","Total perçu (FCFA)"]
    ws3.append(hdrs3)
    for c_idx, h in enumerate(hdrs3, 1):
        cell = ws3.cell(row=2, column=c_idx, value=h)
        cell.font  = HDR_FONT
        cell.fill  = PatternFill("solid", fgColor="065F46")
        cell.alignment = HDR_ALIGN
        cell.border = BORDER

    by_jour = defaultdict(lambda: {"nom":"","profession":"","site":"","periodes":0,"jours":0,"heures":0.0,"montant":0.0})
    for f in feuilles:
        jid = f.journalier_id
        by_jour[jid]["nom"]       = f.journalier.nom_complet
        by_jour[jid]["profession"]= f.journalier.profession or "—"
        by_jour[jid]["site"]      = aff_map.get(jid, "—")
        by_jour[jid]["periodes"]  += 1
        by_jour[jid]["jours"]     += int(f.nb_jours or 0)
        by_jour[jid]["heures"]    += float(f.total_heures or 0)
        by_jour[jid]["montant"]   += float(f.montant_brut or 0)

    grand_total3 = 0
    for row_idx, (jid, d) in enumerate(sorted(by_jour.items(), key=lambda x: x[1]["nom"]), 3):
        row_data = [d["nom"], d["profession"], d["site"],
                    d["periodes"], d["jours"], round(d["heures"],2), round(d["montant"],2)]
        ws3.append(row_data)
        grand_total3 += d["montant"]
        is_even = (row_idx % 2 == 0)
        for c_idx, val in enumerate(row_data, 1):
            cell = ws3.cell(row=row_idx, column=c_idx)
            cell.font   = BODY_FONT
            cell.border = BORDER
            if is_even: cell.fill = EVEN_FILL
            if c_idx == 7:
                cell.number_format = MONEY_FMT; cell.alignment = RIGHT
            if c_idx in (4, 5, 6): cell.alignment = CENTER

    ws3.append([])
    tr3 = ws3.max_row + 1
    ws3.append(["TOTAL GÉNÉRAL","","","","","", grand_total3])
    for c_idx in range(1, 8):
        cell = ws3.cell(row=tr3, column=c_idx)
        cell.font = TOTAL_FONT
        cell.fill = PatternFill("solid", fgColor="065F46")
        cell.border = BORDER
        if c_idx == 7:
            cell.number_format = MONEY_FMT; cell.alignment = RIGHT

    for i, w in enumerate([28, 18, 20, 12, 11, 13, 22], 1):
        ws3.column_dimensions[ws3.cell(1, i).column_letter].width = w

    # ── Export ────────────────────────────────────────────────────────────────
    out = io.BytesIO()
    wb.save(out); out.seek(0)

    parts = ["Paie_Journaliers"]
    if site:       parts.append(site.nom.replace(" ", "_"))
    if date_debut: parts.append(date_debut.strftime("%Y%m%d"))
    if date_fin:   parts.append("au" + date_fin.strftime("%Y%m%d"))
    parts.append(datetime.now().strftime("%Y%m%d"))
    fname = "_".join(parts) + ".xlsx"

    return send_file(out,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True, download_name=fname)

@app.route("/journaliers/paie/payer-selection", methods=["POST"])
@login_required
def journaliers_payer_selection():
    t = get_tenant()
    if not t: return redirect(url_for("login"))
    ids = [int(i) for i in request.form.get("feuille_ids","").split(",") if i.strip().isdigit()]
    nb = 0
    for fid in ids:
        f = FeuillePaieJournalier.query.filter_by(id=fid, tenant_id=t.id, statut="EN_ATTENTE").first()
        if f: f.statut="PAYÉ"; f.date_paiement=datetime.now().date(); nb+=1
    db.session.commit()
    flash(f"{nb} journalier(s) payé(s).", "success")
    return redirect(url_for("journaliers_paie"))

# ── Acomptes ──────────────────────────────────────────────────────────────────
@app.route("/acomptes")
@login_required
def acomptes():
    if current_user.is_super_admin: return redirect(url_for("admin_dashboard"))
    t = get_tenant()
    if not t: return redirect(url_for("login"))
    now = datetime.now()
    mois = request.args.get("mois", now.month, type=int)
    annee = request.args.get("annee", now.year, type=int)
    salarie_id = request.args.get("salarie_id", type=int)
    try:
        query = Acompte.query.filter_by(tenant_id=t.id, annee=annee, mois=mois)
        if salarie_id: query = query.filter_by(salarie_id=salarie_id)
        liste = query.order_by(Acompte.date_acompte.desc()).all()
    except Exception:
        db.create_all(); db.session.rollback(); liste = []
    total_mois       = sum(float(a.montant) for a in liste if a.statut != "ANNULE")
    total_en_attente = sum(float(a.montant) for a in liste if a.statut == "EN_ATTENTE")
    total_deduit     = sum(float(a.montant) for a in liste if a.statut == "DEDUIT")
    salaries_list = Salarie.query.filter_by(tenant_id=t.id, statut="ACTIF").order_by(Salarie.nom).all()
    return render_template("tenant/acomptes.html", tenant=t, liste=liste, salaries=salaries_list,
        mois=mois, annee=annee, now=now, total_mois=total_mois,
        total_en_attente=total_en_attente, total_deduit=total_deduit, MOIS_NOMS=PeriodePaie.MOIS_NOMS)

@app.route("/acomptes/nouveau", methods=["GET","POST"])
@login_required
def acompte_nouveau():
    if current_user.is_super_admin: return redirect(url_for("admin_dashboard"))
    t = get_tenant()
    if not t: return redirect(url_for("login"))
    if not current_user.can_edit: abort(403)
    salaries_list = Salarie.query.filter_by(tenant_id=t.id, statut="ACTIF").order_by(Salarie.nom).all()
    if request.method == "POST":
        salarie_id = request.form.get("salarie_id", type=int)
        montant    = float(request.form.get("montant", 0) or 0)
        date_ac    = _parse_date(request.form.get("date_acompte"))
        mois       = request.form.get("mois", type=int)
        annee      = request.form.get("annee", type=int)
        motif      = request.form.get("motif", "").strip()
        if not salarie_id or montant <= 0 or not date_ac:
            flash("Veuillez remplir tous les champs.", "error")
        else:
            contrat = Contrat.query.filter_by(salarie_id=salarie_id, tenant_id=t.id, actif=True).first()
            if contrat and montant > float(contrat.salaire_base) * 0.5:
                flash(f"Acompte maximum 50% du salaire de base ({float(contrat.salaire_base)*0.5:,.0f} FCFA).".replace(",", " "), "error")
                return render_template("tenant/acompte_form.html", tenant=t, salaries=salaries_list, now=datetime.now())
            db.session.add(Acompte(tenant_id=t.id, salarie_id=salarie_id, montant=montant,
                date_acompte=date_ac, mois=mois, annee=annee, motif=motif, statut="EN_ATTENTE"))
            db.session.commit()
            flash(f"Acompte de {montant:,.0f} FCFA enregistré.".replace(",", " "), "success")
            return redirect(url_for("acomptes", mois=mois, annee=annee))
    return render_template("tenant/acompte_form.html", tenant=t, salaries=salaries_list, now=datetime.now())

@app.route("/acomptes/<int:id>/valider", methods=["POST"])
@login_required
def acompte_valider(id):
    t = get_tenant()
    if not t: return redirect(url_for("login"))
    a = Acompte.query.filter_by(id=id, tenant_id=t.id).first_or_404()
    a.statut = "DEDUIT"; db.session.commit()
    flash("Acompte marqué comme déduit.", "success")
    return redirect(url_for("acomptes", mois=a.mois, annee=a.annee))

@app.route("/acomptes/<int:id>/annuler", methods=["POST"])
@login_required
def acompte_annuler(id):
    t = get_tenant()
    if not t: return redirect(url_for("login"))
    a = Acompte.query.filter_by(id=id, tenant_id=t.id).first_or_404()
    a.statut = "ANNULE"; db.session.commit()
    flash("Acompte annulé.", "success")
    return redirect(url_for("acomptes", mois=a.mois, annee=a.annee))

@app.route("/acomptes/<int:id>/supprimer", methods=["POST"])
@login_required
def acompte_supprimer(id):
    t = get_tenant()
    if not t: return redirect(url_for("login"))
    a = Acompte.query.filter_by(id=id, tenant_id=t.id).first_or_404()
    mois, annee = a.mois, a.annee
    db.session.delete(a); db.session.commit()
    flash("Acompte supprimé.", "success")
    return redirect(url_for("acomptes", mois=mois, annee=annee))

# ── Congés ────────────────────────────────────────────────────────────────────
@app.route("/conges")
@login_required
def conges():
    if current_user.is_super_admin: return redirect(url_for("admin_dashboard"))
    t = get_tenant()
    if not t: return redirect(url_for("login"))
    now   = datetime.now()
    annee = request.args.get("annee", now.year, type=int)
    q     = request.args.get("q", "")

    # ── Calcul du solde — Code du Travail gabonais ───────────────────────────
    # Art. 213 : 2 j/mois pour ≥ 18 ans | 2,5 j/mois pour < 18 ans
    # Allocation = max(Σ bruts 12 mois, dernier brut×12) / 288 × jours acquis
    # Exclusion : prime de transport (Art. 213 al. 3)

    def age_au_31_dec(salarie, annee_ref):
        """Âge du salarié au 31 décembre de l'année de référence."""
        if not salarie.date_naissance:
            return 99  # inconnu → adulte par défaut
        return annee_ref - salarie.date_naissance.year - (
            1 if salarie.date_naissance.replace(year=annee_ref) > datetime(annee_ref,12,31).date() else 0
        )

    def taux_conge(salarie, annee_ref):
        """2.5 j/mois si < 18 ans, sinon 2 j/mois."""
        return 2.5 if age_au_31_dec(salarie, annee_ref) < 18 else 2.0

    def calculer_solde_auto(salarie, annee_ref):
        """Jours acquis proratisés selon le taux applicable."""
        tx = taux_conge(salarie, annee_ref)
        if not salarie.date_embauche:
            return round(12 * tx, 1)
        emb         = salarie.date_embauche
        debut_annee = datetime(annee_ref, 1, 1).date()
        fin_annee   = datetime(annee_ref, 12, 31).date()
        debut_acq   = max(emb, debut_annee)
        fin_acq     = min(datetime.now().date(), fin_annee)
        if fin_acq < debut_acq:
            return 0.0
        mois_trav = min(round((fin_acq - debut_acq).days / 30.44, 1), 12)
        return round(mois_trav * tx, 1)

    def calculer_allocation_conge(salarie, jours_acquis, annee_ref):
        """
        Allocation congés = max(Σbruts12mois, dernierBrut×12) / 288 × jours_acquis
        Prime de transport exclue de la base (Art. 213 al. 3).
        """
        if jours_acquis <= 0:
            return 0.0, 0.0
        # Bulletins des 12 derniers mois
        from datetime import timedelta
        limite = datetime(annee_ref, 12, 31).date()
        debut  = (datetime(annee_ref, 12, 31) - timedelta(days=365)).date()
        buls   = BulletinPaie.query.filter(
            BulletinPaie.tenant_id  == salarie.tenant_id,
            BulletinPaie.salarie_id == salarie.id,
        ).join(PeriodePaie).filter(
            PeriodePaie.annee >= debut.year,
        ).order_by(PeriodePaie.annee.desc(), PeriodePaie.mois.desc()).limit(12).all()

        if not buls:
            # Pas de bulletins → estimer depuis le contrat actif
            contrat = next((c for c in salarie.contrats if c.actif), None)
            if not contrat: return 0.0, 0.0
            last_brut = float(contrat.salaire_base or 0)
            somme_12  = last_brut * 12
        else:
            # Exclure prime_transport de chaque bulletin
            somme_12  = sum(
                float(b.salaire_brut or 0) - float(b.prime_transport or 0)
                for b in buls
            )
            last_brut = float(buls[0].salaire_brut or 0) - float(buls[0].prime_transport or 0)

        # Prendre le plus favorable
        base_methode1 = somme_12  / 288        # Σ 12 mois / 288
        base_methode2 = (last_brut * 12) / 288 # dernier × 12 / 288
        base          = max(base_methode1, base_methode2)
        allocation    = round(base * jours_acquis, 0)
        return round(base, 2), allocation

    salaries_list = Salarie.query.filter_by(tenant_id=t.id, statut="ACTIF")        .options(joinedload(Salarie.categorie)).order_by(Salarie.nom).all()

    soldes = []
    for s in salaries_list:
        if q and q.lower() not in f"{s.nom} {s.prenom} {s.matricule}".lower():
            continue
        solde_db = Conge.query.filter_by(
            tenant_id=t.id, salarie_id=s.id, annee=annee
        ).filter(Conge.date_depart == None).first()

        jours_auto = calculer_solde_auto(s, annee)

        if solde_db:
            acquis    = float(solde_db.jours_acquis or jours_auto)
            pris      = float(solde_db.jours_pris   or 0)
        else:
            acquis    = jours_auto
            pris      = sum(
                float(c.jours_pris or 0)
                for c in Conge.query.filter_by(
                    tenant_id=t.id, salarie_id=s.id, annee=annee, statut="APPROUVÉ"
                ).all()
            )

        taux_j    = taux_conge(s, annee)
        base_all, allocation = calculer_allocation_conge(s, acquis, annee)
        soldes.append({
            "salarie":        s,
            "solde_db":       solde_db,
            "jours_auto":     jours_auto,
            "jours_acquis":   acquis,
            "jours_pris":     pris,
            "jours_restants": round(acquis - pris, 1),
            "alerte":         (acquis - pris) < 5,
            "taux_j":         taux_j,
            "base_allocation":base_all,
            "allocation":     allocation,
            "mineur":         taux_j == 2.5,
        })

    # Demandes (avec date_depart renseignée)
    demandes = Conge.query.filter_by(tenant_id=t.id)        .filter(Conge.date_depart.isnot(None))        .options(joinedload(Conge.salarie))        .order_by(Conge.date_depart.desc()).all()

    annees_dispo = sorted(set(
        [now.year, now.year-1, now.year+1]
        + [c.annee for c in Conge.query.filter_by(tenant_id=t.id).all()]
    ), reverse=True)

    return render_template("tenant/conges.html",
        tenant=t, soldes=soldes, demandes=demandes,
        annee=annee, annees_dispo=annees_dispo, now=now, q=q,
        salaries=salaries_list)

@app.route("/conges/nouveau", methods=["GET","POST"])
@login_required
def conge_nouveau():
    if current_user.is_super_admin: return redirect(url_for("admin_dashboard"))
    t = get_tenant()
    if not t: return redirect(url_for("login"))
    salaries_list = Salarie.query.filter_by(tenant_id=t.id, statut="ACTIF").order_by(Salarie.nom).all()
    if request.method == "POST":
        salarie_id = request.form.get("salarie_id", type=int)
        annee = request.form.get("annee", datetime.now().year, type=int)
        date_dep = _parse_date(request.form.get("date_depart"))
        date_ret = _parse_date(request.form.get("date_retour"))
        type_c = request.form.get("type_conge", "ANNUEL")
        jours = (date_ret - date_dep).days + 1 if date_dep and date_ret else 0
        conge = Conge.query.filter_by(tenant_id=t.id, salarie_id=salarie_id, annee=annee).first()
        if not conge:
            s = Salarie.query.get(salarie_id)
            mois = max(1,(datetime.now().date()-s.date_embauche).days//30) if s.date_embauche else 12
            conge = Conge(tenant_id=t.id, salarie_id=salarie_id, annee=annee,
                jours_acquis=round(min(mois,12)*2.0,1), jours_pris=0, type_conge=type_c, statut="DEMANDÉ")
            db.session.add(conge)
        conge.date_depart  = date_dep
        conge.date_retour  = date_ret
        conge.type_conge   = type_c
        conge.statut       = "DEMANDÉ"
        conge.jours_pris   = float(conge.jours_pris or 0)  # ne pas écraser les jours déjà pris
        db.session.commit()
        flash(f"✅ Demande de congé enregistrée ({jours} jour(s)).", "success")
        return redirect(url_for("conges"))
    return render_template("tenant/conge_form.html", tenant=t, salaries=salaries_list, now=datetime.now())

@app.route("/conges/<int:id>/modifier", methods=["GET","POST"])
@login_required
def conge_modifier(id):
    """Modifier une demande de congé existante."""
    t = get_tenant()
    if not t: return redirect(url_for("login"))
    c = Conge.query.filter_by(id=id, tenant_id=t.id).first_or_404()
    salaries_list = Salarie.query.filter_by(tenant_id=t.id, statut="ACTIF").order_by(Salarie.nom).all()

    if request.method == "POST":
        old_jours = (c.date_retour - c.date_depart).days + 1 if c.date_depart and c.date_retour else 0
        date_dep  = _parse_date(request.form.get("date_depart"))
        date_ret  = _parse_date(request.form.get("date_retour"))
        type_c    = request.form.get("type_conge", "ANNUEL")
        new_jours = (date_ret - date_dep).days + 1 if date_dep and date_ret else 0

        # Si congé APPROUVÉ → ajuster les jours_pris
        if c.statut == "APPROUVÉ":
            c.jours_pris = max(0, float(c.jours_pris or 0) - old_jours + new_jours)

        c.date_depart = date_dep
        c.date_retour = date_ret
        c.type_conge  = type_c
        c.statut      = request.form.get("statut", c.statut)
        db.session.commit()
        flash(f"✅ Congé modifié ({new_jours} jour(s)).", "success")
        return redirect(url_for("conges"))

    return render_template("tenant/conge_form.html",
        tenant=t, salaries=salaries_list,
        conge=c, now=datetime.now(), mode="modifier")


@app.route("/conges/<int:id>/approuver", methods=["POST"])
@login_required
def conge_approuver(id):
    t = get_tenant()
    if not t: return redirect(url_for("login"))
    c = Conge.query.filter_by(id=id, tenant_id=t.id).first_or_404()
    if c.date_depart and c.date_retour:
        jours = (c.date_retour - c.date_depart).days + 1
        # Mettre à jour le solde de l'année
        solde = Conge.query.filter_by(
            tenant_id=t.id, salarie_id=c.salarie_id, annee=c.annee
        ).filter(Conge.date_depart == None).first()
        if not solde:
            s = Salarie.query.get(c.salarie_id)
            mois = max(1,(datetime.now().date()-s.date_embauche).days//30) if s.date_embauche else 12
            solde = Conge(tenant_id=t.id, salarie_id=c.salarie_id, annee=c.annee,
                          jours_acquis=round(min(mois,12)*2.5, 1), jours_pris=0)
            db.session.add(solde)
        solde.jours_pris = float(solde.jours_pris or 0) + jours
        c.jours_pris     = float(c.jours_pris or 0) + jours
    c.statut = "APPROUVÉ"
    db.session.commit()
    flash(f"✅ Congé de {c.salarie.nom_complet} approuvé.", "success")
    return redirect(url_for("conges"))

@app.route("/conges/<int:id>/refuser", methods=["POST"])
@login_required
def conge_refuser(id):
    t = get_tenant()
    if not t: return redirect(url_for("login"))
    c = Conge.query.filter_by(id=id, tenant_id=t.id).first_or_404()
    c.statut="REFUSÉ"; db.session.commit()
    flash("Congé refusé.", "success")
    return redirect(url_for("conges"))

@app.route("/conges/<int:id>/supprimer", methods=["POST"])
@login_required
def conge_supprimer(id):
    t = get_tenant()
    if not t: return redirect(url_for("login"))
    c = Conge.query.filter_by(id=id, tenant_id=t.id).first_or_404()
    db.session.delete(c); db.session.commit()
    flash("Demande supprimée.", "success")
    return redirect(url_for("conges"))


@app.route("/salaries/imprimer")
@login_required
def salaries_imprimer():
    if current_user.is_super_admin: return redirect(url_for("admin_dashboard"))
    t = get_tenant(); 
    if not t: return redirect(url_for("login"))
    salaries_list = Salarie.query.filter_by(tenant_id=t.id).order_by(Salarie.nom).all()
    for s in salaries_list:
        s._contrat_actif = Contrat.query.filter_by(salarie_id=s.id, tenant_id=t.id, actif=True).first()
    return render_template("tenant/salaries_print.html", salaries=salaries_list, tenant=t, now=datetime.now())



@app.route("/api/travailleur/stats-sans-site")
@tenant_required
def api_stats_sans_site():
    t = get_tenant()
    # Salariés actifs sans affectation active
    sal_ids_affectes = {a.salarie_id for a in 
        AffectationSite.query.filter_by(tenant_id=t.id, actif=True).all() if a.salarie_id}
    jour_ids_affectes = {a.journalier_id for a in 
        AffectationSite.query.filter_by(tenant_id=t.id, actif=True).all() if a.journalier_id}
    nb_sal  = Salarie.query.filter_by(tenant_id=t.id, statut="ACTIF").count()
    nb_jour = Journalier.query.filter_by(tenant_id=t.id, statut="ACTIF").count()
    sans_site = (nb_sal - len(sal_ids_affectes)) + (nb_jour - len(jour_ids_affectes))
    return jsonify({"total": max(0, sans_site)})

# ══════════════════════════════════════════════════════════════════════════════
# ── SITES & AFFECTATIONS ──────────────────────────────────────────────────────
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/sites")
@tenant_required
def sites():
    t = get_tenant()
    sites_list = Site.query.filter_by(tenant_id=t.id).order_by(Site.nom).all()
    return render_template("tenant/sites.html", tenant=t, sites=sites_list)

@app.route("/sites/nouveau", methods=["GET","POST"])
@tenant_required
def site_nouveau():
    t = get_tenant()
    if request.method == "POST":
        s = Site(
            tenant_id   = t.id,
            nom         = request.form["nom"].strip(),
            code        = request.form.get("code","").strip().upper() or None,
            adresse     = request.form.get("adresse","").strip() or None,
            ville       = request.form.get("ville","").strip() or None,
            responsable = request.form.get("responsable","").strip() or None,
            telephone   = request.form.get("telephone","").strip() or None,
            description = request.form.get("description","").strip() or None,
        )
        db.session.add(s)
        db.session.commit()
        flash(f"Site « {s.nom} » créé avec succès.", "success")
        return redirect(url_for("sites"))
    return render_template("tenant/site_form.html", tenant=t, site=None)

@app.route("/sites/<int:id>/modifier", methods=["GET","POST"])
@tenant_required
def site_modifier(id):
    t = get_tenant()
    s = Site.query.filter_by(id=id, tenant_id=t.id).first_or_404()
    if request.method == "POST":
        s.nom         = request.form["nom"].strip()
        s.code        = request.form.get("code","").strip().upper() or None
        s.adresse     = request.form.get("adresse","").strip() or None
        s.ville       = request.form.get("ville","").strip() or None
        s.responsable = request.form.get("responsable","").strip() or None
        s.telephone   = request.form.get("telephone","").strip() or None
        s.description = request.form.get("description","").strip() or None
        db.session.commit()
        flash(f"Site « {s.nom} » modifié.", "success")
        return redirect(url_for("sites"))
    return render_template("tenant/site_form.html", tenant=t, site=s)

@app.route("/sites/<int:id>/toggle", methods=["POST"])
@tenant_required
def site_toggle(id):
    t = get_tenant()
    s = Site.query.filter_by(id=id, tenant_id=t.id).first_or_404()
    s.actif = not s.actif
    db.session.commit()
    flash(f"Site « {s.nom} » {'activé' if s.actif else 'désactivé'}.", "success")
    return redirect(url_for("sites"))

@app.route("/sites/<int:id>")
@tenant_required
def site_detail(id):
    t = get_tenant()
    s = Site.query.filter_by(id=id, tenant_id=t.id).first_or_404()

    # Date sélectionnée pour le pointage rapide
    date_str = request.args.get("date_ptg", date.today().strftime("%Y-%m-%d"))
    try:    date_ptg = datetime.strptime(date_str, "%Y-%m-%d").date()
    except: date_ptg = date.today()

    # Affectations actives
    affectations = AffectationSite.query.filter_by(site_id=id, actif=True)        .order_by(AffectationSite.date_debut.desc()).all()

    # Séparer salariés et journaliers affectés
    ids_sal  = [a.salarie_id    for a in affectations if a.salarie_id]
    ids_jour = [a.journalier_id for a in affectations if a.journalier_id]

    salaries_site    = Salarie.query.filter(
        Salarie.tenant_id==t.id, Salarie.statut=="ACTIF",
        Salarie.id.in_(ids_sal)
    ).order_by(Salarie.nom).all() if ids_sal else []

    journaliers_site = Journalier.query.filter(
        Journalier.tenant_id==t.id, Journalier.statut=="ACTIF",
        Journalier.id.in_(ids_jour)
    ).order_by(Journalier.nom).all() if ids_jour else []

    # Pointages du jour pour ce site
    pts_sal  = {p.salarie_id:    p for p in
        Pointage.query.filter_by(tenant_id=t.id, date_pointage=date_ptg)
        .filter(Pointage.salarie_id.in_(ids_sal)).all()} if ids_sal else {}
    pts_jour = {p.journalier_id: p for p in
        Pointage.query.filter_by(tenant_id=t.id, date_pointage=date_ptg)
        .filter(Pointage.journalier_id.in_(ids_jour)).all()} if ids_jour else {}

    # Stats pointage du jour
    nb_presents  = sum(1 for p in list(pts_sal.values())+list(pts_jour.values()) if p.present)
    nb_absents   = sum(1 for p in list(pts_sal.values())+list(pts_jour.values()) if p.absent)
    nb_non_pointes = (len(salaries_site)+len(journaliers_site)) - len(pts_sal) - len(pts_jour)

    # Historique complet
    historique = AffectationSite.query.filter_by(site_id=id)        .order_by(AffectationSite.date_creation.desc()).limit(50).all()

    # Travailleurs disponibles (non affectés à ce site)
    ids_sal_aff  = {a.salarie_id    for a in affectations if a.salarie_id}
    ids_jour_aff = {a.journalier_id for a in affectations if a.journalier_id}
    salaries_dispo    = [x for x in Salarie.query.filter_by(
        tenant_id=t.id, statut="ACTIF").order_by(Salarie.nom).all()
        if x.id not in ids_sal_aff]
    journaliers_dispo = [x for x in Journalier.query.filter_by(
        tenant_id=t.id, statut="ACTIF").order_by(Journalier.nom).all()
        if x.id not in ids_jour_aff]

    # ── KPIs tableau de bord ─────────────────────────────────────────────────
    from datetime import date as _date
    import calendar

    now_d       = _date.today()
    mois_debut  = _date(now_d.year, now_d.month, 1)
    mois_fin    = _date(now_d.year, now_d.month,
                        calendar.monthrange(now_d.year, now_d.month)[1])
    lundi_sem   = now_d - timedelta(days=now_d.weekday())
    samedi_sem  = lundi_sem + timedelta(days=5)

    # Tous les pointages du mois pour ce site (salariés + journaliers)
    ids_all = ids_sal + ids_jour
    pts_mois_sal = Pointage.query.filter_by(tenant_id=t.id)        .filter(Pointage.salarie_id.in_(ids_sal),
                Pointage.date_pointage >= mois_debut,
                Pointage.date_pointage <= mois_fin).all() if ids_sal else []
    pts_mois_jour = Pointage.query.filter_by(tenant_id=t.id)        .filter(Pointage.journalier_id.in_(ids_jour),
                Pointage.date_pointage >= mois_debut,
                Pointage.date_pointage <= mois_fin).all() if ids_jour else []
    pts_mois_tous = pts_mois_sal + pts_mois_jour

    # Pointages de la semaine
    pts_sem_sal  = [p for p in pts_mois_sal  if lundi_sem <= p.date_pointage <= samedi_sem]
    pts_sem_jour = [p for p in pts_mois_jour if lundi_sem <= p.date_pointage <= samedi_sem]

    # KPI — Jours pointés (présences) ce mois
    nb_jours_pointes_mois = sum(1 for p in pts_mois_tous if p.present)
    nb_absences_mois      = sum(1 for p in pts_mois_tous if p.absent)

    # KPI — Taux de présence semaine
    nb_pres_sem = sum(1 for p in pts_sem_sal + pts_sem_jour if p.present)
    nb_abs_sem  = sum(1 for p in pts_sem_sal + pts_sem_jour if p.absent)
    total_ptg_sem = nb_pres_sem + nb_abs_sem
    taux_presence_semaine = round(nb_pres_sem / total_ptg_sem * 100) if total_ptg_sem > 0 else 0

    # KPI — Heures totales semaine
    def total_heures_pt(p):
        return (float(p.heures_normales or 8) +
                float(p.heures_sup_10 or 0) + float(p.heures_sup_30 or 0) +
                float(p.heures_sup_40 or 0) + float(p.heures_sup_70 or 0) +
                float(p.heures_sup or 0))

    heures_semaine = sum(total_heures_pt(p) for p in pts_sem_sal + pts_sem_jour if p.present)

    # KPI — Masse journalière (feuilles de paie journaliers ce mois)
    feuilles_mois = FeuillePaieJournalier.query.filter_by(tenant_id=t.id)        .filter(FeuillePaieJournalier.journalier_id.in_(ids_jour),
                FeuillePaieJournalier.date_debut >= mois_debut,
                FeuillePaieJournalier.date_fin   <= mois_fin).all() if ids_jour else []
    masse_journaliere_mois    = sum(float(f.montant_brut or 0) for f in feuilles_mois)
    feuilles_attente = sum(1 for f in feuilles_mois if f.statut == "EN_ATTENTE")
    feuilles_payees  = sum(1 for f in feuilles_mois if f.statut == "PAYÉ")

    # KPI — Bulletins salariés du mois (dernière période active)
    periode_courante = PeriodePaie.query.filter_by(
        tenant_id=t.id, annee=now_d.year, mois=now_d.month).first()
    bulletins_site = []
    masse_mensuelle = 0
    if periode_courante and ids_sal:
        bulletins_site = BulletinPaie.query.filter_by(
            tenant_id=t.id, periode_id=periode_courante.id
        ).filter(BulletinPaie.salarie_id.in_(ids_sal)).all()
        masse_mensuelle = sum(float(b.net_a_payer or 0) for b in bulletins_site)

    # Évolution présence 7 derniers jours (pour mini-graphique)
    evolution_7j = []
    for i in range(6, -1, -1):
        d = now_d - timedelta(days=i)
        p_d = [p for p in pts_mois_tous if p.date_pointage == d]
        nb_p = sum(1 for p in p_d if p.present)
        nb_a = sum(1 for p in p_d if p.absent)
        evolution_7j.append({
            "date":    d.strftime("%d/%m"),
            "jour":    ["L","Ma","Me","J","V","Sa","Di"][d.weekday()],
            "presents": nb_p,
            "absents":  nb_a,
            "heures":   round(sum(total_heures_pt(p) for p in p_d if p.present), 1),
        })

    return render_template("tenant/site_detail.html",
        tenant=t, site=s,
        affectations=affectations, historique=historique,
        salaries_dispo=salaries_dispo, journaliers_dispo=journaliers_dispo,
        salaries_site=salaries_site, journaliers_site=journaliers_site,
        pts_sal=pts_sal, pts_jour=pts_jour,
        date_ptg=date_ptg,
        date_hier=(date_ptg - timedelta(days=1)).strftime("%Y-%m-%d"),
        date_demain=(date_ptg + timedelta(days=1)).strftime("%Y-%m-%d"),
        nb_presents=nb_presents, nb_absents=nb_absents,
        nb_non_pointes=nb_non_pointes,
        # KPIs tableau de bord
        nb_sal_site=len(salaries_site), nb_jour_site=len(journaliers_site),
        nb_jours_pointes_mois=nb_jours_pointes_mois,
        nb_absences_mois=nb_absences_mois,
        taux_presence_semaine=taux_presence_semaine,
        heures_semaine=round(heures_semaine, 1),
        masse_journaliere_mois=masse_journaliere_mois,
        feuilles_attente=feuilles_attente, feuilles_payees=feuilles_payees,
        masse_mensuelle=masse_mensuelle,
        nb_bulletins_site=len(bulletins_site),
        evolution_7j=evolution_7j,
        mois_nom=["","Jan","Fév","Mar","Avr","Mai","Jun","Jul","Aoû","Sep","Oct","Nov","Déc"][now_d.month],
        today=str(date.today()))

@app.route("/sites/<int:id>/pointage-rapide", methods=["POST"])
@tenant_required
def site_pointage_rapide(id):
    """Sauvegarder le pointage rapide depuis la page d'un site."""
    t = get_tenant()
    s = Site.query.filter_by(id=id, tenant_id=t.id).first_or_404()
    date_str = request.form.get("date_pointage")
    try:    date_p = datetime.strptime(date_str, "%Y-%m-%d").date()
    except: date_p = date.today()

    nb = 0
    for key, val in request.form.items():
        # Salariés
        if key.startswith("sal_present_"):
            sid = int(key.replace("sal_present_", ""))
            present = (val == "1"); absent = not present
            pt = Pointage.query.filter_by(
                tenant_id=t.id, date_pointage=date_p, salarie_id=sid).first()
            if not pt:
                pt = Pointage(tenant_id=t.id, date_pointage=date_p,
                              salarie_id=sid, site_id=id)
                db.session.add(pt)
            pt.present = present; pt.absent = absent
            pt.heures_normales = float(request.form.get(f"sal_h_{sid}", 8) or 8)
            pt.heures_sup_10   = float(request.form.get(f"sal_s10_{sid}", 0) or 0)
            pt.heures_sup_30   = float(request.form.get(f"sal_s30_{sid}", 0) or 0)
            pt.heures_sup_40   = float(request.form.get(f"sal_s40_{sid}", 0) or 0)
            pt.heures_sup_70   = float(request.form.get(f"sal_s70_{sid}", 0) or 0)
            pt.motif_absence   = request.form.get(f"sal_motif_{sid}", "") if absent else None
            nb += 1
        # Journaliers
        elif key.startswith("jour_present_"):
            jid = int(key.replace("jour_present_", ""))
            present = (val == "1"); absent = not present
            pt = Pointage.query.filter_by(
                tenant_id=t.id, date_pointage=date_p, journalier_id=jid).first()
            if not pt:
                pt = Pointage(tenant_id=t.id, date_pointage=date_p,
                              journalier_id=jid, site_id=id)
                db.session.add(pt)
            pt.present = present; pt.absent = absent
            pt.heures_normales = float(request.form.get(f"jour_h_{jid}", 8) or 8)
            pt.heures_sup      = float(request.form.get(f"jour_s_{jid}", 0) or 0)
            pt.motif_absence   = request.form.get(f"jour_motif_{jid}", "") if absent else None
            nb += 1

    db.session.commit()
    flash(f"✅ Pointage du {date_p.strftime('%d/%m/%Y')} sauvegardé — {nb} travailleur(s).", "success")
    return redirect(url_for("site_detail", id=id) + f"?date_ptg={date_str}")

@app.route("/sites/<int:site_id>/affecter", methods=["POST"])
@tenant_required
def site_affecter(site_id):
    """Affecter un ou plusieurs travailleurs à un site."""
    t  = get_tenant()
    s  = Site.query.filter_by(id=site_id, tenant_id=t.id).first_or_404()
    date_debut = request.form.get("date_debut") or str(date.today())
    motif      = request.form.get("motif","").strip() or None
    nb         = 0

    for key in request.form:
        if key.startswith("sal_"):
            sal_id = int(key[4:])
            sal = Salarie.query.filter_by(id=sal_id, tenant_id=t.id).first()
            if not sal: continue
            # Désactiver toute affectation active précédente sur un AUTRE site
            prev = AffectationSite.query.filter_by(
                salarie_id=sal_id, actif=True, tenant_id=t.id).first()
            if prev and prev.site_id != site_id:
                prev.actif    = False
                prev.date_fin = date.today()
                prev.motif    = f"Transféré vers {s.nom}"
            elif prev and prev.site_id == site_id:
                continue  # Déjà sur ce site
            a = AffectationSite(
                tenant_id=t.id, site_id=site_id, salarie_id=sal_id,
                date_debut=date_debut, actif=True, motif=motif,
                cree_par=current_user.email)
            db.session.add(a); nb += 1

        elif key.startswith("jour_"):
            jour_id = int(key[5:])
            jour = Journalier.query.filter_by(id=jour_id, tenant_id=t.id).first()
            if not jour: continue
            prev = AffectationSite.query.filter_by(
                journalier_id=jour_id, actif=True, tenant_id=t.id).first()
            if prev and prev.site_id != site_id:
                prev.actif    = False
                prev.date_fin = date.today()
                prev.motif    = f"Transféré vers {s.nom}"
            elif prev and prev.site_id == site_id:
                continue
            a = AffectationSite(
                tenant_id=t.id, site_id=site_id, journalier_id=jour_id,
                date_debut=date_debut, actif=True, motif=motif,
                cree_par=current_user.email)
            db.session.add(a); nb += 1

    db.session.commit()
    flash(f"{nb} travailleur(s) affecté(s) à « {s.nom} ».", "success")
    return redirect(url_for("site_detail", id=site_id))

@app.route("/sites/affecter-travailleur/<int:affectation_id>/retirer", methods=["POST"])
@tenant_required
def site_retirer(affectation_id):
    """Retirer un travailleur de son site (fin d'affectation)."""
    t = get_tenant()
    a = AffectationSite.query.filter_by(id=affectation_id, tenant_id=t.id).first_or_404()
    motif = request.form.get("motif","").strip() or "Retrait manuel"
    a.actif    = False
    a.date_fin = date.today()
    a.motif    = motif
    db.session.commit()
    flash(f"Affectation terminée pour {a.nom_travailleur}.", "success")
    return redirect(url_for("site_detail", id=a.site_id))

@app.route("/sites/permuter", methods=["POST"])
@tenant_required
def site_permuter():
    """Permuter un travailleur d'un site vers un autre."""
    t         = get_tenant()
    aff_id    = request.form.get("affectation_id", type=int)
    nouveau_site_id = request.form.get("site_destination_id", type=int)
    motif     = request.form.get("motif","Permutation").strip()

    aff_old = AffectationSite.query.filter_by(id=aff_id, tenant_id=t.id, actif=True).first_or_404()
    site_dest = Site.query.filter_by(id=nouveau_site_id, tenant_id=t.id).first_or_404()

    # Fermer l'affectation actuelle
    aff_old.actif    = False
    aff_old.date_fin = date.today()
    aff_old.motif    = f"Permuté vers {site_dest.nom} — {motif}"

    # Créer la nouvelle affectation
    aff_new = AffectationSite(
        tenant_id     = t.id,
        site_id       = nouveau_site_id,
        salarie_id    = aff_old.salarie_id,
        journalier_id = aff_old.journalier_id,
        date_debut    = date.today(),
        actif         = True,
        motif         = f"Permuté depuis {aff_old.site.nom} — {motif}",
        cree_par      = current_user.email,
    )
    db.session.add(aff_new)
    db.session.commit()
    flash(f"{aff_old.nom_travailleur} permuté vers « {site_dest.nom} ».", "success")
    return redirect(url_for("site_detail", id=nouveau_site_id))

@app.route("/api/travailleur/<string:type>/<int:id>/site")
@tenant_required
def api_travailleur_site(type, id):
    """API : retourne le site actuel d'un travailleur."""
    t = get_tenant()
    if type == "salarie":
        a = AffectationSite.query.filter_by(
            salarie_id=id, tenant_id=t.id, actif=True).first()
    else:
        a = AffectationSite.query.filter_by(
            journalier_id=id, tenant_id=t.id, actif=True).first()
    if a:
        return jsonify({"site_id": a.site_id, "site_nom": a.site.nom,
                        "date_debut": str(a.date_debut)})
    return jsonify({"site_id": None, "site_nom": None})


@app.route("/journaliers/imprimer")
@login_required
def journaliers_imprimer():
    if current_user.is_super_admin: return redirect(url_for("admin_dashboard"))
    t = get_tenant()
    if not t: return redirect(url_for("login"))
    return render_template("tenant/journaliers_print.html", journaliers=Journalier.query.filter_by(tenant_id=t.id).order_by(Journalier.nom).all(), tenant=t, now=datetime.now())

# ── Export & API ──────────────────────────────────────────────────────────────
@app.route("/bulletins/export/<int:periode_id>")
@tenant_required
def export_journal(periode_id):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    t=get_tenant()
    p=PeriodePaie.query.filter_by(id=periode_id,tenant_id=t.id).first_or_404()
    buls=BulletinPaie.query.filter_by(periode_id=periode_id,tenant_id=t.id).join(Salarie).order_by(Salarie.nom).all()
    wb=Workbook(); ws=wb.active; ws.title=f"Journal {p.libelle_complet}"
    ws.merge_cells("A1:R1"); ws["A1"]=f"JOURNAL DE PAIE — {p.libelle_complet} — {t.denomination}"
    ws["A1"].font=Font(bold=True,size=13); ws["A1"].alignment=Alignment(horizontal="center")
    hdrs=["Matricule","Nom","Prénom","Emploi","Cat.","Base","Brut","CNSS Sal.","CNAMGS Sal.","TCS","IRPP","Net","Net à Payer","CNSS Pat.","CNAMGS Pat.","FNH","CFP","Statut"]
    for col,h in enumerate(hdrs,1):
        c=ws.cell(row=3,column=col,value=h); c.font=Font(bold=True,color="FFFFFF")
        c.fill=PatternFill("solid",fgColor="1a2332"); c.alignment=Alignment(horizontal="center")
    for row,b in enumerate(buls,4):
        s=b.salarie
        vals=[s.matricule,s.nom,s.prenom,s.emploi,s.categorie.code if s.categorie else "",
              float(b.salaire_base or 0),float(b.salaire_brut or 0),float(b.cnss_salarie or 0),
              float(b.cnamgs_salarie or 0),float(b.tcs or 0),float(b.irpp or 0),
              float(b.salaire_net or 0),float(b.net_a_payer or 0),float(b.cnss_patronale or 0),
              float(b.cnamgs_patronale or 0),float(b.fnh or 0),float(b.cfp or 0),b.statut]
        for col,v in enumerate(vals,1):
            cell=ws.cell(row=row,column=col,value=v)
            if isinstance(v,float): cell.number_format='#,##0'
            if row%2==0: cell.fill=PatternFill("solid",fgColor="F5F5F5")
    out=io.BytesIO(); wb.save(out); out.seek(0)
    return send_file(out,mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,download_name=f"Journal_{p.libelle_mois}_{p.annee}_{t.slug}.xlsx")

@app.route("/api/calculer-bulletin", methods=["POST"])
@login_required
def api_calculer():
    try:
        t = get_tenant()
        data = request.get_json() or {}
        sid = data.pop("salarie_id", None)
        nb_parts = 1.0
        if sid and t:
            s = Salarie.query.filter_by(id=sid, tenant_id=t.id).first()
            if s: nb_parts = float(s.nombre_parts or 1)
        mois  = data.pop("mois_periode", None)
        annee = data.pop("annee_periode", None)
        total_acomptes = 0.0
        if sid and t and mois and annee:
            total_acomptes = float(db.session.query(db.func.sum(Acompte.montant))
                .filter_by(tenant_id=t.id, salarie_id=int(sid), mois=int(mois),
                           annee=int(annee), statut="EN_ATTENTE").scalar() or 0)
        if total_acomptes > 0:
            data["acompte"] = max(float(data.get("acompte", 0)), total_acomptes)
        res = calculer_bulletin(data, nb_parts=nb_parts)
        res["acompte_auto"] = total_acomptes
        return jsonify(res)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/salarie/<int:id>/contrat")
@login_required
def api_contrat(id):
    t=get_tenant()
    s=Salarie.query.filter_by(id=id,tenant_id=t.id).first()
    if not s: return jsonify({})
    c=Contrat.query.filter_by(salarie_id=id,tenant_id=t.id,actif=True).first()
    base={"nom":s.nom_complet,"poste":s.emploi,"matricule":s.matricule,"nombre_parts":float(s.nombre_parts or 1)}
    if c: base["salaire_base"]=float(c.salaire_base); base["poste"]=c.poste or s.emploi
    return jsonify(base)

@app.route("/api/salarie/<int:id>/pointage-mois")
@login_required
def api_pointage_mois(id):
    """Retourne le cumul des heures du pointage pour un salarié sur un mois donné."""
    t = get_tenant()
    if not t: return jsonify({})
    mois  = request.args.get("mois",  type=int)
    annee = request.args.get("annee", type=int)
    if not mois or not annee:
        return jsonify({"erreur": "mois et annee requis"})
    import calendar
    dernier_jour = calendar.monthrange(annee, mois)[1]
    debut = date(annee, mois, 1)
    fin   = date(annee, mois, dernier_jour)
    pts = Pointage.query.filter_by(tenant_id=t.id, salarie_id=id)        .filter(Pointage.date_pointage >= debut, Pointage.date_pointage <= fin,
                Pointage.present == True).all()
    if not pts:
        return jsonify({"nb_jours": 0, "nb_absences": 0,
            "heures_sup_10": 0, "heures_sup_30": 0, "heures_sup_40": 0, "heures_sup_70": 0,
            "heures_normales_total": 0, "total_sup": 0,
            "message": "Aucun pointage pour cette période"})
    nb_jours        = len(pts)
    heures_normales = sum(float(p.heures_normales or 8) for p in pts)
    heures_sup_10   = sum(float(p.heures_sup_10 or 0) for p in pts)
    heures_sup_30   = sum(float(p.heures_sup_30 or 0) for p in pts)
    heures_sup_40   = sum(float(p.heures_sup_40 or 0) for p in pts)
    heures_sup_70   = sum(float(p.heures_sup_70 or 0) for p in pts)
    pts_absents = Pointage.query.filter_by(tenant_id=t.id, salarie_id=id)        .filter(Pointage.date_pointage >= debut, Pointage.date_pointage <= fin,
                Pointage.absent == True).all()
    return jsonify({
        "nb_jours":              nb_jours,
        "nb_absences":           len(pts_absents),
        "heures_normales_total": round(heures_normales, 2),
        "heures_sup_10":         round(heures_sup_10, 2),
        "heures_sup_30":         round(heures_sup_30, 2),
        "heures_sup_40":         round(heures_sup_40, 2),
        "heures_sup_70":         round(heures_sup_70, 2),
        "total_sup":             round(heures_sup_10+heures_sup_30+heures_sup_40+heures_sup_70, 2),
        "message":               f"{nb_jours} jour(s) pointé(s) sur {dernier_jour}"
    })

@app.route("/api/cache/clear", methods=["POST"])
@login_required
def api_cache_clear():
    """Vider le cache du dashboard (bouton rafraîchir)."""
    t = get_tenant()
    if not t: return jsonify({"ok": False})
    _cache_delete(f"{t.id}:")
    return jsonify({"ok": True, "msg": "Cache vidé"})


@app.route("/api/semaine-btp")
@login_required
def api_semaine_btp():
    """
    Calcule la distribution BTP des heures pour un travailleur sur une semaine.
    Params : type (sal|jour), id, date (n'importe quel jour de la semaine)
    """
    t = get_tenant()
    if not t: return jsonify({"erreur": "non connecté"})

    type_w    = request.args.get("type", "sal")
    worker_id = request.args.get("id", type=int)
    date_str  = request.args.get("date", datetime.now().strftime("%Y-%m-%d"))

    try:
        date_ref = datetime.strptime(date_str, "%Y-%m-%d").date()
    except:
        date_ref = datetime.now().date()

    lundi  = date_ref - timedelta(days=date_ref.weekday())
    samedi = lundi + timedelta(days=5)

    if type_w == "sal":
        pts = Pointage.query.filter_by(tenant_id=t.id, salarie_id=worker_id)            .filter(Pointage.date_pointage >= lundi,
                    Pointage.date_pointage <= samedi,
                    Pointage.present == True).all()
    else:
        pts = Pointage.query.filter_by(tenant_id=t.id, journalier_id=worker_id)            .filter(Pointage.date_pointage >= lundi,
                    Pointage.date_pointage <= samedi,
                    Pointage.present == True).all()

    if not pts:
        return jsonify({
            "semaine": f"{lundi.strftime('%d/%m')} → {samedi.strftime('%d/%m/%Y')}",
            "nb_jours": 0, "heures_normales": 0,
            "heures_sup_10": 0, "heures_sup_30": 0,
            "heures_sup_40": 0, "heures_sup_70": 0,
            "message": "Aucun pointage cette semaine"
        })

    from calculs_paie import distribuer_heures_semaine_btp
    jours_data = []
    for p in pts:
        h_norm = float(p.heures_normales or 0)
        if type_w == "sal":
            h_nuit = float(p.heures_sup_40 or 0)
        else:
            h_nuit = 0
        jours_data.append({
            "heures_normales": h_norm,
            "heures_sup_nuit": h_nuit,
            "type_jour": p.type_jour or "NORMAL"
        })

    dist = distribuer_heures_semaine_btp(jours_data)
    dist["semaine"]  = f"{lundi.strftime('%d/%m')} → {samedi.strftime('%d/%m/%Y')}"
    dist["nb_jours"] = len(pts)
    dist["jours_detail"] = [
        {
            "date":     str(p.date_pointage),
            "jour_fr":  ["Lun","Mar","Mer","Jeu","Ven","Sam","Dim"][p.date_pointage.weekday()],
            "heures":   float(p.heures_normales or 0),
            "type_jour": p.type_jour or "NORMAL",
        } for p in sorted(pts, key=lambda x: x.date_pointage)
    ]

    # Montants si salaire connu
    salaire_base = request.args.get("salaire", type=float)
    if salaire_base:
        from calculs_paie import calculer_taux_horaire, COEFF_SUP_10, COEFF_SUP_30
        th = calculer_taux_horaire(salaire_base)
        dist["taux_horaire"]    = round(th, 2)
        dist["montant_10"]      = round(dist["heures_sup_10"] * th * COEFF_SUP_10, 2)
        dist["montant_30"]      = round(dist["heures_sup_30"] * th * COEFF_SUP_30, 2)
        dist["montant_total_sup"] = round(dist["montant_10"] + dist["montant_30"]
                                         + dist["heures_sup_40"] * th * 1.40
                                         + dist["heures_sup_70"] * th * 1.70, 2)

    return jsonify(dist)


@app.route("/api/salarie/<int:id>/acomptes-mois")
@login_required
def api_acomptes_mois(id):
    t = get_tenant()
    mois=request.args.get("mois",type=int); annee=request.args.get("annee",type=int)
    if not t or not mois or not annee: return jsonify({"total":0})
    total = db.session.query(db.func.sum(Acompte.montant))\
            .filter_by(tenant_id=t.id,salarie_id=id,mois=mois,annee=annee,statut="EN_ATTENTE").scalar() or 0
    return jsonify({"total":float(total)})

@app.route("/pointage/recap-semaine")
@login_required
def pointage_recap_semaine():
    """Récapitulatif de présence hebdomadaire par site."""
    if current_user.is_super_admin: return redirect(url_for("admin_dashboard"))
    t = get_tenant()
    if not t: return redirect(url_for("login"))

    # Semaine sélectionnée
    date_str = request.args.get("date", datetime.now().strftime("%Y-%m-%d"))
    try:    date_ref = datetime.strptime(date_str, "%Y-%m-%d").date()
    except: date_ref = datetime.now().date()

    lundi  = date_ref - timedelta(days=date_ref.weekday())
    jours  = [lundi + timedelta(days=i) for i in range(6)]  # lundi→samedi
    samedi = jours[-1]

    # Sites actifs
    sites_list = Site.query.filter_by(tenant_id=t.id, actif=True).order_by(Site.nom).all()

    # Tous les pointages de la semaine
    pts_semaine = Pointage.query.filter_by(tenant_id=t.id)        .filter(Pointage.date_pointage >= lundi,
                Pointage.date_pointage <= samedi).all()

    # Affectations actives (site → workers)
    aff_sal  = {}  # salarie_id  → site_id
    aff_jour = {}  # journalier_id → site_id
    for a in AffectationSite.query.filter_by(tenant_id=t.id, actif=True).all():
        if a.salarie_id:    aff_sal[a.salarie_id]    = a.site_id
        if a.journalier_id: aff_jour[a.journalier_id] = a.site_id

    # ── Construire les données par site et par jour ───────────────────────────
    # Structure : { site_id: { date: { presents, absents, heures, h_sup, non_pointes } } }
    JOURS_FR = ["Lundi","Mardi","Mercredi","Jeudi","Vendredi","Samedi"]

    # Nb de travailleurs affectés à chaque site
    effectif_site = {}
    for s in sites_list:
        nb_sal  = sum(1 for v in aff_sal.values()  if v == s.id)
        nb_jour = sum(1 for v in aff_jour.values() if v == s.id)
        effectif_site[s.id] = nb_sal + nb_jour

    recap = {}
    for s in sites_list:
        recap[s.id] = {
            "site": s,
            "effectif": effectif_site.get(s.id, 0),
            "jours": {},
            "totaux": {"presents": 0, "absents": 0, "heures": 0.0, "h_sup": 0.0},
        }
        for j in jours:
            recap[s.id]["jours"][str(j)] = {
                "date":        j,
                "jour_fr":     JOURS_FR[j.weekday()],
                "presents":    0,
                "absents":     0,
                "non_pointes": effectif_site.get(s.id, 0),
                "heures":      0.0,
                "h_sup":       0.0,
            }

    # Sac sans site
    recap["sans_site"] = {
        "site": None,
        "effectif": 0,
        "jours": {},
        "totaux": {"presents": 0, "absents": 0, "heures": 0.0, "h_sup": 0.0},
    }
    for j in jours:
        recap["sans_site"]["jours"][str(j)] = {
            "date": j, "jour_fr": JOURS_FR[j.weekday()],
            "presents": 0, "absents": 0, "non_pointes": 0,
            "heures": 0.0, "h_sup": 0.0,
        }

    # Remplir avec les pointages réels
    for p in pts_semaine:
        d = str(p.date_pointage)
        site_id = aff_sal.get(p.salarie_id) or aff_jour.get(p.journalier_id)
        key = site_id if site_id and site_id in recap else "sans_site"

        if d not in recap[key]["jours"]:
            continue

        cell = recap[key]["jours"][d]
        if p.present:
            cell["presents"]    += 1
            h_norm = float(p.heures_normales or 8)
            h_sup  = (float(p.heures_sup_10 or 0) + float(p.heures_sup_30 or 0) +
                      float(p.heures_sup_40 or 0) + float(p.heures_sup_70 or 0) +
                      float(p.heures_sup or 0))
            cell["heures"] += h_norm
            cell["h_sup"]  += h_sup
        else:
            cell["absents"] += 1
        # Recalcul non_pointés
        cell["non_pointes"] = max(0,
            recap[key]["effectif"] - cell["presents"] - cell["absents"])

    # Calculer les totaux semaine par site
    for key in recap:
        tot = recap[key]["totaux"]
        for d, cell in recap[key]["jours"].items():
            tot["presents"] += cell["presents"]
            tot["absents"]  += cell["absents"]
            tot["heures"]   += cell["heures"]
            tot["h_sup"]    += cell["h_sup"]

    # Totaux globaux tous sites
    totaux_globaux = {"presents": 0, "absents": 0, "heures": 0.0, "h_sup": 0.0}
    for key in recap:
        t2 = recap[key]["totaux"]
        totaux_globaux["presents"] += t2["presents"]
        totaux_globaux["absents"]  += t2["absents"]
        totaux_globaux["heures"]   += t2["heures"]
        totaux_globaux["h_sup"]    += t2["h_sup"]

    # Filtrer "sans_site" si vide
    if recap["sans_site"]["totaux"]["presents"] == 0 and recap["sans_site"]["totaux"]["absents"] == 0:
        del recap["sans_site"]

    return render_template("tenant/pointage_recap_semaine.html",
        tenant=t,
        jours=jours,
        jours_fr=JOURS_FR,
        recap=recap,
        sites_list=sites_list,
        totaux_globaux=totaux_globaux,
        lundi=lundi,
        samedi=samedi,
        date_ref=date_ref,
        semaine_prec=(lundi - timedelta(days=7)).strftime("%Y-%m-%d"),
        semaine_suiv=(lundi + timedelta(days=7)).strftime("%Y-%m-%d"),
        now=datetime.now())

@app.route("/api/pointage/semaine")
@login_required
def api_pointage_semaine():
    t = get_tenant()
    if not t: return jsonify({})
    date_str = request.args.get("date")
    try: date_sel = datetime.strptime(date_str, "%Y-%m-%d").date()
    except: date_sel = datetime.now().date()
    lundi=date_sel-timedelta(days=date_sel.weekday()); samedi=lundi+timedelta(days=5)
    pts = Pointage.query.filter_by(tenant_id=t.id).filter(Pointage.date_pointage>=lundi,Pointage.date_pointage<=samedi).all()
    stats={}
    for p in pts:
        key=str(p.date_pointage)
        if key not in stats: stats[key]={"presents":0,"absents":0,"heures":0}
        if p.present: stats[key]["presents"]+=1; stats[key]["heures"]+=p.total_heures
        else: stats[key]["absents"]+=1
    return jsonify(stats)


# ══════════════════════════════════════════════════════════════════════════════
# RAPPORT MENSUEL PAR SITE
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/rapports/mensuel-site")
@login_required
def rapport_mensuel_site():
    """Page rapport mensuel par site : pointage + paie journalier + masse salariale."""
    if current_user.is_super_admin: return redirect(url_for("admin_dashboard"))
    t = get_tenant()
    if not t: return redirect(url_for("login"))

    MOIS_NOMS_LONG = ["","Janvier","Février","Mars","Avril","Mai","Juin",
                      "Juillet","Août","Septembre","Octobre","Novembre","Décembre"]

    # Paramètres
    now_d  = datetime.now()
    mois   = request.args.get("mois",  type=int, default=now_d.month)
    annee  = request.args.get("annee", type=int, default=now_d.year)
    site_id= request.args.get("site_id", type=int)

    import calendar
    _, nb_jours_mois = calendar.monthrange(annee, mois)
    mois_debut = date(annee, mois, 1)
    mois_fin   = date(annee, mois, nb_jours_mois)

    sites_list = Site.query.filter_by(tenant_id=t.id, actif=True).order_by(Site.nom).all()
    site_sel   = Site.query.get(site_id) if site_id else None

    # Affectations actives pour ce mois
    aff_sal  = {}   # salarie_id  → site_id
    aff_jour = {}   # journalier_id → site_id
    for a in AffectationSite.query.filter_by(tenant_id=t.id).all():
        if a.actif or (a.date_fin and a.date_fin >= mois_debut):
            if a.salarie_id:    aff_sal[a.salarie_id]    = a.site_id
            if a.journalier_id: aff_jour[a.journalier_id] = a.site_id

    def _build_rapport_site(s):
        """Construit le rapport complet pour un site donné."""
        sid = s.id
        # Travailleurs affectés à ce site
        ids_sal  = [k for k,v in aff_sal.items()  if v == sid]
        ids_jour = [k for k,v in aff_jour.items() if v == sid]

        salaries_aff    = Salarie.query.filter(
            Salarie.tenant_id==t.id, Salarie.id.in_(ids_sal)
        ).order_by(Salarie.nom).all() if ids_sal else []
        journaliers_aff = Journalier.query.filter(
            Journalier.tenant_id==t.id, Journalier.id.in_(ids_jour)
        ).order_by(Journalier.nom).all() if ids_jour else []

        # ── Pointage mensuel ─────────────────────────────────────────────────
        pts_sal  = Pointage.query.filter_by(tenant_id=t.id)            .filter(Pointage.salarie_id.in_(ids_sal),
                    Pointage.date_pointage >= mois_debut,
                    Pointage.date_pointage <= mois_fin).all() if ids_sal else []
        pts_jour = Pointage.query.filter_by(tenant_id=t.id)            .filter(Pointage.journalier_id.in_(ids_jour),
                    Pointage.date_pointage >= mois_debut,
                    Pointage.date_pointage <= mois_fin).all() if ids_jour else []
        pts_tous = pts_sal + pts_jour

        nb_presences  = sum(1 for p in pts_tous if p.present)
        nb_absences   = sum(1 for p in pts_tous if p.absent)
        taux_pres = round(nb_presences / (nb_presences + nb_absences) * 100
                         ) if (nb_presences + nb_absences) > 0 else 0

        # Heures totales
        heures_normales = sum(float(p.heures_normales or 8) for p in pts_tous if p.present)
        heures_sup = sum(
            float(p.heures_sup_10 or 0) + float(p.heures_sup_30 or 0) +
            float(p.heures_sup_40 or 0) + float(p.heures_sup_70 or 0) +
            float(p.heures_sup or 0)
            for p in pts_tous if p.present)

        # Détail par travailleur (pointage)
        detail_sal = []
        for sal in salaries_aff:
            pts_s = [p for p in pts_sal if p.salarie_id == sal.id]
            nb_p  = sum(1 for p in pts_s if p.present)
            nb_a  = sum(1 for p in pts_s if p.absent)
            h_n   = sum(float(p.heures_normales or 8) for p in pts_s if p.present)
            h_s   = sum(float(p.heures_sup_10 or 0)+float(p.heures_sup_30 or 0)+
                        float(p.heures_sup_40 or 0)+float(p.heures_sup_70 or 0)
                        for p in pts_s if p.present)
            detail_sal.append({
                "nom": sal.nom_complet, "matricule": sal.matricule,
                "emploi": sal.emploi or "—", "type": "MENSUEL",
                "nb_presences": nb_p, "nb_absences": nb_a,
                "heures_normales": round(h_n, 1), "heures_sup": round(h_s, 1),
                "taux": round(nb_p/(nb_p+nb_a)*100) if (nb_p+nb_a) > 0 else 0,
            })

        detail_jour = []
        for jour in journaliers_aff:
            pts_j = [p for p in pts_jour if p.journalier_id == jour.id]
            nb_p  = sum(1 for p in pts_j if p.present)
            nb_a  = sum(1 for p in pts_j if p.absent)
            h_n   = sum(float(p.heures_normales or 8) for p in pts_j if p.present)
            h_s   = sum(float(p.heures_sup or 0) for p in pts_j if p.present)
            detail_jour.append({
                "nom": jour.nom_complet, "profession": jour.profession or "—",
                "taux_horaire": float(jour.taux_horaire or 0), "type": "JOURNALIER",
                "nb_presences": nb_p, "nb_absences": nb_a,
                "heures_normales": round(h_n, 1), "heures_sup": round(h_s, 1),
                "taux": round(nb_p/(nb_p+nb_a)*100) if (nb_p+nb_a) > 0 else 0,
            })

        # ── Paie journaliers ─────────────────────────────────────────────────
        feuilles = FeuillePaieJournalier.query.filter_by(tenant_id=t.id)            .filter(FeuillePaieJournalier.journalier_id.in_(ids_jour),
                    FeuillePaieJournalier.date_debut >= mois_debut,
                    FeuillePaieJournalier.date_fin   <= mois_fin).all() if ids_jour else []
        masse_jour_brut   = sum(float(f.montant_brut or 0) for f in feuilles)
        feuilles_payees   = sum(1 for f in feuilles if f.statut == "PAYÉ")
        feuilles_attente  = sum(1 for f in feuilles if f.statut == "EN_ATTENTE")

        detail_feuilles = [{
            "nom":         f.journalier.nom_complet,
            "profession":  f.journalier.profession or "—",
            "date_debut":  f.date_debut.strftime("%d/%m/%Y") if f.date_debut else "",
            "date_fin":    f.date_fin.strftime("%d/%m/%Y")   if f.date_fin   else "",
            "nb_jours":    f.nb_jours,
            "heures":      round(float(f.total_heures or 0), 1),
            "taux":        round(float(f.taux_horaire or 0)),
            "montant":     round(float(f.montant_brut or 0)),
            "statut":      f.statut,
        } for f in feuilles]

        # ── Bulletins salariés ────────────────────────────────────────────────
        periode = PeriodePaie.query.filter_by(
            tenant_id=t.id, annee=annee, mois=mois).first()
        bulletins_site = []
        masse_mensuelle = {}
        detail_bulletins = []
        if periode and ids_sal:
            bulletins_site = BulletinPaie.query.filter_by(
                tenant_id=t.id, periode_id=periode.id
            ).filter(BulletinPaie.salarie_id.in_(ids_sal)).all()
            masse_mensuelle = calculer_masse_salariale(bulletins_site)
            detail_bulletins = [{
                "nom":        b.salarie.nom_complet,
                "matricule":  b.salarie.matricule,
                "emploi":     b.salarie.emploi or "—",
                "brut":       round(float(b.salaire_brut  or 0)),
                "net":        round(float(b.net_a_payer   or 0)),
                "cnss":       round(float(b.cnss_salarie  or 0)),
                "irpp":       round(float(b.irpp          or 0)),
                "statut":     b.statut,
            } for b in sorted(bulletins_site, key=lambda x: x.salarie.nom)]

        return {
            "site": s,
            "effectif_sal":   len(salaries_aff),
            "effectif_jour":  len(journaliers_aff),
            "effectif_total": len(salaries_aff) + len(journaliers_aff),
            # Pointage
            "nb_presences": nb_presences, "nb_absences": nb_absences,
            "taux_presence": taux_pres,
            "heures_normales": round(heures_normales, 1),
            "heures_sup": round(heures_sup, 1),
            "detail_sal": detail_sal,
            "detail_jour": detail_jour,
            # Paie journaliers
            "feuilles": detail_feuilles,
            "masse_jour_brut": round(masse_jour_brut),
            "feuilles_payees": feuilles_payees,
            "feuilles_attente": feuilles_attente,
            # Bulletins mensuels
            "bulletins": detail_bulletins,
            "masse_mensuelle": masse_mensuelle,
        }

    # Construire rapport(s)
    if site_sel:
        rapports = [_build_rapport_site(site_sel)]
    else:
        rapports = [_build_rapport_site(s) for s in sites_list]

    # Totaux globaux
    totaux = {
        "effectif": sum(r["effectif_total"] for r in rapports),
        "presences": sum(r["nb_presences"]   for r in rapports),
        "absences":  sum(r["nb_absences"]    for r in rapports),
        "h_normales": round(sum(r["heures_normales"] for r in rapports), 1),
        "h_sup":      round(sum(r["heures_sup"]     for r in rapports), 1),
        "masse_jour": sum(r["masse_jour_brut"]  for r in rapports),
        "masse_men":  sum(r["masse_mensuelle"].get("total_net", 0) for r in rapports),
    }

    return render_template("tenant/rapport_mensuel_site.html",
        tenant=t, rapports=rapports, totaux=totaux,
        sites=sites_list, site_sel=site_sel,
        mois=mois, annee=annee,
        mois_nom=MOIS_NOMS_LONG[mois],
        MOIS_NOMS=MOIS_NOMS_LONG,
        now=datetime.now())


@app.route("/rapports/mensuel-site/export")
@login_required
def rapport_mensuel_site_export():
    """Export Excel du rapport mensuel par site."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    import io, calendar

    if current_user.is_super_admin: return redirect(url_for("admin_dashboard"))
    t = get_tenant()
    if not t: return redirect(url_for("login"))

    MOIS_NOMS_LONG = ["","Janvier","Février","Mars","Avril","Mai","Juin",
                      "Juillet","Août","Septembre","Octobre","Novembre","Décembre"]

    mois   = request.args.get("mois",  type=int, default=datetime.now().month)
    annee  = request.args.get("annee", type=int, default=datetime.now().year)
    site_id= request.args.get("site_id", type=int)
    mois_nom = MOIS_NOMS_LONG[mois]
    _, nb_jours_mois = calendar.monthrange(annee, mois)
    mois_debut = date(annee, mois, 1)
    mois_fin   = date(annee, mois, nb_jours_mois)

    # Reconstruire le rapport (même logique que la route GET)
    # On rappelle simplement la route interne via redirect vers export dédié
    # Pour éviter la duplication, on re-calcule ici directement

    sites_list = Site.query.filter_by(tenant_id=t.id, actif=True).order_by(Site.nom).all()
    site_sel   = Site.query.get(site_id) if site_id else None
    sites_a_traiter = [site_sel] if site_sel else sites_list

    aff_sal  = {}
    aff_jour = {}
    for a in AffectationSite.query.filter_by(tenant_id=t.id).all():
        if a.actif or (a.date_fin and a.date_fin >= mois_debut):
            if a.salarie_id:    aff_sal[a.salarie_id]    = a.site_id
            if a.journalier_id: aff_jour[a.journalier_id] = a.site_id

    # Styles
    def hdr(ws, row, cols, texts, fill_color="1a2332"):
        for i, txt in enumerate(texts, 1):
            c = ws.cell(row, i, txt)
            c.font      = Font(bold=True, color="FFFFFF", size=9)
            c.fill      = PatternFill("solid", fgColor=fill_color)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border    = Border(**{s: Side(style="thin", color="D1D5DB")
                                    for s in ["left","right","top","bottom"]})
        ws.row_dimensions[row].height = 20

    def titre_section(ws, row, txt, color="374151", ncols=10):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
        c = ws.cell(row, 1, txt)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor=color)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 18

    MONEY = "#,##0"
    thin  = {s: Side(style="thin", color="E5E7EB") for s in ["left","right","top","bottom"]}
    EVEN  = PatternFill("solid", fgColor="F8FAFC")

    wb = Workbook()

    # ══════════════════════════════════════════════════════════════════════════
    # ONGLET 1 — SYNTHÈSE GLOBALE
    # ══════════════════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Synthèse"
    ws.freeze_panes = "A4"

    # Titre principal
    titre_doc = f"RAPPORT MENSUEL — {mois_nom.upper()} {annee} — {t.denomination}"
    if site_sel: titre_doc += f" — {site_sel.nom}"
    ws.merge_cells("A1:K1")
    ws["A1"] = titre_doc
    ws["A1"].font = Font(bold=True, size=13, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="1a2332")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24
    ws.append([f"Édité le {datetime.now().strftime('%d/%m/%Y à %H:%M')}"])
    ws.append([])

    hdr(ws, 3, 11,
        ["Site","Effectif","Mensuels","Journaliers","Présences","Absences",
         "Taux prés.","H.normales","H.sup","Masse journaliers (FCFA)","Net mensuels (FCFA)"])

    grand_tot = {"eff":0,"pres":0,"abs":0,"hn":0,"hs":0,"mj":0,"mn":0}
    for row_i, s in enumerate(sites_a_traiter, 4):
        ids_sal  = [k for k,v in aff_sal.items()  if v == s.id]
        ids_jour = [k for k,v in aff_jour.items() if v == s.id]
        pts = Pointage.query.filter_by(tenant_id=t.id)            .filter(Pointage.date_pointage >= mois_debut,
                    Pointage.date_pointage <= mois_fin)            .filter(db.or_(
                Pointage.salarie_id.in_(ids_sal)    if ids_sal  else db.false(),
                Pointage.journalier_id.in_(ids_jour) if ids_jour else db.false()
            )).all()
        nb_p  = sum(1 for p in pts if p.present)
        nb_a  = sum(1 for p in pts if p.absent)
        taux  = round(nb_p/(nb_p+nb_a)*100) if (nb_p+nb_a) > 0 else 0
        hn    = round(sum(float(p.heures_normales or 8) for p in pts if p.present), 1)
        hs    = round(sum(float(p.heures_sup_10 or 0)+float(p.heures_sup_30 or 0)+
                          float(p.heures_sup_40 or 0)+float(p.heures_sup_70 or 0)+
                          float(p.heures_sup or 0) for p in pts if p.present), 1)
        feuilles = FeuillePaieJournalier.query.filter_by(tenant_id=t.id)            .filter(FeuillePaieJournalier.journalier_id.in_(ids_jour),
                    FeuillePaieJournalier.date_debut >= mois_debut,
                    FeuillePaieJournalier.date_fin   <= mois_fin).all() if ids_jour else []
        mj = round(sum(float(f.montant_brut or 0) for f in feuilles))
        per= PeriodePaie.query.filter_by(tenant_id=t.id, annee=annee, mois=mois).first()
        mn = 0
        if per and ids_sal:
            buls = BulletinPaie.query.filter_by(tenant_id=t.id, periode_id=per.id)                .filter(BulletinPaie.salarie_id.in_(ids_sal)).all()
            mn = round(sum(float(b.net_a_payer or 0) for b in buls))
        eff = len(ids_sal) + len(ids_jour)
        row_data = [s.nom, eff, len(ids_sal), len(ids_jour),
                    nb_p, nb_a, f"{taux}%", hn, hs, mj, mn]
        ws.append(row_data)
        for ci, v in enumerate(row_data, 1):
            c = ws.cell(row_i, ci)
            c.border = Border(**thin)
            if row_i % 2 == 0: c.fill = EVEN
            if ci in (10, 11): c.number_format = MONEY; c.alignment = Alignment(horizontal="right")
            if ci in (5,6,7,8,9): c.alignment = Alignment(horizontal="center")
        grand_tot["eff"]+=eff; grand_tot["pres"]+=nb_p; grand_tot["abs"]+=nb_a
        grand_tot["hn"]+=hn; grand_tot["hs"]+=hs; grand_tot["mj"]+=mj; grand_tot["mn"]+=mn

    # Total
    tr = ws.max_row + 1
    ws.append(["TOTAL", grand_tot["eff"], "", "", grand_tot["pres"], grand_tot["abs"],
                "", round(grand_tot["hn"],1), round(grand_tot["hs"],1),
                grand_tot["mj"], grand_tot["mn"]])
    for ci in range(1, 12):
        c = ws.cell(tr, ci)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1a2332")
        if ci in (10, 11): c.number_format = MONEY; c.alignment = Alignment(horizontal="right")

    for i, w in enumerate([24,10,10,12,10,10,10,12,10,22,22], 1):
        ws.column_dimensions[ws.cell(1,i).column_letter].width = w

    # ══════════════════════════════════════════════════════════════════════════
    # ONGLETS PAR SITE
    # ══════════════════════════════════════════════════════════════════════════
    for s in sites_a_traiter:
        ws_s = wb.create_sheet(s.nom[:28])
        ws_s.freeze_panes = "A4"
        ids_sal  = [k for k,v in aff_sal.items()  if v == s.id]
        ids_jour = [k for k,v in aff_jour.items() if v == s.id]

        # Titre
        ws_s.merge_cells("A1:I1")
        ws_s["A1"] = f"{s.nom} — {mois_nom} {annee} — {t.denomination}"
        ws_s["A1"].font = Font(bold=True, size=12, color="FFFFFF")
        ws_s["A1"].fill = PatternFill("solid", fgColor="1a2332")
        ws_s["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws_s.row_dimensions[1].height = 20
        ws_s.append([])
        row_cur = 2

        # ── Section pointage ──────────────────────────────────────────────────
        row_cur += 1
        titre_section(ws_s, row_cur, f"📅 POINTAGE — {mois_nom} {annee}", "065f46", 9)
        row_cur += 1
        hdr(ws_s, row_cur, 9,
            ["Nom","Type","Emploi/Profession","Présences","Absences","Taux %","H.norm.","H.sup","Total h."],
            "065f46")
        row_cur += 1

        pts_sal  = Pointage.query.filter_by(tenant_id=t.id)            .filter(Pointage.salarie_id.in_(ids_sal),
                    Pointage.date_pointage >= mois_debut,
                    Pointage.date_pointage <= mois_fin).all() if ids_sal else []
        pts_jour = Pointage.query.filter_by(tenant_id=t.id)            .filter(Pointage.journalier_id.in_(ids_jour),
                    Pointage.date_pointage >= mois_debut,
                    Pointage.date_pointage <= mois_fin).all() if ids_jour else []

        salaries_aff    = Salarie.query.filter(Salarie.id.in_(ids_sal)).order_by(Salarie.nom).all() if ids_sal else []
        journaliers_aff = Journalier.query.filter(Journalier.id.in_(ids_jour)).order_by(Journalier.nom).all() if ids_jour else []

        for trv_list, pts_list, typ in [
            (salaries_aff, pts_sal, "Mensuel"),
            (journaliers_aff, pts_jour, "Journalier")
        ]:
            for trv in trv_list:
                if typ == "Mensuel":
                    pts_t = [p for p in pts_list if p.salarie_id == trv.id]
                    emploi = trv.emploi or "—"
                else:
                    pts_t = [p for p in pts_list if p.journalier_id == trv.id]
                    emploi = trv.profession or "—"
                nb_p = sum(1 for p in pts_t if p.present)
                nb_a = sum(1 for p in pts_t if p.absent)
                hn   = round(sum(float(p.heures_normales or 8) for p in pts_t if p.present), 1)
                hs   = round(sum(float(p.heures_sup_10 or 0)+float(p.heures_sup_30 or 0)+
                                 float(p.heures_sup_40 or 0)+float(p.heures_sup_70 or 0)+
                                 float(p.heures_sup or 0) for p in pts_t if p.present), 1)
                taux = round(nb_p/(nb_p+nb_a)*100) if (nb_p+nb_a) > 0 else 0
                row_d = [trv.nom_complet, typ, emploi, nb_p, nb_a, f"{taux}%", hn, hs, round(hn+hs,1)]
                ws_s.append(row_d)
                for ci, v in enumerate(row_d, 1):
                    c = ws_s.cell(row_cur, ci)
                    c.border = Border(**thin)
                    if row_cur % 2 == 0: c.fill = EVEN
                    if ci in (4,5,6,7,8,9): c.alignment = Alignment(horizontal="center")
                row_cur += 1

        ws_s.append([])
        row_cur += 1

        # ── Section paie journaliers ──────────────────────────────────────────
        if ids_jour:
            titre_section(ws_s, row_cur, f"🦺 PAIE JOURNALIERS — {mois_nom} {annee}", "92400e", 9)
            row_cur += 1
            hdr(ws_s, row_cur, 9,
                ["Journalier","Profession","Période du","au","Nb jours","Heures","Taux/h","Montant (FCFA)","Statut"],
                "92400e")
            row_cur += 1
            feuilles = FeuillePaieJournalier.query.filter_by(tenant_id=t.id)                .filter(FeuillePaieJournalier.journalier_id.in_(ids_jour),
                        FeuillePaieJournalier.date_debut >= mois_debut,
                        FeuillePaieJournalier.date_fin   <= mois_fin).all()
            total_feuilles = 0
            for f in feuilles:
                m = round(float(f.montant_brut or 0)); total_feuilles += m
                row_d = [f.journalier.nom_complet, f.journalier.profession or "—",
                         f.date_debut.strftime("%d/%m/%Y") if f.date_debut else "",
                         f.date_fin.strftime("%d/%m/%Y")   if f.date_fin   else "",
                         f.nb_jours, round(float(f.total_heures or 0),1),
                         round(float(f.taux_horaire or 0)), m, f.statut]
                ws_s.append(row_d)
                for ci, v in enumerate(row_d, 1):
                    c = ws_s.cell(row_cur, ci)
                    c.border = Border(**thin)
                    if row_cur % 2 == 0: c.fill = EVEN
                    if ci == 8: c.number_format = MONEY; c.alignment = Alignment(horizontal="right")
                    if ci in (5,6,7): c.alignment = Alignment(horizontal="center")
                row_cur += 1
            # Total
            ws_s.append(["TOTAL JOURNALIERS", "", "", "", "", "", "", total_feuilles, ""])
            for ci in range(1, 10):
                c = ws_s.cell(row_cur, ci)
                c.font = Font(bold=True, color="FFFFFF")
                c.fill = PatternFill("solid", fgColor="92400e")
                if ci == 8: c.number_format = MONEY; c.alignment = Alignment(horizontal="right")
            row_cur += 2; ws_s.append([])

        # ── Section bulletins mensuels ────────────────────────────────────────
        if ids_sal:
            per = PeriodePaie.query.filter_by(tenant_id=t.id, annee=annee, mois=mois).first()
            if per:
                buls = BulletinPaie.query.filter_by(tenant_id=t.id, periode_id=per.id)                    .filter(BulletinPaie.salarie_id.in_(ids_sal)).all()
                if buls:
                    titre_section(ws_s, row_cur,
                        f"📄 BULLETINS DE PAIE — {mois_nom} {annee}", "1e40af", 9)
                    row_cur += 1
                    hdr(ws_s, row_cur, 9,
                        ["Salarié","Matricule","Emploi","Brut (FCFA)","CNSS sal.",
                         "TCS","IRPP","Net à payer (FCFA)","Statut"], "1e40af")
                    row_cur += 1
                    total_brut = total_net = 0
                    for b in sorted(buls, key=lambda x: x.salarie.nom):
                        brut = round(float(b.salaire_brut or 0))
                        net  = round(float(b.net_a_payer  or 0))
                        total_brut += brut; total_net += net
                        row_d = [b.salarie.nom_complet, b.salarie.matricule,
                                 b.salarie.emploi or "—", brut,
                                 round(float(b.cnss_salarie or 0)),
                                 round(float(b.tcs  or 0)),
                                 round(float(b.irpp or 0)), net, b.statut]
                        ws_s.append(row_d)
                        for ci, v in enumerate(row_d, 1):
                            c = ws_s.cell(row_cur, ci)
                            c.border = Border(**thin)
                            if row_cur % 2 == 0: c.fill = EVEN
                            if ci in (4,5,6,7,8): c.number_format = MONEY; c.alignment = Alignment(horizontal="right")
                        row_cur += 1
                    ws_s.append(["TOTAL SALARIÉS","","",total_brut,"","","",total_net,""])
                    for ci in range(1, 10):
                        c = ws_s.cell(row_cur, ci)
                        c.font = Font(bold=True, color="FFFFFF")
                        c.fill = PatternFill("solid", fgColor="1e40af")
                        if ci in (4,8): c.number_format = MONEY; c.alignment = Alignment(horizontal="right")

        for i, w in enumerate([28,12,20,10,10,10,10,20,12], 1):
            ws_s.column_dimensions[ws_s.cell(1,i).column_letter].width = w

    # Export
    out = io.BytesIO(); wb.save(out); out.seek(0)
    fname_parts = [f"Rapport_{mois_nom}_{annee}"]
    if site_sel: fname_parts.append(site_sel.nom.replace(" ","_"))
    fname_parts.append(datetime.now().strftime("%Y%m%d"))
    return send_file(out,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True, download_name="_".join(fname_parts)+".xlsx")





def _gen_excel_cnss(tenant, trim_label, annee, mois_labels,
                    sal_data, total_base_cnss, total_base_cnamgs,
                    tot_cnss_m):
    """
    Génère la feuille CNSS conforme au formulaire officiel gabonais.
    sal_data : liste de dicts {nom_complet, matricule, numero_cnss,
               date_embauche, m1_base_cnss, m2_base_cnss, m3_base_cnss}
    N_HRS = 8 (constante légale)
    """
    import openpyxl, io
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    BD2=Border(left=Side(style="thin",color="E5E7EB"),right=Side(style="thin",color="E5E7EB"),
               top=Side(style="thin",color="E5E7EB"),bottom=Side(style="thin",color="E5E7EB"))
    CTR2=Alignment(horizontal="center",vertical="center",wrap_text=True)
    LFT2=Alignment(horizontal="left",  vertical="center",wrap_text=True)
    RGT2=Alignment(horizontal="right", vertical="center")
    HF=PatternFill("solid",fgColor="1a2332"); HN=Font(bold=True,color="FFFFFF",size=9)
    TF=PatternFill("solid",fgColor="D6EAF8"); GF=PatternFill("solid",fgColor="E8F4FD")
    AF=PatternFill("solid",fgColor="EBF5FB")

    def Cx(ws,r,c,val,font=None,fill=None,align=None,fmt=None,span=None):
        if span and span>1:
            ws.merge_cells(start_row=r,start_column=c,end_row=r,end_column=c+span-1)
        cell=ws.cell(r,c)
        if val  is not None: cell.value=val
        if font is not None: cell.font=font
        if fill is not None: cell.fill=fill
        if align is not None: cell.alignment=align
        if fmt  is not None: cell.number_format=fmt
        cell.border=BD2; return cell

    wb=openpyxl.Workbook(); ws=wb.active; ws.title="CNSS"
    for i,w in enumerate([5,14,12,28,8,12,12,16,14,4,16,14,4,16,14,4,10,10],1):
        ws.column_dimensions[get_column_letter(i)].width=w

    def rh(r,h): ws.row_dimensions[r].height=h

    # ── En-tête employeur ─────────────────────────────────────────────────────
    rh(10,18)
    Cx(ws,10,1,"Matricule employeur",Font(bold=True,size=9),align=LFT2)
    Cx(ws,10,2,tenant.numero_cnss or "—",Font(size=9),align=LFT2)
    Cx(ws,10,5,"Période",Font(bold=True,size=9),align=CTR2)
    Cx(ws,10,6,trim_label,Font(bold=True,size=11),align=CTR2)
    Cx(ws,10,7,"Année",Font(bold=True,size=9),align=CTR2)
    Cx(ws,10,8,annee,Font(bold=True,size=11),align=CTR2)
    Cx(ws,10,13,"CACHET ET SIGNATURE",HN,HF,CTR2,span=4)

    rh(12,20)
    Cx(ws,12,1,"Nom ou Raison Sociale",Font(bold=True,size=9),align=LFT2)
    Cx(ws,12,2,tenant.denomination,Font(bold=True,size=9),align=LFT2,span=3)

    rh(15,16); Cx(ws,15,1,"B.P :",Font(size=9),align=LFT2)
    Cx(ws,15,3,f"VILLE : {getattr(tenant,'ville','Libreville')}",Font(size=9),align=LFT2,span=2)
    rh(17,16); Cx(ws,17,1,"TEL :",Font(size=9),align=LFT2)
    Cx(ws,17,2,getattr(tenant,"telephone",""),Font(size=9),align=LFT2)
    Cx(ws,17,9,"Effectif total",Font(bold=True,size=9),align=CTR2)
    Cx(ws,17,10,len(sal_data),Font(bold=True,size=11),TF,CTR2)
    rh(19,16); Cx(ws,19,1,"Email :",Font(size=9),align=LFT2)

    # ── Résumé cotisations ────────────────────────────────────────────────────
    rh(20,32)
    Cx(ws,20,2,"Rémunération totale plafonnée CNSS",Font(bold=True,size=8),GF,CTR2,span=3)
    Cx(ws,20,5,"Montant déduction Alloc. Familiales",Font(size=8),GF,CTR2,span=3)
    Cx(ws,20,9,"Rémunération totale plafonnée CNAMGS",Font(bold=True,size=8),GF,CTR2,span=3)
    Cx(ws,20,13,"DATE DE RECEPTION",HN,HF,CTR2,span=4)
    rh(21,18)
    Cx(ws,21,2,total_base_cnss,Font(bold=True,size=10),TF,RGT2,"#,##0",span=3)
    Cx(ws,21,5,0,Font(size=9),TF,RGT2,"#,##0",span=3)
    Cx(ws,21,9,total_base_cnamgs,Font(bold=True,size=10),TF,RGT2,"#,##0",span=3)
    rh(22,18)
    Cx(ws,22,2,"Cotisations brutes dues CNSS",Font(bold=True,size=8),align=CTR2,span=3)
    Cx(ws,22,5,"Cotisations nettes dues CNSS",Font(bold=True,size=8),align=CTR2,span=3)
    Cx(ws,22,9,"Cotisations nettes dues CNAMGS",Font(bold=True,size=8),align=CTR2,span=3)
    rh(23,18)
    cot_cnss   = round(total_base_cnss   * 0.23)
    cot_cnamgs = round(total_base_cnamgs * 0.061)
    Cx(ws,23,2,cot_cnss,  Font(bold=True,size=10),TF,RGT2,"#,##0",span=3)
    Cx(ws,23,5,cot_cnss,  Font(bold=True,size=10),TF,RGT2,"#,##0",span=3)
    Cx(ws,23,9,cot_cnamgs,Font(bold=True,size=10),TF,RGT2,"#,##0",span=3)

    # ── En-têtes mois ─────────────────────────────────────────────────────────
    rh(25,14)
    for col,lbl in [(8,mois_labels[0]),(11,mois_labels[1]),(14,mois_labels[2])]:
        Cx(ws,25,col,lbl,HN,HF,CTR2)
    rh(26,36)
    for col,lbl in [(1,"N°"),(2,"N°CNSS /\nN°CNAMGS"),(3,"N° Paie"),
                    (4,"NOM ET PRENOM"),(5,"Taux CNSS"),(6,"EMBAUCHE"),(7,"CESSATION"),
                    (8,"SALAIRE\nPLAFONNE"),(9,"SALAIRE\nDEPLAFONNE"),(10,"Nbre\nHrs"),
                    (11,"SALAIRE\nPLAFONNE"),(12,"SALAIRE\nDEPLAFONNE"),(13,"Nbre\nHrs"),
                    (14,"SALAIRE\nPLAFONNE"),(15,"SALAIRE\nDEPLAFONNE"),(16,"Nbre\nHrs")]:
        Cx(ws,26,col,lbl,HN,HF,CTR2)

    # ── 2 lignes par employé ──────────────────────────────────────────────────
    N_HRS_CNSS = 8  # toujours 8
    dr = 27
    for i, sal in enumerate(sal_data, 1):
        r1 = dr + (i-1)*2
        r2 = dr + (i-1)*2 + 1
        bg = AF if i%2==0 else None
        rh(r1,16); rh(r2,16)

        # Ligne impaire : numéro + taux + données BASE CNSS + heures
        Cx(ws,r1,1,i,Font(size=9),bg,CTR2)
        Cx(ws,r1,5,23,Font(size=9),bg,CTR2)  # Taux CNSS = 23%
        emb = sal.get("date_embauche","")
        Cx(ws,r1,6,emb,Font(size=8),bg,CTR2)
        Cx(ws,r1,7,"",None,bg,CTR2)
        # Mois 1, 2, 3 — BASE CNSS (plafonnée), N_HRS=8
        Cx(ws,r1,8, sal.get("m1_base_cnss",0), Font(size=9),bg,RGT2,"#,##0")
        Cx(ws,r1,9, sal.get("m1_base_cnss",0), Font(size=9),bg,RGT2,"#,##0")  # déplafonné = même valeur
        Cx(ws,r1,10,N_HRS_CNSS,Font(size=9),bg,CTR2)
        Cx(ws,r1,11,sal.get("m2_base_cnss",0), Font(size=9),bg,RGT2,"#,##0")
        Cx(ws,r1,12,sal.get("m2_base_cnss",0), Font(size=9),bg,RGT2,"#,##0")
        Cx(ws,r1,13,N_HRS_CNSS,Font(size=9),bg,CTR2)
        Cx(ws,r1,14,sal.get("m3_base_cnss",0), Font(size=9),bg,RGT2,"#,##0")
        Cx(ws,r1,15,sal.get("m3_base_cnss",0), Font(size=9),bg,RGT2,"#,##0")
        Cx(ws,r1,16,N_HRS_CNSS,Font(size=9),bg,CTR2)

        # Ligne paire : N°CNSS + Matricule + NOM
        Cx(ws,r2,2,sal.get("numero_cnss",""),Font(size=9),bg,CTR2)
        Cx(ws,r2,3,sal.get("matricule",""),  Font(size=9),bg,CTR2)
        Cx(ws,r2,4,sal.get("nom_complet",""),Font(bold=True,size=9),bg,LFT2)
        for c in [1,5,6,7,8,9,10,11,12,13,14,15,16]:
            Cx(ws,r2,c,"",None,bg,None)

    # ── Sous-total ────────────────────────────────────────────────────────────
    rst = dr + len(sal_data)*2; rh(rst,18)
    Cx(ws,rst,1,"SOUS TOTAL À REPORTER PAGE SUIVANTE",Font(bold=True,size=9),TF,LFT2,span=7)
    for col,val in [(8,tot_cnss_m[0]),(11,tot_cnss_m[1]),(14,tot_cnss_m[2])]:
        Cx(ws,rst,col,val,Font(bold=True),TF,RGT2,"#,##0")
        for cc in [col+1,col+2]: Cx(ws,rst,cc,"",None,TF,None)

    # ── RECAP ─────────────────────────────────────────────────────────────────
    rr = rst+2; rh(rr,18)
    Cx(ws,rr,1,"RECAP",HN,HF,CTR2,span=2)
    Cx(ws,rr,3,"TAUX",HN,HF,CTR2)
    Cx(ws,rr,4,23,Font(bold=True,size=10),TF,CTR2)
    Cx(ws,rr,5,"MASSE SALARIALE PLAFONNEE CNSS",Font(bold=True,size=8),TF,LFT2,span=3)
    for col,val in [(8,tot_cnss_m[0]),(11,tot_cnss_m[1]),(14,tot_cnss_m[2])]:
        Cx(ws,rr,col,val,Font(bold=True),TF,RGT2,"#,##0")
        for cc in [col+1,col+2]: Cx(ws,rr,cc,"",None,TF,None)

    rcot = rr+1; rh(rcot,24)
    Cx(ws,rcot,1,"COTISATION GLOBALE DUE (CNSS)",Font(bold=True,size=11),HF,LFT2,span=13)
    Cx(ws,rcot,14,cot_cnss,Font(bold=True,size=13),TF,RGT2,"#,##0",span=3)

    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.getvalue()


def _gen_excel_cnamgs(tenant, trim_label, annee, mois_labels,
                      sal_data, total_base_cnamgs, tot_cnamgs_m):
    """
    Génère la feuille CNAMGS conforme au formulaire officiel gabonais.
    sal_data : liste de dicts {nom_complet, matricule, numero_cnamgs,
               date_embauche, m1_base_cnamgs, m2_base_cnamgs, m3_base_cnamgs}
    N_HRS = 8 (constante légale)
    """
    import openpyxl, io
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    BD2=Border(left=Side(style="thin",color="E5E7EB"),right=Side(style="thin",color="E5E7EB"),
               top=Side(style="thin",color="E5E7EB"),bottom=Side(style="thin",color="E5E7EB"))
    CTR2=Alignment(horizontal="center",vertical="center",wrap_text=True)
    LFT2=Alignment(horizontal="left",  vertical="center",wrap_text=True)
    RGT2=Alignment(horizontal="right", vertical="center")
    HF=PatternFill("solid",fgColor="1a2332"); HN=Font(bold=True,color="FFFFFF",size=9)
    T2=PatternFill("solid",fgColor="D5F5E3"); G2=PatternFill("solid",fgColor="E8F8F5")
    YF=PatternFill("solid",fgColor="FEF9E7")

    def Cx(ws,r,c,val,font=None,fill=None,align=None,fmt=None,span=None):
        if span and span>1:
            ws.merge_cells(start_row=r,start_column=c,end_row=r,end_column=c+span-1)
        cell=ws.cell(r,c)
        if val  is not None: cell.value=val
        if font is not None: cell.font=font
        if fill is not None: cell.fill=fill
        if align is not None: cell.alignment=align
        if fmt  is not None: cell.number_format=fmt
        cell.border=BD2; return cell

    wb=openpyxl.Workbook(); ws=wb.active; ws.title="CNAMGS"
    for i,w in enumerate([5,18,28,14,14,18,10,18,10,18,10],1):
        ws.column_dimensions[get_column_letter(i)].width=w

    def rh(r,h): ws.row_dimensions[r].height=h

    # ── Titre ─────────────────────────────────────────────────────────────────
    rh(5,22)
    Cx(ws,5,4,"DECLARATION TRIMESTRIELLE DE SALAIRES",
       Font(bold=True,size=13,color="1a2332"),None,CTR2,span=6)

    # ── En-tête ───────────────────────────────────────────────────────────────
    rh(8,16)
    Cx(ws,8,4,"Période :",Font(bold=True,size=9),align=CTR2)
    Cx(ws,8,5,trim_label,Font(bold=True,size=11),align=CTR2)
    Cx(ws,8,6,annee,Font(bold=True,size=11),align=CTR2)
    rh(9,16)
    Cx(ws,9,1,"Matricule employeur CNAMGS",Font(bold=True,size=9),align=LFT2)
    Cx(ws,9,2,getattr(tenant,"numero_cnamgs","—"),Font(size=9),align=LFT2,span=2)
    rh(11,20)
    Cx(ws,11,1,"Nom ou Raison Sociale",Font(bold=True,size=9),align=LFT2)
    Cx(ws,11,2,tenant.denomination,Font(bold=True,size=9),align=LFT2,span=3)
    Cx(ws,11,5,trim_label,Font(bold=True,size=11,color="1a2332"),align=CTR2)
    Cx(ws,11,6,annee,Font(bold=True,size=11),align=CTR2)
    Cx(ws,11,8,"CACHET ET SIGNATURE",HN,HF,CTR2,span=4)

    # Taux (sur lignes séparées pour éviter conflits merge)
    rh(13,16)
    Cx(ws,13,4,"Taux de cotisation",Font(bold=True,size=9),T2,LFT2,span=2)
    rh(14,16)
    Cx(ws,14,1,"B.P :",Font(size=9),align=LFT2)
    Cx(ws,14,2,getattr(tenant,"adresse",""),Font(size=9),align=LFT2,span=2)
    Cx(ws,14,4,"Employeur",Font(bold=True,size=9),T2,CTR2)
    Cx(ws,14,5,0.041,Font(bold=True,size=9),T2,RGT2,fmt="0.0%")
    rh(15,16)
    Cx(ws,15,1,"VILLE :",Font(size=9),align=LFT2)
    Cx(ws,15,2,getattr(tenant,"ville","Libreville"),Font(size=9),align=LFT2)
    Cx(ws,15,4,"Travailleur",Font(bold=True,size=9),T2,CTR2)
    Cx(ws,15,5,0.02,Font(bold=True,size=9),T2,RGT2,fmt="0.0%")
    rh(16,16)
    Cx(ws,16,1,"TEL :",Font(size=9),align=LFT2)
    Cx(ws,16,2,getattr(tenant,"telephone",""),Font(size=9),align=LFT2)
    rh(18,16)
    Cx(ws,18,4,"Plafond mensuel CNAMGS",Font(bold=True,size=9),T2,LFT2)
    Cx(ws,18,5,2500000,Font(bold=True,size=9),T2,RGT2,fmt="#,##0")

    # ── Cotisations nettes dues ───────────────────────────────────────────────
    rh(19,16)
    Cx(ws,19,2,"Cotisations nettes dues CNAMGS",Font(bold=True,size=9),align=LFT2,span=4)
    Cx(ws,19,8,"DATE DE RECEPTION",HN,HF,CTR2,span=4)
    rh(20,20)
    Cx(ws,20,2,round(total_base_cnamgs*0.061),Font(bold=True,size=12,color="1e40af"),T2,RGT2,"#,##0",span=4)
    rh(21,16)
    Cx(ws,21,2,"Cotisations payées à la CNAMGS",Font(italic=True,size=9),align=LFT2,span=4)

    # ── RECAP ─────────────────────────────────────────────────────────────────
    rh(24,18)
    Cx(ws,24,1,"Recap.",HN,HF,CTR2)
    Cx(ws,24,2,"Effectif",HN,HF,CTR2)
    Cx(ws,24,3,len(sal_data),Font(bold=True,size=11),T2,CTR2)
    Cx(ws,24,4,"MASSE SALARIALE SOUMISE À COTISATION :",Font(bold=True,size=8),T2,LFT2,span=2)
    Cx(ws,24,6,total_base_cnamgs,Font(bold=True,size=10),T2,RGT2,"#,##0")
    Cx(ws,24,8,"COTISATIONS SOCIALES:",Font(bold=True,size=9),T2,LFT2)
    Cx(ws,24,10,round(total_base_cnamgs*0.061),Font(bold=True,size=10),T2,RGT2,"#,##0")
    rh(25,16)
    Cx(ws,25,8,"Part patronale (4.1%)",Font(size=9),G2,LFT2)
    Cx(ws,25,10,round(total_base_cnamgs*0.041),Font(bold=True,size=9),G2,RGT2,"#,##0")
    rh(26,16)
    Cx(ws,26,8,"Part salariale (2%)",Font(size=9),G2,LFT2)
    Cx(ws,26,10,round(total_base_cnamgs*0.02),Font(bold=True,size=9),G2,RGT2,"#,##0")
    rh(27,18)
    Cx(ws,27,1,"TOTAL À REPORTER PAGE SUIVANTE",HN,HF,LFT2,span=5)
    for col,val in [(6,tot_cnamgs_m[0]),(8,tot_cnamgs_m[1]),(10,tot_cnamgs_m[2])]:
        Cx(ws,27,col,val,Font(bold=True),T2,RGT2,"#,##0")

    # ── Labels mois ───────────────────────────────────────────────────────────
    rh(28,14)
    for col,lbl in [(6,mois_labels[0]),(8,mois_labels[1]),(10,mois_labels[2])]:
        Cx(ws,28,col,lbl,HN,HF,CTR2,span=2)

    # ── En-têtes colonnes employés ────────────────────────────────────────────
    rh(38,14)
    Cx(ws,38,4,"Date",HN,HF,CTR2,span=2)
    for col,lbl in [(6,mois_labels[0]),(8,mois_labels[1]),(10,mois_labels[2])]:
        Cx(ws,38,col,lbl,HN,HF,CTR2,span=2)
    rh(39,36)
    for col,lbl in [(1,"N°"),(2,"Matricule"),(3,"NOM ET PRENOM"),
                    (4,"EMBAUCHE"),(5,"CESSATION"),
                    (6,"Assiette soumise\nà cotisation"),(7,"Nbre\nHrs/Jrs"),
                    (8,"Assiette soumise\nà cotisation"),(9,"Nbre\nHrs/Jrs"),
                    (10,"Assiette soumise\nà cotisation"),(11,"Nbre\nHrs/Jrs")]:
        Cx(ws,39,col,lbl,HN,HF,CTR2)

    # ── 1 ligne par employé — BASE CNAMGS, N_HRS=8 ───────────────────────────
    N_HRS_CNAMGS = 8  # toujours 8
    for i, sal in enumerate(sal_data, 1):
        r = 39 + i; rh(r,18)
        bg = PatternFill("solid",fgColor="E8F8F5") if i%2==0 else None
        Cx(ws,r,1,i,Font(size=9),bg,CTR2)
        Cx(ws,r,2,sal.get("matricule",""),Font(size=9),bg,CTR2)
        Cx(ws,r,3,sal.get("nom_complet",""),Font(bold=True,size=9),bg,LFT2)
        Cx(ws,r,4,sal.get("date_embauche",""),Font(size=8),bg,CTR2)
        Cx(ws,r,5,"",None,bg,CTR2)
        # BASE CNAMGS + toujours 8h
        Cx(ws,r,6, sal.get("m1_base_cnamgs",0),Font(size=9),bg,RGT2,"#,##0")
        Cx(ws,r,7, N_HRS_CNAMGS,Font(size=9),bg,CTR2)
        Cx(ws,r,8, sal.get("m2_base_cnamgs",0),Font(size=9),bg,RGT2,"#,##0")
        Cx(ws,r,9, N_HRS_CNAMGS,Font(size=9),bg,CTR2)
        Cx(ws,r,10,sal.get("m3_base_cnamgs",0),Font(size=9),bg,RGT2,"#,##0")
        Cx(ws,r,11,N_HRS_CNAMGS,Font(size=9),bg,CTR2)

    # ── Note pénalités ────────────────────────────────────────────────────────
    r_note = 39 + len(sal_data) + 2; rh(r_note,60)
    Cx(ws,r_note,1,
       "Au-delà de la date limite, une pénalité est appliquée conformément à la loi :\n"
       "- 25% pour non dépôt de la DTS calculé sur le montant de la DTS du dernier trimestre déclaré ;\n"
       "- 2% pour non paiement des cotisations par mois de retard cumulable au prorata temporis.",
       Font(italic=True,size=8), YF,
       Alignment(horizontal="left",vertical="top",wrap_text=True), span=11)

    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.getvalue()


# ══════════════════════════════════════════════════════════════════════════════
# DÉCLARATIONS SOCIALES & FISCALES
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/declaration-cnss")
@login_required
def declaration_cnss():
    """Déclarations sociales : CNSS/CNAMGS (trimestrielles) + CFP/FNH/TCS/IRPP (mensuelles)."""
    if current_user.is_super_admin: return redirect(url_for("admin_dashboard"))
    t = get_tenant()
    if not t: return redirect(url_for("login"))

    periodes = PeriodePaie.query.filter_by(tenant_id=t.id)        .order_by(PeriodePaie.annee.desc(), PeriodePaie.mois.desc()).all()

    # Période sélectionnée
    pid     = request.args.get("periode_id", type=int)
    periode = PeriodePaie.query.filter_by(id=pid, tenant_id=t.id).first() if pid else               (periodes[0] if periodes else None)

    # Mode : mensuel (CFP/FNH/TCS/IRPP) ou trimestriel (CNSS/CNAMGS)
    mode    = request.args.get("mode", "mensuel")  # mensuel | trimestriel

    bulletins_mois = []
    bulletins_trim = []
    stats_mensuel  = {}
    stats_trim     = {}

    if periode:
        def s(buls, field): return round(sum(float(getattr(b, field) or 0) for b in buls), 2)

        # ── Bulletins du mois sélectionné (CFP, FNH, TCS, IRPP) ─────────────
        bulletins_mois = BulletinPaie.query.filter_by(
            tenant_id=t.id, periode_id=periode.id
        ).options(joinedload(BulletinPaie.salarie)).all()

        stats_mensuel = {
            "nb":           len(bulletins_mois),
            "total_brut":   s(bulletins_mois, "salaire_brut"),
            "total_cfp":    s(bulletins_mois, "cfp"),
            "total_fnh":    s(bulletins_mois, "fnh"),
            "total_tcs":    s(bulletins_mois, "tcs"),
            "total_irpp":   s(bulletins_mois, "irpp"),
        }
        stats_mensuel["total_mensuel"] = (stats_mensuel["total_cfp"]  +
                                          stats_mensuel["total_fnh"]  +
                                          stats_mensuel["total_tcs"]  +
                                          stats_mensuel["total_irpp"])

        # ── Bulletins du trimestre (CNSS/CNAMGS) ─────────────────────────────
        # Trimestre : T1=Jan-Mar, T2=Avr-Jun, T3=Jul-Sep, T4=Oct-Dec
        mois = periode.mois
        trim_debut = ((mois - 1) // 3) * 3 + 1   # 1, 4, 7, 10
        trim_fin   = trim_debut + 2                # 3, 6, 9, 12
        trim_num   = (mois - 1) // 3 + 1          # 1, 2, 3, 4
        trim_label = f"T{trim_num} {periode.annee} ({['Jan-Mar','Avr-Jun','Jul-Sep','Oct-Déc'][trim_num-1]})"

        periodes_trim = PeriodePaie.query.filter_by(
            tenant_id=t.id, annee=periode.annee
        ).filter(
            PeriodePaie.mois >= trim_debut,
            PeriodePaie.mois <= trim_fin
        ).all()
        ids_trim = [p.id for p in periodes_trim]

        bulletins_trim = BulletinPaie.query.filter(
            BulletinPaie.tenant_id == t.id,
            BulletinPaie.periode_id.in_(ids_trim)
        ).options(joinedload(BulletinPaie.salarie),
                  joinedload(BulletinPaie.periode)).all()

        # Regrouper par salarié pour le trimestre
        from collections import defaultdict
        sal_trim = defaultdict(lambda: {
            "salarie": None, "brut": 0, "base_cnss": 0,
            "cnss_sal": 0, "cnss_pat": 0,
            "base_cnamgs": 0, "cnamgs_sal": 0, "cnamgs_pat": 0,
            "mois_list": []
        })
        for b in bulletins_trim:
            k = b.salarie_id
            sal_trim[k]["salarie"]    = b.salarie
            sal_trim[k]["brut"]      += float(b.salaire_brut   or 0)
            sal_trim[k]["base_cnss"] += float(b.base_cnss      or 0)
            sal_trim[k]["cnss_sal"]  += float(b.cnss_salarie   or 0)
            sal_trim[k]["cnss_pat"]  += float(b.cnss_patronale or 0)
            sal_trim[k]["base_cnamgs"]+= float(b.base_cnamgs   or 0)
            sal_trim[k]["cnamgs_sal"]+= float(b.cnamgs_salarie  or 0)
            sal_trim[k]["cnamgs_pat"]+= float(b.cnamgs_patronale or 0)
            sal_trim[k]["mois_list"].append(b.periode.mois if b.periode else 0)

        lignes_trim = sorted(sal_trim.values(), key=lambda x: x["salarie"].nom if x["salarie"] else "")

        stats_trim = {
            "trim_label":    trim_label,
            "trim_num":      trim_num,
            "nb":            len(lignes_trim),
            "mois_couverts": sorted(set(
                m for lg in lignes_trim for m in lg["mois_list"]
            )),
            "total_brut":    sum(lg["brut"]       for lg in lignes_trim),
            "total_cnss_sal":sum(lg["cnss_sal"]   for lg in lignes_trim),
            "total_cnss_pat":sum(lg["cnss_pat"]   for lg in lignes_trim),
            "total_cnamgs_sal":sum(lg["cnamgs_sal"]  for lg in lignes_trim),
            "total_cnamgs_pat":sum(lg["cnamgs_pat"]  for lg in lignes_trim),
        }
        stats_trim["total_cnss"]     = stats_trim["total_cnss_sal"]   + stats_trim["total_cnss_pat"]
        stats_trim["total_cnamgs"]   = stats_trim["total_cnamgs_sal"] + stats_trim["total_cnamgs_pat"]
        stats_trim["total_a_verser"] = stats_trim["total_cnss"] + stats_trim["total_cnamgs"]

    else:
        lignes_trim = []

    MOIS_FR = ["","Jan","Fév","Mar","Avr","Mai","Jun","Jul","Aoû","Sep","Oct","Nov","Déc"]

    return render_template("tenant/declaration_cnss.html",
        tenant=t, periodes=periodes, periode=periode, mode=mode,
        bulletins_mois=bulletins_mois, stats_mensuel=stats_mensuel,
        lignes_trim=lignes_trim, stats_trim=stats_trim,
        MOIS_FR=MOIS_FR)


# ══════════════════════════════════════════════════════════════════════════════
# EXPORT COMPTABLE SAGE 100
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/export/sage/journal/<int:periode_id>")
@tenant_required
def export_sage_journal(periode_id):
    """
    Export du journal de paie mensuel au format Sage 100 (.txt).
    Importable dans Sage 100 Comptabilité via Fichier → Importer → Journal.
    Seuls les bulletins VALIDE sont inclus.
    """
    t = get_tenant()
    periode = PeriodePaie.query.filter_by(id=periode_id, tenant_id=t.id).first_or_404()
    bulletins = (BulletinPaie.query
                 .filter(BulletinPaie.periode_id==periode_id, BulletinPaie.tenant_id==t.id, BulletinPaie.statut.in_(["VALIDE","VALIDÉ"]))
                 .join(Salarie)
                 .order_by(Salarie.nom)
                 .all())

    if not bulletins:
        flash("Aucun bulletin validé pour cette période. Validez les bulletins avant l'export.", "warning")
        return redirect(url_for("bulletins"))

    try:
        from export_comptable import generer_journal_paie, ExportVide
        contenu = generer_journal_paie(bulletins, periode, t)
        nom_fichier = f"journal_paie_{periode.mois:02d}{periode.annee}_{t.sigle or t.id}.txt"
        logger.info(f"[Export Sage] Journal paie — tenant={t.id} période={periode.libelle_complet}")
        return send_file(
            io.BytesIO(contenu),
            mimetype="text/plain",
            as_attachment=True,
            download_name=nom_fichier,
        )
    except Exception as e:
        logger.error(f"[Export Sage] Erreur journal : {e}")
        flash(f"Erreur lors de la génération : {e}", "error")
        return redirect(url_for("bulletins"))


@app.route("/export/sage/livre/<int:periode_id>")
@tenant_required
def export_sage_livre(periode_id):
    """
    Export du livre de paie détaillé par salarié au format CSV (.csv).
    Compatible Excel et importable dans Sage 100.
    Seuls les bulletins VALIDE sont inclus.
    """
    t = get_tenant()
    periode = PeriodePaie.query.filter_by(id=periode_id, tenant_id=t.id).first_or_404()
    bulletins = (BulletinPaie.query
                 .filter(BulletinPaie.periode_id==periode_id, BulletinPaie.tenant_id==t.id, BulletinPaie.statut.in_(["VALIDE","VALIDÉ"]))
                 .join(Salarie)
                 .order_by(Salarie.nom)
                 .all())

    if not bulletins:
        flash("Aucun bulletin validé pour cette période. Validez les bulletins avant l'export.", "warning")
        return redirect(url_for("bulletins"))

    try:
        from export_comptable import generer_livre_paie, ExportVide
        contenu = generer_livre_paie(bulletins, periode, t)
        nom_fichier = f"livre_paie_{periode.mois:02d}{periode.annee}_{t.sigle or t.id}.csv"
        logger.info(f"[Export Sage] Livre paie — tenant={t.id} période={periode.libelle_complet}")
        return send_file(
            io.BytesIO(contenu),
            mimetype="text/csv; charset=utf-8",
            as_attachment=True,
            download_name=nom_fichier,
        )
    except Exception as e:
        logger.error(f"[Export Sage] Erreur livre : {e}")
        flash(f"Erreur lors de la génération : {e}", "error")
        return redirect(url_for("bulletins"))


@app.route("/export/sage/les-deux/<int:periode_id>")
@tenant_required
def export_sage_les_deux(periode_id):
    """
    Export des deux fichiers (journal + livre) dans une archive ZIP.
    Pratique pour envoyer tout au comptable en une fois.
    """
    import zipfile
    t = get_tenant()
    periode = PeriodePaie.query.filter_by(id=periode_id, tenant_id=t.id).first_or_404()
    bulletins = (BulletinPaie.query
                 .filter(BulletinPaie.periode_id==periode_id, BulletinPaie.tenant_id==t.id, BulletinPaie.statut.in_(["VALIDE","VALIDÉ"]))
                 .join(Salarie)
                 .order_by(Salarie.nom)
                 .all())

    if not bulletins:
        flash("Aucun bulletin validé pour cette période.", "warning")
        return redirect(url_for("bulletins"))

    try:
        from export_comptable import generer_journal_paie, generer_livre_paie
        journal = generer_journal_paie(bulletins, periode, t)
        livre   = generer_livre_paie(bulletins, periode, t)

        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
            zf.writestr(
                f"journal_paie_{periode.mois:02d}{periode.annee}_{t.sigle or t.id}.txt",
                journal
            )
            zf.writestr(
                f"livre_paie_{periode.mois:02d}{periode.annee}_{t.sigle or t.id}.csv",
                livre
            )
        zip_buffer.seek(0)

        nom_zip = f"export_sage_{periode.mois:02d}{periode.annee}_{t.sigle or t.id}.zip"
        logger.info(f"[Export Sage] ZIP généré — tenant={t.id}")
        return send_file(
            zip_buffer,
            mimetype="application/zip",
            as_attachment=True,
            download_name=nom_zip,
        )
    except Exception as e:
        logger.error(f"[Export Sage] Erreur ZIP : {e}")
        flash(f"Erreur lors de la génération : {e}", "error")
        return redirect(url_for("bulletins"))


@app.route("/declaration-cnss/export-excel")
@login_required
def declaration_cnss_excel():
    """Export Excel déclarations : mensuel (CFP/FNH/TCS/IRPP) ou trimestriel (CNSS/CNAMGS)."""
    if current_user.is_super_admin: return redirect(url_for("admin_dashboard"))
    t = get_tenant()
    if not t: return redirect(url_for("login"))

    pid     = request.args.get("periode_id", type=int)
    mode    = request.args.get("mode", "mensuel")
    periode = PeriodePaie.query.filter_by(id=pid, tenant_id=t.id).first_or_404()

    MOIS_FR2 = ["","Janvier","Février","Mars","Avril","Mai","Juin",
                "Juillet","Août","Septembre","Octobre","Novembre","Décembre"]

    if mode == "trimestriel":
        # ── Trimestre : récupérer les 3 mois ─────────────────────────────────
        mois=periode.mois; trim_num=(mois-1)//3+1
        trim_debut=((mois-1)//3)*3+1; trim_fin=trim_debut+2
        trim_label=f"T {trim_num}"
        mois_labels=[MOIS_FR2[m] for m in range(trim_debut, trim_fin+1)]

        periodes_trim=PeriodePaie.query.filter_by(tenant_id=t.id, annee=periode.annee)            .filter(PeriodePaie.mois>=trim_debut, PeriodePaie.mois<=trim_fin).all()
        mois_map={p.mois: p.id for p in periodes_trim}

        buls_trim=BulletinPaie.query.filter(
            BulletinPaie.tenant_id==t.id,
            BulletinPaie.periode_id.in_([p.id for p in periodes_trim])
        ).options(joinedload(BulletinPaie.salarie), joinedload(BulletinPaie.periode)).all()

        # Regrouper par salarié
        from collections import defaultdict
        sal_map = defaultdict(lambda: {
            "nom_complet":"","matricule":"","numero_cnss":"","numero_cnamgs":"",
            "date_embauche":"",
            "m1_base_cnss":0,"m2_base_cnss":0,"m3_base_cnss":0,
            "m1_base_cnamgs":0,"m2_base_cnamgs":0,"m3_base_cnamgs":0,
        })
        for b in buls_trim:
            k=b.salarie_id
            sal=b.salarie
            sal_map[k]["nom_complet"]   = sal.nom_complet
            sal_map[k]["matricule"]     = sal.matricule or ""
            sal_map[k]["numero_cnss"]   = sal.numero_cnss or ""
            sal_map[k]["numero_cnamgs"] = sal.numero_cnamgs or ""
            if sal.date_embauche:
                sal_map[k]["date_embauche"] = sal.date_embauche.strftime("%d/%m/%Y")
            m = b.periode.mois if b.periode else 0
            if m == trim_debut:
                sal_map[k]["m1_base_cnss"]   = float(b.base_cnss   or 0)
                sal_map[k]["m1_base_cnamgs"] = float(b.base_cnamgs or 0)
            elif m == trim_debut+1:
                sal_map[k]["m2_base_cnss"]   = float(b.base_cnss   or 0)
                sal_map[k]["m2_base_cnamgs"] = float(b.base_cnamgs or 0)
            elif m == trim_fin:
                sal_map[k]["m3_base_cnss"]   = float(b.base_cnss   or 0)
                sal_map[k]["m3_base_cnamgs"] = float(b.base_cnamgs or 0)

        sal_data = sorted(sal_map.values(), key=lambda x: x["nom_complet"])

        tot_cnss_m  = [sum(s["m1_base_cnss"]   for s in sal_data),
                       sum(s["m2_base_cnss"]   for s in sal_data),
                       sum(s["m3_base_cnss"]   for s in sal_data)]
        tot_cnamgs_m= [sum(s["m1_base_cnamgs"] for s in sal_data),
                       sum(s["m2_base_cnamgs"] for s in sal_data),
                       sum(s["m3_base_cnamgs"] for s in sal_data)]
        total_base_cnss   = sum(tot_cnss_m)
        total_base_cnamgs = sum(tot_cnamgs_m)

        # Générer les deux fichiers et les zip
        import zipfile, io
        cnss_bytes   = _gen_excel_cnss(t, trim_label, periode.annee, mois_labels,
                                       sal_data, total_base_cnss, total_base_cnamgs, tot_cnss_m)
        cnamgs_bytes = _gen_excel_cnamgs(t, trim_label, periode.annee, mois_labels,
                                         sal_data, total_base_cnamgs, tot_cnamgs_m)

        zip_buf = io.BytesIO()
        nom_base = t.denomination.replace(" ","_")[:20]
        with zipfile.ZipFile(zip_buf, 'w', zipfile.ZIP_DEFLATED) as zf:
            zf.writestr(f"CNSS_{nom_base}_{trim_label.replace(' ','')}_{periode.annee}.xlsx",   cnss_bytes)
            zf.writestr(f"CNAMGS_{nom_base}_{trim_label.replace(' ','')}_{periode.annee}.xlsx", cnamgs_bytes)
        zip_buf.seek(0)

        from flask import Response
        nom_zip = f"declarations_trimestrielles_{nom_base}_{trim_label.replace(' ','')}_{periode.annee}.zip"
        return Response(zip_buf.read(), mimetype="application/zip",
            headers={"Content-Disposition": f'attachment; filename="{nom_zip}"'})


# ══════════════════════════════════════════════════════════════════════════════
# DÉCLARATION CNSS/CNAMGS — Export CSV portail électronique
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/declaration-cnss/export-csv")
@login_required
def declaration_cnss_csv():
    """
    Export CSV uploadable directement sur le portail CNSS Gabon (cnss.ga)
    et CNAMGS. Génère une archive ZIP avec les deux fichiers CSV.
    """
    if current_user.is_super_admin: return redirect(url_for("admin_dashboard"))
    t = get_tenant()
    if not t: return redirect(url_for("login"))

    pid     = request.args.get("periode_id", type=int)
    periode = PeriodePaie.query.filter_by(id=pid, tenant_id=t.id).first_or_404()

    # ── Calculer le trimestre ──────────────────────────────────────────────
    from declaration_cnss import calculer_trimestre, generer_csv_cnss, generer_csv_cnamgs
    trim_num, trim_debut, trim_fin, trim_label = calculer_trimestre(periode.mois)

    periodes_trim = PeriodePaie.query.filter_by(
        tenant_id=t.id, annee=periode.annee
    ).filter(
        PeriodePaie.mois >= trim_debut,
        PeriodePaie.mois <= trim_fin
    ).all()

    if not periodes_trim:
        flash("Aucune période trouvée pour ce trimestre.", "warning")
        return redirect(url_for("declaration_cnss", periode_id=pid, mode="trimestriel"))

    buls_trim = BulletinPaie.query.filter(
        BulletinPaie.tenant_id == t.id,
        BulletinPaie.periode_id.in_([p.id for p in periodes_trim])
    ).options(
        joinedload(BulletinPaie.salarie),
        joinedload(BulletinPaie.periode)
    ).all()

    if not buls_trim:
        flash("Aucun bulletin pour ce trimestre. Saisissez et validez les bulletins d'abord.", "warning")
        return redirect(url_for("declaration_cnss", periode_id=pid, mode="trimestriel"))

    # ── Regrouper par salarié ──────────────────────────────────────────────
    from collections import defaultdict
    sal_map = defaultdict(lambda: {
        "nom_complet": "", "matricule": "", "numero_cnss": "",
        "numero_cnamgs": "", "date_embauche": "",
        "m1_base_cnss": 0, "m2_base_cnss": 0, "m3_base_cnss": 0,
        "m1_base_cnamgs": 0, "m2_base_cnamgs": 0, "m3_base_cnamgs": 0,
    })

    for b in buls_trim:
        k   = b.salarie_id
        sal = b.salarie
        sal_map[k]["nom_complet"]   = sal.nom_complet
        sal_map[k]["matricule"]     = sal.matricule or ""
        sal_map[k]["numero_cnss"]   = sal.numero_cnss or ""
        sal_map[k]["numero_cnamgs"] = sal.numero_cnamgs or ""
        if sal.date_embauche:
            sal_map[k]["date_embauche"] = sal.date_embauche.strftime("%d/%m/%Y")
        m = b.periode.mois if b.periode else 0
        if m == trim_debut:
            sal_map[k]["m1_base_cnss"]   = float(b.base_cnss   or 0)
            sal_map[k]["m1_base_cnamgs"] = float(b.base_cnamgs or 0)
        elif m == trim_debut + 1:
            sal_map[k]["m2_base_cnss"]   = float(b.base_cnss   or 0)
            sal_map[k]["m2_base_cnamgs"] = float(b.base_cnamgs or 0)
        elif m == trim_fin:
            sal_map[k]["m3_base_cnss"]   = float(b.base_cnss   or 0)
            sal_map[k]["m3_base_cnamgs"] = float(b.base_cnamgs or 0)

    sal_data = list(sal_map.values())

    # ── Générer les deux CSV ───────────────────────────────────────────────
    try:
        csv_cnss   = generer_csv_cnss(sal_data, periode, t, trim_debut, trim_fin)
        csv_cnamgs = generer_csv_cnamgs(sal_data, periode, t, trim_debut, trim_fin)

        import zipfile
        zip_buf  = io.BytesIO()
        nom_base = (t.sigle or t.denomination[:15]).replace(" ", "_")
        trim_str = f"T{trim_num}_{periode.annee}"

        with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
            zf.writestr(f"CNSS_{nom_base}_{trim_str}.csv",   csv_cnss)
            zf.writestr(f"CNAMGS_{nom_base}_{trim_str}.csv", csv_cnamgs)
            # Ajouter un fichier README avec les instructions d'upload
            zf.writestr(
                "INSTRUCTIONS_UPLOAD.txt",
                _instructions_upload(t, trim_label, periode.annee, len(sal_data))
            )
        zip_buf.seek(0)

        nom_zip = f"declarations_CNSS_CNAMGS_{nom_base}_{trim_str}.zip"
        logger.info(f"[CNSS CSV] Export {trim_str} — {len(sal_data)} salariés — tenant={t.id}")

        return send_file(
            zip_buf,
            mimetype="application/zip",
            as_attachment=True,
            download_name=nom_zip,
        )

    except Exception as e:
        logger.error(f"[CNSS CSV] Erreur export : {e}")
        flash(f"Erreur lors de la génération : {e}", "error")
        return redirect(url_for("declaration_cnss", periode_id=pid, mode="trimestriel"))


def _instructions_upload(tenant, trim_label, annee, nb_salaries) -> str:
    """Génère un fichier texte d'instructions pour l'upload sur les portails."""
    return f"""
INSTRUCTIONS D'UPLOAD — DÉCLARATIONS TRIMESTRIELLES
====================================================
Entreprise : {tenant.denomination}
NIF        : {tenant.nif or "—"}
Trimestre  : {trim_label} {annee}
Salariés   : {nb_salaries}
Généré le  : {datetime.now().strftime("%d/%m/%Y à %H:%M")}


FICHIER CNSS : CNSS_*.csv
─────────────────────────
1. Connectez-vous sur https://cnss.ga
2. Allez dans : Mon Espace → Déclarations → Nouvelle déclaration
3. Choisissez : Déclaration trimestrielle de salaires
4. Cliquez sur "Importer un fichier"
5. Sélectionnez le fichier CNSS_*.csv
6. Vérifiez les montants affichés
7. Validez et téléchargez le reçu


FICHIER CNAMGS : CNAMGS_*.csv
──────────────────────────────
1. Connectez-vous sur le portail CNAMGS
2. Allez dans : Déclarations → Déclaration trimestrielle
3. Importez le fichier CNAMGS_*.csv
4. Vérifiez et validez


MONTANTS À VERSER (rappel) :
────────────────────────────
CNSS  : cotisations salariales (5%) + patronales (18%) = 23% de la base
CNAMGS: cotisations salariales (1,5%) + patronales (6%) = 7,5% de la base

Date limite de dépôt : dernier jour du mois suivant la fin du trimestre
  T1 (Jan-Mar) → 30 Avril
  T2 (Avr-Jun) → 31 Juillet
  T3 (Jul-Sep) → 31 Octobre
  T4 (Oct-Déc) → 31 Janvier


IMPORTANT :
──────────
- Utilisez le fichier CSV, pas le fichier Excel, pour l'upload portail
- Le fichier Excel (généré séparément) est pour vos archives papier
- Gardez le reçu de dépôt comme justificatif

En cas de problème : support@paiegalon.com
""".strip()


# ══════════════════════════════════════════════════════════════════════════════
# API REST v1 — Intégration grandes entreprises
# ══════════════════════════════════════════════════════════════════════════════
# Authentification : X-API-Key: <token>  ou  Authorization: Bearer <oauth_token>
# Tous les endpoints retournent JSON. Préfixe : /api/v1/
# ══════════════════════════════════════════════════════════════════════════════

from api_rest import (api_auth_required, _ok, _err, _paginate,
                      _salarie_dict, _bulletin_dict, _periode_dict,
                      _oauth_tokens, OAUTH_TOKEN_TTL)


# ── OAuth2 : obtenir un token ──────────────────────────────────────────────────

@app.route("/api/v1/oauth/token", methods=["POST"])
@csrf.exempt
def api_oauth_token():
    """
    OAuth2 client_credentials flow.
    Body JSON : { "client_id": "...", "client_secret": "...", "grant_type": "client_credentials" }
    """
    data       = request.get_json(force=True) or {}
    grant_type = data.get("grant_type", "")
    client_id  = data.get("client_id",  "").strip()
    client_sec = data.get("client_secret", "").strip()

    if grant_type != "client_credentials":
        return jsonify(_err("UNSUPPORTED_GRANT", "Seul client_credentials est supporté.")), 400
    if not client_id or not client_sec:
        return jsonify(_err("MISSING_CREDENTIALS", "client_id et client_secret requis.")), 400

    client = OAuthClient.query.filter_by(client_id=client_id, actif=True).first()
    if not client or not hmac.compare_digest(client.client_secret, client_sec):
        return jsonify(_err("INVALID_CLIENT", "Identifiants OAuth invalides.")), 401

    tenant = Tenant.query.filter_by(id=client.tenant_id, statut="ACTIF").first()
    if not tenant:
        return jsonify(_err("TENANT_INACTIVE", "Compte inactif ou suspendu.")), 403

    # Générer un access token
    import secrets as _sec
    access_token = _sec.token_hex(32)
    expires_at   = datetime.utcnow() + timedelta(seconds=OAUTH_TOKEN_TTL)
    _oauth_tokens[access_token] = {
        "tenant_id":  tenant.id,
        "expires_at": expires_at,
        "client_id":  client_id,
    }
    client.derniere_utilisation = datetime.utcnow()
    db.session.commit()

    logger.info(f"[API OAuth] Token émis — client={client_id} tenant={tenant.id}")
    return jsonify({
        "access_token": access_token,
        "token_type":   "Bearer",
        "expires_in":   OAUTH_TOKEN_TTL,
        "scope":        "read write",
    }), 200


@app.route("/api/v1/oauth/revoke", methods=["POST"])
@csrf.exempt
def api_oauth_revoke():
    """Révoque un access token OAuth2."""
    data  = request.get_json(force=True) or {}
    token = data.get("token", "").strip()
    if token in _oauth_tokens:
        del _oauth_tokens[token]
    return jsonify({"success": True, "message": "Token révoqué."}), 200


# ── GET /api/v1/me ─────────────────────────────────────────────────────────────

@app.route("/api/v1/me")
@csrf.exempt
@api_auth_required
def api_me(tenant):
    """Infos sur le tenant authentifié."""
    return _ok({
        "id":            tenant.id,
        "denomination":  tenant.denomination,
        "sigle":         tenant.sigle,
        "nif":           tenant.nif,
        "statut":        tenant.statut,
        "plan":          tenant.plan.nom if tenant.plan else None,
        "date_expiration": str(tenant.date_expiration) if tenant.date_expiration else None,
        "nb_salaries_actifs": tenant.nb_salaries_actifs,
    })


# ── GET /api/v1/salaries ───────────────────────────────────────────────────────

@app.route("/api/v1/salaries")
@csrf.exempt
@api_auth_required
def api_salaries_list(tenant):
    """
    Liste des salariés.
    Filtres : ?statut=ACTIF&categorie_id=1&q=dupont
    Pagination : ?page=1&per_page=25
    """
    q = Salarie.query.filter_by(tenant_id=tenant.id)

    statut = request.args.get("statut")
    if statut:
        q = q.filter(Salarie.statut == statut.upper())

    cat_id = request.args.get("categorie_id", type=int)
    if cat_id:
        q = q.filter(Salarie.categorie_id == cat_id)

    search = request.args.get("q", "").strip()
    if search:
        like = f"%{search}%"
        q = q.filter(
            db.or_(Salarie.nom.ilike(like), Salarie.prenom.ilike(like),
                   Salarie.matricule.ilike(like))
        )

    q = q.order_by(Salarie.nom, Salarie.prenom)
    result, meta = _paginate(q, request)
    return _ok([_salarie_dict(s) for s in result.items], meta)


# ── GET /api/v1/salaries/<id> ──────────────────────────────────────────────────

@app.route("/api/v1/salaries/<int:sal_id>")
@csrf.exempt
@api_auth_required
def api_salarie_detail(tenant, sal_id):
    """Détail complet d'un salarié."""
    s = Salarie.query.filter_by(id=sal_id, tenant_id=tenant.id).first()
    if not s:
        return jsonify(_err("NOT_FOUND", "Salarié introuvable.")), 404
    return _ok(_salarie_dict(s, detail=True))


# ── POST /api/v1/salaries ──────────────────────────────────────────────────────

@app.route("/api/v1/salaries", methods=["POST"])
@csrf.exempt
@api_auth_required
def api_salarie_create(tenant):
    """
    Crée un nouveau salarié.
    Body JSON : { "matricule", "nom", "prenom", "emploi", "date_embauche", ... }
    """
    data = request.get_json(force=True) or {}

    # Champs obligatoires
    for champ in ["matricule", "nom", "prenom", "date_embauche"]:
        if not data.get(champ, "").strip():
            return jsonify(_err("MISSING_FIELD", f"Champ obligatoire manquant : {champ}")), 400

    # Vérifier doublon matricule
    if Salarie.query.filter_by(tenant_id=tenant.id, matricule=data["matricule"].strip()).first():
        return jsonify(_err("DUPLICATE_MATRICULE", f"Le matricule {data['matricule']} existe déjà.")), 409

    try:
        from datetime import date as _date
        date_emb = datetime.strptime(data["date_embauche"], "%Y-%m-%d").date()
    except ValueError:
        return jsonify(_err("INVALID_DATE", "Format date_embauche invalide. Utilisez YYYY-MM-DD.")), 400

    s = Salarie(
        tenant_id      = tenant.id,
        matricule      = data["matricule"].strip().upper(),
        nom            = data["nom"].strip().upper(),
        prenom         = data["prenom"].strip(),
        emploi         = data.get("emploi", "").strip(),
        email          = data.get("email", "").strip(),
        telephone      = data.get("telephone", "").strip(),
        sexe           = data.get("sexe", "").strip().upper(),
        nationalite    = data.get("nationalite", "GABONAISE").strip().upper(),
        date_embauche  = date_emb,
        numero_cnss    = data.get("numero_cnss", "").strip(),
        numero_cnamgs  = data.get("numero_cnamgs", "").strip(),
        nombre_parts   = float(data.get("nombre_parts", 1.0)),
        nb_enfants     = int(data.get("nb_enfants", 0)),
        statut         = "ACTIF",
    )
    cat_id = data.get("categorie_id")
    if cat_id:
        cat = CategorieEmploi.query.filter_by(id=cat_id, tenant_id=tenant.id).first()
        if cat:
            s.categorie_id = cat.id

    db.session.add(s)
    db.session.commit()
    _cache_delete(f"{tenant.id}:")
    logger.info(f"[API] Salarié créé — matricule={s.matricule} tenant={tenant.id}")
    return _ok(_salarie_dict(s, detail=True), status=201)


# ── PUT /api/v1/salaries/<id> ──────────────────────────────────────────────────

@app.route("/api/v1/salaries/<int:sal_id>", methods=["PUT", "PATCH"])
@csrf.exempt
@api_auth_required
def api_salarie_update(tenant, sal_id):
    """Met à jour les informations d'un salarié."""
    s = Salarie.query.filter_by(id=sal_id, tenant_id=tenant.id).first()
    if not s:
        return jsonify(_err("NOT_FOUND", "Salarié introuvable.")), 404

    data = request.get_json(force=True) or {}
    champs_modifiables = [
        "nom", "prenom", "emploi", "email", "telephone", "sexe",
        "nationalite", "numero_cnss", "numero_cnamgs",
        "nombre_parts", "nb_enfants", "statut",
        "situation_matrimoniale",
    ]
    for champ in champs_modifiables:
        if champ in data:
            val = data[champ]
            if isinstance(val, str):
                val = val.strip()
                if champ in ("nom", "nationalite"):
                    val = val.upper()
            setattr(s, champ, val)

    db.session.commit()
    logger.info(f"[API] Salarié modifié — id={sal_id} tenant={tenant.id}")
    return _ok(_salarie_dict(s, detail=True))


# ── GET /api/v1/periodes ───────────────────────────────────────────────────────

@app.route("/api/v1/periodes")
@csrf.exempt
@api_auth_required
def api_periodes_list(tenant):
    """Liste des périodes de paie, tri décroissant."""
    q = PeriodePaie.query.filter_by(tenant_id=tenant.id)\
        .order_by(PeriodePaie.annee.desc(), PeriodePaie.mois.desc())

    statut = request.args.get("statut")
    if statut:
        q = q.filter(PeriodePaie.statut == statut.upper())

    result, meta = _paginate(q, request, default_per_page=12)
    return _ok([_periode_dict(p) for p in result.items], meta)


# ── POST /api/v1/periodes ──────────────────────────────────────────────────────

@app.route("/api/v1/periodes", methods=["POST"])
@csrf.exempt
@api_auth_required
def api_periode_create(tenant):
    """Crée une période de paie. Body : { "annee": 2026, "mois": 7 }"""
    data  = request.get_json(force=True) or {}
    annee = data.get("annee")
    mois  = data.get("mois")

    if not annee or not mois:
        return jsonify(_err("MISSING_FIELD", "annee et mois sont obligatoires.")), 400

    annee, mois = int(annee), int(mois)
    if not (1 <= mois <= 12) or not (2000 <= annee <= 2100):
        return jsonify(_err("INVALID_DATE", "Mois (1-12) ou année invalide.")), 400

    if PeriodePaie.query.filter_by(tenant_id=tenant.id, annee=annee, mois=mois).first():
        return jsonify(_err("DUPLICATE", f"La période {mois:02d}/{annee} existe déjà.")), 409

    MOIS_FR = ["","Janvier","Février","Mars","Avril","Mai","Juin",
               "Juillet","Août","Septembre","Octobre","Novembre","Décembre"]
    p = PeriodePaie(
        tenant_id    = tenant.id,
        annee        = annee,
        mois         = mois,
        libelle_mois = MOIS_FR[mois],
        statut       = "OUVERT",
    )
    db.session.add(p)
    db.session.commit()
    return _ok(_periode_dict(p), status=201)


# ── GET /api/v1/bulletins ──────────────────────────────────────────────────────

@app.route("/api/v1/bulletins")
@csrf.exempt
@api_auth_required
def api_bulletins_list(tenant):
    """
    Liste des bulletins de paie.
    Filtres : ?periode_id=1&salarie_id=5&statut=VALIDÉ&annee=2026&mois=6
    """
    q = BulletinPaie.query.filter_by(tenant_id=tenant.id)\
        .join(Salarie).join(PeriodePaie)

    if request.args.get("periode_id"):
        q = q.filter(BulletinPaie.periode_id == request.args.get("periode_id", type=int))
    if request.args.get("salarie_id"):
        q = q.filter(BulletinPaie.salarie_id == request.args.get("salarie_id", type=int))
    if request.args.get("statut"):
        q = q.filter(BulletinPaie.statut == request.args["statut"])
    if request.args.get("annee"):
        q = q.filter(PeriodePaie.annee == request.args.get("annee", type=int))
    if request.args.get("mois"):
        q = q.filter(PeriodePaie.mois == request.args.get("mois", type=int))

    q = q.order_by(PeriodePaie.annee.desc(), PeriodePaie.mois.desc(), Salarie.nom)
    result, meta = _paginate(q, request)
    return _ok([_bulletin_dict(b) for b in result.items], meta)


# ── GET /api/v1/bulletins/<id> ─────────────────────────────────────────────────

@app.route("/api/v1/bulletins/<int:bul_id>")
@csrf.exempt
@api_auth_required
def api_bulletin_detail(tenant, bul_id):
    """Détail complet d'un bulletin."""
    b = BulletinPaie.query.filter_by(id=bul_id, tenant_id=tenant.id)\
        .options(joinedload(BulletinPaie.salarie),
                 joinedload(BulletinPaie.periode)).first()
    if not b:
        return jsonify(_err("NOT_FOUND", "Bulletin introuvable.")), 404
    return _ok(_bulletin_dict(b, detail=True))


# ── POST /api/v1/bulletins/calculer ───────────────────────────────────────────

@app.route("/api/v1/bulletins/calculer", methods=["POST"])
@csrf.exempt
@api_auth_required
def api_bulletin_calculer(tenant):
    """
    Calcule et crée un bulletin de paie.
    Body : { "salarie_id": 1, "periode_id": 2, "salaire_base": 400000, ... }
    """
    data = request.get_json(force=True) or {}

    sal_id = data.get("salarie_id")
    per_id = data.get("periode_id")
    if not sal_id or not per_id:
        return jsonify(_err("MISSING_FIELD", "salarie_id et periode_id sont obligatoires.")), 400

    s = Salarie.query.filter_by(id=sal_id, tenant_id=tenant.id).first()
    if not s:
        return jsonify(_err("NOT_FOUND", "Salarié introuvable.")), 404

    p = PeriodePaie.query.filter_by(id=per_id, tenant_id=tenant.id).first()
    if not p:
        return jsonify(_err("NOT_FOUND", "Période introuvable.")), 404

    # Vérifier doublon
    if BulletinPaie.query.filter_by(
        tenant_id=tenant.id, salarie_id=sal_id, periode_id=per_id
    ).first():
        return jsonify(_err("DUPLICATE", "Un bulletin existe déjà pour ce salarié et cette période.")), 409

    # Lancer le calcul
    from calculs_paie import calculer_bulletin
    donnees = {
        "salaire_base": float(data.get("salaire_base", 0)),
        "prime_caisse": float(data.get("prime_caisse", 0)),
        "prime_transport": float(data.get("prime_transport", 0)),
        "indem_logement": float(data.get("indem_logement", 0)),
        "heures_sup_10": float(data.get("heures_sup_10", 0)),
        "heures_sup_30": float(data.get("heures_sup_30", 0)),
        "acompte": float(data.get("acompte", 0)),
        "absences": float(data.get("absences", 0)),
        "prime_anciennete": float(data.get("prime_anciennete", 0)),
        "sursalaire": float(data.get("sursalaire", 0)),
    }
    try:
        res = calculer_bulletin(donnees, nb_parts=float(s.nombre_parts or 1))
    except Exception as e:
        return jsonify(_err("CALCUL_ERROR", f"Erreur de calcul : {e}")), 500

    # Créer le bulletin
    b = BulletinPaie(tenant_id=tenant.id, salarie_id=s.id, periode_id=p.id, statut="BROUILLON")
    for k, v in {**donnees, **res}.items():
        if hasattr(b, k):
            setattr(b, k, v)
    db.session.add(b)
    db.session.commit()
    _cache_delete(f"{tenant.id}:")
    logger.info(f"[API] Bulletin créé — sal={sal_id} periode={per_id} tenant={tenant.id}")
    return _ok(_bulletin_dict(b, detail=True), status=201)


# ── POST /api/v1/bulletins/<id>/valider ───────────────────────────────────────

@app.route("/api/v1/bulletins/<int:bul_id>/valider", methods=["POST"])
@csrf.exempt
@api_auth_required
def api_bulletin_valider(tenant, bul_id):
    """Valide un bulletin en brouillon."""
    b = BulletinPaie.query.filter_by(id=bul_id, tenant_id=tenant.id).first()
    if not b:
        return jsonify(_err("NOT_FOUND", "Bulletin introuvable.")), 404
    if b.statut != "BROUILLON":
        return jsonify(_err("INVALID_STATE", f"Le bulletin est déjà en statut '{b.statut}'.")), 400

    b.statut          = "VALIDÉ"
    b.date_validation = datetime.utcnow()
    db.session.commit()
    return _ok(_bulletin_dict(b, detail=True))


# ── GET /api/v1/stats ──────────────────────────────────────────────────────────

@app.route("/api/v1/stats")
@csrf.exempt
@api_auth_required
def api_stats(tenant):
    """
    Statistiques de paie du tenant.
    Optionnel : ?periode_id=X pour les stats d'une période précise.
    """
    stats = {
        "nb_salaries_actifs":   tenant.nb_salaries_actifs,
        "nb_salaries_total":    Salarie.query.filter_by(tenant_id=tenant.id).count(),
        "nb_periodes":          PeriodePaie.query.filter_by(tenant_id=tenant.id).count(),
        "nb_bulletins_total":   BulletinPaie.query.filter_by(tenant_id=tenant.id).count(),
        "nb_bulletins_valides": BulletinPaie.query.filter_by(
            tenant_id=tenant.id, statut="VALIDÉ").count(),
    }

    # Stats d'une période précise
    per_id = request.args.get("periode_id", type=int)
    if per_id:
        buls = BulletinPaie.query.filter_by(
            tenant_id=tenant.id, periode_id=per_id).all()
        stats["periode"] = {
            "nb_bulletins":       len(buls),
            "masse_salariale_brute": sum(float(b.salaire_brut or 0) for b in buls),
            "total_net_a_payer":     sum(float(b.net_a_payer or 0) for b in buls),
            "total_cnss_sal":        sum(float(b.cnss_salarie or 0) for b in buls),
            "total_cnss_pat":        sum(float(b.cnss_patronale or 0) for b in buls),
            "total_irpp":            sum(float(b.irpp or 0) for b in buls),
        }

    return _ok(stats)


# ── Gestion des clients OAuth depuis l'interface ───────────────────────────────

@app.route("/api/clients")
@tenant_required
def api_clients_list():
    """Page de gestion des clients OAuth du tenant."""
    t = get_tenant()
    clients = OAuthClient.query.filter_by(tenant_id=t.id).order_by(
        OAuthClient.date_creation.desc()).all()
    return render_template("tenant/api_clients.html", tenant=t, clients=clients)


@app.route("/api/clients/creer", methods=["POST"])
@tenant_required
def api_client_creer():
    """Crée un nouveau client OAuth2 pour ce tenant."""
    t   = get_tenant()
    nom = request.form.get("nom", "").strip()
    if not nom:
        flash("Le nom du client est obligatoire.", "error")
        return redirect(url_for("api_clients_list"))

    import secrets as _sec
    client = OAuthClient(
        tenant_id     = t.id,
        nom           = nom,
        client_id     = f"pg_{_sec.token_hex(16)}",
        client_secret = _sec.token_hex(32),
        description   = request.form.get("description", "").strip(),
        actif         = True,
    )
    db.session.add(client)
    db.session.commit()
    flash(f"Client OAuth '{nom}' créé. Conservez le secret — il ne sera plus affiché.", "success")
    return redirect(url_for("api_clients_list"))


@app.route("/api/clients/<int:client_id>/supprimer", methods=["POST"])
@tenant_required
def api_client_supprimer(client_id):
    """Désactive un client OAuth."""
    t      = get_tenant()
    client = OAuthClient.query.filter_by(id=client_id, tenant_id=t.id).first_or_404()
    client.actif = False
    db.session.commit()
    flash(f"Client '{client.nom}' désactivé.", "success")
    return redirect(url_for("api_clients_list"))


@app.route("/recherche")
@login_required
def recherche_globale():
    """Recherche globale : salariés, journaliers, bulletins, acomptes."""
    if current_user.is_super_admin: return redirect(url_for("admin_dashboard"))
    t = get_tenant()
    if not t: return redirect(url_for("login"))

    q = request.args.get("q", "").strip()
    if not q:
        return render_template("tenant/recherche.html",
            tenant=t, q="", resultats={}, nb_total=0)

    like = f"%{q}%"
    resultats = {}

    # ── Salariés ────────────────────────────────────────────────────────────
    sals = Salarie.query.filter_by(tenant_id=t.id).filter(
        db.or_(
            Salarie.nom.ilike(like), Salarie.prenom.ilike(like),
            Salarie.matricule.ilike(like), Salarie.emploi.ilike(like),
            Salarie.telephone.ilike(like),
        )
    ).order_by(Salarie.nom).limit(10).all()
    if sals:
        resultats["salaries"] = [{"id": s.id, "titre": s.nom_complet,
            "sous_titre": f"{s.emploi or '—'} · {s.matricule}",
            "badge": s.statut, "lien": f"/salaries/{s.id}",
            "icone": "👤"} for s in sals]

    # ── Journaliers ──────────────────────────────────────────────────────────
    jours = Journalier.query.filter_by(tenant_id=t.id).filter(
        db.or_(
            Journalier.nom.ilike(like), Journalier.prenom.ilike(like),
            Journalier.profession.ilike(like), Journalier.telephone.ilike(like),
        )
    ).order_by(Journalier.nom).limit(10).all()
    if jours:
        resultats["journaliers"] = [{"id": j.id, "titre": j.nom_complet,
            "sous_titre": f"{j.profession or '—'} · {int(j.taux_horaire or 0)} FCFA/h",
            "badge": j.statut, "lien": f"/journaliers/{j.id}",
            "icone": "🦺"} for j in jours]

    # ── Bulletins ────────────────────────────────────────────────────────────
    buls = BulletinPaie.query.filter_by(tenant_id=t.id)        .join(Salarie, BulletinPaie.salarie_id == Salarie.id)        .join(PeriodePaie, BulletinPaie.periode_id == PeriodePaie.id)        .filter(db.or_(
            Salarie.nom.ilike(like), Salarie.prenom.ilike(like),
            Salarie.matricule.ilike(like),
        ))        .order_by(BulletinPaie.date_creation.desc()).limit(10).all()
    if buls:
        resultats["bulletins"] = [{"id": b.id,
            "titre": b.salarie.nom_complet,
            "sous_titre": f"{b.periode.libelle_complet} · Net : {int(b.net_a_payer or 0):,} FCFA",
            "badge": b.statut, "lien": f"/bulletins/{b.id}",
            "icone": "📄"} for b in buls]

    # ── Acomptes ─────────────────────────────────────────────────────────────
    acomps = Acompte.query.filter_by(tenant_id=t.id)        .join(Salarie, Acompte.salarie_id == Salarie.id)        .filter(db.or_(
            Salarie.nom.ilike(like), Salarie.prenom.ilike(like),
            Salarie.matricule.ilike(like),
        ))        .order_by(Acompte.date_acompte.desc()).limit(10).all()
    if acomps:
        resultats["acomptes"] = [{"id": a.id,
            "titre": a.salarie.nom_complet,
            "sous_titre": f"{int(a.montant or 0):,} FCFA · {a.date_acompte.strftime('%d/%m/%Y') if a.date_acompte else ''}",
            "badge": a.statut, "lien": "/acomptes",
            "icone": "💸"} for a in acomps]

    nb_total = sum(len(v) for v in resultats.values())
    return render_template("tenant/recherche.html",
        tenant=t, q=q, resultats=resultats, nb_total=nb_total)


@app.route("/api/recherche-rapide")
@login_required
def api_recherche_rapide():
    """API JSON pour l'autocomplétion dans la barre de recherche."""
    t = get_tenant()
    if not t: return jsonify([])
    q = request.args.get("q", "").strip()
    if len(q) < 2: return jsonify([])

    like = f"%{q}%"
    resultats = []

    # Salariés (5 max)
    for s in Salarie.query.filter_by(tenant_id=t.id).filter(
        db.or_(Salarie.nom.ilike(like), Salarie.prenom.ilike(like),
               Salarie.matricule.ilike(like))
    ).limit(5).all():
        resultats.append({"icone":"👤","titre":s.nom_complet,
            "sous_titre":s.emploi or "Salarié","lien":f"/salaries/{s.id}",
            "categorie":"Salariés"})

    # Journaliers (5 max)
    for j in Journalier.query.filter_by(tenant_id=t.id).filter(
        db.or_(Journalier.nom.ilike(like), Journalier.prenom.ilike(like),
               Journalier.profession.ilike(like))
    ).limit(5).all():
        resultats.append({"icone":"🦺","titre":j.nom_complet,
            "sous_titre":j.profession or "Journalier","lien":f"/journaliers/{j.id}",
            "categorie":"Journaliers"})

    return jsonify(resultats[:10])

# ── Helpers ───────────────────────────────────────────────────────────────────
def _pd(v):
    if not v: return None
    try: return datetime.strptime(v,"%Y-%m-%d").date()
    except: return None

@app.template_filter("fcfa")
def fcfa_filter(v):
    try: return f"{int(float(v)):,}".replace(","," ")+" FCFA"
    except: return "— FCFA"

@app.template_filter("date_fr")
def date_fr_filter(v):
    if not v: return "—"
    if isinstance(v, str):
        try: v = datetime.strptime(v[:10], "%Y-%m-%d").date()
        except: return v
    return v.strftime("%d/%m/%Y")

@app.template_filter("datetime_fr")
def datetime_fr_filter(v):
    """Affiche date ET heure : 28/05/2026 à 17h34"""
    if not v: return "Jamais"
    if isinstance(v, str):
        try: v = datetime.fromisoformat(v)
        except: return v
    return v.strftime("%d/%m/%Y à %Hh%M")

@app.context_processor
def inject_globals(): return {"now":datetime.now(),"enumerate":enumerate}

@app.errorhandler(403)
def forbidden(e): return render_template("auth/403.html"),403

# ── Init DB ───────────────────────────────────────────────────────────────────
def init_db():
    db.create_all()  # Crée toutes les tables dont paiements (nouveau)
    if not Plan.query.first():
        for code,nom,prix,ms,mu,desc in [
            ("STARTER","Starter",15000,10,1,"10 salariés max, 1 utilisateur"),
            ("PRO","Pro",35000,50,3,"50 salariés max, 3 utilisateurs"),
            ("CABINET","Cabinet",100000,None,10,"Illimité, 10 utilisateurs"),
        ]: db.session.add(Plan(code=code,nom=nom,prix_mensuel=prix,max_salaries=ms,max_utilisateurs=mu,description=desc))
    if not RubriquePaie.query.first():
        for code,lib,typ,ts,tp,plaf in [
            ("CNSS","Caisse Nationale Sécurité Sociale","COTISATION",0.05,0.18,1500000),
            ("CNAMGS","Assurance Maladie Garantie Sociale","COTISATION",0.02,0.041,2500000),
            ("TCS","Taxe Complémentaire Salaires","RETENUE",0.05,None,None),
            ("FNH","Fonds National Habitat","COTISATION",None,0.03,1500000),
            ("CFP","Contribution Formation Professionnelle","COTISATION",None,0.005,None),
        ]: db.session.add(RubriquePaie(code=code,libelle=lib,type=typ,taux_salarie=ts,taux_patronal=tp,plafond_mensuel=plaf))
    if not Utilisateur.query.filter_by(role="SUPER_ADMIN").first():
        # En production, définir SUPER_ADMIN_EMAIL + SUPER_ADMIN_PASSWORD
        sa_email = os.environ.get("SUPER_ADMIN_EMAIL", "superadmin@paiegalon.com")
        sa_password = os.environ.get("SUPER_ADMIN_PASSWORD", "")
        if not sa_password:
            sa_password = sec.token_urlsafe(16)
            print(f"[INIT] Mot de passe super admin généré automatiquement : {sa_password}")
            print(f"[INIT] Email : {sa_email}  — Changez-le via l'interface admin.")
        sa = Utilisateur(nom="ADMIN", prenom="Super", email=sa_email, role="SUPER_ADMIN", actif=True)
        sa.set_password(sa_password)
        db.session.add(sa)
    if not Tenant.query.first():
        db.session.flush()
        plan = Plan.query.filter_by(code="PRO").first()
        t = Tenant(slug="demo", denomination="ENTREPRISE DEMO", sigle="DEMO", activite="À RENSEIGNER",
                   nif="", ville="Libreville", pays="Gabon",
                   plan_id=plan.id if plan else None, statut="ACTIF")
        t.token_api = sec.token_hex(32)
        db.session.add(t); db.session.flush()
        for code, lib in [("C1","Ouvriers"),("C2","Techniciens"),("C3","Conducteurs"),("C4","Cadres")]:
            db.session.add(CategorieEmploi(tenant_id=t.id, code=code, libelle=lib))
        demo_email    = os.environ.get("DEMO_EMAIL",    "demo@paiegalon.ga")
        demo_password = os.environ.get("DEMO_PASSWORD", "")
        if not demo_password:
            demo_password = sec.token_urlsafe(16)
            print(f"[INIT] Mot de passe compte démo généré : {demo_password}")
            print(f"[INIT] Email démo : {demo_email}")
        u = Utilisateur(nom="DEMO", prenom="Responsable", email=demo_email,
                        role="TENANT_ADMIN", tenant_id=t.id, actif=True)
        u.set_password(demo_password)
        db.session.add(u)
    db.session.commit()
    print("Base initialisée.\n  Super-admin: superadmin@paiegalon.com / Admin2026!\n  Compte démo: demo@paiegalon.ga / Demo2026!")

with app.app_context():
    try:
        db.create_all()
        for col in ["heures_sup_10","heures_sup_30","heures_sup_40","heures_sup_70"]:
            try:
                db.session.execute(db.text(f"ALTER TABLE pointages ADD COLUMN IF NOT EXISTS {col} NUMERIC(5,2) DEFAULT 0"))
            except Exception:
                db.session.rollback()
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        print(f"Migration init error: {e}")

    # ── Modèle bulletin ──────────────────────────────────────────────────────
    try:
        db.session.execute(db.text(
            "ALTER TABLE tenants ADD COLUMN IF NOT EXISTS modele_bulletin VARCHAR(30) DEFAULT 'classique'"
        ))
        db.session.commit()
    except Exception:
        db.session.rollback()

    # ── Champs sécurité Utilisateur ──────────────────────────────────────────
    for _col_sql in [
        "ALTER TABLE utilisateurs ADD COLUMN IF NOT EXISTS email_verifie BOOLEAN DEFAULT FALSE",
        "ALTER TABLE utilisateurs ADD COLUMN IF NOT EXISTS token_confirmation VARCHAR(200)",
        "ALTER TABLE utilisateurs ADD COLUMN IF NOT EXISTS token_confirmation_expiry TIMESTAMP",
        "ALTER TABLE utilisateurs ADD COLUMN IF NOT EXISTS nouvel_email_en_attente VARCHAR(200)",
        "ALTER TABLE utilisateurs ADD COLUMN IF NOT EXISTS token_changement_email VARCHAR(200)",
        "ALTER TABLE utilisateurs ADD COLUMN IF NOT EXISTS token_changement_expiry TIMESTAMP",
        "ALTER TABLE utilisateurs ADD COLUMN IF NOT EXISTS derniere_activite TIMESTAMP",
        "ALTER TABLE utilisateurs ADD COLUMN IF NOT EXISTS nb_echecs_connexion INTEGER DEFAULT 0",
        "ALTER TABLE utilisateurs ADD COLUMN IF NOT EXISTS compte_bloque_jusqu TIMESTAMP",
        # Marquer tous les users existants comme vérifiés (migration rétroactive)
        "UPDATE utilisateurs SET email_verifie = TRUE WHERE email_verifie IS NULL OR email_verifie = FALSE",
    ]:
        try:
            db.session.execute(db.text(_col_sql))
        except Exception:
            db.session.rollback()
    db.session.commit()

    # ── site_id + horaires dans pointages ───────────────────────────────────
    for _sql in [
        "ALTER TABLE pointages ADD COLUMN IF NOT EXISTS site_id INTEGER REFERENCES sites(id) ON DELETE SET NULL",
        "ALTER TABLE pointages ADD COLUMN IF NOT EXISTS entree_matin VARCHAR(5)",
        "ALTER TABLE pointages ADD COLUMN IF NOT EXISTS sortie_matin VARCHAR(5)",
        "ALTER TABLE pointages ADD COLUMN IF NOT EXISTS entree_apmidi VARCHAR(5)",
        "ALTER TABLE pointages ADD COLUMN IF NOT EXISTS sortie_apmidi VARCHAR(5)",
        "ALTER TABLE pointages ADD COLUMN IF NOT EXISTS entree_sup VARCHAR(5)",
        "ALTER TABLE pointages ADD COLUMN IF NOT EXISTS sortie_sup VARCHAR(5)",
        "ALTER TABLE pointages ADD COLUMN IF NOT EXISTS type_jour VARCHAR(20) DEFAULT 'NORMAL'",
    ]:
        try:
            db.session.execute(db.text(_sql))
            db.session.commit()
        except Exception:
            db.session.rollback()

    # ── Sites & Affectations ──────────────────────────────────────────────
    for _sql in [
        """CREATE TABLE IF NOT EXISTS sites (
            id SERIAL PRIMARY KEY, tenant_id INTEGER REFERENCES tenants(id) ON DELETE CASCADE,
            nom VARCHAR(200) NOT NULL, code VARCHAR(30), adresse VARCHAR(300),
            ville VARCHAR(100), responsable VARCHAR(200), telephone VARCHAR(30),
            description TEXT, actif BOOLEAN DEFAULT TRUE, date_creation TIMESTAMP DEFAULT NOW()
        )""",
        """CREATE TABLE IF NOT EXISTS affectations_sites (
            id SERIAL PRIMARY KEY,
            tenant_id INTEGER REFERENCES tenants(id) ON DELETE CASCADE,
            site_id INTEGER REFERENCES sites(id) ON DELETE CASCADE,
            salarie_id INTEGER REFERENCES salaries(id) ON DELETE SET NULL,
            journalier_id INTEGER REFERENCES journaliers(id) ON DELETE SET NULL,
            date_debut DATE NOT NULL DEFAULT CURRENT_DATE,
            date_fin DATE, actif BOOLEAN DEFAULT TRUE,
            motif VARCHAR(300), date_creation TIMESTAMP DEFAULT NOW(), cree_par VARCHAR(200)
        )""",
        "CREATE INDEX IF NOT EXISTS idx_affect_site ON affectations_sites(site_id, actif)",
        "CREATE INDEX IF NOT EXISTS idx_affect_sal  ON affectations_sites(salarie_id)",
        "CREATE INDEX IF NOT EXISTS idx_affect_jour ON affectations_sites(journalier_id)",
    ]:
        try:
            db.session.execute(db.text(_sql))
        except Exception:
            db.session.rollback()
    db.session.commit()
    for sql in [
        "ALTER TABLE tenants ADD COLUMN IF NOT EXISTS logo_url TEXT",
        "ALTER TABLE tenants ALTER COLUMN logo_url TYPE TEXT",
        "ALTER TABLE tenants ADD COLUMN IF NOT EXISTS notes TEXT",
        "ALTER TABLE tenants ADD COLUMN IF NOT EXISTS secteur VARCHAR(200)",
        "ALTER TABLE tenants ALTER COLUMN slug TYPE VARCHAR(100)",
        "ALTER TABLE tenants ALTER COLUMN denomination TYPE VARCHAR(200)",
        "ALTER TABLE salaries ADD COLUMN IF NOT EXISTS email VARCHAR(200)",
        "ALTER TABLE salaries ALTER COLUMN matricule TYPE VARCHAR(50)",
        "ALTER TABLE salaries ALTER COLUMN nom TYPE VARCHAR(100)",
        "ALTER TABLE salaries ALTER COLUMN prenom TYPE VARCHAR(100)",
        "ALTER TABLE salaries ALTER COLUMN emploi TYPE VARCHAR(150)",
        "ALTER TABLE salaries ALTER COLUMN numero_cnss TYPE VARCHAR(30)",
        "ALTER TABLE salaries ALTER COLUMN numero_cnamgs TYPE VARCHAR(30)",
        "ALTER TABLE journaliers ADD COLUMN IF NOT EXISTS date_embauche DATE",
        "ALTER TABLE journaliers ADD COLUMN IF NOT EXISTS date_debut DATE",
        "ALTER TABLE journaliers ADD COLUMN IF NOT EXISTS date_fin DATE",
        "ALTER TABLE journaliers ADD COLUMN IF NOT EXISTS nationalite VARCHAR(60)",
        "ALTER TABLE journaliers ALTER COLUMN taux_horaire TYPE NUMERIC(12,2)",
        "ALTER TABLE utilisateurs ADD COLUMN IF NOT EXISTS reset_token VARCHAR(200)",
        "ALTER TABLE utilisateurs ADD COLUMN IF NOT EXISTS reset_token_expiry TIMESTAMP",
        "ALTER TABLE bulletins_paie ADD COLUMN IF NOT EXISTS indem_compensatrice_conge NUMERIC(15,2) DEFAULT 0",
        "ALTER TABLE bulletins_paie ADD COLUMN IF NOT EXISTS indem_services_rendus NUMERIC(15,2) DEFAULT 0",
        "ALTER TABLE bulletins_paie ADD COLUMN IF NOT EXISTS indem_compensatrice_preavis NUMERIC(15,2) DEFAULT 0",
        "ALTER TABLE bulletins_paie ADD COLUMN IF NOT EXISTS indem_licenciement NUMERIC(15,2) DEFAULT 0",
            # ✅ Nouvelles colonnes base et taux pour chaque rubrique
        "ALTER TABLE bulletins_paie ADD COLUMN IF NOT EXISTS base_salaire_base NUMERIC(15,2) DEFAULT 0",
        "ALTER TABLE bulletins_paie ADD COLUMN IF NOT EXISTS taux_salaire_base VARCHAR(20) DEFAULT ''",
        "ALTER TABLE bulletins_paie ADD COLUMN IF NOT EXISTS base_heures_sup_10 NUMERIC(15,2) DEFAULT 0",
        "ALTER TABLE bulletins_paie ADD COLUMN IF NOT EXISTS taux_heures_sup_10 VARCHAR(20) DEFAULT ''",
        "ALTER TABLE bulletins_paie ADD COLUMN IF NOT EXISTS base_heures_sup_30 NUMERIC(15,2) DEFAULT 0",
        "ALTER TABLE bulletins_paie ADD COLUMN IF NOT EXISTS taux_heures_sup_30 VARCHAR(20) DEFAULT ''",
        "ALTER TABLE bulletins_paie ADD COLUMN IF NOT EXISTS base_heures_sup_40 NUMERIC(15,2) DEFAULT 0",
        "ALTER TABLE bulletins_paie ADD COLUMN IF NOT EXISTS taux_heures_sup_40 VARCHAR(20) DEFAULT ''",
        "ALTER TABLE bulletins_paie ADD COLUMN IF NOT EXISTS base_heures_sup_70 NUMERIC(15,2) DEFAULT 0",
        "ALTER TABLE bulletins_paie ADD COLUMN IF NOT EXISTS taux_heures_sup_70 VARCHAR(20) DEFAULT ''",
        "ALTER TABLE bulletins_paie ADD COLUMN IF NOT EXISTS base_absences NUMERIC(15,2) DEFAULT 0",
        "ALTER TABLE bulletins_paie ADD COLUMN IF NOT EXISTS taux_absences VARCHAR(20) DEFAULT ''",
        "ALTER TABLE bulletins_paie ADD COLUMN IF NOT EXISTS base_sursalaire NUMERIC(15,2) DEFAULT 0",
        "ALTER TABLE bulletins_paie ADD COLUMN IF NOT EXISTS taux_sursalaire VARCHAR(20) DEFAULT ''",
        "ALTER TABLE bulletins_paie ADD COLUMN IF NOT EXISTS base_prime_caisse NUMERIC(15,2) DEFAULT 0",
        "ALTER TABLE bulletins_paie ADD COLUMN IF NOT EXISTS taux_prime_caisse VARCHAR(20) DEFAULT ''",
        "ALTER TABLE bulletins_paie ADD COLUMN IF NOT EXISTS base_carburant NUMERIC(15,2) DEFAULT 0",
        "ALTER TABLE bulletins_paie ADD COLUMN IF NOT EXISTS taux_carburant VARCHAR(20) DEFAULT ''",
        "ALTER TABLE bulletins_paie ADD COLUMN IF NOT EXISTS base_prime_anciennete NUMERIC(15,2) DEFAULT 0",
        "ALTER TABLE bulletins_paie ADD COLUMN IF NOT EXISTS taux_prime_anciennete VARCHAR(20) DEFAULT ''",
        "ALTER TABLE bulletins_paie ADD COLUMN IF NOT EXISTS base_indem_logement NUMERIC(15,2) DEFAULT 0",
        "ALTER TABLE bulletins_paie ADD COLUMN IF NOT EXISTS taux_indem_logement VARCHAR(20) DEFAULT ''",
        "ALTER TABLE bulletins_paie ADD COLUMN IF NOT EXISTS base_indem_domesticite NUMERIC(15,2) DEFAULT 0",
        "ALTER TABLE bulletins_paie ADD COLUMN IF NOT EXISTS taux_indem_domesticite VARCHAR(20) DEFAULT ''",
        "ALTER TABLE bulletins_paie ADD COLUMN IF NOT EXISTS base_indem_eau_electricite NUMERIC(15,2) DEFAULT 0",
        "ALTER TABLE bulletins_paie ADD COLUMN IF NOT EXISTS taux_indem_eau_electricite VARCHAR(20) DEFAULT ''",
        "ALTER TABLE bulletins_paie ADD COLUMN IF NOT EXISTS base_indem_nourriture NUMERIC(15,2) DEFAULT 0",
        "ALTER TABLE bulletins_paie ADD COLUMN IF NOT EXISTS taux_indem_nourriture VARCHAR(20) DEFAULT ''",
        "ALTER TABLE bulletins_paie ADD COLUMN IF NOT EXISTS base_prime_transport NUMERIC(15,2) DEFAULT 0",
        "ALTER TABLE bulletins_paie ADD COLUMN IF NOT EXISTS taux_prime_transport VARCHAR(20) DEFAULT ''",
        "ALTER TABLE bulletins_paie ADD COLUMN IF NOT EXISTS base_prime_responsabilite NUMERIC(15,2) DEFAULT 0",
        "ALTER TABLE bulletins_paie ADD COLUMN IF NOT EXISTS taux_prime_responsabilite VARCHAR(20) DEFAULT ''",
        "ALTER TABLE bulletins_paie ADD COLUMN IF NOT EXISTS base_prime_rendement NUMERIC(15,2) DEFAULT 0",
        "ALTER TABLE bulletins_paie ADD COLUMN IF NOT EXISTS taux_prime_rendement VARCHAR(20) DEFAULT ''",
        "ALTER TABLE bulletins_paie ADD COLUMN IF NOT EXISTS base_prime_assiduité NUMERIC(15,2) DEFAULT 0",
        "ALTER TABLE bulletins_paie ADD COLUMN IF NOT EXISTS taux_prime_assiduité VARCHAR(20) DEFAULT ''",
        "ALTER TABLE bulletins_paie ADD COLUMN IF NOT EXISTS base_prime_qualite NUMERIC(15,2) DEFAULT 0",
        "ALTER TABLE bulletins_paie ADD COLUMN IF NOT EXISTS taux_prime_qualite VARCHAR(20) DEFAULT ''",
        "ALTER TABLE bulletins_paie ADD COLUMN IF NOT EXISTS base_prime_performance NUMERIC(15,2) DEFAULT 0",
        "ALTER TABLE bulletins_paie ADD COLUMN IF NOT EXISTS taux_prime_performance VARCHAR(20) DEFAULT ''",
        "ALTER TABLE bulletins_paie ADD COLUMN IF NOT EXISTS base_allocations_conge NUMERIC(15,2) DEFAULT 0",
        "ALTER TABLE bulletins_paie ADD COLUMN IF NOT EXISTS taux_allocations_conge VARCHAR(20) DEFAULT ''",
        "ALTER TABLE bulletins_paie ADD COLUMN IF NOT EXISTS base_indem_compensatrice_conge NUMERIC(15,2) DEFAULT 0",
        "ALTER TABLE bulletins_paie ADD COLUMN IF NOT EXISTS taux_indem_compensatrice_conge VARCHAR(20) DEFAULT ''",
        "ALTER TABLE bulletins_paie ADD COLUMN IF NOT EXISTS base_indem_services_rendus NUMERIC(15,2) DEFAULT 0",
        "ALTER TABLE bulletins_paie ADD COLUMN IF NOT EXISTS taux_indem_services_rendus VARCHAR(20) DEFAULT ''",
        "ALTER TABLE bulletins_paie ADD COLUMN IF NOT EXISTS base_indem_compensatrice_preavis NUMERIC(15,2) DEFAULT 0",
        "ALTER TABLE bulletins_paie ADD COLUMN IF NOT EXISTS taux_indem_compensatrice_preavis VARCHAR(20) DEFAULT ''",
        "ALTER TABLE bulletins_paie ADD COLUMN IF NOT EXISTS base_indem_licenciement NUMERIC(15,2) DEFAULT 0",
        "ALTER TABLE bulletins_paie ADD COLUMN IF NOT EXISTS taux_indem_licenciement VARCHAR(20) DEFAULT ''",
    ]:
        try: db.session.execute(db.text(sql)); db.session.commit()
        except Exception: db.session.rollback()
    init_db()
    print("✅ Migrations terminées.")

if __name__=="__main__":
    with app.app_context(): init_db()
    app.run(debug=True,port=5000)
