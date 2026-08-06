"""
blueprints/tenant.py — Toutes les routes tenant :
    dashboard, salariés, bulletins, congés, acomptes, périodes,
    journaliers, pointage, sites, rapports, exports, paiement,
    paramètres, utilisateurs, audit, recherche, simulateur
"""
import os, io, json, hmac, math, logging
import secrets as sec
from datetime import datetime, date, timedelta

from flask import (Blueprint, render_template, request, redirect, url_for,
                   flash, jsonify, send_file, abort, session, Response, current_app)
from flask_login import login_required, current_user, logout_user
from flask_mail import Message
from sqlalchemy.orm import joinedload
from sqlalchemy import func

from models import (db, utcnow, Plan, Tenant, Utilisateur, CategorieEmploi, Salarie,
                    Contrat, PeriodePaie, BulletinPaie, RubriquePaie, Conge,
                    Acompte, Journalier, Pointage, FeuillePaieJournalier,
                    Site, AffectationSite, Paiement, OAuthClient, AuditLog,
                    Prestataire, FacturePrestataire, ComposantPaie, BulletinComposant,
                    AvanceJournalier)
from calculs_paie import (calculer_bulletin, calculer_masse_salariale,
                           calculer_heures_sup_btp, distribuer_heures_semaine_btp,
                           calculer_prime_anciennete_btp, calculer_preavis_btp,
                           calculer_indemnite_services_rendus_btp, ventiler_heures_mois)
from audit import log_action, get_audit_logs
from core import (get_tenant, tenant_required, can_edit, admin_only,
                  require_permission, calculer_parts_irpp, parse_date,
                  cache_get, cache_set, cache_delete, csv_safe,
                  _cache_get, _cache_set, _cache_delete, _parse_date, _pd,
                  send_email_async, plan_required,
                  TTL_KPIS_DASH, TTL_EVOLUTION, TTL_CATS_STATS, TTL_ALERTES)
from i18n import SUPPORTED_LANGUAGES, set_language
from jours_feries import (jours_feries_annee, est_jour_ferie,
                          nom_jour_ferie, type_jour_auto)
from notifications import get_notifications, compter_notifications

# Modèles passés au module notifications (évite les imports circulaires)
_NOTIF_MODELS = {
    "Contrat": Contrat, "Conge": Conge, "Salarie": Salarie,
    "PeriodePaie": PeriodePaie, "FacturePrestataire": FacturePrestataire,
    "Prestataire": Prestataire,
}

logger = logging.getLogger("paiegalon")

bp = Blueprint("tenant", __name__)

# ── Rôles assignables au sein d'un tenant ─────────────────────────────────────
# Liste blanche stricte : un admin de tenant ne peut JAMAIS attribuer le rôle
# plateforme SUPER_ADMIN (sinon escalade de privilèges → accès cross-tenant).
ROLES_TENANT_AUTORISES = {
    "TENANT_ADMIN", "RH", "COMPTABLE", "DIRECTEUR", "GESTIONNAIRE", "LECTURE",
}


@bp.before_request
def _exiger_email_confirme():
    """
    Bloque l'accès aux pages tenant tant que l'email n'est pas confirmé.
    Ne s'applique qu'aux utilisateurs tenant authentifiés et non vérifiés.
    Les routes de déconnexion / confirmation / renvoi appartiennent au
    blueprint `auth` et ne sont donc jamais interceptées ici. Les appels
    JSON (calcul temps réel, simulateur) sont laissés passer pour ne pas
    casser le front : un utilisateur non confirmé ne peut de toute façon
    pas naviguer vers les pages qui les déclenchent.
    """
    if not current_user.is_authenticated:
        return
    if current_user.is_super_admin or getattr(current_user, "email_verifie", True):
        return
    # Laisser passer les requêtes JSON/API internes (pas de blocage HTML utile).
    if request.path.rsplit("/", 1)[-1].startswith("api") or "/api/" in request.path \
       or request.is_json or "application/json" in (request.headers.get("Accept", "")):
        return
    return render_template("auth/email_non_confirme.html", email=current_user.email), 403

@bp.route("/parametres/api/regenerer-token", methods=["POST"])
@tenant_required
@admin_only
def regenerer_token_api():
    """Régénère le token API du tenant. Le token en clair n'est affiché qu'ici,
    une seule fois (il est stocké haché). Réservé aux administrateurs du tenant."""
    t = get_tenant()
    raw = t.generate_token()
    db.session.commit()
    log_action("REGENERATE", "token_api", t.id, "Régénération du token API")
    db.session.commit()
    flash(f"Nouveau token API : {raw} — copiez-le maintenant, il ne sera plus affiché.",
          "success")
    return redirect(url_for("api_v1.api_clients_list"))


@bp.route("/profil/2fa", methods=["POST"])
@login_required
def basculer_2fa():
    """Active ou désactive la double authentification par email pour soi-même."""
    if current_user.is_super_admin:
        flash("La 2FA du super-admin est gérée séparément.", "error")
        return redirect(url_for("tenant.parametres") + "#securite")
    activer = request.form.get("activer") == "1"
    if activer and not os.environ.get("MAIL_PASSWORD"):
        flash("La 2FA nécessite la configuration de l'email serveur. Contactez l'administrateur.", "error")
        return redirect(url_for("tenant.parametres") + "#securite")
    current_user.twofa_active = activer
    db.session.commit()
    log_action("UPDATE", "utilisateur", current_user.id,
               f"2FA {'activée' if activer else 'désactivée'}")
    db.session.commit()
    flash(f"Double authentification {'activée' if activer else 'désactivée'}.", "success")
    return redirect(url_for("tenant.parametres") + "#securite")


@bp.route("/cabinet")
@login_required
def cabinet_dashboard():
    """Tableau de bord d'un cabinet : liste de ses entreprises clientes avec,
    pour chacune, un aperçu (salariés, statut). Réservé aux comptes cabinet."""
    if current_user.is_super_admin:
        return redirect(url_for("admin.admin_dashboard"))
    t = get_tenant()
    if not t:
        return redirect(url_for("auth.login"))
    if not t.est_cabinet:
        # Un tenant normal n'a pas de tableau de bord cabinet.
        return redirect(url_for("tenant.dashboard"))

    from sqlalchemy import func
    entreprises = (Tenant.query.filter_by(cabinet_id=t.id)
                   .order_by(Tenant.denomination).all())

    # Aperçu par entreprise : nombre de salariés actifs + dernière période
    apercu = []
    for e in entreprises:
        nb_sal = Salarie.query.filter_by(tenant_id=e.id, statut="ACTIF").count()
        derniere_periode = (PeriodePaie.query.filter_by(tenant_id=e.id)
                            .order_by(PeriodePaie.annee.desc(), PeriodePaie.mois.desc())
                            .first())
        apercu.append({
            "tenant": e,
            "nb_salaries": nb_sal,
            "derniere_periode": derniere_periode,
        })

    total_salaries = sum(a["nb_salaries"] for a in apercu)

    return render_template("tenant/cabinet_dashboard.html",
        cabinet=t, apercu=apercu,
        nb_entreprises=len(entreprises), total_salaries=total_salaries)


@bp.route("/cabinet/entrer/<int:entreprise_id>")
@login_required
def cabinet_entrer(entreprise_id):
    """Le cabinet 'entre' dans une de ses entreprises pour y travailler.
    Sécurité : on vérifie que l'entreprise appartient bien à ce cabinet."""
    if current_user.is_super_admin:
        return redirect(url_for("admin.admin_dashboard"))
    t = current_user.tenant
    if not t or not t.est_cabinet:
        flash("Réservé aux comptes cabinet.", "error")
        return redirect(url_for("tenant.dashboard"))

    entreprise = Tenant.query.get_or_404(entreprise_id)
    # GARDE-FOU : l'entreprise doit appartenir à CE cabinet.
    if entreprise.cabinet_id != t.id:
        flash("Cette entreprise ne fait pas partie de votre portefeuille.", "error")
        return redirect(url_for("tenant.cabinet_dashboard"))

    session["cabinet_entreprise_id"] = entreprise.id
    log_action("SUPPORT_ACCESS", "tenant", entreprise.id,
               f"Cabinet {t.denomination} accède à l'entreprise {entreprise.denomination}",
               user_id=current_user.id, tenant_id=entreprise.id)
    db.session.commit()
    flash(f"Vous gérez maintenant : {entreprise.denomination}", "success")
    return redirect(url_for("tenant.dashboard"))


@bp.route("/cabinet/sortir")
@login_required
def cabinet_sortir():
    """Le cabinet quitte l'entreprise courante et revient à son portefeuille."""
    session.pop("cabinet_entreprise_id", None)
    return redirect(url_for("tenant.cabinet_dashboard"))


@bp.route("/cabinet/entreprise/nouvelle", methods=["GET", "POST"])
@login_required
def cabinet_entreprise_nouvelle():
    """Le cabinet ajoute une nouvelle entreprise cliente à son portefeuille.
    Crée un tenant rattaché (cabinet_id), avec ses catégories par défaut.
    Pas de nouvel utilisateur : c'est le cabinet qui gère l'entreprise."""
    if current_user.is_super_admin:
        return redirect(url_for("admin.admin_dashboard"))
    t = get_tenant()
    if not t or not t.est_cabinet:
        flash("Réservé aux comptes cabinet.", "error")
        return redirect(url_for("tenant.dashboard"))
    if not current_user.is_tenant_admin:
        flash("Seul l'administrateur du cabinet peut ajouter une entreprise.", "error")
        return redirect(url_for("tenant.cabinet_dashboard"))

    from calculs_paie import CONVENTIONS_DISPONIBLES

    if request.method == "POST":
        denom = request.form.get("denomination", "").strip()
        if not denom:
            flash("La dénomination de l'entreprise est obligatoire.", "error")
            return render_template("tenant/cabinet_entreprise_form.html",
                cabinet=t, conventions=CONVENTIONS_DISPONIBLES)

        convention = request.form.get("convention", "AUCUNE").strip().upper()
        if convention not in CONVENTIONS_DISPONIBLES:
            convention = "AUCUNE"

        # Slug unique
        slug_base = denom.lower().replace(" ", "_")[:30]
        slug = slug_base
        i = 1
        while Tenant.query.filter_by(slug=slug).first():
            slug = f"{slug_base}_{i}"; i += 1

        # Créer l'entreprise rattachée au cabinet.
        # Elle hérite du plan du cabinet (forfait) et de son statut/expiration :
        # tant que le cabinet est à jour, ses entreprises le sont aussi.
        e = Tenant(
            slug=slug, denomination=denom.upper(),
            sigle=request.form.get("sigle", "").strip().upper(),
            activite=request.form.get("activite", "").strip(),
            nif=request.form.get("nif", "").strip(),
            numero_cnss=request.form.get("numero_cnss", "").strip(),
            telephone=request.form.get("telephone", "").strip(),
            ville=request.form.get("ville", "Libreville").strip() or "Libreville",
            pays="Gabon",
            convention=convention,
            cabinet_id=t.id,               # ← rattachement au cabinet
            plan_id=t.plan_id,             # hérite du forfait cabinet
            statut=t.statut,               # suit le statut du cabinet
            date_expiration=t.date_expiration,
        )
        e.generate_token()
        db.session.add(e)
        db.session.flush()

        # Catégories d'emploi par défaut (comme à l'inscription classique)
        for code, lib in [("C1", "Ouvriers"), ("C2", "Techniciens"),
                          ("C3", "Conducteurs de Travaux"), ("C4", "Cadres")]:
            db.session.add(CategorieEmploi(tenant_id=e.id, code=code, libelle=lib))

        db.session.commit()
        log_action("CREATE", "tenant", e.id,
                   f"Entreprise '{e.denomination}' ajoutée au cabinet {t.denomination}",
                   user_id=current_user.id, tenant_id=t.id)
        flash(f"Entreprise « {e.denomination} » ajoutée à votre portefeuille.", "success")
        return redirect(url_for("tenant.cabinet_dashboard"))

    return render_template("tenant/cabinet_entreprise_form.html",
        cabinet=t, conventions=CONVENTIONS_DISPONIBLES)


@bp.route("/dashboard")
@login_required
def dashboard():
    if current_user.is_super_admin: return redirect(url_for("admin.admin_dashboard"))
    t=get_tenant()
    if not t: flash("Aucune entreprise associée.","error"); return redirect(url_for("auth.login"))
    # Mode cabinet : un compte cabinet voit la liste de ses entreprises, pas un
    # dashboard d'entreprise unique. (Sauf s'il est "entré" dans une entreprise
    # via la bascule — étape 3 — auquel cas get_tenant() renverra l'entreprise.)
    if t.est_cabinet:
        return redirect(url_for("tenant.cabinet_dashboard"))
    now=datetime.now()

    # ── KPIs emploi (cache TTL=5min) ─────────────────────────────────────────
    _ck_kpis = f"{t.id}:kpis_emploi"
    kpis_cached = _cache_get(_ck_kpis)
    if kpis_cached:
        nb_actifs, nb_inactifs, nb_total, nb_journaliers, nb_new_mois = kpis_cached
    else:
        from sqlalchemy import func
        _sal_q = db.session.query(
            func.sum(db.cast(Salarie.statut == "ACTIF",   db.Integer)).label("actifs"),
            func.sum(db.cast(Salarie.statut == "INACTIF", db.Integer)).label("inactifs"),
            func.count().label("total"),
        ).filter(Salarie.tenant_id == t.id).one()
        nb_actifs   = int(_sal_q.actifs   or 0)
        nb_inactifs = int(_sal_q.inactifs or 0)
        nb_total    = int(_sal_q.total    or 0)
        nb_journaliers = Journalier.query.filter_by(tenant_id=t.id, statut="ACTIF").count()
        debut_mois  = datetime(now.year, now.month, 1).date()
        nb_new_mois = Salarie.query.filter(Salarie.tenant_id==t.id,
                                           Salarie.date_embauche>=debut_mois).count()
        _cache_set(_ck_kpis, (nb_actifs, nb_inactifs, nb_total, nb_journaliers, nb_new_mois),
                   TTL_KPIS_DASH)
    nb_total_employes = nb_actifs + nb_journaliers
    debut_mois  = datetime(now.year, now.month, 1).date()
    periode = PeriodePaie.query.filter_by(tenant_id=t.id, annee=now.year, mois=now.month).first()
    masse={}; nb_v=nb_b=nb_p=0
    if periode:
        buls = BulletinPaie.query.filter_by(tenant_id=t.id, periode_id=periode.id).all()
        masse = calculer_masse_salariale(buls)
        nb_v = sum(1 for b in buls if b.statut=="VALIDÉ")
        nb_p = sum(1 for b in buls if b.statut=="PAYÉ")
        nb_b = sum(1 for b in buls if b.statut=="BROUILLON")
    _ck_evo = f"{t.id}:evolution_{now.year}_{now.month}"
    evolution = _cache_get(_ck_evo)
    if evolution is None:
        evolution = []
        mois_noms = ["","Jan","Fév","Mar","Avr","Mai","Jun","Jul","Aoû","Sep","Oct","Nov","Déc"]
        for i in range(5, -1, -1):
            m = now.month - i; y = now.year
            while m <= 0: m += 12; y -= 1
            p = PeriodePaie.query.filter_by(tenant_id=t.id, annee=y, mois=m).first()
            if p:
                buls_p = BulletinPaie.query.filter_by(tenant_id=t.id, periode_id=p.id).all()
                total_net    = sum(float(b.net_a_payer    or 0) for b in buls_p)
                total_brut   = sum(float(b.salaire_brut   or 0) for b in buls_p)
                total_charges= sum(float(b.cnss_patronale or 0)+float(b.cnamgs_patronale or 0)
                                   +float(b.fnh or 0)+float(b.cfp or 0) for b in buls_p)
            else:
                total_net=total_brut=total_charges=0; buls_p=[]
            evolution.append({"mois":mois_noms[m],"annee":y,"brut":round(total_brut),
                              "net":round(total_net),"charges":round(total_charges),
                              "nb_bulletins":len(buls_p)})
        _cache_set(_ck_evo, evolution, TTL_EVOLUTION)
    top_salaries = []
    if periode:
        top_salaries = (BulletinPaie.query
            .filter_by(tenant_id=t.id, periode_id=periode.id)
            .options(joinedload(BulletinPaie.salarie))
            .order_by(BulletinPaie.net_a_payer.desc())
            .limit(5).all())
    from sqlalchemy import func
    _ck_cats = f"{t.id}:cats_stats"
    cats_stats = _cache_get(_ck_cats)
    if cats_stats is None:
        cats_stats = db.session.query(
            CategorieEmploi.code, CategorieEmploi.libelle,
            func.count(Salarie.id).label("nb")
        ).join(Salarie, Salarie.categorie_id==CategorieEmploi.id)\
         .filter(Salarie.tenant_id==t.id, Salarie.statut=="ACTIF")\
         .group_by(CategorieEmploi.code, CategorieEmploi.libelle).all()
        # Convertir en liste de tuples simples (JSON sérialisable)
        cats_stats = [(r.code, r.libelle, r.nb) for r in cats_stats]
        _cache_set(_ck_cats, cats_stats, TTL_CATS_STATS)
    derniers = (BulletinPaie.query
        .filter_by(tenant_id=t.id)
        .options(joinedload(BulletinPaie.salarie),
                 joinedload(BulletinPaie.periode))
        .order_by(BulletinPaie.date_creation.desc())
        .limit(6).all())
    # ══════════════════════════════════════════════════════════════════════════
    # ALERTES INTELLIGENTES
    # ══════════════════════════════════════════════════════════════════════════
    import calendar
    alertes = []
    MOIS_NOMS_LONG = ["","Janvier","Février","Mars","Avril","Mai","Juin",
                      "Juillet","Août","Septembre","Octobre","Novembre","Décembre"]
    mois_courant = MOIS_NOMS_LONG[now.month]
    debut_mois_d = date(now.year, now.month, 1)
    fin_mois_d   = date(now.year, now.month, calendar.monthrange(now.year, now.month)[1])

    # ── 1. Quota employés dépassé ou proche ──────────────────────────────────
    q_info = t.quota_employes_info
    if q_info.get("max"):
        pct = round(q_info["actuel"] / q_info["max"] * 100)
        if q_info["plein"]:
            alertes.append({"type":"danger","icone":"🚫","titre":"Limite d'employés atteinte",
                "msg":f"Vous avez atteint la limite de {q_info['max']} employés de votre plan {t.plan.nom}. "
                      f"Impossible d'ajouter de nouveaux travailleurs.",
                "lien":"/parametres","lien_texte":"Changer de plan"})
        elif pct >= 80:
            alertes.append({"type":"warning","icone":"⚠️","titre":"Quota employés bientôt atteint",
                "msg":f"{q_info['actuel']}/{q_info['max']} employés utilisés ({pct}%). "
                      f"Il reste {q_info['max'] - q_info['actuel']} place(s).",
                "lien":"/parametres","lien_texte":"Voir le plan"})

    # ── 2. Période non créée pour le mois courant ─────────────────────────────
    if not periode:
        alertes.append({"type":"warning","icone":"📅","titre":f"Période {mois_courant} {now.year} manquante",
            "msg":f"Aucune période de paie n'est ouverte pour {mois_courant} {now.year}. "
                  f"Les bulletins ne peuvent pas être saisis.",
            "lien":"/periodes","lien_texte":"Créer la période"})

    # ── 3. Période précédente non clôturée ────────────────────────────────────
    mois_prec = now.month - 1 or 12
    annee_prec = now.year if now.month > 1 else now.year - 1
    periode_prec = PeriodePaie.query.filter_by(
        tenant_id=t.id, annee=annee_prec, mois=mois_prec).first()
    if periode_prec and periode_prec.statut not in ("CLÔTURÉE", "CLOTUREE", "PAYÉE"):
        buls_prec = BulletinPaie.query.filter_by(
            tenant_id=t.id, periode_id=periode_prec.id).all()
        nb_non_clos = sum(1 for b in buls_prec if b.statut in ("BROUILLON", "VALIDÉ"))
        if nb_non_clos > 0:
            alertes.append({"type":"warning","icone":"🔓","titre":f"Période {MOIS_NOMS_LONG[mois_prec]} non clôturée",
                "msg":f"{nb_non_clos} bulletin(s) de {MOIS_NOMS_LONG[mois_prec]} {annee_prec} "
                      f"ne sont pas encore payés.",
                "lien":f"/bulletins?periode_id={periode_prec.id}","lien_texte":"Voir les bulletins"})

    # ── 4. Bulletins en brouillon ce mois ─────────────────────────────────────
    if nb_b > 0:
        alertes.append({"type":"warning","icone":"📝","titre":f"{nb_b} brouillon(s) à valider",
            "msg":f"{nb_b} bulletin(s) de {mois_courant} sont en brouillon et doivent être validés.",
            "lien":f"/bulletins?periode_id={periode.id}&statut=BROUILLON" if periode else "/bulletins",
            "lien_texte":"Valider maintenant"})

    # ── 5. Salariés actifs sans bulletin ce mois ──────────────────────────────
    if periode:
        ids_avec_bulletin = {b.salarie_id for b in
            BulletinPaie.query.filter_by(tenant_id=t.id, periode_id=periode.id).all()}
        salaries_sans_bulletin = Salarie.query.filter_by(
            tenant_id=t.id, statut="ACTIF"
        ).filter(~Salarie.id.in_(ids_avec_bulletin)).all() if ids_avec_bulletin else             Salarie.query.filter_by(tenant_id=t.id, statut="ACTIF").all()
        nb_sans = len(salaries_sans_bulletin)
        if nb_sans > 0:
            exemples = ", ".join(s.nom_complet for s in salaries_sans_bulletin[:3])
            if nb_sans > 3: exemples += f" et {nb_sans-3} autre(s)"
            alertes.append({"type":"info","icone":"👤","titre":f"{nb_sans} salarié(s) sans bulletin ce mois",
                "msg":f"{exemples}.",
                "lien":f"/bulletins/saisie","lien_texte":"Saisir un bulletin"})

    # ── 6. Journaliers avec feuilles en attente de paiement ───────────────────
    feuilles_att = FeuillePaieJournalier.query.filter_by(
        tenant_id=t.id, statut="EN_ATTENTE").all()
    nb_feuilles_att = len(feuilles_att)
    if nb_feuilles_att > 0:
        montant_att = sum(float(f.montant_brut or 0) for f in feuilles_att)
        alertes.append({"type":"warning","icone":"🦺","titre":f"{nb_feuilles_att} feuille(s) journaliers non payée(s)",
            "msg":f"Total en attente : {int(montant_att):,} FCFA.",
            "lien":"/journaliers/paie","lien_texte":"Payer maintenant"})

    # ── 7. Journaliers actifs sans pointage cette semaine ─────────────────────
    lundi = (now.date() - timedelta(days=now.weekday()))
    samedi = lundi + timedelta(days=5)
    nb_jour_actifs = Journalier.query.filter_by(tenant_id=t.id, statut="ACTIF").count()
    if nb_jour_actifs > 0:
        ids_pointes_sem = {p.journalier_id for p in
            Pointage.query.filter_by(tenant_id=t.id)
            .filter(Pointage.journalier_id.isnot(None),
                    Pointage.date_pointage >= lundi,
                    Pointage.date_pointage <= samedi).all()}
        nb_non_pointes_jour = nb_jour_actifs - len(ids_pointes_sem)
        if nb_non_pointes_jour > 0 and now.weekday() >= 1:   # après lundi
            alertes.append({"type":"info","icone":"📋","titre":f"{nb_non_pointes_jour} journalier(s) non pointé(s) cette semaine",
                "msg":f"Semaine du {lundi.strftime('%d/%m')} : {nb_non_pointes_jour} journalier(s) "
                      f"sans pointage enregistré.",
                "lien":f"/pointage/recap-semaine","lien_texte":"Voir le récap"})

    # ── 8. Acomptes en attente de déduction ───────────────────────────────────
    acomptes_att = Acompte.query.filter_by(tenant_id=t.id, statut="EN_ATTENTE").count()
    if acomptes_att > 0:
        alertes.append({"type":"info","icone":"💸","titre":f"{acomptes_att} acompte(s) en attente",
            "msg":f"{acomptes_att} acompte(s) n'ont pas encore été déduits des bulletins.",
            "lien":"/acomptes","lien_texte":"Voir les acomptes"})

    # ── 9. Statut abonnement/essai — TOUJOURS affiché ───────────────────────
    exp_date = t.date_expiration
    jours_restants = (exp_date.date() - now.date()).days if exp_date else None

    if t.statut == "ESSAI":
        if jours_restants is None or jours_restants <= 0:
            alertes.append({"type":"danger","icone":"🚫",
                "titre":"Période d'essai expirée",
                "msg":"Votre essai gratuit a expiré. Souscrivez maintenant pour continuer.",
                "lien":"/parametres","lien_texte":"Souscrire maintenant"})
        elif jours_restants <= 7:
            alertes.append({"type":"danger","icone":"⏰",
                "titre":f"Essai : {jours_restants} jour(s) restant(s)",
                "msg":f"Expire le {exp_date.strftime('%d/%m/%Y')}. Souscrivez pour ne pas perdre vos données.",
                "lien":"/parametres","lien_texte":"Souscrire maintenant"})
        elif jours_restants <= 14:
            alertes.append({"type":"warning","icone":"⏳",
                "titre":f"Essai : {jours_restants} jour(s) restant(s)",
                "msg":f"Votre période d'essai se termine le {exp_date.strftime('%d/%m/%Y')}.",
                "lien":"/parametres","lien_texte":"Voir les plans"})
        else:
            alertes.append({"type":"info","icone":"🧪",
                "titre":f"Période d'essai — {jours_restants} jour(s) restant(s)",
                "msg":f"Essai gratuit jusqu'au {exp_date.strftime('%d/%m/%Y')}. Plan actuel : {t.plan.nom}.",
                "lien":"/parametres","lien_texte":"Voir les plans"})

    elif t.statut == "ACTIF":
        if jours_restants is None or jours_restants <= 0:
            alertes.append({"type":"danger","icone":"🔒",
                "titre":"Abonnement expiré",
                "msg":"Votre abonnement a expiré. Renouvelez pour continuer.",
                "lien":"/parametres","lien_texte":"Renouveler"})
        elif jours_restants <= 7:
            alertes.append({"type":"danger","icone":"⏰",
                "titre":f"Abonnement : {jours_restants} jour(s) restant(s)",
                "msg":f"Expire le {exp_date.strftime('%d/%m/%Y')}. Renouvelez dès maintenant.",
                "lien":"/parametres","lien_texte":"Renouveler"})
        elif jours_restants <= 30:
            alertes.append({"type":"warning","icone":"📅",
                "titre":f"Abonnement : {jours_restants} jour(s) restant(s)",
                "msg":f"Expire le {exp_date.strftime('%d/%m/%Y')}.",
                "lien":"/parametres","lien_texte":"Renouveler"})
        else:
            alertes.append({"type":"info","icone":"✅",
                "titre":f"Abonnement actif — {jours_restants} jour(s) restant(s)",
                "msg":f"Plan {t.plan.nom} valide jusqu'au {exp_date.strftime('%d/%m/%Y')}.",
                "lien":"/parametres","lien_texte":"Gérer l'abonnement"})

    elif t.statut == "PAIEMENT_EN_ATTENTE":
        alertes.append({"type":"warning","icone":"💳",
            "titre":"Paiement en attente",
            "msg":"Votre paiement est en cours de traitement. L'accès complet sera rétabli dès confirmation.",
            "lien":"/parametres","lien_texte":"Contacter le support"})

    # Trier par priorité : danger > warning > info
    _prio = {"danger": 0, "warning": 1, "info": 2}
    alertes.sort(key=lambda a: _prio.get(a["type"], 3))

    # Compteurs pour l'en-tête
    nb_alertes_critiques = sum(1 for a in alertes if a["type"] == "danger")
    nb_alertes_warning   = sum(1 for a in alertes if a["type"] == "warning")

    return render_template("tenant/dashboard.html", tenant=t,
        nb_actifs=nb_actifs, nb_inactifs=nb_inactifs, nb_total=nb_total,
        nb_journaliers=nb_journaliers, nb_total_employes=nb_total_employes,
        nb_new_mois=nb_new_mois, periode=periode, masse=masse,
        nb_valides=nb_v, nb_payes=nb_p, nb_brouillon=nb_b,
        evolution=evolution, top_salaries=top_salaries,
        cats_stats=cats_stats, derniers=derniers, alertes=alertes,
        nb_alertes_critiques=nb_alertes_critiques,
        nb_alertes_warning=nb_alertes_warning, now=now)

# ── Salariés ──────────────────────────────────────────────────────────────────
@bp.route("/salaries")
@login_required
def salaries():
    if current_user.is_super_admin: return redirect(url_for("admin.admin_dashboard"))
    t = get_tenant()
    if not t: return redirect(url_for("auth.login"))
    q      = request.args.get("q", "")
    statut = request.args.get("statut", "")
    page   = request.args.get("page", 1, type=int)
    query  = Salarie.query.filter_by(tenant_id=t.id)
    if q:      query = query.filter(db.or_(Salarie.nom.ilike(f"%{q}%"), Salarie.prenom.ilike(f"%{q}%"), Salarie.matricule.ilike(f"%{q}%")))
    if statut: query = query.filter_by(statut=statut)
    query = query.options(joinedload(Salarie.categorie), joinedload(Salarie.contrats))
    pagination = query.order_by(Salarie.nom).paginate(page=page, per_page=25, error_out=False)
    _args  = {k: v for k, v in request.args.items() if k != 'page'}
    _base  = request.path + '?' + '&'.join(f'{k}={v}' for k, v in _args.items())
    _sep   = '&' if _args else '?'
    return render_template("tenant/salaries.html",
        salaries=pagination.items, pagination=pagination,
        categories=CategorieEmploi.query.filter_by(tenant_id=t.id).all(),
        q=q, statut=statut, tenant=t,
        pagination_base=_base + _sep)







@bp.route("/simulateur")
@login_required
def simulateur_paie():
    """Simulateur de paie interactif — avec comparaison scénarios, net→brut, augmentation."""
    if current_user.is_super_admin: return redirect(url_for("admin.admin_dashboard"))
    t = get_tenant()
    if not t: return redirect(url_for("auth.login"))
    salaries_list = Salarie.query.filter_by(tenant_id=t.id, statut="ACTIF")\
        .options(joinedload(Salarie.categorie)).order_by(Salarie.nom).all()
    return render_template("tenant/simulateur.html", tenant=t, salaries=salaries_list,
                           convention=t.convention or "AUCUNE")


@bp.route("/api/simuler-paie/scenarios", methods=["POST"])
@login_required
def api_simuler_scenarios():
    """Compare jusqu'à 3 scénarios de paie côte à côte."""
    if current_user.is_super_admin: return jsonify({"error": "forbidden"}), 403
    t = get_tenant()
    if not t: return jsonify({"error": "non authentifié"}), 401

    data = request.get_json(force=True) or {}
    scenarios = data.get("scenarios", [])
    nb_parts  = float(data.get("nb_parts", 1.0))

    if not scenarios or len(scenarios) < 2:
        return jsonify({"error": "Au moins 2 scénarios requis"}), 400

    from simulation_paie import comparer_scenarios
    result = comparer_scenarios(scenarios, nb_parts=nb_parts)
    return jsonify(result)


@bp.route("/api/simuler-paie/net-vers-brut", methods=["POST"])
@login_required
def api_simuler_net_vers_brut():
    """Calcule le brut nécessaire pour atteindre un net cible."""
    if current_user.is_super_admin: return jsonify({"error": "forbidden"}), 403
    t = get_tenant()
    if not t: return jsonify({"error": "non authentifié"}), 401

    data = request.get_json(force=True) or {}
    net_cible = float(data.get("net_cible", 0))
    nb_parts  = float(data.get("nb_parts",  1.0))
    extras    = data.get("extras", {})

    if not net_cible or net_cible <= 0:
        return jsonify({"error": "Net cible invalide"}), 400

    from simulation_paie import simuler_depuis_net
    result = simuler_depuis_net(net_cible, nb_parts=nb_parts, donnees_extra=extras)
    return jsonify(result)


@bp.route("/api/simuler-paie/augmentation", methods=["POST"])
@login_required
def api_simuler_augmentation():
    """Simule l'impact d'une augmentation de salaire."""
    if current_user.is_super_admin: return jsonify({"error": "forbidden"}), 403
    t = get_tenant()
    if not t: return jsonify({"error": "non authentifié"}), 401

    data = request.get_json(force=True) or {}
    salaire_actuel      = float(data.get("salaire_actuel", 0))
    augmentation_pct    = data.get("augmentation_pct")
    augmentation_montant= data.get("augmentation_montant")
    nb_parts            = float(data.get("nb_parts", 1.0))
    extras              = data.get("extras", {})

    if not salaire_actuel or salaire_actuel <= 0:
        return jsonify({"error": "Salaire actuel invalide"}), 400

    from simulation_paie import simuler_augmentation
    result = simuler_augmentation(
        salaire_actuel       = salaire_actuel,
        augmentation_pct     = float(augmentation_pct) if augmentation_pct else None,
        augmentation_montant = float(augmentation_montant) if augmentation_montant else None,
        nb_parts             = nb_parts,
        donnees_extra        = extras,
    )
    return jsonify(result)


@bp.route("/api/simuler-paie", methods=["POST"])
@login_required
def api_simuler_paie():
    """API JSON : simuler un bulletin de paie complet sans le sauvegarder."""
    t = get_tenant()
    if not t: return jsonify({"erreur": "non connecté"})
    from calculs_paie import calculer_bulletin, calculer_heures_sup_btp
    try:
        d = request.get_json(force=True) or {}
        # Accepter aussi le form-data
        if not d:
            d = {k: request.form.get(k) for k in request.form}

        def flt(key, default=0):
            try: return float(str(d.get(key) or default).replace(",",".") or default)
            except: return float(default)

        # Conversion des heures supplémentaires (saisies en HEURES) → montants.
        # Effectuée quelle que soit la convention : les coefficients +10/+30/+40/+70
        # sont communs au BTP et au Commerce (la sélection de convention côté UI
        # ne change que les libellés et le préremplissage structurel BTP).
        sal_base = flt("salaire_base")
        if sal_base > 0:
            from calculs_paie import calculer_heures_sup_btp
            hs = calculer_heures_sup_btp(sal_base,
                h10=flt("h10"), h30=flt("h30"),
                h40=flt("h40"), h70=flt("h70"),
                h30b=flt("h30b"), convention=t.convention)
            d["heures_sup_10"] = hs["montant_10"]
            d["heures_sup_30"] = hs["montant_30"]
            d["heures_sup_30b"] = hs["montant_30b"]
            d["heures_sup_40"] = hs["montant_40"]
            d["heures_sup_70"] = hs["montant_70"]
        d["convention"] = t.convention

        # Nombre de parts IRPP
        sal_id = d.get("salarie_id")
        nb_parts = 1.0
        if sal_id:
            s = Salarie.query.filter_by(id=int(sal_id), tenant_id=t.id).first()
            if s: nb_parts = float(s.nombre_parts or 1)
        nb_parts = flt("nb_parts") or nb_parts

        # Composants personnalisés (aperçu temps réel) : lus depuis composant_<id>
        comps_live = []
        for comp in ComposantPaie.query.filter_by(tenant_id=t.id, actif=True).all():
            montant = flt(f"composant_{comp.id}")
            if montant:
                comps_live.append({
                    "libelle": comp.libelle, "sens": comp.sens, "montant": montant,
                    "soumis_cnss": comp.soumis_cnss, "soumis_cnamgs": comp.soumis_cnamgs,
                    "soumis_irpp": comp.soumis_irpp})
        d["composants"] = comps_live

        result = calculer_bulletin(d, nb_parts=nb_parts)

        # Enrichir avec détails BTP
        from calculs_paie import calculer_taux_horaire, H_NORMALES_MENSUEL
        if sal_base > 0:
            th = calculer_taux_horaire(sal_base)
            result["taux_horaire"]     = round(th, 2)
            result["h_normales_mensuel"] = H_NORMALES_MENSUEL

        # Ajouter libellés pour l'affichage
        result["gains_detail"] = [
            {"label": "Salaire de base",          "montant": result["salaire_base"]},
            {"label": "H.sup +10%",               "montant": result["heures_sup_10"]},
            {"label": "H.sup +30%",               "montant": result["heures_sup_30"]},
            {"label": "H.sup +30% (repos/férié)", "montant": result.get("heures_sup_30b", 0)},
            {"label": "H.sup +40% (nuit/dim.)",   "montant": result["heures_sup_40"]},
            {"label": "H.sup +70% (fériés)",      "montant": result["heures_sup_70"]},
            {"label": "Sursalaire",               "montant": result["sursalaire"]},
            {"label": "Prime de transport",       "montant": result["prime_transport"]},
            {"label": "Prime de responsabilité",  "montant": result["prime_responsabilite"]},
            {"label": "Indemnité logement",       "montant": result["indem_logement"]},
            {"label": "Prime d'ancienneté",      "montant": result["prime_anciennete"]},
            {"label": "Autres primes",            "montant": sum([
                result["prime_caisse"], result["carburant"],
                result["prime_rendement"], result["prime_qualite"],
                result["prime_performance"], result["prime_assiduité"],
            ])},
        ]
        result["gains_detail"] = [g for g in result["gains_detail"] if g["montant"] > 0]

        result["retenues_detail"] = [
            {"label": f"CNSS salarié (5% / base {int(result['base_cnss']):,} FCFA)", "montant": result["cnss_salarie"]},
            {"label": f"CNAMGS salarié (2% / base {int(result['base_cnamgs']):,} FCFA)", "montant": result["cnamgs_salarie"]},
            {"label": f"TCS (5% / base {int(result['base_tcs']):,} FCFA)",             "montant": result["tcs"]},
            {"label": f"IRPP ({nb_parts} part(s))",                                    "montant": result["irpp"]},
            {"label": "Acompte",                                                        "montant": result["acompte"]},
            {"label": "Retenue absences",                                               "montant": result["absences"]},
        ]
        result["retenues_detail"] = [r for r in result["retenues_detail"] if r["montant"] > 0]

        result["charges_pat_detail"] = [
            {"label": "CNSS patronal (18%)",     "montant": result["cnss_patronale"]},
            {"label": "CNAMGS patronal (4.1%)",  "montant": result["cnamgs_patronale"]},
            {"label": "FNH (3%)",                "montant": result["fnh"]},
            {"label": "CFP (0.5%)",              "montant": result["cfp"]},
        ]

        return jsonify(result)
    except Exception as e:
        import traceback
        return jsonify({"erreur": str(e), "trace": traceback.format_exc()[-300:]})

@bp.route("/salaries/import", methods=["GET","POST"])
@login_required
def salaries_import():
    """Import en masse de salariés depuis un fichier Excel."""
    if current_user.is_super_admin: return redirect(url_for("admin.admin_dashboard"))
    t = get_tenant()
    if not t: return redirect(url_for("auth.login"))

    if request.method == "GET":
        categories = CategorieEmploi.query.filter_by(tenant_id=t.id).all()
        return render_template("tenant/salaries_import.html",
            tenant=t, categories=categories)

    # ── POST : traitement du fichier ─────────────────────────────────────────
    fichier = request.files.get("fichier")
    if not fichier or not fichier.filename.endswith((".xlsx", ".xls")):
        flash("❌ Fichier invalide. Utilisez le modèle Excel fourni (.xlsx).", "error")
        return redirect(url_for("tenant.salaries_import"))

    # Garde anti-DoS (zip-bomb / fichier surdimensionné) : 5 Mo max.
    fichier.seek(0, 2); _taille = fichier.tell(); fichier.seek(0)
    if _taille > 5_000_000:
        flash("❌ Fichier trop volumineux (max 5 Mo).", "error")
        return redirect(url_for("tenant.salaries_import"))

    mode = request.form.get("mode", "ignorer")  # ignorer | ecraser

    import openpyxl
    from datetime import date as date_type

    def parse_date(val):
        if not val: return None
        if isinstance(val, (date_type, datetime)): return val if isinstance(val, date_type) else val.date()
        s = str(val).strip()
        for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d/%m/%y"):
            try: return datetime.strptime(s, fmt).date()
            except: pass
        return None

    def clean(val): return str(val).strip() if val not in (None, "") else None

    try:
        wb = openpyxl.load_workbook(fichier, data_only=True)
        ws = wb.active
    except Exception as e:
        flash(f"❌ Erreur lecture fichier : {e}", "error")
        return redirect(url_for("tenant.salaries_import"))

    # Trouver la ligne d'en-tête (ligne avec "MATRICULE")
    header_row = None
    for row_idx in range(1, 10):
        row_vals = [str(ws.cell(row_idx, c).value or "").upper().strip() for c in range(1, 20)]
        if "MATRICULE" in row_vals:
            header_row = row_idx
            break

    if not header_row:
        flash("❌ Entête non trouvée. Utilisez le modèle Excel fourni.", "error")
        return redirect(url_for("tenant.salaries_import"))

    # Mapper les colonnes
    headers = {}
    for col in range(1, ws.max_column + 1):
        val = str(ws.cell(header_row, col).value or "").upper().strip()
        if val: headers[val] = col

    required = ["MATRICULE", "NOM", "PRENOM", "EMPLOI", "DATE_EMBAUCHE"]
    for req in required:
        if req not in headers:
            flash(f"❌ Colonne obligatoire manquante : {req}", "error")
            return redirect(url_for("tenant.salaries_import"))

    # Charger les catégories du tenant
    cats = {c.code.upper(): c for c in CategorieEmploi.query.filter_by(tenant_id=t.id).all()}

    # ── Vérifier quota ────────────────────────────────────────────────────────
    nb_existants = Salarie.query.filter_by(tenant_id=t.id, statut="ACTIF").count()
    quota = t.quota_employes_info

    # Traiter les lignes de données
    nb_crees = nb_maj = nb_erreurs = nb_ignores = 0
    erreurs   = []
    avertissements = []

    for row_idx in range(header_row + 1, ws.max_row + 1):
        def get(col_name):
            c = headers.get(col_name)
            return ws.cell(row_idx, c).value if c else None

        matricule = clean(get("MATRICULE"))
        if not matricule: continue  # ligne vide

        nom    = clean(get("NOM"))
        prenom = clean(get("PRENOM"))
        emploi = clean(get("EMPLOI"))
        date_e = parse_date(get("DATE_EMBAUCHE"))

        # Validation champs obligatoires
        if not all([nom, prenom, emploi, date_e]):
            erreurs.append(f"Ligne {row_idx} ({matricule}) : champs obligatoires manquants")
            nb_erreurs += 1
            continue

        # Vérifier si matricule existe
        existant = Salarie.query.filter_by(tenant_id=t.id, matricule=matricule).first()

        if existant and mode == "ignorer":
            nb_ignores += 1
            continue

        # Récupérer catégorie
        cat_code = clean(get("CATEGORIE"))
        cat = cats.get(cat_code.upper()) if cat_code else None

        # Salaire de base → créer/maj contrat
        salaire_raw = get("SALAIRE_BASE")
        salaire_base = None
        if salaire_raw:
            try: salaire_base = float(str(salaire_raw).replace(" ","").replace(",","."))
            except: pass

        # Quota check pour nouvelles créations
        if not existant and quota.get("max"):
            if nb_existants + nb_crees >= quota["max"]:
                erreurs.append(f"Quota atteint ({quota['max']} employés). Import arrêté à la ligne {row_idx}.")
                break

        data = dict(
            tenant_id              = t.id,
            matricule              = matricule.upper(),
            nom                    = nom.upper(),
            prenom                 = prenom,
            emploi                 = emploi.upper(),
            date_embauche          = date_e,
            categorie_id           = cat.id if cat else None,
            telephone              = clean(get("TELEPHONE")),
            sexe                   = clean(get("SEXE")),
            date_naissance         = parse_date(get("DATE_NAISSANCE")),
            situation_matrimoniale = clean(get("SITUATION_MAT")),
            nationalite            = clean(get("NATIONALITE")) or "GABONAISE",
            numero_cnss            = clean(get("NUMERO_CNSS")),
            numero_cnamgs          = clean(get("NUMERO_CNAMGS")),
            email                  = clean(get("EMAIL")),
            adresse                = clean(get("ADRESSE")),
            statut                 = "ACTIF",
        )
        try: data["nb_enfants"]   = int(float(str(get("NB_ENFANTS") or 0)))
        except: data["nb_enfants"] = 0
        try: data["nombre_parts"] = float(str(get("NOMBRE_PARTS") or 1).replace(",","."))
        except: data["nombre_parts"] = 1

        try:
            if existant:
                for k, v in data.items():
                    if v is not None: setattr(existant, k, v)
                s_obj = existant
                nb_maj += 1
            else:
                s_obj = Salarie(**data)
                db.session.add(s_obj)
                db.session.flush()
                nb_crees += 1

            # Créer/maj contrat si salaire fourni
            if salaire_base and salaire_base > 0:
                contrat = next((c for c in (s_obj.contrats if s_obj.id else [])), None)
                if not contrat:
                    contrat = Contrat(tenant_id=t.id, salarie_id=s_obj.id,
                                      date_debut=date_e, actif=True)
                    db.session.add(contrat)
                contrat.salaire_base = salaire_base
                contrat.type_contrat = "CDI"
                contrat.actif        = True

        except Exception as e:
            db.session.rollback()
            erreurs.append(f"Ligne {row_idx} ({matricule}) : {str(e)[:80]}")
            nb_erreurs += 1
            continue

    db.session.commit()
    _cache_delete(f"{t.id}:")  # Invalider cache

    # Message résumé
    msg_parts = []
    if nb_crees:   msg_parts.append(f"✅ {nb_crees} salarié(s) créé(s)")
    if nb_maj:     msg_parts.append(f"🔄 {nb_maj} mis à jour")
    if nb_ignores: msg_parts.append(f"⏭️ {nb_ignores} ignoré(s) (déjà existants)")
    if nb_erreurs: msg_parts.append(f"❌ {nb_erreurs} erreur(s)")
    flash(" · ".join(msg_parts) or "Aucune donnée importée.", "success" if not nb_erreurs else "error")

    for err in erreurs[:5]:
        flash(f"⚠️ {err}", "error")

    return redirect(url_for("tenant.salaries"))


@bp.route("/salaries/import/modele")
@login_required
def salaries_import_modele():
    """Génère et télécharge le modèle Excel vierge d'import des salariés.

    Le fichier est produit à la volée (openpyxl) pour rester toujours disponible
    et toujours aligné sur les colonnes réellement attendues par l'import.
    """
    if current_user.is_super_admin:
        return redirect(url_for("admin.admin_dashboard"))
    t = get_tenant()
    if not t:
        return redirect(url_for("auth.login"))

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO
    from flask import send_file

    # Colonnes : les 5 premières sont OBLIGATOIRES, les suivantes optionnelles.
    colonnes = [
        ("MATRICULE", True), ("NOM", True), ("PRENOM", True),
        ("EMPLOI", True), ("DATE_EMBAUCHE", True),
        ("SEXE", False), ("DATE_NAISSANCE", False), ("NATIONALITE", False),
        ("SITUATION_MAT", False), ("NB_ENFANTS", False), ("NOMBRE_PARTS", False),
        ("ADRESSE", False), ("TELEPHONE", False), ("EMAIL", False),
        ("CATEGORIE", False), ("SALAIRE_BASE", False),
        ("NUMERO_CNSS", False), ("NUMERO_CNAMGS", False),
    ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Salariés"

    entete_oblig = Font(bold=True, color="FFFFFF")
    entete_opt   = Font(bold=True, color="0F3D36")
    fill_oblig   = PatternFill("solid", fgColor="0F3D36")
    fill_opt     = PatternFill("solid", fgColor="E8EFEC")

    for idx, (nom, oblig) in enumerate(colonnes, start=1):
        cell = ws.cell(row=1, column=idx, value=nom)
        cell.alignment = Alignment(horizontal="center")
        cell.font = entete_oblig if oblig else entete_opt
        cell.fill = fill_oblig if oblig else fill_opt
        ws.column_dimensions[cell.column_letter].width = max(14, len(nom) + 2)

    # Ligne d'exemple (sera ignorée si supprimée ; sert de guide de format)
    exemple = {
        "MATRICULE": "0001", "NOM": "NDONG", "PRENOM": "Jean",
        "EMPLOI": "Maçon", "DATE_EMBAUCHE": "15/01/2026",
        "SEXE": "M", "DATE_NAISSANCE": "10/06/1990", "NATIONALITE": "Gabonaise",
        "SITUATION_MAT": "MARIE", "NB_ENFANTS": 2, "NOMBRE_PARTS": 2.5,
        "ADRESSE": "Libreville", "TELEPHONE": "077000000", "EMAIL": "",
        "CATEGORIE": "C1", "SALAIRE_BASE": 200000,
        "NUMERO_CNSS": "", "NUMERO_CNAMGS": "",
    }
    for idx, (nom, _) in enumerate(colonnes, start=1):
        ws.cell(row=2, column=idx, value=exemple.get(nom, ""))
    for c in range(1, len(colonnes) + 1):
        ws.cell(row=2, column=c).font = Font(italic=True, color="9CA3AF")

    # Feuille d'instructions, avec les catégories réellement définies pour ce tenant.
    ws2 = wb.create_sheet("Instructions")
    cats = CategorieEmploi.query.filter_by(tenant_id=t.id).all()
    lignes_info = [
        ("Modèle d'import des salariés — PaieGabon", True),
        ("", False),
        ("Colonnes OBLIGATOIRES : MATRICULE, NOM, PRENOM, EMPLOI, DATE_EMBAUCHE", False),
        ("Les autres colonnes sont facultatives.", False),
        ("Format des dates : JJ/MM/AAAA (ex. 15/01/2026).", False),
        ("SEXE : M ou F.", False),
        ("SITUATION_MAT : CELIBATAIRE, MARIE, DIVORCE, VEUF.", False),
        ("SALAIRE_BASE : montant entier en FCFA, sans espaces (ex. 200000).", False),
        ("Supprimez la ligne d'exemple avant l'import.", False),
        ("", False),
        ("Codes CATEGORIE disponibles pour votre entreprise :", True),
    ]
    if cats:
        for c in cats:
            lignes_info.append((f"   {c.code} — {c.libelle}", False))
    else:
        lignes_info.append(("   (aucune catégorie définie — laissez la colonne vide)", False))
    for i, (txt, gras) in enumerate(lignes_info, start=1):
        cell = ws2.cell(row=i, column=1, value=txt)
        if gras:
            cell.font = Font(bold=True, color="0F3D36")
    ws2.column_dimensions["A"].width = 70

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name="modele_import_salaries.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@bp.route("/salaries/nouveau", methods=["GET","POST"])
@login_required
def salarie_nouveau():
    if current_user.is_super_admin: return redirect(url_for("admin.admin_dashboard"))
    t=get_tenant()
    if not t: return redirect(url_for("auth.login"))
    if not current_user.can_edit: abort(403)
    cats=CategorieEmploi.query.filter_by(tenant_id=t.id).all()
    # Vérifier quota dès le GET (bloquer accès au formulaire)
    q = t.quota_employes_info
    if q["max"] and q["plein"]:
        flash(
            f"Limite atteinte — Plan « {t.plan.nom} » : {q['max']} employé(s) maximum "
            f"({q['salaries']} salarié(s) + {q['journaliers']} journalier(s)). "
            f"Passez au plan supérieur.", "error"
        )
        return redirect(url_for("tenant.salaries"))
    if request.method=="POST":
        if not t.peut_ajouter_employe:
            flash(f"Limite atteinte ({t.plan.max_salaries} employés). Passez au plan supérieur.","error")
            return redirect(url_for("tenant.salaries"))
        s=Salarie(tenant_id=t.id,
            matricule=request.form["matricule"].strip().upper(),
            categorie_id=request.form.get("categorie_id") or None,
            nom=request.form["nom"].strip().upper(), prenom=request.form["prenom"].strip(),
            telephone=request.form.get("telephone"), email=request.form.get("email","").strip() or None,
            nationalite=request.form.get("nationalite","GABONAISE"),
            sexe=request.form.get("sexe"),
            date_naissance=_pd(request.form.get("date_naissance")),
            date_embauche=_pd(request.form["date_embauche"]),
            situation_matrimoniale=request.form.get("situation_matrimoniale"),
            nb_enfants=int(request.form.get("nb_enfants") or 0),
            nb_enfants_moins_16ans=int(request.form.get("nb_enfants_moins_16ans") or 0),
            nombre_parts=float(request.form.get("nombre_parts") or 1),
            numero_cnss=request.form.get("numero_cnss"), numero_cnamgs=request.form.get("numero_cnamgs"),
            emploi=request.form.get("emploi"), assujetti_cnamgs=request.form.get("assujetti_cnamgs")=="OUI", statut="ACTIF")
        db.session.add(s)
        sb=float(request.form.get("salaire_base") or 0)
        if sb: db.session.add(Contrat(tenant_id=t.id,salarie=s,type_contrat=request.form.get("type_contrat","CDI"),date_debut=s.date_embauche,salaire_base=sb,poste=s.emploi,actif=True))
        db.session.flush()
        log_action("CREATE", "salarie", s.id,
                   f"Nouveau salarié : {s.nom_complet} (matricule {s.matricule})",
                   apres={"nom": s.nom, "prenom": s.prenom, "matricule": s.matricule,
                          "emploi": s.emploi, "salaire_base": sb})
        db.session.commit(); flash(f"Salarié {s.nom_complet} créé.","success")
        return redirect(url_for("tenant.salarie_detail",id=s.id))
    sites = Site.query.filter_by(tenant_id=t.id, actif=True).order_by(Site.nom).all()
    return render_template("tenant/salarie_form.html", salarie=None, categories=cats,
        action="nouveau", tenant=t, sites=sites, aff_actuelle=None,
        grille_salaires=_grille_tenant(t))

@bp.route("/journaliers/<int:id>/convertir", methods=["GET", "POST"])
@login_required
def journalier_convertir(id):
    """Transforme un journalier en salarié (mensuel).

    Principe : on CRÉE un nouveau salarié à partir des données du journalier, on
    ARCHIVE le journalier (statut CONVERTI) pour préserver tout son historique de
    paie journalière (pointages/avances/paies passés restent attachés au
    journalier), et on reporte son affectation de site active. L'opération est
    neutre pour le quota (−1 journalier actif, +1 salarié actif)."""
    if current_user.is_super_admin:
        return redirect(url_for("admin.admin_dashboard"))
    t = get_tenant()
    if not t:
        return redirect(url_for("auth.login"))
    if not current_user.can_edit:
        abort(403)
    j = Journalier.query.filter_by(id=id, tenant_id=t.id).first_or_404()
    if (j.statut or "").upper() != "ACTIF":
        flash("Ce journalier n'est pas actif — il a peut-être déjà été converti.", "error")
        return redirect(url_for("tenant.journalier_detail", id=j.id))
    cats = CategorieEmploi.query.filter_by(tenant_id=t.id).all()

    # Matricule suggéré (préfixe sur le nom + séquence, garanti unique).
    import re as _re
    base = (_re.sub(r"[^A-Z]", "", (j.nom or "").upper())[:4]) or "SAL"
    _n = 1
    while Salarie.query.filter_by(tenant_id=t.id, matricule=f"{base}{_n:03d}").first():
        _n += 1
    suggestion = f"{base}{_n:03d}"

    if request.method == "POST":
        matricule = (request.form.get("matricule") or "").strip().upper()
        if not matricule:
            flash("Le matricule est obligatoire.", "error")
            return redirect(url_for("tenant.journalier_convertir", id=j.id))
        if Salarie.query.filter_by(tenant_id=t.id, matricule=matricule).first():
            flash(f"Le matricule « {matricule} » existe déjà. Choisissez-en un autre.", "error")
            return redirect(url_for("tenant.journalier_convertir", id=j.id))
        sb = float(request.form.get("salaire_base") or 0)

        s = Salarie(
            tenant_id=t.id, matricule=matricule,
            categorie_id=request.form.get("categorie_id") or None,
            nom=(request.form.get("nom") or j.nom).strip().upper(),
            prenom=(request.form.get("prenom") or j.prenom).strip(),
            telephone=request.form.get("telephone") or j.telephone,
            email=(request.form.get("email") or "").strip() or None,
            nationalite=request.form.get("nationalite") or j.nationalite or "GABONAISE",
            sexe=request.form.get("sexe"),
            date_naissance=_pd(request.form.get("date_naissance")),
            date_embauche=_pd(request.form.get("date_embauche")) or j.date_embauche or date.today(),
            situation_matrimoniale=request.form.get("situation_matrimoniale"),
            nb_enfants=int(request.form.get("nb_enfants") or 0),
            nb_enfants_moins_16ans=int(request.form.get("nb_enfants_moins_16ans") or 0),
            nombre_parts=float(request.form.get("nombre_parts") or 1),
            numero_cnss=request.form.get("numero_cnss"),
            numero_cnamgs=request.form.get("numero_cnamgs"),
            emploi=request.form.get("emploi") or j.profession,
            assujetti_cnamgs=request.form.get("assujetti_cnamgs") == "OUI",
            statut="ACTIF",
        )
        db.session.add(s)
        if sb:
            db.session.add(Contrat(
                tenant_id=t.id, salarie=s,
                type_contrat=request.form.get("type_contrat", "CDI"),
                date_debut=s.date_embauche, salaire_base=sb, poste=s.emploi, actif=True))
        # Archiver le journalier (historique préservé, retiré des listes actives).
        j.statut = "CONVERTI"
        db.session.flush()
        # Reporter l'affectation de site active vers le nouveau salarié.
        aff = AffectationSite.query.filter_by(
            tenant_id=t.id, journalier_id=j.id, actif=True).first()
        if aff:
            db.session.add(AffectationSite(
                tenant_id=t.id, site_id=aff.site_id, salarie_id=s.id, actif=True))
            aff.actif = False
        log_action("UPDATE", "journalier", j.id,
                   f"Journalier {j.nom_complet} converti en salarié (matricule {s.matricule})")
        log_action("CREATE", "salarie", s.id,
                   f"Salarié issu de la conversion du journalier {j.nom_complet}",
                   apres={"nom": s.nom, "prenom": s.prenom, "matricule": s.matricule,
                          "salaire_base": sb, "origine": "journalier"})
        db.session.commit()
        flash(f"{s.nom_complet} est désormais salarié (matricule {s.matricule}). "
              f"L'historique du journalier est conservé.", "success")
        return redirect(url_for("tenant.salarie_detail", id=s.id))

    return render_template("tenant/journalier_convertir.html",
                           journalier=j, categories=cats, tenant=t, suggestion=suggestion)


@bp.route("/salaries/<int:id>")
@login_required
def salarie_detail(id):
    if current_user.is_super_admin: return redirect(url_for("admin.admin_dashboard"))
    t = get_tenant()
    if not t: return redirect(url_for("auth.login"))
    s = Salarie.query.filter_by(id=id, tenant_id=t.id).first_or_404()
    bulletins = BulletinPaie.query.filter_by(salarie_id=id, tenant_id=t.id)        .order_by(BulletinPaie.date_creation.desc()).all()
    contrat = Contrat.query.filter_by(salarie_id=id, tenant_id=t.id, actif=True).first()
    conge   = Conge.query.filter_by(salarie_id=id, tenant_id=t.id,
                                     annee=datetime.now().year).first()
    total_brut = sum(float(b.salaire_brut or 0) for b in bulletins)
    total_net  = sum(float(b.net_a_payer  or 0) for b in bulletins)
    total_cnss = sum(float(b.cnss_salarie or 0) for b in bulletins)
    total_irpp = sum(float(b.irpp         or 0) for b in bulletins)
    nb_mois    = len(bulletins)
    anciennete_jours = (datetime.now().date() - s.date_embauche).days if s.date_embauche else 0
    anciennete_ans   = anciennete_jours // 365
    anciennete_mois  = (anciennete_jours % 365) // 30

    # ── Historique des pointages (30 derniers jours) ──────────────────────────
    nb_jours = request.args.get("nb_jours", type=int, default=30)
    nb_jours = min(max(nb_jours, 7), 90)          # borne 7-90 jours
    date_fin   = datetime.now().date()
    date_debut = date_fin - timedelta(days=nb_jours - 1)

    pts_hist = Pointage.query.filter_by(tenant_id=t.id, salarie_id=id)        .filter(Pointage.date_pointage >= date_debut,
                Pointage.date_pointage <= date_fin)        .order_by(Pointage.date_pointage.desc()).all()

    # Stats synthèse
    nb_presences  = sum(1 for p in pts_hist if p.present)
    nb_absences   = sum(1 for p in pts_hist if p.absent)
    nb_non_pointes = nb_jours - len(pts_hist)
    h_normales_tot = round(sum(float(p.heures_normales or 0) for p in pts_hist if p.present), 1)
    h_sup_tot      = round(sum(
        float(p.heures_sup_10 or 0) + float(p.heures_sup_30 or 0) +
        float(p.heures_sup_40 or 0) + float(p.heures_sup_70 or 0)
        for p in pts_hist if p.present), 1)
    taux_presence = round(nb_presences / (nb_presences + nb_absences) * 100
                          ) if (nb_presences + nb_absences) > 0 else 0

    return render_template("tenant/salarie_detail.html",
        salarie=s, tenant=t, bulletins=bulletins, contrat=contrat, conge=conge,
        total_brut=total_brut, total_net=total_net, total_cnss=total_cnss,
        total_irpp=total_irpp, nb_mois=nb_mois,
        anciennete_ans=anciennete_ans, anciennete_mois=anciennete_mois,
        # Historique pointage
        pts_hist=pts_hist, nb_jours=nb_jours,
        nb_presences=nb_presences, nb_absences=nb_absences,
        nb_non_pointes=nb_non_pointes,
        h_normales_tot=h_normales_tot, h_sup_tot=h_sup_tot,
        taux_presence=taux_presence,
        date_debut_hist=date_debut, date_fin_hist=date_fin)

@bp.route("/salaries/<int:id>/modifier", methods=["GET","POST"])
@login_required
def salarie_modifier(id):
    if current_user.is_super_admin: return redirect(url_for("admin.admin_dashboard"))
    t=get_tenant()
    if not t: return redirect(url_for("auth.login"))
    if not current_user.can_edit: abort(403)
    s = Salarie.query.filter_by(id=id, tenant_id=t.id).first_or_404()
    cats = CategorieEmploi.query.filter_by(tenant_id=t.id).all()
    if request.method=="POST":
        for f,v in [("nom",request.form["nom"].strip().upper()),("prenom",request.form["prenom"].strip()),
            ("telephone",request.form.get("telephone")),
            ("email",request.form.get("email","").strip() or None),
            ("nationalite",request.form.get("nationalite")),
            ("sexe",request.form.get("sexe")),("date_naissance",_pd(request.form.get("date_naissance"))),
            ("situation_matrimoniale",request.form.get("situation_matrimoniale")),
            ("nb_enfants",int(request.form.get("nb_enfants") or 0)),
            ("nombre_parts",calculer_parts_irpp(request.form.get("situation_matrimoniale",""),int(request.form.get("nb_enfants",0) or 0))),
            ("numero_cnss",request.form.get("numero_cnss")),("numero_cnamgs",request.form.get("numero_cnamgs")),
            ("emploi",request.form.get("emploi")),("categorie_id",request.form.get("categorie_id") or None),
            ("mode_paiement",(request.form.get("mode_paiement","ESPECES") or "ESPECES").strip()),
            ("statut",request.form.get("statut","ACTIF")),("date_modification",utcnow())]:
            setattr(s,f,v)
        db.session.commit()
        # ── Affectation site ──────────────────────────────────────────────
        site_id = request.form.get("site_id", type=int)
        if site_id:
            aff_prev = AffectationSite.query.filter_by(
                salarie_id=s.id, tenant_id=t.id, actif=True).first()
            if aff_prev and aff_prev.site_id != site_id:
                aff_prev.actif    = False
                aff_prev.date_fin = date.today()
                aff_prev.motif    = "Réaffecté via formulaire salarié"
            if not aff_prev or aff_prev.site_id != site_id:
                db.session.add(AffectationSite(
                    tenant_id=t.id, site_id=site_id, salarie_id=s.id,
                    date_debut=date.today(), actif=True,
                    cree_par=current_user.email))
            db.session.commit()
        elif request.form.get("retirer_site"):
            aff = AffectationSite.query.filter_by(
                salarie_id=s.id, tenant_id=t.id, actif=True).first()
            if aff:
                aff.actif    = False
                aff.date_fin = date.today()
                aff.motif    = "Retiré via formulaire salarié"
                db.session.commit()
        flash("Fiche mise à jour.", "success")
        log_action("UPDATE", "salarie", s.id, f"Modification fiche salarié {s.nom_complet}")
        db.session.commit()
        return redirect(url_for("tenant.salarie_detail", id=s.id))
    # Récupérer site actuel + liste des sites
    aff_actuelle = AffectationSite.query.filter_by(
        salarie_id=id, tenant_id=t.id, actif=True).first()
    sites = Site.query.filter_by(tenant_id=t.id, actif=True).order_by(Site.nom).all()
    return render_template("tenant/salarie_form.html", salarie=s, categories=cats,
        action="modifier", tenant=t, sites=sites, aff_actuelle=aff_actuelle,
        grille_salaires=_grille_tenant(t))


# ══════════════════════════════════════════════════════════════════════════════
# GESTION DES CONTRATS SALARIÉS
# ══════════════════════════════════════════════════════════════════════════════

@bp.route("/salaries/<int:sal_id>/contrats")
@login_required
@tenant_required
def contrats_salarie(sal_id):
    """Liste de tous les contrats d'un salarié avec historique."""
    t = get_tenant()
    s = Salarie.query.filter_by(id=sal_id, tenant_id=t.id).first_or_404()
    contrats = Contrat.query.filter_by(
        salarie_id=sal_id, tenant_id=t.id
    ).order_by(Contrat.date_debut.desc()).all()
    cats = CategorieEmploi.query.filter_by(tenant_id=t.id).order_by(CategorieEmploi.code).all()
    return render_template("tenant/contrats_salarie.html",
                           salarie=s, contrats=contrats, tenant=t,
                           categories=cats)


@bp.route("/salaries/<int:sal_id>/contrats/nouveau", methods=["GET","POST"])
@login_required
@tenant_required
@can_edit
def contrat_nouveau(sal_id):
    """Créer un nouveau contrat pour un salarié."""
    t = get_tenant()
    s = Salarie.query.filter_by(id=sal_id, tenant_id=t.id).first_or_404()
    cats = CategorieEmploi.query.filter_by(tenant_id=t.id).all()

    if request.method == "POST":
        type_c  = request.form.get("type_contrat", "CDI")
        salaire = float(request.form.get("salaire_base", 0) or 0)
        poste   = request.form.get("poste", "").strip() or s.emploi
        cat_id  = request.form.get("categorie_id", type=int)

        try:
            date_debut = datetime.strptime(request.form.get("date_debut",""), "%Y-%m-%d").date()
        except ValueError:
            flash("Date de début invalide.", "error")
            return render_template("tenant/contrat_form.html",
                                   salarie=s, tenant=t, categories=cats, contrat=None)

        date_fin = None
        if request.form.get("date_fin"):
            try:
                date_fin = datetime.strptime(request.form.get("date_fin"), "%Y-%m-%d").date()
            except ValueError:
                flash("Date de fin invalide.", "error")
                return render_template("tenant/contrat_form.html",
                                       salarie=s, tenant=t, categories=cats, contrat=None)

        if not salaire or salaire <= 0:
            flash("Le salaire de base doit être positif.", "error")
            return render_template("tenant/contrat_form.html",
                                   salarie=s, tenant=t, categories=cats, contrat=None)

        # Désactiver l'ancien contrat actif
        Contrat.query.filter_by(
            salarie_id=sal_id, tenant_id=t.id, actif=True
        ).update({"actif": False})

        # Créer le nouveau contrat
        c = Contrat(
            tenant_id    = t.id,
            salarie_id   = sal_id,
            type_contrat = type_c,
            date_debut   = date_debut,
            date_fin     = date_fin,
            salaire_base = salaire,
            poste        = poste,
            categorie_id = cat_id,
            actif        = True,
        )
        # Mettre à jour le salarié
        s.emploi = poste
        if cat_id:
            s.categorie_id = cat_id

        db.session.add(c)
        db.session.flush()
        log_action("CREATE", "contrat", c.id,
                   f"Nouveau contrat {type_c} pour {s.nom_complet} — "
                   f"salaire {int(salaire):,} FCFA à partir du {date_debut.strftime('%d/%m/%Y')}")
        db.session.commit()
        _cache_delete(f"{t.id}:")
        flash(f"Contrat {type_c} créé pour {s.nom_complet}.", "success")
        return redirect(url_for("tenant.contrats_salarie", sal_id=sal_id))

    return render_template("tenant/contrat_form.html",
                           salarie=s, tenant=t, categories=cats, contrat=None)


@bp.route("/contrats/<int:id>/modifier", methods=["GET","POST"])
@login_required
@tenant_required
@can_edit
def contrat_modifier(id):
    """Modifier un contrat existant."""
    t = get_tenant()
    c = Contrat.query.filter_by(id=id, tenant_id=t.id).first_or_404()
    s = c.salarie
    cats = CategorieEmploi.query.filter_by(tenant_id=t.id).all()

    if request.method == "POST":
        avant = c.to_dict()
        c.type_contrat = request.form.get("type_contrat", c.type_contrat)
        c.poste        = request.form.get("poste", c.poste or "").strip()
        c.categorie_id = request.form.get("categorie_id", type=int) or c.categorie_id

        try:
            c.salaire_base = float(request.form.get("salaire_base", c.salaire_base) or 0)
        except ValueError:
            pass

        if request.form.get("date_fin"):
            try:
                c.date_fin = datetime.strptime(request.form.get("date_fin"), "%Y-%m-%d").date()
            except ValueError:
                flash("Date de fin invalide.", "error")
                return render_template("tenant/contrat_form.html",
                                       salarie=s, tenant=t, categories=cats, contrat=c)
        else:
            c.date_fin = None

        db.session.flush()
        log_action("UPDATE", "contrat", c.id,
                   f"Modification contrat {s.nom_complet}",
                   avant=avant, apres=c.to_dict())
        db.session.commit()
        flash("Contrat mis à jour.", "success")
        return redirect(url_for("tenant.contrats_salarie", sal_id=s.id))

    return render_template("tenant/contrat_form.html",
                           salarie=s, tenant=t, categories=cats, contrat=c)


@bp.route("/contrats/<int:id>/terminer", methods=["POST"])
@login_required
@tenant_required
@can_edit
def contrat_terminer(id):
    """Marquer un contrat comme terminé (date de fin = aujourd'hui)."""
    t  = get_tenant()
    c  = Contrat.query.filter_by(id=id, tenant_id=t.id).first_or_404()
    s  = c.salarie
    motif = request.form.get("motif", "").strip()

    from datetime import date as _date
    c.date_fin = _date.today()
    c.actif    = False

    log_action("UPDATE", "contrat", c.id,
               f"Fin de contrat {s.nom_complet} — {motif or 'Non précisé'}")
    db.session.commit()
    flash(f"Contrat de {s.nom_complet} terminé.", "success")
    return redirect(url_for("tenant.contrats_salarie", sal_id=s.id))


@bp.route("/contrats/<int:id>/supprimer", methods=["POST"])
@login_required
@tenant_required
@can_edit
def contrat_supprimer(id):
    """Supprimer un contrat (uniquement si inactif ou aucun bulletin associé)."""
    t = get_tenant()
    c = Contrat.query.filter_by(id=id, tenant_id=t.id).first_or_404()
    s = c.salarie

    if c.actif:
        flash("Impossible de supprimer le contrat actif. Terminez-le d'abord.", "error")
        return redirect(url_for("tenant.contrats_salarie", sal_id=s.id))

    log_action("DELETE", "contrat", c.id,
               f"Suppression contrat {c.type_contrat} de {s.nom_complet} "
               f"(du {c.date_debut} au {c.date_fin or 'en cours'})")
    db.session.delete(c)
    db.session.commit()
    flash("Contrat supprimé.", "success")
    return redirect(url_for("tenant.contrats_salarie", sal_id=s.id))


@bp.route("/salaries/<int:id>/supprimer", methods=["POST"])
@login_required
def salarie_supprimer(id):
    if current_user.is_super_admin: return redirect(url_for("admin.admin_dashboard"))
    t = get_tenant()
    if not t: return redirect(url_for("auth.login"))
    if not current_user.is_tenant_admin:
        flash("Seul l'administrateur peut supprimer un salarié.", "error")
        return redirect(url_for("tenant.salaries"))
    s = Salarie.query.filter_by(id=id, tenant_id=t.id).first_or_404()
    nom = s.nom_complet
    bulletins_actifs = BulletinPaie.query.filter_by(salarie_id=id).filter(
        BulletinPaie.statut.in_(["VALIDÉ","PAYÉ"])).count()
    if bulletins_actifs > 0:
        flash(f"Impossible de supprimer {nom} : {bulletins_actifs} bulletin(s) validé(s). Passez-le en INACTIF.", "error")
        return redirect(url_for("tenant.salarie_detail", id=id))
    try:
        BulletinPaie.query.filter_by(salarie_id=id).delete()
        Contrat.query.filter_by(salarie_id=id).delete()
        Pointage.query.filter_by(salarie_id=id).delete()
        Acompte.query.filter_by(salarie_id=id).delete()
        Conge.query.filter_by(salarie_id=id).delete()
        db.session.delete(s); db.session.commit()
        log_action("DELETE", "salarie", id, f"Suppression salarié {nom}")
        db.session.commit()
        flash(f"Salarié {nom} supprimé.", "success")
    except Exception as e:
        db.session.rollback(); flash(f"Erreur: {str(e)}", "error")
        return redirect(url_for("tenant.salarie_detail", id=id))
    return redirect(url_for("tenant.salaries"))

# ── Bulletins ─────────────────────────────────────────────────────────────────
@bp.route("/bulletins")
@login_required
def bulletins():
    if current_user.is_super_admin: return redirect(url_for("admin.admin_dashboard"))
    t=get_tenant()
    if not t: return redirect(url_for("auth.login"))
    pid          = request.args.get("periode_id", type=int)
    sf           = request.args.get("statut", "")
    site_filtre_id = request.args.get("site_id", type=int)
    periodes     = PeriodePaie.query.filter_by(tenant_id=t.id)                    .order_by(PeriodePaie.annee.desc(), PeriodePaie.mois.desc()).all()
    sites_list   = Site.query.filter_by(tenant_id=t.id, actif=True).order_by(Site.nom).all()
    site_filtre  = Site.query.filter_by(id=site_filtre_id, tenant_id=t.id).first() if site_filtre_id else None
    ps = None; buls = []; masse = {}; pagination = None

    if pid:
        ps = PeriodePaie.query.filter_by(id=pid, tenant_id=t.id).first_or_404()
        q = BulletinPaie.query.options(
            joinedload(BulletinPaie.salarie),
            joinedload(BulletinPaie.periode),
        ).filter_by(tenant_id=t.id, periode_id=pid)
        if sf:
            q = q.filter_by(statut=sf)

        # ── Filtre par site ──────────────────────────────────────────────────
        if site_filtre_id:
            # Récupérer les IDs des salariés affectés à ce site
            ids_sal = [a.salarie_id for a in AffectationSite.query.filter_by(
                tenant_id=t.id, site_id=site_filtre_id, actif=True
            ).filter(AffectationSite.salarie_id.isnot(None)).all()]
            q = q.filter(BulletinPaie.salarie_id.in_(ids_sal))

        page_bul   = request.args.get("page", 1, type=int)
        # Une seule query base avec le join, puis on pagine
        q_joined   = q.join(Salarie).order_by(Salarie.nom)
        buls_tous  = q_joined.all()
        masse      = calculer_masse_salariale(buls_tous)
        pagination = q_joined.paginate(page=page_bul, per_page=25, error_out=False)
        buls       = pagination.items

    # Affectation site de chaque salarié pour affichage dans le tableau
    aff_sal = {a.salarie_id: a.site for a in AffectationSite.query.filter_by(
        tenant_id=t.id, actif=True
    ).filter(AffectationSite.salarie_id.isnot(None)).all()}

    _args = {k: v for k, v in request.args.items() if k != 'page'}
    _base = request.path + '?' + '&'.join(f'{k}={v}' for k, v in _args.items())
    _sep  = '&' if _args else '?'
    return render_template("tenant/bulletins.html",
        periodes=periodes, periode_sel=ps,
        bulletins=buls, masse=masse, statut_filtre=sf,
        sites=sites_list, site_filtre=site_filtre, aff_sal=aff_sal,
        pagination=pagination if pid else None,
        pagination_base=_base + _sep,
        tenant=t)

@bp.route("/bulletins/bordereau")
@login_required
def bulletins_bordereau():
    """Bordereau de paie imprimable d'une période : liste des salariés, net à
    payer, mode de paiement, avec sous-totaux Espèces / Virement.
    """
    if current_user.is_super_admin:
        return redirect(url_for("admin.admin_dashboard"))
    t = get_tenant()
    if not t:
        return redirect(url_for("auth.login"))

    pid = request.args.get("periode_id", type=int)
    if not pid:
        flash("Choisissez une période pour générer le bordereau.", "error")
        return redirect(url_for("tenant.bulletins"))
    ps = PeriodePaie.query.filter_by(id=pid, tenant_id=t.id).first_or_404()

    statut_f = request.args.get("statut", "")
    q = BulletinPaie.query.options(joinedload(BulletinPaie.salarie)) \
        .filter_by(tenant_id=t.id, periode_id=pid)
    if statut_f:
        q = q.filter_by(statut=statut_f)
    bulletins = q.join(Salarie).order_by(Salarie.nom).all()

    lignes = []
    recap_mode = {"ESPECES": {"total": 0.0, "nb": 0}, "VIREMENT": {"total": 0.0, "nb": 0}}
    total_general = 0.0
    for b in bulletins:
        net = float(b.net_a_payer or 0)
        mode = (b.mode_paiement if b.statut == "PAYÉ" and b.mode_paiement
                else (b.salarie.mode_paiement if b.salarie else "ESPECES")) or "ESPECES"
        if mode not in ("ESPECES", "VIREMENT"):
            mode = "ESPECES"
        recap_mode[mode]["total"] += net
        recap_mode[mode]["nb"]    += 1
        total_general += net
        lignes.append({
            "matricule": b.salarie.matricule if b.salarie else "—",
            "nom": b.salarie.nom_complet if b.salarie else "—",
            "emploi": (b.salarie.emploi if b.salarie else "") or "—",
            "net": net, "mode": mode, "statut": b.statut,
        })

    return render_template("tenant/bulletins_bordereau_print.html",
        tenant=t, periode=ps, lignes=lignes, recap_mode=recap_mode,
        total_general=total_general, nb_total=len(lignes),
        statut=statut_f, now=datetime.now())

@bp.route("/bulletins/generer-lot", methods=["POST"])
@login_required
def bulletins_generer_lot():
    """Génère un bulletin BROUILLON pour chaque salarié actif d'une période.

    Principes de sûreté :
      - ne crée QUE des brouillons (jamais de bulletin validé) ;
      - ne touche JAMAIS un bulletin existant (salarié ignoré, pas écrasé) ;
      - refuse une période clôturée ;
      - ignore les salariés sans contrat actif ou embauchés après la période.
    Le salaire de base vient du contrat actif ; la prime d'ancienneté et les
    acomptes en attente sont appliqués automatiquement. L'utilisateur ajuste
    ensuite les cas particuliers avant de valider.
    """
    if current_user.is_super_admin:
        return redirect(url_for("admin.admin_dashboard"))
    t = get_tenant()
    if not t:
        return redirect(url_for("auth.login"))
    if not current_user.can_edit:
        abort(403)

    pid = request.form.get("periode_id", type=int)
    if not pid:
        flash("Choisissez une période avant de générer les bulletins.", "error")
        return redirect(url_for("tenant.bulletins"))
    periode = PeriodePaie.query.filter_by(id=pid, tenant_id=t.id).first_or_404()
    retour = f"/bulletins?periode_id={pid}"

    if periode.statut not in ("OUVERT", "OUVERTE"):
        flash(f"La période {periode.libelle_mois} {periode.annee} est "
              f"{periode.statut.lower()} : aucun bulletin ne peut y être généré.", "error")
        return redirect(retour)

    import calendar as _cal
    fin_periode = date(periode.annee, periode.mois,
                       _cal.monthrange(periode.annee, periode.mois)[1])

    # Salariés déjà dotés d'un bulletin sur cette période : intouchables.
    deja = {b.salarie_id for b in
            BulletinPaie.query.filter_by(tenant_id=t.id, periode_id=pid).all()}

    salaries = (Salarie.query.filter_by(tenant_id=t.id, statut="ACTIF")
                .order_by(Salarie.nom, Salarie.prenom).all())

    crees = 0
    ignores_existants = ignores_sans_contrat = ignores_non_embauches = 0
    noms_sans_contrat = []

    for s in salaries:
        if s.id in deja:
            ignores_existants += 1
            continue
        if s.date_embauche and s.date_embauche > fin_periode:
            ignores_non_embauches += 1
            continue
        contrat = (Contrat.query
                   .filter_by(salarie_id=s.id, tenant_id=t.id, actif=True)
                   .order_by(Contrat.date_debut.desc()).first())
        if not contrat or not contrat.salaire_base:
            ignores_sans_contrat += 1
            if len(noms_sans_contrat) < 8:
                noms_sans_contrat.append(s.nom_complet)
            continue

        donnees = {"salaire_base": float(contrat.salaire_base)}
        if s.date_embauche:
            donnees["anciennete_annees"] = max(
                0, (fin_periode - s.date_embauche).days // 365)

        # Acomptes en attente du mois : déduits comme dans la saisie unitaire.
        total_ac = float(db.session.query(db.func.sum(Acompte.montant))
                         .filter_by(tenant_id=t.id, salarie_id=s.id,
                                    mois=periode.mois, annee=periode.annee,
                                    statut="EN_ATTENTE").scalar() or 0)
        if total_ac > 0:
            donnees["acompte"] = total_ac

        res = calculer_bulletin(dict(donnees, convention=t.convention),
                                nb_parts=float(s.nombre_parts or 1))
        b = BulletinPaie(tenant_id=t.id, salarie_id=s.id, periode_id=pid)
        for k, v in res.items():
            if not k.startswith("_") and hasattr(b, k):
                setattr(b, k, v)
        b.statut = "BROUILLON"
        b.mode_paiement = s.mode_paiement or "ESPECES"
        db.session.add(b)
        crees += 1

    db.session.commit()
    log_action("GENERATE_BATCH", "bulletin", pid,
               f"Génération en lot {periode.libelle_mois} {periode.annee} : "
               f"{crees} brouillon(s) créé(s)")

    if crees:
        flash(f"{crees} bulletin(s) brouillon créé(s) pour "
              f"{periode.libelle_mois} {periode.annee}. Vérifiez-les avant validation.",
              "success")
    else:
        flash("Aucun bulletin créé : tous les salariés actifs en ont déjà un "
              "sur cette période, ou aucun n'a de contrat actif.", "warning")

    details = []
    if ignores_existants:
        details.append(f"{ignores_existants} salarié(s) avaient déjà un bulletin (inchangés)")
    if ignores_non_embauches:
        details.append(f"{ignores_non_embauches} embauché(s) après la période")
    if ignores_sans_contrat:
        noms = ", ".join(noms_sans_contrat)
        suite = "…" if ignores_sans_contrat > len(noms_sans_contrat) else ""
        details.append(f"{ignores_sans_contrat} sans contrat actif ({noms}{suite})")
    if details:
        flash(" · ".join(details), "info")

    return redirect(retour)


@bp.route("/bulletins/valider-lot", methods=["POST"])
@login_required
def bulletins_valider_lot():
    """Valider une sélection de bulletins ou tous les brouillons d'une période."""
    if current_user.is_super_admin: return redirect(url_for("admin.admin_dashboard"))
    t = get_tenant()
    if not t: return redirect(url_for("auth.login"))
    if not current_user.can_edit: abort(403)

    pid      = request.form.get("periode_id", type=int)
    site_id  = request.form.get("site_id",    type=int)
    action   = request.form.get("action_lot", "valider")
    ids_str  = request.form.get("bulletin_ids", "")
    ids_sel  = [int(x) for x in ids_str.split(",") if x.strip().isdigit()]

    if not pid:
        flash("Période manquante.", "error")
        return redirect(url_for("tenant.bulletins"))

    # ── CORRECTION BUG : si aucun ID sélectionné + action valider → refuser ──
    # Avant : ids_sel vide → validait TOUS les bulletins de la période
    # Maintenant : ids_sel vide → uniquement pour les actions "tout valider" explicites
    if not ids_sel and action == "valider":
        # Vérifier que c'est bien une demande "tout valider" (bouton dédié)
        tout_valider = request.form.get("tout_valider", "0")
        if tout_valider != "1":
            flash("Aucun bulletin sélectionné. Cochez des bulletins ou utilisez 'Tout valider'.", "warning")
            return redirect(f"/bulletins?periode_id={pid}" + (f"&site_id={site_id}" if site_id else ""))

    # Charger les bulletins ciblés
    if ids_sel:
        buls = BulletinPaie.query.filter(
            BulletinPaie.id.in_(ids_sel),
            BulletinPaie.tenant_id == t.id
        ).all()
    else:
        q = BulletinPaie.query.filter_by(tenant_id=t.id, periode_id=pid, statut="BROUILLON")
        if site_id:
            ids_sal = [a.salarie_id for a in AffectationSite.query.filter_by(
                tenant_id=t.id, site_id=site_id, actif=True
            ).filter(AffectationSite.salarie_id.isnot(None)).all()]
            q = q.filter(BulletinPaie.salarie_id.in_(ids_sal))
        buls = q.all()

    nb = 0
    if action == "valider":
        for b in buls:
            if b.statut != "BROUILLON": continue
            b.statut          = "VALIDÉ"
            b.date_validation = utcnow()
            attribuer_numero_bulletin(b)
            for a in Acompte.query.filter_by(
                tenant_id=t.id, salarie_id=b.salarie_id,
                mois=b.periode.mois, annee=b.periode.annee, statut="EN_ATTENTE").all():
                a.statut = "DEDUIT"
            nb += 1
        msg = f"✅ {nb} bulletin(s) validé(s)."
        log_action("VALIDATE", "bulletin", pid,
                   f"Validation de {nb} bulletin(s) — période {pid}")

    elif action == "annuler_validation":
        # ── NOUVEAU : annuler la validation → repasser en BROUILLON ──────────
        for b in buls:
            if b.statut not in ("VALIDÉ", "VALIDE"): continue
            b.statut          = "BROUILLON"
            b.date_validation = None
            # Remettre les acomptes déduits en attente
            for a in Acompte.query.filter_by(
                tenant_id=t.id, salarie_id=b.salarie_id,
                mois=b.periode.mois, annee=b.periode.annee, statut="DEDUIT").all():
                a.statut = "EN_ATTENTE"
            nb += 1
        msg = f"↩️ {nb} bulletin(s) remis en brouillon."
        log_action("CANCEL", "bulletin", pid,
                   f"Annulation validation de {nb} bulletin(s) — période {pid}")

    elif action == "payer":
        for b in buls:
            if b.statut not in ("VALIDÉ", "VALIDE", "BROUILLON"): continue
            b.statut = "PAYÉ"
            nb += 1
        msg = f"💰 {nb} bulletin(s) marqué(s) comme payé(s)."
        log_action("PAY", "bulletin", pid,
                   f"{nb} bulletin(s) marqué(s) payé(s) — période {pid}")

    elif action == "supprimer_brouillons":
        for b in buls:
            if b.statut != "BROUILLON": continue
            db.session.delete(b); nb += 1
        msg = f"🗑️ {nb} brouillon(s) supprimé(s)."
        log_action("DELETE", "bulletin", pid,
                   f"Suppression de {nb} brouillon(s) — période {pid}")

    else:
        flash("Action inconnue.", "error")
        return redirect(f"/bulletins?periode_id={pid}")

    db.session.commit()
    _cache_delete(f"{t.id}:")
    flash(msg, "success")
    redir = f"/bulletins?periode_id={pid}"
    if site_id: redir += f"&site_id={site_id}"
    return redirect(redir)

def attribuer_numero_bulletin(b):
    """
    Attribue un numéro séquentiel immuable au bulletin lors de sa validation.
    Format : BP-<année>-<séquence 6 chiffres>, continu et unique par tenant.
    Ne fait rien si le bulletin a déjà un numéro (immuabilité du document).
    """
    if b.numero:
        return
    annee = b.periode.annee if b.periode else utcnow().year
    dernier = (db.session.query(db.func.max(BulletinPaie.numero_seq))
               .filter(BulletinPaie.tenant_id == b.tenant_id).scalar()) or 0
    b.numero_seq = dernier + 1
    b.numero     = f"BP-{annee}-{b.numero_seq:06d}"


@bp.route("/bulletins/saisie", methods=["GET","POST"])
@login_required
def bulletin_saisie():
    if current_user.is_super_admin: return redirect(url_for("admin.admin_dashboard"))
    t=get_tenant()
    if not t: return redirect(url_for("auth.login"))
    if not current_user.can_edit: abort(403)
    sals=Salarie.query.filter_by(tenant_id=t.id,statut="ACTIF").order_by(Salarie.nom).all()
    pers=PeriodePaie.query.filter_by(tenant_id=t.id,statut="OUVERT").order_by(PeriodePaie.annee.desc(),PeriodePaie.mois.desc()).all()
    if request.method=="POST":
        sid=int(request.form["salarie_id"]); pid=int(request.form["periode_id"])
        s=Salarie.query.filter_by(id=sid,tenant_id=t.id).first_or_404()
        periode = PeriodePaie.query.filter_by(id=pid, tenant_id=t.id).first_or_404()
        acomptes_en_attente = Acompte.query.filter_by(
            tenant_id=t.id, salarie_id=sid,
            mois=periode.mois, annee=periode.annee, statut="EN_ATTENTE").all()
        total_acomptes = sum(float(a.montant) for a in acomptes_en_attente)
        donnees={}
        for k,v in request.form.items():
            # Exclure les champs non-numériques et les champs base_/taux_ (traités séparément)
            if k in ("salarie_id","periode_id","csrf_token","action","nb_jours_travailles"):
                continue
            if k.startswith("base_") or k.startswith("taux_") or k.startswith("composant_"):
                continue
            try:
                donnees[k] = float(v) if v else 0
            except (ValueError, TypeError):
                donnees[k] = 0
        if total_acomptes > 0:
            donnees["acompte"] = max(donnees.get("acompte", 0), total_acomptes)

        # ── Composants personnalisés du tenant (gains/retenues) ────────────────
        # Lecture des montants saisis (champ composant_<id>) ; on construit la
        # liste passée au calcul et on en garde une copie pour la persistance.
        composants_actifs = ComposantPaie.query.filter_by(tenant_id=t.id, actif=True).all()
        composants_saisis = []
        for comp in composants_actifs:
            montant = request.form.get(f"composant_{comp.id}", type=float) or 0
            if montant:
                composants_saisis.append({
                    "composant_id": comp.id, "libelle": comp.libelle, "sens": comp.sens,
                    "montant": montant, "soumis_cnss": comp.soumis_cnss,
                    "soumis_cnamgs": comp.soumis_cnamgs, "soumis_irpp": comp.soumis_irpp,
                    "base": request.form.get(f"base_composant_{comp.id}", type=float),
                    "taux": request.form.get(f"taux_composant_{comp.id}", type=float),
                })
        donnees["composants"] = composants_saisis

        # ── Prime d'ancienneté automatique (selon convention collective) ───────
        # L'ancienneté est calculée à la fin de la période de paie. Le calcul
        # effectif se fait dans calculer_bulletin() et ne s'applique QUE si
        # aucune prime n'a été saisie manuellement (l'override reste prioritaire).
        if s.date_embauche:
            import calendar as _cal2
            _fin_periode = date(periode.annee, periode.mois,
                                _cal2.monthrange(periode.annee, periode.mois)[1])
            _anc = max(0, (_fin_periode - s.date_embauche).days // 365)
            donnees["anciennete_annees"] = _anc
            if _anc >= 2 and not donnees.get("prime_anciennete"):
                from calculs_paie import prime_anciennete as _calc_pa
                _pa = _calc_pa(t.convention, float(donnees.get("salaire_base") or 0), _anc)
                if _pa > 0:
                    flash(
                        f"Prime d'ancienneté calculée automatiquement : "
                        f"{int(_pa):,} FCFA ({_anc} ans d'ancienneté, "
                        f"convention {t.convention}). Vous pouvez la modifier "
                        f"manuellement.".replace(",", " "),
                        "info")

        # ── L6 : Allocation de congé (Code du travail 2021, Art. 225) ──────────
        # Calcul automatique si un congé ANNUEL débute dans le mois de la période
        # et que l'utilisateur n'a pas saisi de valeur manuelle (l'override
        # manuel reste prioritaire). Versée avant le départ en congé (Art. 225).
        if not donnees.get("allocations_conge"):
            import calendar as _cal
            _debut_mois = date(periode.annee, periode.mois, 1)
            _fin_mois   = date(periode.annee, periode.mois,
                               _cal.monthrange(periode.annee, periode.mois)[1])
            jours_conge = sum(
                float(c.jours_pris or 0)
                for c in s.conges
                if (c.type_conge in ("ANNUEL", None))
                   and c.statut in ("APPROUVÉ", "APPROUVE", "PRIS")
                   and c.date_depart and _debut_mois <= c.date_depart <= _fin_mois
            )
            if jours_conge > 0:
                from conges_avance import allocation_conge
                bulletins_12 = (BulletinPaie.query
                    .filter_by(tenant_id=t.id, salarie_id=sid)
                    .filter(BulletinPaie.statut.in_(["VALIDÉ", "VALIDE", "PAYÉ"]),
                            BulletinPaie.periode_id != pid)
                    .order_by(BulletinPaie.date_creation.desc()).limit(12).all())
                alloc = allocation_conge(bulletins_12, jours_conge)
                if alloc > 0:
                    donnees["allocations_conge"] = alloc
                    flash(
                        f"Allocation de congé calculée automatiquement pour "
                        f"{jours_conge:.0f} jour(s) pris : {int(alloc):,} FCFA "
                        f"(moyenne des 12 derniers mois, Art. 225). "
                        f"Vous pouvez la modifier manuellement."
                        .replace(",", " "),
                        "info")

        res=calculer_bulletin(dict(donnees, convention=t.convention),nb_parts=float(s.nombre_parts or 1))
        ex=BulletinPaie.query.filter_by(tenant_id=t.id,salarie_id=sid,periode_id=pid).first()
        # 🔒 Immuabilité : un bulletin validé est un document de paie officiel.
        # Il ne peut pas être réécrit en silence — il faut d'abord annuler sa
        # validation (action tracée), sinon l'historique de paie serait modifiable.
        if ex and ex.statut in ("VALIDÉ", "VALIDE"):
            log_action("BLOCK_EDIT", "bulletin", ex.id,
                       "Tentative de modification d'un bulletin validé (bloquée)")
            flash("Ce bulletin est validé : il ne peut pas être modifié. "
                  "Annulez d'abord sa validation depuis la liste des bulletins, "
                  "puis ressaisissez-le.", "error")
            return redirect(url_for("tenant.bulletin_detail", id=ex.id))
        b=ex or BulletinPaie(tenant_id=t.id,salarie_id=sid,periode_id=pid)
        if not ex: db.session.add(b)
        for k,v in res.items():
            if not k.startswith("_") and hasattr(b,k): setattr(b,k,v)
        b.nb_jours_travailles=int(request.form.get("nb_jours_travailles") or 0)
        # ✅ Sauvegarder base et taux saisis manuellement pour chaque rubrique
        RUBRIQUES_BT = ["salaire_base","heures_sup_10","heures_sup_30","heures_sup_30b","heures_sup_40","heures_sup_70",
            "absences","sursalaire","prime_caisse","carburant","prime_anciennete",
            "indem_logement","indem_domesticite","indem_eau_electricite","indem_nourriture",
            "prime_transport","prime_responsabilite","prime_rendement","prime_assiduité",
            "prime_qualite","prime_performance","allocations_conge",
            "indem_compensatrice_conge","indem_services_rendus",
            "indem_compensatrice_preavis","indem_licenciement"]
        for r in RUBRIQUES_BT:
            base_val = request.form.get(f"base_{r}", "")
            taux_val = request.form.get(f"taux_{r}", "")
            if hasattr(b, f"base_{r}"):
                try: setattr(b, f"base_{r}", float(base_val) if base_val else None)
                except: pass
            if hasattr(b, f"taux_{r}"):
                setattr(b, f"taux_{r}", taux_val.strip()[:20] if taux_val else "")
        action=request.form.get("action","brouillon")
        if action=="valider":
            b.statut="VALIDÉ"; b.date_validation=utcnow()
            attribuer_numero_bulletin(b)
            for a in acomptes_en_attente: a.statut = "DEDUIT"
        else:
            b.statut="BROUILLON"
        # Persistance des composants personnalisés (instantané par bulletin)
        db.session.flush()  # garantit b.id pour un nouveau bulletin
        BulletinComposant.query.filter_by(bulletin_id=b.id).delete()
        for cs in composants_saisis:
            db.session.add(BulletinComposant(
                bulletin_id=b.id, composant_id=cs["composant_id"],
                libelle=cs["libelle"], sens=cs["sens"], montant=cs["montant"],
                base=cs.get("base"), taux=cs.get("taux"),
                soumis_cnss=cs["soumis_cnss"], soumis_cnamgs=cs["soumis_cnamgs"],
                soumis_irpp=cs["soumis_irpp"]))
        db.session.commit()
        if total_acomptes > 0:
            flash(f"Bulletin sauvegardé. Acompte de {int(total_acomptes):,} FCFA déduit automatiquement.".replace(",", " "), "success")
        else:
            flash(f"Bulletin {'validé' if b.statut=='VALIDÉ' else 'sauvegardé'}.","success")
        return redirect(url_for("tenant.bulletin_detail",id=b.id))
    sid=request.args.get("salarie_id",type=int)
    ss=Salarie.query.filter_by(id=sid,tenant_id=t.id).first() if sid else None
    c=Contrat.query.filter_by(salarie_id=sid,tenant_id=t.id,actif=True).first() if sid else None
    acomptes_attente = Acompte.query.filter_by(tenant_id=t.id, salarie_id=sid, statut="EN_ATTENTE").all() if sid else []
    total_acomptes = sum(float(a.montant) for a in acomptes_attente)
    composants_actifs = ComposantPaie.query.filter_by(tenant_id=t.id, actif=True).order_by(
        ComposantPaie.ordre, ComposantPaie.libelle).all()
    return render_template("tenant/bulletin_saisie.html", salaries=sals, periodes=pers, salarie_sel=ss, contrat=c, tenant=t,
        acomptes_attente=acomptes_attente, total_acomptes=total_acomptes,
        composants_actifs=composants_actifs)

@bp.route("/bulletins/<int:id>")
@login_required
def bulletin_detail(id):
    if current_user.is_super_admin: return redirect(url_for("admin.admin_dashboard"))
    t=get_tenant()
    if not t: return redirect(url_for("auth.login"))
    bulletin = BulletinPaie.query.filter_by(id=id,tenant_id=t.id).first_or_404()
    composants = BulletinComposant.query.filter_by(bulletin_id=bulletin.id).all()
    from datetime import date as _date
    return render_template("tenant/bulletin_detail.html",
        bulletin=bulletin, tenant=t, composants=composants,
        today=_date.today().isoformat())

@bp.route("/bulletins/<int:id>/valider", methods=["POST"])
@login_required
def bulletin_valider(id):
    if current_user.is_super_admin: return redirect(url_for("admin.admin_dashboard"))
    t = get_tenant()
    if not t: return redirect(url_for("auth.login"))
    b = BulletinPaie.query.filter_by(id=id, tenant_id=t.id).first_or_404()
    if b.statut == "VALIDÉ":
        flash("Ce bulletin est déjà validé.", "info")
        return redirect(url_for("tenant.bulletin_detail", id=id))
    acomptes = Acompte.query.filter_by(
        tenant_id=t.id, salarie_id=b.salarie_id,
        mois=b.periode.mois, annee=b.periode.annee, statut="EN_ATTENTE").all()
    for a in acomptes: a.statut = "DEDUIT"
    b.statut = "VALIDÉ"; b.date_validation = utcnow()
    attribuer_numero_bulletin(b)
    db.session.commit()
    log_action("VALIDATE", "bulletin", b.id,
               f"Validation bulletin {b.salarie.nom_complet if b.salarie else ''} "
               f"(net {float(b.net_a_payer or 0):,.0f} F)".replace(",", " "))
    db.session.commit()
    flash("Bulletin validé avec succès.", "success")
    return redirect(url_for("tenant.bulletin_detail", id=id))

@bp.route("/bulletins/<int:id>/payer", methods=["POST"])
@login_required
def bulletin_paye(id):
    if current_user.is_super_admin: return redirect(url_for("admin.admin_dashboard"))
    t = get_tenant()
    if not t: return redirect(url_for("auth.login"))
    b = BulletinPaie.query.filter_by(id=id, tenant_id=t.id).first_or_404()
    mode = (request.form.get("mode_paiement") or "").strip()
    if mode not in ("ESPECES", "VIREMENT"):
        mode = (b.salarie.mode_paiement if b.salarie else "ESPECES") or "ESPECES"
    b.mode_paiement = mode
    # Date de paiement : celle choisie dans le formulaire, sinon aujourd'hui.
    from datetime import date as _date, datetime as _dt
    date_pmt = _date.today()
    _df = (request.form.get("date_paiement") or "").strip()
    if _df:
        try:
            date_pmt = _dt.strptime(_df, "%Y-%m-%d").date()
        except ValueError:
            date_pmt = _date.today()
    b.date_paiement = date_pmt
    b.statut = "PAYÉ"; db.session.commit()
    log_action("PAY", "bulletin", b.id,
               f"Bulletin payé {b.salarie.nom_complet if b.salarie else ''}")
    db.session.commit()
    # Interconnexion Caisse : proposer la sortie correspondante (ne bloque jamais).
    try:
        from interco_caisse import proposer_ecriture
        nom = b.salarie.nom_complet if b.salarie else "salarié"
        periode = f"{b.periode.libelle_mois} {b.periode.annee}" if b.periode else ""
        proposer_ecriture(
            t, source_ref=f"bulletin-{b.id}",
            montant=float(b.net_a_payer or 0),
            motif=f"Salaire {nom} — {periode}".strip(" —"),
            compte_suggere="6611", date_operation=b.date_paiement)
    except Exception:
        pass
    flash("Bulletin marqué comme payé.", "success")
    return redirect(url_for("tenant.bulletin_detail", id=id))

@bp.route("/bulletins/<int:id>/supprimer", methods=["POST"])
@login_required
def bulletin_supprimer(id):
    if current_user.is_super_admin:
        b = BulletinPaie.query.get_or_404(id)
        salarie_id = b.salarie_id
        db.session.delete(b); db.session.commit()
        flash("Bulletin supprimé (super admin).", "success")
        return redirect(url_for("tenant.salarie_detail", id=salarie_id))
    t = get_tenant()
    if not t: return redirect(url_for("auth.login"))
    b = BulletinPaie.query.filter_by(id=id, tenant_id=t.id).first_or_404()
    if b.statut == "VALIDÉ":
        flash("Impossible de supprimer un bulletin validé.", "error")
        return redirect(url_for("tenant.bulletin_detail", id=id))
    db.session.delete(b); db.session.commit()
    flash("Bulletin supprimé.", "success")
    return redirect(url_for("tenant.bulletins"))

@bp.route("/bulletins/<int:id>/pdf")
@login_required
def bulletin_pdf(id):
    """Génère et retourne le bulletin en PDF téléchargeable."""
    if current_user.is_super_admin:
        b = BulletinPaie.query.get_or_404(id)
        t = b.salarie.tenant
    else:
        t = get_tenant()
        if not t: return redirect(url_for("auth.login"))
        b = BulletinPaie.query.filter_by(id=id, tenant_id=t.id).first_or_404()
    try:
        from pdf_bulletin import generer_bulletin_pdf
        pdf_bytes = generer_bulletin_pdf(b, t)
        nom_fichier = (
            f"bulletin_{b.salarie.nom}_{b.salarie.prenom}_{b.periode.annee}_{b.periode.mois:02d}.pdf"
            .replace(" ", "_")
        )
        from flask import Response
        return Response(
            pdf_bytes,
            mimetype="application/pdf",
            headers={
                "Content-Disposition": f'attachment; filename="{nom_fichier}"',
                "Content-Length": str(len(pdf_bytes)),
            }
        )
    except Exception as e:
        flash(f"Erreur génération PDF : {e}", "error")
        return redirect(url_for("tenant.bulletin_detail", id=id))


@bp.route("/bulletins/export-zip/<int:periode_id>")
@login_required
def bulletins_export_zip(periode_id):
    """Télécharge tous les bulletins d'une période dans un ZIP (un PDF par salarié)."""
    t = get_tenant()
    if not t:
        return redirect(url_for("auth.login"))
    p = PeriodePaie.query.filter_by(id=periode_id, tenant_id=t.id).first_or_404()
    bulletins = (BulletinPaie.query.filter_by(periode_id=periode_id, tenant_id=t.id)
                 .join(Salarie).order_by(Salarie.nom).all())
    if not bulletins:
        flash("Aucun bulletin à exporter pour cette période.", "error")
        return redirect(url_for("tenant.bulletins"))

    import zipfile
    from pdf_bulletin import generer_bulletin_pdf
    zip_buffer = io.BytesIO()
    erreurs = 0
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        for b in bulletins:
            try:
                pdf_bytes = generer_bulletin_pdf(b, t)
                nom = (f"{b.salarie.nom}_{b.salarie.prenom}_{b.periode.annee}_{b.periode.mois:02d}.pdf"
                       .replace(" ", "_"))
                zf.writestr(nom, pdf_bytes)
            except Exception as e:
                erreurs += 1
                logger.error(f"Erreur PDF bulletin {b.id} : {e}")
    zip_buffer.seek(0)
    data = zip_buffer.read()

    log_action("EXPORT", "bulletin", periode_id,
               f"Export ZIP {len(bulletins)} bulletins — {p.libelle_complet}",
               user_id=current_user.id, tenant_id=t.id)
    db.session.commit()

    nom_zip = f"bulletins_{p.libelle_mois}_{p.annee}_{t.slug}.zip".replace(" ", "_")
    from flask import Response
    return Response(data, mimetype="application/zip",
                    headers={"Content-Disposition": f'attachment; filename="{nom_zip}"',
                             "Content-Length": str(len(data))})


@bp.route("/bulletins/<int:id>/imprimer")
@login_required
def bulletin_imprimer(id):
    """Aperçu imprimable du bulletin (HTML), disponible dès le brouillon."""
    return _bulletin_imprimer_impl(id)


def _recap_sites_salarie_periode(t, salarie_id, annee, mois):
    """Heures travaillées par site pour un salarié sur un mois (multi-chantiers).

    Un salarié peut pointer sur plusieurs chantiers dans le mois. On agrège ses
    heures par site pour les afficher sur le bulletin. Renvoie (recap_sites,
    multi_sites). N'altère pas le calcul de paie : la rémunération couvre déjà
    l'ensemble des heures, tous sites confondus.
    """
    import calendar
    debut = date(annee, mois, 1)
    fin   = date(annee, mois, calendar.monthrange(annee, mois)[1])
    pts = (Pointage.query
           .filter_by(tenant_id=t.id, salarie_id=salarie_id)
           .filter(Pointage.date_pointage >= debut, Pointage.date_pointage <= fin,
                   Pointage.present == True, Pointage.absent == False)
           .options(joinedload(Pointage.site))
           .all())
    agg = {}
    for p in pts:
        h = (float(p.heures_normales or 0) + float(p.heures_sup or 0)
             + float(p.heures_sup_10 or 0) + float(p.heures_sup_30 or 0)
             + float(getattr(p, "heures_sup_30b", 0) or 0)
             + float(p.heures_sup_40 or 0) + float(p.heures_sup_70 or 0))
        nom = (p.site.nom if getattr(p, "site", None) else None) or "Non affecté"
        e = agg.setdefault(nom, {"nom": nom, "heures": 0.0, "jours": 0})
        e["heures"] += h
        e["jours"]  += 1
    recap = sorted(agg.values(), key=lambda x: -x["heures"])
    for e in recap:
        e["heures"] = round(e["heures"], 2)
    return recap, (len(recap) > 1)


def _bulletin_imprimer_impl(id):
    """Aperçu imprimable du bulletin (HTML), disponible dès le brouillon.

    Impression = consultation : accessible à tout utilisateur du tenant
    (pas de restriction can_edit), y compris pour un bulletin en BROUILLON.
    """
    if current_user.is_super_admin:
        b = BulletinPaie.query.get_or_404(id)
        t = b.salarie.tenant
    else:
        t = get_tenant()
        if not t:
            return redirect(url_for("auth.login"))
        b = BulletinPaie.query.filter_by(id=id, tenant_id=t.id).first_or_404()
    # Modèle de bulletin choisi par le tenant, avec repli sécurisé.
    try:
        modele = t.modele_bulletin or "classique"
    except Exception:
        modele = "classique"
    template_map = {
        "classique":   "tenant/bulletin_print.html",
        "moderne":     "tenant/bulletin_print_moderne.html",
        "minimaliste": "tenant/bulletin_print_minimaliste.html",
    }
    template = template_map.get(modele, "tenant/bulletin_print.html")
    # Vérifier que le template existe réellement sur le serveur, sinon repli.
    import os
    tpl_path = os.path.join(os.path.dirname(__file__), "..", "templates", template)
    if not os.path.exists(tpl_path):
        template = "tenant/bulletin_print.html"
    composants = BulletinComposant.query.filter_by(bulletin_id=b.id).all()
    # Répartition par site (multi-chantiers) — affichée seulement si > 1 site.
    from models import PeriodePaie
    per = db.session.get(PeriodePaie, b.periode_id)
    recap_sites, multi_sites = ([], False)
    if per is not None:
        recap_sites, multi_sites = _recap_sites_salarie_periode(t, b.salarie_id, per.annee, per.mois)
    return render_template(template, bulletin=b, tenant=t, composants=composants,
                           recap_sites=recap_sites, multi_sites=multi_sites)


# ✅ ENVOI EMAIL ASYNCHRONE — ne bloque plus le serveur
@bp.route("/bulletins/<int:id>/envoyer-email", methods=["POST"])
@login_required
def bulletin_envoyer_email(id):
    if current_user.is_super_admin: return redirect(url_for("admin.admin_dashboard"))
    t = get_tenant()
    if not t: return redirect(url_for("auth.login"))
    b = BulletinPaie.query.filter_by(id=id, tenant_id=t.id).first_or_404()
    s = b.salarie
    dest_email = request.form.get("email_dest", "").strip()
    if not dest_email and s.email:
        dest_email = s.email
    if not dest_email:
        flash(f"{s.nom_complet} n'a pas d'adresse email. Renseignez-en une dans le formulaire.", "error")
        return redirect(url_for("tenant.bulletin_detail", id=id))
    if not os.environ.get("MAIL_PASSWORD"):
        flash("Email non configuré sur le serveur : ajoutez la clé API Resend "
              "(variable MAIL_PASSWORD) dans les variables d'environnement.", "error")
        return redirect(url_for("tenant.bulletin_detail", id=id))
    try:
        corps = (f"Bonjour {s.prenom},\n\n"
                 f"Veuillez trouver votre bulletin de paie pour : {b.periode.libelle_complet}\n\n"
                 f"Salaire brut : {int(b.salaire_brut or 0)} FCFA\n"
                 f"Net a payer  : {int(b.net_a_payer or 0)} FCFA\n\n"
                 f"Cordialement,\n{t.denomination}")
        msg = Message(
            subject=f"Bulletin de paie {b.periode.libelle_complet} — {t.denomination}",
            recipients=[dest_email],
            body=corps,
            sender=current_app.config["MAIL_DEFAULT_SENDER"]
        )
        # Joindre le bulletin en PDF
        from pdf_bulletin import generer_bulletin_pdf
        pdf_bytes = generer_bulletin_pdf(b, t)
        nom_pdf = (f"bulletin_{s.nom}_{s.prenom}_"
                   f"{b.periode.annee}_{b.periode.mois:02d}.pdf")
        msg.attach(nom_pdf, "application/pdf", pdf_bytes)
        # ✅ Envoi dans un thread séparé → le serveur répond immédiatement
        send_email_async(current_app.extensions["mail"], msg)
        flash(f"Email en cours d'envoi à {dest_email}.", "success")
    except Exception as e:
        flash(f"Erreur préparation email: {str(e)}", "error")
    return redirect(url_for("tenant.bulletin_detail", id=id))

@bp.route("/bulletins/envoyer-tous", methods=["POST"])
@login_required
def bulletins_envoyer_tous():
    if current_user.is_super_admin: return redirect(url_for("admin.admin_dashboard"))
    t = get_tenant()
    if not t: return redirect(url_for("auth.login"))
    periode_id = request.form.get("periode_id", type=int)
    if not periode_id: flash("Période manquante.", "error"); return redirect(url_for("tenant.bulletins"))
    if not os.environ.get("MAIL_PASSWORD"):
        flash("Email non configuré sur le serveur : ajoutez la clé API Resend "
              "(variable MAIL_PASSWORD) dans les variables d'environnement.", "error")
        return redirect(url_for("tenant.bulletins"))
    buls = BulletinPaie.query.filter_by(tenant_id=t.id, periode_id=periode_id).all()
    nb_ok=0; nb_sans_email=0
    for b in buls:
        if not b.salarie.email: nb_sans_email+=1; continue
        try:
            corps = (f"Bonjour {b.salarie.prenom},\n\n"
                     f"Bulletin {b.periode.libelle_complet}\n"
                     f"Net a payer : {int(b.net_a_payer or 0)} FCFA\n\n"
                     f"Cordialement, {t.denomination}")
            msg = Message(subject=f"Bulletin {b.periode.libelle_complet}",
                recipients=[b.salarie.email], body=corps,
                sender=current_app.config["MAIL_DEFAULT_SENDER"])
            from pdf_bulletin import generer_bulletin_pdf
            pdf_bytes = generer_bulletin_pdf(b, t)
            nom_pdf = (f"bulletin_{b.salarie.nom}_{b.salarie.prenom}_"
                       f"{b.periode.annee}_{b.periode.mois:02d}.pdf")
            msg.attach(nom_pdf, "application/pdf", pdf_bytes)
            send_email_async(current_app.extensions["mail"], msg)
            nb_ok+=1
        except Exception as e:
            print(f"Erreur email {b.salarie.email}: {e}")
    flash(f"{nb_ok} email(s) en cours d'envoi. {nb_sans_email} salarié(s) sans email.", "success")
    return redirect(url_for("tenant.bulletins"))

# ── Périodes ──────────────────────────────────────────────────────────────────
@bp.route("/periodes")
@login_required
def periodes():
    if current_user.is_super_admin: return redirect(url_for("admin.admin_dashboard"))
    t=get_tenant()
    if not t: return redirect(url_for("auth.login"))
    periodes_liste = (PeriodePaie.query.filter_by(tenant_id=t.id)
                      .order_by(PeriodePaie.annee.desc(), PeriodePaie.mois.desc()).all())

    # Masse brute et nombre de bulletins par période (une seule requête groupée)
    stats = {}
    for pid, nb, brut, net in (db.session.query(
            BulletinPaie.periode_id,
            db.func.count(BulletinPaie.id),
            db.func.sum(BulletinPaie.salaire_brut),
            db.func.sum(BulletinPaie.net_a_payer))
            .filter_by(tenant_id=t.id)
            .group_by(BulletinPaie.periode_id).all()):
        stats[pid] = {"nb": nb or 0, "brut": float(brut or 0),
                      "net": float(net or 0), "brouillons": 0}

    # Brouillons restants : ce sont eux qui bloquent une clôture sereine
    for pid, nb in (db.session.query(
            BulletinPaie.periode_id, db.func.count(BulletinPaie.id))
            .filter_by(tenant_id=t.id, statut="BROUILLON")
            .group_by(BulletinPaie.periode_id).all()):
        if pid in stats:
            stats[pid]["brouillons"] = nb or 0

    return render_template("tenant/periodes.html", tenant=t,
        periodes=periodes_liste, stats_periodes=stats,
        now=datetime.now())

@bp.route("/periodes/<int:id>")
@login_required
def periode_detail(id):
    """Fiche récapitulative d'une période : masse salariale, retenues
    salariales, charges patronales et accès aux déclarations."""
    if current_user.is_super_admin:
        return redirect(url_for("admin.admin_dashboard"))
    t = get_tenant()
    if not t:
        return redirect(url_for("auth.login"))

    p = PeriodePaie.query.filter_by(id=id, tenant_id=t.id).first_or_404()
    bulletins = (BulletinPaie.query.options(joinedload(BulletinPaie.salarie))
                 .filter_by(tenant_id=t.id, periode_id=p.id).all())

    def somme(champ, source=None):
        return round(sum(float(getattr(b, champ) or 0) for b in (source or bulletins)), 2)

    # Répartition par statut
    par_statut = {"BROUILLON": 0, "VALIDÉ": 0, "PAYÉ": 0}
    for b in bulletins:
        st = "VALIDÉ" if b.statut in ("VALIDÉ", "VALIDE") else b.statut
        par_statut[st] = par_statut.get(st, 0) + 1

    # Masse salariale
    brut = somme("salaire_brut")
    net = somme("salaire_net")
    net_a_payer = somme("net_a_payer")

    # Retenues salariales
    retenues = {
        "cnss": somme("cnss_salarie"),
        "cnamgs": somme("cnamgs_salarie"),
        "tcs": somme("tcs"),
        "irpp": somme("irpp"),
    }
    retenues["total"] = round(sum(retenues.values()), 2)

    # Charges patronales
    patronales = {
        "cnss": somme("cnss_patronale"),
        "cnamgs": somme("cnamgs_patronale"),
        "fnh": somme("fnh"),
        "cfp": somme("cfp"),
    }
    patronales["total"] = round(sum(patronales.values()), 2)

    # Reversements par organisme (part salariale + part patronale)
    organismes = [
        {"nom": "CNSS", "salarial": retenues["cnss"], "patronal": patronales["cnss"],
         "total": round(retenues["cnss"] + patronales["cnss"], 2),
         "note": "Déclaration trimestrielle"},
        {"nom": "CNAMGS", "salarial": retenues["cnamgs"], "patronal": patronales["cnamgs"],
         "total": round(retenues["cnamgs"] + patronales["cnamgs"], 2),
         "note": "Déclaration trimestrielle"},
        {"nom": "TCS", "salarial": retenues["tcs"], "patronal": 0,
         "total": retenues["tcs"], "note": "Taxe complémentaire sur les salaires"},
        {"nom": "IRPP", "salarial": retenues["irpp"], "patronal": 0,
         "total": retenues["irpp"], "note": "Retenue à la source, reversée à la DGI"},
        {"nom": "FNH", "salarial": 0, "patronal": patronales["fnh"],
         "total": patronales["fnh"], "note": "Fonds national de l'habitat"},
        {"nom": "CFP", "salarial": 0, "patronal": patronales["cfp"],
         "total": patronales["cfp"], "note": "Contribution à la formation professionnelle"},
    ]
    total_reversements = round(sum(o["total"] for o in organismes), 2)

    stats = {
        "nb_bulletins": len(bulletins),
        "par_statut": par_statut,
        "nb_brouillons": par_statut.get("BROUILLON", 0),
        "brut": brut, "net": net, "net_a_payer": net_a_payer,
        "base_cnss": somme("base_cnss"),
        "base_cnamgs": somme("base_cnamgs"),
        "base_tcs": somme("base_tcs"),
        "base_irpp": somme("base_irpp"),
        "cout_employeur": round(brut + patronales["total"], 2),
    }

    return render_template("tenant/periode_detail.html", tenant=t, p=p,
        bulletins=bulletins, stats=stats, retenues=retenues,
        patronales=patronales, organismes=organismes,
        total_reversements=total_reversements, now=datetime.now())


@bp.route("/periodes/nouvelle", methods=["POST"])
@tenant_required
@can_edit
def periode_nouvelle():
    t=get_tenant(); annee=int(request.form["annee"]); mois=int(request.form["mois"])
    noms=PeriodePaie.MOIS_NOMS
    if PeriodePaie.query.filter_by(tenant_id=t.id,annee=annee,mois=mois).first(): flash("Période existante.","warning")
    else:
        db.session.add(PeriodePaie(tenant_id=t.id,annee=annee,mois=mois,libelle_mois=noms[mois],
            trimestre=f"T{(mois-1)//3+1}",statut="OUVERT",date_ouverture=utcnow()))
        db.session.commit(); flash(f"Période {noms[mois]} {annee} créée.","success")
    return redirect(url_for("tenant.periodes"))

@bp.route("/periodes/<int:id>/cloturer", methods=["POST"])
@tenant_required
@can_edit
def periode_cloturer(id):
    """Clôture une période, en refusant de le faire à l'aveugle s'il reste des
    brouillons : ce sont des salariés potentiellement non payés. La clôture
    reste possible, mais elle doit alors être confirmée explicitement."""
    t = get_tenant()
    p = PeriodePaie.query.filter_by(id=id, tenant_id=t.id).first_or_404()

    if p.statut != "OUVERT":
        flash("Cette période est déjà clôturée.", "warning")
        return redirect(url_for("tenant.periodes"))

    nb_brouillons = BulletinPaie.query.filter_by(
        tenant_id=t.id, periode_id=p.id, statut="BROUILLON").count()

    if nb_brouillons and request.form.get("confirmer") != "1":
        flash(f"Clôture annulée : {nb_brouillons} bulletin(s) encore en brouillon "
              f"sur {p.libelle_complet}. Validez-les d'abord, ou confirmez la "
              f"clôture depuis la fiche de la période si c'est volontaire.", "error")
        return redirect(url_for("tenant.periode_detail", id=p.id))

    p.statut = "CLÔTURÉ"
    p.date_cloture = utcnow()
    db.session.commit()
    log_action("CLOSE", "periode", p.id,
               f"Clôture de {p.libelle_complet}"
               + (f" avec {nb_brouillons} brouillon(s) restant(s)" if nb_brouillons else ""))

    if nb_brouillons:
        flash(f"Période {p.libelle_complet} clôturée avec {nb_brouillons} "
              f"brouillon(s) non validé(s). Rouvrez-la si vous devez les traiter.",
              "warning")
    else:
        flash(f"Période {p.libelle_complet} clôturée.", "success")
    return redirect(url_for("tenant.periodes"))


@bp.route("/periodes/<int:id>/rouvrir", methods=["POST"])
@tenant_required
@can_edit
def periode_rouvrir(id):
    """Rouvre une période clôturée. Une clôture par erreur reste ainsi
    rattrapable ; l'opération est tracée dans le journal d'audit."""
    t = get_tenant()
    p = PeriodePaie.query.filter_by(id=id, tenant_id=t.id).first_or_404()

    if p.statut == "OUVERT":
        flash("Cette période est déjà ouverte.", "info")
        return redirect(url_for("tenant.periode_detail", id=p.id))

    ancienne_cloture = p.date_cloture
    p.statut = "OUVERT"
    p.date_cloture = None
    db.session.commit()
    log_action("REOPEN", "periode", p.id,
               f"Réouverture de {p.libelle_complet}"
               + (f" (clôturée le {ancienne_cloture:%d/%m/%Y})" if ancienne_cloture else ""))
    flash(f"Période {p.libelle_complet} rouverte. Vous pouvez de nouveau y "
          f"générer et valider des bulletins.", "success")
    return redirect(url_for("tenant.periode_detail", id=p.id))

# ── Paiement abonnement ───────────────────────────────────────────────────────
@bp.route("/offres")
@login_required
def offres():
    """Page 'Nos offres' : le client compare les plans et en choisit un,
    ce qui le mène au paiement avec ce plan présélectionné."""
    if current_user.is_super_admin:
        return redirect(url_for("admin.admin_dashboard"))
    t = get_tenant()
    if not t:
        return redirect(url_for("auth.login"))
    plans = Plan.query.filter_by(actif=True).order_by(Plan.prix_mensuel).all()
    return render_template("tenant/offres.html", tenant=t, plans=plans)


@bp.route("/paiement")
@login_required
def paiement():
    if current_user.is_super_admin: return redirect(url_for("admin.admin_dashboard"))
    t = get_tenant()
    if not t: return redirect(url_for("auth.login"))
    plans = Plan.query.filter_by(actif=True).order_by(Plan.prix_mensuel).all()
    historique = Paiement.query.filter_by(tenant_id=t.id)\
        .order_by(Paiement.date_creation.desc()).limit(10).all()
    from coordonnees_paiement import AIRTEL_MONEY, BANQUE, CONTACT_PAIEMENT
    # Plan présélectionné (?plan=<id>) depuis la page "Nos offres" ; sinon plan actuel
    plan_choisi = request.args.get("plan", type=int) or t.plan_id
    return render_template("tenant/paiement.html", tenant=t, plans=plans,
                           historique=historique, plan_choisi=plan_choisi,
                           airtel=AIRTEL_MONEY, banque=BANQUE, contact=CONTACT_PAIEMENT)


# ── Airtel Money — Initiation ──────────────────────────────────────────────────
@bp.route("/paiement/airtel/initier", methods=["POST"])
@login_required
def paiement_airtel_initier():
    """
    Lance une demande de paiement STK Push Airtel Money.
    Le client reçoit une notification USSD sur son téléphone.
    """
    if current_user.is_super_admin: return redirect(url_for("admin.admin_dashboard"))
    t = get_tenant()
    if not t: return redirect(url_for("auth.login"))

    telephone  = request.form.get("telephone", "").strip()
    duree      = int(request.form.get("duree", 1) or 1)
    plan_id    = request.form.get("plan_id", type=int) or (t.plan_id)

    plan = db.session.get(Plan, plan_id) if plan_id else t.plan
    if not plan:
        flash("Plan introuvable.", "error")
        return redirect(url_for("tenant.paiement"))

    if not telephone:
        flash("Veuillez saisir votre numéro Airtel Money.", "error")
        return redirect(url_for("tenant.paiement"))

    montant = float(plan.prix_mensuel) * duree

    # Générer une référence unique
    import uuid
    reference = f"AM-{t.id}-{uuid.uuid4().hex[:10].upper()}"

    # Enregistrer la tentative en base
    p = Paiement(
        tenant_id=t.id,
        moyen="AIRTEL_MONEY",
        montant=montant,
        duree_mois=duree,
        plan_id=plan.id,
        reference_interne=reference,
        telephone=telephone,
        statut="EN_ATTENTE",
    )
    db.session.add(p)
    db.session.commit()

    # Appeler l'API Airtel
    try:
        from airtel_money import initier_paiement, AirtelConfigError
        resultat = initier_paiement(
            reference=reference,
            telephone=telephone,
            montant=montant,
            description=f"Abonnement PaieGabon {plan.nom} — {duree} mois",
        )
        p.reference_externe = resultat.get("transaction_id")
        import json
        p.reponse_raw = json.dumps(resultat.get("raw", {}))

        if resultat["success"]:
            db.session.commit()
            logger.info(f"[Paiement] Airtel initié — ref={reference} tenant={t.id}")
            flash(
                f"Demande de paiement envoyée sur le {telephone}. "
                "Validez sur votre téléphone dans les 2 minutes.",
                "success"
            )
            return redirect(url_for("tenant.paiement_airtel_attente", reference=reference))
        else:
            p.statut = "ECHEC"
            p.notes  = resultat["message"]
            db.session.commit()
            flash(f"Échec : {resultat['message']}", "error")
            return redirect(url_for("tenant.paiement"))

    except Exception as e:
        p.statut = "ECHEC"
        p.notes  = str(e)
        db.session.commit()
        logger.error(f"[Paiement] Erreur Airtel : {e}")
        flash(f"Erreur de connexion Airtel Money. Réessayez ou contactez le support.", "error")
        return redirect(url_for("tenant.paiement"))


# ── Airtel Money — Page d'attente ──────────────────────────────────────────────
@bp.route("/paiement/airtel/attente/<reference>")
@login_required
def paiement_airtel_attente(reference):
    """
    Page d'attente affichée après l'initiation.
    Fait un polling automatique toutes les 5 secondes via AJAX.
    """
    if current_user.is_super_admin: return redirect(url_for("admin.admin_dashboard"))
    t = get_tenant()
    if not t: return redirect(url_for("auth.login"))
    p = Paiement.query.filter_by(
        reference_interne=reference, tenant_id=t.id
    ).first_or_404()
    return render_template("tenant/paiement_attente.html", paiement=p, tenant=t)


# ── Airtel Money — Vérification statut (AJAX polling) ─────────────────────────
@bp.route("/paiement/airtel/statut/<reference>")
@login_required
def paiement_airtel_statut(reference):
    """
    Endpoint JSON pour le polling côté client.
    Retourne le statut actuel de la transaction.
    """
    t = get_tenant()
    if not t: return jsonify({"statut": "ERREUR", "message": "Non connecté"}), 401

    p = Paiement.query.filter_by(
        reference_interne=reference, tenant_id=t.id
    ).first_or_404()

    # Si déjà confirmé en base, retourner directement
    if p.statut == "SUCCES":
        return jsonify({"statut": "SUCCES", "message": "Paiement confirmé !"})
    if p.statut == "ECHEC":
        return jsonify({"statut": "ECHEC", "message": p.notes or "Paiement refusé."})
    if p.statut == "EXPIRE":
        return jsonify({"statut": "EXPIRE", "message": "Délai dépassé. Recommencez."})

    # Interroger l'API Airtel si on a un transaction_id
    if p.reference_externe:
        try:
            from airtel_money import verifier_statut
            r = verifier_statut(p.reference_externe)
            if r["statut"] == "SUCCESS":
                _activer_abonnement(p)
                return jsonify({"statut": "SUCCES", "message": "Paiement confirmé !"})
            elif r["statut"] in ("FAILED", "EXPIRED"):
                p.statut = "ECHEC" if r["statut"] == "FAILED" else "EXPIRE"
                p.notes  = r.get("message", "")
                db.session.commit()
                return jsonify({"statut": p.statut, "message": p.notes})
        except Exception as e:
            logger.warning(f"[Paiement] Polling Airtel erreur : {e}")

    return jsonify({"statut": "EN_ATTENTE", "message": "En attente de confirmation…"})


# ── Airtel Money — Webhook (callback automatique d'Airtel) ────────────────────
@bp.route("/webhook/airtel", methods=["POST"])
def webhook_airtel():
    """
    Reçoit les notifications automatiques d'Airtel après paiement du client.
    Airtel appelle cette URL avec le résultat de la transaction.
    """
    from airtel_money import valider_signature_webhook
    import json

    payload_bytes = request.get_data()
    signature     = request.headers.get("X-Airtel-Signature", "")

    # Vérifier la signature si configurée
    if not valider_signature_webhook(payload_bytes, signature):
        logger.warning("[Webhook Airtel] Signature invalide — requête ignorée.")
        return jsonify({"status": "SIGNATURE_INVALIDE"}), 401

    try:
        data = request.get_json(force=True) or {}
        logger.info(f"[Webhook Airtel] Reçu : {data}")

        # Extraire les infos de la transaction
        txn   = data.get("transaction", {}) or data.get("data", {}).get("transaction", {})
        ref   = txn.get("id") or txn.get("reference") or data.get("reference", "")
        statut_api = (txn.get("status") or data.get("status", {}).get("code", "")).upper()

        if not ref:
            logger.warning("[Webhook Airtel] Référence manquante dans le payload.")
            return jsonify({"status": "REF_MANQUANTE"}), 400

        # Retrouver le paiement
        p = Paiement.query.filter(
            (Paiement.reference_interne == ref) |
            (Paiement.reference_externe == ref)
        ).first()

        if not p:
            logger.warning(f"[Webhook Airtel] Paiement introuvable pour ref={ref}")
            return jsonify({"status": "INTROUVABLE"}), 404

        if p.statut == "SUCCES":
            # Idempotence — déjà traité
            return jsonify({"status": "DEJA_TRAITE"}), 200

        p.reponse_raw = json.dumps(data)

        if statut_api in ("TS", "SUCCESS", "200"):
            # Confirmation autoritative : on ré-interroge Airtel avant d'activer
            from airtel_money import verifier_statut
            verif = verifier_statut(p.reference_externe or ref)
            if verif.get("statut") != "SUCCESS":
                p.statut = "ECHEC"
                p.notes  = f"Webhook OK mais vérification Airtel = {verif.get('statut')}"
                db.session.commit()
                logger.warning(f"[Webhook Airtel] Vérif. divergente — ref={ref}")
                return jsonify({"status": "NON_CONFIRME"}), 200
            _activer_abonnement(p)
            logger.info(f"[Webhook Airtel] Succès vérifié — ref={ref} tenant={p.tenant_id}")
        else:
            p.statut = "ECHEC"
            p.notes  = f"Code Airtel : {statut_api}"
            db.session.commit()
            logger.info(f"[Webhook Airtel] Échec — ref={ref} code={statut_api}")

        return jsonify({"status": "OK"}), 200

    except Exception as e:
        logger.error(f"[Webhook Airtel] Erreur traitement : {e}")
        db.session.rollback()
        return jsonify({"status": "ERREUR_INTERNE"}), 500


# ── Helper : activer l'abonnement après paiement confirmé ─────────────────────
def _activer_abonnement(paiement: "Paiement"):
    """
    Appelé après confirmation d'un paiement (webhook ou polling).
    Met à jour le tenant : statut ACTIF, date_expiration prolongée.
    Envoie un email de confirmation.
    """
    from datetime import timezone
    p = paiement
    p.statut           = "SUCCES"
    p.date_confirmation = utcnow()

    t = p.tenant
    now = utcnow()

    # Prolonger depuis aujourd'hui ou depuis la date d'expiration si future
    base = t.date_expiration if (t.date_expiration and t.date_expiration > now) else now
    t.date_expiration = base + timedelta(days=30 * p.duree_mois)
    t.statut = "ACTIF"

    if p.plan_id:
        t.plan_id = p.plan_id

    db.session.commit()
    _cache_delete(f"{t.id}:")  # invalider le cache dashboard

    logger.info(
        f"[Abonnement] Tenant {t.id} activé jusqu'au "
        f"{t.date_expiration.strftime('%d/%m/%Y')} — "
        f"{p.duree_mois} mois via {p.moyen}"
    )

    # Email de confirmation
    try:
        msg = Message(
            subject=f"Abonnement PaieGabon activé — {t.denomination}",
            recipients=[u.email for u in t.utilisateurs if u.role == "TENANT_ADMIN" and u.email],
            body=(
                f"Bonjour,\n\n"
                f"Votre paiement de {float(p.montant):,.0f} FCFA a été confirmé.\n"
                f"Abonnement actif jusqu'au : {t.date_expiration.strftime('%d/%m/%Y')}\n"
                f"Référence : {p.reference_interne}\n\n"
                f"Merci de votre confiance.\n"
                f"L'équipe PaieGabon"
            ),
        )
        send_email_async(current_app.extensions["mail"], msg)
    except Exception as e:
        logger.warning(f"[Abonnement] Email de confirmation non envoyé : {e}")


@bp.route("/paiement/confirmer", methods=["POST"])
@login_required
def paiement_confirmer():
    """Paiement manuel : le client déclare avoir payé (Airtel Money ou virement).
    Crée un Paiement EN_ATTENTE que le super-admin validera après vérification."""
    if current_user.is_super_admin: return redirect(url_for("admin.admin_dashboard"))
    t = get_tenant()
    if not t: return redirect(url_for("auth.login"))
    moyen     = request.form.get("moyen", "MANUEL").strip().upper()  # AIRTEL_MONEY | VIREMENT
    reference = request.form.get("reference", "").strip()
    duree     = int(request.form.get("duree", 1) or 1)
    plan_id   = request.form.get("plan_id", type=int) or t.plan_id
    if not reference:
        flash("Veuillez indiquer la référence de la transaction.", "error")
        return redirect(url_for("tenant.paiement"))

    plan = db.session.get(Plan, plan_id) if plan_id else t.plan
    # Montant attendu avec remises par durée (3 mois -5%, 6 mois -10%, 12 mois -15%)
    remises = {1: 1.0, 3: 0.95, 6: 0.90, 12: 0.85}
    coef = remises.get(duree, 1.0)
    montant = round(float(plan.prix_mensuel) * duree * coef) if plan else 0

    import uuid
    ref_interne = f"MAN-{t.id}-{uuid.uuid4().hex[:8].upper()}"
    libelle_moyen = {"AIRTEL_MONEY": "Airtel Money", "VIREMENT": "Virement bancaire"}.get(moyen, moyen)
    p = Paiement(
        tenant_id=t.id, moyen=moyen, montant=montant,
        duree_mois=duree, plan_id=plan.id if plan else None,
        reference_interne=ref_interne, reference_externe=reference,
        statut="EN_ATTENTE",
        notes=f"Paiement {libelle_moyen} déclaré par {current_user.email}",
    )
    db.session.add(p)
    t.statut = "PAIEMENT_EN_ATTENTE"
    db.session.commit()
    log_action("CREATE", "paiement", p.id,
               f"Déclaration paiement {libelle_moyen} — réf {reference}, {duree} mois",
               user_id=current_user.id, tenant_id=t.id)
    flash(f"Paiement déclaré (réf. {reference}). Votre abonnement sera activé "
          f"après vérification, généralement sous 24-48h. Merci !", "success")
    return redirect(url_for("tenant.paiement"))


# ══════════════════════════════════════════════════════════════════════════════
# CINETPAY — Paiement multi-opérateurs (Airtel, Moov, Visa, Mastercard)
# ══════════════════════════════════════════════════════════════════════════════

@bp.route("/paiement/cinetpay/initier", methods=["POST"])
@login_required
def paiement_cinetpay_initier():
    """
    Initie un paiement CinetPay.
    Crée une session et redirige le client vers la page de paiement CinetPay
    où il choisit son moyen : Airtel Money, Moov Money ou carte bancaire.
    """
    if current_user.is_super_admin: return redirect(url_for("admin.admin_dashboard"))
    t = get_tenant()
    if not t: return redirect(url_for("auth.login"))

    duree   = int(request.form.get("duree", 1) or 1)
    plan_id = request.form.get("plan_id", type=int) or t.plan_id
    plan    = db.session.get(Plan, plan_id) if plan_id else t.plan

    if not plan:
        flash("Plan introuvable.", "error")
        return redirect(url_for("tenant.paiement"))

    montant = float(plan.prix_mensuel) * duree

    import uuid
    reference = f"CP-{t.id}-{uuid.uuid4().hex[:10].upper()}"

    # Récupérer l'admin du tenant pour pré-remplir les infos client
    admin = Utilisateur.query.filter_by(tenant_id=t.id, role="TENANT_ADMIN").first()
    nom_client   = admin.nom_complet if admin else t.denomination
    email_client = admin.email if admin else ""

    # Enregistrer la tentative
    p = Paiement(
        tenant_id=t.id,
        moyen="CINETPAY",
        montant=montant,
        duree_mois=duree,
        plan_id=plan.id,
        reference_interne=reference,
        statut="EN_ATTENTE",
    )
    db.session.add(p)
    db.session.commit()

    try:
        from cinetpay import initier_paiement, CinetPayConfigError
        resultat = initier_paiement(
            reference=reference,
            montant=montant,
            description=f"PaieGabon {plan.nom} — {duree} mois — {t.denomination}",
            nom_client=nom_client,
            email_client=email_client,
        )

        import json
        p.reponse_raw = json.dumps(resultat.get("raw", {}))

        if resultat["success"]:
            p.reference_externe = resultat.get("payment_token", "")
            db.session.commit()
            logger.info(f"[CinetPay] Session créée — ref={reference} tenant={t.id}")
            # Rediriger directement vers la page CinetPay
            return redirect(resultat["payment_url"])
        else:
            p.statut = "ECHEC"
            p.notes  = resultat["message"]
            db.session.commit()
            flash(f"Erreur CinetPay : {resultat['message']}", "error")
            return redirect(url_for("tenant.paiement"))

    except Exception as e:
        p.statut = "ECHEC"
        p.notes  = str(e)
        db.session.commit()
        logger.error(f"[CinetPay] Erreur initiation : {e}")
        flash("Erreur de connexion CinetPay. Réessayez ou contactez le support.", "error")
        return redirect(url_for("tenant.paiement"))


@bp.route("/paiement/cinetpay/retour")
@login_required
def paiement_cinetpay_retour():
    """
    Page de retour après la page de paiement CinetPay.
    CinetPay redirige ici après que le client ait terminé (succès ou annulation).
    On affiche un message d'attente pendant que le webhook confirme.
    """
    if current_user.is_super_admin: return redirect(url_for("admin.admin_dashboard"))
    t = get_tenant()
    if not t: return redirect(url_for("auth.login"))

    transaction_id = request.args.get("transaction_id", "")
    # Chercher le paiement par référence interne ou externe
    p = None
    if transaction_id:
        p = Paiement.query.filter(
            (Paiement.reference_interne == transaction_id) |
            (Paiement.reference_externe == transaction_id),
            Paiement.tenant_id == t.id
        ).first()

    # Vérification immédiate du statut
    if p and p.statut == "EN_ATTENTE" and (p.reference_interne or p.reference_externe):
        try:
            from cinetpay import verifier_statut
            ref = p.reference_interne
            r   = verifier_statut(ref)
            if r["statut"] == "ACCEPTED":
                _activer_abonnement(p)
                flash("Paiement confirmé ! Votre abonnement est actif.", "success")
                return redirect(url_for("tenant.dashboard"))
            elif r["statut"] in ("REFUSED", "CANCELLED"):
                p.statut = "ECHEC"
                p.notes  = r.get("message", "Paiement refusé ou annulé.")
                db.session.commit()
        except Exception as e:
            logger.warning(f"[CinetPay] Vérification retour échouée : {e}")

    if p and p.statut == "SUCCES":
        flash("Paiement confirmé ! Votre abonnement est actif.", "success")
        return redirect(url_for("tenant.dashboard"))

    # Afficher la page d'attente (le webhook va confirmer dans quelques secondes)
    return render_template("tenant/paiement_cinetpay_retour.html",
                           paiement=p, tenant=t,
                           transaction_id=transaction_id)


@bp.route("/paiement/cinetpay/statut/<reference>")
@login_required
def paiement_cinetpay_statut(reference):
    """Endpoint JSON pour le polling côté client sur la page de retour."""
    t = get_tenant()
    if not t: return jsonify({"statut": "ERREUR"}), 401

    p = Paiement.query.filter(
        (Paiement.reference_interne == reference),
        Paiement.tenant_id == t.id
    ).first_or_404()

    if p.statut == "SUCCES":
        return jsonify({"statut": "SUCCES", "message": "Paiement confirmé !"})
    if p.statut == "ECHEC":
        return jsonify({"statut": "ECHEC", "message": p.notes or "Paiement refusé."})

    # Vérification active
    try:
        from cinetpay import verifier_statut
        r = verifier_statut(reference)
        if r["statut"] == "ACCEPTED":
            _activer_abonnement(p)
            return jsonify({"statut": "SUCCES", "message": "Paiement confirmé !"})
        elif r["statut"] in ("REFUSED", "CANCELLED"):
            p.statut = "ECHEC"
            p.notes  = r.get("message", "")
            db.session.commit()
            return jsonify({"statut": "ECHEC", "message": p.notes})
    except Exception as e:
        logger.warning(f"[CinetPay] Polling statut erreur : {e}")

    return jsonify({"statut": "EN_ATTENTE", "message": "Vérification en cours…"})


@bp.route("/webhook/cinetpay", methods=["POST"])
def webhook_cinetpay():
    """
    Reçoit les notifications automatiques de CinetPay.

    SÉCURITÉ : on ne fait JAMAIS confiance au statut envoyé dans le corps de la
    requête (falsifiable — le site_id n'est pas un secret). On ré-interroge
    l'API CinetPay (/payment/check) pour obtenir le statut et le montant
    authentiques, et on vérifie que le montant payé correspond au montant
    attendu avant d'activer l'abonnement.
    """
    import json
    from cinetpay import valider_webhook, verifier_statut

    # 1. Lecture tolérante du corps (CinetPay peut envoyer du JSON ou du form-data)
    data = request.get_json(silent=True) or request.form.to_dict()
    logger.info(f"[Webhook CinetPay] Reçu : {data}")

    # 2. Filtre de premier niveau : le site_id doit correspondre
    if not valider_webhook(data):
        return jsonify({"status": "SITE_ID_INVALIDE"}), 401

    # 3. Extraire la référence de transaction
    ref = (data.get("cpm_trans_id") or data.get("transaction_id")
           or data.get("metadata") or "")
    if not ref:
        logger.warning("[Webhook CinetPay] Référence manquante.")
        return jsonify({"status": "REF_MANQUANTE"}), 400

    # 4. Retrouver le paiement en base
    p = Paiement.query.filter_by(reference_interne=ref).first()
    if not p:
        token = data.get("cpm_payment_config") or data.get("payment_token", "")
        p = Paiement.query.filter_by(reference_externe=token).first() if token else None
    if not p:
        logger.warning(f"[Webhook CinetPay] Paiement introuvable ref={ref}")
        return jsonify({"status": "INTROUVABLE"}), 404

    # 5. Idempotence — déjà traité avec succès
    if p.statut == "SUCCES":
        return jsonify({"status": "DEJA_TRAITE"}), 200

    # 6. VÉRIFICATION AUTORITATIVE côté serveur (ne pas croire le corps)
    try:
        verif = verifier_statut(p.reference_interne)
    except Exception as e:
        logger.error(f"[Webhook CinetPay] Échec vérification API : {e}")
        db.session.rollback()
        return jsonify({"status": "VERIF_ERREUR"}), 502

    p.reponse_raw = json.dumps({"webhook": data, "verification": verif.get("raw", {})})

    if verif.get("statut") != "ACCEPTED":
        p.statut = "ECHEC"
        p.notes  = f"Statut CinetPay vérifié : {verif.get('statut')}"
        db.session.commit()
        logger.info(f"[Webhook CinetPay] Non confirmé — ref={ref} statut={verif.get('statut')}")
        return jsonify({"status": "NON_CONFIRME"}), 200

    # 7. Vérifier que le MONTANT payé correspond au montant attendu
    montant_attendu = int(round(float(p.montant or 0)))
    try:
        montant_paye = int(round(float(verif.get("montant") or 0)))
    except (TypeError, ValueError):
        montant_paye = 0
    if montant_paye and montant_paye < montant_attendu:
        p.statut = "ECHEC"
        p.notes  = f"Montant payé ({montant_paye}) < attendu ({montant_attendu}) — rejeté."
        db.session.commit()
        logger.warning(f"[Webhook CinetPay] Montant insuffisant ref={ref} : "
                       f"{montant_paye} < {montant_attendu}")
        return jsonify({"status": "MONTANT_INVALIDE"}), 200

    # 8. Tout est vérifié → activer l'abonnement
    try:
        _activer_abonnement(p)
        logger.info(f"[Webhook CinetPay] Succès vérifié — ref={ref} tenant={p.tenant_id}")
        return jsonify({"status": "OK"}), 200
    except Exception as e:
        logger.error(f"[Webhook CinetPay] Erreur activation : {e}")
        db.session.rollback()
        return jsonify({"status": "ERREUR_INTERNE"}), 500


@bp.route("/parametres")
@tenant_required
def parametres():
    t=get_tenant()
    # Passer tous les plans actifs pour l'onglet abonnement
    plans_dispo = Plan.query.filter_by(actif=True).order_by(Plan.prix_mensuel.asc()).all()
    return render_template("tenant/parametres.html", tenant=t,
        rubriques=RubriquePaie.query.filter_by(actif=True).all(),
        categories=CategorieEmploi.query.filter_by(tenant_id=t.id).all(),
        users=Utilisateur.query.filter_by(tenant_id=t.id).all(),
        plans_dispo=plans_dispo)

@bp.route("/parametres/grille-salaires", methods=["GET", "POST"])
@tenant_required
def grille_salaires():
    """Édition/vérification de la grille de salaires conventionnelle.
    L'auto-remplissage du salaire de base (fiche salarié) n'utilise QUE la grille
    enregistrée ici — jamais la graine brute — pour garantir des montants validés."""
    import json as _json
    from calculs_paie import GRILLE_CATEGORIES_AERIEN, grille_salaire_aerien_seed
    t = get_tenant()
    if not current_user.can_edit:
        abort(403)

    if request.method == "POST":
        grille = {}
        for key, val in request.form.items():
            if not key.startswith("montant_") or not (val or "").strip():
                continue
            try:
                _, code, ech = key.split("_", 2)
                montant = float(val.replace(" ", "").replace("\u202f", "").replace(",", "."))
            except (ValueError, IndexError):
                continue
            if montant > 0:
                grille.setdefault(code, {})[ech] = round(montant, 2)
        t.grille_salaires = _json.dumps(grille, ensure_ascii=False)
        db.session.commit()
        log_action("UPDATE", "tenant", t.id, "Grille de salaires mise à jour")
        flash("Grille de salaires enregistrée. Elle est désormais utilisée pour "
              "pré-remplir le salaire de base des salariés.", "success")
        return redirect(url_for("tenant.grille_salaires"))

    # GET : grille sauvegardée, sinon graine aérienne (à vérifier) si convention AERIEN.
    grille, est_seed = {}, False
    if t.grille_salaires:
        try:
            grille = _json.loads(t.grille_salaires)
        except (ValueError, TypeError):
            grille = {}
    if not grille and (t.convention or "").upper() == "AERIEN":
        grille = grille_salaire_aerien_seed()
        est_seed = True
    return render_template("tenant/grille_salaires.html", tenant=t,
        categories=GRILLE_CATEGORIES_AERIEN, grille=grille,
        est_seed=est_seed, nb_echelons=10)


def _grille_tenant(t):
    """Grille de salaires ENREGISTRÉE du tenant → dict {code: {echelon: montant}}.
    Vide si non renseignée. Utilisée pour le pré-remplissage du salaire de base."""
    import json as _json
    if not t or not t.grille_salaires:
        return {}
    try:
        return _json.loads(t.grille_salaires)
    except (ValueError, TypeError):
        return {}


@bp.route("/parametres/logo", methods=["POST"])
@login_required
def parametres_logo():
    if current_user.is_super_admin: return redirect(url_for("admin.admin_dashboard"))
    t = get_tenant()
    if not t: return redirect(url_for("auth.login"))
    logo_file = request.files.get("logo")
    if logo_file and logo_file.filename:
        import base64
        file_data = logo_file.read()
        if len(file_data) > 1_000_000:
            flash("Fichier trop volumineux. Maximum 1 Mo.", "error")
            return redirect(url_for("tenant.parametres"))

        # ── Validation de l'extension ─────────────────────────────────────────
        ext = logo_file.filename.rsplit(".", 1)[-1].lower() if "." in logo_file.filename else ""
        EXTENSIONS_AUTORISEES = {"png", "jpg", "jpeg", "gif", "webp"}
        if ext not in EXTENSIONS_AUTORISEES:
            flash("Format non autorisé. Utilisez PNG, JPG, JPEG, GIF ou WEBP (pas SVG).", "error")
            return redirect(url_for("tenant.parametres"))

        # ── Validation du MIME réel (magic bytes) — pas seulement l'extension ─
        MAGIC = {
            b"\x89PNG":   "image/png",
            b"\xff\xd8\xff": "image/jpeg",
            b"GIF8":      "image/gif",
            b"RIFF":      None,  # WebP — vérification complémentaire ci-dessous
        }
        detected_mime = None
        for magic, mime in MAGIC.items():
            if file_data[:len(magic)] == magic:
                if magic == b"RIFF" and file_data[8:12] == b"WEBP":
                    detected_mime = "image/webp"
                else:
                    detected_mime = mime
                break
        if not detected_mime:
            flash("Le contenu du fichier ne correspond pas à une image valide.", "error")
            return redirect(url_for("tenant.parametres"))

        b64 = base64.b64encode(file_data).decode("utf-8")
        logo_data = f"data:{detected_mime};base64,{b64}"
        try:
            db.session.execute(db.text("UPDATE tenants SET logo_url = :logo WHERE id = :id"),{"logo": logo_data, "id": t.id})
            db.session.commit(); db.session.expire(t)
            log_action("UPDATE", "parametres", t.id, "Mise à jour du logo de la société")
            db.session.commit()
            flash("Logo mis à jour avec succès.", "success")
        except Exception as e:
            db.session.rollback(); flash(f"Erreur: {str(e)}", "error")
    else:
        flash("Aucun fichier sélectionné.", "error")
    return redirect(url_for("tenant.parametres"))

@bp.route("/parametres/logo/supprimer", methods=["POST"])
@login_required
def parametres_logo_supprimer():
    t = get_tenant()
    if not t: return redirect(url_for("auth.login"))
    t.logo_url = None; db.session.commit()
    log_action("UPDATE", "parametres", t.id, "Suppression du logo de la société")
    db.session.commit()
    flash("Logo supprime.", "success")
    return redirect(url_for("tenant.parametres"))

@bp.route("/parametres/modele-bulletin", methods=["POST"])
@login_required
def parametres_modele_bulletin():
    """Changer le modèle d'impression des bulletins."""
    if current_user.is_super_admin: return redirect(url_for("admin.admin_dashboard"))
    t = get_tenant()
    if not t: return redirect(url_for("auth.login"))
    modele = request.form.get("modele_bulletin", "classique")
    if modele not in ("classique", "moderne", "minimaliste"):
        modele = "classique"
    t.modele_bulletin = modele
    db.session.commit()
    log_action("UPDATE", "parametres", t.id, f"Modèle d'impression bulletins : {modele}")
    db.session.commit()
    flash(f"Modèle d'impression « {modele.capitalize()} » appliqué.", "success")
    return redirect(url_for("tenant.parametres"))

@bp.route("/parametres/societe", methods=["POST"])
@tenant_required
@can_edit
def parametres_societe():
    if not current_user.can_manage_parametres:
        flash("Accès refusé. Seul l'administrateur peut modifier les paramètres.", "error")
        return redirect(url_for("tenant.parametres"))
    t=get_tenant()
    for f in ["denomination","sigle","activite","secteur","nif","numero_cnss","numero_cnamgs","adresse","boite_postale","telephone","ville","region"]:
        try: setattr(t,f,request.form.get(f,"").strip() or None)
        except: pass
    # Convention collective applicable — mise à jour uniquement si le champ est
    # présent dans le formulaire soumis (évite d'écraser la valeur depuis un
    # formulaire qui ne l'inclut pas, ex. la carte « Informations générales »).
    from calculs_paie import CONVENTIONS_DISPONIBLES
    conv_raw = request.form.get("convention")
    if conv_raw is not None:
        conv = (conv_raw or "AUCUNE").upper()
        if conv in CONVENTIONS_DISPONIBLES:
            t.convention = conv
    # Seuil hebdomadaire de déclenchement des heures supplémentaires (dérogation).
    # La loi gabonaise fixe le seuil légal à 40h/semaine ; une dérogation peut le
    # porter jusqu'à 48h. On borne strictement la valeur saisie dans cet intervalle.
    seuil_raw = request.form.get("seuil_heures_sup_hebdo")
    if seuil_raw is not None and str(seuil_raw).strip() != "":
        try:
            seuil = float(str(seuil_raw).replace(",", "."))
            t.seuil_heures_sup_hebdo = max(40.0, min(seuil, 48.0))
        except (ValueError, TypeError):
            flash("Seuil d'heures supplémentaires invalide — valeur inchangée.", "error")
    # Langue de l'interface — idem, seulement si le champ est soumis.
    langue = request.form.get("langue")
    if langue is not None and langue in SUPPORTED_LANGUAGES:
        t.langue = langue
        set_language(langue)
    db.session.commit()
    log_action("UPDATE", "parametres", t.id, "Modification des informations de la société")
    db.session.commit()
    flash("Informations mises à jour." if (t.langue or "fr") == "fr" else "Settings updated.", "success")
    return redirect(url_for("tenant.parametres"))


@bp.route("/parametres/importer-grille-commerce", methods=["POST"])
@tenant_required
@can_edit
def importer_grille_commerce():
    """Crée/complète les catégories d'emploi à partir de la grille conventionnelle COMMERCE."""
    if not current_user.can_manage_parametres:
        flash("Accès refusé. Seul l'administrateur peut modifier les paramètres.", "error")
        return redirect(url_for("tenant.parametres"))
    t = get_tenant()
    from calculs_paie import GRILLE_COMMERCE
    existantes = {c.code for c in CategorieEmploi.query.filter_by(tenant_id=t.id).all()}
    ajout = 0
    maj = 0
    for code, libelle, mensuel, _horaire in GRILLE_COMMERCE:
        if code in existantes:
            cat = CategorieEmploi.query.filter_by(tenant_id=t.id, code=code).first()
            if cat and (cat.salaire_minimum is None or float(cat.salaire_minimum or 0) == 0):
                cat.salaire_minimum = mensuel
                maj += 1
        else:
            db.session.add(CategorieEmploi(
                tenant_id=t.id, code=code, libelle=libelle, salaire_minimum=mensuel,
                description="Grille Convention Collective du Commerce"))
            ajout += 1
    # Bascule la convention du tenant sur COMMERCE si pas déjà fait
    if t.convention != "COMMERCE":
        t.convention = "COMMERCE"
    db.session.commit()
    flash(f"Grille Commerce importée : {ajout} catégorie(s) ajoutée(s), {maj} mise(s) à jour.", "success")
    return redirect(url_for("tenant.parametres"))


@bp.route("/parametres/importer-grille-hydrocarbures", methods=["POST"])
@tenant_required
@can_edit
def importer_grille_hydrocarbures():
    """Crée/complète les catégories A→M à partir de la grille Hydrocarbures."""
    if not current_user.can_manage_parametres:
        flash("Accès refusé. Seul l'administrateur peut modifier les paramètres.", "error")
        return redirect(url_for("tenant.parametres"))
    t = get_tenant()
    from convention_hydrocarbures import GRILLE_HYDROCARBURES
    existantes = {c.code for c in CategorieEmploi.query.filter_by(tenant_id=t.id).all()}
    ajout = 0
    maj = 0
    for code, libelle, mensuel, _horaire in GRILLE_HYDROCARBURES:
        if code in existantes:
            cat = CategorieEmploi.query.filter_by(tenant_id=t.id, code=code).first()
            if cat and (cat.salaire_minimum is None or float(cat.salaire_minimum or 0) == 0):
                cat.salaire_minimum = mensuel
                maj += 1
        else:
            db.session.add(CategorieEmploi(
                tenant_id=t.id, code=code, libelle=libelle, salaire_minimum=mensuel,
                description="Grille Convention Hydrocarbures (Recherche & Exploitation)"))
            ajout += 1
    # Bascule la convention du tenant sur HYDROCARBURES
    if t.convention != "HYDROCARBURES":
        t.convention = "HYDROCARBURES"
    # Congés de base : 2,5 jours ouvrables / mois (Art. 42)
    if hasattr(t, "jours_conge_par_mois"):
        t.jours_conge_par_mois = 2.5
    db.session.commit()
    flash(f"Grille Hydrocarbures importée : {ajout} catégorie(s) ajoutée(s), {maj} mise(s) à jour.", "success")
    return redirect(url_for("tenant.parametres"))


@bp.route("/parametres/importer-grille-petrole", methods=["POST"])
@tenant_required
@can_edit
def importer_grille_petrole():
    """Crée/complète les catégories d'emploi à partir de la grille conventionnelle PÉTROLE.

    ⚠️ Les montants de l'Annexe n°2 (1983) sont obsolètes : les premières
    catégories sont sous le SMIG actuel. On importe donc la STRUCTURE des
    catégories en appliquant un plancher au SMIG légal, et on n'écrase jamais une
    catégorie existante. Les montants doivent être actualisés par l'entreprise.
    """
    if not current_user.can_manage_parametres:
        flash("Accès refusé. Seul l'administrateur peut modifier les paramètres.", "error")
        return redirect(url_for("tenant.parametres"))
    t = get_tenant()
    from calculs_paie import GRILLE_PETROLE, SMIG_GABON
    existantes = {c.code for c in CategorieEmploi.query.filter_by(tenant_id=t.id).all()}
    ajout = 0
    maj = 0
    plancher_applique = False
    for code, libelle, mensuel_1983 in GRILLE_PETROLE:
        # Plancher SMIG : aucune catégorie ne peut être créée sous le minimum légal.
        mensuel = max(int(mensuel_1983), int(SMIG_GABON))
        if mensuel != int(mensuel_1983):
            plancher_applique = True
        if code in existantes:
            cat = CategorieEmploi.query.filter_by(tenant_id=t.id, code=code).first()
            if cat and (cat.salaire_minimum is None or float(cat.salaire_minimum or 0) == 0):
                cat.salaire_minimum = mensuel
                maj += 1
        else:
            db.session.add(CategorieEmploi(
                tenant_id=t.id, code=code, libelle=libelle, salaire_minimum=mensuel,
                description="Grille Convention Pétrole (montants à actualiser)"))
            ajout += 1
    if t.convention != "PETROLE":
        t.convention = "PETROLE"
    db.session.commit()
    msg = f"Grille Pétrole importée : {ajout} catégorie(s) ajoutée(s), {maj} mise(s) à jour."
    if plancher_applique:
        msg += (f" ⚠️ Certains montants de 1983 étaient sous le SMIG "
                f"({int(SMIG_GABON):,} FCFA) et ont été relevés au plancher légal — "
                f"actualisez-les selon votre grille interne.").replace(",", " ")
    flash(msg, "success")
    return redirect(url_for("tenant.parametres"))


@bp.route("/langue/<lang>")
def changer_langue(lang):
    """Change la langue de l'interface — accessible depuis n'importe quelle page."""
    set_language(lang)
    # Sauvegarder sur le tenant si connecté
    if current_user.is_authenticated and not current_user.is_super_admin:
        t = get_tenant()
        if t and lang in SUPPORTED_LANGUAGES:
            t.langue = lang
            db.session.commit()
    # Rediriger vers la page précédente
    return redirect(request.referrer or url_for("tenant.dashboard"))


# ══════════════════════════════════════════════════════════════════════════════
# AUDIT TRAIL — Journal des actions
# ══════════════════════════════════════════════════════════════════════════════

@bp.route("/audit")
@tenant_required
def audit_trail():
    """
    Page du journal d'audit — visible par l'admin du tenant.
    Affiche qui a fait quoi et quand.
    """
    t = get_tenant()
    if not current_user.is_tenant_admin:
        flash("Accès réservé à l'administrateur.", "error")
        return redirect(url_for("tenant.dashboard"))

    page      = request.args.get("page", 1, type=int)
    action    = request.args.get("action", "")
    entite    = request.args.get("entite", "")
    user_id   = request.args.get("user_id", type=int)
    recherche = request.args.get("q", "").strip()
    d_debut   = _parse_date(request.args.get("date_debut", ""))
    d_fin     = _parse_date(request.args.get("date_fin", ""))
    per_page  = 50

    # La date de fin est inclusive (jusqu'à 23:59:59)
    from datetime import datetime as _dt, time as _time
    d_fin_dt = _dt.combine(d_fin, _time.max) if d_fin else None

    logs, total = get_audit_logs(
        tenant_id  = t.id,
        limit      = per_page,
        offset     = (page - 1) * per_page,
        action     = action or None,
        entite     = entite or None,
        user_id    = user_id or None,
        date_debut = d_debut or None,
        date_fin   = d_fin_dt,
        recherche  = recherche or None,
    )

    # Liste des utilisateurs pour le filtre
    users = Utilisateur.query.filter_by(tenant_id=t.id, actif=True).order_by(Utilisateur.nom).all()

    import math
    nb_pages = math.ceil(total / per_page) if total else 1

    return render_template("tenant/audit_trail.html",
        tenant=t, logs=logs, total=total,
        page=page, nb_pages=nb_pages, per_page=per_page,
        action_filtre=action, entite_filtre=entite, user_filtre=user_id,
        recherche=recherche,
        date_debut=request.args.get("date_debut", ""),
        date_fin=request.args.get("date_fin", ""),
        users=users,
        ACTIONS=["CREATE","UPDATE","DELETE","VALIDATE","CANCEL","PAY",
                 "LOGIN","LOGOUT","EXPORT","IMPORT"],
        ENTITES=["salarie","bulletin","conge","acompte","periode","paiement",
                 "pointage","journalier","avance_journalier","feuille_journalier",
                 "prestataire","facture_prestataire","contrat_prestation",
                 "avance_prestataire","utilisateur","parametres"],
    )


@bp.route("/audit/support")
@tenant_required
def audit_support():
    """Journal des interventions support de l'éditeur (Ameriack).
    Présenté séparément du journal courant : transparence sur les accès
    de support, sans parasiter le suivi quotidien du client."""
    t = get_tenant()
    if not current_user.is_tenant_admin:
        flash("Accès réservé à l'administrateur.", "error")
        return redirect(url_for("tenant.dashboard"))

    from audit import get_audit_logs_support
    page = request.args.get("page", 1, type=int)
    per_page = 50
    logs, total = get_audit_logs_support(
        tenant_id=t.id, limit=per_page, offset=(page - 1) * per_page)

    import math
    nb_pages = math.ceil(total / per_page) if total else 1

    return render_template("tenant/audit_support.html",
        tenant=t, logs=logs, total=total,
        page=page, nb_pages=nb_pages, per_page=per_page)


@bp.route("/audit/export")
@tenant_required
def audit_export():
    """Export CSV du journal d'audit."""
    t = get_tenant()
    if not current_user.is_tenant_admin:
        flash("Accès réservé à l'administrateur.", "error")
        return redirect(url_for("tenant.audit_trail"))

    from datetime import datetime as _dt, time as _time
    d_debut = _parse_date(request.args.get("date_debut", ""))
    d_fin   = _parse_date(request.args.get("date_fin", ""))
    logs, _ = get_audit_logs(
        tenant_id=t.id, limit=5000,
        action=request.args.get("action") or None,
        entite=request.args.get("entite") or None,
        user_id=request.args.get("user_id", type=int) or None,
        recherche=request.args.get("q", "").strip() or None,
        date_debut=d_debut or None,
        date_fin=(_dt.combine(d_fin, _time.max) if d_fin else None))

    import csv, io as _io
    output = _io.StringIO()
    writer = csv.writer(output, delimiter=";")
    writer.writerow(["Date","Utilisateur","Rôle","Action","Objet","ID","Description","IP"])
    for l in logs:
        writer.writerow([
            l.date_action.strftime("%d/%m/%Y %H:%M:%S") if l.date_action else "",
            csv_safe(l.user.nom_complet if l.user else "Système"),
            l.user.role_label  if l.user else "—",
            csv_safe(l.action), csv_safe(l.entite or ""), l.entite_id or "",
            csv_safe(l.description or ""), l.ip_address or "",
        ])

    log_action("EXPORT", "audit", None, "Export CSV journal d'audit")
    db.session.commit()

    return send_file(
        io.BytesIO(("\ufeff" + output.getvalue()).encode("utf-8")),
        mimetype="text/csv; charset=utf-8",
        as_attachment=True,
        download_name=f"audit_{t.sigle or t.id}_{datetime.now().strftime('%Y%m%d')}.csv",
    )


@bp.route("/parametres/demande-changement-plan", methods=["POST"])
@login_required
def demande_changement_plan():
    if current_user.is_super_admin: return redirect(url_for("admin.admin_dashboard"))
    t = get_tenant()
    if not t: return redirect(url_for("auth.login"))
    if not current_user.is_tenant_admin: abort(403)
    plan_souhaite_id = request.form.get("plan_id", type=int)
    motif = request.form.get("motif", "").strip()
    plan_souhaite = db.session.get(Plan, plan_souhaite_id) if plan_souhaite_id else None
    if not plan_souhaite:
        flash("Plan invalide.", "error")
        return redirect(url_for("tenant.parametres"))
    # Enregistrer la demande dans les notes + changer statut
    note_demande = (
        f"[DEMANDE CHANGEMENT PLAN — {datetime.now().strftime('%d/%m/%Y %H:%M')}] "
        f"Plan souhaité : {plan_souhaite.nom} ({int(plan_souhaite.prix_mensuel):,} FCFA/mois). "
        f"Motif : {motif or 'Non précisé'}. "
        f"Demandé par : {current_user.nom_complet} ({current_user.email})."
    )
    # Ajouter à la suite des notes existantes
    t.notes = (t.notes or "") + ("\n" if t.notes else "") + note_demande
    t.statut = "PAIEMENT_EN_ATTENTE"
    db.session.commit()
    flash(f"Demande de passage au plan « {plan_souhaite.nom} » enregistrée. L'équipe PaieGabon vous contactera sous 24h pour finaliser.", "success")
    return redirect(url_for("tenant.parametres"))

@bp.route("/parametres/annuler-abonnement", methods=["POST"])
@login_required
def annuler_abonnement():
    if current_user.is_super_admin: return redirect(url_for("admin.admin_dashboard"))
    t = get_tenant()
    if not t: return redirect(url_for("auth.login"))
    if not current_user.is_tenant_admin: abort(403)
    motif = request.form.get("motif", "").strip()
    t.statut = "ANNULATION_DEMANDEE"
    t.notes = f"Annulation demandée le {datetime.now().strftime('%d/%m/%Y')}. Motif: {motif}"
    db.session.commit()
    flash("Demande d annulation enregistrée. L equipe PaieGabon vous contactera sous 48h.", "success")
    return redirect(url_for("tenant.parametres"))

# ── Utilisateurs ──────────────────────────────────────────────────────────────
@bp.route("/utilisateurs")
@login_required
def utilisateurs():
    if current_user.is_super_admin: return redirect(url_for("admin.admin_dashboard"))
    t = get_tenant()
    if not t: return redirect(url_for("auth.login"))
    liste = Utilisateur.query.filter_by(tenant_id=t.id).order_by(Utilisateur.nom).all()
    return render_template("tenant/utilisateurs.html", tenant=t, utilisateurs=liste, users=liste)

@bp.route("/utilisateurs/nouveau", methods=["GET","POST"])
@login_required
def utilisateur_nouveau():
    if current_user.is_super_admin: return redirect(url_for("admin.admin_dashboard"))
    t = get_tenant()
    if not t: return redirect(url_for("auth.login"))
    if not current_user.is_tenant_admin:
        flash("Réservé à l administrateur.", "error")
        return redirect(url_for("tenant.utilisateurs"))
    # ── Vérifier la limite dès le GET (bloquer l'accès au formulaire) ───────
    if t.plan and t.plan.max_utilisateurs:
        nb_actuel = Utilisateur.query.filter_by(tenant_id=t.id, actif=True).count()
        if nb_actuel >= t.plan.max_utilisateurs:
            flash(
                f"Limite atteinte — Plan « {t.plan.nom} » : "
                f"{t.plan.max_utilisateurs} utilisateur(s) maximum "
                f"(vous en avez {nb_actuel}). "
                f"Passez au plan supérieur pour en ajouter d'autres.",
                "error"
            )
            return redirect(url_for("tenant.utilisateurs"))

    if request.method == "GET":
        nb_utilisateurs = Utilisateur.query.filter_by(tenant_id=t.id, actif=True).count()
        return render_template("tenant/utilisateur_form.html", tenant=t,
            nb_utilisateurs=nb_utilisateurs)
    email = request.form.get("email", "").strip().lower()
    nom = request.form.get("nom", "").strip().upper()
    prenom = request.form.get("prenom", "").strip()
    password = request.form.get("password", "")
    role = request.form.get("role", "GESTIONNAIRE").strip().upper()
    # Liste blanche : empêche l'attribution de SUPER_ADMIN ou d'un rôle inconnu.
    if role not in ROLES_TENANT_AUTORISES:
        flash("Rôle invalide.", "error")
        return redirect(url_for("tenant.utilisateurs"))
    if not email or not nom or not password:
        flash("Veuillez remplir tous les champs.", "error")
        return render_template("tenant/utilisateur_form.html", tenant=t)
    if Utilisateur.query.filter_by(email=email).first():
        flash("Email déjà utilisé.", "error")
        return render_template("tenant/utilisateur_form.html", tenant=t)
    u = Utilisateur(nom=nom, prenom=prenom, email=email, role=role, tenant_id=t.id, actif=True)
    u.set_password(password)
    db.session.add(u); db.session.commit()
    log_action("CREATE", "utilisateur", u.id,
               f"Création utilisateur {u.nom_complet} ({u.role})")
    db.session.commit()
    flash(f"Utilisateur {u.nom_complet} créé.", "success")
    return redirect(url_for("tenant.utilisateurs"))

@bp.route("/utilisateurs/<int:id>/toggle", methods=["POST"])
@login_required
def utilisateur_toggle(id):
    """Activer / désactiver un utilisateur."""
    if current_user.is_super_admin: return redirect(url_for("admin.admin_dashboard"))
    t = get_tenant()
    if not t: return redirect(url_for("auth.login"))
    if not current_user.is_tenant_admin:
        flash("Réservé à l'administrateur.", "error")
        return redirect(url_for("tenant.utilisateurs"))
    u = Utilisateur.query.filter_by(id=id, tenant_id=t.id).first_or_404()
    if u.id == current_user.id:
        flash("Vous ne pouvez pas vous désactiver vous-même.", "error")
        return redirect(url_for("tenant.utilisateurs"))
    u.actif = not u.actif
    db.session.commit()
    etat = "activé" if u.actif else "désactivé"
    log_action("UPDATE", "utilisateur", u.id,
               f"Utilisateur {u.nom_complet} {etat} par {current_user.nom_complet}")
    db.session.commit()
    flash(f"Utilisateur {u.nom_complet} {etat}.", "success")
    return redirect(url_for("tenant.utilisateurs"))


@bp.route("/utilisateurs/<int:id>/modifier", methods=["GET","POST"])
@login_required
def utilisateur_modifier(id):
    """Modifier le rôle et les infos d'un utilisateur."""
    if current_user.is_super_admin: return redirect(url_for("admin.admin_dashboard"))
    t = get_tenant()
    if not t: return redirect(url_for("auth.login"))
    if not current_user.is_tenant_admin:
        flash("Réservé à l'administrateur.", "error")
        return redirect(url_for("tenant.utilisateurs"))

    u = Utilisateur.query.filter_by(id=id, tenant_id=t.id).first_or_404()

    if request.method == "POST":
        ancien_role = u.role
        nouveau_role = request.form.get("role", u.role).strip().upper()
        # Liste blanche : empêche l'escalade vers SUPER_ADMIN ou un rôle inconnu.
        if nouveau_role not in ROLES_TENANT_AUTORISES:
            flash("Rôle invalide.", "error")
            return redirect(url_for("tenant.utilisateurs"))

        # Empêcher de retirer son propre rôle admin
        if u.id == current_user.id and nouveau_role != "TENANT_ADMIN":
            flash("Vous ne pouvez pas changer votre propre rôle.", "error")
            return redirect(url_for("tenant.utilisateurs"))

        u.nom    = request.form.get("nom", u.nom).strip().upper()
        u.prenom = request.form.get("prenom", u.prenom).strip()
        u.role   = nouveau_role

        # Changer le mot de passe si fourni
        nouveau_mdp = request.form.get("nouveau_mdp", "").strip()
        if nouveau_mdp:
            if len(nouveau_mdp) < 8:
                flash("Le mot de passe doit faire au moins 8 caractères.", "error")
                return render_template("tenant/utilisateur_form.html",
                                       utilisateur=u, tenant=t, mode="modifier")
            u.set_password(nouveau_mdp)

        db.session.commit()
        log_action("UPDATE", "utilisateur", u.id,
                   f"Modification {u.nom_complet} — rôle : {ancien_role} → {nouveau_role}")
        db.session.commit()
        flash(f"Utilisateur {u.nom_complet} mis à jour.", "success")
        return redirect(url_for("tenant.utilisateurs"))

    return render_template("tenant/utilisateur_form.html",
                           utilisateur=u, tenant=t, mode="modifier")


@bp.route("/utilisateurs/<int:id>/supprimer", methods=["POST"])
@login_required
def utilisateur_supprimer(id):
    """
    Supprime définitivement un utilisateur.
    Règles :
      - Seul l'admin du tenant peut supprimer
      - On ne peut pas se supprimer soi-même
      - On ne peut pas supprimer le dernier admin
    """
    if current_user.is_super_admin: return redirect(url_for("admin.admin_dashboard"))
    t = get_tenant()
    if not t: return redirect(url_for("auth.login"))
    if not current_user.is_tenant_admin:
        flash("Réservé à l'administrateur.", "error")
        return redirect(url_for("tenant.utilisateurs"))

    u = Utilisateur.query.filter_by(id=id, tenant_id=t.id).first_or_404()

    # Règle 1 : pas de suicide
    if u.id == current_user.id:
        flash("Vous ne pouvez pas supprimer votre propre compte.", "error")
        return redirect(url_for("tenant.utilisateurs"))

    # Règle 2 : conserver au moins un admin actif
    if u.role == "TENANT_ADMIN":
        nb_admins = Utilisateur.query.filter_by(
            tenant_id=t.id, role="TENANT_ADMIN", actif=True
        ).filter(Utilisateur.id != u.id).count()
        if nb_admins == 0:
            flash("Impossible de supprimer le seul administrateur actif du compte. "
                  "Activez ou désignez d'abord un autre administrateur.", "error")
            return redirect(url_for("tenant.utilisateurs"))

    nom_sauvegarde = u.nom_complet
    log_action("DELETE", "utilisateur", u.id,
               f"Suppression définitive de {nom_sauvegarde} ({u.role_label})")
    db.session.delete(u)
    db.session.commit()
    flash(f"Utilisateur {nom_sauvegarde} supprimé définitivement.", "success")
    return redirect(url_for("tenant.utilisateurs"))

# ── Journaliers ───────────────────────────────────────────────────────────────
@bp.route("/journaliers")
@login_required
def journaliers():
    if current_user.is_super_admin: return redirect(url_for("admin.admin_dashboard"))
    t = get_tenant()
    if not t: return redirect(url_for("auth.login"))
    q    = request.args.get("q", "")
    page = request.args.get("page", 1, type=int)
    query = Journalier.query.filter_by(tenant_id=t.id)
    if q: query = query.filter(db.or_(Journalier.nom.ilike(f"%{q}%"), Journalier.prenom.ilike(f"%{q}%"), Journalier.profession.ilike(f"%{q}%")))
    pagination = query.order_by(Journalier.nom).paginate(page=page, per_page=25, error_out=False)
    _args = {k: v for k, v in request.args.items() if k != 'page'}
    _base = request.path + '?' + '&'.join(f'{k}={v}' for k, v in _args.items())
    _sep  = '&' if _args else '?'
    return render_template("tenant/journaliers.html",
        tenant=t, journaliers=pagination.items, pagination=pagination, q=q,
        pagination_base=_base + _sep)

@bp.route("/journaliers/nouveau", methods=["GET","POST"])
@login_required
def journalier_nouveau():
    if current_user.is_super_admin: return redirect(url_for("admin.admin_dashboard"))
    t = get_tenant()
    if not t: return redirect(url_for("auth.login"))
    # Vérifier quota dès le GET
    q = t.quota_employes_info
    if q["max"] and q["plein"]:
        flash(
            f"Limite atteinte — Plan « {t.plan.nom} » : {q['max']} employé(s) maximum "
            f"({q['salaries']} salarié(s) + {q['journaliers']} journalier(s)). "
            f"Passez au plan supérieur.", "error"
        )
        return redirect(url_for("tenant.journaliers"))
    if request.method == "POST":
        if not t.peut_ajouter_employe:
            flash(f"Limite atteinte ({t.plan.max_salaries} employés). Passez au plan supérieur.","error")
            return redirect(url_for("tenant.journaliers"))
        j = Journalier(tenant_id=t.id,
            nom=request.form["nom"].strip().upper(),
            prenom=request.form["prenom"].strip(),
            telephone=request.form.get("telephone","").strip(),
            profession=request.form.get("profession","").strip().upper(),
            taux_horaire=float(request.form.get("taux_horaire",0) or 0),
            type_paie=("MENSUEL" if request.form.get("type_paie")=="MENSUEL" else "JOURNALIER"),
            date_embauche=_parse_date(request.form.get("date_embauche")),
            date_debut=   _parse_date(request.form.get("date_debut")),
            date_fin=     _parse_date(request.form.get("date_fin")),
            nationalite=  request.form.get("nationalite","").strip() or None,
            mode_paiement=(request.form.get("mode_paiement","ESPECES") or "ESPECES").strip(),
            statut="ACTIF")
        db.session.add(j); db.session.commit()
        log_action("CREATE", "journalier", j.id,
                   f"Création journalier {j.nom_complet} ({j.type_paie})", apres=j.to_dict())
        db.session.commit()
        flash(f"Journalier {j.nom_complet} créé.", "success")
        return redirect(url_for("tenant.journaliers"))
    return render_template("tenant/journalier_form.html", tenant=t, journalier=None)

@bp.route("/journaliers/<int:id>")
@login_required
def journalier_detail(id):
    """Fiche détail d'un journalier avec historique de pointage."""
    if current_user.is_super_admin: return redirect(url_for("admin.admin_dashboard"))
    t = get_tenant()
    if not t: return redirect(url_for("auth.login"))
    j = Journalier.query.filter_by(id=id, tenant_id=t.id).first_or_404()

    # Feuilles de paie
    feuilles = FeuillePaieJournalier.query.filter_by(
        journalier_id=id, tenant_id=t.id
    ).order_by(FeuillePaieJournalier.date_fin.desc()).all()
    total_percu = sum(float(f.montant_brut or 0) for f in feuilles if f.statut == "PAYÉ")

    # ── Avances du journalier + déduction sur les feuilles ────────────────────
    avances = (AvanceJournalier.query.filter_by(journalier_id=id, tenant_id=t.id)
               .order_by(AvanceJournalier.date_avance.desc()).all())
    total_avances = round(sum(float(a.montant or 0) for a in avances), 2)
    avances_reste = round(sum(a.reste_a_regulariser for a in avances), 2)
    imput_av = _imputer_avances_journalier(feuilles, {id: avances})

    # Affectation site courante
    aff = AffectationSite.query.filter_by(
        journalier_id=id, tenant_id=t.id, actif=True).first()

    # ── Historique des pointages ──────────────────────────────────────────────
    nb_jours = request.args.get("nb_jours", type=int, default=30)
    nb_jours = min(max(nb_jours, 7), 90)
    date_fin   = datetime.now().date()
    date_debut = date_fin - timedelta(days=nb_jours - 1)

    pts_hist = Pointage.query.filter_by(tenant_id=t.id, journalier_id=id)        .filter(Pointage.date_pointage >= date_debut,
                Pointage.date_pointage <= date_fin)        .order_by(Pointage.date_pointage.desc()).all()

    nb_presences   = sum(1 for p in pts_hist if p.present)
    nb_absences    = sum(1 for p in pts_hist if p.absent)
    nb_non_pointes = nb_jours - len(pts_hist)
    h_normales_tot = round(sum(float(p.heures_normales or 0) for p in pts_hist if p.present), 1)
    h_sup_tot      = round(sum(float(p.heures_sup or 0) for p in pts_hist if p.present), 1)
    taux_presence  = round(nb_presences / (nb_presences + nb_absences) * 100
                           ) if (nb_presences + nb_absences) > 0 else 0

    return render_template("tenant/journalier_detail.html",
        journalier=j, tenant=t, feuilles=feuilles,
        total_percu=total_percu, aff=aff,
        avances=avances, total_avances=total_avances, avances_reste=avances_reste,
        imput_av=imput_av,
        pts_hist=pts_hist, nb_jours=nb_jours,
        nb_presences=nb_presences, nb_absences=nb_absences,
        nb_non_pointes=nb_non_pointes,
        h_normales_tot=h_normales_tot, h_sup_tot=h_sup_tot,
        taux_presence=taux_presence,
        date_debut_hist=date_debut, date_fin_hist=date_fin)

@bp.route("/journaliers/<int:id>/modifier", methods=["GET","POST"])
@login_required
def journalier_modifier(id):
    t = get_tenant()
    if not t: return redirect(url_for("auth.login"))
    j = Journalier.query.filter_by(id=id, tenant_id=t.id).first_or_404()
    if request.method == "POST":
        j.nom=request.form["nom"].strip().upper(); j.prenom=request.form["prenom"].strip()
        j.telephone=request.form.get("telephone","").strip()
        j.profession=request.form.get("profession","").strip().upper()
        j.taux_horaire=float(request.form.get("taux_horaire",0) or 0)
        j.type_paie=("MENSUEL" if request.form.get("type_paie")=="MENSUEL" else "JOURNALIER")
        j.date_embauche=_parse_date(request.form.get("date_embauche"))
        j.date_debut=   _parse_date(request.form.get("date_debut"))
        j.date_fin=     _parse_date(request.form.get("date_fin"))
        j.nationalite=  request.form.get("nationalite","").strip() or None
        j.mode_paiement=(request.form.get("mode_paiement","ESPECES") or "ESPECES").strip()
        j.statut=request.form.get("statut","ACTIF")
        # ── Affectation site ──────────────────────────────────────────────
        site_id = request.form.get("site_id", type=int)
        if site_id:
            aff_prev = AffectationSite.query.filter_by(
                journalier_id=j.id, tenant_id=t.id, actif=True).first()
            if aff_prev and aff_prev.site_id != site_id:
                aff_prev.actif    = False
                aff_prev.date_fin = date.today()
                aff_prev.motif    = "Réaffecté via formulaire journalier"
            if not aff_prev or aff_prev.site_id != site_id:
                db.session.add(AffectationSite(
                    tenant_id=t.id, site_id=site_id, journalier_id=j.id,
                    date_debut=date.today(), actif=True,
                    cree_par=current_user.email))
        elif request.form.get("retirer_site"):
            aff = AffectationSite.query.filter_by(
                journalier_id=j.id, tenant_id=t.id, actif=True).first()
            if aff:
                aff.actif    = False
                aff.date_fin = date.today()
                aff.motif    = "Retiré via formulaire journalier"
        db.session.commit()
        log_action("UPDATE", "journalier", j.id,
                   f"Modification journalier {j.nom_complet}", apres=j.to_dict())
        db.session.commit()
        flash("Journalier mis à jour.", "success")
        return redirect(url_for("tenant.journaliers"))
    aff_actuelle = AffectationSite.query.filter_by(
        journalier_id=id, tenant_id=t.id, actif=True).first()
    sites = Site.query.filter_by(tenant_id=t.id, actif=True).order_by(Site.nom).all()
    return render_template("tenant/journalier_form.html", tenant=t, journalier=j,
        sites=sites, aff_actuelle=aff_actuelle)

# ── Pointage ──────────────────────────────────────────────────────────────────
@bp.route("/pointage")
@login_required
def pointage():
    if current_user.is_super_admin: return redirect(url_for("admin.admin_dashboard"))
    t = get_tenant()
    if not t: return redirect(url_for("auth.login"))
    now = datetime.now()
    date_str = request.args.get("date", now.strftime("%Y-%m-%d"))
    try: date_sel = datetime.strptime(date_str, "%Y-%m-%d").date()
    except: date_sel = now.date()
    # ── Filtre par site ───────────────────────────────────────────────────────
    site_filtre_id = request.args.get("site_id", type=int)
    sites_list = Site.query.filter_by(tenant_id=t.id, actif=True).order_by(Site.nom).all()
    site_filtre = Site.query.filter_by(id=site_filtre_id, tenant_id=t.id).first() if site_filtre_id else None

    if site_filtre_id:
        # Salariés affectés à ce site
        ids_sal = [a.salarie_id for a in AffectationSite.query.filter_by(
            tenant_id=t.id, site_id=site_filtre_id, actif=True
        ).filter(AffectationSite.salarie_id.isnot(None)).all()]
        ids_jour = [a.journalier_id for a in AffectationSite.query.filter_by(
            tenant_id=t.id, site_id=site_filtre_id, actif=True
        ).filter(AffectationSite.journalier_id.isnot(None)).all()]
        salaries_list   = Salarie.query.filter(
            Salarie.tenant_id==t.id, Salarie.statut=="ACTIF",
            Salarie.id.in_(ids_sal)
        ).order_by(Salarie.nom).all()
        journaliers_list = Journalier.query.filter(
            Journalier.tenant_id==t.id, Journalier.statut=="ACTIF",
            Journalier.id.in_(ids_jour)
        ).order_by(Journalier.nom).all()
    else:
        salaries_list    = Salarie.query.filter_by(tenant_id=t.id, statut="ACTIF").order_by(Salarie.nom).all()
        journaliers_list = Journalier.query.filter_by(tenant_id=t.id, statut="ACTIF").order_by(Journalier.nom).all()

    pts_salaries    = {p.salarie_id:    p for p in Pointage.query.filter_by(tenant_id=t.id, date_pointage=date_sel).filter(Pointage.salarie_id.isnot(None)).all()}
    pts_journaliers = {p.journalier_id: p for p in Pointage.query.filter_by(tenant_id=t.id, date_pointage=date_sel).filter(Pointage.journalier_id.isnot(None)).all()}
    nb_presents_sal  = sum(1 for p in pts_salaries.values()    if p.present)
    nb_presents_jour = sum(1 for p in pts_journaliers.values() if p.present)
    nb_absents       = sum(1 for p in list(pts_salaries.values())+list(pts_journaliers.values()) if p.absent)
    lundi   = date_sel - timedelta(days=date_sel.weekday())
    semaine = [lundi + timedelta(days=i) for i in range(6)]

    # Affectation site de chaque travailleur pour affichage dans le pointage
    aff_sal  = {a.salarie_id:    a.site for a in AffectationSite.query.filter_by(tenant_id=t.id, actif=True).filter(AffectationSite.salarie_id.isnot(None)).all()}
    aff_jour = {a.journalier_id: a.site for a in AffectationSite.query.filter_by(tenant_id=t.id, actif=True).filter(AffectationSite.journalier_id.isnot(None)).all()}

    return render_template("tenant/pointage.html",
        tenant=t, date_sel=date_sel, semaine=semaine,
        date_hier=(date_sel  - timedelta(days=1)).strftime("%Y-%m-%d"),
        date_demain=(date_sel + timedelta(days=1)).strftime("%Y-%m-%d"),
        salaries=salaries_list, journaliers=journaliers_list,
        pts_salaries=pts_salaries, pts_journaliers=pts_journaliers,
        nb_presents_sal=nb_presents_sal, nb_presents_jour=nb_presents_jour,
        nb_absents=nb_absents, now=now,
        sites=sites_list, site_filtre=site_filtre,
        aff_sal=aff_sal, aff_jour=aff_jour)

@bp.route("/pointage/individuel", methods=["GET","POST"])
@login_required
def pointage_individuel():
    """Pointage d'un seul salarié ou journalier."""
    if current_user.is_super_admin: return redirect(url_for("admin.admin_dashboard"))
    t = get_tenant()
    if not t: return redirect(url_for("auth.login"))

    date_str  = request.args.get("date", datetime.now().strftime("%Y-%m-%d"))
    type_w    = request.args.get("type", "sal")   # "sal" ou "jour"
    worker_id = request.args.get("id", type=int)

    try:   date_sel = datetime.strptime(date_str, "%Y-%m-%d").date()
    except: date_sel = datetime.now().date()

    if request.method == "POST":
        # Sauvegarder le pointage individuel
        date_p = datetime.strptime(
            request.form.get("date_pointage", date_str), "%Y-%m-%d").date()
        wtype  = request.form.get("worker_type", "sal")
        wid    = request.form.get("worker_id", type=int)

        # Validation stricte : le travailleur doit exister ET appartenir au tenant
        if wtype == "sal":
            worker_obj = Salarie.query.filter_by(id=wid, tenant_id=t.id).first()
        else:
            worker_obj = Journalier.query.filter_by(id=wid, tenant_id=t.id).first()
        if not worker_obj:
            flash("Travailleur introuvable ou non autorisé.", "error")
            return redirect(url_for("tenant.pointage"))

        present = request.form.get("present") == "1"
        absent  = not present
        def _hm(val):
            """Valider et retourner un horaire HH:MM ou None."""
            v = (val or "").strip()
            if not v: return None
            import re
            return v if re.match(r"^\d{1,2}:\d{2}$", v) else None

        def _diff_hm(debut, fin):
            """Calculer la différence en heures entre deux horaires HH:MM."""
            try:
                h1,m1 = map(int, debut.split(":")); h2,m2 = map(int, fin.split(":"))
                diff = (h2*60+m2 - h1*60-m1) / 60
                return max(0.0, round(diff, 2))
            except: return 0.0

        # Horaires saisis
        em = _hm(request.form.get("entree_matin"))
        sm = _hm(request.form.get("sortie_matin"))
        ea = _hm(request.form.get("entree_apmidi"))
        sa = _hm(request.form.get("sortie_apmidi"))
        es = _hm(request.form.get("entree_sup"))
        ss = _hm(request.form.get("sortie_sup"))

        # Calcul auto des heures normales depuis les horaires
        h_normales_auto = 0.0
        if em and sm: h_normales_auto += _diff_hm(em, sm)
        if ea and sa: h_normales_auto += _diff_hm(ea, sa)
        h_normales_man = float(request.form.get("heures_normales", 0) or 0)
        # Priorité aux horaires si saisis, sinon saisie manuelle
        heures_normales_final = round(h_normales_auto, 2) if h_normales_auto > 0 else h_normales_man or 8

        # Heures sup depuis horaires
        h_sup_horaire = _diff_hm(es, ss) if (es and ss) else 0.0

        type_jour = request.form.get("type_jour", "NORMAL")

        # Reclasser les heures selon le type de jour
        h_sup_10_man = float(request.form.get("heures_sup_10",0) or 0)
        h_sup_30_man = float(request.form.get("heures_sup_30",0) or 0)
        h_sup_30b_man = float(request.form.get("heures_sup_30b",0) or 0)
        h_sup_40_man = float(request.form.get("heures_sup_40",0) or 0)
        h_sup_70_man = float(request.form.get("heures_sup_70",0) or 0)
        h_sup_30b_final = 0
        _conv_t = (t.convention or "").upper()

        if type_jour == "DIMANCHE":
            if _conv_t in ("PETROLE", "INDUSTRIE"):
                # Pétrole / Industrie : dimanche de jour → case 30b ; la nuit est
                # recalculée par la ventilation mensuelle.
                h_sup_30b_final = round(h_sup_horaire + heures_normales_final, 2)
                h_sup_70_final = 0
            else:
                # Dimanche travaillé : intégralité en +70% (réglementation BTP Gabon)
                h_sup_70_final = round(h_sup_horaire + heures_normales_final, 2)
            h_sup_10_final = 0; h_sup_30_final = 0; h_sup_40_final = 0
            heures_normales_final = 0
        elif type_jour == "FERIE":
            if _conv_t in ("PETROLE", "INDUSTRIE"):
                h_sup_30b_final = round(h_sup_horaire + heures_normales_final, 2)
                h_sup_70_final = 0
            else:
                # Tout va en +70% (jour férié)
                h_sup_70_final = round(h_sup_horaire + heures_normales_final, 2)
            h_sup_10_final = 0; h_sup_30_final = 0; h_sup_40_final = 0
            heures_normales_final = 0
        elif type_jour in ("CHOME_PAYE", "CHOME_RECUPERABLE"):
            # Présent mais jour chômé → heures normales conservées, pas de sup
            h_sup_10_final = 0; h_sup_30_final = 0; h_sup_40_final = 0; h_sup_70_final = 0
        else:
            # NORMAL
            h_sup_10_final = round(h_sup_horaire, 2) if h_sup_horaire > 0 else h_sup_10_man
            h_sup_30_final = h_sup_30_man
            h_sup_30b_final = h_sup_30b_man
            h_sup_40_final = h_sup_40_man
            h_sup_70_final = h_sup_70_man

        kwargs = dict(
            heures_normales = heures_normales_final,
            motif_absence   = request.form.get("motif_absence","") if absent else None,
            entree_matin    = em, sortie_matin  = sm,
            entree_apmidi   = ea, sortie_apmidi = sa,
            entree_sup      = es, sortie_sup    = ss,
            type_jour       = type_jour,
        )
        if wtype == "sal":
            kwargs.update(dict(
                heures_sup_10 = h_sup_10_final,
                heures_sup_30 = h_sup_30_final,
                heures_sup_30b = h_sup_30b_final,
                heures_sup_40 = h_sup_40_final,
                heures_sup_70 = h_sup_70_final,
            ))
            pt = Pointage.query.filter_by(
                tenant_id=t.id, date_pointage=date_p, salarie_id=wid).first()
            if not pt:
                pt = Pointage(tenant_id=t.id, date_pointage=date_p, salarie_id=wid)
                db.session.add(pt)
        else:
            # Journalier : gérer heures_sup selon type_jour
            if type_jour in ("DIMANCHE", "FERIE"):
                kwargs["heures_sup"]      = round(h_sup_horaire + heures_normales_final, 2)
                kwargs["heures_normales"] = 0
            else:
                kwargs["heures_sup"] = float(request.form.get("heures_sup",0) or 0)
            pt = Pointage.query.filter_by(
                tenant_id=t.id, date_pointage=date_p, journalier_id=wid).first()
            if not pt:
                pt = Pointage(tenant_id=t.id, date_pointage=date_p, journalier_id=wid)
                db.session.add(pt)
        pt.present = present
        pt.absent  = absent
        for k, v in kwargs.items():
            setattr(pt, k, v)
        db.session.commit()
        worker_name = worker_obj.nom_complet
        flash(f"✅ Pointage de {worker_name} enregistré.", "success")
        # Rester sur la même page pour pointer la personne suivante
        redir = request.form.get("next_url") or f"/pointage/individuel?date={date_p}&type={wtype}&id={wid}"
        return redirect(redir)

    # GET : charger le travailleur sélectionné
    worker = pt_existant = None
    historique_30j = []
    stats_30j = {"presences": 0, "absences": 0, "h_normales": 0.0,
                 "h_sup": 0.0, "taux": 0}

    if worker_id:
        if type_w == "sal":
            worker = Salarie.query.filter_by(id=worker_id, tenant_id=t.id).first()
            pt_existant = Pointage.query.filter_by(
                tenant_id=t.id, date_pointage=date_sel, salarie_id=worker_id).first()
            if worker:
                date_debut_30 = date_sel - timedelta(days=29)
                historique_30j = Pointage.query.filter_by(
                    tenant_id=t.id, salarie_id=worker_id
                ).filter(
                    Pointage.date_pointage >= date_debut_30,
                    Pointage.date_pointage <= date_sel
                ).order_by(Pointage.date_pointage.desc()).all()
        else:
            worker = Journalier.query.filter_by(id=worker_id, tenant_id=t.id).first()
            pt_existant = Pointage.query.filter_by(
                tenant_id=t.id, date_pointage=date_sel, journalier_id=worker_id).first()
            if worker:
                date_debut_30 = date_sel - timedelta(days=29)
                historique_30j = Pointage.query.filter_by(
                    tenant_id=t.id, journalier_id=worker_id
                ).filter(
                    Pointage.date_pointage >= date_debut_30,
                    Pointage.date_pointage <= date_sel
                ).order_by(Pointage.date_pointage.desc()).all()

        # Stats sur les 30 jours
        if historique_30j:
            nb_p = sum(1 for p in historique_30j if p.present)
            nb_a = sum(1 for p in historique_30j if p.absent)
            hn   = round(sum(float(p.heures_normales or 0) for p in historique_30j if p.present), 1)
            if type_w == "sal":
                hs = round(sum(
                    float(p.heures_sup_10 or 0) + float(p.heures_sup_30 or 0) +
                    float(p.heures_sup_40 or 0) + float(p.heures_sup_70 or 0)
                    for p in historique_30j if p.present), 1)
            else:
                hs = round(sum(float(p.heures_sup or 0) for p in historique_30j if p.present), 1)
            total_ptg = nb_p + nb_a
            stats_30j = {
                "presences": nb_p, "absences": nb_a,
                "h_normales": hn, "h_sup": hs,
                "taux": round(nb_p / total_ptg * 100) if total_ptg > 0 else 0
            }

    # Listes pour la recherche
    salaries_list    = Salarie.query.filter_by(
        tenant_id=t.id, statut="ACTIF").order_by(Salarie.nom).all()
    journaliers_list = Journalier.query.filter_by(
        tenant_id=t.id, statut="ACTIF").order_by(Journalier.nom).all()

    return render_template("tenant/pointage_individuel.html",
        tenant=t, date_sel=date_sel,
        date_hier=(date_sel - timedelta(days=1)).strftime("%Y-%m-%d"),
        date_demain=(date_sel + timedelta(days=1)).strftime("%Y-%m-%d"),
        type_w=type_w, worker=worker, pt_existant=pt_existant,
        historique_30j=historique_30j, stats_30j=stats_30j,
        salaries=salaries_list, journaliers=journaliers_list,
        now=datetime.now())

@bp.route("/pointage/supprimer/<int:ptg_id>", methods=["POST"])
@login_required
def pointage_supprimer(ptg_id):
    """Supprimer un pointage individuel."""
    if current_user.is_super_admin: return redirect(url_for("admin.admin_dashboard"))
    t = get_tenant()
    if not t: return redirect(url_for("auth.login"))
    pt = Pointage.query.filter_by(id=ptg_id, tenant_id=t.id).first_or_404()
    # Mémoriser le contexte pour rediriger au bon endroit
    next_url = request.form.get("next_url", "/pointage")
    date_str = str(pt.date_pointage)
    type_w   = "sal" if pt.salarie_id else "jour"
    wid      = pt.salarie_id or pt.journalier_id
    db.session.delete(pt)
    db.session.commit()
    log_action("DELETE", "pointage", ptg_id, f"Suppression pointage du {date_str}")
    db.session.commit()
    flash("🗑️ Pointage supprimé.", "success")
    return redirect(next_url or f"/pointage/individuel?date={date_str}&type={type_w}&id={wid}")

@bp.route("/pointage/sauvegarder", methods=["POST"])
@login_required
def pointage_sauvegarder():
    t = get_tenant()
    if not t: return redirect(url_for("auth.login"))
    date_str = request.form.get("date_pointage")
    try: date_p = datetime.strptime(date_str, "%Y-%m-%d").date()
    except: flash("Date invalide.", "error"); return redirect(url_for("tenant.pointage"))
    nb = 0
    sel_sal  = {v for k,v in request.form.items() if k.startswith("sel_sal_")}
    sel_jour = {v for k,v in request.form.items() if k.startswith("sel_jour_")}
    for key, val in request.form.items():
        if key.startswith("sal_present_"):
            sid_str = key.replace("sal_present_","")
            if sid_str not in sel_sal: continue
            sid = int(sid_str); present = val == "1"; absent = not present
            pt = Pointage.query.filter_by(tenant_id=t.id, date_pointage=date_p, salarie_id=sid).first()
            if not pt: pt = Pointage(tenant_id=t.id, date_pointage=date_p, salarie_id=sid); db.session.add(pt)
            pt.present=present; pt.absent=absent
            pt.heures_normales = float(request.form.get(f"sal_heures_{sid}", 8) or 8)
            pt.heures_sup_10   = float(request.form.get(f"sal_sup10_{sid}", 0) or 0)
            pt.heures_sup_30   = float(request.form.get(f"sal_sup30_{sid}", 0) or 0)
            pt.heures_sup_30b  = float(request.form.get(f"sal_sup30b_{sid}", 0) or 0)
            pt.heures_sup_40   = float(request.form.get(f"sal_sup40_{sid}", 0) or 0)
            pt.heures_sup_70   = float(request.form.get(f"sal_sup70_{sid}", 0) or 0)
            pt.motif_absence   = request.form.get(f"sal_motif_{sid}", "") if absent else None
            nb += 1
        if key.startswith("jour_present_"):
            jid_str = key.replace("jour_present_","")
            if jid_str not in sel_jour: continue
            jid = int(jid_str); present = val == "1"; absent = not present
            pt = Pointage.query.filter_by(tenant_id=t.id, date_pointage=date_p, journalier_id=jid).first()
            if not pt: pt = Pointage(tenant_id=t.id, date_pointage=date_p, journalier_id=jid); db.session.add(pt)
            pt.present=present; pt.absent=absent
            pt.heures_normales = float(request.form.get(f"jour_heures_{jid}", 8) or 8)
            pt.heures_sup      = float(request.form.get(f"jour_sup_{jid}", 0) or 0)
            pt.motif_absence   = request.form.get(f"jour_motif_{jid}", "") if absent else None
            nb += 1
    db.session.commit()
    if nb:
        log_action("UPDATE", "pointage", None,
                   f"Pointage du {date_p.strftime('%d/%m/%Y')} enregistré ({nb} ligne(s))")
        db.session.commit()
    flash(f"Pointage du {date_p.strftime('%d/%m/%Y')} sauvegardé ({nb} lignes).", "success")
    return redirect(url_for("tenant.pointage", date=date_str))

# ── Paie journaliers ──────────────────────────────────────────────────────────
@bp.route("/journaliers/paie")
@login_required
def journaliers_paie():
    if current_user.is_super_admin: return redirect(url_for("admin.admin_dashboard"))
    t = get_tenant()
    if not t: return redirect(url_for("auth.login"))

    # ── Filtre par site ──────────────────────────────────────────────────────
    site_filtre_id = request.args.get("site_id", type=int)
    sites_list     = Site.query.filter_by(tenant_id=t.id, actif=True).order_by(Site.nom).all()
    site_filtre    = Site.query.filter_by(id=site_filtre_id, tenant_id=t.id).first() if site_filtre_id else None
    statut_filtre  = request.args.get("statut", "")

    if site_filtre_id:
        ids_jour = [a.journalier_id for a in AffectationSite.query.filter_by(
            tenant_id=t.id, site_id=site_filtre_id, actif=True
        ).filter(AffectationSite.journalier_id.isnot(None)).all()]
        journaliers_list = Journalier.query.filter(
            Journalier.tenant_id==t.id, Journalier.statut=="ACTIF",
            Journalier.id.in_(ids_jour)
        ).order_by(Journalier.nom).all()
    else:
        journaliers_list = Journalier.query.filter_by(tenant_id=t.id, statut="ACTIF").order_by(Journalier.nom).all()

    # ── Feuilles filtrées ────────────────────────────────────────────────────
    q_feuilles = FeuillePaieJournalier.query.filter_by(tenant_id=t.id)
    if site_filtre_id:
        ids_jour_all = [a.journalier_id for a in AffectationSite.query.filter_by(
            tenant_id=t.id, site_id=site_filtre_id
        ).filter(AffectationSite.journalier_id.isnot(None)).all()]
        q_feuilles = q_feuilles.filter(FeuillePaieJournalier.journalier_id.in_(ids_jour_all))
    if statut_filtre:
        q_feuilles = q_feuilles.filter_by(statut=statut_filtre)
    page_f      = request.args.get("page", 1, type=int)
    # KPIs sur toutes les feuilles (sans pagination)
    feuilles_tous    = q_feuilles.order_by(FeuillePaieJournalier.date_fin.desc()).all()
    total_en_attente = sum(float(f.montant_brut or 0) for f in feuilles_tous if f.statut == "EN_ATTENTE")
    total_paye       = sum(float(f.montant_brut or 0) for f in feuilles_tous if f.statut == "PAYÉ")
    nb_en_attente    = sum(1 for f in feuilles_tous if f.statut == "EN_ATTENTE")
    q_feuilles = q_feuilles.options(joinedload(FeuillePaieJournalier.journalier))
    pagination_f     = q_feuilles.order_by(FeuillePaieJournalier.date_fin.desc()).paginate(page=page_f, per_page=25, error_out=False)
    feuilles         = pagination_f.items

    # Déduction des avances pour les feuilles affichées
    imput_av = _imputer_avances_journalier(
        feuilles, _avances_par_journalier(t.id, {f.journalier_id for f in feuilles}))

    # Affectation site de chaque journalier (pour affichage dans la liste)
    aff_jour = {a.journalier_id: a.site for a in AffectationSite.query.filter_by(
        tenant_id=t.id, actif=True
    ).filter(AffectationSite.journalier_id.isnot(None)).all()}

    _args = {k: v for k, v in request.args.items() if k != 'page'}
    _base = request.path + '?' + '&'.join(f'{k}={v}' for k, v in _args.items())
    _sep  = '&' if _args else '?'
    return render_template("tenant/journaliers_paie.html",
        tenant=t, feuilles=feuilles, journaliers=journaliers_list,
        sites=sites_list, site_filtre=site_filtre, statut_filtre=statut_filtre,
        total_en_attente=total_en_attente, total_paye=total_paye,
        nb_en_attente=nb_en_attente, aff_jour=aff_jour, imput_av=imput_av,
        pagination=pagination_f, pagination_base=_base + _sep,
        now=datetime.now(), today=date.today().isoformat())

@bp.route("/journaliers/paie/generer", methods=["POST"])
@login_required
def journaliers_paie_generer():
    t = get_tenant()
    if not t: return redirect(url_for("auth.login"))
    date_debut = _parse_date(request.form.get("date_debut"))
    date_fin   = _parse_date(request.form.get("date_fin"))
    if not date_debut or not date_fin: flash("Dates invalides.", "error"); return redirect(url_for("tenant.journaliers_paie"))
    site_id      = request.form.get("site_id", type=int)
    taux_custom  = {}  # taux personnalisés par journalier
    heures_custom = {} # heures manuelles par journalier
    ids_coches   = request.form.getlist("journalier_ids")

    for key, val in request.form.items():
        if key.startswith("taux_") and val:
            try: taux_custom[int(key[5:])] = float(val)
            except: pass
        if key.startswith("heures_") and val:
            try: heures_custom[int(key[7:])] = float(val)
            except: pass

    if ids_coches:
        journaliers_a_payer = Journalier.query.filter(
            Journalier.tenant_id==t.id,
            Journalier.id.in_([int(i) for i in ids_coches])
        ).all()
    elif site_id:
        ids_site = [a.journalier_id for a in AffectationSite.query.filter_by(
            tenant_id=t.id, site_id=site_id, actif=True
        ).filter(AffectationSite.journalier_id.isnot(None)).all()]
        journaliers_a_payer = Journalier.query.filter(
            Journalier.tenant_id==t.id, Journalier.statut=="ACTIF",
            Journalier.id.in_(ids_site)
        ).all()
    else:
        journaliers_a_payer = Journalier.query.filter_by(tenant_id=t.id, statut="ACTIF").all()

    nb = 0
    for j in journaliers_a_payer:
        if str(j.id) not in ids_coches and ids_coches:
            continue
        taux = taux_custom.get(j.id, float(j.taux_horaire or 0))
        if j.id in heures_custom:
            total_h  = heures_custom[j.id]
            nb_jours = 1  # heures manuelles = considéré comme 1 entrée
        else:
            pts = Pointage.query.filter_by(tenant_id=t.id, journalier_id=j.id)                  .filter(Pointage.date_pointage>=date_debut, Pointage.date_pointage<=date_fin,
                          Pointage.present==True).all()
            total_h  = sum(float(p.heures_normales or 0)+float(p.heures_sup or 0) for p in pts)
            nb_jours = len(pts)
        if total_h <= 0 and nb_jours == 0: continue
        if FeuillePaieJournalier.query.filter_by(
            tenant_id=t.id, journalier_id=j.id,
            date_debut=date_debut, date_fin=date_fin).first(): continue
        db.session.add(FeuillePaieJournalier(
            tenant_id=t.id, journalier_id=j.id,
            date_debut=date_debut, date_fin=date_fin,
            nb_jours=nb_jours, total_heures=total_h,
            taux_horaire=taux, montant_brut=round(total_h*taux, 2),
            statut="EN_ATTENTE"))
        nb += 1
    db.session.commit()
    flash(f"{nb} feuille(s) générée(s).", "success")
    redirect_url = url_for("tenant.journaliers_paie")
    if site_id:
        redirect_url += f"?site_id={site_id}"
    return redirect(redirect_url)

@bp.route("/journaliers/paie/generer-mois", methods=["POST"])
@login_required
def journaliers_paie_generer_mois():
    """Génère la paie de FIN DE MOIS pour tous les journaliers de type MENSUEL.

    La période couvre le mois entier (1er → dernier jour). Le montant est calculé
    à partir des pointages présents du mois × taux horaire.
    """
    t = get_tenant()
    if not t: return redirect(url_for("auth.login"))
    import calendar as _cal
    mois  = request.form.get("mois", type=int)
    annee = request.form.get("annee", type=int)
    if not mois or not annee:
        flash("Mois/année manquant.", "error")
        return redirect(url_for("tenant.journaliers_paie"))
    date_debut = date(annee, mois, 1)
    date_fin   = date(annee, mois, _cal.monthrange(annee, mois)[1])

    mensuels = Journalier.query.filter_by(
        tenant_id=t.id, statut="ACTIF", type_paie="MENSUEL").all()
    nb = 0; nb_existant = 0
    for j in mensuels:
        if FeuillePaieJournalier.query.filter_by(
            tenant_id=t.id, journalier_id=j.id,
            date_debut=date_debut, date_fin=date_fin).first():
            nb_existant += 1
            continue
        pts = Pointage.query.filter_by(tenant_id=t.id, journalier_id=j.id).filter(
            Pointage.date_pointage >= date_debut,
            Pointage.date_pointage <= date_fin,
            Pointage.present == True).all()
        total_h  = sum(float(p.heures_normales or 0) + float(p.heures_sup or 0) for p in pts)
        nb_jours = len(pts)
        if total_h <= 0 and nb_jours == 0:
            continue
        taux = float(j.taux_horaire or 0)
        from calculs_paie import arrondi_millier_superieur
        brut = arrondi_millier_superieur(total_h * taux)   # arrondi au millier de F supérieur
        db.session.add(FeuillePaieJournalier(
            tenant_id=t.id, journalier_id=j.id,
            date_debut=date_debut, date_fin=date_fin,
            nb_jours=nb_jours, total_heures=total_h,
            taux_horaire=taux, montant_brut=brut,
            statut="EN_ATTENTE"))
        nb += 1
    db.session.commit()
    if nb:
        log_action("CREATE", "feuille_journalier", None,
                   f"Génération paie mensuelle : {nb} feuille(s) ({mois:02d}/{annee})")
        db.session.commit()
    msg = f"{nb} feuille(s) mensuelle(s) générée(s) pour {mois:02d}/{annee}."
    if nb_existant:
        msg += f" {nb_existant} déjà existante(s) ignorée(s)."
    if not mensuels:
        msg = "Aucun journalier de type « Mensuel » n'est défini."
    flash(msg, "success" if mensuels else "error")
    return redirect(url_for("tenant.journaliers_paie"))


def _avances_par_journalier(tenant_id, journalier_ids=None):
    """Charge les avances et les regroupe par journalier : {jid: [AvanceJournalier]}."""
    from collections import defaultdict
    q = AvanceJournalier.query.filter_by(tenant_id=tenant_id)
    if journalier_ids is not None:
        if not journalier_ids:
            return {}
        q = q.filter(AvanceJournalier.journalier_id.in_(list(journalier_ids)))
    par_j = defaultdict(list)
    for a in q.order_by(AvanceJournalier.date_avance).all():
        par_j[a.journalier_id].append(a)
    return par_j


def _imputer_avances_journalier(feuilles, avances_par_journalier):
    """Déduit les avances des feuilles de paie, par journalier.

    Les feuilles PAYÉ utilisent le montant d'avance figé (`avance_deduite`).
    Pour les feuilles EN_ATTENTE, l'encours d'avances non régularisées du
    journalier est réparti sur ses feuilles (la plus ancienne d'abord), plafonné
    au montant de chaque feuille.

    Renvoie {feuille_id: {"avance": x, "net": y}} (en XAF).
    """
    from collections import defaultdict
    encours = {}
    for jid, avs in avances_par_journalier.items():
        encours[jid] = round(sum(max(0.0, float(a.montant or 0) - float(a.montant_regularise or 0))
                                 for a in avs), 2)
    res = {}
    par_j = defaultdict(list)
    for f in feuilles:
        par_j[f.journalier_id].append(f)
    for jid, fs in par_j.items():
        rem = encours.get(jid, 0.0)
        for f in fs:                       # feuilles déjà payées : déduction figée
            if f.statut == "PAYÉ":
                ded = float(f.avance_deduite or 0)
                res[f.id] = {"avance": ded, "net": round(float(f.montant_a_payer) - ded, 2)}
        for f in sorted([f for f in fs if f.statut != "PAYÉ"],
                        key=lambda x: (x.date_fin, x.id)):
            base = float(f.montant_a_payer)
            ded  = round(min(rem, base), 2)
            rem  = round(rem - ded, 2)
            res[f.id] = {"avance": ded, "net": round(base - ded, 2)}
    return res


@bp.route("/journaliers/<int:id>/avances/nouvelle", methods=["POST"])
@login_required
def journalier_avance_nouvelle(id):
    """Enregistre une avance versée à un journalier."""
    t = get_tenant()
    if not t: return redirect(url_for("auth.login"))
    j = Journalier.query.filter_by(id=id, tenant_id=t.id).first_or_404()
    try:
        montant = float(request.form.get("montant", 0) or 0)
    except (TypeError, ValueError):
        montant = 0
    if montant <= 0:
        flash("Le montant de l'avance doit être supérieur à zéro.", "error")
        return redirect(url_for("tenant.journalier_detail", id=id))
    aff = AffectationSite.query.filter_by(journalier_id=id, tenant_id=t.id, actif=True).first()
    av = AvanceJournalier(
        tenant_id=t.id, journalier_id=j.id,
        site_id=(aff.site_id if aff else None),
        montant=montant,
        date_avance=_parse_date(request.form.get("date_avance", "")) or datetime.now().date(),
        mode_paiement=(request.form.get("mode_paiement", "ESPECES") or "ESPECES").strip(),
        reference=(request.form.get("reference", "") or "").strip() or None,
        motif=(request.form.get("motif", "") or "").strip() or None)
    db.session.add(av); db.session.commit()
    log_action("CREATE", "avance_journalier", av.id,
               f"Avance {montant:,.0f} F — {j.nom_complet}".replace(",", " "),
               apres=av.to_dict())
    db.session.commit()
    flash(f"Avance de {montant:,.0f} F enregistrée pour {j.nom_complet}.".replace(",", " "), "success")
    return redirect(url_for("tenant.journalier_detail", id=id))


@bp.route("/journaliers/avances/<int:aid>/modifier", methods=["POST"])
@login_required
def journalier_avance_modifier(aid):
    """Modifie une avance, tant qu'elle n'est ni validée ni déjà déduite."""
    t = get_tenant()
    if not t: return redirect(url_for("auth.login"))
    av = AvanceJournalier.query.filter_by(id=aid, tenant_id=t.id).first_or_404()
    jid = av.journalier_id
    if not av.est_modifiable:
        flash("Cette avance est validée ou déjà déduite : elle n'est plus modifiable.", "error")
        return redirect(url_for("tenant.journalier_detail", id=jid))
    try:
        montant = float(request.form.get("montant", av.montant) or av.montant)
    except (TypeError, ValueError):
        montant = float(av.montant)
    if montant <= 0:
        flash("Le montant de l'avance doit être supérieur à zéro.", "error")
        return redirect(url_for("tenant.journalier_detail", id=jid))
    av.montant = montant
    d = _parse_date(request.form.get("date_avance", ""))
    if d:
        av.date_avance = d
    av.motif = (request.form.get("motif", "") or "").strip() or None
    db.session.commit()
    log_action("UPDATE", "avance_journalier", av.id,
               f"Modification avance — {av.journalier.nom_complet}", apres=av.to_dict())
    db.session.commit()
    flash("Avance modifiée.", "success")
    return redirect(url_for("tenant.journalier_detail", id=jid))


@bp.route("/journaliers/avances/<int:aid>/valider", methods=["POST"])
@login_required
def journalier_avance_valider(aid):
    """Valide une avance : elle devient figée (plus modifiable ni supprimable)."""
    t = get_tenant()
    if not t: return redirect(url_for("auth.login"))
    av = AvanceJournalier.query.filter_by(id=aid, tenant_id=t.id).first_or_404()
    av.statut = "VALIDEE"
    db.session.commit()
    log_action("VALIDATE", "avance_journalier", av.id,
               f"Validation avance {float(av.montant):,.0f} F — {av.journalier.nom_complet}".replace(",", " "))
    db.session.commit()
    flash("Avance validée — elle est désormais figée.", "success")
    return redirect(url_for("tenant.journalier_detail", id=av.journalier_id))


@bp.route("/journaliers/avances/<int:aid>/supprimer", methods=["POST"])
@login_required
def journalier_avance_supprimer(aid):
    """Supprime une avance (uniquement si non validée et non encore déduite)."""
    t = get_tenant()
    if not t: return redirect(url_for("auth.login"))
    av = AvanceJournalier.query.filter_by(id=aid, tenant_id=t.id).first_or_404()
    jid = av.journalier_id
    if not av.est_modifiable:
        flash("Avance validée ou déjà déduite : suppression impossible.", "error")
        return redirect(url_for("tenant.journalier_detail", id=jid))
    db.session.delete(av); db.session.commit()
    log_action("DELETE", "avance_journalier", aid,
               f"Suppression avance {float(av.montant):,.0f} F".replace(",", " "), avant=av.to_dict())
    db.session.commit()
    flash("Avance supprimée.", "success")
    return redirect(url_for("tenant.journalier_detail", id=jid))


@bp.route("/journaliers/paie/<int:id>/payer", methods=["POST"])
@login_required
def journalier_payer(id):
    t = get_tenant()
    if not t: return redirect(url_for("auth.login"))
    f = FeuillePaieJournalier.query.filter_by(id=id, tenant_id=t.id).first_or_404()
    f.montant_brut = f.montant_a_payer   # fige l'arrondi millier sup. pour les mensuels
    # Déduction des avances non régularisées du journalier (plus ancienne d'abord)
    avs = (AvanceJournalier.query.filter_by(tenant_id=t.id, journalier_id=f.journalier_id)
           .order_by(AvanceJournalier.date_avance).all())
    a_deduire = round(min(
        sum(a.reste_a_regulariser for a in avs),
        float(f.montant_a_payer)), 2)
    f.avance_deduite = a_deduire
    reste = a_deduire
    for a in avs:
        if reste <= 0:
            break
        dispo = a.reste_a_regulariser
        if dispo <= 0:
            continue
        pris = min(dispo, reste)
        a.montant_regularise = float(a.montant_regularise or 0) + pris
        reste = round(reste - pris, 2)
    mode = (request.form.get("mode_paiement") or "").strip()
    if mode not in ("ESPECES", "VIREMENT"):
        mode = (f.journalier.mode_paiement if f.journalier else "ESPECES") or "ESPECES"
    f.mode_paiement = mode
    # Date de paiement : celle choisie dans le formulaire, sinon aujourd'hui.
    date_pmt = date.today()
    _df = (request.form.get("date_paiement") or "").strip()
    if _df:
        try:
            date_pmt = datetime.strptime(_df, "%Y-%m-%d").date()
        except ValueError:
            date_pmt = date.today()
    f.statut = "PAYÉ"; f.date_paiement = date_pmt; db.session.commit()
    net = float(f.montant_brut) - a_deduire
    log_action("PAY", "feuille_journalier", f.id,
               f"Paiement {f.journalier.nom_complet} — net {net:,.0f} F"
               + (f" (avances {a_deduire:,.0f} F)" if a_deduire else ""),
               apres={"montant_brut": float(f.montant_brut), "avance_deduite": a_deduire,
                      "net": net, "periode": f"{f.date_debut}→{f.date_fin}"})
    db.session.commit()
    # Interconnexion Caisse : proposer la sortie correspondante (ne bloque jamais).
    try:
        from interco_caisse import proposer_ecriture
        nom = f.journalier.nom_complet if f.journalier else "journalier"
        proposer_ecriture(
            t, source_ref=f"journalier-{f.id}",
            montant=float(net),
            motif=f"Paie journalier {nom} — {f.date_debut}→{f.date_fin}",
            compte_suggere="6611", date_operation=f.date_paiement)
    except Exception:
        pass
    if a_deduire > 0:
        flash(f"Paiement de {f.journalier.nom_complet} enregistré "
              f"(net {net:,.0f} F après {a_deduire:,.0f} F d'avances).".replace(",", " "), "success")
    else:
        flash(f"Paiement de {f.journalier.nom_complet} enregistré.", "success")
    return redirect(url_for("tenant.journaliers_paie"))

@bp.route("/journaliers/paie/<int:id>/modifier", methods=["POST"])
@login_required
def journalier_feuille_modifier(id):
    t = get_tenant()
    if not t: return redirect(url_for("auth.login"))
    f = FeuillePaieJournalier.query.filter_by(id=id, tenant_id=t.id).first_or_404()
    f.montant_brut = float(request.form.get("montant_brut", f.montant_brut) or f.montant_brut)
    f.observation  = request.form.get("observation", "").strip()
    db.session.commit(); flash("Feuille modifiée.", "success")
    return redirect(url_for("tenant.journaliers_paie"))

@bp.route("/journaliers/paie/<int:id>/supprimer", methods=["POST"])
@login_required
def journalier_feuille_supprimer(id):
    t = get_tenant()
    if not t: return redirect(url_for("auth.login"))
    f = FeuillePaieJournalier.query.filter_by(id=id, tenant_id=t.id).first_or_404()
    db.session.delete(f); db.session.commit()
    flash("Feuille supprimée.", "success")
    return redirect(url_for("tenant.journaliers_paie"))

@bp.route("/journaliers/pointage/modele")
@login_required
def journaliers_pointage_modele():
    """Génère un modèle Excel (grille mensuelle) pré-rempli pour l'import de
    pointages journaliers. Une ligne par journalier du site choisi, une colonne
    par jour du mois. Dimanches et fériés grisés. L'utilisateur saisit les
    heures travaillées ; il téléverse ensuite le fichier via /importer.
    Params : site_id (optionnel), mois (YYYY-MM, défaut = mois en cours).
    """
    if current_user.is_super_admin:
        return redirect(url_for("admin.admin_dashboard"))
    t = get_tenant()
    if not t:
        return redirect(url_for("auth.login"))

    import calendar as _cal
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from io import BytesIO

    # ── Période demandée ───────────────────────────────────────────────
    mois_str = request.args.get("mois", "")  # "2026-07"
    try:
        annee, mois = (int(x) for x in mois_str.split("-"))
        date(annee, mois, 1)  # validation
    except (ValueError, TypeError):
        today = datetime.now()
        annee, mois = today.year, today.month
    nb_jours = _cal.monthrange(annee, mois)[1]

    MOIS_FR = ["", "Janvier", "Février", "Mars", "Avril", "Mai", "Juin", "Juillet",
               "Août", "Septembre", "Octobre", "Novembre", "Décembre"]
    mois_nom = f"{MOIS_FR[mois]} {annee}"

    # ── Journaliers du site (ou tous les actifs si aucun site) ─────────
    site_id = request.args.get("site_id", type=int)
    site_obj = None
    if site_id:
        site_obj = Site.query.filter_by(id=site_id, tenant_id=t.id).first()
        ids_site = [a.journalier_id for a in AffectationSite.query.filter_by(
            tenant_id=t.id, site_id=site_id, actif=True
        ).filter(AffectationSite.journalier_id.isnot(None)).all()]
        journaliers = Journalier.query.filter(
            Journalier.tenant_id == t.id, Journalier.statut == "ACTIF",
            Journalier.id.in_(ids_site)
        ).order_by(Journalier.nom).all() if ids_site else []
    else:
        journaliers = Journalier.query.filter_by(
            tenant_id=t.id, statut="ACTIF").order_by(Journalier.nom).all()

    # ── Styles ─────────────────────────────────────────────────────────
    VERT, BLANC, GRIS_WE = "0F3D36", "FFFFFF", "F3F4F6"
    thin = Side(style="thin", color="D1D5DB")
    bord = Border(left=thin, right=thin, top=thin, bottom=thin)
    f_titre  = Font(name="Arial", size=12, bold=True, color=BLANC)
    f_entete = Font(name="Arial", size=9,  bold=True, color=BLANC)
    f_normal = Font(name="Arial", size=10)
    fill_vert = PatternFill("solid", fgColor=VERT)
    fill_we   = PatternFill("solid", fgColor=GRIS_WE)
    center = Alignment(horizontal="center", vertical="center")

    wb = Workbook()
    ws = wb.active
    ws.title = f"Pointage {MOIS_FR[mois][:3]} {annee}"
    ws.sheet_view.showGridLines = False

    n_cols = 2 + nb_jours  # ID + Nom + jours
    # Titre
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    tc = ws.cell(row=1, column=1,
                 value=f"POINTAGE JOURNALIERS — {mois_nom}"
                       + (f"  ·  {site_obj.nom}" if site_obj else "  ·  Tous sites"))
    tc.font = f_titre; tc.fill = fill_vert
    tc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 24

    # Légende
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    lc = ws.cell(row=2, column=1,
        value="Saisissez le nombre d'HEURES travaillées par jour. "
              "Vide = jour ignoré (non modifié). 0 = absent. "
              "Au-delà de 8h : le surplus compte en heures supplémentaires. "
              "Ne modifiez PAS la colonne ID.")
    lc.font = Font(name="Arial", size=9, italic=True, color="6B7280")
    lc.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[2].height = 40

    # En-têtes (ligne 3)
    for col, lbl in ((1, "ID"), (2, "Nom journalier")):
        c = ws.cell(row=3, column=col, value=lbl)
        c.font = f_entete; c.fill = fill_vert; c.alignment = center; c.border = bord
    for j in range(1, nb_jours + 1):
        d = date(annee, mois, j)
        c = ws.cell(row=3, column=2 + j, value=j)
        c.font = f_entete; c.fill = fill_vert; c.alignment = center; c.border = bord
        # Dimanche ou férié → en-tête rouge (repère visuel)
        if type_jour_auto(d) in ("DIMANCHE", "FERIE"):
            c.fill = PatternFill("solid", fgColor="B91C1C")
    ws.row_dimensions[3].height = 18

    # Lignes journaliers
    for r, j in enumerate(journaliers, start=4):
        cid = ws.cell(row=r, column=1, value=j.id)
        cid.font = f_normal; cid.alignment = center; cid.border = bord
        cnom = ws.cell(row=r, column=2, value=j.nom_complet)
        cnom.font = f_normal; cnom.border = bord
        cnom.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        for jour in range(1, nb_jours + 1):
            d = date(annee, mois, jour)
            c = ws.cell(row=r, column=2 + jour)
            c.font = f_normal; c.alignment = center; c.border = bord
            if type_jour_auto(d) in ("DIMANCHE", "FERIE"):
                c.fill = fill_we  # grisé : repère (mais saisissable, ex. dimanche travaillé)

    # Largeurs + gel des volets
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 24
    for j in range(1, nb_jours + 1):
        ws.column_dimensions[get_column_letter(2 + j)].width = 4.5
    ws.freeze_panes = "C4"

    # ── Sortie fichier ─────────────────────────────────────────────────
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"pointage_{annee}-{mois:02d}"
    if site_obj:
        nom_propre = "".join(c if c.isalnum() else "_" for c in site_obj.nom)
        fname += "_" + nom_propre
    return send_file(
        buf, as_attachment=True, download_name=f"{fname}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@bp.route("/journaliers/pointage/importer", methods=["POST"])
@login_required
def journaliers_pointage_importer():
    """Étape 1/2 — APERÇU. Lit la grille mensuelle téléversée, applique la règle
    journalier (≤8h normal / surplus en sup / dimanche-férié → tout en sup),
    et affiche ce qui SERA créé/modifié. N'écrit RIEN en base. Les données
    validées repartent dans un champ caché vers /confirmer.
    """
    if current_user.is_super_admin:
        return redirect(url_for("admin.admin_dashboard"))
    t = get_tenant()
    if not t:
        return redirect(url_for("auth.login"))

    import openpyxl, json
    import calendar as _cal

    fichier = request.files.get("fichier")
    if not fichier or not fichier.filename.endswith((".xlsx", ".xls")):
        flash("❌ Fichier invalide. Utilisez le modèle Excel fourni (.xlsx).", "error")
        return redirect(url_for("tenant.pointage"))

    # Garde anti-DoS : 5 Mo max (même règle que l'import salariés).
    fichier.seek(0, 2); _taille = fichier.tell(); fichier.seek(0)
    if _taille > 5_000_000:
        flash("❌ Fichier trop volumineux (max 5 Mo).", "error")
        return redirect(url_for("tenant.pointage"))

    try:
        wb = openpyxl.load_workbook(fichier, data_only=True)
        ws = wb.active
    except Exception as e:
        flash(f"❌ Erreur lecture fichier : {e}", "error")
        return redirect(url_for("tenant.pointage"))

    # ── Retrouver l'en-tête : la ligne qui contient "ID" en col.1 ──────
    header_row = None
    for r in range(1, 8):
        v = str(ws.cell(r, 1).value or "").upper().strip()
        if v == "ID":
            header_row = r
            break
    if not header_row:
        flash("❌ En-tête introuvable. Utilisez le modèle téléchargé (colonne ID).", "error")
        return redirect(url_for("tenant.pointage"))

    # ── Déduire l'année/mois à partir des colonnes de jours ────────────
    # Les colonnes 3..N portent les numéros de jour. On récupère aussi le
    # mois/année depuis le TITRE (ligne 1) si présent, sinon mois courant.
    titre = str(ws.cell(1, 1).value or "")
    annee = mois = None
    import re
    MOIS_MAP = {"JANVIER":1,"FÉVRIER":2,"FEVRIER":2,"MARS":3,"AVRIL":4,"MAI":5,
                "JUIN":6,"JUILLET":7,"AOÛT":8,"AOUT":8,"SEPTEMBRE":9,
                "OCTOBRE":10,"NOVEMBRE":11,"DÉCEMBRE":12,"DECEMBRE":12}
    for nom_m, num in MOIS_MAP.items():
        if nom_m in titre.upper():
            mois = num
            break
    m_an = re.search(r"(20\d{2})", titre)
    if m_an:
        annee = int(m_an.group(1))
    if not (annee and mois):
        now = datetime.now(); annee, mois = now.year, now.month
    nb_jours_mois = _cal.monthrange(annee, mois)[1]

    # ── Mapper les colonnes-jours : {col: numéro_de_jour} ──────────────
    col_jour = {}
    for c in range(3, ws.max_column + 1):
        val = ws.cell(header_row, c).value
        try:
            j = int(val)
            if 1 <= j <= nb_jours_mois:
                col_jour[c] = j
        except (TypeError, ValueError):
            continue
    if not col_jour:
        flash("❌ Aucune colonne de jour détectée. Utilisez le modèle téléchargé.", "error")
        return redirect(url_for("tenant.pointage"))

    # ── IDs de journaliers valides pour ce tenant (sécurité) ───────────
    ids_valides = {j.id: j for j in Journalier.query.filter_by(
        tenant_id=t.id).all()}

    # ── Parcours des lignes → construction de l'aperçu ─────────────────
    a_appliquer = []   # [{jid, nom, date, heures, hn, hs, present, type_jour, action}]
    lignes_ignorees = 0
    erreurs = []

    for r in range(header_row + 1, ws.max_row + 1):
        id_val = ws.cell(r, 1).value
        if id_val in (None, ""):
            continue
        try:
            jid = int(id_val)
        except (TypeError, ValueError):
            erreurs.append(f"Ligne {r} : ID « {id_val} » invalide.")
            continue
        j_obj = ids_valides.get(jid)
        if not j_obj:
            erreurs.append(f"Ligne {r} : journalier ID {jid} inconnu (ignoré).")
            continue

        for c, jour in col_jour.items():
            case = ws.cell(r, c).value
            # Case VIDE → on ignore ce jour (ne touche à rien).
            if case in (None, ""):
                continue
            try:
                heures = float(str(case).replace(",", "."))
            except (TypeError, ValueError):
                erreurs.append(f"Ligne {r}, jour {jour} : valeur « {case} » non numérique.")
                continue
            if heures < 0:
                erreurs.append(f"Ligne {r}, jour {jour} : heures négatives.")
                continue

            d = date(annee, mois, jour)
            tj = type_jour_auto(d)  # NORMAL | DIMANCHE | FERIE

            # ── Règle journalier ──────────────────────────────────────
            if heures == 0:
                present = False; hn = 0.0; hs = 0.0
            else:
                present = True
                if tj in ("DIMANCHE", "FERIE"):
                    hn = 0.0; hs = round(heures, 2)         # tout majoré
                elif heures > 8:
                    hn = 8.0; hs = round(heures - 8, 2)     # surplus en sup
                else:
                    hn = round(heures, 2); hs = 0.0

            # Existe déjà ? (pour l'étiquette créé / mis à jour)
            existe = Pointage.query.filter_by(
                tenant_id=t.id, date_pointage=d, journalier_id=jid).first() is not None

            a_appliquer.append({
                "jid": jid, "nom": j_obj.nom_complet,
                "date": d.strftime("%Y-%m-%d"), "jour": jour,
                "heures": heures, "hn": hn, "hs": hs,
                "present": present, "type_jour": tj,
                "action": "maj" if existe else "creer",
            })

    if not a_appliquer:
        flash("Aucune donnée exploitable dans le fichier (toutes les cases sont vides ?).", "warning")
        return redirect(url_for("tenant.pointage"))

    nb_creer = sum(1 for x in a_appliquer if x["action"] == "creer")
    nb_maj   = sum(1 for x in a_appliquer if x["action"] == "maj")

    return render_template("tenant/pointage_import_apercu.html",
        tenant=t, lignes=a_appliquer, nb_creer=nb_creer, nb_maj=nb_maj,
        nb_total=len(a_appliquer), erreurs=erreurs,
        annee=annee, mois=mois,
        payload=json.dumps(a_appliquer))

@bp.route("/journaliers/pointage/importer/confirmer", methods=["POST"])
@login_required
def journaliers_pointage_importer_confirmer():
    """Étape 2/2 — ENREGISTREMENT. Reçoit le JSON validé (champ caché de
    l'aperçu) et applique update-or-create sur chaque Pointage.
    """
    if current_user.is_super_admin:
        return redirect(url_for("admin.admin_dashboard"))
    t = get_tenant()
    if not t:
        return redirect(url_for("auth.login"))

    import json
    try:
        lignes = json.loads(request.form.get("payload", "[]"))
    except (ValueError, TypeError):
        flash("❌ Données d'import illisibles. Recommencez le téléversement.", "error")
        return redirect(url_for("tenant.pointage"))

    if not lignes:
        flash("Aucune donnée à enregistrer.", "warning")
        return redirect(url_for("tenant.pointage"))

    # IDs valides (re-vérif sécurité : on ne fait jamais confiance au client).
    ids_valides = {j.id for j in Journalier.query.filter_by(tenant_id=t.id).all()}

    nb_creer = nb_maj = 0
    for x in lignes:
        try:
            jid = int(x["jid"])
            d   = _parse_date(x["date"])
            hn  = float(x["hn"]); hs = float(x["hs"])
            present = bool(x["present"]); tj = str(x["type_jour"])
        except (KeyError, ValueError, TypeError):
            continue
        if jid not in ids_valides or not d:
            continue

        pt = Pointage.query.filter_by(
            tenant_id=t.id, date_pointage=d, journalier_id=jid).first()
        if pt:
            nb_maj += 1
        else:
            pt = Pointage(tenant_id=t.id, date_pointage=d, journalier_id=jid)
            db.session.add(pt)
            nb_creer += 1

        pt.present         = present
        pt.absent          = not present
        pt.heures_normales = hn
        pt.heures_sup      = hs
        pt.type_jour       = tj

    db.session.commit()
    log_action("IMPORT", "pointage", None,
               f"Import pointages journaliers : {nb_creer} créé(s), {nb_maj} mis à jour")
    flash(f"✅ Import terminé : {nb_creer} pointage(s) créé(s), {nb_maj} mis à jour.", "success")
    return redirect(url_for("tenant.pointage"))

@bp.route("/journaliers/paie/imprimer-sites")
@login_required
def journaliers_paie_imprimer_sites():
    """Impression de la paie journalier GROUPÉE PAR SITE (une page par site).

    Pensée pour distribuer aux chefs de chantier : chaque site a sa propre page
    (saut de page), avec le détail des journaliers, une colonne signature et un
    sous-total. Une page récapitulative tous sites termine le document.
    Filtre par période obligatoire côté formulaire : date_debut / date_fin.
    """
    if current_user.is_super_admin: return redirect(url_for("admin.admin_dashboard"))
    t = get_tenant()
    if not t: return redirect(url_for("auth.login"))
    statut_f   = request.args.get("statut", "")
    date_debut = _parse_date(request.args.get("date_debut", ""))
    date_fin   = _parse_date(request.args.get("date_fin", ""))

    # Feuilles de la période (et statut éventuel). On retient toute feuille qui
    # CHEVAUCHE la période choisie (plus tolérant qu'une inclusion stricte).
    q = FeuillePaieJournalier.query.filter_by(tenant_id=t.id)
    if statut_f:
        q = q.filter_by(statut=statut_f)
    if date_debut:
        q = q.filter(FeuillePaieJournalier.date_fin >= date_debut)
    if date_fin:
        q = q.filter(FeuillePaieJournalier.date_debut <= date_fin)
    feuilles = q.options(joinedload(FeuillePaieJournalier.journalier)).order_by(
        FeuillePaieJournalier.date_fin.desc()).all()
    # ── Sélection optionnelle des journaliers à imprimer ──────────────
    # Le front peut transmettre ?ids=3,7,12 pour ne sortir qu'une partie des
    # journaliers. Absent ou vide → comportement historique : tout le monde.
    ids_param = request.args.get("ids", "").strip()
    if ids_param:
        try:
            ids_voulus = {int(x) for x in ids_param.split(",") if x.strip().isdigit()}
        except ValueError:
            ids_voulus = set()
        if ids_voulus:
            feuilles = [f for f in feuilles if f.journalier_id in ids_voulus]

    # Site de chaque journalier : on prend TOUTES les affectations, en préférant
    # l'affectation active, puis la plus récente. Ainsi un journalier dont
    # l'affectation n'est pas marquée "active" n'est pas perdu, et seuls les
    # journaliers réellement sans aucune affectation tombent dans "Sans site".
    site_par_journalier = {}
    affs = (AffectationSite.query
            .filter_by(tenant_id=t.id)
            .filter(AffectationSite.journalier_id.isnot(None))
            .order_by(AffectationSite.actif.desc(),
                      AffectationSite.date_debut.desc())
            .all())
    for a in affs:
        if a.journalier_id not in site_par_journalier and a.site is not None:
            site_par_journalier[a.journalier_id] = a.site

    sites_list = Site.query.filter_by(tenant_id=t.id, actif=True).order_by(Site.nom).all()

    # Déduction des avances (par journalier) sur les feuilles affichées
    imput_av = _imputer_avances_journalier(
        feuilles, _avances_par_journalier(t.id, {f.journalier_id for f in feuilles}))

    # Regroupement : { site_id : {"site": Site|None, "feuilles": [...], "total": x} }
    groupes = {}
    SANS_SITE = 0
    for f in feuilles:
        s = site_par_journalier.get(f.journalier_id)
        key = s.id if s else SANS_SITE
        if key not in groupes:
            groupes[key] = {"site": s, "feuilles": [], "total": 0.0,
                            "total_brut": 0.0, "total_avance": 0.0,
                            "recap_mode": {"ESPECES": {"total": 0.0, "nb": 0},
                                           "VIREMENT": {"total": 0.0, "nb": 0}}}
        groupes[key]["feuilles"].append(f)
        net = imput_av.get(f.id, {}).get("net", float(f.montant_a_payer or 0))
        groupes[key]["total"]       += net
        groupes[key]["total_brut"]  += float(f.montant_a_payer or 0)
        groupes[key]["total_avance"]+= imput_av.get(f.id, {}).get("avance", 0.0)
        mode = (f.mode_paiement if f.statut == "PAYÉ" and f.mode_paiement
                else (f.journalier.mode_paiement if f.journalier else "ESPECES")) or "ESPECES"
        if mode not in ("ESPECES", "VIREMENT"):
            mode = "ESPECES"
        groupes[key]["recap_mode"][mode]["total"] += net
        groupes[key]["recap_mode"][mode]["nb"]    += 1

    # Ordonner : sites par nom (selon sites_list), puis "Sans site" à la fin
    groupes_ordonnes = []
    for s in sites_list:
        if s.id in groupes:
            groupes_ordonnes.append(groupes[s.id])
    if SANS_SITE in groupes:
        groupes_ordonnes.append(groupes[SANS_SITE])

    total_general = sum(g["total"] for g in groupes_ordonnes)
    nb_total = sum(len(g["feuilles"]) for g in groupes_ordonnes)
    # Récap global par mode (somme des récaps de chaque site)
    recap_global = {"ESPECES": {"total": 0.0, "nb": 0}, "VIREMENT": {"total": 0.0, "nb": 0}}
    for g in groupes_ordonnes:
        for m in ("ESPECES", "VIREMENT"):
            recap_global[m]["total"] += g["recap_mode"][m]["total"]
            recap_global[m]["nb"]    += g["recap_mode"][m]["nb"]
    return render_template("tenant/journaliers_paie_sites_print.html",
        tenant=t, groupes=groupes_ordonnes, total_general=total_general,
        recap_global=recap_global,
        nb_total=nb_total, statut=statut_f, date_debut=date_debut, date_fin=date_fin,
        imput_av=imput_av, now=datetime.now())


@bp.route("/journaliers/paie/imprimer")
@login_required
def journaliers_paie_imprimer():
    """Page imprimable des feuilles de paie journalier (avec colonne signature).

    Filtres optionnels : site_id, statut, date_debut, date_fin.
    """
    if current_user.is_super_admin: return redirect(url_for("admin.admin_dashboard"))
    t = get_tenant()
    if not t: return redirect(url_for("auth.login"))
    site_id    = request.args.get("site_id", type=int)
    statut_f   = request.args.get("statut", "")
    date_debut = _parse_date(request.args.get("date_debut", ""))
    date_fin   = _parse_date(request.args.get("date_fin", ""))
    site       = Site.query.filter_by(id=site_id, tenant_id=t.id).first() if site_id else None

    q = FeuillePaieJournalier.query.filter_by(tenant_id=t.id)
    if site_id:
        ids_j = [a.journalier_id for a in AffectationSite.query.filter_by(
            tenant_id=t.id, site_id=site_id).filter(
            AffectationSite.journalier_id.isnot(None)).all()]
        q = q.filter(FeuillePaieJournalier.journalier_id.in_(ids_j))
    if statut_f:
        q = q.filter_by(statut=statut_f)
    if date_debut:
        q = q.filter(FeuillePaieJournalier.date_debut >= date_debut)
    if date_fin:
        q = q.filter(FeuillePaieJournalier.date_fin <= date_fin)
    feuilles = q.options(joinedload(FeuillePaieJournalier.journalier)).order_by(
        FeuillePaieJournalier.date_fin.desc()).all()
    imput_av = _imputer_avances_journalier(
        feuilles, _avances_par_journalier(t.id, {f.journalier_id for f in feuilles}))
    total = sum(imput_av.get(f.id, {}).get("net", float(f.montant_a_payer or 0)) for f in feuilles)
    # ── Sous-totaux par mode de paiement (net réellement décaissé) ──────
    recap_mode = {"ESPECES": {"total": 0.0, "nb": 0}, "VIREMENT": {"total": 0.0, "nb": 0}}
    for f in feuilles:
        mode = (f.mode_paiement if f.statut == "PAYÉ" and f.mode_paiement
                else (f.journalier.mode_paiement if f.journalier else "ESPECES")) or "ESPECES"
        if mode not in recap_mode:
            mode = "ESPECES"
        net = imput_av.get(f.id, {}).get("net", float(f.montant_a_payer or 0))
        recap_mode[mode]["total"] += net
        recap_mode[mode]["nb"]    += 1
    return render_template("tenant/journaliers_paie_print.html",
        tenant=t, feuilles=feuilles, site=site, statut=statut_f,
        date_debut=date_debut, date_fin=date_fin, total=total,
        recap_mode=recap_mode,
        imput_av=imput_av, now=datetime.now())


@bp.route("/journaliers/paie/export")
@login_required
def journaliers_paie_export():
    """Export Excel des feuilles de paie journalier — filtré par site et/ou période."""
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    import io, calendar

    t = get_tenant()
    if not t: return redirect(url_for("auth.login"))

    # ── Paramètres de filtre ──────────────────────────────────────────────────
    site_id    = request.args.get("site_id",    type=int)
    statut_f   = request.args.get("statut",     "")
    date_debut = _parse_date(request.args.get("date_debut", ""))
    date_fin   = _parse_date(request.args.get("date_fin",   ""))
    site       = Site.query.filter_by(id=site_id, tenant_id=t.id).first() if site_id else None

    # ── Requête ───────────────────────────────────────────────────────────────
    q = FeuillePaieJournalier.query.filter_by(tenant_id=t.id)
    if site_id:
        ids_j = [a.journalier_id for a in AffectationSite.query.filter_by(
            tenant_id=t.id, site_id=site_id
        ).filter(AffectationSite.journalier_id.isnot(None)).all()]
        q = q.filter(FeuillePaieJournalier.journalier_id.in_(ids_j))
    if date_debut:
        q = q.filter(FeuillePaieJournalier.date_debut >= date_debut)
    if date_fin:
        q = q.filter(FeuillePaieJournalier.date_fin   <= date_fin)
    if statut_f:
        q = q.filter_by(statut=statut_f)
    feuilles = q.order_by(
        FeuillePaieJournalier.date_fin.desc(),
        FeuillePaieJournalier.journalier_id
    ).all()

    # ── Affectation site de chaque journalier ─────────────────────────────────
    aff_map = {}
    for a in AffectationSite.query.filter_by(tenant_id=t.id).filter(
            AffectationSite.journalier_id.isnot(None)).all():
        if a.journalier_id not in aff_map:
            aff_map[a.journalier_id] = a.site.nom if a.site else "—"

    # ── Styles communs ────────────────────────────────────────────────────────
    HDR_FONT   = Font(bold=True, color="FFFFFF", size=9)
    HDR_FILL   = PatternFill("solid", fgColor="1a2332")
    HDR_ALIGN  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    BODY_FONT  = Font(size=9)
    EVEN_FILL  = PatternFill("solid", fgColor="F7F8FA")
    TOTAL_FONT = Font(bold=True, size=10, color="FFFFFF")
    TOTAL_FILL = PatternFill("solid", fgColor="1a2332")
    MONEY_FMT  = '#,##0'
    thin       = Side(style="thin", color="D1D5DB")
    BORDER     = Border(left=thin, right=thin, top=thin, bottom=thin)
    CENTER     = Alignment(horizontal="center")
    RIGHT      = Alignment(horizontal="right")

    wb = Workbook()

    # ══════════════════════════════════════════════════════════════════════════
    # ONGLET 1 — Détail complet
    # ══════════════════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Détail"
    ws.freeze_panes = "A4"

    # Titre
    titre = (f"PAIE JOURNALIERS — {t.denomination}"
             + (f" — {site.nom}" if site else "")
             + (f" — {date_debut.strftime('%d/%m/%Y')} au {date_fin.strftime('%d/%m/%Y')}" if date_debut and date_fin else "")
             + f" — Édité le {datetime.now().strftime('%d/%m/%Y à %H:%M')}")
    ws.merge_cells("A1:K1")
    ws["A1"] = titre
    ws["A1"].font = Font(bold=True, size=12, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="1a2332")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    ws.append([])  # ligne vide

    # En-têtes
    hdrs = ["Journalier","Profession","Site","Période du","au",
            "Nb jours","Total heures","Taux/h (FCFA)","Montant brut (FCFA)","Statut","Date paiement"]
    ws.append(hdrs)
    for c_idx, h in enumerate(hdrs, 1):
        cell = ws.cell(row=3, column=c_idx, value=h)
        cell.font  = HDR_FONT
        cell.fill  = HDR_FILL
        cell.alignment = HDR_ALIGN
        cell.border = BORDER
    ws.row_dimensions[3].height = 18

    # Données
    total_montant  = 0
    total_jours    = 0
    total_heures   = 0
    for row_idx, f in enumerate(feuilles, 4):
        site_nom  = aff_map.get(f.journalier_id, "—")
        montant   = float(f.montant_brut  or 0)
        heures    = float(f.total_heures  or 0)
        jours     = int(f.nb_jours or 0)
        total_montant += montant
        total_jours   += jours
        total_heures  += heures
        row_data = [
            csv_safe(f.journalier.nom_complet),
            csv_safe(f.journalier.profession or "—"),
            csv_safe(site_nom),
            f.date_debut.strftime("%d/%m/%Y") if f.date_debut else "",
            f.date_fin.strftime("%d/%m/%Y")   if f.date_fin   else "",
            jours,
            round(heures, 2),
            float(f.taux_horaire or 0),
            montant,
            csv_safe(f.statut),
            f.date_paiement.strftime("%d/%m/%Y") if f.date_paiement else "",
        ]
        ws.append(row_data)
        is_even = (row_idx % 2 == 0)
        for c_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=c_idx)
            cell.font   = BODY_FONT
            cell.border = BORDER
            if is_even: cell.fill = EVEN_FILL
            # Formats numériques
            if c_idx in (6, 7):   cell.alignment = CENTER
            if c_idx in (8, 9):
                cell.number_format = MONEY_FMT
                cell.alignment     = RIGHT
            # Statut coloré
            if c_idx == 10:
                cell.alignment = CENTER
                if val == "PAYÉ":
                    cell.font = Font(bold=True, color="065F46", size=9)
                else:
                    cell.font = Font(bold=True, color="92400E", size=9)

    # Ligne totaux
    ws.append([])
    tr = ws.max_row + 1
    totals = ["", "", "", "", "TOTAL", total_jours, round(total_heures,2),
              "", total_montant, "", ""]
    ws.append(totals)
    for c_idx, val in enumerate(totals, 1):
        cell = ws.cell(row=tr, column=c_idx)
        cell.font   = TOTAL_FONT
        cell.fill   = TOTAL_FILL
        cell.border = BORDER
        if c_idx in (8, 9):
            cell.number_format = MONEY_FMT
            cell.alignment     = RIGHT
        if c_idx == 5:
            cell.alignment = RIGHT

    # Largeurs colonnes
    col_widths = [28, 18, 20, 13, 13, 10, 13, 16, 20, 14, 15]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ══════════════════════════════════════════════════════════════════════════
    # ONGLET 2 — Récap par site
    # ══════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Récap par site")
    ws2.freeze_panes = "A3"

    ws2.merge_cells("A1:F1")
    ws2["A1"] = f"RÉCAPITULATIF PAR SITE — {t.denomination}"
    ws2["A1"].font = Font(bold=True, size=12, color="FFFFFF")
    ws2["A1"].fill = PatternFill("solid", fgColor="374151")
    ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 20

    hdrs2 = ["Site","Nb journaliers","Nb jours total","Total heures","Montant brut (FCFA)","Statut"]
    ws2.append(hdrs2)
    for c_idx, h in enumerate(hdrs2, 1):
        cell = ws2.cell(row=2, column=c_idx, value=h)
        cell.font  = HDR_FONT
        cell.fill  = PatternFill("solid", fgColor="374151")
        cell.alignment = HDR_ALIGN
        cell.border = BORDER

    # Grouper par site
    from collections import defaultdict
    by_site = defaultdict(lambda: {"journaliers": set(), "jours": 0, "heures": 0.0,
                                    "montant": 0.0, "nb_payes": 0, "nb_total": 0})
    for f in feuilles:
        s_nom = aff_map.get(f.journalier_id, "Sans site")
        by_site[s_nom]["journaliers"].add(f.journalier_id)
        by_site[s_nom]["jours"]    += int(f.nb_jours or 0)
        by_site[s_nom]["heures"]   += float(f.total_heures or 0)
        by_site[s_nom]["montant"]  += float(f.montant_brut or 0)
        by_site[s_nom]["nb_total"] += 1
        if f.statut == "PAYÉ": by_site[s_nom]["nb_payes"] += 1

    grand_total = 0
    for row_idx, (s_nom, data) in enumerate(sorted(by_site.items()), 3):
        pct_paye = int(data["nb_payes"] / data["nb_total"] * 100) if data["nb_total"] else 0
        statut_txt = f"{data['nb_payes']}/{data['nb_total']} payé(s) ({pct_paye}%)"
        row_data = [
            csv_safe(s_nom),
            len(data["journaliers"]),
            data["jours"],
            round(data["heures"], 2),
            round(data["montant"], 2),
            statut_txt,
        ]
        ws2.append(row_data)
        grand_total += data["montant"]
        is_even = (row_idx % 2 == 0)
        for c_idx, val in enumerate(row_data, 1):
            cell = ws2.cell(row=row_idx, column=c_idx)
            cell.font   = BODY_FONT
            cell.border = BORDER
            if is_even: cell.fill = EVEN_FILL
            if c_idx == 5:
                cell.number_format = MONEY_FMT
                cell.alignment     = RIGHT
            if c_idx in (2, 3, 4):
                cell.alignment = CENTER

    # Total récap
    ws2.append([])
    tr2 = ws2.max_row + 1
    ws2.append(["TOTAL GÉNÉRAL", "", "", "", grand_total, ""])
    for c_idx in range(1, 7):
        cell = ws2.cell(row=tr2, column=c_idx)
        cell.font   = TOTAL_FONT
        cell.fill   = PatternFill("solid", fgColor="374151")
        cell.border = BORDER
        if c_idx == 5:
            cell.number_format = MONEY_FMT
            cell.alignment     = RIGHT

    for i, w in enumerate([28, 16, 14, 14, 22, 22], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # ══════════════════════════════════════════════════════════════════════════
    # ONGLET 3 — Récap par journalier
    # ══════════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Récap par journalier")
    ws3.freeze_panes = "A3"

    ws3.merge_cells("A1:G1")
    ws3["A1"] = f"RÉCAPITULATIF PAR JOURNALIER — {t.denomination}"
    ws3["A1"].font = Font(bold=True, size=12, color="FFFFFF")
    ws3["A1"].fill = PatternFill("solid", fgColor="065F46")
    ws3["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 20

    hdrs3 = ["Journalier","Profession","Site","Nb périodes","Nb jours","Total heures","Total perçu (FCFA)"]
    ws3.append(hdrs3)
    for c_idx, h in enumerate(hdrs3, 1):
        cell = ws3.cell(row=2, column=c_idx, value=h)
        cell.font  = HDR_FONT
        cell.fill  = PatternFill("solid", fgColor="065F46")
        cell.alignment = HDR_ALIGN
        cell.border = BORDER

    by_jour = defaultdict(lambda: {"nom":"","profession":"","site":"","periodes":0,"jours":0,"heures":0.0,"montant":0.0})
    for f in feuilles:
        jid = f.journalier_id
        by_jour[jid]["nom"]       = f.journalier.nom_complet
        by_jour[jid]["profession"]= f.journalier.profession or "—"
        by_jour[jid]["site"]      = aff_map.get(jid, "—")
        by_jour[jid]["periodes"]  += 1
        by_jour[jid]["jours"]     += int(f.nb_jours or 0)
        by_jour[jid]["heures"]    += float(f.total_heures or 0)
        by_jour[jid]["montant"]   += float(f.montant_brut or 0)

    grand_total3 = 0
    for row_idx, (jid, d) in enumerate(sorted(by_jour.items(), key=lambda x: x[1]["nom"]), 3):
        row_data = [csv_safe(d["nom"]), csv_safe(d["profession"]), csv_safe(d["site"]),
                    d["periodes"], d["jours"], round(d["heures"],2), round(d["montant"],2)]
        ws3.append(row_data)
        grand_total3 += d["montant"]
        is_even = (row_idx % 2 == 0)
        for c_idx, val in enumerate(row_data, 1):
            cell = ws3.cell(row=row_idx, column=c_idx)
            cell.font   = BODY_FONT
            cell.border = BORDER
            if is_even: cell.fill = EVEN_FILL
            if c_idx == 7:
                cell.number_format = MONEY_FMT; cell.alignment = RIGHT
            if c_idx in (4, 5, 6): cell.alignment = CENTER

    ws3.append([])
    tr3 = ws3.max_row + 1
    ws3.append(["TOTAL GÉNÉRAL","","","","","", grand_total3])
    for c_idx in range(1, 8):
        cell = ws3.cell(row=tr3, column=c_idx)
        cell.font = TOTAL_FONT
        cell.fill = PatternFill("solid", fgColor="065F46")
        cell.border = BORDER
        if c_idx == 7:
            cell.number_format = MONEY_FMT; cell.alignment = RIGHT

    for i, w in enumerate([28, 18, 20, 12, 11, 13, 22], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    # ── Export ────────────────────────────────────────────────────────────────
    out = io.BytesIO()
    wb.save(out); out.seek(0)

    parts = ["Paie_Journaliers"]
    if site:       parts.append(site.nom.replace(" ", "_"))
    if date_debut: parts.append(date_debut.strftime("%Y%m%d"))
    if date_fin:   parts.append("au" + date_fin.strftime("%Y%m%d"))
    parts.append(datetime.now().strftime("%Y%m%d"))
    fname = "_".join(parts) + ".xlsx"

    return send_file(out,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True, download_name=fname)

@bp.route("/journaliers/paie/payer-selection", methods=["POST"])
@login_required
def journaliers_payer_selection():
    t = get_tenant()
    if not t: return redirect(url_for("auth.login"))
    ids = [int(i) for i in request.form.get("feuille_ids","").split(",") if i.strip().isdigit()]
    nb = 0
    for fid in ids:
        f = FeuillePaieJournalier.query.filter_by(id=fid, tenant_id=t.id, statut="EN_ATTENTE").first()
        if f: f.statut="PAYÉ"; f.date_paiement=datetime.now().date(); nb+=1
    db.session.commit()
    flash(f"{nb} journalier(s) payé(s).", "success")
    return redirect(url_for("tenant.journaliers_paie"))

# ── Acomptes ──────────────────────────────────────────────────────────────────
@bp.route("/acomptes")
@login_required
def acomptes():
    if current_user.is_super_admin: return redirect(url_for("admin.admin_dashboard"))
    t = get_tenant()
    if not t: return redirect(url_for("auth.login"))
    now = datetime.now()
    mois = request.args.get("mois", now.month, type=int)
    annee = request.args.get("annee", now.year, type=int)
    salarie_id = request.args.get("salarie_id", type=int)
    try:
        query = Acompte.query.filter_by(tenant_id=t.id, annee=annee, mois=mois)
        if salarie_id: query = query.filter_by(salarie_id=salarie_id)
        liste = query.order_by(Acompte.date_acompte.desc()).all()
    except Exception:
        db.create_all(); db.session.rollback(); liste = []
    total_mois       = sum(float(a.montant) for a in liste if a.statut != "ANNULE")
    total_en_attente = sum(float(a.montant) for a in liste if a.statut == "EN_ATTENTE")
    total_deduit     = sum(float(a.montant) for a in liste if a.statut == "DEDUIT")
    salaries_list = Salarie.query.filter_by(tenant_id=t.id, statut="ACTIF").order_by(Salarie.nom).all()
    return render_template("tenant/acomptes.html", tenant=t, liste=liste, salaries=salaries_list,
        mois=mois, annee=annee, now=now, total_mois=total_mois,
        total_en_attente=total_en_attente, total_deduit=total_deduit, MOIS_NOMS=PeriodePaie.MOIS_NOMS)

@bp.route("/acomptes/imprimer", methods=["POST"])
@login_required
def acomptes_imprimer():
    """Page imprimable de la liste des acomptes sélectionnés (avec colonne signature)."""
    if current_user.is_super_admin: return redirect(url_for("admin.admin_dashboard"))
    t = get_tenant()
    if not t: return redirect(url_for("auth.login"))
    ids = [int(i) for i in request.form.getlist("acompte_ids") if str(i).isdigit()]
    mois  = request.form.get("mois", type=int)
    annee = request.form.get("annee", type=int)
    if not ids:
        flash("Sélectionnez au moins un acompte à imprimer.", "error")
        return redirect(url_for("tenant.acomptes", mois=mois, annee=annee))
    liste = Acompte.query.filter(
        Acompte.tenant_id == t.id,
        Acompte.id.in_(ids)
    ).options(joinedload(Acompte.salarie)).order_by(Acompte.date_acompte).all()
    total = sum(float(a.montant or 0) for a in liste if a.statut != "ANNULE")
    return render_template("tenant/acomptes_print.html",
        tenant=t, liste=liste, total=total, mois=mois, annee=annee,
        MOIS_NOMS=PeriodePaie.MOIS_NOMS, now=datetime.now())


@bp.route("/acomptes/nouveau", methods=["GET","POST"])
@login_required
def acompte_nouveau():
    if current_user.is_super_admin: return redirect(url_for("admin.admin_dashboard"))
    t = get_tenant()
    if not t: return redirect(url_for("auth.login"))
    if not current_user.can_edit: abort(403)
    salaries_list = Salarie.query.filter_by(tenant_id=t.id, statut="ACTIF").order_by(Salarie.nom).all()
    if request.method == "POST":
        salarie_id = request.form.get("salarie_id", type=int)
        montant    = float(request.form.get("montant", 0) or 0)
        date_ac    = _parse_date(request.form.get("date_acompte"))
        mois       = request.form.get("mois", type=int)
        annee      = request.form.get("annee", type=int)
        motif      = request.form.get("motif", "").strip()
        if not salarie_id or montant <= 0 or not date_ac:
            flash("Veuillez remplir tous les champs.", "error")
        else:
            # Valider que le salarié appartient bien au tenant (anti-IDOR).
            sal = Salarie.query.filter_by(id=salarie_id, tenant_id=t.id).first()
            if not sal:
                flash("Salarié introuvable.", "error")
                return render_template("tenant/acompte_form.html", tenant=t, salaries=salaries_list, now=datetime.now())
            contrat = Contrat.query.filter_by(salarie_id=salarie_id, tenant_id=t.id, actif=True).first()
            if contrat and montant > float(contrat.salaire_base) * 0.5:
                flash(f"Acompte maximum 50% du salaire de base ({float(contrat.salaire_base)*0.5:,.0f} FCFA).".replace(",", " "), "error")
                return render_template("tenant/acompte_form.html", tenant=t, salaries=salaries_list, now=datetime.now())
            ac = Acompte(tenant_id=t.id, salarie_id=salarie_id, montant=montant,
                date_acompte=date_ac, mois=mois, annee=annee, motif=motif, statut="EN_ATTENTE")
            db.session.add(ac)
            db.session.commit()
            log_action("CREATE", "acompte", ac.id,
                       f"Acompte {montant:,.0f} F — {sal.nom_complet}".replace(",", " "))
            db.session.commit()
            flash(f"Acompte de {montant:,.0f} FCFA enregistré.".replace(",", " "), "success")
            return redirect(url_for("tenant.acomptes", mois=mois, annee=annee))
    return render_template("tenant/acompte_form.html", tenant=t, salaries=salaries_list, now=datetime.now())

@bp.route("/acomptes/<int:id>/valider", methods=["POST"])
@login_required
def acompte_valider(id):
    t = get_tenant()
    if not t: return redirect(url_for("auth.login"))
    a = Acompte.query.filter_by(id=id, tenant_id=t.id).first_or_404()
    a.statut = "DEDUIT"; db.session.commit()
    log_action("VALIDATE", "acompte", a.id, f"Acompte déduit ({float(a.montant or 0):,.0f} F)".replace(",", " "))
    db.session.commit()
    flash("Acompte marqué comme déduit.", "success")
    return redirect(url_for("tenant.acomptes", mois=a.mois, annee=a.annee))

@bp.route("/acomptes/<int:id>/annuler", methods=["POST"])
@login_required
def acompte_annuler(id):
    t = get_tenant()
    if not t: return redirect(url_for("auth.login"))
    a = Acompte.query.filter_by(id=id, tenant_id=t.id).first_or_404()
    a.statut = "ANNULE"; db.session.commit()
    log_action("CANCEL", "acompte", a.id, "Acompte annulé")
    db.session.commit()
    flash("Acompte annulé.", "success")
    return redirect(url_for("tenant.acomptes", mois=a.mois, annee=a.annee))

@bp.route("/acomptes/<int:id>/supprimer", methods=["POST"])
@login_required
def acompte_supprimer(id):
    t = get_tenant()
    if not t: return redirect(url_for("auth.login"))
    a = Acompte.query.filter_by(id=id, tenant_id=t.id).first_or_404()
    mois, annee = a.mois, a.annee
    db.session.delete(a); db.session.commit()
    log_action("DELETE", "acompte", id, "Suppression acompte")
    db.session.commit()
    flash("Acompte supprimé.", "success")
    return redirect(url_for("tenant.acomptes", mois=mois, annee=annee))

# ── Congés avancés ────────────────────────────────────────────────────────────

@bp.route("/conges/bilan")
@tenant_required
def conges_bilan():
    """Bilan congés de tous les salariés actifs avec jours acquis, pris, restants."""
    t = get_tenant()
    annee = request.args.get("annee", date.today().year, type=int)
    salaries_actifs = Salarie.query.filter_by(
        tenant_id=t.id, statut="ACTIF"
    ).options(
        joinedload(Salarie.conges),
        joinedload(Salarie.contrats),
    ).order_by(Salarie.nom).all()

    from conges_avance import bilan_conges_tenant
    bilan = bilan_conges_tenant(salaries_actifs, annee)

    # Stats globales
    total_acquis   = sum(b["jours_acquis"]   for b in bilan)
    total_pris     = sum(b["jours_pris"]     for b in bilan)
    total_restants = sum(b["jours_restants"] for b in bilan)
    nb_alertes     = sum(1 for b in bilan if b["alerte"])

    return render_template("tenant/conges_bilan.html",
        tenant=t, bilan=bilan, annee=annee,
        total_acquis=total_acquis, total_pris=total_pris,
        total_restants=total_restants, nb_alertes=nb_alertes,
        annees=list(range(date.today().year - 2, date.today().year + 2)),
    )


@bp.route("/conges/planning")
@tenant_required
def conges_planning():
    """Planning visuel des congés — vue calendrier mensuel."""
    t    = get_tenant()
    mois = request.args.get("mois", date.today().month, type=int)
    annee= request.args.get("annee", date.today().year,  type=int)

    tous_conges = Conge.query.filter_by(tenant_id=t.id)\
        .options(joinedload(Conge.salarie))\
        .filter(Conge.statut.in_(["APPROUVÉ","APPROUVE","DEMANDÉ","DEMANDE","PRIS"]))\
        .all()

    from conges_avance import planning_absences
    planning = planning_absences(tous_conges, annee, mois)

    # Infos du mois pour le calendrier
    import calendar
    cal = calendar.monthcalendar(annee, mois)
    nb_jours = calendar.monthrange(annee, mois)[1]
    MOIS_FR = ["","Janvier","Février","Mars","Avril","Mai","Juin",
               "Juillet","Août","Septembre","Octobre","Novembre","Décembre"]

    return render_template("tenant/conges_planning.html",
        tenant=t, planning=planning, mois=mois, annee=annee,
        mois_nom=MOIS_FR[mois], cal=cal, nb_jours=nb_jours,
        MOIS_FR=MOIS_FR,
    )


@bp.route("/salaries/<int:sal_id>/solde-tout-compte")
@tenant_required
def solde_tout_compte(sal_id):
    """Calcul du solde de tout compte (indemnité congés + licenciement)."""
    t = get_tenant()
    s = Salarie.query.filter_by(id=sal_id, tenant_id=t.id)\
        .options(joinedload(Salarie.conges),
                 joinedload(Salarie.contrats)).first_or_404()

    date_cessation_str = request.args.get("date_cessation", "")
    try:
        date_cessation = datetime.strptime(date_cessation_str, "%Y-%m-%d").date() \
            if date_cessation_str else date.today()
    except ValueError:
        date_cessation = date.today()

    # 12 derniers bulletins
    bulletins_12 = BulletinPaie.query.filter_by(
        tenant_id=t.id, salarie_id=sal_id
    ).filter(
        BulletinPaie.statut.in_(["VALIDÉ","VALIDE","PAYÉ"])
    ).order_by(BulletinPaie.date_creation.desc()).limit(12).all()

    from conges_avance import calculer_solde_tout_compte
    cause = request.args.get("cause", "LICENCIEMENT")
    solde = calculer_solde_tout_compte(
        s, bulletins_12, date_cessation, convention=t.convention,
        cause=cause, jours_conge_par_mois=t.jours_conge_par_mois
    )

    return render_template("tenant/solde_tout_compte.html",
        tenant=t, salarie=s, solde=solde, date_cessation=date_cessation,
    )


def _doc_response(pdf_bytes, nom_fichier):
    """Helper : renvoie un PDF en téléchargement."""
    from flask import Response
    return Response(pdf_bytes, mimetype="application/pdf",
                    headers={"Content-Disposition": f'attachment; filename="{nom_fichier}"',
                             "Content-Length": str(len(pdf_bytes))})


@bp.route("/salaries/<int:sal_id>/document/<type_doc>")
@login_required
def salarie_document(sal_id, type_doc):
    """
    Génère un document RH PDF pour un salarié.
    type_doc : attestation-travail | certificat-travail | attestation-salaire | solde-tout-compte
    """
    t = get_tenant()
    if not t:
        return redirect(url_for("auth.login"))
    s = (Salarie.query.filter_by(id=sal_id, tenant_id=t.id)
         .options(joinedload(Salarie.conges), joinedload(Salarie.contrats))
         .first_or_404())

    from documents_rh import (attestation_travail, certificat_travail,
                              attestation_salaire, solde_tout_compte_pdf)
    base_nom = f"{s.nom}_{s.prenom}".replace(" ", "_")

    try:
        if type_doc == "attestation-travail":
            pdf = attestation_travail(s, t)
            nom = f"attestation_travail_{base_nom}.pdf"

        elif type_doc == "certificat-travail":
            pdf = certificat_travail(s, t)
            nom = f"certificat_travail_{base_nom}.pdf"

        elif type_doc == "attestation-salaire":
            # Récupérer le dernier bulletin pour les montants
            dernier = (BulletinPaie.query.filter_by(tenant_id=t.id, salarie_id=sal_id)
                       .order_by(BulletinPaie.date_creation.desc()).first())
            brut = float(dernier.salaire_brut) if dernier else None
            net  = float(dernier.net_a_payer) if dernier else None
            if brut is None:
                contrat = next((c for c in s.contrats if c.actif), None)
                brut = float(contrat.salaire_base) if contrat else None
            pdf = attestation_salaire(s, t, brut, net)
            nom = f"attestation_salaire_{base_nom}.pdf"

        elif type_doc == "solde-tout-compte":
            date_cess_str = request.args.get("date_cessation", "")
            date_cess = parse_date(date_cess_str) or s.date_cessation or date.today()
            bulletins_12 = (BulletinPaie.query.filter_by(tenant_id=t.id, salarie_id=sal_id)
                            .filter(BulletinPaie.statut.in_(["VALIDÉ", "VALIDE", "PAYÉ"]))
                            .order_by(BulletinPaie.date_creation.desc()).limit(12).all())
            from conges_avance import calculer_solde_tout_compte
            cause = request.args.get("cause", "LICENCIEMENT")
            solde = calculer_solde_tout_compte(
                s, bulletins_12, date_cess, convention=t.convention,
                cause=cause, jours_conge_par_mois=t.jours_conge_par_mois
            )
            pdf = solde_tout_compte_pdf(s, t, solde, date_cess)
            nom = f"solde_tout_compte_{base_nom}.pdf"

        else:
            flash("Type de document inconnu.", "error")
            return redirect(url_for("tenant.salarie_detail", id=sal_id))

        log_action("EXPORT", "salarie", sal_id,
                   f"Document {type_doc} généré pour {s.nom_complet}",
                   user_id=current_user.id, tenant_id=t.id)
        db.session.commit()
        return _doc_response(pdf, nom)

    except Exception as e:
        logger.error(f"Erreur génération document {type_doc} : {e}")
        flash(f"Erreur lors de la génération du document : {e}", "error")
        return redirect(url_for("tenant.salarie_detail", id=sal_id))


@bp.route("/api/conges/jours-acquis/<int:sal_id>")
@login_required
def api_jours_acquis(sal_id):
    """Retourne les jours acquis d'un salarié (JSON)."""
    t = get_tenant()
    if not t: return jsonify({"error": "non authentifié"}), 401
    s = Salarie.query.filter_by(id=sal_id, tenant_id=t.id).first_or_404()

    from conges_avance import calculer_jours_acquis
    result = calculer_jours_acquis(s.date_embauche, convention=getattr(s.tenant, "convention", None))
    # Sérialiser les dates
    for k in ["periode_debut","periode_fin"]:
        if result.get(k):
            result[k] = str(result[k])
    return jsonify(result)


# ── Congés ────────────────────────────────────────────────────────────────────
@bp.route("/conges")
@login_required
def conges():
    if current_user.is_super_admin: return redirect(url_for("admin.admin_dashboard"))
    t = get_tenant()
    if not t: return redirect(url_for("auth.login"))
    now   = datetime.now()
    annee = request.args.get("annee", now.year, type=int)
    q     = request.args.get("q", "")

    # ── Calcul du solde — Code du Travail gabonais ───────────────────────────
    # Art. 213 : 2 j/mois pour ≥ 18 ans | 2,5 j/mois pour < 18 ans
    # Allocation = max(Σ bruts 12 mois, dernier brut×12) / 288 × jours acquis
    # Exclusion : prime de transport (Art. 213 al. 3)

    def age_au_31_dec(salarie, annee_ref):
        """Âge du salarié au 31 décembre de l'année de référence."""
        if not salarie.date_naissance:
            return 99  # inconnu → adulte par défaut
        return annee_ref - salarie.date_naissance.year - (
            1 if salarie.date_naissance.replace(year=annee_ref) > datetime(annee_ref,12,31).date() else 0
        )

    def taux_conge(salarie, annee_ref):
        """2.5 j/mois si < 18 ans, sinon 2 j/mois."""
        return 2.5 if age_au_31_dec(salarie, annee_ref) < 18 else 2.0

    def calculer_solde_auto(salarie, annee_ref):
        """Jours acquis proratisés selon le taux applicable."""
        tx = taux_conge(salarie, annee_ref)
        if not salarie.date_embauche:
            return round(12 * tx, 1)
        emb         = salarie.date_embauche
        debut_annee = datetime(annee_ref, 1, 1).date()
        fin_annee   = datetime(annee_ref, 12, 31).date()
        debut_acq   = max(emb, debut_annee)
        fin_acq     = min(datetime.now().date(), fin_annee)
        if fin_acq < debut_acq:
            return 0.0
        mois_trav = min(round((fin_acq - debut_acq).days / 30.44, 1), 12)
        return round(mois_trav * tx, 1)

    salaries_list = Salarie.query.filter_by(tenant_id=t.id, statut="ACTIF")        .options(joinedload(Salarie.categorie), joinedload(Salarie.contrats)).order_by(Salarie.nom).all()

    # ── PRÉCHARGEMENT pour éviter les requêtes N+1 ───────────────────────────
    from datetime import timedelta
    sal_ids = [s.id for s in salaries_list]

    # 1. Tous les soldes de congés de l'année en une requête
    soldes_db_map = {}
    if sal_ids:
        for c in Conge.query.filter(
            Conge.tenant_id == t.id, Conge.salarie_id.in_(sal_ids),
            Conge.annee == annee, Conge.date_depart == None
        ).all():
            soldes_db_map[c.salarie_id] = c

    # 2. Tous les congés approuvés de l'année (pour cumul jours pris) en une requête
    pris_map = {}
    if sal_ids:
        for c in Conge.query.filter(
            Conge.tenant_id == t.id, Conge.salarie_id.in_(sal_ids),
            Conge.annee == annee, Conge.statut == "APPROUVÉ"
        ).all():
            pris_map.setdefault(c.salarie_id, 0.0)
            pris_map[c.salarie_id] += float(c.jours_pris or 0)

    # 3. Tous les bulletins des 12 derniers mois en une requête, groupés par salarié
    buls_map = {}
    if sal_ids:
        debut_periode = (datetime(annee, 12, 31) - timedelta(days=365)).date()
        buls_query = (BulletinPaie.query
            .filter(BulletinPaie.tenant_id == t.id,
                    BulletinPaie.salarie_id.in_(sal_ids))
            .join(PeriodePaie)
            .filter(PeriodePaie.annee >= debut_periode.year)
            .order_by(PeriodePaie.annee.desc(), PeriodePaie.mois.desc())
            .all())
        for b in buls_query:
            buls_map.setdefault(b.salarie_id, []).append(b)

    def calculer_allocation_conge(salarie, jours_acquis, annee_ref):
        """
        Allocation congés = max(Σbruts12mois, dernierBrut×12) / 288 × jours_acquis
        Prime de transport exclue de la base (Art. 213 al. 3).
        Utilise les données préchargées (buls_map) — aucune requête SQL ici.
        """
        if jours_acquis <= 0:
            return 0.0, 0.0
        buls = buls_map.get(salarie.id, [])[:12]
        if not buls:
            contrat = next((c for c in salarie.contrats if c.actif), None)
            if not contrat: return 0.0, 0.0
            last_brut = float(contrat.salaire_base or 0)
            somme_12  = last_brut * 12
        else:
            somme_12  = sum(
                float(b.salaire_brut or 0) - float(b.prime_transport or 0)
                for b in buls
            )
            last_brut = float(buls[0].salaire_brut or 0) - float(buls[0].prime_transport or 0)
        base_methode1 = somme_12  / 288
        base_methode2 = (last_brut * 12) / 288
        base          = max(base_methode1, base_methode2)
        allocation    = round(base * jours_acquis, 0)
        return round(base, 2), allocation

    soldes = []
    for s in salaries_list:
        if q and q.lower() not in f"{s.nom} {s.prenom} {s.matricule}".lower():
            continue
        solde_db = soldes_db_map.get(s.id)

        jours_auto = calculer_solde_auto(s, annee)

        if solde_db:
            acquis    = float(solde_db.jours_acquis or jours_auto)
            pris      = float(solde_db.jours_pris   or 0)
        else:
            acquis    = jours_auto
            pris      = pris_map.get(s.id, 0.0)

        taux_j    = taux_conge(s, annee)
        base_all, allocation = calculer_allocation_conge(s, acquis, annee)
        soldes.append({
            "salarie":        s,
            "solde_db":       solde_db,
            "jours_auto":     jours_auto,
            "jours_acquis":   acquis,
            "jours_pris":     pris,
            "jours_restants": round(acquis - pris, 1),
            "alerte":         (acquis - pris) < 5,
            "taux_j":         taux_j,
            "base_allocation":base_all,
            "allocation":     allocation,
            "mineur":         taux_j == 2.5,
        })

    # Demandes (avec date_depart renseignée) — paginées
    page_conges = request.args.get("page", 1, type=int)
    q_demandes = (Conge.query.filter_by(tenant_id=t.id)
                  .filter(Conge.date_depart.isnot(None))
                  .options(joinedload(Conge.salarie))
                  .order_by(Conge.date_depart.desc()))
    pagination = q_demandes.paginate(page=page_conges, per_page=25, error_out=False)
    demandes = pagination.items

    _args = {k: v for k, v in request.args.items() if k != "page"}
    _base = request.path + "?" + "&".join(f"{k}={v}" for k, v in _args.items())
    _sep  = "&" if _args else "?"

    annees_dispo = sorted(set(
        [now.year, now.year-1, now.year+1]
        + [c.annee for c in Conge.query.filter_by(tenant_id=t.id).all()]
    ), reverse=True)

    return render_template("tenant/conges.html",
        tenant=t, soldes=soldes, demandes=demandes,
        pagination=pagination, pagination_base=_base + _sep,
        annee=annee, annees_dispo=annees_dispo, now=now, q=q,
        salaries=salaries_list)

@bp.route("/conges/nouveau", methods=["GET","POST"])
@login_required
def conge_nouveau():
    if current_user.is_super_admin: return redirect(url_for("admin.admin_dashboard"))
    t = get_tenant()
    if not t: return redirect(url_for("auth.login"))
    salaries_list = Salarie.query.filter_by(tenant_id=t.id, statut="ACTIF").order_by(Salarie.nom).all()
    if request.method == "POST":
        salarie_id = request.form.get("salarie_id", type=int)
        annee = request.form.get("annee", datetime.now().year, type=int)
        date_dep = _parse_date(request.form.get("date_depart"))
        date_ret = _parse_date(request.form.get("date_retour"))
        type_c = request.form.get("type_conge", "ANNUEL")
        # Congé de maternité : si une date présumée d'accouchement est fournie,
        # les 14 semaines légales (Art. 208) sont calculées automatiquement.
        if type_c == "MATERNITE":
            date_acc = _parse_date(request.form.get("date_accouchement"))
            if date_acc:
                from conges_avance import calculer_conge_maternite
                mat = calculer_conge_maternite(
                    date_acc,
                    naissances_multiples=(request.form.get("naissances_multiples") == "1"),
                    complications=(request.form.get("complications_grossesse") == "1"),
                )
                date_dep, date_ret = mat["date_debut"], mat["date_fin"]
        jours = (date_ret - date_dep).days + 1 if date_dep and date_ret else 0
        conge = Conge.query.filter_by(tenant_id=t.id, salarie_id=salarie_id, annee=annee).first()
        if not conge:
            s = Salarie.query.filter_by(id=salarie_id, tenant_id=t.id).first()
            mois = max(1,(datetime.now().date()-s.date_embauche).days//30) if s.date_embauche else 12
            conge = Conge(tenant_id=t.id, salarie_id=salarie_id, annee=annee,
                jours_acquis=round(min(mois,12)*2.0,1), jours_pris=0, type_conge=type_c, statut="DEMANDÉ")
            db.session.add(conge)
        conge.date_depart  = date_dep
        conge.date_retour  = date_ret
        conge.type_conge   = type_c
        conge.statut       = "DEMANDÉ"
        conge.jours_pris   = float(conge.jours_pris or 0)  # ne pas écraser les jours déjà pris
        db.session.commit()
        log_action("CREATE", "conge", conge.id,
                   f"Demande de congé {conge.salarie.nom_complet if conge.salarie else ''} "
                   f"({jours} j) — {conge.date_debut}→{conge.date_fin}")
        db.session.commit()
        flash(f"✅ Demande de congé enregistrée ({jours} jour(s)).", "success")
        return redirect(url_for("tenant.conges"))
    return render_template("tenant/conge_form.html", tenant=t, salaries=salaries_list, now=datetime.now())

@bp.route("/conges/<int:id>/modifier", methods=["GET","POST"])
@login_required
def conge_modifier(id):
    """Modifier une demande de congé existante."""
    t = get_tenant()
    if not t: return redirect(url_for("auth.login"))
    c = Conge.query.filter_by(id=id, tenant_id=t.id).first_or_404()
    salaries_list = Salarie.query.filter_by(tenant_id=t.id, statut="ACTIF").order_by(Salarie.nom).all()

    if request.method == "POST":
        old_jours = (c.date_retour - c.date_depart).days + 1 if c.date_depart and c.date_retour else 0
        date_dep  = _parse_date(request.form.get("date_depart"))
        date_ret  = _parse_date(request.form.get("date_retour"))
        type_c    = request.form.get("type_conge", "ANNUEL")
        new_jours = (date_ret - date_dep).days + 1 if date_dep and date_ret else 0

        # Si congé APPROUVÉ → ajuster les jours_pris
        if c.statut == "APPROUVÉ":
            c.jours_pris = max(0, float(c.jours_pris or 0) - old_jours + new_jours)

        c.date_depart = date_dep
        c.date_retour = date_ret
        c.type_conge  = type_c
        c.statut      = request.form.get("statut", c.statut)
        db.session.commit()
        log_action("UPDATE", "conge", c.id,
                   f"Modification congé — {c.salarie.nom_complet if c.salarie else ''} ({new_jours} j)")
        db.session.commit()
        flash(f"✅ Congé modifié ({new_jours} jour(s)).", "success")
        return redirect(url_for("tenant.conges"))

    return render_template("tenant/conge_form.html",
        tenant=t, salaries=salaries_list,
        conge=c, now=datetime.now(), mode="modifier")


@bp.route("/conges/<int:id>/approuver", methods=["POST"])
@login_required
def conge_approuver(id):
    t = get_tenant()
    if not t: return redirect(url_for("auth.login"))
    c = Conge.query.filter_by(id=id, tenant_id=t.id).first_or_404()
    if c.date_depart and c.date_retour:
        jours = (c.date_retour - c.date_depart).days + 1
        # Mettre à jour le solde de l'année
        solde = Conge.query.filter_by(
            tenant_id=t.id, salarie_id=c.salarie_id, annee=c.annee
        ).filter(Conge.date_depart == None).first()
        if not solde:
            s = Salarie.query.filter_by(id=c.salarie_id, tenant_id=t.id).first()
            mois = max(1,(datetime.now().date()-s.date_embauche).days//30) if s.date_embauche else 12
            solde = Conge(tenant_id=t.id, salarie_id=c.salarie_id, annee=c.annee,
                          jours_acquis=round(min(mois,12)*2.5, 1), jours_pris=0)
            db.session.add(solde)
        solde.jours_pris = float(solde.jours_pris or 0) + jours
        c.jours_pris     = float(c.jours_pris or 0) + jours
    c.statut = "APPROUVÉ"
    db.session.commit()
    log_action("VALIDATE", "conge", c.id,
               f"Congé approuvé — {c.salarie.nom_complet if c.salarie else ''}")
    db.session.commit()
    flash(f"✅ Congé de {c.salarie.nom_complet} approuvé.", "success")
    return redirect(url_for("tenant.conges"))

@bp.route("/conges/<int:id>/refuser", methods=["POST"])
@login_required
def conge_refuser(id):
    t = get_tenant()
    if not t: return redirect(url_for("auth.login"))
    c = Conge.query.filter_by(id=id, tenant_id=t.id).first_or_404()
    c.statut="REFUSÉ"; db.session.commit()
    log_action("CANCEL", "conge", c.id,
               f"Congé refusé — {c.salarie.nom_complet if c.salarie else ''}")
    db.session.commit()
    flash("Congé refusé.", "success")
    return redirect(url_for("tenant.conges"))

@bp.route("/conges/<int:id>/supprimer", methods=["POST"])
@login_required
def conge_supprimer(id):
    t = get_tenant()
    if not t: return redirect(url_for("auth.login"))
    c = Conge.query.filter_by(id=id, tenant_id=t.id).first_or_404()
    db.session.delete(c); db.session.commit()
    log_action("DELETE", "conge", id,
               f"Suppression demande de congé — {c.salarie.nom_complet if c.salarie else ''}")
    db.session.commit()
    flash("Demande supprimée.", "success")
    return redirect(url_for("tenant.conges"))


@bp.route("/salaries/imprimer")
@login_required
def salaries_imprimer():
    if current_user.is_super_admin: return redirect(url_for("admin.admin_dashboard"))
    t = get_tenant(); 
    if not t: return redirect(url_for("auth.login"))
    salaries_list = Salarie.query.filter_by(tenant_id=t.id).order_by(Salarie.nom).all()
    for s in salaries_list:
        s._contrat_actif = Contrat.query.filter_by(salarie_id=s.id, tenant_id=t.id, actif=True).first()
    return render_template("tenant/salaries_print.html", salaries=salaries_list, tenant=t, now=datetime.now())


# ══════════════════════════════════════════════════════════════════════════════
# ── IMPRESSION DES POINTAGES (salariés & journaliers) ─────────────────────────
# ══════════════════════════════════════════════════════════════════════════════

_MOIS_FR = ["", "Janvier", "Février", "Mars", "Avril", "Mai", "Juin", "Juillet",
            "Août", "Septembre", "Octobre", "Novembre", "Décembre"]
_JOURS_FR = ["Lundi", "Mardi", "Mercredi", "Jeudi", "Vendredi", "Samedi", "Dimanche"]
_TYPE_JOUR_LABEL = {
    "NORMAL": "Ordinaire", "DIMANCHE": "Dimanche", "FERIE": "Férié travaillé",
    "CHOME_PAYE": "Férié chômé payé", "CHOME_RECUPERABLE": "Férié récupérable",
}


def _pointages_mois_contexte(t, pointages, convention):
    """Construit le contexte d'impression d'un relevé de pointage mensuel.

    Retourne un dict : lignes par jour, totaux ventilés (heures normales et
    supplémentaires) et, pour la convention BTP, le détail de répartition
    semaine par semaine (utile en cas de réclamation du travailleur).
    """
    pts = sorted(pointages, key=lambda p: p.date_pointage)
    conv = (convention or "").upper()

    # Carte des heures de nuit PAR JOUR (BTP) : calculée depuis les horaires
    # réels via la même fonction que la ventilation (pointage_vers_jours), afin
    # que la colonne « DONT NUIT » de chaque ligne soit cohérente avec le total
    # +40 % affiché en bas (somme des lignes = total).
    nuit_par_date = {}
    if conv in ("BTP", "PETROLE", "INDUSTRIE", "AERIEN"):
        from calculs_paie import pointage_vers_jours
        for j in pointage_vers_jours(pts):
            d = j.get("date")
            if d is not None:
                nuit_par_date[d] = nuit_par_date.get(d, 0.0) + float(j.get("heures_nuit") or 0)

    lignes = []
    for p in pts:
        hn   = float(p.heures_normales or 0)
        hsup = float(p.heures_sup or 0)     # heures sup "simples" (journaliers, non majorées)
        h10  = float(p.heures_sup_10 or 0)
        h30  = float(p.heures_sup_30 or 0)
        h30b = float(getattr(p, "heures_sup_30b", 0) or 0)
        h40  = float(p.heures_sup_40 or 0)
        h70  = float(p.heures_sup_70 or 0)
        absent = bool(p.absent)
        present = bool(p.present) and not absent
        total_jour = hn + hsup + h10 + h30 + h30b + h40 + h70
        tj = (p.type_jour or "NORMAL").upper()
        # Nuit du jour : depuis l'horaire (BTP/Pétrole/Industrie), sinon valeur stockée (journaliers)
        if conv in ("BTP", "PETROLE", "INDUSTRIE", "AERIEN"):
            nuit_jour = nuit_par_date.get(p.date_pointage, 0.0)
        else:
            nuit_jour = h40
        lignes.append({
            "date": p.date_pointage,
            "jour_sem": _JOURS_FR[p.date_pointage.weekday()],
            "type_label": _TYPE_JOUR_LABEL.get(tj, tj.title()),
            "present": present, "absent": absent,
            "motif": p.motif_absence or "",
            "site": (p.site.nom if getattr(p, "site", None) else ""),
            "heures_travaillees": round(total_jour, 2),
            "heures_nuit": round(nuit_jour, 2),
            "observation": p.observation or "",
        })

    pts_travailles = [p for p in pts if p.present and not p.absent]
    pts_absents    = [p for p in pts if p.absent]

    if conv in ("BTP", "PETROLE", "INDUSTRIE", "AERIEN") and pts_travailles:
        from calculs_paie import ventiler_heures_mois, pointage_vers_jours
        v = ventiler_heures_mois(conv, pointage_vers_jours(pts), seuil_normales=t.seuil_hs)
        totaux = {
            "heures_normales": v["heures_normales"],
            "heures_sup_10":   v["heures_sup_10"],
            "heures_sup_30":   v["heures_sup_30"],
            "heures_sup_30b":  v.get("heures_sup_30b", 0.0),
            "heures_sup_40":   v["heures_sup_40"],
            "heures_sup_70":   v["heures_sup_70"],
        }
        detail_semaines = v.get("detail_semaines", [])
    else:
        totaux = {
            "heures_normales": sum(float(p.heures_normales or 0) for p in pts_travailles),
            "heures_sup_10":   sum(float(p.heures_sup_10 or 0) for p in pts_travailles),
            "heures_sup_30":   sum(float(p.heures_sup_30 or 0) for p in pts_travailles),
            "heures_sup_30b":  sum(float(getattr(p, "heures_sup_30b", 0) or 0) for p in pts_travailles),
            "heures_sup_40":   sum(float(p.heures_sup_40 or 0) for p in pts_travailles),
            "heures_sup_70":   sum(float(p.heures_sup_70 or 0) for p in pts_travailles),
        }
        detail_semaines = []

    totaux = {k: round(v, 2) for k, v in totaux.items()}
    # Heures sup "simples" (colonne heures_sup) : utilisées par les journaliers,
    # non majorées. Nulles pour les salariés BTP (qui utilisent les tranches).
    heures_sup_simple = round(sum(float(p.heures_sup or 0) for p in pts_travailles), 2)
    totaux["heures_sup_simple"] = heures_sup_simple
    totaux["total_sup"] = round(totaux["heures_sup_10"] + totaux["heures_sup_30"]
                                + totaux.get("heures_sup_30b", 0)
                                + totaux["heures_sup_40"] + totaux["heures_sup_70"]
                                + heures_sup_simple, 2)
    totaux["total_general"] = round(totaux["heures_normales"] + totaux["total_sup"], 2)
    # Total des heures de nuit = somme des heures de nuit de chaque ligne (cohérent
    # avec la colonne « dont nuit » du tableau).
    totaux["heures_nuit"] = round(sum(l["heures_nuit"] for l in lignes), 2)
    # Heures supplémentaires DE JOUR = total des sup. moins la part de nuit, pour
    # trois catégories additives sans double comptage :
    #   heures normales + heures sup. de jour + heures de nuit = total travaillé.
    totaux["heures_sup_jour"] = round(max(0.0, totaux["total_sup"] - totaux["heures_nuit"]), 2)
    totaux["nb_jours"]    = len(pts_travailles)
    totaux["nb_absences"] = len(pts_absents)

    # Taux de majoration RÉELS de la convention, pour étiqueter correctement les
    # tranches (BTP ≠ Pétrole ≠ Industrie). Évite d'afficher « +10 % » à tort.
    from calculs_paie import coeffs_heures_sup
    _c = coeffs_heures_sup(conv)
    coeffs_pct = {k: int(round((float(v) - 1) * 100)) for k, v in _c.items()}

    # Répartition par site (multi-chantiers) : un salarié/journalier peut
    # travailler sur plusieurs sites dans le mois. On agrège ses heures par site
    # à partir des lignes (jours travaillés uniquement).
    sites_agg = {}
    for p, l in zip(pts, lignes):
        if l["absent"]:
            continue
        nom = (p.site.nom if getattr(p, "site", None) else None) or "Non affecté"
        e = sites_agg.setdefault(nom, {"nom": nom, "heures": 0.0, "nuit": 0.0, "jours": 0})
        e["heures"] += l["heures_travaillees"]
        e["nuit"]   += l["heures_nuit"]
        e["jours"]  += 1
    recap_sites = sorted(sites_agg.values(), key=lambda x: -x["heures"])
    for e in recap_sites:
        e["heures"] = round(e["heures"], 2)
        e["nuit"]   = round(e["nuit"], 2)
    multi_sites = len(recap_sites) > 1

    return {"lignes": lignes, "totaux": totaux, "detail_semaines": detail_semaines,
            "convention": conv, "coeffs_pct": coeffs_pct,
            "recap_sites": recap_sites, "multi_sites": multi_sites}


def _resoudre_mois_annee():
    now = datetime.now()
    mois  = request.args.get("mois",  type=int) or now.month
    annee = request.args.get("annee", type=int) or now.year
    mois  = min(max(mois, 1), 12)
    return mois, annee


@bp.route("/salaries/<int:id>/pointages/imprimer")
@login_required
def salarie_pointages_imprimer(id):
    """Relevé mensuel imprimable des pointages d'un salarié (totaux + répartition)."""
    if current_user.is_super_admin:
        return redirect(url_for("admin.admin_dashboard"))
    t = get_tenant()
    if not t:
        return redirect(url_for("auth.login"))
    s = Salarie.query.filter_by(id=id, tenant_id=t.id).first_or_404()
    mois, annee = _resoudre_mois_annee()
    import calendar
    debut = date(annee, mois, 1)
    fin   = date(annee, mois, calendar.monthrange(annee, mois)[1])
    pts = (Pointage.query
           .filter_by(tenant_id=t.id, salarie_id=id)
           .filter(Pointage.date_pointage >= debut, Pointage.date_pointage <= fin)
           .options(joinedload(Pointage.site))
           .order_by(Pointage.date_pointage).all())
    ctx = _pointages_mois_contexte(t, pts, t.convention)
    return render_template("tenant/pointages_print.html",
        tenant=t, now=datetime.now(),
        personne={"nom_complet": s.nom_complet,
                  "reference": ("Matricule : " + s.matricule) if s.matricule else (s.emploi or ""),
                  "type": "Salarié"},
        mois=mois, annee=annee, mois_libelle=_MOIS_FR[mois], **ctx)


@bp.route("/journaliers/<int:id>/pointages/imprimer")
@login_required
def journalier_pointages_imprimer(id):
    """Relevé mensuel imprimable des pointages d'un journalier (totaux d'heures)."""
    if current_user.is_super_admin:
        return redirect(url_for("admin.admin_dashboard"))
    t = get_tenant()
    if not t:
        return redirect(url_for("auth.login"))
    j = Journalier.query.filter_by(id=id, tenant_id=t.id).first_or_404()
    mois, annee = _resoudre_mois_annee()
    import calendar
    debut = date(annee, mois, 1)
    fin   = date(annee, mois, calendar.monthrange(annee, mois)[1])
    pts = (Pointage.query
           .filter_by(tenant_id=t.id, journalier_id=id)
           .filter(Pointage.date_pointage >= debut, Pointage.date_pointage <= fin)
           .options(joinedload(Pointage.site))
           .order_by(Pointage.date_pointage).all())
    # Les journaliers ne relèvent pas de la ventilation conventionnelle BTP :
    # on cumule directement les colonnes pointées.
    ctx = _pointages_mois_contexte(t, pts, convention=None)
    return render_template("tenant/pointages_print.html",
        tenant=t, now=datetime.now(),
        personne={"nom_complet": j.nom_complet,
                  "reference": (j.profession or "Journalier"),
                  "type": "Journalier"},
        mois=mois, annee=annee, mois_libelle=_MOIS_FR[mois], **ctx)


@bp.route("/api/travailleur/stats-sans-site")
@tenant_required
def api_stats_sans_site():
    t = get_tenant()
    # Salariés actifs sans affectation active
    sal_ids_affectes = {a.salarie_id for a in 
        AffectationSite.query.filter_by(tenant_id=t.id, actif=True).all() if a.salarie_id}
    jour_ids_affectes = {a.journalier_id for a in 
        AffectationSite.query.filter_by(tenant_id=t.id, actif=True).all() if a.journalier_id}
    nb_sal  = Salarie.query.filter_by(tenant_id=t.id, statut="ACTIF").count()
    nb_jour = Journalier.query.filter_by(tenant_id=t.id, statut="ACTIF").count()
    sans_site = (nb_sal - len(sal_ids_affectes)) + (nb_jour - len(jour_ids_affectes))
    return jsonify({"total": max(0, sans_site)})

# ══════════════════════════════════════════════════════════════════════════════
# ── SITES & AFFECTATIONS ──────────────────────────────────────────────────────
# ══════════════════════════════════════════════════════════════════════════════

@bp.route("/composants")
@tenant_required
def composants():
    t = get_tenant()
    composants_list = ComposantPaie.query.filter_by(tenant_id=t.id).order_by(
        ComposantPaie.ordre, ComposantPaie.libelle).all()
    return render_template("tenant/composants.html", tenant=t, composants=composants_list)


@bp.route("/composants/nouveau", methods=["GET", "POST"])
@tenant_required
def composant_nouveau():
    t = get_tenant()
    if not current_user.can_edit: abort(403)
    if request.method == "POST":
        libelle = request.form.get("libelle", "").strip()
        if not libelle:
            flash("Le libellé est obligatoire.", "error")
            return render_template("tenant/composant_form.html", tenant=t, composant=None)
        c = ComposantPaie(
            tenant_id     = t.id,
            libelle       = libelle,
            sens          = "RETENUE" if request.form.get("sens") == "RETENUE" else "GAIN",
            soumis_cnss   = request.form.get("soumis_cnss")   == "on",
            soumis_cnamgs = request.form.get("soumis_cnamgs") == "on",
            soumis_irpp   = request.form.get("soumis_irpp")   == "on",
            ordre         = request.form.get("ordre", type=int) or 0,
            actif         = True,
        )
        db.session.add(c)
        db.session.commit()
        flash(f"Composant « {c.libelle} » créé.", "success")
        return redirect(url_for("tenant.composants"))
    return render_template("tenant/composant_form.html", tenant=t, composant=None)


@bp.route("/composants/<int:id>/modifier", methods=["GET", "POST"])
@tenant_required
def composant_modifier(id):
    t = get_tenant()
    if not current_user.can_edit: abort(403)
    c = ComposantPaie.query.filter_by(id=id, tenant_id=t.id).first_or_404()
    if request.method == "POST":
        libelle = request.form.get("libelle", "").strip()
        if not libelle:
            flash("Le libellé est obligatoire.", "error")
            return render_template("tenant/composant_form.html", tenant=t, composant=c)
        c.libelle       = libelle
        c.sens          = "RETENUE" if request.form.get("sens") == "RETENUE" else "GAIN"
        c.soumis_cnss   = request.form.get("soumis_cnss")   == "on"
        c.soumis_cnamgs = request.form.get("soumis_cnamgs") == "on"
        c.soumis_irpp   = request.form.get("soumis_irpp")   == "on"
        c.ordre         = request.form.get("ordre", type=int) or 0
        db.session.commit()
        flash(f"Composant « {c.libelle} » modifié.", "success")
        return redirect(url_for("tenant.composants"))
    return render_template("tenant/composant_form.html", tenant=t, composant=c)


@bp.route("/composants/<int:id>/toggle", methods=["POST"])
@tenant_required
def composant_toggle(id):
    t = get_tenant()
    if not current_user.can_edit: abort(403)
    c = ComposantPaie.query.filter_by(id=id, tenant_id=t.id).first_or_404()
    c.actif = not c.actif
    db.session.commit()
    flash(f"Composant « {c.libelle} » {'activé' if c.actif else 'désactivé'}.", "success")
    return redirect(url_for("tenant.composants"))


@bp.route("/sites")
@tenant_required
def sites():
    t = get_tenant()
    sites_list = Site.query.filter_by(tenant_id=t.id).order_by(Site.nom).all()
    return render_template("tenant/sites.html", tenant=t, sites=sites_list)

@bp.route("/sites/nouveau", methods=["GET","POST"])
@tenant_required
def site_nouveau():
    t = get_tenant()
    if request.method == "POST":
        s = Site(
            tenant_id   = t.id,
            nom         = request.form["nom"].strip(),
            code        = request.form.get("code","").strip().upper() or None,
            adresse     = request.form.get("adresse","").strip() or None,
            ville       = request.form.get("ville","").strip() or None,
            responsable = request.form.get("responsable","").strip() or None,
            telephone   = request.form.get("telephone","").strip() or None,
            description = request.form.get("description","").strip() or None,
        )
        db.session.add(s)
        db.session.commit()
        flash(f"Site « {s.nom} » créé avec succès.", "success")
        return redirect(url_for("tenant.sites"))
    return render_template("tenant/site_form.html", tenant=t, site=None)

@bp.route("/sites/<int:id>/modifier", methods=["GET","POST"])
@tenant_required
def site_modifier(id):
    t = get_tenant()
    s = Site.query.filter_by(id=id, tenant_id=t.id).first_or_404()
    if request.method == "POST":
        s.nom         = request.form["nom"].strip()
        s.code        = request.form.get("code","").strip().upper() or None
        s.adresse     = request.form.get("adresse","").strip() or None
        s.ville       = request.form.get("ville","").strip() or None
        s.responsable = request.form.get("responsable","").strip() or None
        s.telephone   = request.form.get("telephone","").strip() or None
        s.description = request.form.get("description","").strip() or None
        db.session.commit()
        flash(f"Site « {s.nom} » modifié.", "success")
        return redirect(url_for("tenant.sites"))
    return render_template("tenant/site_form.html", tenant=t, site=s)

@bp.route("/sites/<int:id>/toggle", methods=["POST"])
@tenant_required
def site_toggle(id):
    t = get_tenant()
    s = Site.query.filter_by(id=id, tenant_id=t.id).first_or_404()
    s.actif = not s.actif
    db.session.commit()
    flash(f"Site « {s.nom} » {'activé' if s.actif else 'désactivé'}.", "success")
    return redirect(url_for("tenant.sites"))

@bp.route("/sites/<int:id>")
@tenant_required
def site_detail(id):
    t = get_tenant()
    s = Site.query.filter_by(id=id, tenant_id=t.id).first_or_404()

    # Date sélectionnée pour le pointage rapide
    date_str = request.args.get("date_ptg", date.today().strftime("%Y-%m-%d"))
    try:    date_ptg = datetime.strptime(date_str, "%Y-%m-%d").date()
    except: date_ptg = date.today()

    # Affectations actives
    affectations = AffectationSite.query.filter_by(site_id=id, actif=True)        .order_by(AffectationSite.date_debut.desc()).all()

    # Séparer salariés et journaliers affectés
    ids_sal  = [a.salarie_id    for a in affectations if a.salarie_id]
    ids_jour = [a.journalier_id for a in affectations if a.journalier_id]

    salaries_site    = Salarie.query.filter(
        Salarie.tenant_id==t.id, Salarie.statut=="ACTIF",
        Salarie.id.in_(ids_sal)
    ).order_by(Salarie.nom).all() if ids_sal else []

    journaliers_site = Journalier.query.filter(
        Journalier.tenant_id==t.id, Journalier.statut=="ACTIF",
        Journalier.id.in_(ids_jour)
    ).order_by(Journalier.nom).all() if ids_jour else []

    # Pointages du jour pour ce site
    pts_sal  = {p.salarie_id:    p for p in
        Pointage.query.filter_by(tenant_id=t.id, date_pointage=date_ptg)
        .filter(Pointage.salarie_id.in_(ids_sal)).all()} if ids_sal else {}
    pts_jour = {p.journalier_id: p for p in
        Pointage.query.filter_by(tenant_id=t.id, date_pointage=date_ptg)
        .filter(Pointage.journalier_id.in_(ids_jour)).all()} if ids_jour else {}

    # Stats pointage du jour
    nb_presents  = sum(1 for p in list(pts_sal.values())+list(pts_jour.values()) if p.present)
    nb_absents   = sum(1 for p in list(pts_sal.values())+list(pts_jour.values()) if p.absent)
    nb_non_pointes = (len(salaries_site)+len(journaliers_site)) - len(pts_sal) - len(pts_jour)

    # Historique complet
    historique = AffectationSite.query.filter_by(site_id=id)        .order_by(AffectationSite.date_creation.desc()).limit(50).all()

    # Travailleurs disponibles (non affectés à ce site)
    ids_sal_aff  = {a.salarie_id    for a in affectations if a.salarie_id}
    ids_jour_aff = {a.journalier_id for a in affectations if a.journalier_id}
    salaries_dispo    = [x for x in Salarie.query.filter_by(
        tenant_id=t.id, statut="ACTIF").order_by(Salarie.nom).all()
        if x.id not in ids_sal_aff]
    journaliers_dispo = [x for x in Journalier.query.filter_by(
        tenant_id=t.id, statut="ACTIF").order_by(Journalier.nom).all()
        if x.id not in ids_jour_aff]

    # ── KPIs tableau de bord ─────────────────────────────────────────────────
    from datetime import date as _date
    import calendar

    now_d       = _date.today()
    mois_debut  = _date(now_d.year, now_d.month, 1)
    mois_fin    = _date(now_d.year, now_d.month,
                        calendar.monthrange(now_d.year, now_d.month)[1])
    lundi_sem   = now_d - timedelta(days=now_d.weekday())
    samedi_sem  = lundi_sem + timedelta(days=5)

    # Tous les pointages du mois pour ce site (salariés + journaliers)
    ids_all = ids_sal + ids_jour
    pts_mois_sal = Pointage.query.filter_by(tenant_id=t.id)        .filter(Pointage.salarie_id.in_(ids_sal),
                Pointage.date_pointage >= mois_debut,
                Pointage.date_pointage <= mois_fin).all() if ids_sal else []
    pts_mois_jour = Pointage.query.filter_by(tenant_id=t.id)        .filter(Pointage.journalier_id.in_(ids_jour),
                Pointage.date_pointage >= mois_debut,
                Pointage.date_pointage <= mois_fin).all() if ids_jour else []
    pts_mois_tous = pts_mois_sal + pts_mois_jour

    # Pointages de la semaine
    pts_sem_sal  = [p for p in pts_mois_sal  if lundi_sem <= p.date_pointage <= samedi_sem]
    pts_sem_jour = [p for p in pts_mois_jour if lundi_sem <= p.date_pointage <= samedi_sem]

    # KPI — Jours pointés (présences) ce mois
    nb_jours_pointes_mois = sum(1 for p in pts_mois_tous if p.present)
    nb_absences_mois      = sum(1 for p in pts_mois_tous if p.absent)

    # KPI — Taux de présence semaine
    nb_pres_sem = sum(1 for p in pts_sem_sal + pts_sem_jour if p.present)
    nb_abs_sem  = sum(1 for p in pts_sem_sal + pts_sem_jour if p.absent)
    total_ptg_sem = nb_pres_sem + nb_abs_sem
    taux_presence_semaine = round(nb_pres_sem / total_ptg_sem * 100) if total_ptg_sem > 0 else 0

    # KPI — Heures totales semaine
    def total_heures_pt(p):
        return (float(p.heures_normales or 8) +
                float(p.heures_sup_10 or 0) + float(p.heures_sup_30 or 0) +
                float(p.heures_sup_40 or 0) + float(p.heures_sup_70 or 0) +
                float(p.heures_sup or 0))

    heures_semaine = sum(total_heures_pt(p) for p in pts_sem_sal + pts_sem_jour if p.present)

    # KPI — Masse journalière (feuilles de paie journaliers ce mois)
    feuilles_mois = FeuillePaieJournalier.query.filter_by(tenant_id=t.id)        .filter(FeuillePaieJournalier.journalier_id.in_(ids_jour),
                FeuillePaieJournalier.date_debut >= mois_debut,
                FeuillePaieJournalier.date_fin   <= mois_fin).all() if ids_jour else []
    masse_journaliere_mois    = sum(float(f.montant_brut or 0) for f in feuilles_mois)
    feuilles_attente = sum(1 for f in feuilles_mois if f.statut == "EN_ATTENTE")
    feuilles_payees  = sum(1 for f in feuilles_mois if f.statut == "PAYÉ")

    # KPI — Bulletins salariés du mois (dernière période active)
    periode_courante = PeriodePaie.query.filter_by(
        tenant_id=t.id, annee=now_d.year, mois=now_d.month).first()
    bulletins_site = []
    masse_mensuelle = 0
    if periode_courante and ids_sal:
        bulletins_site = BulletinPaie.query.filter_by(
            tenant_id=t.id, periode_id=periode_courante.id
        ).filter(BulletinPaie.salarie_id.in_(ids_sal)).all()
        masse_mensuelle = sum(float(b.net_a_payer or 0) for b in bulletins_site)

    # Évolution présence 7 derniers jours (pour mini-graphique)
    evolution_7j = []
    for i in range(6, -1, -1):
        d = now_d - timedelta(days=i)
        p_d = [p for p in pts_mois_tous if p.date_pointage == d]
        nb_p = sum(1 for p in p_d if p.present)
        nb_a = sum(1 for p in p_d if p.absent)
        evolution_7j.append({
            "date":    d.strftime("%d/%m"),
            "jour":    ["L","Ma","Me","J","V","Sa","Di"][d.weekday()],
            "presents": nb_p,
            "absents":  nb_a,
            "heures":   round(sum(total_heures_pt(p) for p in p_d if p.present), 1),
        })

    return render_template("tenant/site_detail.html",
        tenant=t, site=s,
        affectations=affectations, historique=historique,
        salaries_dispo=salaries_dispo, journaliers_dispo=journaliers_dispo,
        salaries_site=salaries_site, journaliers_site=journaliers_site,
        pts_sal=pts_sal, pts_jour=pts_jour,
        date_ptg=date_ptg,
        date_hier=(date_ptg - timedelta(days=1)).strftime("%Y-%m-%d"),
        date_demain=(date_ptg + timedelta(days=1)).strftime("%Y-%m-%d"),
        nb_presents=nb_presents, nb_absents=nb_absents,
        nb_non_pointes=nb_non_pointes,
        # KPIs tableau de bord
        nb_sal_site=len(salaries_site), nb_jour_site=len(journaliers_site),
        nb_jours_pointes_mois=nb_jours_pointes_mois,
        nb_absences_mois=nb_absences_mois,
        taux_presence_semaine=taux_presence_semaine,
        heures_semaine=round(heures_semaine, 1),
        masse_journaliere_mois=masse_journaliere_mois,
        feuilles_attente=feuilles_attente, feuilles_payees=feuilles_payees,
        masse_mensuelle=masse_mensuelle,
        nb_bulletins_site=len(bulletins_site),
        evolution_7j=evolution_7j,
        mois_nom=["","Jan","Fév","Mar","Avr","Mai","Jun","Jul","Aoû","Sep","Oct","Nov","Déc"][now_d.month],
        today=str(date.today()))

@bp.route("/sites/<int:id>/pointage-rapide", methods=["POST"])
@tenant_required
def site_pointage_rapide(id):
    """Sauvegarder le pointage rapide depuis la page d'un site."""
    t = get_tenant()
    s = Site.query.filter_by(id=id, tenant_id=t.id).first_or_404()
    date_str = request.form.get("date_pointage")
    try:    date_p = datetime.strptime(date_str, "%Y-%m-%d").date()
    except: date_p = date.today()

    nb = 0
    for key, val in request.form.items():
        # Salariés
        if key.startswith("sal_present_"):
            sid = int(key.replace("sal_present_", ""))
            present = (val == "1"); absent = not present
            pt = Pointage.query.filter_by(
                tenant_id=t.id, date_pointage=date_p, salarie_id=sid).first()
            if not pt:
                pt = Pointage(tenant_id=t.id, date_pointage=date_p,
                              salarie_id=sid, site_id=id)
                db.session.add(pt)
            pt.present = present; pt.absent = absent
            pt.heures_normales = float(request.form.get(f"sal_h_{sid}", 8) or 8)
            pt.heures_sup_10   = float(request.form.get(f"sal_s10_{sid}", 0) or 0)
            pt.heures_sup_30   = float(request.form.get(f"sal_s30_{sid}", 0) or 0)
            pt.heures_sup_30b  = float(request.form.get(f"sal_s30b_{sid}", 0) or 0)
            pt.heures_sup_40   = float(request.form.get(f"sal_s40_{sid}", 0) or 0)
            pt.heures_sup_70   = float(request.form.get(f"sal_s70_{sid}", 0) or 0)
            pt.motif_absence   = request.form.get(f"sal_motif_{sid}", "") if absent else None
            nb += 1
        # Journaliers
        elif key.startswith("jour_present_"):
            jid = int(key.replace("jour_present_", ""))
            present = (val == "1"); absent = not present
            pt = Pointage.query.filter_by(
                tenant_id=t.id, date_pointage=date_p, journalier_id=jid).first()
            if not pt:
                pt = Pointage(tenant_id=t.id, date_pointage=date_p,
                              journalier_id=jid, site_id=id)
                db.session.add(pt)
            pt.present = present; pt.absent = absent
            pt.heures_normales = float(request.form.get(f"jour_h_{jid}", 8) or 8)
            pt.heures_sup      = float(request.form.get(f"jour_s_{jid}", 0) or 0)
            pt.motif_absence   = request.form.get(f"jour_motif_{jid}", "") if absent else None
            nb += 1

    db.session.commit()
    flash(f"✅ Pointage du {date_p.strftime('%d/%m/%Y')} sauvegardé — {nb} travailleur(s).", "success")
    return redirect(url_for("tenant.site_detail", id=id) + f"?date_ptg={date_str}")

@bp.route("/sites/<int:site_id>/affecter", methods=["POST"])
@tenant_required
def site_affecter(site_id):
    """Affecter un ou plusieurs travailleurs à un site."""
    t  = get_tenant()
    s  = Site.query.filter_by(id=site_id, tenant_id=t.id).first_or_404()
    date_debut = request.form.get("date_debut") or str(date.today())
    motif      = request.form.get("motif","").strip() or None
    nb         = 0

    for key in request.form:
        if key.startswith("sal_"):
            sal_id = int(key[4:])
            sal = Salarie.query.filter_by(id=sal_id, tenant_id=t.id).first()
            if not sal: continue
            # Désactiver toute affectation active précédente sur un AUTRE site
            prev = AffectationSite.query.filter_by(
                salarie_id=sal_id, actif=True, tenant_id=t.id).first()
            if prev and prev.site_id != site_id:
                prev.actif    = False
                prev.date_fin = date.today()
                prev.motif    = f"Transféré vers {s.nom}"
            elif prev and prev.site_id == site_id:
                continue  # Déjà sur ce site
            a = AffectationSite(
                tenant_id=t.id, site_id=site_id, salarie_id=sal_id,
                date_debut=date_debut, actif=True, motif=motif,
                cree_par=current_user.email)
            db.session.add(a); nb += 1

        elif key.startswith("jour_"):
            jour_id = int(key[5:])
            jour = Journalier.query.filter_by(id=jour_id, tenant_id=t.id).first()
            if not jour: continue
            prev = AffectationSite.query.filter_by(
                journalier_id=jour_id, actif=True, tenant_id=t.id).first()
            if prev and prev.site_id != site_id:
                prev.actif    = False
                prev.date_fin = date.today()
                prev.motif    = f"Transféré vers {s.nom}"
            elif prev and prev.site_id == site_id:
                continue
            a = AffectationSite(
                tenant_id=t.id, site_id=site_id, journalier_id=jour_id,
                date_debut=date_debut, actif=True, motif=motif,
                cree_par=current_user.email)
            db.session.add(a); nb += 1

    db.session.commit()
    flash(f"{nb} travailleur(s) affecté(s) à « {s.nom} ».", "success")
    return redirect(url_for("tenant.site_detail", id=site_id))

@bp.route("/sites/affecter-travailleur/<int:affectation_id>/retirer", methods=["POST"])
@tenant_required
def site_retirer(affectation_id):
    """Retirer un travailleur de son site (fin d'affectation)."""
    t = get_tenant()
    a = AffectationSite.query.filter_by(id=affectation_id, tenant_id=t.id).first_or_404()
    motif = request.form.get("motif","").strip() or "Retrait manuel"
    a.actif    = False
    a.date_fin = date.today()
    a.motif    = motif
    db.session.commit()
    flash(f"Affectation terminée pour {a.nom_travailleur}.", "success")
    return redirect(url_for("tenant.site_detail", id=a.site_id))

@bp.route("/sites/permuter", methods=["POST"])
@tenant_required
def site_permuter():
    """Permuter un travailleur d'un site vers un autre."""
    t         = get_tenant()
    aff_id    = request.form.get("affectation_id", type=int)
    nouveau_site_id = request.form.get("site_destination_id", type=int)
    motif     = request.form.get("motif","Permutation").strip()

    aff_old = AffectationSite.query.filter_by(id=aff_id, tenant_id=t.id, actif=True).first_or_404()
    site_dest = Site.query.filter_by(id=nouveau_site_id, tenant_id=t.id).first_or_404()

    # Fermer l'affectation actuelle
    aff_old.actif    = False
    aff_old.date_fin = date.today()
    aff_old.motif    = f"Permuté vers {site_dest.nom} — {motif}"

    # Créer la nouvelle affectation
    aff_new = AffectationSite(
        tenant_id     = t.id,
        site_id       = nouveau_site_id,
        salarie_id    = aff_old.salarie_id,
        journalier_id = aff_old.journalier_id,
        date_debut    = date.today(),
        actif         = True,
        motif         = f"Permuté depuis {aff_old.site.nom} — {motif}",
        cree_par      = current_user.email,
    )
    db.session.add(aff_new)
    db.session.commit()
    flash(f"{aff_old.nom_travailleur} permuté vers « {site_dest.nom} ».", "success")
    return redirect(url_for("tenant.site_detail", id=nouveau_site_id))

@bp.route("/api/travailleur/<string:type>/<int:id>/site")
@tenant_required
def api_travailleur_site(type, id):
    """API : retourne le site actuel d'un travailleur."""
    t = get_tenant()
    if type == "salarie":
        a = AffectationSite.query.filter_by(
            salarie_id=id, tenant_id=t.id, actif=True).first()
    else:
        a = AffectationSite.query.filter_by(
            journalier_id=id, tenant_id=t.id, actif=True).first()
    if a:
        return jsonify({"site_id": a.site_id, "site_nom": a.site.nom,
                        "date_debut": str(a.date_debut)})
    return jsonify({"site_id": None, "site_nom": None})


@bp.route("/journaliers/imprimer")
@login_required
def journaliers_imprimer():
    if current_user.is_super_admin: return redirect(url_for("admin.admin_dashboard"))
    t = get_tenant()
    if not t: return redirect(url_for("auth.login"))
    return render_template("tenant/journaliers_print.html", journaliers=Journalier.query.filter_by(tenant_id=t.id).order_by(Journalier.nom).all(), tenant=t, now=datetime.now())

# ── Export & API ──────────────────────────────────────────────────────────────
@bp.route("/bulletins/export/<int:periode_id>")
@tenant_required
def export_journal(periode_id):
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, PatternFill, Alignment
    t=get_tenant()
    p=PeriodePaie.query.filter_by(id=periode_id,tenant_id=t.id).first_or_404()
    buls=BulletinPaie.query.filter_by(periode_id=periode_id,tenant_id=t.id).join(Salarie).order_by(Salarie.nom).all()
    wb=Workbook(); ws=wb.active; ws.title=f"Journal {p.libelle_complet}"
    ws.merge_cells("A1:R1"); ws["A1"]=f"JOURNAL DE PAIE — {p.libelle_complet} — {t.denomination}"
    ws["A1"].font=Font(bold=True,size=13); ws["A1"].alignment=Alignment(horizontal="center")
    hdrs=["Matricule","Nom","Prénom","Emploi","Cat.","Base","Brut","CNSS Sal.","CNAMGS Sal.","TCS","IRPP","Net","Net à Payer","CNSS Pat.","CNAMGS Pat.","FNH","CFP","Statut"]
    for col,h in enumerate(hdrs,1):
        c=ws.cell(row=3,column=col,value=h); c.font=Font(bold=True,color="FFFFFF")
        c.fill=PatternFill("solid",fgColor="1a2332"); c.alignment=Alignment(horizontal="center")
    for row,b in enumerate(buls,4):
        s=b.salarie
        vals=[s.matricule,s.nom,s.prenom,s.emploi,s.categorie.code if s.categorie else "",
              float(b.salaire_base or 0),float(b.salaire_brut or 0),float(b.cnss_salarie or 0),
              float(b.cnamgs_salarie or 0),float(b.tcs or 0),float(b.irpp or 0),
              float(b.salaire_net or 0),float(b.net_a_payer or 0),float(b.cnss_patronale or 0),
              float(b.cnamgs_patronale or 0),float(b.fnh or 0),float(b.cfp or 0),b.statut]
        for col,v in enumerate(vals,1):
            cell=ws.cell(row=row,column=col,value=csv_safe(v) if isinstance(v,str) else v)
            if isinstance(v,float): cell.number_format='#,##0'
            if row%2==0: cell.fill=PatternFill("solid",fgColor="F5F5F5")
    out=io.BytesIO(); wb.save(out); out.seek(0)
    return send_file(out,mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,download_name=f"Journal_{p.libelle_mois}_{p.annee}_{t.slug}.xlsx")

@bp.route("/api/calculer-bulletin", methods=["POST"])
@login_required
def api_calculer():
    try:
        t = get_tenant()
        data = request.get_json() or {}
        sid = data.pop("salarie_id", None)
        nb_parts = 1.0
        if sid and t:
            s = Salarie.query.filter_by(id=sid, tenant_id=t.id).first()
            if s: nb_parts = float(s.nombre_parts or 1)
        mois  = data.pop("mois_periode", None)
        annee = data.pop("annee_periode", None)
        # ── Ancienneté du salarié (pour la prime d'ancienneté automatique) ──
        # Référence : fin de la période de paie si connue, sinon aujourd'hui.
        if sid and t:
            s_anc = Salarie.query.filter_by(id=sid, tenant_id=t.id).first()
            if s_anc and s_anc.date_embauche:
                from datetime import date as _date
                try:
                    if mois and annee:
                        m, a = int(mois), int(annee)
                        ref = (_date(a + 1, 1, 1) if m == 12 else _date(a, m + 1, 1))
                        ref = ref - timedelta(days=1)      # dernier jour du mois
                    else:
                        ref = _date.today()
                except (TypeError, ValueError):
                    ref = _date.today()
                data["anciennete_annees"] = max(0, (ref - s_anc.date_embauche).days // 365)
        total_acomptes = 0.0
        if sid and t and mois and annee:
            total_acomptes = float(db.session.query(db.func.sum(Acompte.montant))
                .filter_by(tenant_id=t.id, salarie_id=int(sid), mois=int(mois),
                           annee=int(annee), statut="EN_ATTENTE").scalar() or 0)
        if total_acomptes > 0:
            data["acompte"] = max(float(data.get("acompte", 0)), total_acomptes)
        # Composants personnalisés (aperçu temps réel) : lus depuis composant_<id>
        if t:
            comps_live = []
            for comp in ComposantPaie.query.filter_by(tenant_id=t.id, actif=True).all():
                try:
                    montant = float(data.get(f"composant_{comp.id}") or 0)
                except (TypeError, ValueError):
                    montant = 0.0
                if montant:
                    comps_live.append({
                        "libelle": comp.libelle, "sens": comp.sens, "montant": montant,
                        "soumis_cnss": comp.soumis_cnss, "soumis_cnamgs": comp.soumis_cnamgs,
                        "soumis_irpp": comp.soumis_irpp})
            data["composants"] = comps_live
        if t:
            data["convention"] = t.convention
        res = calculer_bulletin(data, nb_parts=nb_parts)
        res["acompte_auto"] = total_acomptes
        return jsonify(res)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@bp.route("/api/salarie/<int:id>/contrat")
@login_required
def api_contrat(id):
    t=get_tenant()
    s=Salarie.query.filter_by(id=id,tenant_id=t.id).first()
    if not s: return jsonify({})
    c=Contrat.query.filter_by(salarie_id=id,tenant_id=t.id,actif=True).first()
    base={"nom":s.nom_complet,"poste":s.emploi,"matricule":s.matricule,"nombre_parts":float(s.nombre_parts or 1)}
    if c: base["salaire_base"]=float(c.salaire_base); base["poste"]=c.poste or s.emploi
    return jsonify(base)

@bp.route("/api/salarie/<int:id>/pointage-mois")
@login_required
def api_pointage_mois(id):
    """Retourne le cumul des heures du pointage pour un salarié sur un mois donné."""
    t = get_tenant()
    if not t: return jsonify({})
    mois  = request.args.get("mois",  type=int)
    annee = request.args.get("annee", type=int)
    if not mois or not annee:
        return jsonify({"erreur": "mois et annee requis"})
    import calendar
    dernier_jour = calendar.monthrange(annee, mois)[1]
    debut = date(annee, mois, 1)
    fin   = date(annee, mois, dernier_jour)
    pts = Pointage.query.filter_by(tenant_id=t.id, salarie_id=id)        .filter(Pointage.date_pointage >= debut, Pointage.date_pointage <= fin).all()
    pts_travailles = [p for p in pts if p.present and not p.absent]
    if not pts_travailles:
        return jsonify({"nb_jours": 0, "nb_absences": 0,
            "heures_sup_10": 0, "heures_sup_30": 0, "heures_sup_40": 0, "heures_sup_70": 0,
            "heures_normales_total": 0, "total_sup": 0,
            "message": "Aucun pointage pour cette période"})
    nb_jours = len(pts_travailles)

    if (t.convention or "").upper() in ("BTP", "PETROLE", "INDUSTRIE", "AERIEN"):
        # Ventilation réglementaire (BTP/Pétrole/Industrie) : semaine par semaine, ligne par ligne
        from calculs_paie import ventiler_heures_mois, pointage_vers_jours
        v = ventiler_heures_mois(t.convention, pointage_vers_jours(pts), seuil_normales=t.seuil_hs)
        heures_normales = v["heures_normales"]
        heures_sup_10   = v["heures_sup_10"]
        heures_sup_30   = v["heures_sup_30"]
        heures_sup_30b  = v.get("heures_sup_30b", 0.0)
        heures_sup_40   = v["heures_sup_40"]
        heures_sup_70   = v["heures_sup_70"]
        detail_semaines = v["detail_semaines"]
    else:
        # Autres conventions : cumul direct des colonnes déjà ventilées par jour
        heures_normales = sum(float(p.heures_normales or 8) for p in pts_travailles)
        heures_sup_10   = sum(float(p.heures_sup_10 or 0) for p in pts_travailles)
        heures_sup_30   = sum(float(p.heures_sup_30 or 0) for p in pts_travailles)
        heures_sup_30b  = sum(float(getattr(p, "heures_sup_30b", 0) or 0) for p in pts_travailles)
        heures_sup_40   = sum(float(p.heures_sup_40 or 0) for p in pts_travailles)
        heures_sup_70   = sum(float(p.heures_sup_70 or 0) for p in pts_travailles)
        detail_semaines = []
    pts_absents = Pointage.query.filter_by(tenant_id=t.id, salarie_id=id)        .filter(Pointage.date_pointage >= debut, Pointage.date_pointage <= fin,
                Pointage.absent == True).all()
    return jsonify({
        "nb_jours":              nb_jours,
        "nb_absences":           len(pts_absents),
        "heures_normales_total": round(heures_normales, 2),
        "heures_sup_10":         round(heures_sup_10, 2),
        "heures_sup_30":         round(heures_sup_30, 2),
        "heures_sup_30b":        round(heures_sup_30b, 2),
        "heures_sup_40":         round(heures_sup_40, 2),
        "heures_sup_70":         round(heures_sup_70, 2),
        "total_sup":             round(heures_sup_10+heures_sup_30+heures_sup_30b+heures_sup_40+heures_sup_70, 2),
        "detail_semaines":       detail_semaines,
        "message":               f"{nb_jours} jour(s) pointé(s) sur {dernier_jour}"
    })

@bp.route("/api/cache/clear", methods=["POST"])
@login_required
def api_cache_clear():
    """Vider le cache du dashboard (bouton rafraîchir)."""
    t = get_tenant()
    if not t: return jsonify({"ok": False})
    _cache_delete(f"{t.id}:")
    return jsonify({"ok": True, "msg": "Cache vidé"})


@bp.route("/api/jour-ferie")
@login_required
def api_jour_ferie():
    """
    Indique si une date est un jour férié / dimanche, pour pré-remplir
    automatiquement le type de jour dans le pointage.
    Param : date (format YYYY-MM-DD)
    Retour : {type_jour, est_ferie, nom, est_dimanche}
    """
    t = get_tenant()
    if not t:
        return jsonify({"erreur": "non connecté"}), 401
    date_str = request.args.get("date", "")
    d = _parse_date(date_str)
    if not d:
        return jsonify({"erreur": "date invalide"}), 400
    nom = nom_jour_ferie(d)
    return jsonify({
        "date":         d.strftime("%Y-%m-%d"),
        "type_jour":    type_jour_auto(d),
        "est_ferie":    nom is not None,
        "nom":          nom,
        "est_dimanche": d.weekday() == 6,
    })


@bp.route("/notifications")
@login_required
def notifications_page():
    """Page listant tous les rappels et notifications."""
    if current_user.is_super_admin:
        return redirect(url_for("admin.admin_dashboard"))
    t = get_tenant()
    if not t:
        return redirect(url_for("auth.login"))
    notifs = get_notifications(t, db, _NOTIF_MODELS)
    # Grouper par catégorie pour l'affichage
    par_categorie = {}
    for n in notifs:
        par_categorie.setdefault(n["categorie"], []).append(n)
    labels_cat = {
        "contrat": "Contrats", "conge": "Congés",
        "facture": "Factures prestataires", "periode": "Périodes de paie",
    }
    nb_critiques = sum(1 for n in notifs if n["type"] == "danger")
    return render_template("tenant/notifications.html",
        tenant=t, notifs=notifs, par_categorie=par_categorie,
        labels_cat=labels_cat, total=len(notifs), nb_critiques=nb_critiques)


@bp.route("/api/notifications/count")
@login_required
def api_notifications_count():
    """Compteur pour la cloche (rafraîchi en arrière-plan)."""
    t = get_tenant()
    if not t:
        return jsonify({"total": 0, "critiques": 0})
    total, critiques = compter_notifications(t, db, _NOTIF_MODELS)
    return jsonify({"total": total, "critiques": critiques})


@bp.route("/jours-feries")
@login_required
def jours_feries_page():
    """Affiche le calendrier des jours fériés gabonais pour une année."""
    if current_user.is_super_admin:
        return redirect(url_for("admin.admin_dashboard"))
    t = get_tenant()
    if not t:
        return redirect(url_for("auth.login"))
    annee = request.args.get("annee", datetime.now().year, type=int)
    feries = sorted(jours_feries_annee(annee).items())
    jours_sem = ["Lundi", "Mardi", "Mercredi", "Jeudi",
                 "Vendredi", "Samedi", "Dimanche"]
    feries_fmt = [{
        "date":      d,
        "nom":       nom,
        "jour_sem":  jours_sem[d.weekday()],
        "est_we":    d.weekday() >= 5,
    } for d, nom in feries]
    annees_dispo = list(range(datetime.now().year - 1, datetime.now().year + 3))
    return render_template("tenant/jours_feries.html",
                           tenant=t, annee=annee, feries=feries_fmt,
                           annees_dispo=annees_dispo)


@bp.route("/api/semaine-btp")
@login_required
def api_semaine_btp():
    """
    Calcule la distribution BTP des heures pour un travailleur sur une semaine.
    Params : type (sal|jour), id, date (n'importe quel jour de la semaine)
    """
    t = get_tenant()
    if not t: return jsonify({"erreur": "non connecté"})

    type_w    = request.args.get("type", "sal")
    worker_id = request.args.get("id", type=int)
    date_str  = request.args.get("date", datetime.now().strftime("%Y-%m-%d"))

    try:
        date_ref = datetime.strptime(date_str, "%Y-%m-%d").date()
    except:
        date_ref = datetime.now().date()

    lundi  = date_ref - timedelta(days=date_ref.weekday())
    samedi = lundi + timedelta(days=5)

    if type_w == "sal":
        pts = Pointage.query.filter_by(tenant_id=t.id, salarie_id=worker_id)            .filter(Pointage.date_pointage >= lundi,
                    Pointage.date_pointage <= samedi,
                    Pointage.present == True).all()
    else:
        pts = Pointage.query.filter_by(tenant_id=t.id, journalier_id=worker_id)            .filter(Pointage.date_pointage >= lundi,
                    Pointage.date_pointage <= samedi,
                    Pointage.present == True).all()

    if not pts:
        return jsonify({
            "semaine": f"{lundi.strftime('%d/%m')} → {samedi.strftime('%d/%m/%Y')}",
            "nb_jours": 0, "heures_normales": 0,
            "heures_sup_10": 0, "heures_sup_30": 0,
            "heures_sup_40": 0, "heures_sup_70": 0,
            "message": "Aucun pointage cette semaine"
        })

    from calculs_paie import distribuer_heures_semaine_btp
    jours_data = []
    for p in pts:
        h_norm = float(p.heures_normales or 0)
        if type_w == "sal":
            h_nuit = float(p.heures_sup_40 or 0)
        else:
            h_nuit = 0
        jours_data.append({
            "heures_normales": h_norm,
            "heures_sup_nuit": h_nuit,
            "type_jour": p.type_jour or "NORMAL"
        })

    dist = distribuer_heures_semaine_btp(jours_data, seuil_normales=t.seuil_hs)
    dist["semaine"]  = f"{lundi.strftime('%d/%m')} → {samedi.strftime('%d/%m/%Y')}"
    dist["nb_jours"] = len(pts)
    dist["jours_detail"] = [
        {
            "date":     str(p.date_pointage),
            "jour_fr":  ["Lun","Mar","Mer","Jeu","Ven","Sam","Dim"][p.date_pointage.weekday()],
            "heures":   float(p.heures_normales or 0),
            "type_jour": p.type_jour or "NORMAL",
        } for p in sorted(pts, key=lambda x: x.date_pointage)
    ]

    # Montants si salaire connu
    salaire_base = request.args.get("salaire", type=float)
    if salaire_base:
        from calculs_paie import calculer_taux_horaire, COEFF_SUP_10, COEFF_SUP_30
        th = calculer_taux_horaire(salaire_base)
        dist["taux_horaire"]    = round(th, 2)
        dist["montant_10"]      = round(dist["heures_sup_10"] * th * COEFF_SUP_10, 2)
        dist["montant_30"]      = round(dist["heures_sup_30"] * th * COEFF_SUP_30, 2)
        dist["montant_total_sup"] = round(dist["montant_10"] + dist["montant_30"]
                                         + dist["heures_sup_40"] * th * 1.40
                                         + dist["heures_sup_70"] * th * 1.70, 2)

    return jsonify(dist)


@bp.route("/api/salarie/<int:id>/acomptes-mois")
@login_required
def api_acomptes_mois(id):
    t = get_tenant()
    mois=request.args.get("mois",type=int); annee=request.args.get("annee",type=int)
    if not t or not mois or not annee: return jsonify({"total":0})
    total = db.session.query(db.func.sum(Acompte.montant))\
            .filter_by(tenant_id=t.id,salarie_id=id,mois=mois,annee=annee,statut="EN_ATTENTE").scalar() or 0
    return jsonify({"total":float(total)})

@bp.route("/pointage/recap-semaine")
@login_required
def pointage_recap_semaine():
    """Récapitulatif de présence hebdomadaire par site."""
    if current_user.is_super_admin: return redirect(url_for("admin.admin_dashboard"))
    t = get_tenant()
    if not t: return redirect(url_for("auth.login"))

    # Semaine sélectionnée
    date_str = request.args.get("date", datetime.now().strftime("%Y-%m-%d"))
    try:    date_ref = datetime.strptime(date_str, "%Y-%m-%d").date()
    except: date_ref = datetime.now().date()

    lundi  = date_ref - timedelta(days=date_ref.weekday())
    jours  = [lundi + timedelta(days=i) for i in range(6)]  # lundi→samedi
    samedi = jours[-1]

    # Sites actifs
    sites_list = Site.query.filter_by(tenant_id=t.id, actif=True).order_by(Site.nom).all()

    # Tous les pointages de la semaine
    pts_semaine = Pointage.query.filter_by(tenant_id=t.id)        .filter(Pointage.date_pointage >= lundi,
                Pointage.date_pointage <= samedi).all()

    # Affectations actives (site → workers)
    aff_sal  = {}  # salarie_id  → site_id
    aff_jour = {}  # journalier_id → site_id
    for a in AffectationSite.query.filter_by(tenant_id=t.id, actif=True).all():
        if a.salarie_id:    aff_sal[a.salarie_id]    = a.site_id
        if a.journalier_id: aff_jour[a.journalier_id] = a.site_id

    # ── Construire les données par site et par jour ───────────────────────────
    # Structure : { site_id: { date: { presents, absents, heures, h_sup, non_pointes } } }
    JOURS_FR = ["Lundi","Mardi","Mercredi","Jeudi","Vendredi","Samedi"]

    # Nb de travailleurs affectés à chaque site
    effectif_site = {}
    for s in sites_list:
        nb_sal  = sum(1 for v in aff_sal.values()  if v == s.id)
        nb_jour = sum(1 for v in aff_jour.values() if v == s.id)
        effectif_site[s.id] = nb_sal + nb_jour

    recap = {}
    for s in sites_list:
        recap[s.id] = {
            "site": s,
            "effectif": effectif_site.get(s.id, 0),
            "jours": {},
            "totaux": {"presents": 0, "absents": 0, "heures": 0.0, "h_sup": 0.0},
        }
        for j in jours:
            recap[s.id]["jours"][str(j)] = {
                "date":        j,
                "jour_fr":     JOURS_FR[j.weekday()],
                "presents":    0,
                "absents":     0,
                "non_pointes": effectif_site.get(s.id, 0),
                "heures":      0.0,
                "h_sup":       0.0,
            }

    # Sac sans site
    recap["sans_site"] = {
        "site": None,
        "effectif": 0,
        "jours": {},
        "totaux": {"presents": 0, "absents": 0, "heures": 0.0, "h_sup": 0.0},
    }
    for j in jours:
        recap["sans_site"]["jours"][str(j)] = {
            "date": j, "jour_fr": JOURS_FR[j.weekday()],
            "presents": 0, "absents": 0, "non_pointes": 0,
            "heures": 0.0, "h_sup": 0.0,
        }

    # Remplir avec les pointages réels
    for p in pts_semaine:
        d = str(p.date_pointage)
        site_id = aff_sal.get(p.salarie_id) or aff_jour.get(p.journalier_id)
        key = site_id if site_id and site_id in recap else "sans_site"

        if d not in recap[key]["jours"]:
            continue

        cell = recap[key]["jours"][d]
        if p.present:
            cell["presents"]    += 1
            h_norm = float(p.heures_normales or 8)
            h_sup  = (float(p.heures_sup_10 or 0) + float(p.heures_sup_30 or 0) +
                      float(p.heures_sup_40 or 0) + float(p.heures_sup_70 or 0) +
                      float(p.heures_sup or 0))
            cell["heures"] += h_norm
            cell["h_sup"]  += h_sup
        else:
            cell["absents"] += 1
        # Recalcul non_pointés
        cell["non_pointes"] = max(0,
            recap[key]["effectif"] - cell["presents"] - cell["absents"])

    # Calculer les totaux semaine par site
    for key in recap:
        tot = recap[key]["totaux"]
        for d, cell in recap[key]["jours"].items():
            tot["presents"] += cell["presents"]
            tot["absents"]  += cell["absents"]
            tot["heures"]   += cell["heures"]
            tot["h_sup"]    += cell["h_sup"]

    # Totaux globaux tous sites
    totaux_globaux = {"presents": 0, "absents": 0, "heures": 0.0, "h_sup": 0.0}
    for key in recap:
        t2 = recap[key]["totaux"]
        totaux_globaux["presents"] += t2["presents"]
        totaux_globaux["absents"]  += t2["absents"]
        totaux_globaux["heures"]   += t2["heures"]
        totaux_globaux["h_sup"]    += t2["h_sup"]

    # Filtrer "sans_site" si vide
    if recap["sans_site"]["totaux"]["presents"] == 0 and recap["sans_site"]["totaux"]["absents"] == 0:
        del recap["sans_site"]

    return render_template("tenant/pointage_recap_semaine.html",
        tenant=t,
        jours=jours,
        jours_fr=JOURS_FR,
        recap=recap,
        sites_list=sites_list,
        totaux_globaux=totaux_globaux,
        lundi=lundi,
        samedi=samedi,
        date_ref=date_ref,
        semaine_prec=(lundi - timedelta(days=7)).strftime("%Y-%m-%d"),
        semaine_suiv=(lundi + timedelta(days=7)).strftime("%Y-%m-%d"),
        now=datetime.now())

@bp.route("/api/pointage/semaine")
@login_required
def api_pointage_semaine():
    t = get_tenant()
    if not t: return jsonify({})
    date_str = request.args.get("date")
    try: date_sel = datetime.strptime(date_str, "%Y-%m-%d").date()
    except: date_sel = datetime.now().date()
    lundi=date_sel-timedelta(days=date_sel.weekday()); samedi=lundi+timedelta(days=5)
    pts = Pointage.query.filter_by(tenant_id=t.id).filter(Pointage.date_pointage>=lundi,Pointage.date_pointage<=samedi).all()
    stats={}
    for p in pts:
        key=str(p.date_pointage)
        if key not in stats: stats[key]={"presents":0,"absents":0,"heures":0}
        if p.present: stats[key]["presents"]+=1; stats[key]["heures"]+=p.total_heures
        else: stats[key]["absents"]+=1
    return jsonify(stats)


# ══════════════════════════════════════════════════════════════════════════════
# RAPPORT MENSUEL PAR SITE
# ══════════════════════════════════════════════════════════════════════════════

@bp.route("/rapports/mensuel-site")
@login_required
def rapport_mensuel_site():
    """Page rapport mensuel par site : pointage + paie journalier + masse salariale."""
    if current_user.is_super_admin: return redirect(url_for("admin.admin_dashboard"))
    t = get_tenant()
    if not t: return redirect(url_for("auth.login"))

    MOIS_NOMS_LONG = ["","Janvier","Février","Mars","Avril","Mai","Juin",
                      "Juillet","Août","Septembre","Octobre","Novembre","Décembre"]

    # Paramètres
    now_d  = datetime.now()
    mois   = request.args.get("mois",  type=int, default=now_d.month)
    annee  = request.args.get("annee", type=int, default=now_d.year)
    site_id= request.args.get("site_id", type=int)

    import calendar
    _, nb_jours_mois = calendar.monthrange(annee, mois)
    mois_debut = date(annee, mois, 1)
    mois_fin   = date(annee, mois, nb_jours_mois)

    sites_list = Site.query.filter_by(tenant_id=t.id, actif=True).order_by(Site.nom).all()
    site_sel   = Site.query.filter_by(id=site_id, tenant_id=t.id).first() if site_id else None

    # Affectations actives pour ce mois
    aff_sal  = {}   # salarie_id  → site_id
    aff_jour = {}   # journalier_id → site_id
    for a in AffectationSite.query.filter_by(tenant_id=t.id).all():
        if a.actif or (a.date_fin and a.date_fin >= mois_debut):
            if a.salarie_id:    aff_sal[a.salarie_id]    = a.site_id
            if a.journalier_id: aff_jour[a.journalier_id] = a.site_id

    def _build_rapport_site(s):
        """Construit le rapport complet pour un site donné."""
        sid = s.id
        # Travailleurs affectés à ce site
        ids_sal  = [k for k,v in aff_sal.items()  if v == sid]
        ids_jour = [k for k,v in aff_jour.items() if v == sid]

        salaries_aff    = Salarie.query.filter(
            Salarie.tenant_id==t.id, Salarie.id.in_(ids_sal)
        ).order_by(Salarie.nom).all() if ids_sal else []
        journaliers_aff = Journalier.query.filter(
            Journalier.tenant_id==t.id, Journalier.id.in_(ids_jour)
        ).order_by(Journalier.nom).all() if ids_jour else []

        # ── Pointage mensuel ─────────────────────────────────────────────────
        pts_sal  = Pointage.query.filter_by(tenant_id=t.id)            .filter(Pointage.salarie_id.in_(ids_sal),
                    Pointage.date_pointage >= mois_debut,
                    Pointage.date_pointage <= mois_fin).all() if ids_sal else []
        pts_jour = Pointage.query.filter_by(tenant_id=t.id)            .filter(Pointage.journalier_id.in_(ids_jour),
                    Pointage.date_pointage >= mois_debut,
                    Pointage.date_pointage <= mois_fin).all() if ids_jour else []
        pts_tous = pts_sal + pts_jour

        nb_presences  = sum(1 for p in pts_tous if p.present)
        nb_absences   = sum(1 for p in pts_tous if p.absent)
        taux_pres = round(nb_presences / (nb_presences + nb_absences) * 100
                         ) if (nb_presences + nb_absences) > 0 else 0

        # Heures totales
        heures_normales = sum(float(p.heures_normales or 8) for p in pts_tous if p.present)
        heures_sup = sum(
            float(p.heures_sup_10 or 0) + float(p.heures_sup_30 or 0) +
            float(p.heures_sup_40 or 0) + float(p.heures_sup_70 or 0) +
            float(p.heures_sup or 0)
            for p in pts_tous if p.present)

        # Détail par travailleur (pointage)
        detail_sal = []
        for sal in salaries_aff:
            pts_s = [p for p in pts_sal if p.salarie_id == sal.id]
            nb_p  = sum(1 for p in pts_s if p.present)
            nb_a  = sum(1 for p in pts_s if p.absent)
            h_n   = sum(float(p.heures_normales or 8) for p in pts_s if p.present)
            h_s   = sum(float(p.heures_sup_10 or 0)+float(p.heures_sup_30 or 0)+
                        float(p.heures_sup_40 or 0)+float(p.heures_sup_70 or 0)
                        for p in pts_s if p.present)
            detail_sal.append({
                "nom": sal.nom_complet, "matricule": sal.matricule,
                "emploi": sal.emploi or "—", "type": "MENSUEL",
                "nb_presences": nb_p, "nb_absences": nb_a,
                "heures_normales": round(h_n, 1), "heures_sup": round(h_s, 1),
                "taux": round(nb_p/(nb_p+nb_a)*100) if (nb_p+nb_a) > 0 else 0,
            })

        detail_jour = []
        for jour in journaliers_aff:
            pts_j = [p for p in pts_jour if p.journalier_id == jour.id]
            nb_p  = sum(1 for p in pts_j if p.present)
            nb_a  = sum(1 for p in pts_j if p.absent)
            h_n   = sum(float(p.heures_normales or 8) for p in pts_j if p.present)
            h_s   = sum(float(p.heures_sup or 0) for p in pts_j if p.present)
            detail_jour.append({
                "nom": jour.nom_complet, "profession": jour.profession or "—",
                "taux_horaire": float(jour.taux_horaire or 0), "type": "JOURNALIER",
                "nb_presences": nb_p, "nb_absences": nb_a,
                "heures_normales": round(h_n, 1), "heures_sup": round(h_s, 1),
                "taux": round(nb_p/(nb_p+nb_a)*100) if (nb_p+nb_a) > 0 else 0,
            })

        # ── Paie journaliers ─────────────────────────────────────────────────
        feuilles = FeuillePaieJournalier.query.filter_by(tenant_id=t.id)            .filter(FeuillePaieJournalier.journalier_id.in_(ids_jour),
                    FeuillePaieJournalier.date_debut >= mois_debut,
                    FeuillePaieJournalier.date_fin   <= mois_fin).all() if ids_jour else []
        masse_jour_brut   = sum(float(f.montant_brut or 0) for f in feuilles)
        feuilles_payees   = sum(1 for f in feuilles if f.statut == "PAYÉ")
        feuilles_attente  = sum(1 for f in feuilles if f.statut == "EN_ATTENTE")

        detail_feuilles = [{
            "nom":         f.journalier.nom_complet,
            "profession":  f.journalier.profession or "—",
            "date_debut":  f.date_debut.strftime("%d/%m/%Y") if f.date_debut else "",
            "date_fin":    f.date_fin.strftime("%d/%m/%Y")   if f.date_fin   else "",
            "nb_jours":    f.nb_jours,
            "heures":      round(float(f.total_heures or 0), 1),
            "taux":        round(float(f.taux_horaire or 0)),
            "montant":     round(float(f.montant_brut or 0)),
            "statut":      f.statut,
        } for f in feuilles]

        # ── Bulletins salariés ────────────────────────────────────────────────
        periode = PeriodePaie.query.filter_by(
            tenant_id=t.id, annee=annee, mois=mois).first()
        bulletins_site = []
        masse_mensuelle = {}
        detail_bulletins = []
        if periode and ids_sal:
            bulletins_site = BulletinPaie.query.filter_by(
                tenant_id=t.id, periode_id=periode.id
            ).filter(BulletinPaie.salarie_id.in_(ids_sal)).all()
            masse_mensuelle = calculer_masse_salariale(bulletins_site)
            detail_bulletins = [{
                "nom":        b.salarie.nom_complet,
                "matricule":  b.salarie.matricule,
                "emploi":     b.salarie.emploi or "—",
                "brut":       round(float(b.salaire_brut  or 0)),
                "net":        round(float(b.net_a_payer   or 0)),
                "cnss":       round(float(b.cnss_salarie  or 0)),
                "irpp":       round(float(b.irpp          or 0)),
                "statut":     b.statut,
            } for b in sorted(bulletins_site, key=lambda x: x.salarie.nom)]

        return {
            "site": s,
            "effectif_sal":   len(salaries_aff),
            "effectif_jour":  len(journaliers_aff),
            "effectif_total": len(salaries_aff) + len(journaliers_aff),
            # Pointage
            "nb_presences": nb_presences, "nb_absences": nb_absences,
            "taux_presence": taux_pres,
            "heures_normales": round(heures_normales, 1),
            "heures_sup": round(heures_sup, 1),
            "detail_sal": detail_sal,
            "detail_jour": detail_jour,
            # Paie journaliers
            "feuilles": detail_feuilles,
            "masse_jour_brut": round(masse_jour_brut),
            "feuilles_payees": feuilles_payees,
            "feuilles_attente": feuilles_attente,
            # Bulletins mensuels
            "bulletins": detail_bulletins,
            "masse_mensuelle": masse_mensuelle,
        }

    # Construire rapport(s)
    if site_sel:
        rapports = [_build_rapport_site(site_sel)]
    else:
        rapports = [_build_rapport_site(s) for s in sites_list]

    # Totaux globaux
    totaux = {
        "effectif": sum(r["effectif_total"] for r in rapports),
        "presences": sum(r["nb_presences"]   for r in rapports),
        "absences":  sum(r["nb_absences"]    for r in rapports),
        "h_normales": round(sum(r["heures_normales"] for r in rapports), 1),
        "h_sup":      round(sum(r["heures_sup"]     for r in rapports), 1),
        "masse_jour": sum(r["masse_jour_brut"]  for r in rapports),
        "masse_men":  sum(r["masse_mensuelle"].get("total_net", 0) for r in rapports),
    }

    return render_template("tenant/rapport_mensuel_site.html",
        tenant=t, rapports=rapports, totaux=totaux,
        sites=sites_list, site_sel=site_sel,
        mois=mois, annee=annee,
        mois_nom=MOIS_NOMS_LONG[mois],
        MOIS_NOMS=MOIS_NOMS_LONG,
        now=datetime.now())


@bp.route("/rapports/mensuel-site/export")
@login_required
def rapport_mensuel_site_export():
    """Export Excel du rapport mensuel par site."""
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    import io, calendar

    if current_user.is_super_admin: return redirect(url_for("admin.admin_dashboard"))
    t = get_tenant()
    if not t: return redirect(url_for("auth.login"))

    MOIS_NOMS_LONG = ["","Janvier","Février","Mars","Avril","Mai","Juin",
                      "Juillet","Août","Septembre","Octobre","Novembre","Décembre"]

    mois   = request.args.get("mois",  type=int, default=datetime.now().month)
    annee  = request.args.get("annee", type=int, default=datetime.now().year)
    site_id= request.args.get("site_id", type=int)
    mois_nom = MOIS_NOMS_LONG[mois]
    _, nb_jours_mois = calendar.monthrange(annee, mois)
    mois_debut = date(annee, mois, 1)
    mois_fin   = date(annee, mois, nb_jours_mois)

    # Reconstruire le rapport (même logique que la route GET)
    # On rappelle simplement la route interne via redirect vers export dédié
    # Pour éviter la duplication, on re-calcule ici directement

    sites_list = Site.query.filter_by(tenant_id=t.id, actif=True).order_by(Site.nom).all()
    site_sel   = Site.query.filter_by(id=site_id, tenant_id=t.id).first() if site_id else None
    sites_a_traiter = [site_sel] if site_sel else sites_list

    aff_sal  = {}
    aff_jour = {}
    for a in AffectationSite.query.filter_by(tenant_id=t.id).all():
        if a.actif or (a.date_fin and a.date_fin >= mois_debut):
            if a.salarie_id:    aff_sal[a.salarie_id]    = a.site_id
            if a.journalier_id: aff_jour[a.journalier_id] = a.site_id

    # Styles
    def hdr(ws, row, cols, texts, fill_color="1a2332"):
        for i, txt in enumerate(texts, 1):
            c = ws.cell(row, i, txt)
            c.font      = Font(bold=True, color="FFFFFF", size=9)
            c.fill      = PatternFill("solid", fgColor=fill_color)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border    = Border(**{s: Side(style="thin", color="D1D5DB")
                                    for s in ["left","right","top","bottom"]})
        ws.row_dimensions[row].height = 20

    def titre_section(ws, row, txt, color="374151", ncols=10):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
        c = ws.cell(row, 1, txt)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor=color)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 18

    MONEY = "#,##0"
    thin  = {s: Side(style="thin", color="E5E7EB") for s in ["left","right","top","bottom"]}
    EVEN  = PatternFill("solid", fgColor="F8FAFC")

    wb = Workbook()

    # ══════════════════════════════════════════════════════════════════════════
    # ONGLET 1 — SYNTHÈSE GLOBALE
    # ══════════════════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Synthèse"
    ws.freeze_panes = "A4"

    # Titre principal
    titre_doc = f"RAPPORT MENSUEL — {mois_nom.upper()} {annee} — {t.denomination}"
    if site_sel: titre_doc += f" — {site_sel.nom}"
    ws.merge_cells("A1:K1")
    ws["A1"] = titre_doc
    ws["A1"].font = Font(bold=True, size=13, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="1a2332")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24
    ws.append([f"Édité le {datetime.now().strftime('%d/%m/%Y à %H:%M')}"])
    ws.append([])

    hdr(ws, 3, 11,
        ["Site","Effectif","Mensuels","Journaliers","Présences","Absences",
         "Taux prés.","H.normales","H.sup","Masse journaliers (FCFA)","Net mensuels (FCFA)"])

    grand_tot = {"eff":0,"pres":0,"abs":0,"hn":0,"hs":0,"mj":0,"mn":0}
    for row_i, s in enumerate(sites_a_traiter, 4):
        ids_sal  = [k for k,v in aff_sal.items()  if v == s.id]
        ids_jour = [k for k,v in aff_jour.items() if v == s.id]
        pts = Pointage.query.filter_by(tenant_id=t.id)            .filter(Pointage.date_pointage >= mois_debut,
                    Pointage.date_pointage <= mois_fin)            .filter(db.or_(
                Pointage.salarie_id.in_(ids_sal)    if ids_sal  else db.false(),
                Pointage.journalier_id.in_(ids_jour) if ids_jour else db.false()
            )).all()
        nb_p  = sum(1 for p in pts if p.present)
        nb_a  = sum(1 for p in pts if p.absent)
        taux  = round(nb_p/(nb_p+nb_a)*100) if (nb_p+nb_a) > 0 else 0
        hn    = round(sum(float(p.heures_normales or 8) for p in pts if p.present), 1)
        hs    = round(sum(float(p.heures_sup_10 or 0)+float(p.heures_sup_30 or 0)+
                          float(p.heures_sup_40 or 0)+float(p.heures_sup_70 or 0)+
                          float(p.heures_sup or 0) for p in pts if p.present), 1)
        feuilles = FeuillePaieJournalier.query.filter_by(tenant_id=t.id)            .filter(FeuillePaieJournalier.journalier_id.in_(ids_jour),
                    FeuillePaieJournalier.date_debut >= mois_debut,
                    FeuillePaieJournalier.date_fin   <= mois_fin).all() if ids_jour else []
        mj = round(sum(float(f.montant_brut or 0) for f in feuilles))
        per= PeriodePaie.query.filter_by(tenant_id=t.id, annee=annee, mois=mois).first()
        mn = 0
        if per and ids_sal:
            buls = BulletinPaie.query.filter_by(tenant_id=t.id, periode_id=per.id)                .filter(BulletinPaie.salarie_id.in_(ids_sal)).all()
            mn = round(sum(float(b.net_a_payer or 0) for b in buls))
        eff = len(ids_sal) + len(ids_jour)
        row_data = [csv_safe(s.nom), eff, len(ids_sal), len(ids_jour),
                    nb_p, nb_a, f"{taux}%", hn, hs, mj, mn]
        ws.append(row_data)
        for ci, v in enumerate(row_data, 1):
            c = ws.cell(row_i, ci)
            c.border = Border(**thin)
            if row_i % 2 == 0: c.fill = EVEN
            if ci in (10, 11): c.number_format = MONEY; c.alignment = Alignment(horizontal="right")
            if ci in (5,6,7,8,9): c.alignment = Alignment(horizontal="center")
        grand_tot["eff"]+=eff; grand_tot["pres"]+=nb_p; grand_tot["abs"]+=nb_a
        grand_tot["hn"]+=hn; grand_tot["hs"]+=hs; grand_tot["mj"]+=mj; grand_tot["mn"]+=mn

    # Total
    tr = ws.max_row + 1
    ws.append(["TOTAL", grand_tot["eff"], "", "", grand_tot["pres"], grand_tot["abs"],
                "", round(grand_tot["hn"],1), round(grand_tot["hs"],1),
                grand_tot["mj"], grand_tot["mn"]])
    for ci in range(1, 12):
        c = ws.cell(tr, ci)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1a2332")
        if ci in (10, 11): c.number_format = MONEY; c.alignment = Alignment(horizontal="right")

    for i, w in enumerate([24,10,10,12,10,10,10,12,10,22,22], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ══════════════════════════════════════════════════════════════════════════
    # ONGLETS PAR SITE
    # ══════════════════════════════════════════════════════════════════════════
    for s in sites_a_traiter:
        ws_s = wb.create_sheet(s.nom[:28])
        ws_s.freeze_panes = "A4"
        ids_sal  = [k for k,v in aff_sal.items()  if v == s.id]
        ids_jour = [k for k,v in aff_jour.items() if v == s.id]

        # Titre
        ws_s.merge_cells("A1:I1")
        ws_s["A1"] = f"{s.nom} — {mois_nom} {annee} — {t.denomination}"
        ws_s["A1"].font = Font(bold=True, size=12, color="FFFFFF")
        ws_s["A1"].fill = PatternFill("solid", fgColor="1a2332")
        ws_s["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws_s.row_dimensions[1].height = 20
        ws_s.append([])
        row_cur = 2

        # ── Section pointage ──────────────────────────────────────────────────
        row_cur += 1
        titre_section(ws_s, row_cur, f"📅 POINTAGE — {mois_nom} {annee}", "065f46", 9)
        row_cur += 1
        hdr(ws_s, row_cur, 9,
            ["Nom","Type","Emploi/Profession","Présences","Absences","Taux %","H.norm.","H.sup","Total h."],
            "065f46")
        row_cur += 1

        pts_sal  = Pointage.query.filter_by(tenant_id=t.id)            .filter(Pointage.salarie_id.in_(ids_sal),
                    Pointage.date_pointage >= mois_debut,
                    Pointage.date_pointage <= mois_fin).all() if ids_sal else []
        pts_jour = Pointage.query.filter_by(tenant_id=t.id)            .filter(Pointage.journalier_id.in_(ids_jour),
                    Pointage.date_pointage >= mois_debut,
                    Pointage.date_pointage <= mois_fin).all() if ids_jour else []

        salaries_aff    = Salarie.query.filter(Salarie.id.in_(ids_sal)).order_by(Salarie.nom).all() if ids_sal else []
        journaliers_aff = Journalier.query.filter(Journalier.id.in_(ids_jour)).order_by(Journalier.nom).all() if ids_jour else []

        for trv_list, pts_list, typ in [
            (salaries_aff, pts_sal, "Mensuel"),
            (journaliers_aff, pts_jour, "Journalier")
        ]:
            for trv in trv_list:
                if typ == "Mensuel":
                    pts_t = [p for p in pts_list if p.salarie_id == trv.id]
                    emploi = trv.emploi or "—"
                else:
                    pts_t = [p for p in pts_list if p.journalier_id == trv.id]
                    emploi = trv.profession or "—"
                nb_p = sum(1 for p in pts_t if p.present)
                nb_a = sum(1 for p in pts_t if p.absent)
                hn   = round(sum(float(p.heures_normales or 8) for p in pts_t if p.present), 1)
                hs   = round(sum(float(p.heures_sup_10 or 0)+float(p.heures_sup_30 or 0)+
                                 float(p.heures_sup_40 or 0)+float(p.heures_sup_70 or 0)+
                                 float(p.heures_sup or 0) for p in pts_t if p.present), 1)
                taux = round(nb_p/(nb_p+nb_a)*100) if (nb_p+nb_a) > 0 else 0
                row_d = [trv.nom_complet, typ, emploi, nb_p, nb_a, f"{taux}%", hn, hs, round(hn+hs,1)]
                ws_s.append(row_d)
                for ci, v in enumerate(row_d, 1):
                    c = ws_s.cell(row_cur, ci)
                    c.border = Border(**thin)
                    if row_cur % 2 == 0: c.fill = EVEN
                    if ci in (4,5,6,7,8,9): c.alignment = Alignment(horizontal="center")
                row_cur += 1

        ws_s.append([])
        row_cur += 1

        # ── Section paie journaliers ──────────────────────────────────────────
        if ids_jour:
            titre_section(ws_s, row_cur, f"🦺 PAIE JOURNALIERS — {mois_nom} {annee}", "92400e", 9)
            row_cur += 1
            hdr(ws_s, row_cur, 9,
                ["Journalier","Profession","Période du","au","Nb jours","Heures","Taux/h","Montant (FCFA)","Statut"],
                "92400e")
            row_cur += 1
            feuilles = FeuillePaieJournalier.query.filter_by(tenant_id=t.id)                .filter(FeuillePaieJournalier.journalier_id.in_(ids_jour),
                        FeuillePaieJournalier.date_debut >= mois_debut,
                        FeuillePaieJournalier.date_fin   <= mois_fin).all()
            total_feuilles = 0
            for f in feuilles:
                m = round(float(f.montant_brut or 0)); total_feuilles += m
                row_d = [f.journalier.nom_complet, f.journalier.profession or "—",
                         f.date_debut.strftime("%d/%m/%Y") if f.date_debut else "",
                         f.date_fin.strftime("%d/%m/%Y")   if f.date_fin   else "",
                         f.nb_jours, round(float(f.total_heures or 0),1),
                         round(float(f.taux_horaire or 0)), m, f.statut]
                ws_s.append(row_d)
                for ci, v in enumerate(row_d, 1):
                    c = ws_s.cell(row_cur, ci)
                    c.border = Border(**thin)
                    if row_cur % 2 == 0: c.fill = EVEN
                    if ci == 8: c.number_format = MONEY; c.alignment = Alignment(horizontal="right")
                    if ci in (5,6,7): c.alignment = Alignment(horizontal="center")
                row_cur += 1
            # Total
            ws_s.append(["TOTAL JOURNALIERS", "", "", "", "", "", "", total_feuilles, ""])
            for ci in range(1, 10):
                c = ws_s.cell(row_cur, ci)
                c.font = Font(bold=True, color="FFFFFF")
                c.fill = PatternFill("solid", fgColor="92400e")
                if ci == 8: c.number_format = MONEY; c.alignment = Alignment(horizontal="right")
            row_cur += 2; ws_s.append([])

        # ── Section bulletins mensuels ────────────────────────────────────────
        if ids_sal:
            per = PeriodePaie.query.filter_by(tenant_id=t.id, annee=annee, mois=mois).first()
            if per:
                buls = BulletinPaie.query.filter_by(tenant_id=t.id, periode_id=per.id)                    .filter(BulletinPaie.salarie_id.in_(ids_sal)).all()
                if buls:
                    titre_section(ws_s, row_cur,
                        f"📄 BULLETINS DE PAIE — {mois_nom} {annee}", "1e40af", 9)
                    row_cur += 1
                    hdr(ws_s, row_cur, 9,
                        ["Salarié","Matricule","Emploi","Brut (FCFA)","CNSS sal.",
                         "TCS","IRPP","Net à payer (FCFA)","Statut"], "1e40af")
                    row_cur += 1
                    total_brut = total_net = 0
                    for b in sorted(buls, key=lambda x: x.salarie.nom):
                        brut = round(float(b.salaire_brut or 0))
                        net  = round(float(b.net_a_payer  or 0))
                        total_brut += brut; total_net += net
                        row_d = [b.salarie.nom_complet, b.salarie.matricule,
                                 b.salarie.emploi or "—", brut,
                                 round(float(b.cnss_salarie or 0)),
                                 round(float(b.tcs  or 0)),
                                 round(float(b.irpp or 0)), net, b.statut]
                        ws_s.append(row_d)
                        for ci, v in enumerate(row_d, 1):
                            c = ws_s.cell(row_cur, ci)
                            c.border = Border(**thin)
                            if row_cur % 2 == 0: c.fill = EVEN
                            if ci in (4,5,6,7,8): c.number_format = MONEY; c.alignment = Alignment(horizontal="right")
                        row_cur += 1
                    ws_s.append(["TOTAL SALARIÉS","","",total_brut,"","","",total_net,""])
                    for ci in range(1, 10):
                        c = ws_s.cell(row_cur, ci)
                        c.font = Font(bold=True, color="FFFFFF")
                        c.fill = PatternFill("solid", fgColor="1e40af")
                        if ci in (4,8): c.number_format = MONEY; c.alignment = Alignment(horizontal="right")

        for i, w in enumerate([28,12,20,10,10,10,10,20,12], 1):
            ws_s.column_dimensions[get_column_letter(i)].width = w

    # Export
    out = io.BytesIO(); wb.save(out); out.seek(0)
    fname_parts = [f"Rapport_{mois_nom}_{annee}"]
    if site_sel: fname_parts.append(site_sel.nom.replace(" ","_"))
    fname_parts.append(datetime.now().strftime("%Y%m%d"))
    return send_file(out,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True, download_name="_".join(fname_parts)+".xlsx")





def _gen_excel_cnss(tenant, trim_label, annee, mois_labels,
                    sal_data, total_base_cnss, total_base_cnamgs,
                    tot_cnss_m):
    """
    Génère la feuille CNSS conforme au formulaire officiel gabonais.
    sal_data : liste de dicts {nom_complet, matricule, numero_cnss,
               date_embauche, m1_base_cnss, m2_base_cnss, m3_base_cnss}
    N_HRS = 8 (constante légale)
    """
    import openpyxl, io
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    BD2=Border(left=Side(style="thin",color="E5E7EB"),right=Side(style="thin",color="E5E7EB"),
               top=Side(style="thin",color="E5E7EB"),bottom=Side(style="thin",color="E5E7EB"))
    CTR2=Alignment(horizontal="center",vertical="center",wrap_text=True)
    LFT2=Alignment(horizontal="left",  vertical="center",wrap_text=True)
    RGT2=Alignment(horizontal="right", vertical="center")
    HF=PatternFill("solid",fgColor="1a2332"); HN=Font(bold=True,color="FFFFFF",size=9)
    TF=PatternFill("solid",fgColor="D6EAF8"); GF=PatternFill("solid",fgColor="E8F4FD")
    AF=PatternFill("solid",fgColor="EBF5FB")

    def Cx(ws,r,c,val,font=None,fill=None,align=None,fmt=None,span=None):
        if span and span>1:
            ws.merge_cells(start_row=r,start_column=c,end_row=r,end_column=c+span-1)
        cell=ws.cell(r,c)
        if val  is not None: cell.value=val
        if font is not None: cell.font=font
        if fill is not None: cell.fill=fill
        if align is not None: cell.alignment=align
        if fmt  is not None: cell.number_format=fmt
        cell.border=BD2; return cell

    wb=openpyxl.Workbook(); ws=wb.active; ws.title="CNSS"
    for i,w in enumerate([5,14,12,28,8,12,12,16,14,4,16,14,4,16,14,4,10,10],1):
        ws.column_dimensions[get_column_letter(i)].width=w

    def rh(r,h): ws.row_dimensions[r].height=h

    # ── En-tête employeur ─────────────────────────────────────────────────────
    rh(10,18)
    Cx(ws,10,1,"Matricule employeur",Font(bold=True,size=9),align=LFT2)
    Cx(ws,10,2,tenant.numero_cnss or "—",Font(size=9),align=LFT2)
    Cx(ws,10,5,"Période",Font(bold=True,size=9),align=CTR2)
    Cx(ws,10,6,trim_label,Font(bold=True,size=11),align=CTR2)
    Cx(ws,10,7,"Année",Font(bold=True,size=9),align=CTR2)
    Cx(ws,10,8,annee,Font(bold=True,size=11),align=CTR2)
    Cx(ws,10,13,"CACHET ET SIGNATURE",HN,HF,CTR2,span=4)

    rh(12,20)
    Cx(ws,12,1,"Nom ou Raison Sociale",Font(bold=True,size=9),align=LFT2)
    Cx(ws,12,2,tenant.denomination,Font(bold=True,size=9),align=LFT2,span=3)

    rh(15,16); Cx(ws,15,1,"B.P :",Font(size=9),align=LFT2)
    Cx(ws,15,3,f"VILLE : {getattr(tenant,'ville','Libreville')}",Font(size=9),align=LFT2,span=2)
    rh(17,16); Cx(ws,17,1,"TEL :",Font(size=9),align=LFT2)
    Cx(ws,17,2,getattr(tenant,"telephone",""),Font(size=9),align=LFT2)
    Cx(ws,17,9,"Effectif total",Font(bold=True,size=9),align=CTR2)
    Cx(ws,17,10,len(sal_data),Font(bold=True,size=11),TF,CTR2)
    rh(19,16); Cx(ws,19,1,"Email :",Font(size=9),align=LFT2)

    # ── Résumé cotisations ────────────────────────────────────────────────────
    rh(20,32)
    Cx(ws,20,2,"Rémunération totale plafonnée CNSS",Font(bold=True,size=8),GF,CTR2,span=3)
    Cx(ws,20,5,"Montant déduction Alloc. Familiales",Font(size=8),GF,CTR2,span=3)
    Cx(ws,20,9,"Rémunération totale plafonnée CNAMGS",Font(bold=True,size=8),GF,CTR2,span=3)
    Cx(ws,20,13,"DATE DE RECEPTION",HN,HF,CTR2,span=4)
    rh(21,18)
    Cx(ws,21,2,total_base_cnss,Font(bold=True,size=10),TF,RGT2,"#,##0",span=3)
    Cx(ws,21,5,0,Font(size=9),TF,RGT2,"#,##0",span=3)
    Cx(ws,21,9,total_base_cnamgs,Font(bold=True,size=10),TF,RGT2,"#,##0",span=3)
    rh(22,18)
    Cx(ws,22,2,"Cotisations brutes dues CNSS",Font(bold=True,size=8),align=CTR2,span=3)
    Cx(ws,22,5,"Cotisations nettes dues CNSS",Font(bold=True,size=8),align=CTR2,span=3)
    Cx(ws,22,9,"Cotisations nettes dues CNAMGS",Font(bold=True,size=8),align=CTR2,span=3)
    rh(23,18)
    cot_cnss   = round(total_base_cnss   * 0.23)
    cot_cnamgs = round(total_base_cnamgs * 0.061)
    Cx(ws,23,2,cot_cnss,  Font(bold=True,size=10),TF,RGT2,"#,##0",span=3)
    Cx(ws,23,5,cot_cnss,  Font(bold=True,size=10),TF,RGT2,"#,##0",span=3)
    Cx(ws,23,9,cot_cnamgs,Font(bold=True,size=10),TF,RGT2,"#,##0",span=3)

    # ── En-têtes mois ─────────────────────────────────────────────────────────
    rh(25,14)
    for col,lbl in [(8,mois_labels[0]),(11,mois_labels[1]),(14,mois_labels[2])]:
        Cx(ws,25,col,lbl,HN,HF,CTR2)
    rh(26,36)
    for col,lbl in [(1,"N°"),(2,"N°CNSS /\nN°CNAMGS"),(3,"N° Paie"),
                    (4,"NOM ET PRENOM"),(5,"Taux CNSS"),(6,"EMBAUCHE"),(7,"CESSATION"),
                    (8,"SALAIRE\nPLAFONNE"),(9,"SALAIRE\nDEPLAFONNE"),(10,"Nbre\nHrs"),
                    (11,"SALAIRE\nPLAFONNE"),(12,"SALAIRE\nDEPLAFONNE"),(13,"Nbre\nHrs"),
                    (14,"SALAIRE\nPLAFONNE"),(15,"SALAIRE\nDEPLAFONNE"),(16,"Nbre\nHrs")]:
        Cx(ws,26,col,lbl,HN,HF,CTR2)

    # ── 2 lignes par employé ──────────────────────────────────────────────────
    N_HRS_CNSS = 8  # toujours 8
    dr = 27
    for i, sal in enumerate(sal_data, 1):
        r1 = dr + (i-1)*2
        r2 = dr + (i-1)*2 + 1
        bg = AF if i%2==0 else None
        rh(r1,16); rh(r2,16)

        # Ligne impaire : numéro + taux + données BASE CNSS + heures
        Cx(ws,r1,1,i,Font(size=9),bg,CTR2)
        Cx(ws,r1,5,23,Font(size=9),bg,CTR2)  # Taux CNSS = 23%
        emb = sal.get("date_embauche","")
        Cx(ws,r1,6,emb,Font(size=8),bg,CTR2)
        Cx(ws,r1,7,"",None,bg,CTR2)
        # Mois 1, 2, 3 — BASE CNSS (plafonnée), N_HRS=8
        Cx(ws,r1,8, sal.get("m1_base_cnss",0), Font(size=9),bg,RGT2,"#,##0")
        Cx(ws,r1,9, sal.get("m1_base_cnss",0), Font(size=9),bg,RGT2,"#,##0")  # déplafonné = même valeur
        Cx(ws,r1,10,N_HRS_CNSS,Font(size=9),bg,CTR2)
        Cx(ws,r1,11,sal.get("m2_base_cnss",0), Font(size=9),bg,RGT2,"#,##0")
        Cx(ws,r1,12,sal.get("m2_base_cnss",0), Font(size=9),bg,RGT2,"#,##0")
        Cx(ws,r1,13,N_HRS_CNSS,Font(size=9),bg,CTR2)
        Cx(ws,r1,14,sal.get("m3_base_cnss",0), Font(size=9),bg,RGT2,"#,##0")
        Cx(ws,r1,15,sal.get("m3_base_cnss",0), Font(size=9),bg,RGT2,"#,##0")
        Cx(ws,r1,16,N_HRS_CNSS,Font(size=9),bg,CTR2)

        # Ligne paire : N°CNSS + Matricule + NOM
        Cx(ws,r2,2,sal.get("numero_cnss",""),Font(size=9),bg,CTR2)
        Cx(ws,r2,3,sal.get("matricule",""),  Font(size=9),bg,CTR2)
        Cx(ws,r2,4,sal.get("nom_complet",""),Font(bold=True,size=9),bg,LFT2)
        for c in [1,5,6,7,8,9,10,11,12,13,14,15,16]:
            Cx(ws,r2,c,"",None,bg,None)

    # ── Sous-total ────────────────────────────────────────────────────────────
    rst = dr + len(sal_data)*2; rh(rst,18)
    Cx(ws,rst,1,"SOUS TOTAL À REPORTER PAGE SUIVANTE",Font(bold=True,size=9),TF,LFT2,span=7)
    for col,val in [(8,tot_cnss_m[0]),(11,tot_cnss_m[1]),(14,tot_cnss_m[2])]:
        Cx(ws,rst,col,val,Font(bold=True),TF,RGT2,"#,##0")
        for cc in [col+1,col+2]: Cx(ws,rst,cc,"",None,TF,None)

    # ── RECAP ─────────────────────────────────────────────────────────────────
    rr = rst+2; rh(rr,18)
    Cx(ws,rr,1,"RECAP",HN,HF,CTR2,span=2)
    Cx(ws,rr,3,"TAUX",HN,HF,CTR2)
    Cx(ws,rr,4,23,Font(bold=True,size=10),TF,CTR2)
    Cx(ws,rr,5,"MASSE SALARIALE PLAFONNEE CNSS",Font(bold=True,size=8),TF,LFT2,span=3)
    for col,val in [(8,tot_cnss_m[0]),(11,tot_cnss_m[1]),(14,tot_cnss_m[2])]:
        Cx(ws,rr,col,val,Font(bold=True),TF,RGT2,"#,##0")
        for cc in [col+1,col+2]: Cx(ws,rr,cc,"",None,TF,None)

    rcot = rr+1; rh(rcot,24)
    Cx(ws,rcot,1,"COTISATION GLOBALE DUE (CNSS)",Font(bold=True,size=11),HF,LFT2,span=13)
    Cx(ws,rcot,14,cot_cnss,Font(bold=True,size=13),TF,RGT2,"#,##0",span=3)

    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.getvalue()


def _gen_excel_cnamgs(tenant, trim_label, annee, mois_labels,
                      sal_data, total_base_cnamgs, tot_cnamgs_m):
    """
    Génère la feuille CNAMGS conforme au formulaire officiel gabonais.
    sal_data : liste de dicts {nom_complet, matricule, numero_cnamgs,
               date_embauche, m1_base_cnamgs, m2_base_cnamgs, m3_base_cnamgs}
    N_HRS = 8 (constante légale)
    """
    import openpyxl, io
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    BD2=Border(left=Side(style="thin",color="E5E7EB"),right=Side(style="thin",color="E5E7EB"),
               top=Side(style="thin",color="E5E7EB"),bottom=Side(style="thin",color="E5E7EB"))
    CTR2=Alignment(horizontal="center",vertical="center",wrap_text=True)
    LFT2=Alignment(horizontal="left",  vertical="center",wrap_text=True)
    RGT2=Alignment(horizontal="right", vertical="center")
    HF=PatternFill("solid",fgColor="1a2332"); HN=Font(bold=True,color="FFFFFF",size=9)
    T2=PatternFill("solid",fgColor="D5F5E3"); G2=PatternFill("solid",fgColor="E8F8F5")
    YF=PatternFill("solid",fgColor="FEF9E7")

    def Cx(ws,r,c,val,font=None,fill=None,align=None,fmt=None,span=None):
        if span and span>1:
            ws.merge_cells(start_row=r,start_column=c,end_row=r,end_column=c+span-1)
        cell=ws.cell(r,c)
        if val  is not None: cell.value=val
        if font is not None: cell.font=font
        if fill is not None: cell.fill=fill
        if align is not None: cell.alignment=align
        if fmt  is not None: cell.number_format=fmt
        cell.border=BD2; return cell

    wb=openpyxl.Workbook(); ws=wb.active; ws.title="CNAMGS"
    for i,w in enumerate([5,18,28,14,14,18,10,18,10,18,10],1):
        ws.column_dimensions[get_column_letter(i)].width=w

    def rh(r,h): ws.row_dimensions[r].height=h

    # ── Titre ─────────────────────────────────────────────────────────────────
    rh(5,22)
    Cx(ws,5,4,"DECLARATION TRIMESTRIELLE DE SALAIRES",
       Font(bold=True,size=13,color="1a2332"),None,CTR2,span=6)

    # ── En-tête ───────────────────────────────────────────────────────────────
    rh(8,16)
    Cx(ws,8,4,"Période :",Font(bold=True,size=9),align=CTR2)
    Cx(ws,8,5,trim_label,Font(bold=True,size=11),align=CTR2)
    Cx(ws,8,6,annee,Font(bold=True,size=11),align=CTR2)
    rh(9,16)
    Cx(ws,9,1,"Matricule employeur CNAMGS",Font(bold=True,size=9),align=LFT2)
    Cx(ws,9,2,getattr(tenant,"numero_cnamgs","—"),Font(size=9),align=LFT2,span=2)
    rh(11,20)
    Cx(ws,11,1,"Nom ou Raison Sociale",Font(bold=True,size=9),align=LFT2)
    Cx(ws,11,2,tenant.denomination,Font(bold=True,size=9),align=LFT2,span=3)
    Cx(ws,11,5,trim_label,Font(bold=True,size=11,color="1a2332"),align=CTR2)
    Cx(ws,11,6,annee,Font(bold=True,size=11),align=CTR2)
    Cx(ws,11,8,"CACHET ET SIGNATURE",HN,HF,CTR2,span=4)

    # Taux (sur lignes séparées pour éviter conflits merge)
    rh(13,16)
    Cx(ws,13,4,"Taux de cotisation",Font(bold=True,size=9),T2,LFT2,span=2)
    rh(14,16)
    Cx(ws,14,1,"B.P :",Font(size=9),align=LFT2)
    Cx(ws,14,2,getattr(tenant,"adresse",""),Font(size=9),align=LFT2,span=2)
    Cx(ws,14,4,"Employeur",Font(bold=True,size=9),T2,CTR2)
    Cx(ws,14,5,0.041,Font(bold=True,size=9),T2,RGT2,fmt="0.0%")
    rh(15,16)
    Cx(ws,15,1,"VILLE :",Font(size=9),align=LFT2)
    Cx(ws,15,2,getattr(tenant,"ville","Libreville"),Font(size=9),align=LFT2)
    Cx(ws,15,4,"Travailleur",Font(bold=True,size=9),T2,CTR2)
    Cx(ws,15,5,0.02,Font(bold=True,size=9),T2,RGT2,fmt="0.0%")
    rh(16,16)
    Cx(ws,16,1,"TEL :",Font(size=9),align=LFT2)
    Cx(ws,16,2,getattr(tenant,"telephone",""),Font(size=9),align=LFT2)
    rh(18,16)
    Cx(ws,18,4,"Plafond mensuel CNAMGS",Font(bold=True,size=9),T2,LFT2)
    Cx(ws,18,5,2500000,Font(bold=True,size=9),T2,RGT2,fmt="#,##0")

    # ── Cotisations nettes dues ───────────────────────────────────────────────
    rh(19,16)
    Cx(ws,19,2,"Cotisations nettes dues CNAMGS",Font(bold=True,size=9),align=LFT2,span=4)
    Cx(ws,19,8,"DATE DE RECEPTION",HN,HF,CTR2,span=4)
    rh(20,20)
    Cx(ws,20,2,round(total_base_cnamgs*0.061),Font(bold=True,size=12,color="1e40af"),T2,RGT2,"#,##0",span=4)
    rh(21,16)
    Cx(ws,21,2,"Cotisations payées à la CNAMGS",Font(italic=True,size=9),align=LFT2,span=4)

    # ── RECAP ─────────────────────────────────────────────────────────────────
    rh(24,18)
    Cx(ws,24,1,"Recap.",HN,HF,CTR2)
    Cx(ws,24,2,"Effectif",HN,HF,CTR2)
    Cx(ws,24,3,len(sal_data),Font(bold=True,size=11),T2,CTR2)
    Cx(ws,24,4,"MASSE SALARIALE SOUMISE À COTISATION :",Font(bold=True,size=8),T2,LFT2,span=2)
    Cx(ws,24,6,total_base_cnamgs,Font(bold=True,size=10),T2,RGT2,"#,##0")
    Cx(ws,24,8,"COTISATIONS SOCIALES:",Font(bold=True,size=9),T2,LFT2)
    Cx(ws,24,10,round(total_base_cnamgs*0.061),Font(bold=True,size=10),T2,RGT2,"#,##0")
    rh(25,16)
    Cx(ws,25,8,"Part patronale (4.1%)",Font(size=9),G2,LFT2)
    Cx(ws,25,10,round(total_base_cnamgs*0.041),Font(bold=True,size=9),G2,RGT2,"#,##0")
    rh(26,16)
    Cx(ws,26,8,"Part salariale (2%)",Font(size=9),G2,LFT2)
    Cx(ws,26,10,round(total_base_cnamgs*0.02),Font(bold=True,size=9),G2,RGT2,"#,##0")
    rh(27,18)
    Cx(ws,27,1,"TOTAL À REPORTER PAGE SUIVANTE",HN,HF,LFT2,span=5)
    for col,val in [(6,tot_cnamgs_m[0]),(8,tot_cnamgs_m[1]),(10,tot_cnamgs_m[2])]:
        Cx(ws,27,col,val,Font(bold=True),T2,RGT2,"#,##0")

    # ── Labels mois ───────────────────────────────────────────────────────────
    rh(28,14)
    for col,lbl in [(6,mois_labels[0]),(8,mois_labels[1]),(10,mois_labels[2])]:
        Cx(ws,28,col,lbl,HN,HF,CTR2,span=2)

    # ── En-têtes colonnes employés ────────────────────────────────────────────
    rh(38,14)
    Cx(ws,38,4,"Date",HN,HF,CTR2,span=2)
    for col,lbl in [(6,mois_labels[0]),(8,mois_labels[1]),(10,mois_labels[2])]:
        Cx(ws,38,col,lbl,HN,HF,CTR2,span=2)
    rh(39,36)
    for col,lbl in [(1,"N°"),(2,"Matricule"),(3,"NOM ET PRENOM"),
                    (4,"EMBAUCHE"),(5,"CESSATION"),
                    (6,"Assiette soumise\nà cotisation"),(7,"Nbre\nHrs/Jrs"),
                    (8,"Assiette soumise\nà cotisation"),(9,"Nbre\nHrs/Jrs"),
                    (10,"Assiette soumise\nà cotisation"),(11,"Nbre\nHrs/Jrs")]:
        Cx(ws,39,col,lbl,HN,HF,CTR2)

    # ── 1 ligne par employé — BASE CNAMGS, N_HRS=8 ───────────────────────────
    N_HRS_CNAMGS = 8  # toujours 8
    for i, sal in enumerate(sal_data, 1):
        r = 39 + i; rh(r,18)
        bg = PatternFill("solid",fgColor="E8F8F5") if i%2==0 else None
        Cx(ws,r,1,i,Font(size=9),bg,CTR2)
        Cx(ws,r,2,sal.get("matricule",""),Font(size=9),bg,CTR2)
        Cx(ws,r,3,sal.get("nom_complet",""),Font(bold=True,size=9),bg,LFT2)
        Cx(ws,r,4,sal.get("date_embauche",""),Font(size=8),bg,CTR2)
        Cx(ws,r,5,"",None,bg,CTR2)
        # BASE CNAMGS + toujours 8h
        Cx(ws,r,6, sal.get("m1_base_cnamgs",0),Font(size=9),bg,RGT2,"#,##0")
        Cx(ws,r,7, N_HRS_CNAMGS,Font(size=9),bg,CTR2)
        Cx(ws,r,8, sal.get("m2_base_cnamgs",0),Font(size=9),bg,RGT2,"#,##0")
        Cx(ws,r,9, N_HRS_CNAMGS,Font(size=9),bg,CTR2)
        Cx(ws,r,10,sal.get("m3_base_cnamgs",0),Font(size=9),bg,RGT2,"#,##0")
        Cx(ws,r,11,N_HRS_CNAMGS,Font(size=9),bg,CTR2)

    # ── Note pénalités ────────────────────────────────────────────────────────
    r_note = 39 + len(sal_data) + 2; rh(r_note,60)
    Cx(ws,r_note,1,
       "Au-delà de la date limite, une pénalité est appliquée conformément à la loi :\n"
       "- 25% pour non dépôt de la DTS calculé sur le montant de la DTS du dernier trimestre déclaré ;\n"
       "- 2% pour non paiement des cotisations par mois de retard cumulable au prorata temporis.",
       Font(italic=True,size=8), YF,
       Alignment(horizontal="left",vertical="top",wrap_text=True), span=11)

    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.getvalue()


# ══════════════════════════════════════════════════════════════════════════════
# DÉCLARATIONS SOCIALES & FISCALES
# ══════════════════════════════════════════════════════════════════════════════

@bp.route("/declaration-cnss")
@login_required
def declaration_cnss():
    """Déclarations sociales : CNSS/CNAMGS (trimestrielles) + CFP/FNH/TCS/IRPP (mensuelles)."""
    if current_user.is_super_admin: return redirect(url_for("admin.admin_dashboard"))
    t = get_tenant()
    if not t: return redirect(url_for("auth.login"))

    periodes = PeriodePaie.query.filter_by(tenant_id=t.id)        .order_by(PeriodePaie.annee.desc(), PeriodePaie.mois.desc()).all()

    # Période sélectionnée
    pid     = request.args.get("periode_id", type=int)
    periode = PeriodePaie.query.filter_by(id=pid, tenant_id=t.id).first() if pid else               (periodes[0] if periodes else None)

    # Mode : mensuel (CFP/FNH/TCS/IRPP) ou trimestriel (CNSS/CNAMGS)
    mode    = request.args.get("mode", "mensuel")  # mensuel | trimestriel

    bulletins_mois = []
    bulletins_trim = []
    stats_mensuel  = {}
    stats_trim     = {}

    if periode:
        def s(buls, field): return round(sum(float(getattr(b, field) or 0) for b in buls), 2)

        # ── Bulletins du mois sélectionné (CFP, FNH, TCS, IRPP) ─────────────
        bulletins_mois = BulletinPaie.query.filter_by(
            tenant_id=t.id, periode_id=periode.id
        ).options(joinedload(BulletinPaie.salarie)).all()

        stats_mensuel = {
            "nb":           len(bulletins_mois),
            "total_brut":   s(bulletins_mois, "salaire_brut"),
            "total_cfp":    s(bulletins_mois, "cfp"),
            "total_fnh":    s(bulletins_mois, "fnh"),
            "total_tcs":    s(bulletins_mois, "tcs"),
            "total_irpp":   s(bulletins_mois, "irpp"),
        }
        stats_mensuel["total_mensuel"] = (stats_mensuel["total_cfp"]  +
                                          stats_mensuel["total_fnh"]  +
                                          stats_mensuel["total_tcs"]  +
                                          stats_mensuel["total_irpp"])

        # ── Bulletins du trimestre (CNSS/CNAMGS) ─────────────────────────────
        # Trimestre : T1=Jan-Mar, T2=Avr-Jun, T3=Jul-Sep, T4=Oct-Dec
        mois = periode.mois
        trim_debut = ((mois - 1) // 3) * 3 + 1   # 1, 4, 7, 10
        trim_fin   = trim_debut + 2                # 3, 6, 9, 12
        trim_num   = (mois - 1) // 3 + 1          # 1, 2, 3, 4
        trim_label = f"T{trim_num} {periode.annee} ({['Jan-Mar','Avr-Jun','Jul-Sep','Oct-Déc'][trim_num-1]})"

        periodes_trim = PeriodePaie.query.filter_by(
            tenant_id=t.id, annee=periode.annee
        ).filter(
            PeriodePaie.mois >= trim_debut,
            PeriodePaie.mois <= trim_fin
        ).all()
        ids_trim = [p.id for p in periodes_trim]

        bulletins_trim = BulletinPaie.query.filter(
            BulletinPaie.tenant_id == t.id,
            BulletinPaie.periode_id.in_(ids_trim)
        ).options(joinedload(BulletinPaie.salarie),
                  joinedload(BulletinPaie.periode)).all()

        # Regrouper par salarié pour le trimestre
        from collections import defaultdict
        sal_trim = defaultdict(lambda: {
            "salarie": None, "brut": 0, "base_cnss": 0,
            "cnss_sal": 0, "cnss_pat": 0,
            "base_cnamgs": 0, "cnamgs_sal": 0, "cnamgs_pat": 0,
            "mois_list": []
        })
        for b in bulletins_trim:
            k = b.salarie_id
            sal_trim[k]["salarie"]    = b.salarie
            sal_trim[k]["brut"]      += float(b.salaire_brut   or 0)
            sal_trim[k]["base_cnss"] += float(b.base_cnss      or 0)
            sal_trim[k]["cnss_sal"]  += float(b.cnss_salarie   or 0)
            sal_trim[k]["cnss_pat"]  += float(b.cnss_patronale or 0)
            sal_trim[k]["base_cnamgs"]+= float(b.base_cnamgs   or 0)
            sal_trim[k]["cnamgs_sal"]+= float(b.cnamgs_salarie  or 0)
            sal_trim[k]["cnamgs_pat"]+= float(b.cnamgs_patronale or 0)
            sal_trim[k]["mois_list"].append(b.periode.mois if b.periode else 0)

        lignes_trim = sorted(sal_trim.values(), key=lambda x: x["salarie"].nom if x["salarie"] else "")

        stats_trim = {
            "trim_label":    trim_label,
            "trim_num":      trim_num,
            "nb":            len(lignes_trim),
            "mois_couverts": sorted(set(
                m for lg in lignes_trim for m in lg["mois_list"]
            )),
            "total_brut":    sum(lg["brut"]       for lg in lignes_trim),
            "total_cnss_sal":sum(lg["cnss_sal"]   for lg in lignes_trim),
            "total_cnss_pat":sum(lg["cnss_pat"]   for lg in lignes_trim),
            "total_cnamgs_sal":sum(lg["cnamgs_sal"]  for lg in lignes_trim),
            "total_cnamgs_pat":sum(lg["cnamgs_pat"]  for lg in lignes_trim),
        }
        stats_trim["total_cnss"]     = stats_trim["total_cnss_sal"]   + stats_trim["total_cnss_pat"]
        stats_trim["total_cnamgs"]   = stats_trim["total_cnamgs_sal"] + stats_trim["total_cnamgs_pat"]
        stats_trim["total_a_verser"] = stats_trim["total_cnss"] + stats_trim["total_cnamgs"]

    else:
        lignes_trim = []

    MOIS_FR = ["","Jan","Fév","Mar","Avr","Mai","Jun","Jul","Aoû","Sep","Oct","Nov","Déc"]

    return render_template("tenant/declaration_cnss.html",
        tenant=t, periodes=periodes, periode=periode, mode=mode,
        bulletins_mois=bulletins_mois, stats_mensuel=stats_mensuel,
        lignes_trim=lignes_trim, stats_trim=stats_trim,
        MOIS_FR=MOIS_FR)


# ══════════════════════════════════════════════════════════════════════════════
# DÉCLARATION ANNUELLE DES SALAIRES (DAS) — réservée à l'abonnement Cabinet
# ══════════════════════════════════════════════════════════════════════════════

@bp.route("/declaration-das")
@login_required
@plan_required("CABINET", "CABINET_COMPTABLE")
def declaration_das():
    """Déclaration Annuelle des Salaires (DGI) — synthèse annuelle par exercice."""
    if current_user.is_super_admin:
        return redirect(url_for("admin.admin_dashboard"))
    t = get_tenant()
    if not t:
        return redirect(url_for("auth.login"))

    # Exercices disponibles (années ayant au moins une période)
    annees = sorted({p.annee for p in PeriodePaie.query.filter_by(tenant_id=t.id).all()},
                    reverse=True)
    annee = request.args.get("annee", type=int) or (annees[0] if annees else date.today().year)

    lignes, totaux, erreur = [], {}, None
    lignes_hono, tot_hono = [], {}
    from declaration_das import agreger_das, agreger_honoraires, DASVide
    import models as _models
    try:
        lignes, totaux = agreger_das(t, annee, models=_models)
    except DASVide as e:
        erreur = str(e)
    # Volet honoraires (optionnel — ne bloque jamais la DAS salaires)
    try:
        lignes_hono, tot_hono = agreger_honoraires(t, annee, models=_models)
    except Exception:
        lignes_hono, tot_hono = [], {}

    return render_template("tenant/declaration_das.html",
        tenant=t, annees=annees, annee=annee,
        lignes=lignes, totaux=totaux, erreur=erreur,
        lignes_hono=lignes_hono, tot_hono=tot_hono)


@bp.route("/declaration-das/excel")
@login_required
@plan_required("CABINET", "CABINET_COMPTABLE")
def declaration_das_excel():
    """Télécharge la DAS de l'exercice au format Excel."""
    if current_user.is_super_admin:
        return redirect(url_for("admin.admin_dashboard"))
    t = get_tenant()
    if not t:
        return redirect(url_for("auth.login"))

    annee = request.args.get("annee", type=int) or date.today().year
    from declaration_das import generer_das_excel, DASVide
    import models as _models
    try:
        contenu = generer_das_excel(t, annee, models=_models)
    except DASVide as e:
        flash(str(e), "error")
        return redirect(url_for("tenant.declaration_das", annee=annee))

    from flask import Response
    slug = (t.sigle or t.denomination or "entreprise").replace(" ", "_")[:30]
    nom = f"DAS_{slug}_{annee}.xlsx"
    return Response(
        contenu,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{nom}"'},
    )


# ══════════════════════════════════════════════════════════════════════════════
# EXPORT COMPTABLE SAGE 100
# ══════════════════════════════════════════════════════════════════════════════

@bp.route("/export/sage/journal/<int:periode_id>")
@tenant_required
def export_sage_journal(periode_id):
    if not current_user.can_export_sage:
        flash("Accès refusé. Seuls les Administrateurs et Comptables peuvent exporter vers Sage.", "error")
        return redirect(url_for("tenant.bulletins"))
    """
    Export du journal de paie mensuel au format Sage 100 (.txt).
    Importable dans Sage 100 Comptabilité via Fichier → Importer → Journal.
    Seuls les bulletins VALIDE sont inclus.
    """
    t = get_tenant()
    periode = PeriodePaie.query.filter_by(id=periode_id, tenant_id=t.id).first_or_404()
    bulletins = (BulletinPaie.query
                 .filter(BulletinPaie.periode_id==periode_id, BulletinPaie.tenant_id==t.id, BulletinPaie.statut.in_(["VALIDE","VALIDÉ"]))
                 .join(Salarie)
                 .order_by(Salarie.nom)
                 .all())

    if not bulletins:
        flash("Aucun bulletin validé pour cette période. Validez les bulletins avant l'export.", "warning")
        return redirect(url_for("tenant.bulletins"))

    try:
        from export_comptable import generer_journal_paie, ExportVide
        contenu = generer_journal_paie(bulletins, periode, t)
        nom_fichier = f"journal_paie_{periode.mois:02d}{periode.annee}_{t.sigle or t.id}.txt"
        logger.info(f"[Export Sage] Journal paie — tenant={t.id} période={periode.libelle_complet}")
        log_action("EXPORT", "bulletin", periode_id,
                   f"Export Sage Journal de paie — {periode.libelle_complet} ({len(bulletins)} bulletins)")
        db.session.commit()
        return send_file(
            io.BytesIO(contenu),
            mimetype="text/plain",
            as_attachment=True,
            download_name=nom_fichier,
        )
    except Exception as e:
        logger.error(f"[Export Sage] Erreur journal : {e}")
        flash(f"Erreur lors de la génération : {e}", "error")
        return redirect(url_for("tenant.bulletins"))


@bp.route("/export/sage/livre/<int:periode_id>")
@tenant_required
def export_sage_livre(periode_id):
    """
    Export du livre de paie détaillé par salarié au format CSV (.csv).
    Compatible Excel et importable dans Sage 100.
    Seuls les bulletins VALIDE sont inclus.
    """
    t = get_tenant()
    periode = PeriodePaie.query.filter_by(id=periode_id, tenant_id=t.id).first_or_404()
    bulletins = (BulletinPaie.query
                 .filter(BulletinPaie.periode_id==periode_id, BulletinPaie.tenant_id==t.id, BulletinPaie.statut.in_(["VALIDE","VALIDÉ"]))
                 .join(Salarie)
                 .order_by(Salarie.nom)
                 .all())

    if not bulletins:
        flash("Aucun bulletin validé pour cette période. Validez les bulletins avant l'export.", "warning")
        return redirect(url_for("tenant.bulletins"))

    try:
        from export_comptable import generer_livre_paie, ExportVide
        contenu = generer_livre_paie(bulletins, periode, t)
        nom_fichier = f"livre_paie_{periode.mois:02d}{periode.annee}_{t.sigle or t.id}.csv"
        logger.info(f"[Export Sage] Livre paie — tenant={t.id} période={periode.libelle_complet}")
        return send_file(
            io.BytesIO(contenu),
            mimetype="text/csv; charset=utf-8",
            as_attachment=True,
            download_name=nom_fichier,
        )
    except Exception as e:
        logger.error(f"[Export Sage] Erreur livre : {e}")
        flash(f"Erreur lors de la génération : {e}", "error")
        return redirect(url_for("tenant.bulletins"))


@bp.route("/export/sage/les-deux/<int:periode_id>")
@tenant_required
def export_sage_les_deux(periode_id):
    """
    Export des deux fichiers (journal + livre) dans une archive ZIP.
    Pratique pour envoyer tout au comptable en une fois.
    """
    import zipfile
    t = get_tenant()
    periode = PeriodePaie.query.filter_by(id=periode_id, tenant_id=t.id).first_or_404()
    bulletins = (BulletinPaie.query
                 .filter(BulletinPaie.periode_id==periode_id, BulletinPaie.tenant_id==t.id, BulletinPaie.statut.in_(["VALIDE","VALIDÉ"]))
                 .join(Salarie)
                 .order_by(Salarie.nom)
                 .all())

    if not bulletins:
        flash("Aucun bulletin validé pour cette période.", "warning")
        return redirect(url_for("tenant.bulletins"))

    try:
        from export_comptable import generer_journal_paie, generer_livre_paie
        journal = generer_journal_paie(bulletins, periode, t)
        livre   = generer_livre_paie(bulletins, periode, t)

        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
            zf.writestr(
                f"journal_paie_{periode.mois:02d}{periode.annee}_{t.sigle or t.id}.txt",
                journal
            )
            zf.writestr(
                f"livre_paie_{periode.mois:02d}{periode.annee}_{t.sigle or t.id}.csv",
                livre
            )
        zip_buffer.seek(0)

        nom_zip = f"export_sage_{periode.mois:02d}{periode.annee}_{t.sigle or t.id}.zip"
        logger.info(f"[Export Sage] ZIP généré — tenant={t.id}")
        return send_file(
            zip_buffer,
            mimetype="application/zip",
            as_attachment=True,
            download_name=nom_zip,
        )
    except Exception as e:
        logger.error(f"[Export Sage] Erreur ZIP : {e}")
        flash(f"Erreur lors de la génération : {e}", "error")
        return redirect(url_for("tenant.bulletins"))


# ══════════════════════════════════════════════════════════════════════════════
# RAPPORT PDF MENSUEL
# ══════════════════════════════════════════════════════════════════════════════

@bp.route("/rapport/pdf/<int:periode_id>")
@tenant_required
def rapport_pdf(periode_id):
    """
    Génère le rapport PDF mensuel complet organisé par site :
      - Salariés + journaliers regroupés par site
      - Récapitulatif global et charges à verser
    """
    t = get_tenant()
    if not current_user.can_export and not current_user.is_tenant_admin:
        flash("Accès non autorisé.", "error")
        return redirect(url_for("tenant.bulletins"))

    periode = PeriodePaie.query.filter_by(id=periode_id, tenant_id=t.id).first_or_404()

    # ── Bulletins salariés validés ────────────────────────────────────────────
    bulletins = (BulletinPaie.query
        .filter(BulletinPaie.periode_id == periode_id,
                BulletinPaie.tenant_id  == t.id,
                BulletinPaie.statut.in_(["VALIDÉ","VALIDE","PAYÉ"]))
        .options(joinedload(BulletinPaie.salarie))
        .join(Salarie).order_by(Salarie.nom).all())

    # ── Feuilles de paie journaliers du même mois ─────────────────────────────
    from datetime import date as _date
    debut_mois = _date(periode.annee, periode.mois, 1)
    import calendar
    dernier_jour = calendar.monthrange(periode.annee, periode.mois)[1]
    fin_mois = _date(periode.annee, periode.mois, dernier_jour)

    feuilles = (FeuillePaieJournalier.query
        .filter_by(tenant_id=t.id)
        .filter(FeuillePaieJournalier.date_debut >= debut_mois,
                FeuillePaieJournalier.date_fin   <= fin_mois)
        .options(joinedload(FeuillePaieJournalier.journalier))
        .all())

    if not bulletins and not feuilles:
        flash("Aucun bulletin ni feuille de paie pour cette période.", "warning")
        return redirect(url_for("tenant.bulletins"))

    # ── Sites et affectations ─────────────────────────────────────────────────
    sites = Site.query.filter_by(tenant_id=t.id, actif=True).order_by(Site.nom).all()
    affectations = AffectationSite.query.filter_by(tenant_id=t.id, actif=True).all()

    # ── Évolution 6 mois ──────────────────────────────────────────────────────
    MOIS_FR = ["","Jan","Fév","Mar","Avr","Mai","Jun","Jul","Aoû","Sep","Oct","Nov","Déc"]
    evolution = []
    for i in range(5, -1, -1):
        mois  = (periode.mois - i - 1) % 12 + 1
        annee = periode.annee - ((i + 12 - periode.mois) // 12) if i >= periode.mois else periode.annee
        buls_m = BulletinPaie.query.join(PeriodePaie).filter(
            BulletinPaie.tenant_id == t.id,
            PeriodePaie.mois == mois, PeriodePaie.annee == annee,
        ).all()
        evolution.append({
            "mois":  MOIS_FR[mois],
            "annee": annee,
            "brut":  sum(float(b.salaire_brut or 0) for b in buls_m),
            "net":   sum(float(b.net_a_payer  or 0) for b in buls_m),
        })

    try:
        from rapport_pdf import generer_rapport_mensuel
        pdf_bytes = generer_rapport_mensuel(
            bulletins            = bulletins,
            periode              = periode,
            tenant               = t,
            feuilles_journaliers = feuilles,
            sites                = sites,
            affectations         = affectations,
            evolution            = evolution,
        )

        nom_fichier = f"rapport_paie_{periode.mois:02d}{periode.annee}_{t.sigle or t.id}.pdf"
        log_action("EXPORT", "rapport", periode_id,
                   f"Rapport PDF — {periode.libelle_complet} — "
                   f"{len(bulletins)} salariés, {len(feuilles)} journaliers, "
                   f"{len(sites)} sites")
        db.session.commit()
        logger.info(f"[Rapport PDF] Généré — tenant={t.id} période={periode.libelle_complet}")
        return send_file(
            io.BytesIO(pdf_bytes),
            mimetype="application/pdf",
            as_attachment=True,
            download_name=nom_fichier,
        )
    except Exception as e:
        logger.error(f"[Rapport PDF] Erreur : {e}")
        flash(f"Erreur lors de la génération du PDF : {e}", "error")
        return redirect(url_for("tenant.bulletins"))


@bp.route("/rapport/pdf/<int:periode_id>/envoyer-email", methods=["POST"])
@tenant_required
def rapport_pdf_email(periode_id):
    """
    Génère le rapport PDF et l'envoie par email aux admins et directeurs du tenant.
    Peut être déclenché manuellement ou automatiquement en fin de mois.
    """
    t = get_tenant()
    if not current_user.is_tenant_admin:
        flash("Accès réservé à l'administrateur.", "error")
        return redirect(url_for("tenant.bulletins"))

    periode = PeriodePaie.query.filter_by(id=periode_id, tenant_id=t.id).first_or_404()
    bulletins_p = (BulletinPaie.query
        .filter(BulletinPaie.periode_id == periode_id,
                BulletinPaie.tenant_id  == t.id,
                BulletinPaie.statut.in_(["VALIDÉ","VALIDE","PAYÉ"]))
        .options(joinedload(BulletinPaie.salarie))
        .join(Salarie).order_by(Salarie.nom).all())

    if not bulletins_p:
        flash("Aucun bulletin validé — rapport non envoyé.", "warning")
        return redirect(url_for("tenant.bulletins"))

    # Destinataires : admins et directeurs actifs
    destinataires = [
        u.email for u in Utilisateur.query.filter_by(tenant_id=t.id, actif=True).all()
        if u.email and u.role in ("TENANT_ADMIN", "DIRECTEUR")
    ]
    email_extra = request.form.get("email_extra", "").strip()
    if email_extra:
        destinataires.append(email_extra)

    if not destinataires:
        flash("Aucun destinataire configuré.", "error")
        return redirect(url_for("tenant.bulletins"))

    try:
        from rapport_pdf import generer_rapport_mensuel
        evolution = []  # Simplifié pour l'envoi auto
        pdf_bytes = generer_rapport_mensuel(bulletins_p, periode, t, evolution)
        nom_fichier = f"rapport_paie_{periode.mois:02d}{periode.annee}_{t.sigle or t.id}.pdf"

        from calculs_paie import calculer_masse_salariale
        masse = calculer_masse_salariale(bulletins_p)

        msg = Message(
            subject=f"[PaieGabon] Rapport mensuel — {t.denomination} — {periode.libelle_complet}",
            recipients=destinataires,
            body=(
                f"Bonjour,\n\n"
                f"Veuillez trouver en pièce jointe le rapport mensuel de paie de {t.denomination} "
                f"pour la période {periode.libelle_complet} {periode.annee}.\n\n"
                f"Résumé :\n"
                f"  • Nombre de bulletins : {len(bulletins_p)}\n"
                f"  • Masse salariale brute : {int(masse.get('total_brut',0)):,} FCFA\n"
                f"  • Net total à payer : {int(masse.get('total_net',0)):,} FCFA\n"
                f"  • Coût total employeur : {int(masse.get('total_brut',0) + masse.get('total_charges_pat',0)):,} FCFA\n\n"
                f"Ce rapport est confidentiel et destiné à usage interne uniquement.\n\n"
                f"Cordialement,\nPaieGabon SaaS"
            ),
        )
        msg.attach(nom_fichier, "application/pdf", pdf_bytes)
        send_email_async(current_app.extensions["mail"], msg)

        log_action("EXPORT", "rapport", periode_id,
                   f"Rapport PDF envoyé par email à {len(destinataires)} destinataire(s)")
        db.session.commit()

        flash(f"Rapport envoyé par email à {len(destinataires)} destinataire(s).", "success")
        logger.info(f"[Rapport PDF Email] Envoyé à {destinataires} — tenant={t.id}")

    except Exception as e:
        logger.error(f"[Rapport PDF Email] Erreur : {e}")
        flash(f"Erreur lors de l'envoi : {e}", "error")

    return redirect(url_for("tenant.bulletins", periode_id=periode_id))


@bp.route("/declaration-cnss/export-excel")
@login_required
def declaration_cnss_excel():
    """Export Excel déclarations : mensuel (CFP/FNH/TCS/IRPP) ou trimestriel (CNSS/CNAMGS)."""
    if current_user.is_super_admin: return redirect(url_for("admin.admin_dashboard"))
    t = get_tenant()
    if not t: return redirect(url_for("auth.login"))

    pid     = request.args.get("periode_id", type=int)
    mode    = request.args.get("mode", "mensuel")
    periode = PeriodePaie.query.filter_by(id=pid, tenant_id=t.id).first_or_404()

    MOIS_FR2 = ["","Janvier","Février","Mars","Avril","Mai","Juin",
                "Juillet","Août","Septembre","Octobre","Novembre","Décembre"]

    if mode == "trimestriel":
        # ── Trimestre : récupérer les 3 mois ─────────────────────────────────
        mois=periode.mois; trim_num=(mois-1)//3+1
        trim_debut=((mois-1)//3)*3+1; trim_fin=trim_debut+2
        trim_label=f"T {trim_num}"
        mois_labels=[MOIS_FR2[m] for m in range(trim_debut, trim_fin+1)]

        periodes_trim=PeriodePaie.query.filter_by(tenant_id=t.id, annee=periode.annee)            .filter(PeriodePaie.mois>=trim_debut, PeriodePaie.mois<=trim_fin).all()
        mois_map={p.mois: p.id for p in periodes_trim}

        buls_trim=BulletinPaie.query.filter(
            BulletinPaie.tenant_id==t.id,
            BulletinPaie.periode_id.in_([p.id for p in periodes_trim])
        ).options(joinedload(BulletinPaie.salarie), joinedload(BulletinPaie.periode)).all()

        # Regrouper par salarié
        from collections import defaultdict
        sal_map = defaultdict(lambda: {
            "nom_complet":"","matricule":"","numero_cnss":"","numero_cnamgs":"",
            "date_embauche":"",
            "m1_base_cnss":0,"m2_base_cnss":0,"m3_base_cnss":0,
            "m1_base_cnamgs":0,"m2_base_cnamgs":0,"m3_base_cnamgs":0,
        })
        for b in buls_trim:
            k=b.salarie_id
            sal=b.salarie
            sal_map[k]["nom_complet"]   = sal.nom_complet
            sal_map[k]["matricule"]     = sal.matricule or ""
            sal_map[k]["numero_cnss"]   = sal.numero_cnss or ""
            sal_map[k]["numero_cnamgs"] = sal.numero_cnamgs or ""
            if sal.date_embauche:
                sal_map[k]["date_embauche"] = sal.date_embauche.strftime("%d/%m/%Y")
            m = b.periode.mois if b.periode else 0
            if m == trim_debut:
                sal_map[k]["m1_base_cnss"]   = float(b.base_cnss   or 0)
                sal_map[k]["m1_base_cnamgs"] = float(b.base_cnamgs or 0)
            elif m == trim_debut+1:
                sal_map[k]["m2_base_cnss"]   = float(b.base_cnss   or 0)
                sal_map[k]["m2_base_cnamgs"] = float(b.base_cnamgs or 0)
            elif m == trim_fin:
                sal_map[k]["m3_base_cnss"]   = float(b.base_cnss   or 0)
                sal_map[k]["m3_base_cnamgs"] = float(b.base_cnamgs or 0)

        sal_data = sorted(sal_map.values(), key=lambda x: x["nom_complet"])

        tot_cnss_m  = [sum(s["m1_base_cnss"]   for s in sal_data),
                       sum(s["m2_base_cnss"]   for s in sal_data),
                       sum(s["m3_base_cnss"]   for s in sal_data)]
        tot_cnamgs_m= [sum(s["m1_base_cnamgs"] for s in sal_data),
                       sum(s["m2_base_cnamgs"] for s in sal_data),
                       sum(s["m3_base_cnamgs"] for s in sal_data)]
        total_base_cnss   = sum(tot_cnss_m)
        total_base_cnamgs = sum(tot_cnamgs_m)

        # Générer les deux fichiers et les zip
        import zipfile, io
        cnss_bytes   = _gen_excel_cnss(t, trim_label, periode.annee, mois_labels,
                                       sal_data, total_base_cnss, total_base_cnamgs, tot_cnss_m)
        cnamgs_bytes = _gen_excel_cnamgs(t, trim_label, periode.annee, mois_labels,
                                         sal_data, total_base_cnamgs, tot_cnamgs_m)

        zip_buf = io.BytesIO()
        nom_base = t.denomination.replace(" ","_")[:20]
        with zipfile.ZipFile(zip_buf, 'w', zipfile.ZIP_DEFLATED) as zf:
            zf.writestr(f"CNSS_{nom_base}_{trim_label.replace(' ','')}_{periode.annee}.xlsx",   cnss_bytes)
            zf.writestr(f"CNAMGS_{nom_base}_{trim_label.replace(' ','')}_{periode.annee}.xlsx", cnamgs_bytes)
        zip_buf.seek(0)

        from flask import Response
        nom_zip = f"declarations_trimestrielles_{nom_base}_{trim_label.replace(' ','')}_{periode.annee}.zip"
        return Response(zip_buf.read(), mimetype="application/zip",
            headers={"Content-Disposition": f'attachment; filename="{nom_zip}"'})


# ══════════════════════════════════════════════════════════════════════════════
# DÉCLARATION CNSS/CNAMGS — Export CSV portail électronique
# ══════════════════════════════════════════════════════════════════════════════

@bp.route("/declaration-cnss/export-csv")
@login_required
def declaration_cnss_csv():
    """
    Export CSV uploadable directement sur le portail CNSS Gabon (cnss.ga)
    et CNAMGS. Génère une archive ZIP avec les deux fichiers CSV.
    """
    if current_user.is_super_admin: return redirect(url_for("admin.admin_dashboard"))
    t = get_tenant()
    if not t: return redirect(url_for("auth.login"))

    pid     = request.args.get("periode_id", type=int)
    periode = PeriodePaie.query.filter_by(id=pid, tenant_id=t.id).first_or_404()

    # ── Calculer le trimestre ──────────────────────────────────────────────
    from declaration_cnss import calculer_trimestre, generer_csv_cnss, generer_csv_cnamgs
    trim_num, trim_debut, trim_fin, trim_label = calculer_trimestre(periode.mois)

    periodes_trim = PeriodePaie.query.filter_by(
        tenant_id=t.id, annee=periode.annee
    ).filter(
        PeriodePaie.mois >= trim_debut,
        PeriodePaie.mois <= trim_fin
    ).all()

    if not periodes_trim:
        flash("Aucune période trouvée pour ce trimestre.", "warning")
        return redirect(url_for("tenant.declaration_cnss", periode_id=pid, mode="trimestriel"))

    buls_trim = BulletinPaie.query.filter(
        BulletinPaie.tenant_id == t.id,
        BulletinPaie.periode_id.in_([p.id for p in periodes_trim])
    ).options(
        joinedload(BulletinPaie.salarie),
        joinedload(BulletinPaie.periode)
    ).all()

    if not buls_trim:
        flash("Aucun bulletin pour ce trimestre. Saisissez et validez les bulletins d'abord.", "warning")
        return redirect(url_for("tenant.declaration_cnss", periode_id=pid, mode="trimestriel"))

    # ── Regrouper par salarié ──────────────────────────────────────────────
    from collections import defaultdict
    sal_map = defaultdict(lambda: {
        "nom_complet": "", "matricule": "", "numero_cnss": "",
        "numero_cnamgs": "", "date_embauche": "",
        "m1_base_cnss": 0, "m2_base_cnss": 0, "m3_base_cnss": 0,
        "m1_base_cnamgs": 0, "m2_base_cnamgs": 0, "m3_base_cnamgs": 0,
    })

    for b in buls_trim:
        k   = b.salarie_id
        sal = b.salarie
        sal_map[k]["nom_complet"]   = sal.nom_complet
        sal_map[k]["matricule"]     = sal.matricule or ""
        sal_map[k]["numero_cnss"]   = sal.numero_cnss or ""
        sal_map[k]["numero_cnamgs"] = sal.numero_cnamgs or ""
        if sal.date_embauche:
            sal_map[k]["date_embauche"] = sal.date_embauche.strftime("%d/%m/%Y")
        m = b.periode.mois if b.periode else 0
        if m == trim_debut:
            sal_map[k]["m1_base_cnss"]   = float(b.base_cnss   or 0)
            sal_map[k]["m1_base_cnamgs"] = float(b.base_cnamgs or 0)
        elif m == trim_debut + 1:
            sal_map[k]["m2_base_cnss"]   = float(b.base_cnss   or 0)
            sal_map[k]["m2_base_cnamgs"] = float(b.base_cnamgs or 0)
        elif m == trim_fin:
            sal_map[k]["m3_base_cnss"]   = float(b.base_cnss   or 0)
            sal_map[k]["m3_base_cnamgs"] = float(b.base_cnamgs or 0)

    sal_data = list(sal_map.values())

    # ── Générer les deux CSV ───────────────────────────────────────────────
    try:
        csv_cnss   = generer_csv_cnss(sal_data, periode, t, trim_debut, trim_fin)
        csv_cnamgs = generer_csv_cnamgs(sal_data, periode, t, trim_debut, trim_fin)

        import zipfile
        zip_buf  = io.BytesIO()
        nom_base = (t.sigle or t.denomination[:15]).replace(" ", "_")
        trim_str = f"T{trim_num}_{periode.annee}"

        with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
            zf.writestr(f"CNSS_{nom_base}_{trim_str}.csv",   csv_cnss)
            zf.writestr(f"CNAMGS_{nom_base}_{trim_str}.csv", csv_cnamgs)
            # Ajouter un fichier README avec les instructions d'upload
            zf.writestr(
                "INSTRUCTIONS_UPLOAD.txt",
                _instructions_upload(t, trim_label, periode.annee, len(sal_data))
            )
        zip_buf.seek(0)

        nom_zip = f"declarations_CNSS_CNAMGS_{nom_base}_{trim_str}.zip"
        logger.info(f"[CNSS CSV] Export {trim_str} — {len(sal_data)} salariés — tenant={t.id}")

        return send_file(
            zip_buf,
            mimetype="application/zip",
            as_attachment=True,
            download_name=nom_zip,
        )

    except Exception as e:
        logger.error(f"[CNSS CSV] Erreur export : {e}")
        flash(f"Erreur lors de la génération : {e}", "error")
        return redirect(url_for("tenant.declaration_cnss", periode_id=pid, mode="trimestriel"))


def _instructions_upload(tenant, trim_label, annee, nb_salaries) -> str:
    """Génère un fichier texte d'instructions pour l'upload sur les portails."""
    return f"""
INSTRUCTIONS D'UPLOAD — DÉCLARATIONS TRIMESTRIELLES
====================================================
Entreprise : {tenant.denomination}
NIF        : {tenant.nif or "—"}
Trimestre  : {trim_label} {annee}
Salariés   : {nb_salaries}
Généré le  : {datetime.now().strftime("%d/%m/%Y à %H:%M")}


FICHIER CNSS : CNSS_*.csv
─────────────────────────
1. Connectez-vous sur https://cnss.ga
2. Allez dans : Mon Espace → Déclarations → Nouvelle déclaration
3. Choisissez : Déclaration trimestrielle de salaires
4. Cliquez sur "Importer un fichier"
5. Sélectionnez le fichier CNSS_*.csv
6. Vérifiez les montants affichés
7. Validez et téléchargez le reçu


FICHIER CNAMGS : CNAMGS_*.csv
──────────────────────────────
1. Connectez-vous sur le portail CNAMGS
2. Allez dans : Déclarations → Déclaration trimestrielle
3. Importez le fichier CNAMGS_*.csv
4. Vérifiez et validez


MONTANTS À VERSER (rappel) :
────────────────────────────
CNSS  : cotisations salariales (5%) + patronales (18%) = 23% de la base
CNAMGS: cotisations salariales (1,5%) + patronales (6%) = 7,5% de la base

Date limite de dépôt : dernier jour du mois suivant la fin du trimestre
  T1 (Jan-Mar) → 30 Avril
  T2 (Avr-Jun) → 31 Juillet
  T3 (Jul-Sep) → 31 Octobre
  T4 (Oct-Déc) → 31 Janvier


IMPORTANT :
──────────
- Utilisez le fichier CSV, pas le fichier Excel, pour l'upload portail
- Le fichier Excel (généré séparément) est pour vos archives papier
- Gardez le reçu de dépôt comme justificatif

En cas de problème : support@paiegalon.com
""".strip()


# ══════════════════════════════════════════════════════════════════════════════
# API REST v1 — Intégration grandes entreprises
# ══════════════════════════════════════════════════════════════════════════════
# Authentification : X-API-Key: <token>  ou  Authorization: Bearer <oauth_token>
# Tous les endpoints retournent JSON. Préfixe : /api/v1/
# ══════════════════════════════════════════════════════════════════════════════

from api_rest import (api_auth_required, _ok, _err, _paginate,
                      _salarie_dict, _bulletin_dict, _periode_dict,
                      OAUTH_TOKEN_TTL)


# ══════════════════════════════════════════════════════════════════════════════
# RECHERCHE GLOBALE
# ══════════════════════════════════════════════════════════════════════════════
@bp.route("/recherche")
@login_required
def recherche_globale():
    """Recherche globale : salariés, journaliers, bulletins, acomptes."""
    if current_user.is_super_admin:
        return redirect(url_for("admin.admin_dashboard"))
    t = get_tenant()
    if not t:
        return redirect(url_for("auth.login"))

    q = request.args.get("q", "").strip()
    if not q:
        return render_template("tenant/recherche.html",
                               tenant=t, q="", resultats={}, nb_total=0)

    like = f"%{q}%"
    resultats = {}

    # ── Salariés ──────────────────────────────────────────────────────────────
    sals = (Salarie.query.filter_by(tenant_id=t.id)
            .filter(db.or_(
                Salarie.nom.ilike(like), Salarie.prenom.ilike(like),
                Salarie.matricule.ilike(like), Salarie.emploi.ilike(like),
                Salarie.telephone.ilike(like)))
            .order_by(Salarie.nom).limit(10).all())
    if sals:
        resultats["salaries"] = [{"id": s.id, "titre": s.nom_complet,
            "sous_titre": f"{s.emploi or '—'} · {s.matricule}",
            "badge": s.statut, "lien": f"/salaries/{s.id}",
            "icone": "👤"} for s in sals]

    # ── Journaliers ───────────────────────────────────────────────────────────
    jours = (Journalier.query.filter_by(tenant_id=t.id)
             .filter(db.or_(
                 Journalier.nom.ilike(like), Journalier.prenom.ilike(like),
                 Journalier.profession.ilike(like), Journalier.telephone.ilike(like)))
             .order_by(Journalier.nom).limit(10).all())
    if jours:
        resultats["journaliers"] = [{"id": j.id, "titre": j.nom_complet,
            "sous_titre": f"{j.profession or '—'} · {int(j.taux_horaire or 0)} FCFA/h",
            "badge": j.statut, "lien": f"/journaliers/{j.id}",
            "icone": "🦺"} for j in jours]

    # ── Bulletins ─────────────────────────────────────────────────────────────
    buls = (BulletinPaie.query.filter_by(tenant_id=t.id)
            .join(Salarie, BulletinPaie.salarie_id == Salarie.id)
            .join(PeriodePaie, BulletinPaie.periode_id == PeriodePaie.id)
            .filter(db.or_(
                Salarie.nom.ilike(like), Salarie.prenom.ilike(like),
                Salarie.matricule.ilike(like)))
            .order_by(BulletinPaie.date_creation.desc()).limit(10).all())
    if buls:
        resultats["bulletins"] = [{"id": b.id,
            "titre": b.salarie.nom_complet,
            "sous_titre": f"{b.periode.libelle_complet} · Net : {int(b.net_a_payer or 0):,} FCFA",
            "badge": b.statut, "lien": f"/bulletins/{b.id}",
            "icone": "📄"} for b in buls]

    # ── Acomptes ──────────────────────────────────────────────────────────────
    acomps = (Acompte.query.filter_by(tenant_id=t.id)
              .join(Salarie, Acompte.salarie_id == Salarie.id)
              .filter(db.or_(
                  Salarie.nom.ilike(like), Salarie.prenom.ilike(like),
                  Salarie.matricule.ilike(like)))
              .order_by(Acompte.date_acompte.desc()).limit(10).all())
    if acomps:
        resultats["acomptes"] = [{"id": a.id,
            "titre": a.salarie.nom_complet,
            "sous_titre": f"{int(a.montant or 0):,} FCFA · {a.date_acompte.strftime('%d/%m/%Y') if a.date_acompte else ''}",
            "badge": a.statut, "lien": "/acomptes",
            "icone": "💸"} for a in acomps]

    # ── Prestataires / sous-traitants ──────────────────────────────────────────
    prests = (Prestataire.query.filter_by(tenant_id=t.id)
              .filter(db.or_(
                  Prestataire.raison_sociale.ilike(like), Prestataire.code.ilike(like),
                  Prestataire.activite.ilike(like), Prestataire.telephone.ilike(like)))
              .order_by(Prestataire.raison_sociale).limit(10).all())
    if prests:
        resultats["prestataires"] = [{"id": pr.id, "titre": pr.raison_sociale,
            "sous_titre": f"{pr.categorie_label} · {pr.code}",
            "badge": pr.statut if hasattr(pr, "statut") else None,
            "lien": f"/prestataires/{pr.id}", "icone": "🛠️"} for pr in prests]

    nb_total = sum(len(v) for v in resultats.values())
    return render_template("tenant/recherche.html",
                           tenant=t, q=q, resultats=resultats, nb_total=nb_total)


@bp.route("/api/recherche-rapide")
@login_required
def api_recherche_rapide():
    """API JSON pour l'autocomplétion dans la barre de recherche."""
    t = get_tenant()
    if not t:
        return jsonify([])
    q = request.args.get("q", "").strip()
    if len(q) < 2:
        return jsonify([])

    like = f"%{q}%"
    resultats = []

    for s in (Salarie.query.filter_by(tenant_id=t.id)
              .filter(db.or_(Salarie.nom.ilike(like), Salarie.prenom.ilike(like),
                             Salarie.matricule.ilike(like)))
              .limit(5).all()):
        resultats.append({"icone": "👤", "titre": s.nom_complet,
            "sous_titre": s.emploi or "Salarié", "lien": f"/salaries/{s.id}",
            "categorie": "Salariés"})

    for j in (Journalier.query.filter_by(tenant_id=t.id)
              .filter(db.or_(Journalier.nom.ilike(like), Journalier.prenom.ilike(like),
                             Journalier.profession.ilike(like)))
              .limit(5).all()):
        resultats.append({"icone": "🦺", "titre": j.nom_complet,
            "sous_titre": j.profession or "Journalier", "lien": f"/journaliers/{j.id}",
            "categorie": "Journaliers"})

    for pr in (Prestataire.query.filter_by(tenant_id=t.id)
               .filter(db.or_(Prestataire.raison_sociale.ilike(like),
                              Prestataire.code.ilike(like),
                              Prestataire.activite.ilike(like)))
               .limit(5).all()):
        resultats.append({"icone": "🛠️", "titre": pr.raison_sociale,
            "sous_titre": pr.categorie_label or "Prestataire", "lien": f"/prestataires/{pr.id}",
            "categorie": "Prestataires"})

    return jsonify(resultats[:15])


# ═══════════════════════════════════════════════════════════════════════════
#  POINTAGE SALARIÉS — import mensuel (format LONG : 1 ligne / salarié / jour)
#  Circuit : modèle Excel → remplir → téléverser (aperçu) → confirmer (brouillons)
#  Le classement des heures (10/30/40/70) et la détection dimanche/férié sont
#  délégués à ventiler_heures_mois() ; aucune règle de paie n'est réécrite ici.
# ═══════════════════════════════════════════════════════════════════════════
_MOIS_FR_SAL = ["", "Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
                "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"]
_JOURS_FR_SAL = ["Lun", "Mar", "Mer", "Jeu", "Ven", "Sam", "Dim"]
# Motifs d'absence reconnus (comptés, jamais déduits automatiquement)
_MOTIFS_ABSENCE = ["CONGE", "MALADIE", "INJUSTIFIEE"]


def _sal_periode_demandee():
    """Lit le mois demandé (YYYY-MM) ou le mois courant. Retourne (annee, mois)."""
    mois_str = request.args.get("mois", "") or request.form.get("mois", "")
    try:
        annee, mois = (int(x) for x in mois_str.split("-"))
        date(annee, mois, 1)
        return annee, mois
    except (ValueError, TypeError):
        today = datetime.now()
        return today.year, today.month


@bp.route("/salaries/pointage/modele")
@login_required
def salaries_pointage_modele():
    """Génère le modèle Excel (format LONG) pré-rempli : une ligne par salarié
    actif et par jour du mois, dimanches et fériés déjà marqués. L'utilisateur
    saisit les heures travaillées, la nuit, et un éventuel motif d'absence."""
    if current_user.is_super_admin:
        return redirect(url_for("admin.admin_dashboard"))
    t = get_tenant()
    if not t:
        return redirect(url_for("auth.login"))

    import calendar as _cal
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation
    from io import BytesIO

    annee, mois = _sal_periode_demandee()
    nb_jours = _cal.monthrange(annee, mois)[1]
    mois_nom = f"{_MOIS_FR_SAL[mois]} {annee}"
    feries = {d: n for d, n in jours_feries_annee(annee).items() if d.month == mois}

    salaries = (Salarie.query.filter_by(tenant_id=t.id, statut="ACTIF")
                .order_by(Salarie.nom, Salarie.prenom).all())

    VERT, BLANC = "0F3D36", "FFFFFF"
    FERIE_FILL, DIM_FILL = "FCE4D6", "FFF2CC"
    thin = Side(style="thin", color="D1D5DB")
    bord = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")

    wb = Workbook()
    ws = wb.active
    ws.title = f"Pointage {_MOIS_FR_SAL[mois][:3]} {annee}"
    ws.sheet_view.showGridLines = False

    ws["A1"] = f"POINTAGE SALARIÉS — {mois_nom}"
    ws["A1"].font = Font(name="Arial", size=13, bold=True, color=VERT)
    ws["A2"] = ("Saisissez « Heures travaillées » (et « Dont nuit » si concerné) pour chaque jour. "
                "Les dimanches et fériés sont pré-marqués. Pour une absence, indiquez le motif "
                "(CONGE, MALADIE, INJUSTIFIEE) au lieu des heures.")
    ws["A2"].font = Font(name="Arial", size=9, italic=True, color="666666")

    entetes = ["Matricule", "Nom", "Date", "Jour", "Type",
               "Heures travaillées", "Dont nuit", "Absence (motif)"]
    for i, h in enumerate(entetes, 1):
        c = ws.cell(row=4, column=i, value=h)
        c.font = Font(name="Arial", size=10, bold=True, color=BLANC)
        c.fill = PatternFill("solid", fgColor=VERT)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = bord
    ws.row_dimensions[4].height = 30

    # Validation du motif d'absence (liste déroulante)
    dv = DataValidation(type="list", formula1='"CONGE,MALADIE,INJUSTIFIEE"', allow_blank=True)
    ws.add_data_validation(dv)

    r = 5
    for s in salaries:
        nom_complet = f"{s.nom} {s.prenom or ''}".strip()
        for d in range(1, nb_jours + 1):
            dd = date(annee, mois, d)
            if dd in feries:
                tj = "FERIE"
            elif dd.weekday() == 6:
                tj = "DIMANCHE"
            else:
                tj = "NORMAL"
            ws.cell(row=r, column=1, value=s.matricule)
            ws.cell(row=r, column=2, value=nom_complet)
            ws.cell(row=r, column=3, value=dd.strftime("%d/%m/%Y"))
            ws.cell(row=r, column=4, value=_JOURS_FR_SAL[dd.weekday()])
            ws.cell(row=r, column=5, value=tj)
            for col in range(1, 9):
                cell = ws.cell(row=r, column=col)
                cell.border = bord
                cell.font = Font(name="Arial", size=9)
                if tj == "FERIE":
                    cell.fill = PatternFill("solid", fgColor=FERIE_FILL)
                elif tj == "DIMANCHE":
                    cell.fill = PatternFill("solid", fgColor=DIM_FILL)
            dv.add(ws.cell(row=r, column=8))
            r += 1
        r += 1  # ligne vide entre salariés (lisibilité)

    for col, w in {"A": 12, "B": 24, "C": 13, "D": 7, "E": 11,
                   "F": 18, "G": 11, "H": 18}.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A5"

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    from flask import send_file
    nom_fichier = f"pointage_salaries_{annee}_{mois:02d}.xlsx"
    return send_file(bio, as_attachment=True, download_name=nom_fichier,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


def _lire_pointage_salaries(fichier):
    """Lit le fichier Excel de pointage (format LONG) et regroupe les lignes par
    salarié (matricule). Retourne (data, erreurs) où data est un dict
    matricule -> {"nom":..., "jours":[...], "absences":{motif:count}}.
    Chaque jour = {"date": date, "heures": float, "heures_nuit": float}."""
    from openpyxl import load_workbook
    erreurs = []
    try:
        wb = load_workbook(fichier, data_only=True, read_only=True)
    except Exception:
        return None, ["Fichier illisible : n'est pas un classeur Excel valide."]
    ws = wb.active

    # Repérer la ligne d'en-tête (celle qui contient "Matricule")
    entete_row = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=12, values_only=True), 1):
        cells = [str(c).strip().lower() if c is not None else "" for c in row]
        if "matricule" in cells:
            entete_row = i
            break
    if entete_row is None:
        return None, ["En-tête introuvable : le fichier doit contenir une colonne « Matricule »."]

    data = {}
    for row in ws.iter_rows(min_row=entete_row + 1, values_only=True):
        if not row or all(c is None for c in row):
            continue
        mat = (str(row[0]).strip() if row[0] is not None else "")
        if not mat:
            continue
        nom = (str(row[1]).strip() if len(row) > 1 and row[1] is not None else "")
        date_str = (str(row[2]).strip() if len(row) > 2 and row[2] is not None else "")
        heures = row[5] if len(row) > 5 else None
        nuit = row[6] if len(row) > 6 else None
        motif = (str(row[7]).strip().upper() if len(row) > 7 and row[7] is not None else "")

        # Date : accepte "JJ/MM/AAAA" ou un vrai objet date Excel
        d = None
        if isinstance(date_str, str) and "/" in date_str:
            try:
                jj, mm, aa = (int(x) for x in date_str.split("/"))
                d = date(aa, mm, jj)
            except (ValueError, TypeError):
                d = None
        if d is None and hasattr(row[2], "year"):
            d = row[2] if isinstance(row[2], date) else None
        if d is None:
            continue  # ligne sans date exploitable : ignorée silencieusement

        entree = data.setdefault(mat, {"nom": nom, "jours": [], "absences": {}})
        # Absence : on compte par motif, on ne pose PAS d'heures ce jour-là
        if motif:
            m = motif if motif in _MOTIFS_ABSENCE else "AUTRE"
            entree["absences"][m] = entree["absences"].get(m, 0) + 1
            continue
        try:
            h = float(heures) if heures not in (None, "") else 0.0
        except (ValueError, TypeError):
            h = 0.0
        try:
            hn = float(nuit) if nuit not in (None, "") else 0.0
        except (ValueError, TypeError):
            hn = 0.0
        if h <= 0 and hn <= 0:
            continue  # jour non travaillé, rien à ventiler
        entree["jours"].append({"date": d, "heures": h, "heures_nuit": hn})

    return data, erreurs


@bp.route("/salaries/pointage/importer", methods=["POST"])
@login_required
def salaries_pointage_importer():
    """Étape APERÇU : lit le fichier, ventile les heures par salarié via
    ventiler_heures_mois(), compte les absences, et affiche un récapitulatif.
    Rien n'est enregistré ici."""
    if current_user.is_super_admin:
        return redirect(url_for("admin.admin_dashboard"))
    t = get_tenant()
    if not t:
        return redirect(url_for("auth.login"))
    if not current_user.can_edit:
        abort(403)

    annee, mois = _sal_periode_demandee()
    fichier = request.files.get("fichier")
    if not fichier or not fichier.filename:
        flash("Choisissez un fichier de pointage à téléverser.", "error")
        return redirect(url_for("tenant.salaries_pointage"))

    data, erreurs = _lire_pointage_salaries(fichier)
    if erreurs:
        for e in erreurs:
            flash(e, "error")
        return redirect(url_for("tenant.salaries_pointage"))
    if not data:
        flash("Aucune donnée de pointage trouvée dans le fichier.", "warning")
        return redirect(url_for("tenant.salaries_pointage"))

    feries = set(jours_feries_annee(annee).keys())
    salaries = {s.matricule: s for s in
                Salarie.query.filter_by(tenant_id=t.id, statut="ACTIF").all()}

    apercu = []
    inconnus = []
    for mat, info in data.items():
        s = salaries.get(mat)
        if not s:
            inconnus.append(mat)
            continue
        vent = ventiler_heures_mois(t.convention, info["jours"], feries=feries)
        total_abs = sum(info["absences"].values())
        apercu.append({
            "matricule": mat,
            "nom": info["nom"] or s.nom_complet,
            "salarie_id": s.id,
            "heures_sup_10": vent.get("heures_sup_10", 0),
            "heures_sup_30": vent.get("heures_sup_30", 0),
            "heures_sup_30b": vent.get("heures_sup_30b", 0),
            "heures_sup_40": vent.get("heures_sup_40", 0),
            "heures_sup_70": vent.get("heures_sup_70", 0),
            "nb_jours_travailles": len(info["jours"]),
            "absences": info["absences"],
            "total_absences": total_abs,
        })
    apercu.sort(key=lambda x: x["nom"])

    # Mémoriser en session pour l'étape de confirmation (données légères)
    session["pointage_sal"] = {
        "annee": annee, "mois": mois,
        "lignes": [{k: v for k, v in a.items() if k != "absences"} for a in apercu],
    }

    return render_template("tenant/salaries_pointage_apercu.html",
        apercu=apercu, annee=annee, mois=mois, tenant=t,
        mois_nom=f"{_MOIS_FR_SAL[mois]} {annee}",
        convention=t.convention, inconnus=inconnus)


@bp.route("/salaries/pointage/confirmer", methods=["POST"])
@login_required
def salaries_pointage_confirmer():
    """Étape CONFIRMATION : crée/complète les bulletins BROUILLON de la période
    avec les heures ventilées. Ne touche jamais un bulletin existant ni une
    période clôturée. Les absences sont affichées, jamais déduites d'office."""
    if current_user.is_super_admin:
        return redirect(url_for("admin.admin_dashboard"))
    t = get_tenant()
    if not t:
        return redirect(url_for("auth.login"))
    if not current_user.can_edit:
        abort(403)

    stock = session.get("pointage_sal")
    if not stock or not stock.get("lignes"):
        flash("Session d'import expirée. Recommencez le téléversement.", "error")
        return redirect(url_for("tenant.salaries_pointage"))

    annee, mois = stock["annee"], stock["mois"]
    periode = PeriodePaie.query.filter_by(tenant_id=t.id, annee=annee, mois=mois).first()
    if not periode:
        flash("Aucune période de paie ouverte pour ce mois. Créez-la d'abord.", "error")
        return redirect(url_for("tenant.salaries_pointage"))
    if periode.statut not in ("OUVERT", "OUVERTE"):
        flash(f"La période {periode.libelle_mois} {periode.annee} est clôturée.", "error")
        return redirect(url_for("tenant.salaries_pointage"))

    import calendar as _cal
    fin_periode = date(annee, mois, _cal.monthrange(annee, mois)[1])

    deja = {b.salarie_id for b in
            BulletinPaie.query.filter_by(tenant_id=t.id, periode_id=periode.id).all()}

    crees = maj = ignores = 0
    for ligne in stock["lignes"]:
        s = Salarie.query.filter_by(id=ligne["salarie_id"], tenant_id=t.id).first()
        if not s:
            continue
        heures = {
            "heures_sup_10": ligne.get("heures_sup_10", 0),
            "heures_sup_30": ligne.get("heures_sup_30", 0),
            "heures_sup_30b": ligne.get("heures_sup_30b", 0),
            "heures_sup_40": ligne.get("heures_sup_40", 0),
            "heures_sup_70": ligne.get("heures_sup_70", 0),
        }
        contrat = (Contrat.query
                   .filter_by(salarie_id=s.id, tenant_id=t.id, actif=True)
                   .order_by(Contrat.date_debut.desc()).first())
        if not contrat or not contrat.salaire_base:
            ignores += 1
            continue

        if s.id in deja:
            # Bulletin existant : on met à jour SEULEMENT les heures sup, sans
            # écraser les autres ajustements déjà faits.
            b = BulletinPaie.query.filter_by(
                tenant_id=t.id, periode_id=periode.id, salarie_id=s.id).first()
            if b and b.statut == "BROUILLON":
                donnees = {"salaire_base": float(contrat.salaire_base), **heures,
                           "convention": t.convention}
                if s.date_embauche:
                    donnees["anciennete_annees"] = max(0, (fin_periode - s.date_embauche).days // 365)
                res = calculer_bulletin(donnees, nb_parts=float(s.nombre_parts or 1))
                for k, v in res.items():
                    if not k.startswith("_") and hasattr(b, k):
                        setattr(b, k, v)
                maj += 1
            else:
                ignores += 1
            continue

        # Nouveau brouillon
        donnees = {"salaire_base": float(contrat.salaire_base), **heures,
                   "convention": t.convention}
        if s.date_embauche:
            donnees["anciennete_annees"] = max(0, (fin_periode - s.date_embauche).days // 365)
        res = calculer_bulletin(donnees, nb_parts=float(s.nombre_parts or 1))
        b = BulletinPaie(tenant_id=t.id, salarie_id=s.id, periode_id=periode.id)
        for k, v in res.items():
            if not k.startswith("_") and hasattr(b, k):
                setattr(b, k, v)
        b.statut = "BROUILLON"
        b.mode_paiement = s.mode_paiement or "ESPECES"
        db.session.add(b)
        crees += 1

    db.session.commit()
    session.pop("pointage_sal", None)
    log_action("IMPORT_POINTAGE_SAL", "bulletin", periode.id,
               f"Import pointage salariés {periode.libelle_mois} {annee} : "
               f"{crees} créé(s), {maj} mis à jour")
    flash(f"Pointage importé : {crees} bulletin(s) créé(s), {maj} mis à jour. "
          f"Vérifiez et complétez les absences avant validation.", "success")
    if ignores:
        flash(f"{ignores} salarié(s) ignoré(s) (sans contrat actif, ou bulletin déjà validé).", "info")
    return redirect(url_for("tenant.bulletins", periode_id=periode.id))


@bp.route("/salaries/pointage")
@login_required
def salaries_pointage():
    """Page d'accueil de l'import : choix du mois, téléchargement du modèle,
    téléversement du fichier rempli."""
    if current_user.is_super_admin:
        return redirect(url_for("admin.admin_dashboard"))
    t = get_tenant()
    if not t:
        return redirect(url_for("auth.login"))
    annee, mois = _sal_periode_demandee()
    return render_template("tenant/salaries_pointage.html",
        annee=annee, mois=mois, tenant=t,
        mois_nom=f"{_MOIS_FR_SAL[mois]} {annee}",
        convention=t.convention)
