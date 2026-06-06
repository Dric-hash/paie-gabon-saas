"""
blueprints/admin.py — Super-Admin : gestion tenants, plans, stats, import
"""
import os
from datetime import datetime, timedelta

from flask import (Blueprint, render_template, request, redirect,
                   url_for, flash, jsonify, send_file, abort, session, current_app)
from flask_login import login_user, login_required, current_user
import io

from models import (db, Plan, Tenant, Utilisateur, CategorieEmploi, Salarie,
                    Contrat, PeriodePaie, BulletinPaie, RubriquePaie, Conge,
                    Acompte, Journalier, Pointage, FeuillePaieJournalier,
                    Site, AffectationSite, Paiement)
from calculs_paie import calculer_bulletin, calculer_masse_salariale
from audit import log_action
from core import super_admin_required, get_tenant, cache_delete

bp = Blueprint("admin", __name__)

@bp.route("/admin")
@super_admin_required
def admin_dashboard():
    from datetime import timedelta
    now     = datetime.utcnow()
    MOIS_FR = ["","Jan","Fév","Mar","Avr","Mai","Jun","Jul","Aoû","Sep","Oct","Nov","Déc"]

    # ── Tous les tenants (base de tout) ─────────────────────────────────────
    tenants      = Tenant.query.options(joinedload(Tenant.plan)).order_by(Tenant.date_inscription.desc()).all()
    nb_tenants   = len(tenants)
    nb_actifs    = sum(1 for t in tenants if t.statut == "ACTIF")
    nb_essai     = sum(1 for t in tenants if t.statut == "ESSAI")
    nb_suspendus = sum(1 for t in tenants if t.statut == "SUSPENDU")

    # ── MRR / ARR ────────────────────────────────────────────────────────────
    mrr = sum((float(t.plan.prix_mensuel) if t.plan else 0) for t in tenants if t.statut == "ACTIF")
    arr = round(mrr * 12)

    # ── Croissance MoM (mois en cours vs mois précédent) ────────────────────
    debut_mois_c = datetime(now.year, now.month, 1)
    debut_mois_p = (debut_mois_c.replace(day=1) - timedelta(days=1)).replace(day=1)
    inscrits_ce_mois  = sum(1 for t in tenants if t.date_inscription and t.date_inscription >= debut_mois_c)
    inscrits_mois_prec= sum(1 for t in tenants if t.date_inscription and debut_mois_p <= t.date_inscription < debut_mois_c)
    croissance_pct    = round(((inscrits_ce_mois - inscrits_mois_prec) / max(inscrits_mois_prec, 1)) * 100)

    # ── Churn rate (suspendus / total actifs + suspendus) ────────────────────
    churn_rate = round(nb_suspendus / max(nb_actifs + nb_suspendus, 1) * 100, 1)

    # ── ARPU (revenu moyen par client actif) ─────────────────────────────────
    arpu = round(mrr / max(nb_actifs, 1))

    # ── Taux conversion ──────────────────────────────────────────────────────
    taux_conversion = round((nb_actifs / max(nb_tenants, 1)) * 100)

    # ── Essais urgents (< 7 jours) ───────────────────────────────────────────
    essais_urgents = [
        t for t in tenants
        if t.statut == "ESSAI" and t.date_expiration and t.date_expiration <= now + timedelta(days=7)
    ]

    # ── Activité totale ──────────────────────────────────────────────────────
    total_sal = db.session.query(db.func.count(Salarie.id)).scalar() or 0
    total_bul = db.session.query(db.func.count(BulletinPaie.id)).scalar() or 0
    total_ptg = db.session.query(db.func.count(Pointage.id)).scalar() or 0

    # ── Score d'activité par tenant — UNE SEULE requête agrégée ─────────────
    # Avant : 4 requêtes × N tenants = N+1 massif
    # Maintenant : 3 requêtes GROUP BY quel que soit le nombre de tenants
    from sqlalchemy import func as sqlfunc

    # Salariés actifs par tenant
    sal_counts = dict(
        db.session.query(Salarie.tenant_id, sqlfunc.count(Salarie.id))
        .filter(Salarie.statut == "ACTIF")
        .group_by(Salarie.tenant_id)
        .all()
    )
    # Bulletins total par tenant
    bul_counts = dict(
        db.session.query(BulletinPaie.tenant_id, sqlfunc.count(BulletinPaie.id))
        .group_by(BulletinPaie.tenant_id)
        .all()
    )
    # Bulletins ce mois par tenant
    bul_mois_counts = dict(
        db.session.query(BulletinPaie.tenant_id, sqlfunc.count(BulletinPaie.id))
        .join(PeriodePaie, BulletinPaie.periode_id == PeriodePaie.id)
        .filter(PeriodePaie.annee == now.year, PeriodePaie.mois == now.month)
        .group_by(BulletinPaie.tenant_id)
        .all()
    )
    # Pointages ce mois par tenant
    ptg_mois_counts = dict(
        db.session.query(Pointage.tenant_id, sqlfunc.count(Pointage.id))
        .filter(Pointage.date_pointage >= debut_mois_c.date())
        .group_by(Pointage.tenant_id)
        .all()
    )

    sal_par_tenant   = sal_counts
    bul_par_tenant   = bul_counts
    score_par_tenant = {}
    for t in tenants:
        nb_s     = sal_counts.get(t.id, 0)
        nb_b_mois= bul_mois_counts.get(t.id, 0)
        nb_p_mois= ptg_mois_counts.get(t.id, 0)
        score = min(100, int(
            min(nb_b_mois / max(nb_s, 1) * 50, 50) +
            min(nb_p_mois / max(nb_s * 20, 1) * 30, 30) +
            min(nb_s / 10 * 20, 20)
        ))
        score_par_tenant[t.id] = score

    # ── Revenus 6 mois glissants + nb inscriptions ───────────────────────────
    revenus_6mois     = []
    inscrits_6mois    = []
    bul_6mois         = []
    mois_labels       = []
    for i in range(5, -1, -1):
        d     = now - timedelta(days=i * 30)
        debut = datetime(d.year, d.month, 1)
        fin   = (debut + timedelta(days=32)).replace(day=1)
        mois_labels.append(f"{MOIS_FR[d.month]} {str(d.year)[2:]}")
        rev = sum(
            (float(t.plan.prix_mensuel) if t.plan else 0)
            for t in tenants
            if t.statut == "ACTIF" and t.date_inscription and t.date_inscription <= fin
        )
        revenus_6mois.append(int(rev))
        inscrits_6mois.append(sum(
            1 for t in tenants if t.date_inscription and debut <= t.date_inscription < fin
        ))
        nb_bul = BulletinPaie.query.join(PeriodePaie).filter(
            PeriodePaie.annee == d.year, PeriodePaie.mois == d.month
        ).count()
        bul_6mois.append(nb_bul)

    # ── Répartition par plan ─────────────────────────────────────────────────
    plans_tous = Plan.query.filter_by(actif=True).all()
    repartition_plans = [
        {"nom": p.nom, "nb": sum(1 for t in tenants if t.plan_id == p.id and t.statut == "ACTIF"),
         "prix": float(p.prix_mensuel)}
        for p in plans_tous
        if sum(1 for t in tenants if t.plan_id == p.id and t.statut == "ACTIF") > 0
    ]

    # ── Top tenants les plus actifs ───────────────────────────────────────────
    top_tenants = sorted(
        [t for t in tenants if t.statut == "ACTIF"],
        key=lambda t: score_par_tenant.get(t.id, 0),
        reverse=True
    )[:5]

    # ── Nouvelles inscriptions ce mois ───────────────────────────────────────
    nouveaux_ce_mois = [t for t in tenants if t.date_inscription and t.date_inscription >= debut_mois_c]

    # ── Dernières actions audit (tous tenants) ───────────────────────────────
    from audit import get_audit_logs_admin
    derniers_logs = get_audit_logs_admin(limit=20)

    return render_template("admin/dashboard.html",
        tenants=tenants, nb_tenants=nb_tenants,
        nb_actifs=nb_actifs, nb_essai=nb_essai, nb_suspendus=nb_suspendus,
        mrr=mrr, arr=arr, arpu=arpu, churn_rate=churn_rate,
        croissance_pct=croissance_pct, inscrits_ce_mois=inscrits_ce_mois,
        taux_conversion=taux_conversion,
        total_sal=total_sal, total_bul=total_bul, total_ptg=total_ptg,
        essais_urgents=essais_urgents,
        revenus_6mois=revenus_6mois, inscrits_6mois=inscrits_6mois,
        mois_labels=mois_labels, repartition_plans=repartition_plans,
        bul_6mois=bul_6mois, bul_labels=mois_labels,
        sal_par_tenant=sal_par_tenant, bul_par_tenant=bul_par_tenant,
        score_par_tenant=score_par_tenant, top_tenants=top_tenants,
        nouveaux_ce_mois=nouveaux_ce_mois,
        bulletins_par_tenant=bul_par_tenant,
        derniers_logs=derniers_logs,
        now=now)

@bp.route("/admin/tenants")
@super_admin_required
def admin_tenants():
    q=request.args.get("q",""); statut=request.args.get("statut","")
    query=Tenant.query
    if q: query=query.filter(Tenant.denomination.ilike(f"%{q}%"))
    if statut: query=query.filter_by(statut=statut)
    return render_template("admin/tenants.html", tenants=query.order_by(Tenant.date_inscription.desc()).all(),
        plans=Plan.query.all(), q=q, statut=statut)

@bp.route("/admin/tenants/<int:id>")
@super_admin_required
def admin_tenant_detail(id):
    t = Tenant.query.get_or_404(id)
    # Dernières périodes pour l'historique activité
    periodes_recentes = PeriodePaie.query.filter_by(tenant_id=id)        .order_by(PeriodePaie.annee.desc(), PeriodePaie.mois.desc()).limit(6).all()
    nb_journaliers = Journalier.query.filter_by(tenant_id=id).count()
    return render_template("admin/tenant_detail.html", tenant=t,
        nb_salaries=Salarie.query.filter_by(tenant_id=id).count(),
        nb_bulletins=BulletinPaie.query.filter_by(tenant_id=id).count(),
        users=Utilisateur.query.filter_by(tenant_id=id).all(),
        plans=Plan.query.all(),
        periodes_recentes=periodes_recentes,
        nb_journaliers=nb_journaliers,
        now=datetime.utcnow())

@bp.route("/admin/tenants/<int:id>/statut", methods=["POST"])
@super_admin_required
def admin_tenant_statut(id):
    t=Tenant.query.get_or_404(id)
    t.statut=request.form.get("statut",t.statut)
    if request.form.get("plan_id"): t.plan_id=int(request.form["plan_id"])
    # Date d'expiration essai
    date_exp = request.form.get("date_expiration","").strip()
    if date_exp:
        try:
            from datetime import datetime as _dt
            t.date_expiration = _dt.strptime(date_exp, "%Y-%m-%d")
        except: pass
    elif request.form.get("clear_expiration"):
        t.date_expiration = None
    db.session.commit(); flash(f"{t.denomination} mis à jour.","success")
    return redirect(url_for("admin.admin_tenant_detail",id=id))

@bp.route("/admin/tenants/<int:id>/notes", methods=["POST"])
@super_admin_required
def admin_tenant_notes(id):
    t = Tenant.query.get_or_404(id)
    t.notes = request.form.get("notes", "").strip()
    db.session.commit()
    flash("Notes mises à jour.", "success")
    return redirect(url_for("admin.admin_tenant_detail", id=id))


@bp.route("/admin/tenants/<int:id>/regenerer-token", methods=["POST"])
@super_admin_required
def admin_regenerer_token(id):
    """Génère ou régénère le token API d'un tenant."""
    t = Tenant.query.get_or_404(id)
    ancien = t.token_api
    t.token_api = sec.token_hex(32)
    db.session.commit()
    action = "généré" if not ancien else "régénéré"
    flash(f"Token API {action} pour {t.denomination}. "
          f"Transmettez-le de façon sécurisée au développeur RH du client.", "success")
    logger.info(f"[SuperAdmin] Token API {action} — tenant={t.id} par {current_user.email}")
    return redirect(url_for("admin.admin_tenant_detail", id=id))

@bp.route("/admin/tenants/<int:id>/impersonate")
@super_admin_required
def admin_impersonate(id):
    u=Utilisateur.query.filter_by(tenant_id=id,role="TENANT_ADMIN").first()
    if not u: flash("Aucun admin trouvé.","error"); return redirect(url_for("admin.admin_tenant_detail",id=id))
    logout_user(); login_user(u)
    flash(f"Connecté en tant que {u.nom_complet} ({u.tenant.denomination})","warning")
    return redirect(url_for("tenant.dashboard"))

@bp.route("/admin/plans", methods=["GET","POST"])
@super_admin_required
def admin_plans():
    if request.method=="POST":
        p=Plan(code=request.form["code"].strip().upper(),
               nom=request.form["nom"].strip(),
               prix_mensuel=float(request.form.get("prix_mensuel",0)),
               max_salaries=int(request.form["max_salaries"]) if request.form.get("max_salaries") else None,
               max_utilisateurs=int(request.form["max_utilisateurs"]) if request.form.get("max_utilisateurs") else None,
               description=request.form.get("description",""))
        db.session.add(p); db.session.commit(); flash("Plan créé avec succès.","success")
    plans = Plan.query.order_by(Plan.prix_mensuel.asc()).all()
    # Nb clients actifs + revenus par plan
    stats_plans = {}
    for p in plans:
        nb_actifs  = Tenant.query.filter_by(plan_id=p.id, statut="ACTIF").count()
        nb_essai   = Tenant.query.filter_by(plan_id=p.id, statut="ESSAI").count()
        nb_total   = Tenant.query.filter_by(plan_id=p.id).count()
        revenus    = nb_actifs * float(p.prix_mensuel)
        stats_plans[p.id] = {"nb_actifs": nb_actifs, "nb_essai": nb_essai,
                              "nb_total": nb_total, "revenus": revenus}
    return render_template("admin/plans.html", plans=plans, stats_plans=stats_plans)

@bp.route("/admin/plans/<int:id>/modifier", methods=["POST"])
@super_admin_required
def admin_plan_modifier(id):
    p = Plan.query.get_or_404(id)
    p.nom          = request.form.get("nom", p.nom).strip()
    p.prix_mensuel = float(request.form.get("prix_mensuel", p.prix_mensuel))
    p.max_salaries     = int(request.form["max_salaries"]) if request.form.get("max_salaries") else None
    p.max_utilisateurs = int(request.form["max_utilisateurs"]) if request.form.get("max_utilisateurs") else None
    p.description  = request.form.get("description", p.description or "")
    db.session.commit()
    flash(f"Plan « {p.nom} » modifié avec succès.", "success")
    return redirect(url_for("admin.admin_plans"))

@bp.route("/admin/plans/<int:id>/toggle", methods=["POST"])
@super_admin_required
def admin_plan_toggle(id):
    p = Plan.query.get_or_404(id)
    p.actif = not p.actif
    db.session.commit()
    etat = "activé" if p.actif else "désactivé"
    flash(f"Plan « {p.nom} » {etat}.", "success")
    return redirect(url_for("admin.admin_plans"))

@bp.route("/admin/plans/<int:id>/supprimer", methods=["POST"])
@super_admin_required
def admin_plan_supprimer(id):
    p = Plan.query.get_or_404(id)
    nb_clients = Tenant.query.filter_by(plan_id=id).count()
    if nb_clients > 0:
        flash(f"Impossible de supprimer : {nb_clients} entreprise(s) utilisent ce plan.", "error")
        return redirect(url_for("admin.admin_plans"))
    nom = p.nom
    db.session.delete(p); db.session.commit()
    flash(f"Plan « {nom} » supprimé.", "success")
    return redirect(url_for("admin.admin_plans"))

@bp.route("/admin/stats")
@super_admin_required
def admin_stats():
    tenants = Tenant.query.order_by(Tenant.denomination).all()
    stats = []
    total_revenus  = 0
    total_salaries = 0
    total_bulletins= 0
    total_journaliers = 0
    nb_actif = nb_essai = nb_suspendu = nb_attente = 0

    for t in tenants:
        nb_sal = Salarie.query.filter_by(tenant_id=t.id, statut="ACTIF").count()
        nb_bul = BulletinPaie.query.filter_by(tenant_id=t.id).count()
        nb_per = PeriodePaie.query.filter_by(tenant_id=t.id).count()
        nb_jou = Journalier.query.filter_by(tenant_id=t.id).count()
        rev    = float(t.plan.prix_mensuel) if t.plan and t.statut == "ACTIF" else 0
        total_revenus   += rev
        total_salaries  += nb_sal
        total_bulletins += nb_bul
        total_journaliers += nb_jou
        if   t.statut == "ACTIF":                nb_actif   += 1
        elif t.statut == "ESSAI":                nb_essai   += 1
        elif t.statut == "SUSPENDU":             nb_suspendu+= 1
        elif t.statut == "PAIEMENT_EN_ATTENTE":  nb_attente += 1
        stats.append({
            "tenant": t, "nb_salaries": nb_sal, "nb_bulletins": nb_bul,
            "nb_periodes": nb_per, "nb_journaliers": nb_jou, "revenus": rev
        })

    # Top 8 pour graphique — trié en Python
    top8 = sorted(stats, key=lambda x: x["nb_bulletins"], reverse=True)[:8]
    top8_labels = [s["tenant"].denomination[:18] for s in top8]
    top8_data   = [s["nb_bulletins"] for s in top8]

    taux_conv = round(nb_actif / len(tenants) * 100) if tenants else 0

    return render_template("admin/stats.html",
        stats=stats,
        total_revenus=total_revenus,
        total_salaries=total_salaries,
        total_bulletins=total_bulletins,
        total_journaliers=total_journaliers,
        nb_actifs=nb_actif,
        nb_essai=nb_essai,
        nb_suspendu=nb_suspendu,
        nb_attente=nb_attente,
        taux_conv=taux_conv,
        top8_labels=top8_labels,
        top8_data=top8_data,
        now=datetime.utcnow())

@bp.route("/admin/import", methods=["GET","POST"])
@login_required
@super_admin_required
def admin_import_excel():
    from werkzeug.utils import secure_filename
    tenants = Tenant.query.order_by(Tenant.denomination).all()
    resultats = None
    if request.method == "POST":
        tenant_id = request.form.get("tenant_id", type=int)
        tenant = Tenant.query.get_or_404(tenant_id)
        f = request.files.get("excel_file")
        if not f or not f.filename.endswith(".xlsx"):
            flash("Fichier invalide. Utilisez un fichier .xlsx", "error")
            return render_template("admin/import_excel.html", tenants=tenants)
        imp_societe  = "import_societe"  in request.form
        imp_salaries = "import_salaries" in request.form
        imp_bulletins= "import_bulletins"in request.form
        ecraser      = "ecraser"         in request.form
        try:
            from openpyxl import load_workbook
            from datetime import datetime as dt2, date as d2
            wb = load_workbook(io.BytesIO(f.read()), read_only=True, data_only=True)
            nb_sal=nb_bul=nb_per=nb_err=0
            def nv(v,d=0):
                try: return float(v) if v is not None else d
                except: return d
            def dv(v):
                if v is None: return None
                if isinstance(v,dt2): return v.date()
                if isinstance(v,d2): return v
                try: return dt2.strptime(str(v)[:10],"%Y-%m-%d").date()
                except: return None
            if imp_societe and "INFOS SOCIETE" in wb.sheetnames:
                ws=wb["INFOS SOCIETE"]; infos={}
                for row in ws.iter_rows(values_only=True):
                    if row[1] and row[2]: infos[str(row[1]).strip().upper()]=row[2]
                tenant.denomination=str(infos.get("DENOMINATION SOCIALE",tenant.denomination)).strip().upper()
                tenant.sigle=str(infos.get("SIGLE",tenant.sigle or "")).strip()
                tenant.activite=str(infos.get("ACTIVITE",tenant.activite or "")).strip()
                tenant.nif=str(infos.get("NIF",tenant.nif or "")).strip()
                tenant.adresse=str(infos.get("ADRESSE",tenant.adresse or "")).strip()
                tenant.ville=str(infos.get("VILLE",tenant.ville or "Libreville")).strip()
            cats={}
            for code,lib in [("C1","Ouvriers"),("C2","Techniciens"),("C3","Conducteurs"),("C4","Cadres")]:
                cat=CategorieEmploi.query.filter_by(tenant_id=tenant.id,code=code).first()
                if not cat:
                    cat=CategorieEmploi(tenant_id=tenant.id,code=code,libelle=lib)
                    db.session.add(cat); db.session.flush()
                cats[code]=cat
            salaries_map={}
            if imp_salaries and "INFOS SALARIES" in wb.sheetnames:
                ws=wb["INFOS SALARIES"]; header=None
                for row in ws.iter_rows(values_only=True):
                    if not any(v is not None for v in row): continue
                    if header is None: header=row; continue
                    if row[0] is None: continue
                    matricule=str(row[0]).strip().upper()
                    if not matricule: continue
                    sal=Salarie.query.filter_by(tenant_id=tenant.id,matricule=matricule).first()
                    if sal and not ecraser: salaries_map[matricule]=sal; continue
                    if not sal: sal=Salarie(tenant_id=tenant.id,matricule=matricule); db.session.add(sal)
                    sal.nom=str(row[1]).strip().upper() if row[1] else "—"
                    sal.prenom=str(row[2]).strip() if row[2] else "—"
                    sal.telephone=str(row[3]).strip() if row[3] else None
                    sal.nationalite=str(row[5]).strip().upper() if row[5] else "GABONAISE"
                    sal.sexe=str(row[6]).strip().upper() if row[6] else "M"
                    sal.date_naissance=dv(row[7])
                    sal.date_embauche=dv(row[9]) or d2(2024,8,1)
                    sal.date_cessation=dv(row[10])
                    sal.situation_matrimoniale=str(row[11]).strip().upper() if row[11] else None
                    sal.nb_enfants=int(row[12]) if row[12] and str(row[12]).replace('.','').isdigit() else 0
                    sal.nombre_parts=float(row[13]) if row[13] else 1.0
                    sal.numero_cnss=str(row[14]).strip() if row[14] else None
                    sal.numero_cnamgs=str(row[15]).strip() if row[15] else None
                    sal.emploi=str(row[16]).strip().upper() if row[16] else None
                    sal.nb_enfants_moins_16ans=int(row[18]) if row[18] and str(row[18]).replace('.','').isdigit() else 0
                    sal.assujetti_cnamgs=str(row[19]).strip().upper()=="OUI" if row[19] else True
                    sal.statut="INACTIF" if dv(row[10]) else "ACTIF"
                    cat_code=str(row[17]).strip().upper() if row[17] else "C1"
                    sal.categorie_id=cats.get(cat_code,cats.get("C1")).id
                    db.session.flush(); salaries_map[matricule]=sal; nb_sal+=1
            if not imp_salaries:
                for s in Salarie.query.filter_by(tenant_id=tenant.id).all():
                    salaries_map[s.matricule]=s
            MOIS={"JANVIER":1,"FÉVRIER":2,"FEVRIER":2,"MARS":3,"AVRIL":4,"MAI":5,"JUIN":6,
                  "JUILLET":7,"AOÛT":8,"AOUT":8,"SEPTEMBRE":9,"OCTOBRE":10,"NOVEMBRE":11,"DÉCEMBRE":12,"DECEMBRE":12}
            periodes_cache={}
            if imp_bulletins and "DONNEES DU BULLETIN" in wb.sheetnames:
                ws=wb["DONNEES DU BULLETIN"]
                for i,row in enumerate(ws.iter_rows(values_only=True)):
                    if i==0: continue
                    if not any(v is not None for v in row): continue
                    if row[4] is None: continue
                    matricule=str(row[4]).strip().upper().replace(' - ','-').replace('- ','-').replace(' -','-')
                    if matricule not in salaries_map: nb_err+=1; continue
                    annee=int(row[1]) if row[1] else None
                    mois_str=str(row[2]).strip().upper() if row[2] else ""
                    mois=MOIS.get(mois_str)
                    if not annee or not mois: nb_err+=1; continue
                    pk=f"{annee}-{mois}"
                    if pk not in periodes_cache:
                        p=PeriodePaie.query.filter_by(tenant_id=tenant.id,annee=annee,mois=mois).first()
                        if not p:
                            mois_nom=[k for k,v in MOIS.items() if v==mois and len(k)>4][0]
                            p=PeriodePaie(tenant_id=tenant.id,annee=annee,mois=mois,
                                libelle_mois=mois_nom,trimestre=f"T{(mois-1)//3+1}",statut="CLÔTURÉ")
                            db.session.add(p); db.session.flush(); nb_per+=1
                        periodes_cache[pk]=p
                    sal=salaries_map[matricule]
                    bul=BulletinPaie.query.filter_by(tenant_id=tenant.id,salarie_id=sal.id,periode_id=periodes_cache[pk].id).first()
                    if bul and not ecraser: continue
                    if not bul: bul=BulletinPaie(tenant_id=tenant.id,salarie_id=sal.id,periode_id=periodes_cache[pk].id); db.session.add(bul)
                    bul.nb_jours_travailles=int(nv(row[42]))
                    bul.salaire_base=nv(row[5]); bul.heures_sup_10=nv(row[7]); bul.heures_sup_30=nv(row[9])
                    bul.heures_sup_40=nv(row[11]); bul.heures_sup_70=nv(row[13]); bul.absences=nv(row[15])
                    bul.sursalaire=nv(row[17]); bul.prime_caisse=nv(row[19]); bul.carburant=nv(row[21])
                    bul.prime_anciennete=nv(row[23]); bul.indem_logement=nv(row[25])
                    bul.indem_domesticite=nv(row[26]); bul.indem_eau_electricite=nv(row[27])
                    bul.indem_nourriture=nv(row[28]); bul.prime_rendement=nv(row[29])
                    bul.prime_assiduité=nv(row[31]); bul.prime_qualite=nv(row[33])
                    bul.prime_performance=nv(row[35]); bul.prime_transport=nv(row[37])
                    bul.prime_responsabilite=nv(row[39]); bul.allocations_conge=nv(row[41])
                    bul.salaire_brut=nv(row[53]); bul.base_cnss=nv(row[54])
                    bul.cnss_salarie=nv(row[55]); bul.cnss_patronale=nv(row[56])
                    bul.base_cnamgs=nv(row[59]); bul.cnamgs_salarie=nv(row[60]); bul.cnamgs_patronale=nv(row[61])
                    bul.fnh=nv(row[62]); bul.cfp=nv(row[63]); bul.base_tcs=nv(row[72]); bul.tcs=nv(row[73])
                    bul.net_avant_irpp=nv(row[74]); bul.base_irpp=nv(row[75]); bul.irpp=nv(row[76])
                    bul.salaire_net=nv(row[77]); bul.prime_panier=nv(row[78]); bul.indem_transport=nv(row[79])
                    bul.indem_representation=nv(row[80]); bul.prime_salisure=nv(row[81])
                    bul.acompte=nv(row[82]); bul.net_a_payer=nv(row[83]) if len(row)>83 else 0
                    bul.statut="VALIDÉ"; bul.date_validation=datetime.utcnow(); nb_bul+=1
            db.session.commit()
            resultats={"nb_salaries":nb_sal,"nb_bulletins":nb_bul,"nb_periodes":nb_per,"erreurs":nb_err}
            flash(f"Import réussi ! {nb_sal} salariés, {nb_bul} bulletins, {nb_per} périodes.","success")
        except Exception as e:
            db.session.rollback()
            flash(f"Erreur : {str(e)}","error")
    return render_template("admin/import_excel.html", tenants=tenants, resultats=resultats)

@bp.route("/admin/update-taux", methods=["POST"])
@super_admin_required
def admin_update_taux():
    taux = {"CNSS":(0.05,0.18),"CNAMGS":(0.02,0.041),"FNH":(0.0,0.03),"TCS":(0.05,0.0),"CFP":(0.0,0.005)}
    nb = 0
    for code,(sal,pat) in taux.items():
        for r in RubriquePaie.query.filter_by(code=code).all():
            r.taux_salarie=sal; r.taux_patronal=pat; nb+=1
    db.session.commit()
    flash(f"Taux mis a jour ({nb} rubriques).", "success")
    return redirect(url_for("admin.admin_rubriques"))

@bp.route("/admin/rubriques", methods=["GET","POST"])
@super_admin_required
def admin_rubriques():
    if request.method=="POST":
        r=RubriquePaie(code=request.form["code"].strip().upper(),
            libelle=request.form["libelle"].strip(), type=request.form.get("type","COTISATION"),
            taux_salarie=float(request.form["taux_salarie"]) if request.form.get("taux_salarie") else None,
            taux_patronal=float(request.form["taux_patronal"]) if request.form.get("taux_patronal") else None,
            plafond_mensuel=float(request.form["plafond_mensuel"]) if request.form.get("plafond_mensuel") else None)
        db.session.add(r); db.session.commit(); flash("Rubrique créée.","success")
    return render_template("admin/rubriques.html", rubriques=RubriquePaie.query.all())

@bp.route("/admin/tenants/<int:id>/supprimer", methods=["POST"])
@super_admin_required
def admin_tenant_supprimer(id):
    t = Tenant.query.get_or_404(id)
    nom = t.denomination
    try:
        # ── 1. Pointages journaliers (FK bloquante) ───────────────────────
        for j in Journalier.query.filter_by(tenant_id=id).all():
            Pointage.query.filter_by(journalier_id=j.id).delete()
            FeuillePaieJournalier.query.filter_by(journalier_id=j.id).delete()
            AffectationSite.query.filter_by(journalier_id=j.id).delete()
        Journalier.query.filter_by(tenant_id=id).delete()

        # ── 2. Données salariés ────────────────────────────────────────────
        for s in Salarie.query.filter_by(tenant_id=id).all():
            BulletinPaie.query.filter_by(salarie_id=s.id).delete()
            Contrat.query.filter_by(salarie_id=s.id).delete()
            Pointage.query.filter_by(salarie_id=s.id).delete()
            Acompte.query.filter_by(salarie_id=s.id).delete()
            Conge.query.filter_by(salarie_id=s.id).delete()
            AffectationSite.query.filter_by(salarie_id=s.id).delete()
        Salarie.query.filter_by(tenant_id=id).delete()

        # ── 3. Reste du tenant ─────────────────────────────────────────────
        AffectationSite.query.filter_by(tenant_id=id).delete()
        from models import Site
        Site.query.filter_by(tenant_id=id).delete()
        PeriodePaie.query.filter_by(tenant_id=id).delete()
        CategorieEmploi.query.filter_by(tenant_id=id).delete()
        Acompte.query.filter_by(tenant_id=id).delete()
        Conge.query.filter_by(tenant_id=id).delete()
        Utilisateur.query.filter_by(tenant_id=id).delete()

        db.session.delete(t)
        db.session.commit()
        flash(f"Entreprise « {nom} » supprimée définitivement.", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"Erreur: {str(e)}", "error")
    return redirect(url_for("admin.admin_tenants"))

