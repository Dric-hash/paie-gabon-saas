"""
migration_import.py — Script générique d'import Excel vers PaieGabon SaaS
=========================================================================
Usage :
    python migration_import.py --excel MON_FICHIER.xlsx \
                               --email admin@monentreprise.ga \
                               --password MonMotDePasse2026! \
                               [--slug mon_entreprise]

Ce script importe pour N'IMPORTE QUELLE SOCIÉTÉ :
  1. Les infos société  (feuille "INFOS SOCIETE")
  2. Les salariés       (feuille "INFOS SALARIES")
  3. Les bulletins      (feuille "DONNEES DU BULLETIN")

Si le tenant existe déjà (même slug), il est mis à jour sans doublons.
"""

import os, sys, argparse
from datetime import datetime, date
from openpyxl import load_workbook

# ── ARGUMENTS EN LIGNE DE COMMANDE ────────────────────────────────────────────
parser = argparse.ArgumentParser(description="Import Excel générique vers PaieGabon")
parser.add_argument("--excel",    required=True, help="Chemin vers le fichier .xlsx")
parser.add_argument("--email",    required=True, help="Email de l'admin de l'entreprise")
parser.add_argument("--password", required=True, help="Mot de passe de l'admin")
parser.add_argument("--slug",     default=None,  help="Slug unique (généré automatiquement si absent)")
parser.add_argument("--plan",     default="PRO", help="Code du plan (STARTER, PRO, CABINET)")
args = parser.parse_args()

EXCEL_FILE     = args.excel
ADMIN_EMAIL    = args.email.strip().lower()
ADMIN_PASSWORD = args.password
PLAN_CODE      = args.plan.upper()

# ── INITIALISATION FLASK ───────────────────────────────────────────────────────
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from app import app
from models import db, Plan, Tenant, Utilisateur, CategorieEmploi, \
                   Salarie, Contrat, PeriodePaie, BulletinPaie
import secrets as sec_mod

# ── HELPERS ───────────────────────────────────────────────────────────────────
def n(val, default=0):
    try:
        if val is None: return default
        return float(val)
    except: return default

def d(val):
    if val is None: return None
    if isinstance(val, datetime): return val.date()
    if isinstance(val, date): return val
    try: return datetime.strptime(str(val)[:10], "%Y-%m-%d").date()
    except: return None

MOIS_NOMS = {
    "JANVIER":1,"FÉVRIER":2,"FEVRIER":2,"MARS":3,"AVRIL":4,"MAI":5,"JUIN":6,
    "JUILLET":7,"AOÛT":8,"AOUT":8,"SEPTEMBRE":9,"OCTOBRE":10,"NOVEMBRE":11,
    "DÉCEMBRE":12,"DECEMBRE":12
}

def migrer():
    print("=" * 65)
    print("  IMPORT EXCEL → PAIEGALON SAAS  (script générique)")
    print("=" * 65)

    if not os.path.exists(EXCEL_FILE):
        print(f"❌ Fichier introuvable : {EXCEL_FILE}")
        sys.exit(1)

    print(f"\n📂 Fichier : {EXCEL_FILE}")
    wb = load_workbook(EXCEL_FILE, read_only=True, data_only=True)

    with app.app_context():

        # ── ÉTAPE 1 : TENANT ──────────────────────────────────────────────────
        print("\n🏢 Étape 1 — Lecture des infos société...")

        # Lire feuille INFOS SOCIETE
        infos = {}
        if "INFOS SOCIETE" in wb.sheetnames:
            ws_soc = wb["INFOS SOCIETE"]
            for row in ws_soc.iter_rows(values_only=True):
                if row[1] and row[2]:
                    infos[str(row[1]).strip().upper()] = str(row[2]).strip()

        denomination = infos.get("DENOMINATION SOCIALE", infos.get("DENOMINATION", "NOUVELLE ENTREPRISE")).upper()

        # Générer le slug
        slug_base = (args.slug or denomination.lower()
                     .replace(" ", "_").replace("'", "").replace("-", "_"))[:30]
        slug = slug_base
        i = 1
        # Vérifier unicité sauf si tenant existe déjà
        while True:
            existing_slug = Tenant.query.filter_by(slug=slug).first()
            if not existing_slug:
                break
            # Si même dénomination → c'est probablement le même tenant
            if existing_slug.denomination == denomination:
                break
            slug = f"{slug_base}_{i}"; i += 1

        tenant = Tenant.query.filter_by(slug=slug).first()
        if tenant:
            print(f"   ℹ️  Tenant existant trouvé : {tenant.denomination} (id={tenant.id}) — mise à jour")
        else:
            plan = Plan.query.filter_by(code=PLAN_CODE).first() or Plan.query.first()
            tenant = Tenant(slug=slug, statut="ACTIF", plan_id=plan.id if plan else None)
            db.session.add(tenant)
            print(f"   ➕ Nouveau tenant créé : slug={slug}")

        tenant.denomination  = denomination
        tenant.sigle         = infos.get("SIGLE", denomination[:5])
        tenant.activite      = infos.get("ACTIVITE", infos.get("SECTEUR", ""))
        tenant.nif           = infos.get("NIF", "")
        tenant.adresse       = infos.get("ADRESSE", "")
        tenant.boite_postale = infos.get("BOÎTE POSTALE", infos.get("BOITE POSTALE", ""))
        tenant.ville         = infos.get("VILLE", "Libreville")
        tenant.region        = infos.get("REGION", "ESTUAIRE")
        tenant.pays          = infos.get("PAYS", "Gabon")
        if not tenant.token_api:
            tenant.token_api = sec_mod.token_hex(32)

        db.session.flush()
        print(f"   ✅ Tenant : {tenant.denomination} (id={tenant.id})")

        # ── ÉTAPE 2 : ADMIN ───────────────────────────────────────────────────
        print(f"\n👤 Étape 2 — Compte administrateur : {ADMIN_EMAIL}")
        admin = Utilisateur.query.filter_by(email=ADMIN_EMAIL).first()
        if not admin:
            nom_parts = ADMIN_EMAIL.split("@")[0].upper()
            admin = Utilisateur(
                nom=nom_parts, prenom="Admin",
                email=ADMIN_EMAIL, role="TENANT_ADMIN",
                tenant_id=tenant.id, actif=True)
            admin.set_password(ADMIN_PASSWORD)
            db.session.add(admin)
            print(f"   ✅ Compte créé")
        else:
            admin.tenant_id = tenant.id
            admin.set_password(ADMIN_PASSWORD)
            print(f"   ℹ️  Compte existant mis à jour")

        # ── ÉTAPE 3 : CATÉGORIES ──────────────────────────────────────────────
        print("\n📋 Étape 3 — Catégories d'emploi...")
        cats = {}
        for code, libelle in [
            ("C1", "Ouvriers / Conducteurs / Chauffeurs"),
            ("C2", "Techniciens"),
            ("C3", "Conducteurs de Travaux"),
            ("C4", "Cadres / Responsables / Gérants"),
        ]:
            cat = CategorieEmploi.query.filter_by(tenant_id=tenant.id, code=code).first()
            if not cat:
                cat = CategorieEmploi(tenant_id=tenant.id, code=code, libelle=libelle)
                db.session.add(cat)
                db.session.flush()
            cats[code] = cat
        print(f"   ✅ 4 catégories OK")

        # ── ÉTAPE 4 : SALARIÉS ────────────────────────────────────────────────
        nb_sal = 0
        salaries_map = {}

        if "INFOS SALARIES" in wb.sheetnames:
            print("\n👷 Étape 4 — Import des salariés...")
            ws_sal = wb["INFOS SALARIES"]
            header_sal = None

            for row in ws_sal.iter_rows(values_only=True):
                if not any(v is not None for v in row): continue
                if header_sal is None: header_sal = row; continue
                if row[0] is None: continue

                matricule = str(row[0]).strip().upper()
                if not matricule: continue

                sal = Salarie.query.filter_by(tenant_id=tenant.id, matricule=matricule).first()
                if not sal:
                    sal = Salarie(tenant_id=tenant.id, matricule=matricule)
                    db.session.add(sal)

                sal.nom                    = str(row[1]).strip().upper() if row[1] else "—"
                sal.prenom                 = str(row[2]).strip() if row[2] else "—"
                sal.telephone              = str(row[3]).strip() if row[3] else None
                sal.adresse                = str(row[4]).strip() if row[4] else None
                sal.nationalite            = str(row[5]).strip().upper() if row[5] else "GABONAISE"
                sal.sexe                   = str(row[6]).strip().upper() if row[6] else "M"
                sal.date_naissance         = d(row[7])
                sal.date_embauche          = d(row[9]) or date(2024, 1, 1)
                sal.date_cessation         = d(row[10])
                sal.situation_matrimoniale = str(row[11]).strip().upper() if row[11] else None
                sal.nb_enfants             = int(row[12]) if row[12] and str(row[12]).replace('.','').isdigit() else 0
                sal.nombre_parts           = float(row[13]) if row[13] else 1.0
                sal.numero_cnss            = str(row[14]).strip() if row[14] else None
                sal.numero_cnamgs          = str(row[15]).strip() if row[15] else None
                sal.emploi                 = str(row[16]).strip().upper() if row[16] else None
                sal.nb_enfants_moins_16ans = int(row[18]) if row[18] and str(row[18]).replace('.','').isdigit() else 0
                sal.assujetti_cnamgs       = str(row[19]).strip().upper() == "OUI" if row[19] else True
                sal.statut                 = "INACTIF" if d(row[10]) else "ACTIF"

                cat_code = str(row[17]).strip().upper() if row[17] else "C1"
                sal.categorie_id = cats.get(cat_code, cats.get("C1")).id

                db.session.flush()
                salaries_map[matricule] = sal
                nb_sal += 1
                statut_icon = "✅" if sal.statut == "ACTIF" else "⚪"
                print(f"   {statut_icon} {matricule} — {sal.nom} {sal.prenom}")

            print(f"\n   📊 {nb_sal} salariés traités")
        else:
            print("\n   ⚠️  Feuille 'INFOS SALARIES' absente — import salariés ignoré")
            for s in Salarie.query.filter_by(tenant_id=tenant.id).all():
                salaries_map[s.matricule] = s

        # ── ÉTAPE 5 : BULLETINS ───────────────────────────────────────────────
        nb_bul = 0
        nb_skip = 0
        periodes_cache = {}

        if "DONNEES DU BULLETIN" in wb.sheetnames:
            print("\n📄 Étape 5 — Import des bulletins de paie...")
            ws_bul = wb["DONNEES DU BULLETIN"]

            for i, row in enumerate(ws_bul.iter_rows(values_only=True)):
                if i == 0: continue
                if not any(v is not None for v in row): continue
                if row[4] is None: continue

                matricule = str(row[4]).strip().upper().replace(' - ','-').replace('- ','-').replace(' -','-')
                if matricule not in salaries_map:
                    nb_skip += 1; continue

                annee = int(row[1]) if row[1] else None
                mois_str = str(row[2]).strip().upper() if row[2] else ""
                mois = MOIS_NOMS.get(mois_str)
                if not annee or not mois:
                    nb_skip += 1; continue

                periode_key = f"{annee}-{mois}"
                if periode_key not in periodes_cache:
                    p = PeriodePaie.query.filter_by(tenant_id=tenant.id, annee=annee, mois=mois).first()
                    if not p:
                        mois_nom = [k for k,v in MOIS_NOMS.items() if v == mois and len(k) > 4][0]
                        p = PeriodePaie(tenant_id=tenant.id, annee=annee, mois=mois,
                                        libelle_mois=mois_nom, trimestre=f"T{(mois-1)//3+1}",
                                        statut="CLÔTURÉ")
                        db.session.add(p); db.session.flush()
                    periodes_cache[periode_key] = p

                sal = salaries_map[matricule]
                bul = BulletinPaie.query.filter_by(
                    tenant_id=tenant.id, salarie_id=sal.id,
                    periode_id=periodes_cache[periode_key].id).first()
                if not bul:
                    bul = BulletinPaie(tenant_id=tenant.id, salarie_id=sal.id,
                                       periode_id=periodes_cache[periode_key].id)
                    db.session.add(bul)

                bul.nb_jours_travailles   = int(n(row[42]))
                bul.salaire_base          = n(row[5])
                bul.heures_sup_10         = n(row[7])
                bul.heures_sup_30         = n(row[9])
                bul.heures_sup_40         = n(row[11])
                bul.heures_sup_70         = n(row[13])
                bul.absences              = n(row[15])
                bul.sursalaire            = n(row[17])
                bul.prime_caisse          = n(row[19])
                bul.carburant             = n(row[21])
                bul.prime_anciennete      = n(row[23])
                bul.indem_logement        = n(row[25])
                bul.indem_domesticite     = n(row[26])
                bul.indem_eau_electricite = n(row[27])
                bul.indem_nourriture      = n(row[28])
                bul.prime_rendement       = n(row[29])
                bul.prime_assiduité       = n(row[31])
                bul.prime_qualite         = n(row[33])
                bul.prime_performance     = n(row[35])
                bul.prime_transport       = n(row[37])
                bul.prime_responsabilite  = n(row[39])
                bul.allocations_conge     = n(row[41])
                bul.salaire_brut          = n(row[53])
                bul.base_cnss             = n(row[54])
                bul.cnss_salarie          = n(row[55])
                bul.cnss_patronale        = n(row[56])
                bul.base_cnamgs           = n(row[59])
                bul.cnamgs_salarie        = n(row[60])
                bul.cnamgs_patronale      = n(row[61])
                bul.fnh                   = n(row[62])
                bul.cfp                   = n(row[63])
                bul.base_tcs              = n(row[72])
                bul.tcs                   = n(row[73])
                bul.net_avant_irpp        = n(row[74])
                bul.base_irpp             = n(row[75])
                bul.irpp                  = n(row[76])
                bul.salaire_net           = n(row[77])
                bul.prime_panier          = n(row[78])
                bul.indem_transport       = n(row[79])
                bul.indem_representation  = n(row[80])
                bul.prime_salisure        = n(row[81])
                bul.acompte               = n(row[82])
                bul.net_a_payer           = n(row[83]) if len(row) > 83 else 0
                bul.statut                = "VALIDÉ"
                bul.date_validation       = datetime.utcnow()
                nb_bul += 1

            print(f"   📊 {nb_bul} bulletins importés · {nb_skip} ignorés")
        else:
            print("\n   ⚠️  Feuille 'DONNEES DU BULLETIN' absente — import bulletins ignoré")

        # ── COMMIT ────────────────────────────────────────────────────────────
        print("\n💾 Enregistrement en base de données...")
        db.session.commit()

        # ── RÉSUMÉ ────────────────────────────────────────────────────────────
        print("\n" + "=" * 65)
        print("  ✅ IMPORT TERMINÉ AVEC SUCCÈS !")
        print("=" * 65)
        print(f"\n  Entreprise   : {tenant.denomination}")
        print(f"  Slug         : {tenant.slug}")
        print(f"  Salariés     : {nb_sal}")
        print(f"  Bulletins    : {nb_bul}")
        print(f"  Périodes     : {len(periodes_cache)}")
        print(f"\n  Connexion    : {ADMIN_EMAIL}")
        print(f"  Mot de passe : {ADMIN_PASSWORD}")
        print(f"\n  URL : https://ameriack-paie.up.railway.app")
        print("=" * 65)

if __name__ == "__main__":
    migrer()
