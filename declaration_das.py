# -*- coding: utf-8 -*-
"""
declaration_das.py — Déclaration Annuelle des Salaires (DAS) — Gabon (DGI/DGCC).

Génère le fichier Excel de la DAS à partir des bulletins de paie validés d'un
exercice. Réservé à l'abonnement Cabinet (100 000 FCFA) côté routes.

Stratégie : calcul des cumuls annuels en Python et écriture de VALEURS via
openpyxl (aucune dépendance à LibreOffice au runtime — comme l'export CNSS).
La structure des colonnes reproduit fidèlement le fichier officiel
(onglets "source_salaires" / ID19 / ID20).

── Ventilation imposable / non imposable (cohérente avec calculs_paie) ───────────
  salaire_brut inclut déjà : salaire de base, heures sup, primes, avantages en
  nature (logement/eau-élec/domesticité/nourriture), allocations de congé et
  indemnités de rupture. Les indemnités non imposables (transport "net",
  représentation, prime de panier, prime de salissure) sont versées EN SUS.

  Colonnes DAS :
    T  Salaire brut de présence = salaire_brut − avantages nature − brut de congé
    U  Logement      = Σ indem_logement
    V  Eau & élec.   = Σ indem_eau_electricite
    W  Domesticité   = Σ indem_domesticite
    X  Nourriture    = Σ indem_nourriture
    Y  Indemnités imposables (657) = 0 (déjà incluses dans le brut)
    Z  Salaire brut de congé = Σ (allocations_conge + indem_compensatrice_conge)
    AA TOTAL (1 à 5) = T + U + V + W + X + Y + Z = salaire_brut
    AB TCTS = Σ tcs   AC IRPP = Σ irpp   AD CFP = Σ cfp   AE FNH = Σ fnh
    AF TOTAL impôts retenus = AB+AC+AD+AE
    AG Logement (NI) = 0           AH Transport (NI) = Σ indem_transport
    AI Domesticité (NI) = 0        AJ Autres (NI) = Σ (représentation+panier+salissure)
    AK TOTAL non imposables = AG+AH+AI+AJ
"""

from datetime import date
from io import BytesIO

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from core import csv_safe


class DASVide(Exception):
    """Levée quand aucun bulletin validé n'existe pour l'exercice demandé."""
    pass


# ── Couleurs de marque ────────────────────────────────────────────────────────
VERT   = "0F3D36"
OR     = "B8862F"
GRIS   = "F3F4F6"
BLANC  = "FFFFFF"


# ── Helpers de mapping vers les codes officiels ───────────────────────────────
def _code_nationalite(nat: str) -> int:
    n = (nat or "").upper()
    if not n or "GABON" in n:
        return 1
    cemac = ("CAMEROUN", "CENTRAFRIQUE", "CENTRAFRICAIN", "CONGO", "TCHAD",
             "GUINEE EQUATORIALE", "EQUATO")
    if any(p in n for p in cemac):
        return 2
    return 4  # par défaut : non africain (ajustable manuellement)


def _code_sexe(sexe: str) -> int:
    s = (sexe or "").upper()
    if s.startswith("M"):
        return 1
    if s.startswith("F"):
        return 2
    return 0


def _code_situation(sit: str) -> int:
    s = (sit or "").upper()
    if "MARI" in s:
        return 1
    if "CELIB" in s or "CÉLIB" in s:
        return 2
    if "VEU" in s:
        return 3
    if "DIVORC" in s:
        return 4
    return 0


def _age(date_naissance, annee: int):
    if not date_naissance:
        return None
    return annee - date_naissance.year


def _presence(date_embauche, date_cessation, annee: int):
    """Renvoie (jour_debut, mois_debut, jour_fin, mois_fin) de présence dans l'exercice."""
    debut_an = date(annee, 1, 1)
    fin_an = date(annee, 12, 31)
    deb = date_embauche if (date_embauche and date_embauche > debut_an) else debut_an
    fin = date_cessation if (date_cessation and date_cessation < fin_an) else fin_an
    if deb.year != annee:
        deb = debut_an
    if fin.year != annee:
        fin = fin_an
    return deb.day, deb.month, fin.day, fin.month


# ══════════════════════════════════════════════════════════════════════════════
# CONSTANTES DE VENTILATION — AJUSTABLES
# ──────────────────────────────────────────────────────────────────────────────
# Ces constantes pilotent la répartition des éléments du bulletin dans les
# colonnes de la DAS. Pour modifier le traitement fiscal d'une rubrique, il
# suffit de la déplacer d'une liste à l'autre ci-dessous — aucune autre
# modification de code n'est nécessaire.
#
# Rappel : `salaire_brut` inclut DÉJÀ base + heures sup + primes + avantages en
# nature + allocations de congé + indemnités de rupture. Les rubriques listées
# dans AVANTAGES_NATURE et BRUT_CONGE sont donc SOUSTRAITES du brut pour obtenir
# le "salaire brut de présence" (afin de ne pas les compter deux fois).
# Les rubriques NON_IMPOSABLE_* sont versées EN SUS du brut (déclaratif seul).
# ══════════════════════════════════════════════════════════════════════════════

# Statuts de bulletins retenus pour la DAS
STATUTS_RETENUS = ("VALIDÉ", "VALIDE", "PAYÉ", "PAYE")

# Avantages en nature → colonnes U/V/W/X (imposables, inclus dans le brut)
AVANTAGES_NATURE = {
    "logement":        ["indem_logement"],
    "eau_electricite": ["indem_eau_electricite"],
    "domesticite":     ["indem_domesticite"],
    "nourriture":      ["indem_nourriture"],
}

# Salaire brut de congé → colonne Z (imposable, inclus dans le brut)
BRUT_CONGE = ["allocations_conge", "indem_compensatrice_conge"]

# Indemnités imposables (partie 657) → colonne Y
# Vide par défaut : tout l'imposable est déjà dans le salaire_brut.
INDEMNITES_IMPOSABLES_657 = []

# Indemnités NON imposables → colonnes AG/AH/AI/AJ (versées en sus du brut)
NON_IMPOSABLE_LOGEMENT    = []
NON_IMPOSABLE_TRANSPORT   = ["indem_transport"]
NON_IMPOSABLE_DOMESTICITE = []
NON_IMPOSABLE_AUTRES      = ["indem_representation", "prime_panier", "prime_salisure"]

# Impôts retenus à la source → colonnes AB/AC/AD/AE
IMPOTS_RETENUS = {
    "tcts": ["tcs"], "irpp": ["irpp"], "cfp": ["cfp"], "fnh": ["fnh"],
}

# ── Honoraires (prestataires) — taux et filtres ───────────────────────────────
TAUX_TVA              = 0.18      # TVA Gabon
TAUX_RETENUE_LOCAL    = 0.095     # retenue à la source, prestataire résident
TAUX_RETENUE_ETRANGER = 0.20      # retenue à la source, prestataire étranger
# Statuts de factures retenus (on exclut les factures annulées)
STATUTS_FACTURES_RETENUS = ("EN_ATTENTE", "PAYEE", "PARTIELLE")
# Catégories de prestataires concernées par la déclaration d'honoraires
# (les fournisseurs de biens sont exclus par défaut).
CATEGORIES_HONORAIRES = ("FREELANCE", "SOUS_TRAITANT")


def _somme_champs(buls, champs, f):
    """Somme, sur une liste de bulletins, l'ensemble des `champs` indiqués."""
    return sum(f(b, c) for b in buls for c in champs)


def agreger_das(tenant, annee: int, db=None, models=None):
    """
    Construit la liste des lignes DAS (une par salarié) + le total entreprise.

    db / models : passés par l'appelant (blueprint) pour éviter les imports
    circulaires. Si None, importe le module `models` global.
    """
    if models is None:
        import models as models  # noqa
    BulletinPaie = models.BulletinPaie
    PeriodePaie = models.PeriodePaie
    Salarie = models.Salarie
    if db is None:
        db = models.db

    # Périodes de l'exercice
    periodes = PeriodePaie.query.filter_by(tenant_id=tenant.id, annee=annee).all()
    pids = [p.id for p in periodes]
    if not pids:
        raise DASVide(f"Aucune période de paie pour l'exercice {annee}.")

    bulletins = (BulletinPaie.query
                 .filter(BulletinPaie.tenant_id == tenant.id,
                         BulletinPaie.periode_id.in_(pids),
                         BulletinPaie.statut.in_(STATUTS_RETENUS))
                 .all())
    if not bulletins:
        raise DASVide(f"Aucun bulletin validé pour l'exercice {annee}.")

    # Regroupement par salarié
    par_sal = {}
    for b in bulletins:
        par_sal.setdefault(b.salarie_id, []).append(b)

    def f(b, champ):
        return float(getattr(b, champ) or 0)

    lignes = []
    for sal_id, buls in par_sal.items():
        s = db.session.get(Salarie, sal_id)
        if not s:
            continue

        brut_total = sum(f(b, "salaire_brut") for b in buls)
        av_logement = _somme_champs(buls, AVANTAGES_NATURE["logement"], f)
        av_eau = _somme_champs(buls, AVANTAGES_NATURE["eau_electricite"], f)
        av_dom = _somme_champs(buls, AVANTAGES_NATURE["domesticite"], f)
        av_nour = _somme_champs(buls, AVANTAGES_NATURE["nourriture"], f)
        av_nature = av_logement + av_eau + av_dom + av_nour
        brut_conge = _somme_champs(buls, BRUT_CONGE, f)
        indem_imposables = _somme_champs(buls, INDEMNITES_IMPOSABLES_657, f)
        brut_presence = max(brut_total - av_nature - brut_conge - indem_imposables, 0)

        tcts = _somme_champs(buls, IMPOTS_RETENUS["tcts"], f)
        irpp = _somme_champs(buls, IMPOTS_RETENUS["irpp"], f)
        cfp = _somme_champs(buls, IMPOTS_RETENUS["cfp"], f)
        fnh = _somme_champs(buls, IMPOTS_RETENUS["fnh"], f)
        total_impots = tcts + irpp + cfp + fnh

        ni_logement = _somme_champs(buls, NON_IMPOSABLE_LOGEMENT, f)
        ni_transport = _somme_champs(buls, NON_IMPOSABLE_TRANSPORT, f)
        ni_dom = _somme_champs(buls, NON_IMPOSABLE_DOMESTICITE, f)
        ni_autres = _somme_champs(buls, NON_IMPOSABLE_AUTRES, f)
        ni_total = ni_logement + ni_transport + ni_dom + ni_autres

        pj, pm, fj, fm = _presence(s.date_embauche, s.date_cessation, annee)
        cat_code = s.categorie.code if s.categorie else ""

        lignes.append({
            "matricule":      s.matricule or "",
            "cnss":           s.numero_cnss or "",
            "nif":            s.nif or "",
            "nom":            s.nom or "",
            "prenom":         s.prenom or "",
            "profession":     s.emploi or (s.categorie.libelle if s.categorie else ""),
            "code_emploi":    s.code_emploi or cat_code,
            "code_niveau":    s.niveau or "",
            "nationalite":    _code_nationalite(s.nationalite),
            "age":            _age(s.date_naissance, annee),
            "sexe":           _code_sexe(s.sexe),
            "situation":      _code_situation(s.situation_matrimoniale),
            "nb_enfants":     int(s.nb_enfants or 0),
            "presence_pj": pj, "presence_pm": pm, "presence_fj": fj, "presence_fm": fm,
            "brut_presence":  round(brut_presence, 0),
            "av_logement":    round(av_logement, 0),
            "av_eau":         round(av_eau, 0),
            "av_dom":         round(av_dom, 0),
            "av_nour":        round(av_nour, 0),
            "indem_imposables": round(indem_imposables, 0),
            "brut_conge":     round(brut_conge, 0),
            "total_1a5":      round(brut_total, 0),
            "tcts":           round(tcts, 0),
            "irpp":           round(irpp, 0),
            "cfp":            round(cfp, 0),
            "fnh":            round(fnh, 0),
            "total_impots":   round(total_impots, 0),
            "ni_logement":    round(ni_logement, 0),
            "ni_transport":   round(ni_transport, 0),
            "ni_dom":         round(ni_dom, 0),
            "ni_autres":      round(ni_autres, 0),
            "ni_total":       round(ni_total, 0),
            "nb_bulletins":   len(buls),
        })

    lignes.sort(key=lambda l: (l["nom"], l["prenom"]))

    totaux = {
        "nb_salaries":  len(lignes),
        "brut_presence": sum(l["brut_presence"] for l in lignes),
        "av_nature":    sum(l["av_logement"] + l["av_eau"] + l["av_dom"] + l["av_nour"] for l in lignes),
        "brut_conge":   sum(l["brut_conge"] for l in lignes),
        "total_1a5":    sum(l["total_1a5"] for l in lignes),
        "tcts":         sum(l["tcts"] for l in lignes),
        "irpp":         sum(l["irpp"] for l in lignes),
        "cfp":          sum(l["cfp"] for l in lignes),
        "fnh":          sum(l["fnh"] for l in lignes),
        "total_impots": sum(l["total_impots"] for l in lignes),
        "ni_total":     sum(l["ni_total"] for l in lignes),
    }
    return lignes, totaux


# ── Agrégation des honoraires (prestataires) ──────────────────────────────────
def agreger_honoraires(tenant, annee: int, db=None, models=None):
    """
    Construit la liste des lignes d'honoraires (une par prestataire) + totaux,
    à partir des factures prestataires de l'exercice.

    Renvoie ([], {...}) si aucune facture éligible — ce volet est optionnel et
    ne bloque jamais la DAS salaires.
    """
    if models is None:
        import models as models  # noqa
    Prestataire = models.Prestataire
    FacturePrestataire = models.FacturePrestataire
    ContratPrestation = models.ContratPrestation
    if db is None:
        db = models.db

    factures = (FacturePrestataire.query
                .filter(FacturePrestataire.tenant_id == tenant.id,
                        FacturePrestataire.statut.in_(STATUTS_FACTURES_RETENUS))
                .all())
    # Filtrer sur l'exercice (année de la facture)
    factures = [fa for fa in factures if fa.date_facture and fa.date_facture.year == annee]
    if not factures:
        return [], _totaux_honoraires([])

    par_prest = {}
    for fa in factures:
        par_prest.setdefault(fa.prestataire_id, []).append(fa)

    lignes = []
    for prest_id, facs in par_prest.items():
        p = db.session.get(Prestataire, prest_id)
        if not p:
            continue
        # Catégorie concernée par les honoraires ?
        if (p.categorie or "").upper() not in CATEGORIES_HONORAIRES:
            # On inclut tout de même si une retenue à la source a été pratiquée.
            if not any(float(fa.montant_retenue or 0) > 0 for fa in facs):
                continue

        ht = sum(float(fa.montant_ht or 0) for fa in facs)
        tva = sum(float(fa.montant_tva or 0) for fa in facs)
        retenue = sum(float(fa.montant_retenue or 0) for fa in facs)
        resident = bool(p.resident)
        retenue_local = retenue if resident else 0.0
        retenue_etranger = 0.0 if resident else retenue
        total = ht + tva

        # Période couverte par les factures de l'exercice
        dates = sorted(fa.date_facture for fa in facs if fa.date_facture)
        deb, fin = dates[0], dates[-1]
        # Type de prestation : objet du dernier contrat si disponible
        type_prestation = ""
        contrat = (ContratPrestation.query
                   .filter_by(tenant_id=tenant.id, prestataire_id=prest_id)
                   .order_by(ContratPrestation.date_debut.desc()).first())
        if contrat:
            type_prestation = contrat.objet or ""

        lignes.append({
            "nif":            p.nif or "",
            "raison_sociale": p.raison_sociale or "",
            "prenom":         "",   # le modèle Prestataire ne stocke pas de prénom séparé
            "activite":       p.activite or "",
            "type_prestation": type_prestation or p.categorie_label,
            "telephone":      p.telephone or "",
            "ville":          p.ville or "",
            "pays":           p.pays or "Gabon",
            "presence_pj": deb.day, "presence_pm": deb.month,
            "presence_fj": fin.day, "presence_fm": fin.month,
            "assujetti_tva":  "Oui" if p.assujetti_tva else "Non",
            "resident":       resident,
            "montant_ht":     round(ht, 0),
            "tva":            round(tva, 0),
            "retenue_local":  round(retenue_local, 0),
            "retenue_etranger": round(retenue_etranger, 0),
            "retenue_total":  round(retenue, 0),
            "total":          round(total, 0),
            "nb_factures":    len(facs),
        })

    lignes.sort(key=lambda l: l["raison_sociale"])
    return lignes, _totaux_honoraires(lignes)


def _totaux_honoraires(lignes):
    return {
        "nb_prestataires":  len(lignes),
        "montant_ht":       sum(l["montant_ht"] for l in lignes),
        "tva":              sum(l["tva"] for l in lignes),
        "retenue_local":    sum(l["retenue_local"] for l in lignes),
        "retenue_etranger": sum(l["retenue_etranger"] for l in lignes),
        "retenue_total":    sum(l["retenue_total"] for l in lignes),
        "total":            sum(l["total"] for l in lignes),
    }


# ── Génération du classeur Excel ──────────────────────────────────────────────
def _style_entete(ws, cell, texte, taille=9, fond=VERT, couleur=BLANC, bold=True):
    cell.value = texte
    cell.font = Font(name="Arial", size=taille, bold=bold, color=couleur)
    cell.fill = PatternFill("solid", fgColor=fond)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="D1D5DB")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def generer_das_excel(tenant, annee: int, models=None) -> bytes:
    """Génère la DAS au format Excel (.xlsx) et renvoie les octets du fichier."""
    lignes_hono, tot_hono = agreger_honoraires(tenant, annee, models=models)
    try:
        lignes, totaux = agreger_das(tenant, annee, models=models)
    except DASVide:
        # Pas de salaires : on autorise une DAS « honoraires seuls » si des
        # prestataires existent ; sinon on relaie l'erreur d'absence de données.
        if not lignes_hono:
            raise
        lignes, totaux = [], {
            "nb_salaries": 0, "brut_presence": 0, "av_nature": 0, "brut_conge": 0,
            "total_1a5": 0, "tcts": 0, "irpp": 0, "cfp": 0, "fnh": 0,
            "total_impots": 0, "ni_total": 0,
        }

    wb = openpyxl.Workbook()

    # ════════════════════════════════════════════════════════════════════════
    # Feuille 1 — Paramètres
    # ════════════════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Paramètres"
    ws.sheet_view.showGridLines = False
    titre = ws.cell(row=1, column=1)
    titre.value = "DÉCLARATION ANNUELLE DES SALAIRES (DAS)"
    titre.font = Font(name="Arial", size=14, bold=True, color=VERT)
    ws.cell(row=2, column=1, value=f"Exercice comptable : {annee}").font = Font(
        name="Arial", size=11, bold=True, color=OR)

    infos = [
        ("Société",            tenant.denomination or ""),
        ("Sigle",              tenant.sigle or ""),
        ("NIF",                tenant.nif or ""),
        ("N° Statistique",     ""),
        ("N° CNSS employeur",  tenant.numero_cnss or ""),
        ("N° CNAMGS",          tenant.numero_cnamgs or ""),
        ("Activité",           tenant.activite or ""),
        ("Adresse / Siège",    tenant.adresse or ""),
        ("Boîte postale",      tenant.boite_postale or ""),
        ("Ville",              tenant.ville or ""),
        ("Téléphone",          tenant.telephone or ""),
        ("Exercice",           annee),
        ("Date de la déclaration", date.today().strftime("%d/%m/%Y")),
        ("Nombre de salariés déclarés", totaux["nb_salaries"]),
    ]
    r = 4
    for label, val in infos:
        c1 = ws.cell(row=r, column=1, value=label)
        c1.font = Font(name="Arial", size=9, bold=True, color="374151")
        c2 = ws.cell(row=r, column=2, value=csv_safe(val) if isinstance(val, str) else val)
        c2.font = Font(name="Arial", size=9)
        r += 1
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 42

    # ════════════════════════════════════════════════════════════════════════
    # Feuille 2 — ID19 (détail par salarié)
    # ════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("ID19 - Détail salaires")
    ws2.sheet_view.showGridLines = False

    # Définition des colonnes : (clé, libellé, largeur, format_nombre)
    NUM = "#,##0"
    cols = [
        ("matricule",     "N° Matricule",        14, None),
        ("nif",           "NIF salarié",         16, None),
        ("cnss",          "N° CNSS",             16, None),
        ("nom",           "Nom",                 18, None),
        ("prenom",        "Prénom",              16, None),
        ("profession",    "Profession",          22, None),
        ("code_emploi",   "Code emploi",         10, None),
        ("code_niveau",   "Code niveau",         10, None),
        ("nationalite",   "Nat.",                 6, "0"),
        ("age",           "Âge",                  6, "0"),
        ("sexe",          "Sexe",                 6, "0"),
        ("situation",     "Sit. fam.",            7, "0"),
        ("nb_enfants",    "Enf.",                 6, "0"),
        ("presence_pj",   "Début (j)",            8, "0"),
        ("presence_pm",   "Début (m)",            8, "0"),
        ("presence_fj",   "Fin (j)",              7, "0"),
        ("presence_fm",   "Fin (m)",              7, "0"),
        ("brut_presence", "Brut présence",       14, NUM),
        ("av_logement",   "AN Logement",         12, NUM),
        ("av_eau",        "AN Eau/Élec.",        12, NUM),
        ("av_dom",        "AN Domest.",          12, NUM),
        ("av_nour",       "AN Nourriture",       13, NUM),
        ("indem_imposables", "Ind. imposables",  14, NUM),
        ("brut_conge",    "Brut congé",          13, NUM),
        ("total_1a5",     "TOTAL (1 à 5)",       15, NUM),
        ("tcts",          "TCTS",                12, NUM),
        ("irpp",          "IRPP",                13, NUM),
        ("cfp",           "CFP",                 11, NUM),
        ("fnh",           "FNH",                 11, NUM),
        ("total_impots",  "Total retenues",      14, NUM),
        ("ni_logement",   "NI Logement",         12, NUM),
        ("ni_transport",  "NI Transport",        12, NUM),
        ("ni_dom",        "NI Domest.",          11, NUM),
        ("ni_autres",     "NI Autres",           12, NUM),
        ("ni_total",      "Total NI",            13, NUM),
    ]

    # Titre + en-têtes
    ws2.cell(row=1, column=1, value=f"ID19 — Détail des salaires {annee} — {tenant.denomination or ''}").font = \
        Font(name="Arial", size=11, bold=True, color=VERT)
    HEADER_ROW = 3
    for j, (key, label, width, _fmt) in enumerate(cols, start=1):
        _style_entete(ws2, ws2.cell(row=HEADER_ROW, column=j), label)
        ws2.column_dimensions[get_column_letter(j)].width = width
    ws2.freeze_panes = ws2.cell(row=HEADER_ROW + 1, column=4)

    thin = Side(style="thin", color="E5E7EB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    data_start = HEADER_ROW + 1
    for i, ligne in enumerate(lignes):
        rr = data_start + i
        for j, (key, label, width, fmt) in enumerate(cols, start=1):
            _v = ligne.get(key)
            cell = ws2.cell(row=rr, column=j, value=csv_safe(_v) if isinstance(_v, str) else _v)
            cell.font = Font(name="Arial", size=8)
            cell.border = border
            if fmt:
                cell.number_format = fmt
            if key in ("matricule", "cnss", "nom", "prenom", "profession"):
                cell.alignment = Alignment(horizontal="left")
            else:
                cell.alignment = Alignment(horizontal="center" if fmt == "0" or fmt is None
                                           else "right")
        # Ligne alternée
        if i % 2 == 1:
            for j in range(1, len(cols) + 1):
                ws2.cell(row=rr, column=j).fill = PatternFill("solid", fgColor=GRIS)

    # Ligne TOTAL (formules SUM sur les colonnes numériques)
    total_row = data_start + len(lignes)
    ws2.cell(row=total_row, column=1, value="TOTAL").font = Font(name="Arial", size=9, bold=True, color=BLANC)
    for j, (key, label, width, fmt) in enumerate(cols, start=1):
        cell = ws2.cell(row=total_row, column=j)
        cell.fill = PatternFill("solid", fgColor=VERT)
        cell.font = Font(name="Arial", size=9, bold=True, color=BLANC)
        cell.border = border
        if fmt == NUM and len(lignes) > 0:
            col_letter = get_column_letter(j)
            cell.value = f"=SUM({col_letter}{data_start}:{col_letter}{total_row - 1})"
            cell.number_format = NUM
            cell.alignment = Alignment(horizontal="right")

    # ════════════════════════════════════════════════════════════════════════
    # Feuille 3 — ID20 (récapitulatif entreprise)
    # ════════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("ID20 - Récapitulatif")
    ws3.sheet_view.showGridLines = False
    ws3.cell(row=1, column=1, value=f"ID20 — Récapitulatif DAS {annee}").font = \
        Font(name="Arial", size=12, bold=True, color=VERT)
    ws3.cell(row=2, column=1, value=tenant.denomination or "").font = Font(name="Arial", size=10, bold=True)

    recap = [
        ("Nombre de salariés déclarés", totaux["nb_salaries"], "0"),
        ("Salaire brut de présence",    totaux["brut_presence"], NUM),
        ("Avantages en nature",         totaux["av_nature"], NUM),
        ("Salaire brut de congé",       totaux["brut_conge"], NUM),
        ("TOTAL imposable (1 à 5)",     totaux["total_1a5"], NUM),
        ("TCTS retenue",                totaux["tcts"], NUM),
        ("IRPP retenu",                 totaux["irpp"], NUM),
        ("CFP",                         totaux["cfp"], NUM),
        ("FNH",                         totaux["fnh"], NUM),
        ("TOTAL impôts retenus à la source", totaux["total_impots"], NUM),
        ("TOTAL indemnités non imposables",  totaux["ni_total"], NUM),
    ]
    rr = 4
    for label, val, fmt in recap:
        c1 = ws3.cell(row=rr, column=1, value=label)
        c1.font = Font(name="Arial", size=9, bold=("TOTAL" in label))
        c2 = ws3.cell(row=rr, column=2, value=val)
        c2.font = Font(name="Arial", size=9, bold=("TOTAL" in label))
        c2.number_format = fmt
        c2.alignment = Alignment(horizontal="right")
        if "TOTAL" in label:
            for col in (1, 2):
                ws3.cell(row=rr, column=col).fill = PatternFill("solid", fgColor=GRIS)
        rr += 1
    ws3.column_dimensions["A"].width = 36
    ws3.column_dimensions["B"].width = 20

    # Bloc honoraires dans le récapitulatif (si présents)
    if tot_hono["nb_prestataires"] > 0:
        rr += 1
        sep = ws3.cell(row=rr, column=1, value="HONORAIRES & COMMISSIONS (art. 189 CGI)")
        sep.font = Font(name="Arial", size=10, bold=True, color=OR)
        rr += 1
        recap_h = [
            ("Nombre de prestataires déclarés",       tot_hono["nb_prestataires"], "0"),
            ("Montant total HT des prestations",      tot_hono["montant_ht"], NUM),
            ("TVA",                                   tot_hono["tva"], NUM),
            ("Retenue à la source — prestataires locaux (ID23)",    tot_hono["retenue_local"], NUM),
            ("Retenue à la source — prestataires étrangers (ID24)", tot_hono["retenue_etranger"], NUM),
            ("TOTAL retenues honoraires",             tot_hono["retenue_total"], NUM),
        ]
        for label, val, fmt in recap_h:
            c1 = ws3.cell(row=rr, column=1, value=label)
            c1.font = Font(name="Arial", size=9, bold=("TOTAL" in label))
            c2 = ws3.cell(row=rr, column=2, value=val)
            c2.font = Font(name="Arial", size=9, bold=("TOTAL" in label))
            c2.number_format = fmt
            c2.alignment = Alignment(horizontal="right")
            if "TOTAL" in label:
                for col in (1, 2):
                    ws3.cell(row=rr, column=col).fill = PatternFill("solid", fgColor=GRIS)
            rr += 1

    # ════════════════════════════════════════════════════════════════════════
    # Feuille 4 — ID23/ID24 (honoraires versés) — uniquement si prestataires
    # ════════════════════════════════════════════════════════════════════════
    if lignes_hono:
        ws4 = wb.create_sheet("ID23-24 - Honoraires")
        ws4.sheet_view.showGridLines = False
        ws4.cell(row=1, column=1,
                 value=f"ID23 / ID24 — Honoraires, commissions versés {annee} — {tenant.denomination or ''}").font = \
            Font(name="Arial", size=11, bold=True, color=VERT)
        ws4.cell(row=2, column=1,
                 value="ID23 = versés au Gabon (résidents) · ID24 = versés hors du Gabon (étrangers) — art. 189 CGI").font = \
            Font(name="Arial", size=8, italic=True, color="6B7280")

        cols_h = [
            ("nif",             "NIF prestataire",   16, None),
            ("raison_sociale",  "Nom / Raison sociale", 26, None),
            ("activite",        "Activité",          18, None),
            ("type_prestation", "Type de prestation", 22, None),
            ("telephone",       "Téléphone",         14, None),
            ("ville",           "Ville",             12, None),
            ("pays",            "Pays",              11, None),
            ("presence_pj",     "Début (j)",          8, "0"),
            ("presence_pm",     "Début (m)",          8, "0"),
            ("presence_fj",     "Fin (j)",            7, "0"),
            ("presence_fm",     "Fin (m)",            7, "0"),
            ("assujetti_tva",   "TVA ?",              7, None),
            ("montant_ht",      "Montant HT",        14, NUM),
            ("tva",             "TVA",               12, NUM),
            ("retenue_local",   "RAS locaux (9,5%)", 13, NUM),
            ("retenue_etranger","RAS étrangers (20%)", 14, NUM),
            ("total",           "Total TTC",         14, NUM),
        ]
        HROW = 4
        for j, (key, label, width, _fmt) in enumerate(cols_h, start=1):
            _style_entete(ws4, ws4.cell(row=HROW, column=j), label)
            ws4.column_dimensions[get_column_letter(j)].width = width
        ws4.freeze_panes = ws4.cell(row=HROW + 1, column=3)

        h_start = HROW + 1
        for i, lh in enumerate(lignes_hono):
            rr2 = h_start + i
            for j, (key, label, width, fmt) in enumerate(cols_h, start=1):
                _v = lh.get(key)
                cell = ws4.cell(row=rr2, column=j, value=csv_safe(_v) if isinstance(_v, str) else _v)
                cell.font = Font(name="Arial", size=8)
                cell.border = border
                if fmt:
                    cell.number_format = fmt
                if key in ("nif", "raison_sociale", "activite", "type_prestation", "telephone", "ville", "pays"):
                    cell.alignment = Alignment(horizontal="left")
                else:
                    cell.alignment = Alignment(horizontal="right" if fmt == NUM else "center")
            if i % 2 == 1:
                for j in range(1, len(cols_h) + 1):
                    ws4.cell(row=rr2, column=j).fill = PatternFill("solid", fgColor=GRIS)

        trow = h_start + len(lignes_hono)
        ws4.cell(row=trow, column=1, value="TOTAL").font = Font(name="Arial", size=9, bold=True, color=BLANC)
        for j, (key, label, width, fmt) in enumerate(cols_h, start=1):
            cell = ws4.cell(row=trow, column=j)
            cell.fill = PatternFill("solid", fgColor=VERT)
            cell.font = Font(name="Arial", size=9, bold=True, color=BLANC)
            cell.border = border
            if fmt == NUM:
                cl = get_column_letter(j)
                cell.value = f"=SUM({cl}{h_start}:{cl}{trow - 1})"
                cell.number_format = NUM
                cell.alignment = Alignment(horizontal="right")

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


# ══════════════════════════════════════════════════════════════════════════════
# CONTRÔLES DE COHÉRENCE (reproduits de la feuille SOCIETE du modèle officiel)
# ══════════════════════════════════════════════════════════════════════════════
def controles_coherence_das(lignes, totaux):
    """Vérifie la cohérence de la DAS avant dépôt, comme le modèle officiel DGI.
    Retourne une liste de dicts : {libelle, statut ('OK'|'ERREUR'|'ATTENTION'), detail}.
    Ne bloque rien : aide au contrôle."""
    res = []

    def _ajouter(libelle, ok, detail="", niveau_ko="ERREUR"):
        res.append({"libelle": libelle,
                    "statut": "OK" if ok else niveau_ko,
                    "detail": detail})

    # 1. Salariés de moins de 15 ans (âge invalide)
    moins15 = [l for l in lignes if isinstance(l.get("age"), int) and 0 < l["age"] < 15]
    _ajouter("Salariés de moins de 15 ans", not moins15,
             ", ".join(f"{l['nom']} {l['prenom']}" for l in moins15))

    # 2. Salariés de plus de 65 ans
    plus65 = [l for l in lignes if isinstance(l.get("age"), int) and l["age"] > 65]
    _ajouter("Salariés de plus de 65 ans", not plus65,
             ", ".join(f"{l['nom']} {l['prenom']}" for l in plus65),
             niveau_ko="ATTENTION")

    # 3. NIF manquant (indispensable pour l'ID19/ID21)
    sans_nif = [l for l in lignes if not (l.get("nif") or "").strip()]
    _ajouter("NIF renseigné pour tous les salariés", not sans_nif,
             f"{len(sans_nif)} salarié(s) sans NIF"
             + (" : " + ", ".join(f"{l['nom']} {l['prenom']}" for l in sans_nif[:5]) if sans_nif else ""))

    # 4. Matricules en double (lignes mélangées)
    from collections import Counter
    mats = Counter((l.get("matricule") or "").strip() for l in lignes if (l.get("matricule") or "").strip())
    doublons = [m for m, n in mats.items() if n > 1]
    _ajouter("Aucun matricule en double", not doublons,
             "Matricules dupliqués : " + ", ".join(doublons) if doublons else "")

    # 5. Cohérence du total imposable (somme des lignes = total déclaré)
    somme_total = round(sum(l.get("total_1a5", 0) for l in lignes), 0)
    tot_declare = round(totaux.get("total_1a5", 0), 0)
    _ajouter("Cohérence du total imposable (1 à 5)", abs(somme_total - tot_declare) < 1,
             f"Somme lignes {int(somme_total):,} ≠ total {int(tot_declare):,}".replace(",", " ")
             if abs(somme_total - tot_declare) >= 1 else "")

    # 6. Cohérence des totaux d'impôts (TCS + IRPP + FNH)
    somme_tcts = round(sum(l.get("tcts", 0) for l in lignes), 0)
    _ajouter("Cohérence du total TCS", abs(somme_tcts - round(totaux.get("tcts", 0), 0)) < 1)
    somme_irpp = round(sum(l.get("irpp", 0) for l in lignes), 0)
    _ajouter("Cohérence du total IRPP", abs(somme_irpp - round(totaux.get("irpp", 0), 0)) < 1)
    somme_fnh = round(sum(l.get("fnh", 0) for l in lignes), 0)
    _ajouter("Cohérence du total FNH", abs(somme_fnh - round(totaux.get("fnh", 0), 0)) < 1)

    # 7. Âge non renseigné (contrôle de complétude)
    sans_age = [l for l in lignes if not isinstance(l.get("age"), int) or l["age"] <= 0]
    _ajouter("Date de naissance renseignée (âge calculable)", not sans_age,
             f"{len(sans_age)} salarié(s) sans âge", niveau_ko="ATTENTION")

    return res
